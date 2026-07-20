# -*- coding: utf-8 -*-
"""
Lógica compartida de la Conciliación Comercial vs Contable (P&L + reconciliación).

SIN Streamlit: lo usan tanto la vista del dashboard como el script
`planilla_pyl_formulas.py`. Recibe DataFrames ya cargados (la fuente la pone
el caller: cargar_hoja en la app, gspread en el script).

- construir_dataframes(df_ar, df_glosas, nc_detalle, nc2canal) -> bundle
- calcular(bundle, mes, canal, kam) -> dict con el P&L y la reconciliación
- construir_workbook(bundle) -> openpyxl Workbook (versión con fórmulas)
"""
import re
import unicodedata
from collections import defaultdict

import pandas as pd

EQUIPO = {"trinidad", "ignacia", "claudia", "nicole"}
# Corrección de línea de negocio: el sheet mete "Marketing" en Páginas Propias;
# en el RAW contable (fuente única) su tipo_negocio es "Marketing". Clave = _norm(canal).
NEGOCIO_FIX = {"marketing": "Marketing"}
# Scope de la conciliación (alineado con el crossover de devoluciones que analiza Gabriela):
# líneas de negocio + canal explícito. Reemplaza el viejo "canales con KAM".
SCOPE_NEG = {"marketplace", "fidelizacion", "paginas propias"}
SCOPE_CANAL = {"unionx b2b"}
MES_NOM = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
# Último mes cerrado con datos comercial + contable en el Sheet. Subir a medida que
# se cargan los meses (define el selector de Mes y el tope de las consultas al RAW).
MES_MAX = 6  # jun-2026
# Selector de período: agregados (YTD/Q1/Q2/1S) + meses individuales cargados.
MESES_OPT = ["YTD", "Q1", "Q2", "1S"] + MES_NOM[1:MES_MAX + 1]


def _meses_periodo(mes):
    """Set de meses (int) que abarca la selección de período, intersectado con lo
    cargado (1..MES_MAX). YTD=todo; 1S=Ene–Jun; Q1=Ene–Mar; Q2=Abr–Jun; mes puntual={n}."""
    disp = set(range(1, MES_MAX + 1))
    if mes == "YTD":
        return set(disp)
    if mes == "1S":
        return {m for m in disp if m <= 6}
    if mes == "Q1":
        return disp & {1, 2, 3}
    if mes == "Q2":
        return disp & {4, 5, 6}
    if mes in MES_NOM:
        return {MES_NOM.index(mes)} & disp
    return set(disp)
MESES_ES = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
            "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
            "noviembre": 11, "diciembre": 12}
# columnas por POSICIÓN en 'Análisis de Resultados'
IX = {"ano": 0, "canal": 2, "kam": 3, "mes": 4,
      "VentaC": 8, "CostoC": 9, "ComVentaC": 11, "ComEnvioC": 12, "MktC": 13,
      "VentaK": 18, "CostoK": 19, "ComVentaK": 21, "ComEnvioK": 22, "MktK": 23}
COMP = ["Venta", "Costo", "ComVenta", "ComEnvio", "Mkt"]
# líneas del P&L (etiqueta, componente)
LINEAS = [("Venta", "Venta"), ("Costo de Venta", "Costo"), ("Margen Directo", None),
          ("Comisión Venta", "ComVenta"), ("Comisión Envío", "ComEnvio"),
          ("Marketing", "Mkt"), ("Contribución", None)]


def _norm(s):
    s = str(s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def num(s):
    s = str(s).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except (ValueError, AttributeError):
        return 0.0


def _mes_num(txt):
    t = _norm(txt)
    for nombre, n in MESES_ES.items():
        if nombre in t:
            return n
    m = re.search(r"\b(\d{1,2})\b", t)
    return int(m.group(1)) if m and 1 <= int(m.group(1)) <= 12 else None


def _origen_glosa(glosa, mes_arch):
    t = _norm(glosa)
    mes = next((n for nombre, n in MESES_ES.items() if re.search(rf"\b{nombre}\b", t)), None)
    ym = re.search(r"\b(20\d{2})\b", t)
    anio = int(ym.group(1)) if ym else None
    if anio is None and mes is not None and mes_arch is not None and mes > mes_arch:
        anio = 2025
    if anio is None:
        anio = 2026
    if mes is None:
        mes = mes_arch
    return anio, mes


def construir_dataframes(df_ar, df_glosas, nc_detalle, nc2canal):
    """df_ar: 'Análisis de Resultados' (acceso posicional). df_glosas: 'Detalle Glosas 2026'.
    nc_detalle: parquet NC Odoo (NC, Mes NC, Fecha venta original, Neto). nc2canal: dict NC->canal."""
    df = df_ar.copy()
    df.columns = range(df.shape[1])
    for k, i in IX.items():
        if k not in ("canal", "kam"):
            df[i] = df[i].apply(num)
    df["_ck"] = df[IX["canal"]].apply(_norm)
    df = df[(df[IX["ano"]] == 2026) & (df[IX["mes"]] >= 1) & (df[IX["mes"]] <= MES_MAX)].copy()

    ck_kams = defaultdict(set)
    for _, r in df.iterrows():
        for p in re.split(r"[/,]", str(r[IX["kam"]])):
            k = _norm(p)
            if k:
                ck_kams[r["_ck"]].add(k)
    es_conkam = {ck: bool(v & EQUIPO) for ck, v in ck_kams.items()}
    nombre = df.groupby("_ck")[IX["canal"]].agg(lambda s: s.mode().iat[0] if len(s.mode()) else s.iloc[0]).to_dict()
    kam_dom = {}
    for ck, d in df.groupby("_ck"):
        vt = d.groupby(d[IX["kam"]].apply(lambda x: str(x).strip()))[IX["VentaC"]].sum()
        cand = [(kam, v) for kam, v in vt.items() if _norm(kam) in EQUIPO]
        kam_dom[ck] = (max(cand, key=lambda x: x[1])[0] if cand else (vt.idxmax() if len(vt) else ""))
    negocio_dom = {}  # ck -> línea de negocio (col 1 = Negocio)
    for ck, d in df.groupby("_ck"):
        negs = d[1].astype(str).str.strip()
        negocio_dom[ck] = (negs.mode().iat[0] if len(negs.mode()) else (negs.iloc[0] if len(negs) else ""))
        # El sheet clasifica "Marketing" como Páginas Propias; en el RAW (fuente) es "Marketing".
        if ck in NEGOCIO_FIX:
            negocio_dom[ck] = NEGOCIO_FIX[ck]
    # Scope = líneas de negocio del crossover + canal UnionX B2B (ya NO "canal con KAM").
    es_scope = {ck: (_norm(negocio_dom.get(ck, "")) in SCOPE_NEG) or (ck in SCOPE_CANAL)
                for ck in negocio_dom}
    df["_conkam"] = df["_ck"].map(es_scope).fillna(False)
    dk = df[df["_conkam"]].copy()

    datos = []
    for (m, ck), d in dk.groupby([IX["mes"], "_ck"]):
        row = {"Mes": MES_NOM[int(m)], "Canal": nombre.get(ck, ck), "KAM": str(kam_dom.get(ck, "")).strip(),
               "Negocio": str(negocio_dom.get(ck, "")).strip()}
        for comp in COMP:
            row[f"{comp}_Com"] = float(d[IX[comp + "C"]].sum())
            row[f"{comp}_Cont"] = float(d[IX[comp + "K"]].sum())
        datos.append(row)
    datos = pd.DataFrame(datos).sort_values(["Canal", "Mes"]).reset_index(drop=True)
    canales = sorted(datos["Canal"].unique())
    kams = sorted(k for k in datos["KAM"].unique() if k)
    canal2negocio = dict(zip(datos["Canal"], datos["Negocio"]))
    negocios = sorted(set(n for n in canal2negocio.values() if n))

    # NC detalle
    def _orig_fecha(fv):
        s = str(fv)[:10]
        if len(s) >= 7 and s[:4].isdigit():
            return int(s[:4]), int(s[5:7])
        return 0, 0
    nc_rows = []
    if nc_detalle is not None and len(nc_detalle):
        ncd = nc_detalle.copy()
        ncd["Canal"] = ncd["NC"].map(nc2canal).fillna("(no id)")
        ncd["_ck"] = ncd["Canal"].apply(_norm)
        ncd = ncd[ncd["_ck"].map(lambda c: es_scope.get(c, False))]
        for _, r in ncd.iterrows():
            oa, om = _orig_fecha(r.get("Fecha venta original", ""))
            nc_rows.append({"Mes": MES_NOM[int(r["Mes NC"])], "Canal": nombre.get(r["_ck"], r["Canal"]),
                            "KAM": str(kam_dom.get(r["_ck"], "")).strip(),
                            "OrigenAnio": oa, "OrigenMes": MES_NOM[om] if 1 <= om <= 12 else "",
                            "Neto": float(r["Neto"])})
    nc_tab = (pd.DataFrame(nc_rows) if nc_rows else
              pd.DataFrame(columns=["Mes", "Canal", "KAM", "OrigenAnio", "OrigenMes", "Neto"]))
    if len(nc_tab):
        nc_tab = nc_tab.groupby(["Mes", "Canal", "KAM", "OrigenAnio", "OrigenMes"], as_index=False)["Neto"].sum()

    # Comisiones detalle (glosas) + aporte del canal (reclasificación venta→menor comisión)
    # Aporte por glosa: Falabella "Oportunidades Únicas", Ripley "Opex" (Paris no se ve por glosa).
    crows, arows = [], []
    APORTE_PAT = ("oportunidad", "opex")
    if df_glosas is not None and len(df_glosas):
        for _, r in df_glosas.iterrows():
            cat = _norm(r.get("Categoría Analítica", ""))
            if "comis" not in cat and cat != "envio":
                continue
            ck = _norm(r.get("Canal", ""))
            if not es_scope.get(ck, False):
                continue
            m = _mes_num(r.get("Mes", ""))
            gl = _norm(r.get("Glosa", ""))
            oa, om = _origen_glosa(r.get("Glosa", ""), m)
            mes_n = MES_NOM[m] if m else "?"
            canal_n = nombre.get(ck, str(r.get("Canal", "")).strip())
            kam_n = str(kam_dom.get(ck, "")).strip()
            monto = num(r.get("Monto ($)", ""))
            crows.append({"Mes": mes_n, "Canal": canal_n, "KAM": kam_n, "OrigenAnio": oa,
                          "OrigenMes": MES_NOM[om] if om and 1 <= om <= 12 else "", "Monto": monto})
            if any(p in gl for p in APORTE_PAT):
                arows.append({"Mes": mes_n, "Canal": canal_n, "KAM": kam_n, "Monto": monto})
    com_tab = (pd.DataFrame(crows) if crows else
               pd.DataFrame(columns=["Mes", "Canal", "KAM", "OrigenAnio", "OrigenMes", "Monto"]))
    if len(com_tab):
        com_tab = com_tab.groupby(["Mes", "Canal", "KAM", "OrigenAnio", "OrigenMes"], as_index=False)["Monto"].sum()
    aporte_tab = (pd.DataFrame(arows) if arows else pd.DataFrame(columns=["Mes", "Canal", "KAM", "Monto"]))
    if len(aporte_tab):
        aporte_tab = aporte_tab.groupby(["Mes", "Canal", "KAM"], as_index=False)["Monto"].sum()

    return {"datos": datos, "nc_tab": nc_tab, "com_tab": com_tab, "aporte_tab": aporte_tab,
            "canales": canales, "kams": kams, "canal2negocio": canal2negocio, "negocios": negocios}


# ------------------------------------------------------------------ cálculo (display)
def _match(v, sel, todos):
    return True if sel == todos else str(v) == str(sel)


def calcular(bundle, mes="YTD", canal="TODOS", kam="TODOS", negocio="TODOS"):
    datos, nc_tab, com_tab = bundle["datos"], bundle["nc_tab"], bundle["com_tab"]
    c2n = bundle.get("canal2negocio", {})
    _cset = None if negocio == "TODOS" else {c for c, n in c2n.items() if n == negocio}
    _okc = lambda v: _match(v, canal, "TODOS") and (_cset is None or v in _cset)
    MSET = _meses_periodo(mes)  # meses (int) del período seleccionado
    _inp = lambda mnom: (MES_NOM.index(mnom) if mnom in MES_NOM else -1) in MSET
    f = datos[datos.apply(lambda r: _inp(r["Mes"]) and _okc(r["Canal"])
                          and _match(r["KAM"], kam, "TODOS"), axis=1)]
    g = lambda c: float(f[c].sum()) if len(f) else 0.0

    lineas = []
    vals = {}
    for label, comp in LINEAS:
        if comp is None:
            continue
        vals[label] = (g(f"{comp}_Com"), g(f"{comp}_Cont"))
    venta_c, venta_k = vals["Venta"]
    margen_c = venta_c - vals["Costo de Venta"][0]
    margen_k = venta_k - vals["Costo de Venta"][1]
    contrib_c = margen_c - vals["Comisión Venta"][0] - vals["Comisión Envío"][0] - vals["Marketing"][0]
    contrib_k = margen_k - vals["Comisión Venta"][1] - vals["Comisión Envío"][1] - vals["Marketing"][1]
    vals["Margen Directo"] = (margen_c, margen_k)
    vals["Contribución"] = (contrib_c, contrib_k)
    for label, _ in LINEAS:
        com, con = vals[label]
        lineas.append({"Línea": label, "Comercial": com, "Contable": con,
                       "Δ $ (Com−Cont)": com - con,
                       "Δ %": ((com - con) / abs(con)) if con else None})
    pyl = pd.DataFrame(lineas)

    # NC del/otro relativo al período: NC registrada dentro del período; "del" = origen
    # 2026 dentro del período, "otro" = registrada en el período pero origen fuera (2025
    # o mes 2026 fuera del período).
    def _flag(row, kind):
        if not (_okc(row["Canal"]) and _match(row["KAM"], kam, "TODOS")):
            return False
        if not _inp(row["Mes"]):
            return False
        oa, om = row["OrigenAnio"], row["OrigenMes"]
        in_per = (oa == 2026 and _inp(om))
        return (not in_per) if kind == "otro" else in_per
    nc_del = float(nc_tab[nc_tab.apply(lambda r: _flag(r, "del"), axis=1)]["Neto"].sum()) if len(nc_tab) else 0.0
    nc_otro = float(nc_tab[nc_tab.apply(lambda r: _flag(r, "otro"), axis=1)]["Neto"].sum()) if len(nc_tab) else 0.0

    def _cflag(row):
        if not (_okc(row["Canal"]) and _match(row["KAM"], kam, "TODOS")):
            return False
        if not _inp(row["Mes"]):
            return False
        oa, om = row["OrigenAnio"], row["OrigenMes"]
        return not (oa == 2026 and _inp(om))
    com_otro = float(com_tab[com_tab.apply(_cflag, axis=1)]["Monto"].sum()) if len(com_tab) else 0.0

    aporte_tab = bundle.get("aporte_tab")
    if aporte_tab is not None and len(aporte_tab):
        am = aporte_tab[aporte_tab.apply(lambda r: _inp(r["Mes"])
                        and _okc(r["Canal"]) and _match(r["KAM"], kam, "TODOS"), axis=1)]
        aporte_canal = float(am["Monto"].sum())
    else:
        aporte_canal = 0.0

    # reconciliación
    venta_aj = venta_c - nc_otro
    m_com = (margen_c / venta_c) if venta_c else 0.0
    m_cont = (margen_k / venta_k) if venta_k else 0.0
    margen_aj_com = venta_aj * m_com
    costeo = venta_aj * (m_com - m_cont)
    margen_aj = venta_aj * m_cont
    com_comercial = vals["Comisión Venta"][0] + vals["Comisión Envío"][0] + vals["Marketing"][0]
    contrib_aj = margen_aj - com_comercial
    no_caida = vals["Comisión Venta"][0] - vals["Comisión Venta"][1]

    return {
        "pyl": pyl,
        "contrib_com": contrib_c, "contrib_cont": contrib_k, "delta_contrib": contrib_c - contrib_k,
        "venta_com": venta_c, "nc_otro": nc_otro, "nc_del": nc_del, "venta_aj": venta_aj,
        "margen_aj_com": margen_aj_com, "costeo": costeo, "margen_aj": margen_aj, "margen_cont_real": margen_k,
        "com_otro": com_otro, "no_caida": no_caida, "aporte_canal": aporte_canal,
        "contrib_aj": contrib_aj, "por_explicar": contrib_aj - contrib_k,
    }


# ------------------------------------------------------------------ Detalle P&L (RAW + glosas)
def _mes_a_int(mes):
    """'Ene'..'May' -> 1..5; 'YTD' -> None."""
    return None if mes == "YTD" else (MES_NOM.index(mes) if mes in MES_NOM else None)


def calcular_detalle(raw_comp, glosas_comp, mes="YTD", canal="TODOS", kam="TODOS", negocio="TODOS"):
    """P&L comercial detallado desde el RAW (+ glosas), por mes/canal/kam.

    raw_comp: filas {Canal, KAM, Tipo('ing_prod'|'ing_envio'|'nc'), Mes(int),
              OrigenAnio, OrigenMes, Venta, Costo}.
    glosas_comp: filas {Canal, KAM, Mes(int), Cat('venta'|'envio'|'mkt'|'otro'),
                 OrigenAnio, OrigenMes, Monto}.
    Regla período (intrínseca a cada fila): origen 2026 y OrigenMes == Mes de registro.
    NC otro 2026 = origen 2026 pero otro mes; NC otro 2025 = origen <= 2025.
    """
    mset = _meses_periodo(mes)  # meses (int) del período

    def _flt(df):
        if df is None or not len(df):
            return df if df is not None else pd.DataFrame()
        m = df
        if canal != "TODOS":
            m = m[m["Canal"] == canal]
        if kam != "TODOS":
            m = m[m["KAM"] == kam]
        if negocio != "TODOS" and "Negocio" in m.columns:
            m = m[m["Negocio"] == negocio]
        m = m[m["Mes"].isin(mset)]
        return m

    rc = _flt(raw_comp)
    s = lambda d, col: float(d[col].sum()) if len(d) else 0.0

    prod = rc[rc["Tipo"] == "ing_prod"] if len(rc) else rc
    env = rc[rc["Tipo"] == "ing_envio"] if len(rc) else rc
    nc = rc[rc["Tipo"] == "nc"] if len(rc) else rc
    ing_prod, costo_prod = s(prod, "Venta"), s(prod, "Costo")
    ing_env, costo_env = s(env, "Venta"), s(env, "Costo")
    if len(nc):
        # del período = origen 2026 dentro del período; otro 2026 = 2026 fuera del período; otro = <=2025
        per = nc[(nc["OrigenAnio"] == 2026) & (nc["OrigenMes"].isin(mset))]
        o26 = nc[(nc["OrigenAnio"] == 2026) & (~nc["OrigenMes"].isin(mset))]
        o25 = nc[nc["OrigenAnio"] <= 2025]
    else:
        per = o26 = o25 = nc
    nc_per_v, nc_per_c = s(per, "Venta"), s(per, "Costo")
    nc_o26_v, nc_o26_c = s(o26, "Venta"), s(o26, "Costo")
    nc_o25_v, nc_o25_c = s(o25, "Venta"), s(o25, "Costo")

    costo = costo_prod + costo_env + nc_per_c + nc_o26_c + nc_o25_c
    venta_neta = ing_prod + ing_env + nc_per_v + nc_o26_v + nc_o25_v
    margen_directo = venta_neta - costo

    gc = _flt(glosas_comp)
    if len(gc):
        es_per = (gc["OrigenAnio"] == 2026) & (gc["OrigenMes"].isin(mset))
        per_g = gc[es_per]
        com_v = float(per_g[per_g["Cat"] == "venta"]["Monto"].sum())
        com_e = float(per_g[per_g["Cat"] == "envio"]["Monto"].sum())
        com_m = float(per_g[per_g["Cat"] == "mkt"]["Monto"].sum())
        glosa_otro = float(gc[~es_per]["Monto"].sum())
    else:
        com_v = com_e = com_m = glosa_otro = 0.0

    margen_contrib = margen_directo - com_v - com_e - com_m - glosa_otro
    return {
        "ing_prod": ing_prod, "ing_env": ing_env,
        "nc_per": nc_per_v, "nc_o2026": nc_o26_v, "nc_o2025": nc_o25_v,
        "nc_per_c": nc_per_c, "nc_o2026_c": nc_o26_c, "nc_o2025_c": nc_o25_c,
        "costo": costo, "costo_ing": costo_prod + costo_env, "costo_nc": nc_per_c + nc_o26_c + nc_o25_c,
        "margen_directo": margen_directo,
        "com_venta": com_v, "com_envio": com_e, "com_mkt": com_m, "glosa_otro": glosa_otro,
        "margen_contrib": margen_contrib, "venta_neta": venta_neta,
    }


# ------------------------------------------------------------------ Bloque B2B (Distribución / Nicolás)
NEG_IX = 1  # columna Negocio en 'Análisis de Resultados'


def construir_b2b(df_ar, df_meta):
    """Distribución (B2B): resultado contable (= comercial) por mes×canal + meta total.
    El presupuesto de Distribución no viene abierto por canal: es el total (cargado
    bajo 'Paris tienda'); se compara el total Distribución vs ese total."""
    df = df_ar.copy()
    df.columns = range(df.shape[1])
    for i in (IX["ano"], IX["mes"], 18, 19, 21, 22, 23, 25):
        df[i] = df[i].apply(num)
    df["_neg"] = df[NEG_IX].apply(_norm)
    d = df[(df[IX["ano"]] == 2026) & (df[IX["mes"]] >= 1) & (df[IX["mes"]] <= MES_MAX) & (df["_neg"] == "distribucion")]
    rows = []
    for (m, canal), g in d.groupby([IX["mes"], IX["canal"]]):
        rows.append({"Mes": MES_NOM[int(m)], "Canal": str(canal).strip(),
                     "Venta": float(g[18].sum()), "Costo": float(g[19].sum()),
                     "Comisión Venta": float(g[21].sum()), "Comisión Envío": float(g[22].sum()),
                     "Marketing": float(g[23].sum()), "Contribución": float(g[25].sum())})
    cols = ["Mes", "Canal", "Venta", "Costo", "Comisión Venta", "Comisión Envío", "Marketing", "Contribución"]
    datos_b2b = pd.DataFrame(rows) if rows else pd.DataFrame(columns=cols)

    meta_rows = []
    if df_meta is not None and len(df_meta):
        def col(name):
            for c in df_meta.columns:
                if str(c).strip().lower() == name.lower():
                    return c
            return None
        cA, cM, cN, cMV, cMC = (col("AÑO"), col("Mes"), col("Negocio"), col("Meta Venta"), col("Meta Contribución"))
        for _, r in df_meta.iterrows():
            try:
                if int(num(r[cA])) == 2026 and 1 <= int(num(r[cM])) <= MES_MAX and _norm(r[cN]) == "distribucion":
                    meta_rows.append({"Mes": MES_NOM[int(num(r[cM]))],
                                      "MetaVenta": num(r[cMV]), "MetaContrib": num(r[cMC])})
            except (TypeError, ValueError):
                pass
    meta_b2b = (pd.DataFrame(meta_rows).groupby("Mes", as_index=False).sum() if meta_rows
                else pd.DataFrame(columns=["Mes", "MetaVenta", "MetaContrib"]))
    return {"datos_b2b": datos_b2b, "meta_b2b": meta_b2b}


def calcular_b2b(b2b, mes="YTD"):
    d, mt = b2b["datos_b2b"], b2b["meta_b2b"]
    nombres = {MES_NOM[i] for i in _meses_periodo(mes)}  # nombres de mes del período
    d = d[d["Mes"].isin(nombres)]
    mt = mt[mt["Mes"].isin(nombres)]
    money_cols = ["Venta", "Costo", "Comisión Venta", "Comisión Envío", "Marketing", "Contribución"]
    pl_canal = (d.groupby("Canal", as_index=False)[money_cols].sum().sort_values("Venta", ascending=False)
                if len(d) else pd.DataFrame(columns=["Canal"] + money_cols))
    if len(pl_canal):
        pl_canal.insert(2, "Margen", pl_canal["Venta"] - pl_canal["Costo"])
    tot = {k: float(d[k].sum()) if len(d) else 0.0 for k in money_cols}
    tot["Margen"] = tot["Venta"] - tot["Costo"]
    meta_v = float(mt["MetaVenta"].sum()) if len(mt) else 0.0
    meta_c = float(mt["MetaContrib"].sum()) if len(mt) else 0.0
    return {"pl_canal": pl_canal, "tot": tot, "meta_venta": meta_v, "meta_contrib": meta_c,
            "cumpl_venta": (tot["Venta"] / meta_v) if meta_v else None,
            "cumpl_contrib": (tot["Contribución"] / meta_c) if meta_c else None}


# ------------------------------------------------------------------ Excel con fórmulas
def construir_workbook(bundle):
    """Workbook openpyxl con selectores Mes/Canal/KAM y todo por fórmula (idéntico al
    script planilla_pyl_formulas)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    datos, nc_tab, com_tab = bundle["datos"], bundle["nc_tab"], bundle["com_tab"]
    canales, kams = bundle["canales"], bundle["kams"]
    PL = "'P&L'!$B$3"
    THIN = Side(style="thin", color="D0D0D0")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    AZUL = PatternFill("solid", fgColor="1E40AF"); GRIS = PatternFill("solid", fgColor="F1F5F9")
    AMBAR = PatternFill("solid", fgColor="FEF3C7"); VERDE = PatternFill("solid", fgColor="DCFCE7")
    MONEY = '#,##0;[Red]-#,##0'; PCT = '0.0%;[Red]-0.0%'

    wb = Workbook()
    wsd = wb.active
    wsd.title = "Datos"
    dcols = ["Mes", "Canal", "KAM"] + [f"{c}_{s}" for c in COMP for s in ("Com", "Cont")]
    wsd.append(dcols)
    for _, r in datos.iterrows():
        wsd.append([r[c] for c in dcols])
    nD = len(datos)
    DC = {name: get_column_letter(i + 1) for i, name in enumerate(dcols)}
    drng = lambda name: f"Datos!${DC[name]}$2:${DC[name]}${nD+1}"

    def _detalle(ws, tab, valcol):
        ws.append(["Mes", "Canal", "KAM", "OrigenAnio", "OrigenMes", valcol, "DelPeríodo", "OtroPeríodo"])
        n = len(tab)
        for _, r in tab.iterrows():
            ws.append([r["Mes"], r["Canal"], r["KAM"], r["OrigenAnio"], r["OrigenMes"], r[valcol], None, None])
        for i in range(2, n + 2):
            ws[f"G{i}"] = (f'=IF({PL}="YTD",IF(D{i}=2026,F{i},0),'
                          f'IF(AND(A{i}={PL},D{i}=2026,E{i}={PL}),F{i},0))')
            ws[f"H{i}"] = (f'=IF({PL}="YTD",IF(D{i}=2025,F{i},0),'
                          f'IF(AND(A{i}={PL},NOT(AND(D{i}=2026,E{i}={PL}))),F{i},0))')
        return n

    wsn = wb.create_sheet("NC_detalle"); nN = _detalle(wsn, nc_tab, "Neto")
    wsc = wb.create_sheet("Comisiones_detalle"); nC = _detalle(wsc, com_tab, "Monto")

    # Aporte del canal (reclasificación venta→menor comisión) por glosa
    wsa = wb.create_sheet("Aporte_detalle")
    wsa.append(["Mes", "Canal", "KAM", "Monto"])
    aporte_tab = bundle.get("aporte_tab")
    nA = 0
    if aporte_tab is not None and len(aporte_tab):
        for _, r in aporte_tab.iterrows():
            wsa.append([r["Mes"], r["Canal"], r["KAM"], r["Monto"]]); nA += 1

    wsl = wb.create_sheet("Listas")
    wsl.append(["Meses", "Canales", "KAMs"])
    canal_opt = ["TODOS"] + canales; kam_opt = ["TODOS"] + kams
    for i in range(max(len(MESES_OPT), len(canal_opt), len(kam_opt))):
        wsl.append([MESES_OPT[i] if i < len(MESES_OPT) else None,
                    canal_opt[i] if i < len(canal_opt) else None,
                    kam_opt[i] if i < len(kam_opt) else None])

    ws = wb.create_sheet("P&L", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "P&L Comercial vs Contable — UnionX H1 2026 (MP/Fidelización/Páginas Propias + UnionX B2B)"
    ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
    ws.merge_cells("A1:E1")
    ws["A3"], ws["A4"], ws["A5"] = "Mes:", "Canal:", "KAM:"
    for c in ("A3", "A4", "A5"):
        ws[c].font = Font(bold=True)
    ws["B3"], ws["B4"], ws["B5"] = "YTD", "TODOS", "TODOS"
    for c in ("B3", "B4", "B5"):
        ws[c].fill = AMBAR; ws[c].font = Font(bold=True)
        ws[c].alignment = Alignment(horizontal="center"); ws[c].border = BORDER
    ws["G3"] = '=IF(B3="YTD","*",B3)'; ws["G4"] = '=IF(B4="TODOS","*",B4)'; ws["G5"] = '=IF(B5="TODOS","*",B5)'
    ws["F3"] = "(filtro→)"
    for c in ("F3", "G3", "G4", "G5"):
        ws[c].font = Font(size=8, color="94A3B8")
    dvm = DataValidation(type="list", formula1=f"=Listas!$A$2:$A${1+len(MESES_OPT)}", allow_blank=False)
    dvc = DataValidation(type="list", formula1=f"=Listas!$B$2:$B${1+len(canal_opt)}", allow_blank=False)
    dvk = DataValidation(type="list", formula1=f"=Listas!$C$2:$C${1+len(kam_opt)}", allow_blank=False)
    for dv, cell in ((dvm, "B3"), (dvc, "B4"), (dvk, "B5")):
        ws.add_data_validation(dv); dv.add(ws[cell])

    sumifs = lambda comp: (f'=SUMIFS({drng(comp)},{drng("Mes")},$G$3,'
                           f'{drng("Canal")},$G$4,{drng("KAM")},$G$5)')
    hdr = 7
    for col, txt in zip("ABCDE", ["Línea", "Comercial", "Contable", "Δ $ (Com−Cont)", "Δ %"]):
        cell = ws[f"{col}{hdr}"]; cell.value, cell.fill = txt, AZUL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center" if col != "A" else "left"); cell.border = BORDER
    rr = [hdr + 1]; R = {}

    def put(label, b, c, subtotal=False, resta=False):
        i = rr[0]
        ws[f"A{i}"] = ("(−) " if resta else ("= " if subtotal else "")) + label
        ws[f"B{i}"], ws[f"C{i}"] = b, c
        ws[f"D{i}"] = f"=B{i}-C{i}"; ws[f"E{i}"] = f'=IF(C{i}=0,"",D{i}/ABS(C{i}))'
        for col in "BCD":
            ws[f"{col}{i}"].number_format = MONEY
        ws[f"E{i}"].number_format = PCT
        for col in "ABCDE":
            ws[f"{col}{i}"].border = BORDER
            if subtotal:
                ws[f"{col}{i}"].font = Font(bold=True); ws[f"{col}{i}"].fill = GRIS
        R[label] = i; rr[0] += 1

    put("Venta", sumifs("Venta_Com"), sumifs("Venta_Cont"))
    put("Costo de Venta", sumifs("Costo_Com"), sumifs("Costo_Cont"), resta=True)
    put("Margen Directo", f"=B{R['Venta']}-B{R['Costo de Venta']}", f"=C{R['Venta']}-C{R['Costo de Venta']}", subtotal=True)
    put("Comisión Venta", sumifs("ComVenta_Com"), sumifs("ComVenta_Cont"), resta=True)
    put("Comisión Envío", sumifs("ComEnvio_Com"), sumifs("ComEnvio_Cont"), resta=True)
    put("Marketing", sumifs("Mkt_Com"), sumifs("Mkt_Cont"), resta=True)
    put("Contribución",
        f"=B{R['Margen Directo']}-B{R['Comisión Venta']}-B{R['Comisión Envío']}-B{R['Marketing']}",
        f"=C{R['Margen Directo']}-C{R['Comisión Venta']}-C{R['Comisión Envío']}-C{R['Marketing']}", subtotal=True)

    rV, rM, rCV, rCE, rMK, rK = (R["Venta"], R["Margen Directo"], R["Comisión Venta"],
                                 R["Comisión Envío"], R["Marketing"], R["Contribución"])
    ncG = lambda c: f"NC_detalle!${c}$2:${c}${nN+1}"
    coG = lambda c: f"Comisiones_detalle!${c}$2:${c}${nC+1}"
    nc_del = f'=SUMIFS({ncG("G")},{ncG("B")},$G$4,{ncG("C")},$G$5)'
    nc_otro = f'=SUMIFS({ncG("H")},{ncG("B")},$G$4,{ncG("C")},$G$5)'
    com_otro = f'=SUMIFS({coG("H")},{coG("B")},$G$4,{coG("C")},$G$5)'
    apG = lambda c: f"Aporte_detalle!${c}$2:${c}${nA+1}"
    aporte_body = (f'SUMIFS({apG("D")},{apG("A")},$G$3,{apG("B")},$G$4,{apG("C")},$G$5)' if nA else "0")
    mcom = f"IF(B{rV}=0,0,B{rM}/B{rV})"; mcont = f"IF(C{rV}=0,0,C{rM}/C{rV})"

    er = rK + 2
    ws[f"A{er}"] = "EXPLICACIÓN — RECONCILIACIÓN PASO A PASO (Comercial → ajustes → Contable)"
    ws[f"A{er}"].font = Font(bold=True, color="1E40AF"); ws.merge_cells(f"A{er}:E{er}")
    rr2 = [er + 1]

    def line(label, expr=None, bold=False, fill=None, sub=False, italic=False):
        i = rr2[0]
        ws[f"A{i}"] = ("      • " if sub else "") + label
        if expr is not None:
            ws[f"D{i}"] = expr; ws[f"D{i}"].number_format = MONEY
        fnt = dict(bold=True) if bold else (dict(size=10, color="475569") if sub else
              (dict(italic=True, size=9, color="64748B") if italic else {}))
        if fnt:
            ws[f"A{i}"].font = Font(**fnt)
            if expr is not None:
                ws[f"D{i}"].font = Font(**fnt)
        if fill:
            ws[f"A{i}"].fill = fill
            if expr is not None:
                ws[f"D{i}"].fill = fill
        rr2[0] += 1
        return i

    line("Δ Contribución (Comercial − Contable)", f"=D{rK}", bold=True, fill=VERDE)
    line("① VENTA", bold=True)
    line("Venta comercial", f"=B{rV}")
    r_nco = line("(−) Devoluciones NC de otro período (mes→otros+2025 · YTD→2025)", nc_otro, sub=True)
    r_vaj = line("= Venta ajustada (descontadas NC de otro período)", f"=B{rV}-D{r_nco}", bold=True)
    line("(memo) Devoluciones NC del período", nc_del, sub=True)
    line("② MARGEN DIRECTO sobre la venta ajustada", bold=True)
    line("Margen a tasa COMERCIAL (% comercial × venta ajustada)", f"={mcom}*D{r_vaj}")
    line("(−) por diferencia de % de margen (costeo) × venta ajustada", f"=({mcom}-{mcont})*D{r_vaj}", sub=True)
    r_maj = line("= Margen directo ajustado (venta ajustada × % margen contable)", f"={mcont}*D{r_vaj}", bold=True)
    line("(comparar) Margen Directo Contable real del P&L", f"=C{rM}", italic=True)
    line("③ COMISIONES (timing)", bold=True)
    line("Comisiones de otro período (glosas, incl. 2025)", com_otro, sub=True)
    r_noc = line("Comisiones por caer (no caída: comercial esperada − contable)", f"=B{rCV}-C{rCV}", sub=True)
    r_ap = line("del cual: aporte del canal (Oportunidades Únicas/Falabella, Opex/Ripley)", f"=-({aporte_body})", italic=True)
    line("del cual: provisión / aún por caer (resto)", f"=D{r_noc}-D{r_ap}", italic=True)
    line("④ CONTRIBUCIÓN", bold=True)
    r_caj = line("Contribución ajustada (margen ajustado − comisiones comerciales)",
                 f"=D{r_maj}-(B{rCV}+B{rCE}+B{rMK})", bold=True)
    line("Contribución contable (real del P&L)", f"=C{rK}")
    line("= Diferencia por EXPLICAR (ajustada − contable)", f"=D{r_caj}-C{rK}", bold=True, fill=AMBAR)

    ws.column_dimensions["A"].width = 50
    for c in "BCD":
        ws.column_dimensions[c].width = 17
    ws.column_dimensions["E"].width = 9
    ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 10
    for wsx in (wsd, wsn, wsc, wsa):
        wsx.freeze_panes = "A2"; wsx.auto_filter.ref = wsx.dimensions
        for cell in wsx[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = AZUL
        for cc in wsx.iter_cols(min_row=2):
            vals = [c.value for c in cc if isinstance(c.value, (int, float))]
            if vals and cc[0].column_letter in ("F", "G", "H"):
                for cell in cc:
                    cell.number_format = MONEY
                wsx.column_dimensions[cc[0].column_letter].width = 14
        wsx.column_dimensions["A"].width = 7
        wsx.column_dimensions["B"].width = 24
        wsx.column_dimensions["C"].width = 16
    return wb

