"""Descarga RAW de ventas (Excel/CSV.gz/Parquet, 40 columnas).

Streaming-first: para rangos grandes evita to_excel() (que materializa todo
en RAM y mata Streamlit Cloud). Excel solo se ofrece para rangos <= 90 dias.
"""
import gc
import io
from datetime import datetime, date

import pandas as pd
import streamlit as st

from views.shared import get_service

EXCEL_LIMITE_DIAS = 90
EXCEL_AVISO_FILAS = 100_000

# Columnas de ID largo (16+ dígitos). Excel trunca a 15 dígitos significativos si
# las trata como número → el último dígito pasa a 0. Se fuerzan a formato TEXTO.
COLS_ID_TEXTO = ("Pedido", "Documento", "Marketplace Reference", "Yuju Pack Id")


def _cargar_reconciliacion():
    """Candidatos del reconciliador (ventas Odoo detectadas fuera del RAW). Opcional:
    si no existe el parquet, no agrega pestañas. NO afecta la hoja RAW."""
    from pathlib import Path
    p = Path(__file__).resolve().parent.parent / "data" / "reconciliacion" / "candidatos.parquet"
    if not p.exists():
        return None
    try:
        return pd.read_parquet(p)
    except Exception:
        return None


def _excel_bytes(df: pd.DataFrame) -> bytes:
    """Solo para df pequenios. Materializa en RAM. Las columnas de ID largo se
    escriben como TEXTO (formato '@') para que Excel no las trunque a 15 dígitos.
    Agrega pestañas 'Reconciliación' + 'Resumen Reconciliación' si hay candidatos
    (sin tocar la hoja RAW ni su formato)."""
    from openpyxl.utils import get_column_letter
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as w:
        df.to_excel(w, index=False, sheet_name='RAW')
        ws = w.sheets['RAW']
        for j, col in enumerate(df.columns, start=1):
            if col in COLS_ID_TEXTO:
                letter = get_column_letter(j)
                for cell in ws[letter][1:]:  # saltar encabezado
                    cell.number_format = '@'
        # Pestañas nuevas del reconciliador (no modifican la hoja RAW).
        # ANEXO en dry-run: ventas de Odoo detectadas que NO están en el RAW
        # (cruce de meses / fulfillment). NO son parte del RAW; son para revisión.
        rec = _cargar_reconciliacion()
        if rec is not None and len(rec):
            rec = rec.copy()
            # 'periodo' es la VENTANA que barrió el reconciliador (no una fecha de venta).
            # Renombrar para que no se confunda con la fecha del documento.
            if 'periodo' in rec.columns:
                rec = rec.rename(columns={'periodo': 'Ventana detección (Odoo)'})
            rec.to_excel(w, index=False, sheet_name='Reconciliación')
            resumen = (rec.groupby(['canal', 'estado'], as_index=False)
                       .agg(Lineas=('venta_neta', 'size'), Venta_neta=('venta_neta', 'sum')))
            resumen.to_excel(w, index=False, sheet_name='Resumen Reconciliación')
            # IDs como texto también en la pestaña Reconciliación
            wsr = w.sheets['Reconciliación']
            for j, col in enumerate(rec.columns, start=1):
                if col in ('pedido', 'pedido_marketplace', 'sku'):
                    letter = get_column_letter(j)
                    for cell in wsr[letter][1:]:
                        cell.number_format = '@'
    return output.getvalue()


def _csv_gz_streamed(df: pd.DataFrame) -> bytes:
    """CSV gzip comprimido on-the-fly. ~5x mas chico que Excel."""
    buf = io.BytesIO()
    df.to_csv(buf, index=False, compression='gzip')
    return buf.getvalue()


def _parquet_bytes(df: pd.DataFrame) -> bytes:
    """Parquet zstd. ~20x mas chico, requiere herramientas para abrir."""
    buf = io.BytesIO()
    df.to_parquet(buf, index=False, compression='zstd', compression_level=3)
    return buf.getvalue()


def render():
    st.title("⬇️ Descargar RAW de Ventas")
    st.caption("Excel / CSV.gz / Parquet — 40 columnas RAW (formato histórico)")

    col_d1, col_d2 = st.columns([2, 1])
    with col_d1:
        rango_dl = st.date_input(
            "Período a descargar",
            value=(datetime.now().date().replace(day=1), datetime.now().date()),
            max_value=datetime.now().date(),
            format="YYYY-MM-DD",
            key="rango_dl",
        )

    if not (isinstance(rango_dl, tuple) and len(rango_dl) == 2):
        st.info("Selecciona un rango (desde — hasta).")
        return

    d1, d2 = rango_dl
    dias = (d2 - d1).days + 1

    # Filtro opcional por canal. Vacío = todos los canales.
    try:
        canales_disp = get_service().listar_canales()
    except Exception:
        canales_disp = []
    canales_sel = st.multiselect(
        "Canales (opcional — vacío = todos)",
        options=canales_disp,
        default=[],
        key="canales_dl",
        help="Filtra la descarga a uno o más canales. Si lo dejas vacío, baja todos.",
    )

    # Formato segun tamano
    excel_permitido = dias <= EXCEL_LIMITE_DIAS
    formatos = []
    if excel_permitido:
        formatos.append("Excel (.xlsx)")
    formatos.extend(["CSV comprimido (.csv.gz)", "Parquet (.parquet)"])

    with col_d2:
        st.write("")
        fmt = st.selectbox("Formato", formatos, key="fmt_dl")

    if fmt.startswith("CSV"):
        st.caption("⚠️ Los IDs largos (Marketplace Reference / Yuju Pack Id) se "
                   "conservan completos en el archivo, pero **Excel los trunca al abrir "
                   "un CSV**. Para verlos enteros usa **Excel (.xlsx)** o **Parquet**.")

    # Aviso si rango grande
    if dias > EXCEL_LIMITE_DIAS:
        st.warning(
            f"Rango de **{dias} días** > {EXCEL_LIMITE_DIAS} días. "
            f"Excel no disponible (rompe la app por uso de RAM). "
            f"Usa **CSV.gz** (se abre con Excel y pesa 5x menos) o **Parquet** (20x menos)."
        )
    elif dias > 30:
        st.info(f"Rango de **{dias} días**. Excel funciona pero CSV.gz es más rápido y liviano.")

    if not st.button("📥 Generar descarga", width='stretch', type="primary"):
        return

    with st.spinner(f'Consultando {dias} días de ventas...'):
        df_raw = get_service().descargar_raw(d1.strftime('%Y-%m-%d'), d2.strftime('%Y-%m-%d'))

    if canales_sel and 'Canal' in df_raw.columns:
        df_raw = df_raw[df_raw['Canal'].isin(canales_sel)].copy()

    n = len(df_raw)
    if n == 0:
        aviso = " para los canales seleccionados" if canales_sel else ""
        st.warning(f"Sin datos en el período{aviso}.")
        return

    if canales_sel:
        st.caption(f"Filtrado a {len(canales_sel)} canal(es): {', '.join(canales_sel)}")

    if n > EXCEL_AVISO_FILAS and fmt.startswith("Excel"):
        st.error(
            f"❌ {n:,} filas es demasiado para Excel en Streamlit Cloud. "
            f"Cambia a CSV.gz o Parquet."
        )
        return

    # Sufijo de canal para el nombre de archivo (1 canal → nombre; varios → "Ncanales").
    if len(canales_sel) == 1:
        slug = "".join(ch if ch.isalnum() else "-" for ch in canales_sel[0]).strip("-")
        canal_sfx = f"_{slug}"
    elif len(canales_sel) > 1:
        canal_sfx = f"_{len(canales_sel)}canales"
    else:
        canal_sfx = ""

    with st.spinner(f'Empaquetando {n:,} filas como {fmt}...'):
        try:
            if fmt.startswith("Excel"):
                data = _excel_bytes(df_raw)
                fname = f"Raw_ventas_{d1}_{d2}{canal_sfx}.xlsx"
                mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif fmt.startswith("CSV"):
                data = _csv_gz_streamed(df_raw)
                fname = f"Raw_ventas_{d1}_{d2}{canal_sfx}.csv.gz"
                mime = 'application/gzip'
            else:
                data = _parquet_bytes(df_raw)
                fname = f"Raw_ventas_{d1}_{d2}{canal_sfx}.parquet"
                mime = 'application/octet-stream'
        except MemoryError:
            st.error(
                "❌ Sin memoria suficiente. Reduce el rango o usa Parquet (20x más liviano)."
            )
            return
        finally:
            del df_raw
            gc.collect()

    size_mb = len(data) / (1024 * 1024)
    st.success(f"✅ {n:,} filas listas — {size_mb:.1f} MB")
    st.download_button(
        label=f"💾 Descargar {fmt} ({size_mb:.1f} MB)",
        data=data,
        file_name=fname,
        mime=mime,
        width='stretch',
    )
