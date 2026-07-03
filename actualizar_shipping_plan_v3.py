"""V3: CF limpio + fórmulas INDEX/MATCH desde tabla Refs_Embarques + amarillos a NB-01."""
import pandas as pd, sys, re
from collections import defaultdict
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

P = "data/planillas/Shipping Plan September V.02_actualizado.xlsx"
SV = "data/comex/Shipping Plan B-Jun.30th.xls"

TRANSITO = {'SZ': 52, 'NB': 45}
CRD_ETD = 7
PTO_BOD = 7

AZUL_H = PatternFill('solid', fgColor='FF1F3864')
ROJO   = PatternFill('solid', fgColor='FFE6B8B8')
AZUL_S = PatternFill('solid', fgColor='FFBDD7EE')
VERDE  = PatternFill('solid', fgColor='FFC6EFCE')
GRIS   = PatternFill('solid', fgColor='FFF2F2F2')
NULL_FILL = PatternFill(fill_type=None)
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
NORMAL = Font(size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Parse Steven
def parsear_steven(sheet):
    df = pd.read_excel(SV, sheet_name=sheet, header=None)
    items, seccion, hdr = [], None, None
    for _, row in df.iterrows():
        vals = [v for v in row.values if pd.notna(v)]
        if not vals: continue
        first = str(vals[0]).strip()
        m = re.match(r'^\d+\.(.+)$', first)
        if m and len(vals) <= 2: seccion = m.group(1).strip(); continue
        if first == 'No' and len(vals)>1 and 'Model' in str(vals[1]):
            hdr = [str(v).strip() for v in row.values]; continue
        try: int(first)
        except: continue
        if hdr is None or seccion is None: continue
        it = {col: row.values[j] for j,col in enumerate(hdr) if j<len(row.values)}
        it['__seccion'] = seccion
        items.append(it)
    return items

steven = parsear_steven('SHENZHEN') + parsear_steven('NINGBO')
emb_crd = {}
for it in steven:
    finish = it.get('Finish Time')
    if isinstance(finish, datetime):
        e = it['__seccion']
        if e not in emb_crd or finish > emb_crd[e]:
            emb_crd[e] = finish

# Embarques en orden (NB-01 tiene CRD = max de items existentes en Steven)
EMBARQUES = [
    ('SZ-04,40HQ',           'SZ'),
    ('SZ-03-08,40HQ',        'SZ'),
    ('SZ-01(623)',           'SZ'),
    ('SZ-02(623)',           'SZ'),
    ('IMP OP 350-26 40HQ',   'NB'),
    ('NB-01',                'NB'),
    ('Rest items',           'SZ'),  # asume Shenzhen para fechas
]

# --- Open workbook
wb = openpyxl.load_workbook(P)

# === 1) Hoja Refs_Embarques (creada o reescrita)
if 'Refs_Embarques' in wb.sheetnames:
    del wb['Refs_Embarques']
ws_ref = wb.create_sheet('Refs_Embarques')
ws_ref.sheet_view.showGridLines = False

# Headers
ws_ref['A1'] = 'TABLA DE REFERENCIA — Embarques Steven Plan B'
ws_ref.merge_cells('A1:H1')
ws_ref['A1'].font = Font(bold=True, size=12, color='FF1F3864')

ws_ref['A2'] = 'Modificar CRD o Tránsito acá → toda la planilla se recalcula.'
ws_ref['A2'].font = Font(italic=True, size=10, color='FF7F7F7F')
ws_ref.merge_cells('A2:H2')

hdrs = ['Embarque', 'Origen', 'Tránsito (días)', 'CRD→ETD (días)', 'ETA Pto→Bod (días)', 'CRD', 'ETD', 'ETA Puerto', 'ETA Bodega']
for j, h in enumerate(hdrs):
    c = ws_ref.cell(row=4, column=j+1, value=h)
    c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

for i, (e, origen) in enumerate(EMBARQUES):
    r = 5 + i
    ws_ref.cell(row=r, column=1, value=e)
    ws_ref.cell(row=r, column=2, value=origen)
    ws_ref.cell(row=r, column=3, value=TRANSITO[origen])
    ws_ref.cell(row=r, column=4, value=CRD_ETD)
    ws_ref.cell(row=r, column=5, value=PTO_BOD)
    crd = emb_crd.get(e)
    ws_ref.cell(row=r, column=6, value=crd)
    # ETD = CRD + col D, ETA Pto = ETD + col C, ETA Bod = ETA Pto + col E
    ws_ref.cell(row=r, column=7, value=f'=F{r}+D{r}')
    ws_ref.cell(row=r, column=8, value=f'=G{r}+C{r}')
    ws_ref.cell(row=r, column=9, value=f'=H{r}+E{r}')
    for col in range(1, 10):
        cc = ws_ref.cell(row=r, column=col)
        cc.font = NORMAL; cc.border = BORDER
        if col >= 6:
            cc.number_format = 'DD-MM-YYYY'; cc.alignment = CENTER
        elif col in (3,4,5):
            cc.alignment = CENTER
        elif col == 2:
            cc.alignment = CENTER
        else:
            cc.alignment = LEFT

# Anchos
for col, w in [('A',22),('B',8),('C',15),('D',15),('E',18),('F',12),('G',12),('H',12),('I',12)]:
    ws_ref.column_dimensions[col].width = w

last_ref_row = 4 + len(EMBARQUES)
RANGO_EMB = f"Refs_Embarques!$A$5:$A${last_ref_row}"
RANGO_CRD = f"Refs_Embarques!$F$5:$F${last_ref_row}"
RANGO_ETD = f"Refs_Embarques!$G$5:$G${last_ref_row}"
RANGO_ETP = f"Refs_Embarques!$H$5:$H${last_ref_row}"
RANGO_ETB = f"Refs_Embarques!$I$5:$I${last_ref_row}"

def formula_fecha(emb_cell, rango_target):
    return f'=IFERROR(INDEX({rango_target},MATCH({emb_cell},{RANGO_EMB},0)),"")'

# === 2) Detail_SZ + Detail_NB: limpiar fills + fórmulas en O/P/Q/R + CF en M
print("=== Detail_SZ + Detail_NB: fórmulas indexadas + CF ===")
for sh in ['Detail_SZ', 'Detail_NB']:
    ws = wb[sh]
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    col_units = headers.get('Units')        # H = 8
    col_cant  = headers.get('Cantidad cargada')   # M = 13
    col_emb   = headers.get('Embarque (Steven)')  # N = 14
    col_crd   = headers.get('CRD')          # O? — en estructura previa era O=15
    col_etd   = headers.get('ETD')
    col_etp   = headers.get('ETA Puerto')
    col_etb   = headers.get('ETA Bodega')
    print(f"  [{sh}] cols: units={col_units} cant={col_cant} emb={col_emb} CRD={col_crd} ETD={col_etd} ETP={col_etp} ETB={col_etb}")

    last = ws.max_row
    cantL = get_column_letter(col_cant)
    embL = get_column_letter(col_emb)
    unitsL = get_column_letter(col_units)

    # LIMPIAR fills duros en cant + en filas de datos (col M)
    for r in range(2, last+1):
        cell = ws.cell(row=r, column=col_cant)
        cell.fill = NULL_FILL
        cell.number_format = '#,##0'
        cell.alignment = RIGHT
        # Reemplazar fechas por fórmulas indexadas si hay embarque
        emb_val = ws.cell(row=r, column=col_emb).value
        if emb_val:
            for col_idx, rango in [(col_crd, RANGO_CRD),(col_etd, RANGO_ETD),(col_etp, RANGO_ETP),(col_etb, RANGO_ETB)]:
                cc = ws.cell(row=r, column=col_idx, value=formula_fecha(f'{embL}{r}', rango))
                cc.font = NORMAL; cc.border = BORDER; cc.alignment = CENTER
                cc.number_format = 'DD-MM-YYYY'
        else:
            for col_idx in (col_crd, col_etd, col_etp, col_etb):
                cc = ws.cell(row=r, column=col_idx)
                cc.value = None
                cc.fill = NULL_FILL

    # Eliminar CF previo en col M
    new_rules = {}
    for k, v in dict(ws.conditional_formatting._cf_rules).items():
        if k.sqref != f'{cantL}2:{cantL}{last}':
            new_rules[k] = v
    ws.conditional_formatting._cf_rules = new_rules

    # Re-aplicar CF
    rng = f'{cantL}2:{cantL}{last}'
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'AND(ISNUMBER({cantL}2),{cantL}2={unitsL}2)'], fill=VERDE, stopIfTrue=True))
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'AND(ISNUMBER({cantL}2),{cantL}2>{unitsL}2)'], fill=AZUL_S, stopIfTrue=True))
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'AND(ISNUMBER({cantL}2),{cantL}2<{unitsL}2,{cantL}2>0)'], fill=ROJO, stopIfTrue=True))
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'OR({cantL}2="",ISBLANK({cantL}2))'], fill=GRIS, stopIfTrue=True))
    print(f"    CF re-aplicado a {rng} (4 reglas)")

# === 3) Late_Arrivals + Pend_Confirmation + Missing: re-aplicar fechas como fórmulas
print("\n=== Otras hojas: re-aplicar fórmulas indexadas ===")
for sh, units_col_letter in [('Late_Arrivals','E'), ('Pend_Confirmation','D'), ('Missing','D')]:
    ws = wb[sh]
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    col_cant = headers.get('Cantidad cargada')
    col_emb  = headers.get('Embarque (Steven)')
    col_crd  = headers.get('CRD')
    col_etd  = headers.get('ETD')
    col_etp  = headers.get('ETA Puerto')
    col_etb  = headers.get('ETA Bodega')
    if not col_emb:
        print(f"  [{sh}] no tiene Embarque (Steven), skip"); continue
    embL = get_column_letter(col_emb)
    cantL = get_column_letter(col_cant)
    last = ws.max_row

    for r in range(2, last+1):
        emb_val = ws.cell(row=r, column=col_emb).value
        # limpiar fill duro en cant
        ws.cell(row=r, column=col_cant).fill = NULL_FILL
        ws.cell(row=r, column=col_cant).number_format = '#,##0'
        if emb_val:
            for col_idx, rango in [(col_crd, RANGO_CRD),(col_etd, RANGO_ETD),(col_etp, RANGO_ETP),(col_etb, RANGO_ETB)]:
                cc = ws.cell(row=r, column=col_idx, value=formula_fecha(f'{embL}{r}', rango))
                cc.font = NORMAL; cc.border = BORDER; cc.alignment = CENTER
                cc.number_format = 'DD-MM-YYYY'
        else:
            for col_idx in (col_crd, col_etd, col_etp, col_etb):
                cc = ws.cell(row=r, column=col_idx)
                cc.value = None; cc.fill = NULL_FILL

    # Re-aplicar CF en Cantidad cargada vs Units (col D o E según hoja)
    new_rules = {}
    for k, v in dict(ws.conditional_formatting._cf_rules).items():
        if k.sqref != f'{cantL}2:{cantL}{last}':
            new_rules[k] = v
    ws.conditional_formatting._cf_rules = new_rules
    u = units_col_letter
    rng = f'{cantL}2:{cantL}{last}'
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({cantL}2),{cantL}2={u}2)'], fill=VERDE, stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({cantL}2),{cantL}2>{u}2)'], fill=AZUL_S, stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({cantL}2),{cantL}2<{u}2,{cantL}2>0)'], fill=ROJO, stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR({cantL}2="",ISBLANK({cantL}2))'], fill=GRIS, stopIfTrue=True))
    print(f"  [{sh}] CF + fechas indexadas listas")

# === 4) Missing: mover los 3 amarillos LHSHNUCO de Rest items → NB-01
print("\n=== Missing: mover amarillos a NB-01 ===")
ws = wb['Missing']
col_emb = None; col_cant = None
for c in range(1, ws.max_column+1):
    if ws.cell(row=1, column=c).value == 'Embarque (Steven)': col_emb = c
    if ws.cell(row=1, column=c).value == 'Cantidad cargada': col_cant = c
movidos = 0
for r in range(2, ws.max_row+1):
    sku = str(ws.cell(row=r, column=2).value or '').strip()
    if sku.startswith('LHSHNUCO-'):
        ws.cell(row=r, column=col_emb).value = 'NB-01'
        movidos += 1
        print(f"  R{r} {sku} → NB-01")
print(f"  Total movidos: {movidos}")

# === Guardar
wb.save(P)
print(f"\n[OK] Guardado: {P}")
