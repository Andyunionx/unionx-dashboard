# -*- coding: utf-8 -*-
"""Recupera al RAW las ventas Kitchen/abc(La Polar) que se cayeron por cruce de meses
(factura/orden creada después del freeze). Fuente: Odoo (sale.order confirmadas con
factura, client_order_ref en la lista de faltantes de Gabriela). Overlay al histórico.

Uso: python recuperar_kitchen_abc.py [--apply]
"""
import argparse, json, os, sys, shutil
from pathlib import Path
import pandas as pd
import xmlrpc.client

ROOT = Path(__file__).resolve().parent
HIST = ROOT / "data" / "historico" / "ventas_historico.parquet"
MES = ROOT / "data" / "historico" / "ventas_mes_actual.parquet"
MISS = Path(os.environ.get("MISS_JSON", "C:/Users/andre/AppData/Local/Temp/claude/g--Mi-unidad-TRABAJO-RESPALDO-OPERACIONES-UNION-X---IA/dff4caa3-b773-4c75-b3ac-1164eef810c4/scratchpad/miss_orders.json"))

DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


def _odoo():
    for p in (".env", "eerr-finanzas/.env"):
        f = ROOT / p
        if f.exists():
            for ln in f.read_text(encoding="utf-8").splitlines():
                if "=" in ln and not ln.startswith("#"):
                    k, v = ln.split("=", 1); os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
    URL, DB, U = "https://unionxb2b.odoo.com", "bmya-innovatek-sh-prd-6981800", "andres@grupoeter.cl"
    PWD = os.environ["ANDRES_ODOO_PASSWORD"]
    uid = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/common").authenticate(DB, U, PWD, {})
    return xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/object"), DB, uid, PWD


def _matriz():
    import openpyxl
    wb = openpyxl.load_workbook(ROOT / "data" / "planillas" / "Matriz productos.xlsx", read_only=True, data_only=True)
    ws = wb["Productos"] if "Productos" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True)); idx = {str(h).strip() if h else "": i for i, h in enumerate(rows[0])}
    iS = idx.get("SKU")
    m = {}
    for r in rows[1:]:
        if iS is None or len(r) <= iS or not r[iS]:
            continue
        g = lambda c: (r[idx[c]] if c in idx and idx[c] < len(r) and r[idx[c]] not in (None, "") else "")
        m[str(r[iS]).strip().lower()] = {"producto": g("Producto"), "categoria_macro": g("Categoría macro"),
            "categoria_padre": g("Categoría padre"), "categoria_hijo": g("Categoría hijo"),
            "categoria_comercial": g("Categoría comercial"), "marca": g("Marca"), "proveedor": g("Proveedor"),
            "pack": g("Pack"), "estado_sku": g("In/out")}
    return m


def main(apply=False):
    from clasificar_marca import clasificar_tipo_marca
    miss = json.load(open(MISS))
    # ref -> canal (para saber a qué canal pertenece cada orden; Walmart no se puede
    # inferir del name porque es S-number). Kitchen/abc usan 'CANAL número' (name);
    # Walmart usa el ref 13-díg (convención fulfillment del RAW).
    ref2canal = {}
    for x in miss["kitchen"]:
        ref2canal[str(x[0])] = "Kitchen Center"
    for r in miss["abc"]:
        ref2canal[str(r)] = "Abc"
    for r in miss.get("walmart", []):
        ref2canal[str(r)] = "Walmart"
    allref = list(ref2canal.keys())
    M, DB, uid, PWD = _odoo()
    so = M.execute_kw(DB, uid, PWD, "sale.order", "search_read",
        [[["client_order_ref", "in", allref], ["state", "in", ["sale", "done"]]]],
        {"fields": ["id", "name", "client_order_ref", "date_order", "invoice_ids", "user_id"]})
    print(f"[Odoo] {len(so)} órdenes 'sale' con factura")
    oids = [o["id"] for o in so]
    lines = M.execute_kw(DB, uid, PWD, "sale.order.line", "search_read",
        [[["order_id", "in", oids], ["product_uom_qty", ">", 0]]],
        {"fields": ["order_id", "product_id", "product_uom_qty", "price_total", "price_subtotal", "purchase_price"]})
    pids = list({l["product_id"][0] for l in lines if l.get("product_id")})
    prod = {p["id"]: p for p in M.execute_kw(DB, uid, PWD, "product.product", "read", [pids], {"fields": ["default_code", "name"]})}
    invids = list({i for o in so for i in o.get("invoice_ids", [])})
    inv = {i["id"]: i for i in M.execute_kw(DB, uid, PWD, "account.move", "read", [invids], {"fields": ["name", "invoice_date"]})}
    o_by_id = {o["id"]: o for o in so}
    matriz = _matriz()

    rows = []
    for l in lines:
        o = o_by_id[l["order_id"][0]]
        ref = str(o.get("client_order_ref") or "")
        canal = ref2canal.get(ref, "")
        if not canal:
            continue
        # pedido con formato estándar 'CANAL número' (= sale.order.name) para Kitchen/abc;
        # Walmart mantiene el ref 13-díg (convención fulfillment del RAW).
        pedido = str(o["name"]) if canal in ("Kitchen Center", "Abc") else ref
        fv = str(o["date_order"])[:10]; fdt = pd.to_datetime(fv)
        pid = l["product_id"][0] if l.get("product_id") else None
        sku = str((prod.get(pid) or {}).get("default_code") or "").strip()
        pnom = (prod.get(pid) or {}).get("name", "")
        inv0 = inv.get((o.get("invoice_ids") or [None])[0], {})
        mat = matriz.get(sku.lower(), {})
        qty = float(l["product_uom_qty"]); vb = float(l["price_total"]); vn = float(l["price_subtotal"])
        cu = float(l.get("purchase_price") or 0); ct = cu * qty
        marca = mat.get("marca", "")
        rows.append({
            "tipo_movimiento": "Venta", "bodega": "Bodega Carrascal N°9-10",
            "documento": inv0.get("name", ""), "fecha_documento": str(inv0.get("invoice_date") or fv),
            "pedido": pedido, "estado_pedido": "sale", "tipo_despacho": "",
            "sku": sku, "canal": canal, "fecha_venta": fv, "hora_venta": "",
            "producto": mat.get("producto") or (pnom.split("] ")[-1] if "]" in pnom else pnom),
            "categoria_macro": mat.get("categoria_macro", ""), "categoria_padre": mat.get("categoria_padre", ""),
            "categoria_hijo": mat.get("categoria_hijo", ""), "categoria_comercial": mat.get("categoria_comercial", ""),
            "estado_sku": str(mat.get("estado_sku", "")).lower(), "pack": mat.get("pack", "No"), "marca": marca,
            "proveedor": mat.get("proveedor", ""), "tipo_marca": clasificar_tipo_marca(marca),
            "tipo_compra": "Importación", "tipo_negocio": "Marketplace",
            "kam": (o.get("user_id") or [0, "Clau"])[1].split()[0] if o.get("user_id") else "Clau",
            "estado_canal": "In", "anio_venta": fdt.year, "mes_venta": fdt.month,
            "semana_venta": int(fdt.isocalendar().week), "dia_semana": DIAS[fdt.weekday()], "hora_venta_num": 0,
            "cantidad": qty, "venta_bruta": vb, "costo_unitario": cu, "costo_total": ct,
            "margen_front": vn - ct, "comision_pct": 0.0, "comision": 0.0, "logistica": 0.0,
            "marketing": 0.0, "margen_final": vn - ct, "venta_neta": vn,
        })
    nuevas = pd.DataFrame(rows)
    print(f"[build] {len(nuevas)} líneas · venta bruta ${nuevas['venta_bruta'].sum():,.0f}".replace(",", ".")
          + f" · neta ${nuevas['venta_neta'].sum():,.0f}".replace(",", "."))
    print("  por mes (fecha_venta=date_order Odoo):")
    print(nuevas.groupby(["mes_venta", "canal"]).agg(n=("sku", "size"), vb=("venta_bruta", "sum")).to_string())

    # dedup vs RAW existente (histórico + mes_actual) por (pedido, sku, round(vb))
    hist = pd.read_parquet(HIST); mes = pd.read_parquet(MES)
    def key(df):
        return set(zip(df["pedido"].astype(str), df["sku"].astype(str), df["venta_bruta"].round().astype("Int64").astype(str)))
    existe = key(hist) | key(mes)
    nuevas["_k"] = list(zip(nuevas["pedido"].astype(str), nuevas["sku"].astype(str), nuevas["venta_bruta"].round().astype("Int64").astype(str)))
    ya = nuevas["_k"].isin(existe)
    print(f"\n  ya en RAW (dedup): {ya.sum()} · nuevas a cargar: {(~ya).sum()}")
    add = nuevas[~ya].drop(columns="_k")
    # split por destino: mes<=6 -> histórico; mes==7 -> mes_actual
    to_hist = add[add["mes_venta"] <= 6].copy()
    to_mes = add[add["mes_venta"] >= 7].copy()
    print(f"  → histórico (ene-jun): {len(to_hist)} líneas ${to_hist['venta_bruta'].sum():,.0f}".replace(",", ".")
          + f" | → mes_actual (jul): {len(to_mes)} líneas ${to_mes['venta_bruta'].sum():,.0f}".replace(",", "."))

    if apply:
        # SOLO histórico (congelado, no se auto-sana). Julio queda para el extract diario
        # que regenera mes_actual (si le hago overlay, mañana lo pisa).
        base = pd.read_parquet(HIST)
        extra2 = to_hist.reindex(columns=base.columns)
        if "es_despacho" in base.columns:
            extra2["es_despacho"] = False
        bak = HIST.with_suffix(".parquet.bak_kitchenabc")
        if not bak.exists():
            shutil.copy2(HIST, bak)
        out = pd.concat([base, extra2], ignore_index=True)
        out.to_parquet(HIST, index=False)
        print(f"  ✓ HISTÓRICO: +{len(extra2)} filas (backup {bak.name})")
        print(f"  ⏭  julio ({len(to_mes)} líneas ${to_mes['venta_bruta'].sum():,.0f}".replace(",", ".")
              + ") NO se escribe: lo captura el extract diario en mes_actual.")
    else:
        print("\n  → dry-run (usa --apply para escribir)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(); ap.add_argument("--apply", action="store_true")
    main(apply=ap.parse_args().apply)
