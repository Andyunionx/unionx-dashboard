"""Carga una orden de compra (purchase.order) en Odoo desde el Pre-costeo de COMEX.

Lee `Pre-costeo_x_CBM_<EMB>.xlsx` (hoja Productos) y crea una PO en Odoo con:
  - partner_id = Shenzhen Topwill Electronic Co. Ltd (id 1664)
  - partner_ref = embarque (ej. 26TP0320)
  - currency = CLP
  - price_unit = Costo Internado Unit (CLP) por SKU
  - date_planned = ETA de la tarifa

Por defecto deja la PO en estado `draft` (RFQ). Con --confirm pasa a `purchase`.

Uso:
    python cargar_po_comex_odoo.py --precosteo agente-comex/data/output/26TP0320/Pre-costeo_x_CBM_26TP0320.xlsx
    python cargar_po_comex_odoo.py --precosteo <path> --eta 2026-05-22 --confirm
"""
import argparse
import os
import sys
import xmlrpc.client
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
PROJECT_ROOT = Path(__file__).parent
env = PROJECT_ROOT / '.env'
if env.exists():
    for line in env.read_text(encoding='utf-8').splitlines():
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

ODOO_URL = os.environ.get('ODOO_URL', 'https://unionxb2b.odoo.com')
ODOO_DB = os.environ.get('ODOO_DB', 'bmya-innovatek-sh-prd-6981800')
ODOO_USER = os.environ.get('ODOO_USER') or 'andres@grupoeter.cl'
ODOO_PWD = os.environ.get('ODOO_API_KEY') or os.environ.get('ANDRES_ODOO_PASSWORD')

PARTNER_STEVEN = 1664  # Shenzhen Topwill Electronic Co. Ltd
CURRENCY_CLP = 45
COMPANY_ID = 1
USER_ID = 8  # Andrés Browne
PICKING_TYPE = 1  # Bodega Carrascal N°9-10: Receipts


def conectar_odoo():
    common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common')
    uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PWD, {})
    if not uid:
        raise RuntimeError("No pude autenticar en Odoo")
    models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object')
    return uid, models


def leer_precosteo(path: Path):
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

    # Embarque desde Resumen
    embarque = None
    eta = None
    for row in wb['Resumen'].iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        if str(row[0]).lower() == 'embarque':
            embarque = str(row[1])
        elif str(row[0]).lower() == 'eta':
            eta = str(row[1])

    # Productos
    ws = wb['Productos']
    headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(headers)}

    # Patrones de gastos (red de seguridad para precosteos viejos generados antes
    # del fix del parser PI — ver costear_embarque.py:244). Match con word boundary
    # para evitar falsos positivos tipo 'ff' matcheando dentro de 'SIMCAFFRCO'.
    GASTOS_PATTERNS = [
        r'\bdelivery\b', r'\bmonitor\b', r'\bloading\b', r'\bff\b', r'\bform f\b',
        r'\blocal charge\b', r'\bvehicle\b', r'\bvechile\b',
        r'\bcleaning\b', r'\bcleaing\b', r'\bcustoms\b', r'\btransport\b',
        r'\bstorage\b', r'\bsamples\b', r'\bice bag\b', r'\bcooler samples\b',
    ]
    import re as _re
    gastos_re = _re.compile('|'.join(GASTOS_PATTERNS), _re.IGNORECASE)

    productos = []
    descartados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[idx.get('SKU', 1)] in (None, '', 'None'):
            continue
        sku_raw = str(row[idx['SKU']]).strip()
        model_raw = str(row[idx.get('Model', 0)] or '').strip()
        qty = float(row[idx['Qty']] or 0)
        if not sku_raw:
            continue

        # Solo aplicar filtro de gasto si el SKU es claramente sospechoso
        # (vacío, o nombre largo con palabras de gasto). Si el SKU tiene formato
        # de código real (corto, mayúsculas + guiones + dígitos), NO descartar.
        es_sku_sospechoso = (
            len(sku_raw) > 18 or ' ' in sku_raw or
            sku_raw.lower() in ('none', 'nan')
        )
        # Aplicar match solo a model+descripcion (no a SKU codificado).
        # Excepción: si SKU es texto largo/con espacios, también lo evaluamos.
        texto_eval = model_raw
        if es_sku_sospechoso:
            texto_eval = f"{model_raw} {sku_raw}"
        es_gasto = bool(gastos_re.search(texto_eval)) and qty <= 1
        if es_gasto:
            descartados.append(f"{model_raw or sku_raw}")
            continue
        productos.append({
            'sku': sku_raw,
            'model': row[idx.get('Model', 0)],
            'descripcion': row[idx.get('Descripcion', 2)],
            'qty': qty,
            'price_unit_clp': float(row[idx['Costo Internado Unit (CLP)']] or 0),
        })
    if descartados:
        print(f"  [filtro gastos] descartados {len(descartados)} líneas: {descartados[:5]}")
    return embarque, eta, productos


def buscar_productos(uid, models, skus: list[str]):
    """Devuelve dict {sku: product_id} buscando en product.product por default_code."""
    if not skus:
        return {}
    prods = models.execute_kw(
        ODOO_DB, uid, ODOO_PWD, 'product.product', 'search_read',
        [[('default_code', 'in', skus)]],
        {'fields': ['id', 'default_code', 'name', 'uom_po_id']}
    )
    mapping = {p['default_code']: p for p in prods}
    return mapping


def crear_po(uid, models, embarque: str, eta: str, productos: list[dict], productos_odoo: dict, confirmar: bool):
    # Líneas
    lineas = []
    no_encontrados = []
    total_clp = 0
    date_planned = eta + ' 12:00:00' if eta and len(eta) == 10 else (eta or datetime.now().strftime('%Y-%m-%d 12:00:00'))

    for p in productos:
        if p['sku'] not in productos_odoo:
            no_encontrados.append(p['sku'])
            continue
        odoo_prod = productos_odoo[p['sku']]
        # Nombre línea: model + descripción del PI (pero sin BR para evitar problemas)
        nombre_linea = (p['descripcion'] or odoo_prod['name'])[:200]
        lineas.append((0, 0, {
            'product_id': odoo_prod['id'],
            'name': nombre_linea,
            'product_qty': p['qty'],
            'price_unit': round(p['price_unit_clp'], 2),
            'date_planned': date_planned,
            'product_uom': odoo_prod['uom_po_id'][0] if odoo_prod.get('uom_po_id') else 1,
        }))
        total_clp += p['qty'] * p['price_unit_clp']

    if not lineas:
        raise RuntimeError("No quedaron líneas para crear PO (ningún SKU encontrado en Odoo)")

    print(f"\n[PO] Creando RFQ en Odoo...")
    print(f"   Embarque:   {embarque}")
    print(f"   Proveedor:  Shenzhen Topwill Electronic Co. Ltd (id {PARTNER_STEVEN})")
    print(f"   Moneda:     CLP")
    print(f"   Líneas:     {len(lineas)} (de {len(productos)} en precosteo)")
    print(f"   Total CLP:  ${total_clp:,.0f}")
    print(f"   ETA:        {date_planned}")

    po_vals = {
        'partner_id': PARTNER_STEVEN,
        'partner_ref': embarque,
        'currency_id': CURRENCY_CLP,
        'company_id': COMPANY_ID,
        'user_id': USER_ID,
        'picking_type_id': PICKING_TYPE,
        'date_planned': date_planned,
        'notes': f'<p>Embarque <b>{embarque}</b>. RFQ creado desde COMEX con costo internado CLP.</p>',
        'order_line': lineas,
    }
    po_id = models.execute_kw(ODOO_DB, uid, ODOO_PWD, 'purchase.order', 'create', [po_vals])
    print(f"   PO creada: id={po_id}")

    # Leer name asignado
    po = models.execute_kw(ODOO_DB, uid, ODOO_PWD, 'purchase.order', 'read', [po_id, ['name', 'state', 'amount_total']])[0]
    print(f"   Name:  {po['name']}")
    print(f"   State: {po['state']}")
    print(f"   Total Odoo: ${po['amount_total']:,.0f}")

    if confirmar:
        print(f"\n[CONFIRM] Confirmando PO a estado 'purchase'...")
        models.execute_kw(ODOO_DB, uid, ODOO_PWD, 'purchase.order', 'button_confirm', [[po_id]])
        po2 = models.execute_kw(ODOO_DB, uid, ODOO_PWD, 'purchase.order', 'read', [po_id, ['state']])[0]
        print(f"   Nuevo state: {po2['state']}")

    return po_id, po['name'], no_encontrados


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--precosteo', required=True, help='Path al Pre-costeo_x_CBM_<EMB>.xlsx')
    parser.add_argument('--eta', help='ETA bodega YYYY-MM-DD (override del Tarifas)')
    parser.add_argument('--confirm', action='store_true', help='Confirmar PO (state=purchase)')
    args = parser.parse_args()

    if not ODOO_PWD:
        print("[ERROR] Falta ANDRES_ODOO_PASSWORD/ODOO_API_KEY en .env")
        return 1

    precosteo_path = Path(args.precosteo)
    if not precosteo_path.exists():
        print(f"[ERROR] No existe: {precosteo_path}")
        return 1

    print(f"[1] Leyendo precosteo: {precosteo_path.name}")
    embarque, eta_tarifa, productos = leer_precosteo(precosteo_path)
    print(f"   Embarque: {embarque}")
    print(f"   ETA (tarifa): {eta_tarifa}")
    print(f"   Productos (sin samples): {len(productos)}")

    eta_final = args.eta or eta_tarifa

    print(f"\n[2] Conectando a Odoo...")
    uid, models = conectar_odoo()
    print(f"   Auth OK (uid={uid})")

    print(f"\n[3] Buscando productos en Odoo por SKU...")
    skus = [p['sku'] for p in productos]
    productos_odoo = buscar_productos(uid, models, skus)
    print(f"   Encontrados: {len(productos_odoo)} / {len(skus)}")

    # Verificar idempotencia: ¿ya existe una PO ACTIVA con este partner_ref?
    # Las canceladas (state=cancel) se ignoran para permitir recrear tras ajustes.
    existing = models.execute_kw(ODOO_DB, uid, ODOO_PWD, 'purchase.order', 'search_read',
        [[('partner_ref', '=', embarque)]],
        {'fields': ['id', 'name', 'state', 'amount_total']})
    activas = [po for po in existing if po['state'] != 'cancel']
    if activas:
        print(f"\n[!] Ya existe PO ACTIVA con partner_ref='{embarque}':")
        for po in activas:
            print(f"   id={po['id']} name={po['name']} state={po['state']} total=${po['amount_total']:,.0f}")
        print(f"   Aborto. Si quieres recrear, cancelá esa PO primero.")
        return 2
    if existing:
        print(f"\n[i] {len(existing)} PO(s) canceladas previas para {embarque}: "
              f"{', '.join(po['name'] for po in existing)}. Procedo a crear una nueva.")

    po_id, po_name, no_enc = crear_po(uid, models, embarque, eta_final, productos, productos_odoo, args.confirm)

    if no_enc:
        print(f"\n[!] SKUs NO encontrados en Odoo ({len(no_enc)}):")
        for sku in no_enc:
            print(f"   {sku}")
        # Mandar mail consolidado automático a Felipe + equipo
        try:
            from notificar_skus_faltantes import detectar_faltantes, enviar_mail
            emb_corto = embarque.replace('26TP', '')
            faltantes = detectar_faltantes([emb_corto])
            if faltantes:
                msg_id = enviar_mail(faltantes)
                print(f"   [mail] aviso a Felipe + equipo enviado | id={msg_id}")
        except Exception as e:
            print(f"   [WARN] mail SKUs faltantes falló: {type(e).__name__}: {e}")

    print(f"\n[OK] PO {po_name} (id={po_id}) creada en Odoo.")
    print(f"     URL: {ODOO_URL}/web#id={po_id}&model=purchase.order&view_type=form")
    return 0


if __name__ == '__main__':
    sys.exit(main() or 0)
