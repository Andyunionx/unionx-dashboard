"""Genera el archivo Maestra_Tarifa_Fija.xlsx con los valores FIJOS del costeo.

Se ejecuta UNA SOLA VEZ (o cuando cambien valores fijos como Flete Terrestre).
Output: C:\\Users\\andre\\OneDrive\\Documentos\\Claude\\Projects\\COMEX\\Maestra_Tarifa_Fija.xlsx

Después del costeo automático, `generar_tarifas_embarque.py` lee este archivo y
combina con: puerto (Steven), flete (forwarder), dólar (SII), ETA → produce el
Tarifas_Base_COMEX-XXXX.xlsx específico del embarque.
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

OUTPUT = Path('C:/Users/andre/OneDrive/Documentos/Claude/Projects/COMEX/Maestra_Tarifa_Fija.xlsx')

# ============================================================
# Valores fijos (extraídos del template del usuario)
# ============================================================
INLAND_CHILE_FIJOS = [
    ('Agente Aduana (%)',           0.0016,  '% del CIF en USD'),
    ('Gastos Puerto STI',           20000,   'CLP fijo'),
    ('Flete Terrestre',             415000,  'CLP fijo'),
    ('Seimex',                      238529,  'CLP fijo'),
    ('Desconsolidación Craft',      50000,   'CLP fijo'),
    ('Seguro Carga Contempora',     60000,   'CLP fijo'),
    ('Gastos Despacho',             43658,   'CLP fijo'),
    ('Gate In Maersk',              145000,  'CLP fijo'),
]

BENCHMARKS = [
    ('SZ (Shenzhen)',  '≤16%',  'Benchmark histórico 6m'),
    ('NB (Ningbo)',    '≤18%',  'Benchmark histórico 6m'),
    ('XI (Xiamen)',    '≤17%',  'Benchmark histórico 6m'),
    ('AIR (Aéreo)',    '≤35%',  'Mayor por urgencia'),
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tarifa Fija'

    # Estilos
    bold = Font(bold=True, size=11)
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    fixed_fill = PatternFill('solid', fgColor='E7E6E6')
    title_font = Font(bold=True, size=14, color='1F4E78')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

    row = 1
    ws.cell(row, 1, 'MAESTRA TARIFA FIJA — COMEX').font = title_font
    row += 1
    ws.cell(row, 1, 'Valores que NO cambian entre embarques. Lo variable se carga por embarque.').font = Font(italic=True, size=9)
    row += 2

    # Sección 1: Inland Chile
    ws.cell(row, 1, 'INLAND CHILE (CIF → Internado)').font = bold
    row += 1
    for c, h in enumerate(['Concepto', 'Valor', 'Unidad'], 1):
        cell = ws.cell(row, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    for concepto, valor, unidad in INLAND_CHILE_FIJOS:
        ws.cell(row, 1, concepto).border = border
        cv = ws.cell(row, 2, valor)
        cv.border = border
        cv.fill = fixed_fill
        cv.number_format = '#,##0.0000' if isinstance(valor, float) and valor < 1 else '#,##0'
        ws.cell(row, 3, unidad).border = border
        row += 1

    row += 1

    # Sección 2: Benchmarks
    ws.cell(row, 1, 'BENCHMARKS EFICIENCIA (referencia)').font = bold
    row += 1
    for c, h in enumerate(['Puerto', 'Umbral Máx Sobrecosto', 'Notas'], 1):
        cell = ws.cell(row, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    for puerto, umbral, notas in BENCHMARKS:
        ws.cell(row, 1, puerto).border = border
        cv = ws.cell(row, 2, umbral)
        cv.border = border
        cv.fill = fixed_fill
        ws.cell(row, 3, notas).border = border
        row += 1

    row += 1

    # Sección 3: Parámetros adicionales del flujo
    ws.cell(row, 1, 'PARÁMETROS COSTEO').font = bold
    row += 1
    extras = [
        ('Gift Box markup',         1.03,    'multiplicador (3% adicional)'),
        ('Comisión Steven',         0.03,    '% sobre P×Q + delivery + local charge + long vehicle'),
        ('Steven base incluye FF',  'NO',    'Form F NO entra en base Steven'),
        ('Capacidad 40HQ ref',      68,      'CBM (PL define el real)'),
    ]
    for c, h in enumerate(['Concepto', 'Valor', 'Notas'], 1):
        cell = ws.cell(row, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    for concepto, valor, notas in extras:
        ws.cell(row, 1, concepto).border = border
        cv = ws.cell(row, 2, valor)
        cv.border = border
        cv.fill = fixed_fill
        ws.cell(row, 3, notas).border = border
        row += 1

    # Auto-ancho
    for col_idx in range(1, 4):
        max_len = max((len(str(ws.cell(r, col_idx).value or '')) for r in range(1, row)), default=10)
        ws.column_dimensions[chr(64 + col_idx)].width = max(max_len + 2, 14)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))
    print(f"[OK] Generado: {OUTPUT}")
    print(f"      {len(INLAND_CHILE_FIJOS)} conceptos Inland Chile")
    print(f"      {len(BENCHMARKS)} benchmarks por puerto")
    print(f"      {len(extras)} parámetros adicionales")


if __name__ == '__main__':
    build()
