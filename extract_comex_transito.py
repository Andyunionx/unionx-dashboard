#!/usr/bin/env python3
"""
Extractor COMEX: ordenes en transito desde Google Sheet
'Importaciones UnionX Integrada' (owner: martin@unionx.cl).

Sheet ID: 1RpxZ69Wnfcots006Hp5fzawYxhUsscW03O_hD3psjHA
Pestañas relevantes:
  - 'Confirmado EN TRANSITO' (foco principal: ordenes en barco/aereo activas)
  - 'EN BODEGA' (historico, ya recibidas)

Columnas:
  F | SKU | Variante | PI | STATUS | Tipo Transporte | NRO PEDIDO | Cantidad |
  Costo Uni USD | GIFT BOX + Envio | COSTO INGRESO CLP CHILE |
  Fecha Embarque | Fecha ETA CHILE | Fecha ETA bodega | ODOO

Output:
- data/comex/transito.parquet (limpia, normalizada)
- data/comex/transito_resumen.json (KPIs)

Requiere: el sheet compartido en lectura con la service account
union-x-revenue-bot@union-x-revenue.iam.gserviceaccount.com
"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

PROJECT_ROOT = Path(__file__).parent
OUT_DIR = PROJECT_ROOT / 'data' / 'comex'
OUT_DIR.mkdir(parents=True, exist_ok=True)

SHEET_ID = '1RpxZ69Wnfcots006Hp5fzawYxhUsscW03O_hD3psjHA'
CREDENTIALS = PROJECT_ROOT / 'credentials.json'

# Nombres exactos de pestañas (validar al primer run)
TAB_TRANSITO = 'Confirmado EN TRANSITO'
TAB_BODEGA = 'EN BODEGA'


def _conectar_sheets():
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = [
        'https://www.googleapis.com/auth/spreadsheets.readonly',
        'https://www.googleapis.com/auth/drive.readonly',
    ]
    creds = Credentials.from_service_account_file(str(CREDENTIALS), scopes=scopes)
    return gspread.authorize(creds)


def _parse_num(s):
    """Limpia numero formato Chile/EU: '  500 ', '19,20', '$21.888,000', '63,51%'."""
    if s is None or s == '' or pd.isna(s):
        return None
    s = str(s).strip()
    if s == '':
        return None
    # Remover %, $, espacios
    s = s.replace('$', '').replace('%', '').replace(' ', '').replace('\xa0', '')
    if not s:
        return None
    # Determinar formato: si tiene puntos y coma, asumir EU (1.234,56)
    # Si solo coma, coma es decimal
    if ',' in s and '.' in s:
        # Asumir: punto miles, coma decimal
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        # Coma es decimal
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def _parse_fecha(s):
    """Parsea fecha formato DD/MM/YYYY o D/M/YYYY."""
    if not s or pd.isna(s):
        return None
    s = str(s).strip()
    if not s:
        return None
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_pi_grupo(pi):
    """Extrae año del numero de PI '26TP0309PI' -> año 2026, codigo TP0309."""
    if not pi:
        return None, None, None
    m = re.match(r'^(\d{2})(\w+)PI?$', str(pi).strip())
    if not m:
        return None, None, str(pi)
    año = 2000 + int(m.group(1))
    codigo = m.group(2)
    return año, codigo, str(pi)


def cargar_pestaña(client, tab_name: str) -> pd.DataFrame:
    """Carga una pestaña como DataFrame normalizado."""
    print(f"[1] Abriendo sheet + pestaña '{tab_name}'...", flush=True)
    sh = client.open_by_key(SHEET_ID)
    try:
        ws = sh.worksheet(tab_name)
    except Exception as e:
        print(f"[ERROR] Pestaña '{tab_name}' no encontrada: {e}", flush=True)
        print(f"        Pestañas disponibles: {[w.title for w in sh.worksheets()]}", flush=True)
        return pd.DataFrame()

    rows = ws.get_all_values()
    if len(rows) < 2:
        return pd.DataFrame()

    headers = [h.strip() for h in rows[0]]
    df = pd.DataFrame(rows[1:], columns=headers)
    print(f"   Filas crudas: {len(df):,} | columnas: {list(df.columns)[:10]}...", flush=True)
    return df


def cargar_desde_md_dump(md_path: Path) -> pd.DataFrame:
    """Fallback: parsea el dump markdown del MCP Drive cuando service account no tiene acceso.

    El dump es un JSON {fileContent: "...markdown..."} con la tabla en formato:
    | col1 | col2 | ... |
    | :-: | :-: | ... |
    | val1 | val2 | ... |
    """
    if not md_path.exists():
        return pd.DataFrame()
    content = json.loads(md_path.read_text(encoding='utf-8'))['fileContent']
    # Solo filas con TRANSITO (filtro grueso por palabra)
    lines = content.split('\\n')
    rows = []
    headers = None
    for line in lines:
        if not line.strip().startswith('|'):
            continue
        # Separar por |
        cells = [c.strip() for c in line.strip().strip('|').split('|')]
        # Saltear separadores tipo ':-: | :-: |'
        if all(c in ('', ':-:', ':--', '---', '--:') for c in cells):
            continue
        # Header (primera fila con 'F' / 'SKU')
        if headers is None:
            if 'SKU' in cells or 'sku' in [c.lower() for c in cells]:
                headers = cells
            continue
        # Solo data
        if len(cells) == len(headers):
            rows.append(cells)

    if not rows or not headers:
        return pd.DataFrame()
    df = pd.DataFrame(rows, columns=headers)
    print(f"   [fallback markdown] Filas crudas: {len(df):,}", flush=True)
    return df


def _enriquecer_clp_desde_precosteos(df: pd.DataFrame) -> pd.DataFrame:
    """Para filas con costo_ingreso_clp vacío, busca en los Pre-costeo Excel
    (agente-comex/data/output/26TPXXXX/Pre-costeo_x_CBM_*.xlsx) el costo
    internado unitario CLP por SKU y lo aplica."""
    import openpyxl
    precosteos_dir = PROJECT_ROOT / 'agente-comex' / 'data' / 'output'
    if not precosteos_dir.exists():
        return df

    df['costo_ingreso_clp_num'] = pd.to_numeric(df['costo_ingreso_clp'], errors='coerce').fillna(0)
    skus_sin_clp = df[df['costo_ingreso_clp_num'] == 0][['sku', 'pi_codigo']].drop_duplicates()
    if skus_sin_clp.empty:
        df.drop(columns='costo_ingreso_clp_num', inplace=True)
        return df

    # Mapear: para cada precosteo Excel, leer SKU -> CLP unit
    enriched = 0
    for sub in precosteos_dir.iterdir():
        if not sub.is_dir():
            continue
        for xlsx in sub.glob('Pre-costeo_x_CBM_*.xlsx'):
            try:
                wb = openpyxl.load_workbook(str(xlsx), data_only=True, read_only=True)
                if 'Productos' not in wb.sheetnames:
                    continue
                ws = wb['Productos']
                hdrs = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
                if 'SKU' not in hdrs or 'Costo Internado Unit (CLP)' not in hdrs:
                    continue
                i_sku = hdrs.index('SKU')
                i_clp = hdrs.index('Costo Internado Unit (CLP)')
                # Embarque desde filename: Pre-costeo_x_CBM_26TP0320.xlsx -> TP0320
                m = re.search(r'(\d{2}TP[A-Z]*\d+)', xlsx.stem)
                if not m:
                    continue
                pi_full = m.group(1).upper()
                pi_codigo = pi_full[2:] if pi_full.startswith(('24', '25', '26')) else pi_full

                mapping = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    sku = row[i_sku]
                    clp = row[i_clp]
                    if sku and clp:
                        mapping[str(sku).strip()] = float(clp)
                wb.close()

                # Aplicar: costo_ingreso_clp = total CLP por línea = qty * unit_clp_precosteo
                mask = (df['pi_codigo'] == pi_codigo) & (df['costo_ingreso_clp_num'] == 0) & (df['sku'].astype(str).isin(mapping.keys()))
                for idx in df[mask].index:
                    sku = str(df.at[idx, 'sku'])
                    qty = pd.to_numeric(df.at[idx, 'cantidad'], errors='coerce') or 0
                    total_clp = int(round(qty * mapping[sku]))
                    df.at[idx, 'costo_ingreso_clp'] = str(total_clp)
                    enriched += 1
            except Exception as e:
                print(f"   [WARN] precosteo {xlsx.name}: {type(e).__name__}", flush=True)

    df.drop(columns='costo_ingreso_clp_num', inplace=True)
    if enriched:
        print(f"   [enriquecimiento] {enriched} filas con costo_ingreso_clp vacío rellenadas desde precosteos COMEX", flush=True)
    return df


def normalizar_transito(df: pd.DataFrame) -> pd.DataFrame:
    """Limpia + tipos correctos para la pestaña TRANSITO."""
    if df.empty:
        return df

    # Normalizar nombres de columnas (estaban en el sheet sin underscore consistente)
    rename = {}
    for c in df.columns:
        cl = c.strip()
        cl_norm = cl.lower()
        if cl_norm == 'f':
            rename[c] = 'flag'
        elif cl_norm == 'sku':
            rename[c] = 'sku'
        elif cl_norm == 'variante':
            rename[c] = 'producto'
        elif cl_norm == 'pi':
            rename[c] = 'pi'
        elif cl_norm == 'status':
            rename[c] = 'status'
        elif 'transporte' in cl_norm:
            rename[c] = 'transporte'
        elif 'nro pedido' in cl_norm or cl_norm == 'nro pedido':
            rename[c] = 'nro_pedido'
        elif cl_norm == 'cantidad':
            rename[c] = 'cantidad'
        elif 'costo uni' in cl_norm:
            rename[c] = 'costo_unitario_usd'
        elif 'gift' in cl_norm:
            rename[c] = 'gift_box_envio'
        elif 'costo ingreso' in cl_norm or 'costo ingreso clp' in cl_norm:
            rename[c] = 'costo_ingreso_clp'
        elif 'embarque' in cl_norm:
            rename[c] = 'fecha_embarque'
        elif 'eta chile' in cl_norm:
            rename[c] = 'fecha_eta_chile'
        elif 'eta bodega' in cl_norm:
            rename[c] = 'fecha_eta_bodega'
        elif cl_norm == 'odoo':
            rename[c] = 'odoo'
    df = df.rename(columns=rename)
    # Eliminar columnas sin nombre o duplicadas (sheet tiene celdas extra vacias)
    df = df.loc[:, [c for c in df.columns if c and not c.startswith('Unnamed')]]
    df = df.loc[:, ~df.columns.duplicated()]

    # Filtrar solo filas con TRANSITO (excluir BODEGA u otros)
    if 'status' in df.columns:
        df = df[df['status'].str.upper().str.strip() == 'TRANSITO'].copy()

    # Limpiar/normalizar
    df['sku'] = df['sku'].astype(str).str.strip()
    df = df[df['sku'].str.len() > 0]
    df['producto'] = df['producto'].astype(str).str.strip()
    df['pi'] = df['pi'].astype(str).str.strip()

    # Numericos
    df['cantidad'] = df['cantidad'].apply(_parse_num)
    df['costo_unitario_usd'] = df['costo_unitario_usd'].apply(_parse_num)
    df['gift_box_envio'] = df['gift_box_envio'].apply(_parse_num)

    # Costo total USD (cantidad × unitario)
    df['costo_total_usd'] = df['cantidad'] * df['costo_unitario_usd']

    # Fechas
    df['fecha_embarque'] = df['fecha_embarque'].apply(_parse_fecha)
    df['fecha_eta_chile'] = df['fecha_eta_chile'].apply(_parse_fecha)
    df['fecha_eta_bodega'] = df['fecha_eta_bodega'].apply(_parse_fecha)

    # Año + codigo PI
    df[['pi_anio', 'pi_codigo', 'pi_full']] = df['pi'].apply(
        lambda x: pd.Series(_parse_pi_grupo(x))
    )

    # Filtrar filas vacias
    df = df[df['cantidad'].notna() & (df['cantidad'] > 0)].copy()
    df = df.reset_index(drop=True)
    return df


def main():
    import os
    print(f"=== Extraccion COMEX transito — {datetime.now()} ===\n", flush=True)

    df_raw = pd.DataFrame()

    # Intentar primero via service account
    if CREDENTIALS.exists():
        try:
            client = _conectar_sheets()
            df_raw = cargar_pestaña(client, TAB_TRANSITO)
        except Exception as e:
            err_str = str(e)
            if '403' in err_str or 'permission' in err_str.lower():
                print(f"[WARN] Service account sin acceso al sheet (compartir con "
                      f"union-x-revenue-bot@union-x-revenue.iam.gserviceaccount.com)", flush=True)
            else:
                print(f"[WARN] Auth sheets fallo: {e}", flush=True)

    # Fallback: dump markdown del MCP Drive
    if df_raw.empty:
        md_path = Path(os.environ.get('COMEX_MD_DUMP',
                                        PROJECT_ROOT / 'agente-comex' / 'data' / 'transito_dump.json'))
        if md_path.exists():
            print(f"[FALLBACK] Parseando dump local {md_path.name}", flush=True)
            df_raw = cargar_desde_md_dump(md_path)

    if df_raw.empty:
        print("[ERROR] Sin datos. Compartir sheet con la service account o proveer COMEX_MD_DUMP")
        sys.exit(1)

    print(f"\n[2] Normalizando...", flush=True)
    df = normalizar_transito(df_raw)

    # Enriquecer costo_ingreso_clp vacío desde precosteos COMEX generados por
    # _REACTIVAR_NUEVO_PC/costear_embarque.py (cubre el lag entre que se hace el
    # precosteo y que Martin cargue al sheet).
    df = _enriquecer_clp_desde_precosteos(df)

    print(f"   Filas validas en TRANSITO: {len(df):,}", flush=True)
    print(f"   PIs unicos: {df['pi'].nunique()}", flush=True)
    print(f"   SKUs unicos: {df['sku'].nunique()}", flush=True)
    print(f"   USD total: ${df['costo_total_usd'].sum()/1e6:.2f}M", flush=True)
    print(f"   Unidades total: {df['cantidad'].sum():,.0f}", flush=True)

    # Estadisticas por PI
    print(f"\n[3] Top PIs por USD:", flush=True)
    pi_top = df.groupby(['pi', 'fecha_embarque', 'fecha_eta_chile', 'fecha_eta_bodega'], dropna=False).agg(
        unidades=('cantidad', 'sum'),
        usd=('costo_total_usd', 'sum'),
        skus=('sku', 'nunique'),
    ).sort_values('usd', ascending=False).head(10)
    for idx, row in pi_top.iterrows():
        pi, fe, eta_cl, eta_bo = idx
        print(f"   {pi} | embarque={fe} -> ETA CL={eta_cl} -> bodega={eta_bo} | "
              f"{row['unidades']:>7,.0f} unid | ${row['usd']/1e3:>7,.1f}K | {row['skus']} SKUs", flush=True)

    # Guardar
    out_p = OUT_DIR / 'transito.parquet'
    df.to_parquet(out_p, compression='zstd', compression_level=9, index=False)
    print(f"\n[4] {out_p}: {len(df):,} filas", flush=True)

    # Resumen JSON
    resumen = {
        'generado_en': datetime.now().isoformat(),
        'total_filas': len(df),
        'total_pis': int(df['pi'].nunique()),
        'total_skus': int(df['sku'].nunique()),
        'total_unidades': float(df['cantidad'].sum()),
        'total_usd': float(df['costo_total_usd'].sum()) if df['costo_total_usd'].notna().any() else 0,
        'eta_proxima': str(df['fecha_eta_bodega'].min()) if df['fecha_eta_bodega'].notna().any() else None,
        'eta_lejana': str(df['fecha_eta_bodega'].max()) if df['fecha_eta_bodega'].notna().any() else None,
        'transporte_breakdown': df.groupby('transporte', dropna=False).size().to_dict() if 'transporte' in df.columns else {},
    }
    with open(OUT_DIR / 'transito_resumen.json', 'w', encoding='utf-8') as f:
        json.dump(resumen, f, indent=2, default=str)

    print(f"\n[OK] COMEX transito extraido")


if __name__ == '__main__':
    main()
