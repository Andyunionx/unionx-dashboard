"""Motor + herramienta de cotización AÉREA de importaciones (skill `cotizacion-aereo`).

Regla de peso cobrable (chargeable weight) de carga aérea:
    PV (peso volumétrico) = CBM × 167     (factor IATA: 1 m³ = 167 kg)
    Peso cobrable         = MAX(GW, PV)    (discrimina volumen vs peso)
    Flete                 = peso cobrable × tarifa_usd_kg

Internación Chile:
    internacion = FIJO_base + ad_valorem% × CIF + almacenaje_usd_kg × peso_cobrable
    default FIJO_base = USD 1500, ad_valorem = 0% (TLC China), almacenaje = 0.

Uso:
    python cotizar_aereo.py            # genera la herramienta de 5 pestañas
    from cotizar_aereo import cotizar  # como librería (motor puro)
"""
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

FACTOR_AEREO = 167.0  # kg por m³ (IATA: 6000 cm³/kg)


# ── Motor puro (para cálculo/validación en Python) ───────────────────────────
def cotizar(productos, tarifa_usd_kg, *, internacion=None, tc_clp=None, factor=FACTOR_AEREO):
    """Cotiza un embarque aéreo. productos: dicts con nombre, costo_unit_usd,
    uds_ctn, cbm_ctn, gw_ctn, cantidad. Devuelve (lineas, resumen)."""
    if internacion is None:
        internacion = {"modo": "fijo", "usd": 1500.0}
    lineas = []
    for p in productos:
        cant = float(p["cantidad"]); uds = float(p["uds_ctn"])
        cajas = cant / uds if uds else 0.0
        cbm_total = cajas * float(p["cbm_ctn"]); gw_total = cajas * float(p["gw_ctn"])
        pv = cbm_total * factor; peso_cobrable = max(pv, gw_total)
        fob = float(p["costo_unit_usd"]) * cant
        lineas.append({"nombre": p["nombre"], "cantidad": cant, "cbm_total": cbm_total,
                       "gw_total": gw_total, "pv": pv, "peso_cobrable": peso_cobrable,
                       "manda": "Volumen (PV)" if pv >= gw_total else "Peso (GW)", "fob": fob})
    sum_pv = sum(l["pv"] for l in lineas); sum_gw = sum(l["gw_total"] for l in lineas)
    cw = max(sum_pv, sum_gw); flete_total = cw * tarifa_usd_kg
    base_peso = sum(l["peso_cobrable"] for l in lineas) or 1.0
    for l in lineas:
        l["flete_usd"] = flete_total * (l["peso_cobrable"] / base_peso)
        l["cif_usd"] = l["fob"] + l["flete_usd"]
    cif_total = sum(l["cif_usd"] for l in lineas)
    intern_total = _internacion_total(internacion, cif_total, cw, tc_clp)
    base_cif = cif_total or 1.0
    for l in lineas:
        l["intern_usd"] = intern_total * (l["cif_usd"] / base_cif)
        l["internado_total_usd"] = l["fob"] + l["flete_usd"] + l["intern_usd"]
        l["internado_unit_usd"] = l["internado_total_usd"] / l["cantidad"] if l["cantidad"] else 0.0
        if tc_clp:
            l["internado_unit_clp"] = l["internado_unit_usd"] * tc_clp
    fob_total = sum(l["fob"] for l in lineas)
    resumen = {"peso_cobrable_envio": cw, "fob_total": fob_total, "flete_total": flete_total,
               "intern_total": intern_total, "cif_total": cif_total,
               "internado_total_usd": sum(l["internado_total_usd"] for l in lineas),
               "sobrecosto_pct": (flete_total + intern_total) / fob_total * 100 if fob_total else 0.0}
    return lineas, resumen


def _internacion_total(cfg, cif_total_usd, peso_cobrable, tc_clp):
    if cfg.get("modo", "fijo") == "fijo":
        return float(cfg.get("usd", 1500.0))
    tc = tc_clp or cfg.get("tc_clp", 950.0)
    desconsol = float(cfg.get("desconsolidacion_clp", 135000.0))
    terrestre = float(cfg.get("terrestre_clp", 100000.0))
    ag_pct = float(cfg.get("agente_pct", 0.005)); ag_min = float(cfg.get("agente_min_clp", 80000.0))
    adval = float(cfg.get("ad_valorem_pct", 0.0)); alm = float(cfg.get("almacenaje_usd_kg", 0.0))
    cif_clp = cif_total_usd * tc
    fijo_clp = desconsol + terrestre + max(ag_pct * cif_clp, ag_min)
    return fijo_clp / tc + adval * cif_total_usd + alm * peso_cobrable


# ── HERRAMIENTA (5 pestañas: Resumen · Costeo · Optimizacion · Comparativo · Mix) ──
def escribir_herramienta(productos, out_path, *, tarifa_default=6.0,
                         internacion_fija_usd=1500.0, tc_default=950.0,
                         factor=FACTOR_AEREO, titulo="Cotización Aérea — Herramienta"):
    """Workbook con fórmulas vivas:
      1) Resumen      — variables (amarillas) + resultados globales
      2) Costeo       — costeo por producto + costo Odoo + gap
      3) Optimizacion — curva internado_unit(N) single-product + break-even vs Odoo (+ gráfico)
      4) Comparativo  — aéreo vs Odoo por producto, veredicto
      5) Mix          — qué productos consolidar (premium estructural) + curva de consolidación
    productos: dicts con nombre, model, sku, costo_unit_usd, uds_ctn, cbm_ctn, gw_ctn,
               cantidad, odoo_cost_clp.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import LineChart, Reference

    AZUL = PatternFill("solid", fgColor="4884FC"); AMAR = PatternFill("solid", fgColor="FFF2CC")
    CALC = PatternFill("solid", fgColor="E2EFDA"); DATO = PatternFill("solid", fgColor="FDF2E9")
    HEADF = Font(color="FFFFFF", bold=True); bold = Font(bold=True); ital = Font(italic=True, color="808080")
    thin = Side(style="thin", color="D9D9D9"); border = Border(thin, thin, thin, thin)
    cen = Alignment("center", "center"); izq = Alignment("left", "center", wrap_text=True)

    wb = Workbook()
    R = wb.active; R.title = "Resumen"
    C = wb.create_sheet("Costeo"); O = wb.create_sheet("Optimizacion")
    K = wb.create_sheet("Comparativo"); M = wb.create_sheet("Mix")
    n = len(productos); f = 4; last = f + n - 1

    # ══ 1) RESUMEN ══
    R.merge_cells("A1:B1"); R["A1"] = titulo
    R["A1"].font = Font(bold=True, size=13, color="1F3864")
    R["A3"] = "VARIABLES (edita las amarillas)"; R["A3"].font = bold
    def rin(r, lab, val, fmt=None):
        R.cell(r, 1, lab).font = bold
        c = R.cell(r, 2, val); c.fill = AMAR; c.border = border; c.alignment = cen
        if fmt: c.number_format = fmt
    rin(4, "Tarifa aérea USD/kg", tarifa_default, "#,##0.00")
    rin(5, "Factor volumétrico kg/m³", factor, "#,##0")
    rin(6, "Tipo de cambio CLP/USD", tc_default, "#,##0")
    rin(7, "Modo internación (fijo/desglose)", "fijo")
    rin(8, "Internación FIJA USD", internacion_fija_usd, "#,##0.00")
    R.cell(9, 1, "— o desglose de partidas —").font = ital
    rin(10, "Desconsolidación CLP", 135000, "#,##0")
    rin(11, "Terrestre aeropuerto CLP", 100000, "#,##0")
    rin(12, "Agente aduana %", 0.005, "0.00%")
    rin(13, "Agente aduana mín CLP", 80000, "#,##0")
    rin(14, "Ad valorem % (TLC China=0)", 0.0, "0.0%")
    rin(15, "Almacenaje USD/kg", 0.0, "#,##0.00")
    dv = DataValidation(type="list", formula1='"fijo,desglose"', allow_blank=False)
    R.add_data_validation(dv); dv.add(R["B7"])
    R["A17"] = "RESULTADOS GLOBALES (calculado)"; R["A17"].font = Font(bold=True, color="1F3864")
    def rca(r, lab, formula, fmt="#,##0.00"):
        R.cell(r, 1, lab).font = bold
        c = R.cell(r, 2, formula); c.fill = CALC; c.border = border; c.alignment = cen; c.number_format = fmt
    rca(18, "Peso cobrable envío kg", f"=MAX(SUM(Costeo!$K${f}:$K${last}),SUM(Costeo!$J${f}:$J${last}))", "#,##0.0")
    rca(19, "FOB total USD", f"=SUM(Costeo!$N${f}:$N${last})")
    rca(20, "Flete aéreo total USD", "=B18*B4")
    rca(21, "CIF total USD", "=B19+B20")
    rca(22, "Internación desglose USD", "=(B10+B11+MAX(B12*B21*B6,B13))/B6+B14*B21+B15*B18")
    rca(23, "INTERNACIÓN TOTAL USD", '=IF(B7="fijo",B8,B22)')
    rca(24, "Costo internado TOTAL USD", "=B19+B20+B23")
    rca(25, "Sobrecosto % (flete+intern/FOB)", "=(B20+B23)/B19", "0.0%")
    rca(26, "Costo internado TOTAL CLP", "=B24*B6", "#,##0")
    R.column_dimensions["A"].width = 34; R.column_dimensions["B"].width = 16

    # ══ 2) COSTEO ══
    C.merge_cells("A1:V1"); C["A1"] = "Costeo por producto — variables en 'Resumen' · Odoo = costo actual de libro"
    C["A1"].font = Font(bold=True, size=12, color="1F3864")
    hdr = ["Producto", "Model", "SKU", "Cant.", "Costo unit USD", "Uds/CTN", "CBM/CTN", "GW/CTN",
           "CBM tot", "GW tot kg", "PV kg", "Peso cobrable", "Manda", "FOB USD", "Flete USD",
           "CIF USD", "Internac USD", "Internado tot USD", "Internado u USD", "Internado u CLP",
           "Odoo CLP", "Gap% vs Odoo"]
    for j, h in enumerate(hdr, 1):
        c = C.cell(3, j, h); c.fill = AZUL; c.font = HEADF; c.alignment = cen; c.border = border
    for i, p in enumerate(productos):
        r = f + i
        C.cell(r, 1, p["nombre"]).alignment = izq
        C.cell(r, 2, p.get("model", "")); C.cell(r, 3, p.get("sku", ""))
        C.cell(r, 4, p.get("cantidad", "")).fill = AMAR
        for col, key in [(5, "costo_unit_usd"), (6, "uds_ctn"), (7, "cbm_ctn"), (8, "gw_ctn")]:
            C.cell(r, col, p[key]).fill = DATO
        C.cell(r, 9,  f"=D{r}/F{r}*G{r}"); C.cell(r, 10, f"=D{r}/F{r}*H{r}")
        C.cell(r, 11, f"=I{r}*Resumen!$B$5"); C.cell(r, 12, f"=MAX(J{r},K{r})")
        C.cell(r, 13, f'=IF(K{r}>=J{r},"Volumen (PV)","Peso (GW)")')
        C.cell(r, 14, f"=E{r}*D{r}")
        C.cell(r, 15, f"=Resumen!$B$20*L{r}/SUM($L${f}:$L${last})")
        C.cell(r, 16, f"=N{r}+O{r}"); C.cell(r, 17, f"=Resumen!$B$23*P{r}/Resumen!$B$21")
        C.cell(r, 18, f"=P{r}+Q{r}"); C.cell(r, 19, f"=R{r}/D{r}"); C.cell(r, 20, f"=S{r}*Resumen!$B$6")
        cu = C.cell(r, 21, p.get("odoo_cost_clp") if p.get("odoo_cost_clp") else 0); cu.fill = AMAR
        C.cell(r, 22, f'=IF(U{r}=0,"nuevo",(T{r}-U{r})/U{r})')
        for j in range(1, 23):
            cc = C.cell(r, j); cc.border = border; cc.alignment = izq if j == 1 else cen
        for j in (4, 6): C.cell(r, j).number_format = "#,##0"
        for j in (5, 9, 15, 16, 17, 18, 19): C.cell(r, j).number_format = "#,##0.00"
        C.cell(r, 7).number_format = "0.000"
        for j in (10, 11, 12): C.cell(r, j).number_format = "#,##0.0"
        for j in (20, 21): C.cell(r, j).number_format = "#,##0"
        C.cell(r, 22).number_format = "0.0%"
    tr = last + 1
    C.cell(tr, 1, "TOTAL").font = bold
    for j, col in [(4, "D"), (9, "I"), (10, "J"), (11, "K"), (12, "L"),
                   (14, "N"), (15, "O"), (16, "P"), (17, "Q"), (18, "R")]:
        c = C.cell(tr, j, f"=SUM({col}{f}:{col}{last})"); c.font = bold
        c.number_format = "#,##0" if j == 4 else ("#,##0.0" if j in (10, 11, 12) else "#,##0.00")
    for j in range(1, 23):
        C.cell(tr, j).fill = PatternFill("solid", fgColor="D6E4FF"); C.cell(tr, j).border = border
    anchos = [30, 7, 14, 7, 12, 7, 8, 7, 8, 9, 8, 11, 12, 10, 10, 10, 11, 14, 13, 13, 10, 11]
    for j, w in enumerate(anchos, 1): C.column_dimensions[get_column_letter(j)].width = w
    C.freeze_panes = "A4"
    C.cell(tr + 2, 1, "Amarillo = editas tú (Cant. + Odoo CLP). Naranja = datos PI/PL. Variables en 'Resumen'.").font = ital

    # ══ 3) OPTIMIZACION (single-product) ══
    O.merge_cells("A1:F1"); O["A1"] = "Optimización single-product — costo internado unitario vs unidades"
    O["A1"].font = Font(bold=True, size=12, color="1F3864")
    O["A3"] = "Producto (elige del menú)"; O["A3"].font = bold
    csel = O.cell(3, 2, productos[0]["nombre"]); csel.fill = AMAR; csel.border = border
    dv2 = DataValidation(type="list", formula1=f"=Costeo!$A${f}:$A${last}", allow_blank=False)
    O.add_data_validation(dv2); dv2.add(O["B3"])
    O.cell(4, 1, "Internación asignada USD (si traes SOLO este por aire)").font = bold
    cia = O.cell(4, 2, "=Resumen!$B$23"); cia.fill = AMAR; cia.border = border; cia.number_format = "#,##0.00"
    def ocalc(r, lab, formula, fmt="#,##0"):
        O.cell(r, 1, lab).font = bold
        c = O.cell(r, 2, formula); c.fill = CALC; c.border = border; c.number_format = fmt
    MI = f"MATCH($B$3,Costeo!$A${f}:$A${last},0)"
    ocalc(6, "Costo unit USD", f"=INDEX(Costeo!$E${f}:$E${last},{MI})", "#,##0.00")
    ocalc(7, "Peso cobrable/u kg", f"=INDEX(Costeo!$L${f}:$L${last},{MI})/INDEX(Costeo!$D${f}:$D${last},{MI})", "#,##0.000")
    ocalc(8, "Costo actual Odoo CLP", f"=INDEX(Costeo!$U${f}:$U${last},{MI})", "#,##0")
    ocalc(9, "Tarifa USD/kg", "=Resumen!$B$4", "#,##0.00")
    ocalc(10, "TC CLP/USD", "=Resumen!$B$6", "#,##0")
    ocalc(11, "FOB unit CLP", "=B6*B10"); ocalc(12, "Flete unit CLP", "=B7*B9*B10")
    ocalc(13, "Asíntota CLP (mínimo internado alcanzable)", "=B11+B12")
    ocalc(14, "Internación asignada CLP", "=B4*B10")
    ocalc(15, "Break-even N* vs Odoo (uds)",
          '=IF(B8>B13,B14/(B8-B13),"no alcanza (aire > Odoo a cualquier volumen)")', "#,##0")
    hrow = 17
    for j, h in enumerate(["Unidades N", "Internado unit CLP", "Sobrecosto %", "vs Odoo CLP", "vs Odoo %", "Odoo ref CLP"], 1):
        c = O.cell(hrow, j, h); c.fill = AZUL; c.font = HEADF; c.alignment = cen; c.border = border
    Ns = [50, 100, 150, 200, 300, 400, 500, 750, 1000, 1500, 2000, 3000, 5000]
    for i, N in enumerate(Ns):
        r = hrow + 1 + i
        O.cell(r, 1, N).number_format = "#,##0"
        O.cell(r, 2, f"=$B$13+$B$14/A{r}").number_format = "#,##0"
        O.cell(r, 3, f"=(B{r}-$B$11)/$B$11").number_format = "0.0%"
        O.cell(r, 4, f"=B{r}-$B$8").number_format = "#,##0"
        O.cell(r, 5, f'=IF($B$8=0,"",(B{r}-$B$8)/$B$8)').number_format = "0.0%"
        O.cell(r, 6, "=$B$8").number_format = "#,##0"
        for j in range(1, 7): O.cell(r, j).border = border; O.cell(r, j).alignment = cen
    lastN = hrow + len(Ns)
    ch = LineChart(); ch.title = "Costo internado unitario vs unidades"
    ch.y_axis.title = "CLP/unidad"; ch.x_axis.title = "Unidades"; ch.height = 8; ch.width = 16
    ch.add_data(Reference(O, min_col=2, max_col=2, min_row=hrow, max_row=lastN), titles_from_data=True)
    ch.add_data(Reference(O, min_col=6, max_col=6, min_row=hrow, max_row=lastN), titles_from_data=True)
    ch.set_categories(Reference(O, min_col=1, min_row=hrow + 1, max_row=lastN))
    O.add_chart(ch, "H3")
    O.column_dimensions["A"].width = 42
    for col in "BCDEF": O.column_dimensions[col].width = 16

    # ══ 4) COMPARATIVO ══
    K.merge_cells("A1:J1"); K["A1"] = "Comparativo aéreo vs costo actual Odoo (por producto)"
    K["A1"].font = Font(bold=True, size=12, color="1F3864")
    kh = ["Producto", "SKU", "FOB unit CLP", "Flete unit CLP", "Internado unit CLP",
          "Odoo CLP", "Gap CLP", "Gap %", "Break-even N* (solo aire)", "Veredicto"]
    for j, h in enumerate(kh, 1):
        c = K.cell(3, j, h); c.fill = AZUL; c.font = HEADF; c.alignment = cen; c.border = border
    for i in range(n):
        r2 = f + i; r = f + i
        K.cell(r, 1, f"=Costeo!A{r2}").alignment = izq
        K.cell(r, 2, f"=Costeo!C{r2}")
        K.cell(r, 3, f"=Costeo!E{r2}*Resumen!$B$6")
        K.cell(r, 4, f"=(Costeo!L{r2}/Costeo!D{r2})*Resumen!$B$4*Resumen!$B$6")
        K.cell(r, 5, f"=Costeo!T{r2}"); K.cell(r, 6, f"=Costeo!U{r2}")
        K.cell(r, 7, f"=E{r}-F{r}"); K.cell(r, 8, f'=IF(F{r}=0,"",(E{r}-F{r})/F{r})')
        K.cell(r, 9, f'=IF(F{r}=0,"nuevo",IF(F{r}>(C{r}+D{r}),(Resumen!$B$8*Resumen!$B$6)/(F{r}-(C{r}+D{r})),"no alcanza"))')
        K.cell(r, 10, f'=IF(F{r}=0,"nuevo (sin Odoo)",IF(E{r}<=F{r},"aire <= costo actual",IF(H{r}<0.15,"premium bajo (<15%)","premium alto")))')
        for j in range(1, 11):
            cc = K.cell(r, j); cc.border = border; cc.alignment = izq if j in (1, 10) else cen
        for j in (3, 4, 5, 6, 7, 9): K.cell(r, j).number_format = "#,##0"
        K.cell(r, 8).number_format = "0.0%"
    anchos_k = [30, 14, 12, 12, 15, 11, 11, 9, 18, 26]
    for j, w in enumerate(anchos_k, 1): K.column_dimensions[get_column_letter(j)].width = w
    K.freeze_panes = "A4"

    # ══ 5) MIX ══
    M.merge_cells("A1:G1"); M["A1"] = "Optimización de MIX — qué productos consolidar en un envío aéreo"
    M["A1"].font = Font(bold=True, size=12, color="1F3864")
    M.merge_cells("A3:G3")
    M["A3"] = ("A. ¿Qué productos poner en el mix?  El costo marginal de sumar un producto al envío = "
               "FOB + flete (el fijo ya lo paga el ancla). Se compara esa ASÍNTOTA vs Odoo.")
    M["A3"].font = bold; M["A3"].alignment = izq
    mh = ["Producto", "Costo unit CLP", "Flete unit CLP", "Asíntota CLP (marginal)",
          "Odoo CLP", "Premium estructural %", "Decisión mix"]
    for j, h in enumerate(mh, 1):
        c = M.cell(4, j, h); c.fill = AZUL; c.font = HEADF; c.alignment = cen; c.border = border
    for i in range(n):
        r2 = f + i; r = 5 + i
        M.cell(r, 1, f"=Costeo!A{r2}").alignment = izq
        M.cell(r, 2, f"=Costeo!E{r2}*Resumen!$B$6")
        M.cell(r, 3, f"=(Costeo!L{r2}/Costeo!D{r2})*Resumen!$B$4*Resumen!$B$6")
        M.cell(r, 4, f"=B{r}+C{r}"); M.cell(r, 5, f"=Costeo!U{r2}")
        M.cell(r, 6, f'=IF(E{r}=0,"",(D{r}-E{r})/E{r})')
        M.cell(r, 7, f'=IF(E{r}=0,"nuevo (sin Odoo)",IF(D{r}<=E{r},"fly (marginal <= Odoo)",'
                     f'IF(F{r}<0.15,"fly (premium bajo)",IF(F{r}<0.4,"evaluar","marítimo (premium alto)"))))')
        for j in range(1, 8):
            cc = M.cell(r, j); cc.border = border; cc.alignment = izq if j in (1, 7) else cen
        for j in (2, 3, 4, 5): M.cell(r, j).number_format = "#,##0"
        M.cell(r, 6).number_format = "0.0%"
    base = 5 + n
    M.merge_cells(f"A{base+1}:G{base+1}")
    M.cell(base + 1, 1, "B. Beneficio de consolidar (llenar el envío): la internación fija se reparte entre más unidades").font = bold
    def mbase(r, lab, formula, fmt="#,##0"):
        M.cell(r, 1, lab).font = bold
        c = M.cell(r, 2, formula); c.fill = CALC; c.border = border; c.number_format = fmt
    mbase(base + 2, "Unidades base (mix actual)", f"=SUM(Costeo!$D${f}:$D${last})")
    mbase(base + 3, "FOB base USD", "=Resumen!$B$19", "#,##0.00")
    mbase(base + 4, "Flete base USD", "=Resumen!$B$20", "#,##0.00")
    mbase(base + 5, "Internación fija USD", "=Resumen!$B$23", "#,##0.00")
    ub, fb, lb, ib = f"$B${base+2}", f"$B${base+3}", f"$B${base+4}", f"$B${base+5}"
    hh = base + 7
    for j, h in enumerate(["Escala ×", "Unidades", "FOB USD", "Flete USD", "Internación USD",
                           "Internación USD/u", "Internado total USD", "Sobrecosto % vs FOB"], 1):
        c = M.cell(hh, j, h); c.fill = AZUL; c.font = HEADF; c.alignment = cen; c.border = border
    for i, s in enumerate([0.25, 0.5, 1, 2, 3, 5, 10]):
        r = hh + 1 + i
        M.cell(r, 1, s).number_format = "0.00"
        M.cell(r, 2, f"={ub}*A{r}").number_format = "#,##0"
        M.cell(r, 3, f"={fb}*A{r}").number_format = "#,##0"
        M.cell(r, 4, f"={lb}*A{r}").number_format = "#,##0"
        M.cell(r, 5, f"={ib}").number_format = "#,##0"
        M.cell(r, 6, f"=E{r}/B{r}").number_format = "#,##0.00"
        M.cell(r, 7, f"=C{r}+D{r}+E{r}").number_format = "#,##0"
        M.cell(r, 8, f"=(D{r}+E{r})/C{r}").number_format = "0.0%"
        for j in range(1, 9): M.cell(r, j).border = border; M.cell(r, j).alignment = cen
    lastS = hh + 7
    ch2 = LineChart(); ch2.title = "Sobrecosto % vs escala del envío"
    ch2.y_axis.title = "Sobrecosto %"; ch2.x_axis.title = "Escala ×"; ch2.height = 7; ch2.width = 14
    ch2.add_data(Reference(M, min_col=8, max_col=8, min_row=hh, max_row=lastS), titles_from_data=True)
    ch2.set_categories(Reference(M, min_col=1, min_row=hh + 1, max_row=lastS))
    M.add_chart(ch2, f"A{lastS+2}")
    note = lastS + 18
    M.cell(note, 1, "Cómo leer el mix:").font = bold
    for k, txt in enumerate([
        "1. El envío aéreo se JUSTIFICA con el producto ancla (mayor costo de oportunidad de quiebre).",
        "2. Pagado el fijo, sumar productos cuesta solo FOB+flete → mira la columna 'Decisión mix'.",
        "3. 'fly' = premium estructural bajo (proyectores). 'marítimo' = premium alto (baratos/voluminosos).",
        "4. Consolidar más unidades baja la internación/u (tabla B), pero NO baja el premium de flete.",
        "5. El internado real en mix (con el fijo repartido a cantidades actuales) está en 'Costeo'.",
    ], start=1):
        M.cell(note + k, 1, txt).alignment = izq
    M.column_dimensions["A"].width = 34
    for col in "BCDEF": M.column_dimensions[col].width = 15
    M.column_dimensions["G"].width = 22; M.column_dimensions["H"].width = 16

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# Productos con SKU + costo actual Odoo (CLP) — de PI/PL + Odoo std_price (10-ago-2026)
PRODUCTOS_AEREO = [
    dict(nombre="Secador Multistyler T-Care 5 en 1", model="HD4",   sku="TCMULTSTY5N1", costo_unit_usd=24.00, uds_ctn=6,  cbm_ctn=0.079, gw_ctn=12.10, odoo_cost_clp=21659),
    dict(nombre="Proyector Smart Lumix 9000lm",      model="YG361", sku="LVPROLUM",     costo_unit_usd=50.00, uds_ctn=8,  cbm_ctn=0.083, gw_ctn=17.60, odoo_cost_clp=89443),
    dict(nombre="Proyector Smart Orion 12000lm",     model="AC251", sku="LVPROORI",     costo_unit_usd=60.00, uds_ctn=8,  cbm_ctn=0.083, gw_ctn=19.50, odoo_cost_clp=71625),
    dict(nombre="Proyector Smart Hexa 8000lm",       model="AC201", sku="LVPROHEX",     costo_unit_usd=40.50, uds_ctn=10, cbm_ctn=0.067, gw_ctn=13.50, odoo_cost_clp=37187),
    dict(nombre="Audífono Gamer Zenit Elite",        model="KW-406", sku="LVAUDGM-WH",  costo_unit_usd=19.00, uds_ctn=10, cbm_ctn=0.076, gw_ctn=8.80,  odoo_cost_clp=17185),
    dict(nombre="Proyector Smart Nova 9000lm",       model="AC261", sku="LVPRONOV",     costo_unit_usd=50.00, uds_ctn=8,  cbm_ctn=0.085, gw_ctn=18.60, odoo_cost_clp=56881),
    dict(nombre="Monitor Gamer NexView 15.6\"",       model="TP59", sku="LVMONEXV-15",  costo_unit_usd=37.00, uds_ctn=8,  cbm_ctn=0.073, gw_ctn=12.50, odoo_cost_clp=37116),
    dict(nombre="Mousepad Battlemat XXL 90x40",      model="TP95",  sku="LVMPADXXL-BK", costo_unit_usd=3.40,  uds_ctn=20, cbm_ctn=0.058, gw_ctn=15.00, odoo_cost_clp=3299),
    dict(nombre="Secador Multistyler AeroFlow Beige", model="HD4",  sku="TCMULTSTY5N1-BG", costo_unit_usd=24.00, uds_ctn=6,  cbm_ctn=0.079, gw_ctn=12.10, odoo_cost_clp=21706),
    dict(nombre="Teclado Gamer NovaBlade TKL",       model="KS201", sku="LVTECOFSL-GR", costo_unit_usd=13.80, uds_ctn=20, cbm_ctn=0.051, gw_ctn=13.00, odoo_cost_clp=12029),
]


def _build():
    prods = [dict(p, cantidad=p["uds_ctn"] * 20) for p in PRODUCTOS_AEREO]  # placeholder 20 cajas c/u
    out = escribir_herramienta(prods, r"C:\Users\andre\Downloads\Cotizacion_Aereo.xlsx",
                               tarifa_default=6.0, internacion_fija_usd=1500.0, tc_default=950.0)
    print(f"Herramienta 5-pestañas generada: {out}")
    print("  Resumen · Costeo · Optimizacion · Comparativo · Mix")


if __name__ == "__main__":
    _build()
