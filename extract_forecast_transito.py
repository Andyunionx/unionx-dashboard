#!/usr/bin/env python3
"""
Extrae las columnas "Transito MES" del FORECAST FINAL SKU 26-27 V2.xlsx
y guarda como parquet local para uso en la vista de Cobertura.

Columnas buscadas: "Transito AGO26", "TRANSITO SEP26", "Tránsito OCT26", etc.

Output:
  data/planificacion/snapshots/planif_forecast_transito.parquet
  Formato: [sku, mes (YYYY-MM), unidades]

Uso: python extract_forecast_transito.py
     (requiere el Excel en data/planificacion/FORECAST FINAL SKU 26-27 V2.xlsx)
"""
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).parent
EXCEL_PATH   = PROJECT_ROOT / 'data' / 'planificacion' / 'FORECAST FINAL SKU 26-27 V2.xlsx'
OUTPUT_PATH  = PROJECT_ROOT / 'data' / 'planificacion' / 'snapshots' / 'planif_forecast_transito.parquet'

MESES_ES_A_NUM = {
    'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12
}


def _parse_header_transito(h: str) -> str | None:
    """'Transito AGO26' / 'TRÁNSITO SEP26' → '2026-08'. None si no parsea."""
    if not h:
        return None
    s = str(h).strip().upper()
    # Acepta variantes: TRANSITO, TRÁNSITO, TRAN.
    prefijos = ('TRANSITO', 'TRÁNSITO', 'TRANSITO ', 'TRÁNSITO ', 'TRANS ')
    for pref in prefijos:
        if s.startswith(pref):
            tail = s[len(pref):].strip()
            break
    else:
        return None
    # tail = 'AGO26' o 'AGO27' o similar
    if len(tail) < 5:
        return None
    mmm = tail[:3]
    yy  = tail[3:].strip()
    if mmm not in MESES_ES_A_NUM or not yy.isdigit():
        return None
    return f'{2000 + int(yy):04d}-{MESES_ES_A_NUM[mmm]:02d}'


def extract_transito() -> pd.DataFrame:
    if not EXCEL_PATH.exists():
        print(f"[ERROR] Excel no encontrado: {EXCEL_PATH}", flush=True)
        sys.exit(1)

    print(f"[1] Abriendo Excel...", flush=True)
    t0 = time.time()
    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    print(f"    OK en {time.time() - t0:.1f}s", flush=True)

    ws = wb['FCST BASE SKU MACRO']
    rows_iter = ws.iter_rows(values_only=True)

    next(rows_iter); next(rows_iter)          # skip filas 1-2
    hdr = next(rows_iter)                      # fila 3 = headers
    idx_sku = 4                                # SKU en columna 4

    # Detectar columnas de tránsito
    tr_cols = []
    for idx, h in enumerate(hdr):
        mes_iso = _parse_header_transito(h)
        if mes_iso:
            tr_cols.append((idx, mes_iso, str(h).strip()))

    if not tr_cols:
        print("[WARN] No se encontraron columnas de Transito en el Excel.", flush=True)
        wb.close()
        return pd.DataFrame(columns=['sku', 'mes', 'unidades'])

    print(f"[2] Detectadas {len(tr_cols)} columnas de tránsito:", flush=True)
    for _, mes_iso, lbl in tr_cols:
        print(f"      {mes_iso}  ←  '{lbl}'", flush=True)

    long_rows = []
    n_filas   = 0
    for row in rows_iter:
        n_filas += 1
        sku = row[idx_sku] if idx_sku < len(row) else None
        if not sku:
            continue
        sku_str = str(sku).strip()
        if not sku_str or sku_str.lower() in ('nan', 'none', ''):
            continue
        for col_idx, mes_iso, _ in tr_cols:
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None:
                continue
            try:
                v_f = float(v)
            except Exception:
                continue
            if v_f <= 0:
                continue
            long_rows.append({'sku': sku_str, 'mes': mes_iso, 'unidades': v_f})

    wb.close()
    print(f"[3] Filas leídas: {n_filas:,}. Registros tránsito (>0): {len(long_rows):,}", flush=True)

    df = pd.DataFrame(long_rows)
    if df.empty:
        print("[WARN] Sin datos de tránsito en el Excel.", flush=True)
        return df

    print(f"[4] Resumen por mes:", flush=True)
    g = df.groupby('mes').agg(n_skus=('sku', 'nunique'), total=('unidades', 'sum')).reset_index()
    for _, r in g.iterrows():
        print(f"      {r['mes']}: {r['n_skus']:>4} SKUs, {r['total']:>10,.0f} uds", flush=True)
    return df


if __name__ == '__main__':
    df = extract_transito()
    if not df.empty:
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        df['ts_actualizado'] = datetime.now().isoformat(timespec='seconds')
        df.to_parquet(OUTPUT_PATH, index=False)
        print(f"\n✅ Guardado en {OUTPUT_PATH}", flush=True)
        print(f"   {len(df):,} filas, {df['sku'].nunique()} SKUs únicos", flush=True)
    else:
        print("\n⚠️  Sin datos para guardar.", flush=True)
