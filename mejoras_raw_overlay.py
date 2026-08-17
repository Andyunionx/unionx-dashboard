# -*- coding: utf-8 -*-
"""
Overlay de mejoras al RAW (Nicole, 16-jun-2026) aplicables sin re-extraer Odoo.
`aplicar_mejoras(df)` se usa como post-proceso en el extract diario y también como
backfill one-shot vía CLI.

  P1) Pisar producto + atributos por SKU desde 'Matriz productos.xlsx'
      (1 descripción/categoría/marca por SKU, columnas completas).
  P5) Flag `es_despacho` (líneas de envío sin COGS → margen 100% esperado).
  P3) Backfill: en filas Devolución, fecha_venta = fecha de la VENTA ORIGINAL
      (desde cruce Odoo nc_detalle_h1). Para mes_actual el extract ya lo trae bien
      desde ventas_service; el backfill cubre el histórico congelado.

Uso CLI:  python mejoras_raw_overlay.py [--apply] [--parquet <ruta>]
"""
import argparse
import unicodedata
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parent
MATRIZ = ROOT / "data" / "planillas" / "Matriz productos.xlsx"
NC_DET = ROOT / "data" / "contabilidad" / "nc_detalle_h1.parquet"

MATRIZ_MAP = {
    "Producto": "producto", "Categoría macro": "categoria_macro",
    "Categoría padre": "categoria_padre", "Categoría hijo": "categoria_hijo",
    "Categoría comercial": "categoria_comercial", "Marca": "marca",
    "Proveedor": "proveedor", "Pack": "pack", "In/out": "estado_sku",
}
DESPACHO_KEYS = ("despacho", "flete", "envio", "envío", "shipping")

# Marca inferible por prefijo de SKU cuando no hay marca en ningún lado ni en la Matriz.
# Solo se aplica a filas con marca vacía (nunca pisa una marca existente).
PREFIJO_MARCA = {"LH": "Lhotse", "DN": "Dinasty"}

# Override canal → tipo_negocio (P6). Rellena tipo_negocio SOLO cuando viene vacío
# de Odoo (canal nuevo sin equipo de venta configurado). Nunca pisa un valor existente.
# Latam Pass agregado 22-jul-2026 (OK Andrés): canal de fidelización nuevo.
CANAL_TIPO_NEGOCIO = {
    "latam pass": "Fidelización",
    "shopping latam pass": "Fidelización",  # canal nuevo (Andrés 28-jul): fidelización
    # respaldo de canales de fidelización conocidos (por si Odoo los manda vacíos)
    "travel duty": "Fidelización", "global reward": "Fidelización",
    "celmedia": "Fidelización", "banco bice": "Fidelización", "cmr": "Fidelización",
    "friends": "Fidelización", "gluky": "Fidelización", "relacional": "Fidelización",
    "sawa": "Fidelización",
}


def _norm(s):
    s = str(s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _cargar_matriz():
    import openpyxl
    wb = openpyxl.load_workbook(MATRIZ, read_only=True, data_only=True)
    ws = wb["Productos"] if "Productos" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    idx = {str(h).strip() if h else "": i for i, h in enumerate(rows[0])}
    iSKU = idx.get("SKU")
    m = {}
    for r in rows[1:]:
        if iSKU is None or len(r) <= iSKU or not r[iSKU]:
            continue
        rec = {}
        for mcol, pcol in MATRIZ_MAP.items():
            i = idx.get(mcol)
            if i is not None and i < len(r) and r[i] not in (None, ""):
                rec[pcol] = r[i]
        m[_norm(str(r[iSKU]).strip())] = rec
    return m


def heal_marca(df, verbose=True):
    """P1b — rellena marca vacía sin re-extraer: (1) self-heal con la marca conocida
    del mismo SKU (moda), (2) inferencia por prefijo de SKU (PREFIJO_MARCA).
    Nunca pisa una marca ya existente. Solo toca la columna `marca`."""
    log = print if verbose else (lambda *a, **k: None)
    if "marca" not in df.columns or "sku" not in df.columns:
        return df
    marca = df["marca"].astype(str).str.strip()
    marca = marca.mask(marca.str.lower().isin({"none", "nan", "false", "0", "-", "sin marca"}), "")
    sku = df["sku"].astype(str).str.strip()
    sku_valido = sku.ne("") & sku.str.lower().ne("false")
    vacia = marca.eq("") & sku_valido
    n0 = int(vacia.sum())
    if n0 == 0:
        df["marca"] = marca
        log("  [P1b] marca: nada que rellenar")
        return df
    # 1) self-heal — moda de la marca conocida por SKU
    conocida = marca.ne("")
    if conocida.any():
        moda = marca[conocida].groupby(sku[conocida]).agg(
            lambda s: s.mode().iat[0] if len(s.mode()) else s.iloc[0])
        fill = sku.map(moda)
        heal = vacia & fill.notna()
        marca = marca.where(~heal, fill)
        n1 = int(heal.sum())
    else:
        n1 = 0
    # 2) prefijo de SKU (solo lo que sigue vacío)
    resta = marca.eq("") & sku_valido
    pref = sku.str.upper().str[:2].map(PREFIJO_MARCA)
    pmask = resta & pref.notna()
    marca = marca.where(~pmask, pref)
    n2 = int(pmask.sum())
    # 3) token de marca propia en el NOMBRE del producto (Purito, Levo, UMA, ...).
    #    Cubre productos nuevos que entran sin marca y que aún no están en la Matriz
    #    (catálogo→Yuju→Odoo). Reportado por Sebastián 21-jul (Purito, Pack490).
    n3 = 0
    if "producto" in df.columns:
        resta2 = marca.eq("") & sku_valido
        if resta2.any():
            from clasificar_marca import marca_desde_texto
            inf = df.loc[resta2, "producto"].apply(marca_desde_texto)
            hit = inf[inf.ne("")]
            marca.loc[hit.index] = hit
            n3 = int(len(hit))
    df["marca"] = marca
    log(f"  [P1b] marca: {n1} por SKU + {n2} por prefijo + {n3} por nombre "
        f"(quedan {n0 - n1 - n2 - n3} sin marca)")
    return df


def unificar_descripcion_por_sku(df, verbose=True):
    """P1d — un solo `producto` por SKU en TODAS las filas del RAW.

    Motivo (Nicole 12-ago): las filas com/log/mkt (Sheet 'Raw extras') traen el
    nombre del marketplace y las de Venta el de Odoo; mismo SKU con textos
    distintos rompe cualquier tabla dinámica agrupada por nombre de producto
    (64% de los SKU de ML jul-26 tenían ≥2 nombres). Canónico = nombre más
    frecuente entre las filas de Venta (Odoo, base del 'vs 2025'); si el SKU no
    tiene fila de Venta con nombre, se usa la moda global de ese SKU.
    """
    log = print if verbose else (lambda *a, **k: None)
    if "sku" not in df.columns or "producto" not in df.columns:
        return df
    sk = df["sku"].astype(str).str.strip()
    prod = df["producto"].astype(str).str.strip()
    valido = sk.ne("") & ~sk.str.lower().isin(["nan", "none"])
    es_venta = (df["tipo_movimiento"].astype(str).eq("Venta")
                if "tipo_movimiento" in df.columns else pd.Series(True, index=df.index))
    tmp = pd.DataFrame({"_sk": sk, "_prod": prod})
    def _moda(s):
        m = s.mode()
        return m.iat[0] if not m.empty else s.iloc[0]
    canon_v = tmp[valido & es_venta & prod.ne("")].groupby("_sk")["_prod"].agg(_moda)
    canon_all = tmp[valido & prod.ne("")].groupby("_sk")["_prod"].agg(_moda)
    canon = canon_v.combine_first(canon_all)
    nuevo = sk.map(canon)
    mask = valido & nuevo.notna() & nuevo.ne("")
    n_change = int((mask & (prod != nuevo)).sum())
    df.loc[mask, "producto"] = nuevo[mask].values
    log(f"  [P1d] descripción única por SKU: {n_change:,} filas homologadas "
        f"({int(canon.size):,} SKUs)")
    return df


def aplicar_mejoras(df, con_nc_backfill=True, verbose=True):
    """Aplica P1 (atributos por SKU), P5 (es_despacho) y opcional P3 (backfill NC).
    Vectorizado y sin copia para soportar el histórico (~414k filas)."""
    log = print if verbose else (lambda *a, **k: None)

    # P1 — pisar atributos por SKU (merge vectorizado, 1 join en vez de 9 .map)
    try:
        matriz = _cargar_matriz()  # {sku_norm: {pcol: val}}
        mrows = [{"_sk": k, **{p: v for p, v in rec.items()}} for k, rec in matriz.items()]
        mdf = pd.DataFrame(mrows).rename(columns={c: f"_m_{c}" for c in set(MATRIZ_MAP.values())})
        df["_sk"] = df["sku"].astype(str).str.strip().str.lower()
        df = df.merge(mdf, on="_sk", how="left")
        n_match = int(df["_m_producto"].notna().sum()) if "_m_producto" in df.columns else 0
        for pcol in set(MATRIZ_MAP.values()):
            mc = f"_m_{pcol}"
            if mc in df.columns and pcol in df.columns:
                df[pcol] = df[mc].where(df[mc].notna(), df[pcol])
                df[pcol] = df[pcol].astype(str).replace({"None": "", "nan": ""})
        df = df.drop(columns=[c for c in df.columns if c.startswith("_m_")] + ["_sk"])
        log(f"  [P1] atributos pisados desde Matriz ({n_match:,} filas con SKU en Matriz)")
    except Exception as e:
        log(f"  [P1] omitido: {type(e).__name__}: {e}")
        df = df.drop(columns=[c for c in df.columns if c.startswith("_m_") or c == "_sk"], errors="ignore")

    # P1b — self-heal de marca vacía (mismo SKU + prefijo)
    df = heal_marca(df, verbose=verbose)

    # P1d — descripción única por SKU (evita que un pivote por nombre se parta
    # cuando com/log/mkt traen otro texto para el mismo SKU). Nicole 12-ago.
    df = unificar_descripcion_por_sku(df, verbose=verbose)

    # P1c — tipo_marca SIEMPRE derivado de la marca (Propia/Otras marcas). Evita que
    # quede el crudo In/Out/No aplica de Odoo cuando el paso de clasificación del
    # extract falla (le pasó a julio; reportado por Felipe 20-jul). Regla: clasificar_marca.
    if "marca" in df.columns:
        try:
            from clasificar_marca import clasificar_tipo_marca
            df["tipo_marca"] = df["marca"].apply(clasificar_tipo_marca)
            log("  [P1c] tipo_marca derivado de marca (Propia/Otras marcas)")
        except Exception as e:
            log(f"  [P1c] omitido: {type(e).__name__}: {e}")

    # P6 — override canal → tipo_negocio (solo rellena vacíos; nunca pisa Odoo)
    if "canal" in df.columns and "tipo_negocio" in df.columns:
        canal_n = df["canal"].astype(str).str.strip().str.lower()
        tn = df["tipo_negocio"].astype(str).str.strip()
        vacio = tn.isin(["", "none", "nan", "0"])
        mapped = canal_n.map(CANAL_TIPO_NEGOCIO)
        fix = vacio & mapped.notna()
        df.loc[fix, "tipo_negocio"] = mapped[fix]
        log(f"  [P6] tipo_negocio por canal: {int(fix.sum()):,} filas rellenadas")

    # P5 — flag es_despacho (vectorizado, sin concatenar todo)
    pat = "despacho|flete|env[ií]o|shipping"
    marca_l = df["marca"].astype(str).str.lower()
    prod_l = df["producto"].astype(str).str.lower()
    df["es_despacho"] = marca_l.str.contains(pat, regex=True, na=False) | prod_l.str.contains(pat, regex=True, na=False)
    log(f"  [P5] es_despacho: {int(df['es_despacho'].sum()):,} filas")

    # P5c — envíos NO son proveedores nacionales (no tienen proveedor). Se saca el
    # despacho de tipo_compra='Compra' → 'Envío' para que no infle el margen directo
    # de proveedores nacionales en el reporte de rentabilidad (Andrés 14-jul).
    if "tipo_compra" in df.columns:
        env = df["es_despacho"] & df["tipo_compra"].astype(str).str.strip().str.lower().ne("envío")
        n_env = int(env.sum())
        df.loc[env, "tipo_compra"] = "Envío"
        log(f"  [P5c] tipo_compra='Envío' en {n_env:,} filas de despacho (fuera de proveedores nacionales)")

    # P5b — sanity de costo: filas con cruce de campo (cantidad≈venta → costo_total absurdo).
    # Se corrige a cantidad=1 y costo_total=costo_unitario (evita margen −billones).
    try:
        vn = pd.to_numeric(df["venta_neta"], errors="coerce")
        ct = pd.to_numeric(df["costo_total"], errors="coerce")
        cu = pd.to_numeric(df["costo_unitario"], errors="coerce")
        cant = pd.to_numeric(df["cantidad"], errors="coerce")
        # Firma REAL del bug: 'cantidad' cruzada con venta_neta → cantidad enorme
        # (el caso original tenía ~63.017). Se exige cantidad > 1000 para NO tocar
        # ítems B2B legítimos vendidos a $1 con costo alto (costo≫venta es real ahí),
        # que antes se corrompían a cantidad=1 (Nicole 17-ago, pedido S273209).
        absurd = (ct > (10 * vn.abs() + 100000)) & (cant > 1000)
        if absurd.any():
            df.loc[absurd, "cantidad"] = 1
            df.loc[absurd, "costo_total"] = cu[absurd]
            df.loc[absurd, "margen_front"] = vn[absurd] - cu[absurd]
            df.loc[absurd, "margen_final"] = vn[absurd] - cu[absurd]
            log(f"  [P5b] {int(absurd.sum())} filas con costo absurdo corregidas (cantidad=1)")
    except Exception as e:
        log(f"  [P5b] omitido: {e}")

    # P3 — backfill fecha_venta NC desde el cruce Odoo (solo histórico)
    if con_nc_backfill and NC_DET.exists() and "tipo_movimiento" in df.columns:
        nc = pd.read_parquet(NC_DET)
        nc_orig = {}
        for _, r in nc.iterrows():
            fo = str(r.get("Fecha venta original", ""))[:10]
            if len(fo) == 10 and fo[:4].isdigit():
                nc_orig[str(r["NC"]).strip()] = fo
        dev = df["tipo_movimiento"] == "Devolución"
        nueva = df.loc[dev, "documento"].astype(str).str.strip().map(nc_orig)
        mask = dev & nueva.notna()
        if mask.any():
            df.loc[mask, "fecha_venta"] = nueva[nueva.notna()].values
            fv = pd.to_datetime(df.loc[mask, "fecha_venta"], errors="coerce")
            df.loc[mask, "anio_venta"] = fv.dt.year
            df.loc[mask, "mes_venta"] = fv.dt.month
            df.loc[mask, "semana_venta"] = fv.dt.isocalendar().week.astype("Int64")
            df.loc[mask, "dia_semana"] = fv.dt.dayofweek
        log(f"  [P3] backfill fecha NC: {int(mask.sum()):,} filas")

    for c in ("anio_venta", "mes_venta", "semana_venta", "dia_semana"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype("int64")
    return df


def main(apply=False, parquet="data/historico/ventas_historico.parquet"):
    pq = ROOT / parquet
    df = pd.read_parquet(pq)
    print(f"Parquet: {parquet}  ({len(df):,} filas)")
    df = aplicar_mejoras(df)
    out = pq if apply else pq.with_name(pq.stem + "_MEJORADO.parquet")
    df.to_parquet(out, index=False)
    print(f"{'✓ APLICADO' if apply else '→ prueba'}: {out.relative_to(ROOT)}  ({len(df):,} filas, {df.shape[1]} cols)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--parquet", default="data/historico/ventas_historico.parquet")
    a = ap.parse_args()
    main(apply=a.apply, parquet=a.parquet)
