"""Genera Tarifas_Base_COMEX-XXXX.xlsx para un embarque específico.

Combina:
  - Maestra_Tarifa_Fija.xlsx (valores fijos: Inland Chile, benchmarks, parámetros)
  - Variables por embarque: puerto origen, flete USD, dólar aduana, fecha ETA

Uso:
    python generar_tarifas_embarque.py --embarque 0320 --puerto SZ --flete 2840 --dolar 857.58 --eta 2026-03-20
    python generar_tarifas_embarque.py --embarque 0320 --puerto SZ --flete 2840  # dolar auto SII, eta hoy+30

Output: data/comex/embarques/Tarifas_Base_COMEX-XXXX.xlsx (y copia a la carpeta COMEX en OneDrive)
"""
import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

sys.stdout.reconfigure(encoding='utf-8')
PROJECT_ROOT = Path(__file__).parent
MAESTRA_FIJA = Path('C:/Users/andre/OneDrive/Documentos/Claude/Projects/COMEX/Maestra_Tarifa_Fija.xlsx')
OUTPUT_DIR = Path('C:/Users/andre/OneDrive/Documentos/Claude/Projects/COMEX')


def dolar_sii_observado(fecha: str | None = None) -> float | None:
    """Obtiene el dólar observado del SII para la fecha indicada (default hoy).
    Usa API pública mindicador.cl.
    """
    try:
        if fecha is None:
            url = 'https://mindicador.cl/api/dolar'
        else:
            d = datetime.strptime(fecha, '%Y-%m-%d')
            url = f'https://mindicador.cl/api/dolar/{d.strftime("%d-%m-%Y")}'
        r = requests.get(url, timeout=10)
        if r.status_code == 200:
            data = r.json()
            if data.get('serie'):
                return float(data['serie'][0]['valor'])
    except Exception as e:
        print(f"[WARN] dólar SII falló: {e}", flush=True)
    return None


def cargar_fijos() -> dict:
    """Lee Maestra_Tarifa_Fija.xlsx y devuelve dict con todos los valores fijos."""
    wb = load_workbook(str(MAESTRA_FIJA), read_only=True, data_only=True)
    ws = wb['Tarifa Fija']
    fijos = {}
    section = None
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        cell = str(row[0]).strip()
        if cell in ('INLAND CHILE (CIF → Internado)', 'BENCHMARKS EFICIENCIA (referencia)', 'PARÁMETROS COSTEO'):
            section = cell
            continue
        if cell in ('Concepto', 'Puerto'):
            continue
        # filas de datos
        if row[1] is not None:
            fijos[cell] = {'valor': row[1], 'unidad': row[2] or '', 'section': section}
    wb.close()
    return fijos


def generar(embarque: str, puerto: str, flete_usd: float, dolar: float | None, eta: str):
    fijos = cargar_fijos()
    if dolar is None:
        dolar = dolar_sii_observado(eta) or dolar_sii_observado(None)
        if dolar is None:
            print("[ERROR] No pude obtener dólar SII. Pasa --dolar manualmente.")
            sys.exit(1)
        print(f"[dolar] Auto SII para {eta or 'hoy'}: ${dolar:,.2f} CLP/USD", flush=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tarifas'

    bold = Font(bold=True, size=11)
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    variable_fill = PatternFill('solid', fgColor='FFFF99')
    fixed_fill = PatternFill('solid', fgColor='E7E6E6')
    title = Font(bold=True, size=14, color='1F4E78')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    row = 1
    ws.cell(row, 1, f'TARIFAS COMEX - EMBARQUE {embarque}').font = title
    row += 1
    ws.cell(row, 1, 'Celdas amarillas = valores variables del embarque').font = Font(italic=True, size=9)
    row += 1
    ws.cell(row, 1, 'NOTA: Gastos Inland China se extraen del PI automáticamente').font = Font(italic=True, size=9, color='AA0000')
    row += 2

    # Sección 1: DATOS DEL EMBARQUE
    ws.cell(row, 1, 'DATOS DEL EMBARQUE').font = bold
    row += 1
    for c, h in enumerate(['Concepto', 'Valor', 'Unidad'], 1):
        cell = ws.cell(row, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    for label, val, unidad in [
        ('Puerto Origen',    puerto,                   'SZ, NB, XI, AIR'),
        ('Dólar Aduana',     round(dolar, 2),          'CLP por USD'),
        ('Fecha ETA',        eta,                      'Fecha estimada llegada'),
    ]:
        ws.cell(row, 1, label).border = border
        cv = ws.cell(row, 2, val); cv.border = border; cv.fill = variable_fill
        ws.cell(row, 3, unidad).border = border
        row += 1

    row += 1

    # Sección 2: FLETE MARÍTIMO
    ws.cell(row, 1, 'FLETE MARÍTIMO').font = bold
    row += 1
    for c, h in enumerate(['Concepto', 'Valor', 'Unidad'], 1):
        cell = ws.cell(row, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    for label, val, unidad in [
        ('Flete Total 40HQ',  flete_usd,  'USD total contenedor'),
        ('Capacidad 40HQ',    fijos.get('Capacidad 40HQ ref', {}).get('valor', 68), 'CBM (PL define real)'),
    ]:
        ws.cell(row, 1, label).border = border
        cv = ws.cell(row, 2, val); cv.border = border
        cv.fill = variable_fill if label == 'Flete Total 40HQ' else fixed_fill
        ws.cell(row, 3, unidad).border = border
        row += 1

    row += 1

    # Sección 3: INLAND CHILE (FIJOS)
    ws.cell(row, 1, 'INLAND CHILE (CIF → Internado)').font = bold
    row += 1
    for c, h in enumerate(['Concepto', 'Valor', 'Unidad'], 1):
        cell = ws.cell(row, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    for concepto in ['Agente Aduana (%)', 'Gastos Puerto STI', 'Flete Terrestre', 'Seimex',
                      'Desconsolidación Craft', 'Seguro Carga Contempora', 'Gastos Despacho',
                      'Gate In Maersk']:
        info = fijos.get(concepto, {})
        ws.cell(row, 1, concepto).border = border
        cv = ws.cell(row, 2, info.get('valor', 0)); cv.border = border; cv.fill = fixed_fill
        cv.number_format = '#,##0.0000' if concepto == 'Agente Aduana (%)' else '#,##0'
        ws.cell(row, 3, info.get('unidad', '')).border = border
        row += 1

    row += 1

    # Sección 4: BENCHMARKS
    ws.cell(row, 1, 'BENCHMARKS EFICIENCIA (referencia)').font = bold
    row += 1
    for c, h in enumerate(['Puerto', 'Umbral Máx Sobrecosto', 'Notas'], 1):
        cell = ws.cell(row, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    for puerto_b in ['SZ (Shenzhen)', 'NB (Ningbo)', 'XI (Xiamen)', 'AIR (Aéreo)']:
        info = fijos.get(puerto_b, {})
        ws.cell(row, 1, puerto_b).border = border
        cv = ws.cell(row, 2, info.get('valor', '')); cv.border = border; cv.fill = fixed_fill
        ws.cell(row, 3, info.get('unidad', '')).border = border
        row += 1

    # Auto-ancho
    for col_idx in range(1, 4):
        max_len = max((len(str(ws.cell(r, col_idx).value or '')) for r in range(1, row)), default=10)
        ws.column_dimensions[chr(64 + col_idx)].width = max(max_len + 2, 14)

    # Guardar en 2 ubicaciones: data/comex/embarques/ y carpeta COMEX OneDrive
    local_dir = PROJECT_ROOT / 'data' / 'comex' / 'embarques'
    local_dir.mkdir(parents=True, exist_ok=True)
    local_path = local_dir / f'Tarifas_Base_COMEX-{embarque}.xlsx'
    onedrive_path = OUTPUT_DIR / f'Tarifas_Base_COMEX-{embarque}.xlsx'

    wb.save(str(local_path))
    wb.save(str(onedrive_path))
    print(f"[OK] Generados:")
    print(f"     {local_path}")
    print(f"     {onedrive_path}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--embarque', required=True, help='Ej. 0320, 0424')
    p.add_argument('--puerto', required=True, choices=['SZ', 'NB', 'XI', 'AIR'])
    p.add_argument('--flete', required=True, type=float, help='Flete Total 40HQ USD')
    p.add_argument('--dolar', type=float, default=None,
                   help='Dólar aduana CLP/USD (default: auto SII por fecha eta)')
    p.add_argument('--eta', default=(datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'),
                   help='Fecha ETA (YYYY-MM-DD), default hoy+30 días')
    args = p.parse_args()
    generar(args.embarque, args.puerto, args.flete, args.dolar, args.eta)


if __name__ == '__main__':
    main()
