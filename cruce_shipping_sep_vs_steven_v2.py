"""Cruce v2: una fila por SKU UnionX, con detalle de Steven por embarque concatenado."""
import pandas as pd, sys, re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

UX = "data/planillas/Shipping Plan September V.02.xlsx"
SV = "data/comex/Shipping Plan B-Jun.30th.xls"
OUT = "data/comex/Cruce_Shipping_Sep_vs_StevenB_v2.xlsx"


def parsear_steven(sheet):
    df = pd.read_excel(SV, sheet_name=sheet, header=None)
    items = []
    seccion = None; hdr = None
    for i, row in df.iterrows():
        vals = [v for v in row.values if pd.notna(v)]
        if not vals: continue
        first = str(vals[0]).strip()
        m = re.match(r'^\d+\.(.+)$', first)
        if m and len(vals) <= 2:
            seccion = m.group(1).strip(); continue
        if first == 'No' and len(vals)>1 and 'Model' in str(vals[1]):
            hdr = [str(v).strip() for v in row.values]; continue
        try: int(first)
        except: continue
        if hdr is None or seccion is None: continue
        it = {col: row.values[j] for j,col in enumerate(hdr) if j<len(row.values)}
        it['__seccion'] = seccion
        items.append(it)
    return items

steven = parsear_steven('SHENZHEN') + parsear_steven('NINGBO')

# Index Steven por SKU
from collections import defaultdict
steven_idx = defaultdict(list)
for it in steven:
    sku = str(it.get('SKU','')).strip()
    qty = pd.to_numeric(it.get('Qty'), errors='coerce')
    if not pd.notna(qty): continue
    steven_idx[sku].append({
        'embarque': it['__seccion'],
        'model': str(it.get('Model','')).strip(),
        'qty': int(qty),
    })

# UnionX V02 Detail
ux_sz = pd.read_excel(UX, sheet_name='Detail_SZ')
ux_nb = pd.read_excel(UX, sheet_name='Detail_NB')
ux = pd.concat([ux_sz, ux_nb], ignore_index=True)
ux['SKU'] = ux['SKU'].astype(str).str.strip()
ux['Units'] = pd.to_numeric(ux['Units'], errors='coerce').fillna(0).astype(int)

# Resultados por SKU UnionX
rows = []
for _, r in ux.iterrows():
    sku = str(r['SKU']).strip()
    units_ux = int(r['Units'])
    matches = steven_idx.get(sku, [])

    if not matches:
        rows.append({
            'Cont UnionX': r['Container'],
            'CRD': str(r['CRD'])[:10] if pd.notna(r['CRD']) else '',
            'ETA': str(r['ETA'])[:10] if pd.notna(r['ETA']) else '',
            'Brand': r['Brand'],
            'SKU': sku,
            'Descripción': str(r['Description'])[:55] if pd.notna(r['Description']) else '',
            'Units UnionX': units_ux,
            'Detalle Steven': '(no encontrado)',
            'Qty Steven (total)': 0,
            'Δ (Steven - UnionX)': -units_ux,
            'Estado': '❌ NO EN STEVEN',
        })
        continue

    # Concatenar detalle por embarque
    qty_total_steven = sum(m['qty'] for m in matches)
    detalle = ' | '.join(f"{m['embarque']}: {m['qty']}" for m in matches)
    diff = qty_total_steven - units_ux
    if diff == 0: estado = '✅ OK'
    elif diff > 0: estado = f'🟦 SOBRA {diff}'
    else: estado = f'🔴 FALTA {abs(diff)}'

    rows.append({
        'Cont UnionX': r['Container'],
        'CRD': str(r['CRD'])[:10] if pd.notna(r['CRD']) else '',
        'ETA': str(r['ETA'])[:10] if pd.notna(r['ETA']) else '',
        'Brand': r['Brand'],
        'SKU': sku,
        'Descripción': str(r['Description'])[:55] if pd.notna(r['Description']) else '',
        'Units UnionX': units_ux,
        'Detalle Steven': detalle,
        'Qty Steven (total)': qty_total_steven,
        'Δ (Steven - UnionX)': diff,
        'Estado': estado,
    })

df_res = pd.DataFrame(rows)
print(f"Total SKUs UnionX: {len(df_res)}")
print()

# Imprimir tabla legible
for _, r in df_res.iterrows():
    print(f"{r['Cont UnionX']:<6} {r['SKU']:<18} {r['Brand'][:10]:<10} units_UX={r['Units UnionX']:>5}  →  {r['Estado']:<25}  | {r['Detalle Steven']}")

# Resumen por estado
print(f"\n{'='*80}")
print("Resumen por estado")
print('='*80)
g_estado = df_res.groupby(df_res['Estado'].str.split().str[0]).agg(
    n=('SKU','count'),
    units_ux=('Units UnionX','sum'),
    units_steven=('Qty Steven (total)','sum'),
).reset_index()
print(g_estado.to_string(index=False))

# Generar Excel con formato
AZUL_H = PatternFill('solid', fgColor='FF1F3864')
ROJO   = PatternFill('solid', fgColor='FFE6B8B8')
AZUL_S = PatternFill('solid', fgColor='FFBDD7EE')
VERDE  = PatternFill('solid', fgColor='FFC6EFCE')
GRIS   = PatternFill('solid', fgColor='FFF2F2F2')
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
BLACK_B = Font(bold=True, size=11)
NORMAL = Font(size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cruce UnionX vs Steven"
ws.sheet_view.showGridLines = False

ws.merge_cells('A1:K1')
ws['A1'] = 'CRUCE: Shipping Plan September V.02 (UnionX) vs Shipping Plan B-Jun.30th (Steven)'
ws['A1'].fill = AZUL_H; ws['A1'].font = Font(color='FFFFFFFF', bold=True, size=14)
ws['A1'].alignment = CENTER
ws.row_dimensions[1].height = 28

ws.merge_cells('A2:K2')
ws['A2'] = '34 SKUs UnionX  |  Embarques Steven detectados: SZ-04, SZ-03-08, SZ-01(623), SZ-02(623), IMP OP 350-26, NB-01, Rest items'
ws['A2'].fill = GRIS; ws['A2'].font = Font(italic=True, size=10)
ws['A2'].alignment = CENTER
ws.row_dimensions[2].height = 18

headers = list(df_res.columns)
for j, h in enumerate(headers):
    c = ws.cell(row=4, column=j+1, value=h)
    c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

for i, (_, r) in enumerate(df_res.iterrows()):
    row = 5+i
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=j+1, value=r[h])
        c.font = NORMAL; c.border = BORDER
        if h in ('Units UnionX','Qty Steven (total)','Δ (Steven - UnionX)'):
            c.alignment = RIGHT
            if h == 'Δ (Steven - UnionX)': c.number_format = '+#,##0;-#,##0;0'
            else: c.number_format = '#,##0'
        elif h == 'Estado':
            c.alignment = CENTER; c.font = BLACK_B
            est = str(r[h])
            if '❌' in est or 'FALTA' in est: c.fill = ROJO
            elif 'SOBRA' in est: c.fill = AZUL_S
            elif '✅' in est: c.fill = VERDE
        else:
            c.alignment = LEFT if h == 'Descripción' or h == 'Detalle Steven' else CENTER

# Anchos
widths = {'A':8,'B':11,'C':11,'D':10,'E':17,'F':56,'G':10,'H':50,'I':10,'J':10,'K':18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = 'A5'

wb.save(OUT)
print(f"\n[OK] Excel: {OUT}")
