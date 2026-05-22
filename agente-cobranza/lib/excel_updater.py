"""
Actualiza hojas específicas de un Excel preservando las que NO se tocan
(las que tienen fórmulas, tablas dinámicas, data manual del usuario).

Diseño:
  - Lee el archivo existente con openpyxl (`keep_vba=False` porque son xlsx).
  - Para cada hoja a actualizar:
      * Si existe, la borra y crea nueva con el mismo nombre.
      * Escribe las filas de data.
  - Para cada hoja en `hojas_preservar`: la deja intacta.
  - Aplica fórmulas XLOOKUP opcionales después del rewrite.
  - Guarda en una nueva ruta (no sobrescribe el input, así si algo falla
    el original queda intacto).

LIMITACIÓN CONOCIDA (recálculo de fórmulas):
  En Linux (GitHub Actions) no podemos usar `win32com.client.DispatchEx` que
  hace recalcular Excel desde COM (es Windows-only). Quedan 2 escenarios:

    a) Excel/Drive recalcula al abrir el archivo → la mayoría de las fórmulas
       quedan OK porque la próxima vez que un humano abre el Excel se recalcula.
    b) Para fórmulas que necesitan recalc inmediato → usar fallback con
       LibreOffice headless (TODO si surge necesidad).

  Por ahora confiamos en (a). Si Andrés reporta que alguna fórmula no
  recalcula bien, agregamos LibreOffice como step del workflow.
"""
from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def actualizar_excel(
    path_original: Path,
    path_destino: Path,
    hojas_data: dict[str, list[list[Any]]],
    hojas_preservar: list[str] | None = None,
    xlookup_setup: list[dict] | None = None,
) -> Path:
    """Actualiza un Excel reemplazando hojas específicas, preservando las demás.

    Args:
      path_original: archivo Excel base (no se modifica).
      path_destino:  donde guardar el resultado.
      hojas_data:    dict {nombre_hoja: filas} — cada nombre se REEMPLAZA.
                     `filas[0]` debe ser el header.
      hojas_preservar: nombres de hojas que NO se tocan (fórmulas, dinámicas).
                       Si una hoja no está acá NI en hojas_data, queda igual igualmente
                       (sólo borramos las que vamos a reescribir).
      xlookup_setup: lista de dicts con `{hoja, columna, formula}` para setear
                     fórmulas XLOOKUP después del update.

    Returns: path_destino
    """
    hojas_preservar = hojas_preservar or []
    xlookup_setup = xlookup_setup or []

    # Copia el archivo original a destino, después modificamos destino
    path_original = Path(path_original)
    path_destino = Path(path_destino)
    path_destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(path_original, path_destino)

    wb = load_workbook(path_destino)

    # Lista de hojas ANTES de borrar (para preservar orden)
    hojas_originales = list(wb.sheetnames)

    # Reemplazar las hojas que el config pide actualizar
    for nombre_hoja, filas in hojas_data.items():
        # Si la hoja existe, borrarla
        if nombre_hoja in wb.sheetnames:
            del wb[nombre_hoja]
        # Crear nueva al final (después reordenamos)
        ws = wb.create_sheet(title=nombre_hoja)
        for row in filas:
            ws.append(row)
        # Formato basic: bold en header
        if filas:
            from openpyxl.styles import Font
            for cell in ws[1]:
                cell.font = Font(name=cell.font.name or "Calibri",
                                  size=cell.font.size or 11,
                                  bold=True)

    # Reordenar hojas para mantener el orden original (las nuevas van al final
    # de lo que existía; las preservadas quedan en sus posiciones originales)
    nuevo_orden = []
    for nombre in hojas_originales:
        if nombre in wb.sheetnames:
            nuevo_orden.append(nombre)
    # Agregar las hojas nuevas (data) que no estaban en originales
    for nombre in hojas_data:
        if nombre not in nuevo_orden and nombre in wb.sheetnames:
            nuevo_orden.append(nombre)

    # openpyxl no tiene método directo, usamos move_sheet
    for idx, nombre in enumerate(nuevo_orden):
        actual_idx = wb.sheetnames.index(nombre)
        if actual_idx != idx:
            wb.move_sheet(nombre, offset=idx - actual_idx)

    # ─── Aplicar XLOOKUP customizado por cliente ──────────────────────────
    for lookup in xlookup_setup:
        nombre_hoja = lookup.get("hoja")
        columna = lookup.get("columna")
        formula_tpl = lookup.get("formula")
        if not (nombre_hoja and columna and formula_tpl):
            continue
        if nombre_hoja not in wb.sheetnames:
            print(f"[WARN] XLOOKUP: hoja '{nombre_hoja}' no existe, skip", flush=True)
            continue
        ws: Worksheet = wb[nombre_hoja]
        # Aplicar fórmula a cada fila de data (skip header)
        max_row = ws.max_row
        for fila in range(2, max_row + 1):
            formula = formula_tpl.replace("{fila}", str(fila))
            ws[f"{columna}{fila}"] = formula

    wb.save(path_destino)
    return path_destino
