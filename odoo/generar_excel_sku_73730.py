"""
Genera Excel con el detalle de movimientos del SKU 73730 en 2026.
Hojas: Resumen, Ventas, Traslados, Otros, Detalle completo.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'archive', 'Junior Revenue'))
from odoo_connection import OdooConnection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

odoo = OdooConnection(
    url="https://unionxb2b.odoo.com",
    username="andres@grupoeter.cl",
    password="ROTATED-2026-05-07",
    db_name="bmya-innovatek-sh-prd-6981800"
)
odoo.login()

SKU = "73730"
DESDE = "2026-01-01 00:00:00"
HASTA = "2026-05-05 23:59:59"
OUTPUT = r"G:\Mi unidad\TRABAJO\RESPALDO\OPERACIONES\UNION X - IA\data\outputs\Movimiento_SKU_73730_2026.xlsx"

prod = odoo.search_read("product.product", [["default_code", "=", SKU]],
                        ["id", "default_code", "name"])[0]

# Locations CA1
ca1_locs = odoo.search_read(
    "stock.location",
    [["warehouse_id", "=", 1], ["usage", "=", "internal"]],
    ["id", "complete_name"]
)
ca1_internal_ids = {l['id'] for l in ca1_locs}
ca1_stock_ids = {l['id'] for l in ca1_locs
                  if l['complete_name'] == 'CA1/Stock' or l['complete_name'].startswith('CA1/Stock/')}

# Stock actual
stock_actual_ca1total = sum(q['quantity'] for q in odoo.search_read(
    "stock.quant",
    [["product_id", "=", prod['id']], ["location_id", "in", list(ca1_internal_ids)]],
    ["quantity"]))
stock_actual_ca1stock = sum(q['quantity'] for q in odoo.search_read(
    "stock.quant",
    [["product_id", "=", prod['id']], ["location_id", "in", list(ca1_stock_ids)]],
    ["quantity"]))

# Movimientos
moves = odoo.search_read(
    "stock.move",
    [["product_id", "=", prod['id']], ["state", "=", "done"],
     ["date", ">=", DESDE], ["date", "<=", HASTA],
     "|", ["location_id", "in", list(ca1_internal_ids)],
     ["location_dest_id", "in", list(ca1_internal_ids)]],
    ["id", "date", "product_uom_qty", "quantity", "reference", "origin",
     "location_id", "location_dest_id", "picking_type_id"], limit=5000)

loc_ids = list({m['location_id'][0] for m in moves} | {m['location_dest_id'][0] for m in moves})
locs_info = {l['id']: l for l in odoo.read("stock.location", loc_ids,
                                           ["id", "complete_name", "usage"])}
ptype_ids = list({m['picking_type_id'][0] for m in moves if m.get('picking_type_id')})
ptypes = {pt['id']: pt for pt in odoo.read("stock.picking.type", ptype_ids,
                                            ["id", "name", "code"])}

ventas, traslados_out, otros_out = [], [], []
compras, traslados_in, otros_in = [], [], []

for m in moves:
    qty = m.get('quantity') or m.get('product_uom_qty') or 0.0
    src_id, dst_id = m['location_id'][0], m['location_dest_id'][0]
    if src_id in ca1_internal_ids and dst_id in ca1_internal_ids:
        continue
    src, dst = locs_info.get(src_id, {}), locs_info.get(dst_id, {})
    pt = ptypes.get(m['picking_type_id'][0]) if m.get('picking_type_id') else None
    info = {
        "date": m['date'][:10],
        "qty": qty,
        "ref": m.get('reference') or '',
        "origin": m.get('origin') or '',
        "src": src.get('complete_name'),
        "dst": dst.get('complete_name'),
        "pt": pt['name'] if pt else '',
    }
    if src_id in ca1_internal_ids:
        if dst.get('usage') == 'customer': ventas.append(info)
        elif dst.get('usage') == 'internal': traslados_out.append(info)
        else: otros_out.append(info)
    else:
        if src.get('usage') == 'supplier': compras.append(info)
        elif src.get('usage') == 'internal': traslados_in.append(info)
        else: otros_in.append(info)

def total(lst): return sum(x['qty'] for x in lst)
total_ent = total(compras) + total(traslados_in) + total(otros_in)
total_sal = total(ventas) + total(traslados_out) + total(otros_out)
stock_inicial = stock_actual_ca1total - (total_ent - total_sal)

# ============ EXCEL ============
wb = Workbook()
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

def autosize(ws):
    for col in ws.columns:
        try:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, 60)
        except Exception:
            pass

# Hoja 1: Resumen
ws = wb.active
ws.title = "Resumen"
ws["A1"] = f"Movimiento SKU {SKU} — {prod['name']}"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:C1")
ws["A2"] = "Bodega: CA1 (Carrascal N°9-10) | Período: 01/01/2026 — 05/05/2026"
ws["A2"].font = Font(italic=True, color="595959")
ws.merge_cells("A2:C2")

ws["A4"] = "Concepto"; ws["B4"] = "Unidades"; ws["C4"] = "Detalle"
style_header_row(ws, 4, 3)

filas = [
    ("Stock inicial al 01/01/2026", stock_inicial, "CA1 total (Stock + Input + Output)"),
    ("(+) Entradas por compra (proveedor)", total(compras), "Recepciones desde proveedor"),
    ("(+) Entradas desde otras bodegas", total(traslados_in), "Traslados internos hacia CA1"),
    ("(+) Ajustes positivos / producción", total(otros_in), "Inventory adjustments y otros"),
    ("(−) Ventas a clientes", -total(ventas), "Delivery Orders → Customers"),
    ("(−) Traslados a otras bodegas", -total(traslados_out), "Picking → otra warehouse"),
    ("(−) Devolución proveedor / ajustes", -total(otros_out), "Devoluciones y ajustes negativos"),
    ("Stock actual CA1 total", stock_actual_ca1total, f"De los cuales en CA1/Stock: {stock_actual_ca1stock:g}"),
]
for i, (concepto, qty, det) in enumerate(filas, start=5):
    ws.cell(row=i, column=1, value=concepto).font = SUB_FONT if "Stock" in concepto else Font()
    ws.cell(row=i, column=2, value=qty).alignment = Alignment(horizontal="right")
    ws.cell(row=i, column=3, value=det)
    if "Stock" in concepto:
        for c in range(1, 4):
            ws.cell(row=i, column=c).fill = TOTAL_FILL
    for c in range(1, 4):
        ws.cell(row=i, column=c).border = BORDER

ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 50

# Hojas de detalle
def hoja_detalle(nombre, lista, columnas):
    ws = wb.create_sheet(nombre)
    headers = list(columnas.keys())
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))
    for i, item in enumerate(sorted(lista, key=lambda z: z['date']), start=2):
        for j, key in enumerate(columnas.values(), start=1):
            ws.cell(row=i, column=j, value=item.get(key, ''))
            ws.cell(row=i, column=j).border = BORDER
    # Total
    if lista:
        last = len(lista) + 2
        ws.cell(row=last, column=1, value="TOTAL").font = SUB_FONT
        ws.cell(row=last, column=2, value=sum(x['qty'] for x in lista)).font = SUB_FONT
        for c in range(1, len(headers) + 1):
            ws.cell(row=last, column=c).fill = TOTAL_FILL
            ws.cell(row=last, column=c).border = BORDER
    autosize(ws)

cols_estandar = {
    "Fecha": "date", "Cantidad": "qty", "Referencia": "ref",
    "Pedido / Origen": "origin", "Origen": "src", "Destino": "dst", "Tipo": "pt",
}

hoja_detalle("Ventas a clientes", ventas, cols_estandar)
hoja_detalle("Traslados a otras bodegas", traslados_out, cols_estandar)
hoja_detalle("Compras a proveedor", compras, cols_estandar)
hoja_detalle("Ajustes y devoluciones", otros_out + otros_in, cols_estandar)

# Hoja completa con todos los movimientos
ws_all = wb.create_sheet("Detalle completo")
headers_all = ["Fecha", "Categoría", "Cantidad", "Referencia", "Pedido", "Origen", "Destino", "Tipo"]
for j, h in enumerate(headers_all, start=1):
    ws_all.cell(row=1, column=j, value=h)
style_header_row(ws_all, 1, len(headers_all))

todos = []
for x in ventas: todos.append(("Venta cliente", x))
for x in traslados_out: todos.append(("Traslado salida", x))
for x in otros_out: todos.append(("Otro salida", x))
for x in compras: todos.append(("Compra proveedor", x))
for x in traslados_in: todos.append(("Traslado entrada", x))
for x in otros_in: todos.append(("Otro entrada", x))

for i, (cat, x) in enumerate(sorted(todos, key=lambda z: z[1]['date']), start=2):
    ws_all.cell(row=i, column=1, value=x['date'])
    ws_all.cell(row=i, column=2, value=cat)
    ws_all.cell(row=i, column=3, value=x['qty'])
    ws_all.cell(row=i, column=4, value=x['ref'])
    ws_all.cell(row=i, column=5, value=x['origin'])
    ws_all.cell(row=i, column=6, value=x['src'])
    ws_all.cell(row=i, column=7, value=x['dst'])
    ws_all.cell(row=i, column=8, value=x['pt'])
    for c in range(1, len(headers_all) + 1):
        ws_all.cell(row=i, column=c).border = BORDER
autosize(ws_all)

wb.save(OUTPUT)
print(f"Excel generado: {OUTPUT}")
