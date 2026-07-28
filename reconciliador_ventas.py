# -*- coding: utf-8 -*-
"""
Reconciliador de ventas Odoo ↔ RAW (sale.order-based), semi-automático.

A diferencia del enfoque por factura, reconcilia a nivel SALE.ORDER, así cubre
también el FULFILLMENT (Yuju: ML/Falabella/Ripley/Paris/Walmart), que NO genera
boleta y entra al RAW por el pedido.

Detecta órdenes confirmadas (state sale/done) del período que NO están en el RAW
(ni por name 'CANAL número' ni por client_order_ref), y las prepara para overlay.

GUARDRAILS (regla de Andrés):
  - Solo escribe el HISTÓRICO congelado (mes <= --freeze). NUNCA mes_actual (pulso).
  - NO toca scripts del pulso, sus descargables ni la reportería: solo AGREGA filas.
  - Dry-run por defecto (reporte + Excel de candidatos). --apply para overlayar.

Uso:
  python reconciliador_ventas.py --desde 2026-06-01 --hasta 2026-06-30
  python reconciliador_ventas.py --desde 2026-06-01 --hasta 2026-06-30 --apply --freeze 6
"""
import argparse, os, re, shutil
from pathlib import Path
import pandas as pd
import xmlrpc.client

ROOT = Path(__file__).resolve().parent
HIST = ROOT / "data" / "historico" / "ventas_historico.parquet"
MES = ROOT / "data" / "historico" / "ventas_mes_actual.parquet"
DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


def _odoo():
    for p in (".env", "eerr-finanzas/.env"):
        f = ROOT / p
        if f.exists():
            for ln in f.read_text(encoding="utf-8").splitlines():
                if "=" in ln and not ln.startswith("#"):
                    k, v = ln.split("=", 1); os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
    URL, DB, U = "https://unionxb2b.odoo.com", "bmya-innovatek-sh-prd-6981800", "andres@grupoeter.cl"
    PWD = os.environ["ANDRES_ODOO_PASSWORD"]
    uid = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/common").authenticate(DB, U, PWD, {})
    return xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/object"), DB, uid, PWD


def _matriz():
    import openpyxl
    wb = openpyxl.load_workbook(ROOT / "data" / "planillas" / "Matriz productos.xlsx", read_only=True, data_only=True)
    ws = wb["Productos"] if "Productos" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True)); idx = {str(h).strip() if h else "": i for i, h in enumerate(rows[0])}
    iS = idx.get("SKU"); m = {}
    for r in rows[1:]:
        if iS is None or len(r) <= iS or not r[iS]:
            continue
        g = lambda c: (r[idx[c]] if c in idx and idx[c] < len(r) and r[idx[c]] not in (None, "") else "")
        m[str(r[iS]).strip().lower()] = {"producto": g("Producto"), "categoria_macro": g("Categoría macro"),
            "categoria_padre": g("Categoría padre"), "categoria_hijo": g("Categoría hijo"),
            "categoria_comercial": g("Categoría comercial"), "marca": g("Marca"), "proveedor": g("Proveedor"),
            "pack": g("Pack"), "estado_sku": g("In/out")}
    return m


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--desde", required=True); ap.add_argument("--hasta", required=True)
    ap.add_argument("--apply", action="store_true"); ap.add_argument("--freeze", type=int, default=6)
    a = ap.parse_args()
    import duckdb
    from clasificar_marca import clasificar_tipo_marca
    fmt = lambda v: f"${v:,.0f}".replace(",", ".")
    M, DB, uid, PWD = _odoo()

    # 1) órdenes confirmadas del período
    print(f"[Odoo] sale.order confirmadas {a.desde}..{a.hasta}...")
    so, off = [], 0
    while True:
        b = M.execute_kw(DB, uid, PWD, "sale.order", "search_read",
            [[["state", "in", ["sale", "done"]], ["date_order", ">=", a.desde + " 00:00:00"],
              ["date_order", "<=", a.hasta + " 23:59:59"]]],
            {"fields": ["id", "name", "client_order_ref", "channel", "channel_order_reference",
                        "date_order", "amount_total", "partner_id"],
             "limit": 2000, "offset": off})
        if not b:
            break
        so += b; off += 2000
        if len(b) < 2000:
            break
    print(f"  {len(so)} órdenes")

    # 2) tokens de pedido en el RAW (año completo)
    con = duckdb.connect()
    def toks(P):
        s = con.execute(f"SELECT string_agg(CAST(pedido AS VARCHAR),'|') FROM read_parquet('{P.as_posix()}') "
                        f"WHERE CAST(anio_venta AS INT)=2026").fetchone()[0] or ""
        return set(re.split(r"[|,;\s]+", s))
    pedset = toks(HIST) | toks(MES)

    # 3) ausentes: ni name ni client_order_ref ni sus tokens en el RAW
    def presente(o):
        cand = [str(o.get("name") or ""), str(o.get("client_order_ref") or "")]
        cand += re.split(r"[,\s]+", str(o.get("name") or "")) + re.split(r"[,\s]+", str(o.get("client_order_ref") or ""))
        return any(c and c in pedset for c in cand)
    miss = [o for o in so if float(o.get("amount_total") or 0) > 0 and not presente(o)]
    print(f"  ausentes del RAW: {len(miss)} · {fmt(sum(o['amount_total'] for o in miss))}")

    # canal: se deriva del PARTNER vía Maestra Canales (Empresa→Canal), igual que el
    # extract. NO del campo 'channel' de Odoo (ese trae 'Falabella Chile' y no calza).
    import openpyxl
    mc = openpyxl.load_workbook(ROOT / "data" / "planillas" / "Maestra Canales.xlsx", read_only=True, data_only=True).active
    mcrows = list(mc.iter_rows(values_only=True)); mci = {str(h).strip(): i for i, h in enumerate(mcrows[0])}
    _emp, _can = mci.get("Empresa", 0), mci.get("Canal", 1)
    empresa2canal = {}
    for r in mcrows[1:]:
        if r[_emp] and r[_can]:
            empresa2canal[re.sub(r"\s+", " ", str(r[_emp]).strip().lower())] = str(r[_can]).strip()

    def canal_de(o):
        nm = str(o.get("name") or "").upper()
        if nm.startswith("KITCHEN"):
            return "Kitchen Center"
        if "POLAR" in nm:
            return "Abc"
        partner = re.sub(r"\s+", " ", str((o.get("partner_id") or [0, ""])[1]).strip().lower())
        c = empresa2canal.get(partner)
        if c and c != "Web":
            return c
        if c == "Web" or o.get("channel") == "Web":
            # web → prefijo del channel_order_reference (igual que _estandarizar_canal del extract)
            ref = str(o.get("channel_order_reference") or "").upper()
            if ref.startswith("LH"):
                return "Lhotse web"
            if ref.startswith("SH"):
                return "Simplit web"
            return "UnionX web"
        return f"(REVISAR: {(o.get('partner_id') or [0, ''])[1]})"
    import collections
    by = collections.defaultdict(lambda: [0, 0.0])
    for o in miss:
        c = canal_de(o); by[c][0] += 1; by[c][1] += o["amount_total"]
    print("\n  ausentes por canal:")
    for c, (n, v) in sorted(by.items(), key=lambda x: -x[1][1]):
        print(f"    {str(c):<22} {n:>4} · {fmt(v)}")

    if not miss:
        return
    # 4) construir filas (solo del subconjunto ausente)
    oids = [o["id"] for o in miss]; o_by_id = {o["id"]: o for o in miss}
    lines = []
    for i in range(0, len(oids), 500):
        lines += M.execute_kw(DB, uid, PWD, "sale.order.line", "search_read",
            [[["order_id", "in", oids[i:i+500]], ["product_uom_qty", ">", 0]]],
            {"fields": ["order_id", "product_id", "product_uom_qty", "price_total", "price_subtotal", "purchase_price"]})
    pids = list({l["product_id"][0] for l in lines if l.get("product_id")})
    prod = {}
    for i in range(0, len(pids), 500):
        for p in M.execute_kw(DB, uid, PWD, "product.product", "read", [pids[i:i+500]], {"fields": ["default_code", "name"]}):
            prod[p["id"]] = p
    matriz = _matriz()
    rows = []
    for l in lines:
        o = o_by_id[l["order_id"][0]]; canal = canal_de(o)
        pedido = str(o["name"]) if canal in ("Kitchen Center", "Abc") else str(o.get("client_order_ref") or o["name"])
        fv = str(o["date_order"])[:10]; fdt = pd.to_datetime(fv)
        pid = l["product_id"][0] if l.get("product_id") else None
        sku = str((prod.get(pid) or {}).get("default_code") or "").strip()
        mat = matriz.get(sku.lower(), {}); marca = mat.get("marca", "")
        qty = float(l["product_uom_qty"]); vb = float(l["price_total"]); vn = float(l["price_subtotal"])
        cu = float(l.get("purchase_price") or 0); ct = cu * qty
        rows.append({"tipo_movimiento": "Venta", "bodega": "", "documento": "", "fecha_documento": fv,
            "pedido": pedido, "estado_pedido": "sale", "tipo_despacho": "", "sku": sku, "canal": canal,
            "fecha_venta": fv, "hora_venta": "", "producto": mat.get("producto", ""),
            "categoria_macro": mat.get("categoria_macro", ""), "categoria_padre": mat.get("categoria_padre", ""),
            "categoria_hijo": mat.get("categoria_hijo", ""), "categoria_comercial": mat.get("categoria_comercial", ""),
            "estado_sku": str(mat.get("estado_sku", "")).lower(), "pack": mat.get("pack", "No"), "marca": marca,
            "proveedor": mat.get("proveedor", ""), "tipo_marca": clasificar_tipo_marca(marca),
            "tipo_compra": "Importación", "tipo_negocio": "Marketplace", "kam": "", "estado_canal": "In",
            "anio_venta": fdt.year, "mes_venta": fdt.month, "semana_venta": int(fdt.isocalendar().week),
            "dia_semana": DIAS[fdt.weekday()], "hora_venta_num": 0, "cantidad": qty, "venta_bruta": vb,
            "costo_unitario": cu, "costo_total": ct, "margen_front": vn - ct, "comision_pct": 0.0,
            "comision": 0.0, "logistica": 0.0, "marketing": 0.0, "margen_final": vn - ct, "venta_neta": vn})
    nuevas = pd.DataFrame(rows)
    hist = pd.read_parquet(HIST); mes = pd.read_parquet(MES)
    key = lambda df: set(zip(df["pedido"].astype(str), df["sku"].astype(str), df["venta_bruta"].round().astype("Int64").astype(str)))
    existe = key(hist) | key(mes)
    nuevas["_k"] = list(zip(nuevas["pedido"].astype(str), nuevas["sku"].astype(str), nuevas["venta_bruta"].round().astype("Int64").astype(str)))
    add = nuevas[~nuevas["_k"].isin(existe)].drop(columns="_k")
    # No overlayar canales sin resolver (REVISAR): quedan en el Excel para revisión manual.
    revisar = add[add["canal"].astype(str).str.startswith("(REVISAR")]
    if len(revisar):
        print(f"  ⚠️  {len(revisar)} líneas '(REVISAR)' NO se overlayan (quedan en el Excel): {fmt(revisar['venta_bruta'].sum())}")
    add_ok = add[~add["canal"].astype(str).str.startswith("(REVISAR")]
    to_hist = add_ok[add_ok["mes_venta"] <= a.freeze]
    print(f"\n  líneas nuevas: {len(add)} · a histórico (mes<={a.freeze}): {len(to_hist)} · {fmt(to_hist['venta_bruta'].sum())}")
    out = ROOT / "data" / "outputs" / f"reconciliacion_ventas_{a.desde}_{a.hasta}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        add.to_excel(w, index=False, sheet_name="candidatos")
    print(f"  candidatos → {out.relative_to(ROOT)}")
    if a.apply and len(to_hist):
        base = pd.read_parquet(HIST); ex = to_hist.reindex(columns=base.columns)
        if "es_despacho" in base.columns:
            ex["es_despacho"] = False
        bak = HIST.with_suffix(".parquet.bak_reconvta")
        if not bak.exists():
            shutil.copy2(HIST, bak)
        pd.concat([base, ex], ignore_index=True).to_parquet(HIST, index=False)
        print(f"  ✓ overlay {len(ex)} filas al histórico (backup {bak.name}) — NO se tocó mes_actual/pulso")
    else:
        print("  → dry-run (revisar candidatos; --apply para overlayar histórico)")


if __name__ == "__main__":
    main()
