"""
Lee el Excel aprobado, aplica cambios en Odoo y confirma la factura.
Usa xmlrpc directo para write() ya que execute_kw espera args=[ids, values].

Salvaguardas (auditoría 2026-07):
- Antes de escribir se re-verifica en Odoo: move en draft, line_ids pertenecen
  al move y siguen en 42410104. Nada se escribe si algo no calza.
- El estado posted se detecta leyendo `state` (no por substring del error).
- Solo se postea si TODAS las líneas catchall del move tienen decisión válida.
- Código de cuenta inválido en columna M → error de bloque (no se ignora).
- El RUT del proveedor viaja en la fila separadora → memoria por bloque.
"""
from __future__ import annotations
import xmlrpc.client
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
import openpyxl
from .clasificador import CUENTAS_DESTINO
from .memoria import registrar_aprobacion, registrar_correccion

SEP_MARKER = "▼FACTURA"
COL_LINE_ID = 2; COL_GLOSA = 3; COL_CTA_PROP_COD = 6
COL_APROBADO = 12; COL_CTA_CORRECTA = 13
COL_SEP_MOVE_ID = 2; COL_SEP_RUT = 13
CUENTA_CATCHALL_ID = 1377

CODIGO_A_INFO  = {info["codigo"]: info for info in CUENTAS_DESTINO.values()}
CODIGO_A_CLAVE = {info["codigo"]: clave for clave, info in CUENTAS_DESTINO.items()}
CODIGO_CATCHALL = "42410104"


def _models(client):
    uid = client.authenticate()
    proxy = xmlrpc.client.ServerProxy(
        f"{client.url}/xmlrpc/2/object", allow_none=True,
        transport=xmlrpc.client.SafeTransport())
    return proxy, uid


def _odoo_write(client, model: str, ids: list, values: dict) -> bool:
    proxy, uid = _models(client)
    return proxy.execute_kw(client.db, uid, client.password,
                            model, "write", [ids, values], {})


def _odoo_action_post(client, move_id: int) -> dict:
    proxy, uid = _models(client)
    return proxy.execute_kw(client.db, uid, client.password,
                            "account.move", "action_post", [[move_id]], {})


def _leer_estado_move(client, move_id: int) -> Optional[dict]:
    """Lee estado real del move + sus líneas catchall vigentes. None si no existe."""
    proxy, uid = _models(client)
    moves = proxy.execute_kw(client.db, uid, client.password,
        "account.move", "read", [[move_id]], {"fields": ["state", "line_ids", "name"]})
    if not moves:
        return None
    move = moves[0]
    lineas = proxy.execute_kw(client.db, uid, client.password,
        "account.move.line", "search_read",
        [[["move_id", "=", move_id], ["display_type", "in", [False, "product"]]]],
        {"fields": ["id", "account_id"]})
    move["catchall_ids"] = {l["id"] for l in lineas
                            if l.get("account_id") and l["account_id"][0] == CUENTA_CATCHALL_ID}
    move["all_line_ids"] = {l["id"] for l in lineas}
    return move


@dataclass
class LineaAplicada:
    line_id: int; glosa: str; cuenta_final_codigo: str
    cuenta_final_odoo_id: int; fue_corregida: bool; cambio_cuenta: bool


@dataclass
class FacturaAplicada:
    move_id: int
    lineas: list[LineaAplicada] = field(default_factory=list)
    confirmada: bool = False
    errores: list[str] = field(default_factory=list)


@dataclass
class ResultadoAplicacion:
    facturas: list[FacturaAplicada] = field(default_factory=list)
    errores_globales: list[str] = field(default_factory=list)
    dry_run: bool = True
    partner_rut: Optional[str] = None
    partner_nombre: Optional[str] = None

    @property
    def ok(self) -> bool:
        return not self.errores_globales and all(not f.errores for f in self.facturas)

    @property
    def total_lineas(self) -> int:
        return sum(len(f.lineas) for f in self.facturas)


def _leer_excel(ruta) -> list[dict]:
    """
    Parsea el Excel de respuesta. Bloques: {move_id, rut, nombre, lineas, invalidas}.
    - Filas sin decisión (APROBADO vacío) se ignoran (incluye filas grises informativas).
    - NO sin código válido en col M, o código desconocido → va a `invalidas` (bloquea post).
    """
    wb = openpyxl.load_workbook(ruta)
    if "Propuesta" not in wb.sheetnames:
        raise ValueError("El Excel no tiene hoja 'Propuesta' — no es un archivo de distribución")
    ws = wb["Propuesta"]
    bloques = []; bloque = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] is None:
            continue
        marker = str(row[0] or "").strip()
        if marker == SEP_MARKER:
            if bloque:
                bloques.append(bloque)
            try:
                move_id = int(row[COL_SEP_MOVE_ID - 1])
            except (TypeError, ValueError):
                move_id = 0
            rut = str(row[COL_SEP_RUT - 1] or "").strip()
            nombre = str(row[COL_GLOSA - 1] or "").strip()
            bloque = {"move_id": move_id, "rut": rut, "nombre": nombre,
                      "lineas": [], "invalidas": []}
            continue
        if bloque is None:
            continue
        try:
            line_id = int(row[COL_LINE_ID - 1])
        except (TypeError, ValueError):
            continue
        glosa = str(row[COL_GLOSA - 1] or "").strip()
        aprobado_raw = str(row[COL_APROBADO - 1] or "").strip().upper()
        cuenta_prop = str(row[COL_CTA_PROP_COD - 1] or "").strip()
        cuenta_corr = str(row[COL_CTA_CORRECTA - 1] or "").strip().split(".")[0]
        if aprobado_raw not in ("SI", "SÍ", "S", "NO"):
            continue  # sin decisión (o fila informativa gris)
        aprobado = aprobado_raw in ("SI", "SÍ", "S")
        cuenta_final = cuenta_prop if aprobado else cuenta_corr

        if cuenta_final not in CODIGO_A_INFO and cuenta_final != CODIGO_CATCHALL:
            bloque["invalidas"].append(
                f"línea {line_id}: {'NO sin código válido en col M' if not aprobado else 'cuenta propuesta inválida'}"
                f" ('{cuenta_final or '—'}')")
            continue

        info = CODIGO_A_INFO.get(cuenta_final, {})
        odoo_id = info.get("odoo_id", 0)
        cambio = (odoo_id != CUENTA_CATCHALL_ID and odoo_id != 0)
        bloque["lineas"].append(LineaAplicada(
            line_id=line_id, glosa=glosa, cuenta_final_codigo=cuenta_final,
            cuenta_final_odoo_id=odoo_id, fue_corregida=not aprobado,
            cambio_cuenta=cambio))
    if bloque:
        bloques.append(bloque)
    return bloques


def _aplicar_bloque(odoo_client, move_id, lineas, dry_run, invalidas=None):
    fa = FacturaAplicada(move_id=move_id)
    invalidas = invalidas or []

    # ── Verificación de estado real en Odoo ─────────────────────────────
    estado = _leer_estado_move(odoo_client, move_id)
    if estado is None:
        fa.errores.append(f"Move {move_id} no existe en Odoo")
        print(f"    ✗ Move {move_id} no existe en Odoo")
        return fa
    if estado["state"] == "posted":
        fa.confirmada = True
        print(f"    ℹ️  Factura {estado.get('name', move_id)} ya está confirmada (posted) — skip")
        return fa
    if estado["state"] != "draft":
        fa.errores.append(f"Move {move_id} en estado '{estado['state']}' — no se toca")
        print(f"    ⚠️  Factura {move_id} en estado '{estado['state']}' — skip")
        return fa

    fuera_del_move = [l.line_id for l in lineas if l.line_id not in estado["all_line_ids"]]
    if fuera_del_move:
        fa.errores.append(f"Líneas {fuera_del_move} no pertenecen al move {move_id} "
                          "(¿Excel reordenado o desactualizado?) — bloque abortado")
        print(f"    ✗ {fa.errores[-1]}")
        return fa

    lineas_a_cambiar = [l for l in lineas if l.cambio_cuenta and l.cuenta_final_odoo_id
                        and l.line_id in estado["catchall_ids"]]
    ya_movidas = [l for l in lineas if l.cambio_cuenta and l.line_id not in estado["catchall_ids"]]
    for l in lineas:
        if l in ya_movidas:
            print(f"    ℹ️  Línea {l.line_id}: ya no está en 42410104 — skip")
        elif l.cambio_cuenta:
            print(f"    → Línea {l.line_id}: → {l.cuenta_final_codigo}")
        else:
            print(f"    ≡ Línea {l.line_id}: mantiene 42410104")
        fa.lineas.append(l)

    if lineas_a_cambiar and not dry_run:
        try:
            orm_commands = [(1, l.line_id, {"account_id": l.cuenta_final_odoo_id})
                            for l in lineas_a_cambiar]
            _odoo_write(odoo_client, "account.move", [move_id], {"line_ids": orm_commands})
            for l in lineas_a_cambiar:
                print(f"    ✓ Línea {l.line_id}: → {l.cuenta_final_odoo_id} ({l.cuenta_final_codigo})")
        except Exception as e:
            msg = f"Error escribiendo en factura {move_id}: {e}"
            print(f"    ✗ {msg}"); fa.errores.append(msg)
            return fa
    elif lineas_a_cambiar:
        for l in lineas_a_cambiar:
            print(f"    [DRY] Línea {l.line_id}: → {l.cuenta_final_codigo}")

    # ── Post solo si el bloque quedó completamente resuelto ─────────────
    decididas = {l.line_id for l in lineas}
    pendientes = estado["catchall_ids"] - decididas - {l.line_id for l in lineas_a_cambiar}
    # las que se mantienen en 42410104 con SI explícito cuentan como resueltas
    mantenidas = {l.line_id for l in lineas if not l.cambio_cuenta}
    pendientes -= mantenidas

    if invalidas:
        fa.errores.append(f"Decisiones inválidas: {'; '.join(invalidas)} — factura NO confirmada")
        print(f"    ⚠️  {fa.errores[-1]}")
    elif pendientes:
        print(f"    ⚠️  Factura {move_id}: {len(pendientes)} línea(s) en 42410104 sin decisión "
              f"— se aplicaron los cambios pero NO se confirma")
    elif fa.errores:
        print(f"    ⚠️  Factura {move_id} NO confirmada por errores")
    elif move_id:
        if not dry_run:
            try:
                _odoo_action_post(odoo_client, move_id)
                fa.confirmada = True
                print(f"    ✓ Factura {move_id} CONFIRMADA (posted)")
            except Exception as e:
                msg = f"Error confirmando {move_id}: {e}"
                print(f"    ✗ {msg}"); fa.errores.append(msg)
        else:
            fa.confirmada = True
            print(f"    [DRY] Factura {move_id} sería confirmada")
    return fa


def aplicar_distribucion(odoo_client, ruta_excel, aprobado_por="analista@unionx.cl",
                          dry_run=True, directorio_memoria=None) -> ResultadoAplicacion:
    dir_memo = Path(directorio_memoria) if directorio_memoria else None
    try:
        bloques = _leer_excel(ruta_excel)
    except Exception as e:
        resultado = ResultadoAplicacion(dry_run=dry_run)
        resultado.errores_globales.append(f"Excel ilegible: {e}")
        return resultado

    resultado = ResultadoAplicacion(dry_run=dry_run)
    if not bloques:
        resultado.errores_globales.append("El Excel no tiene bloques de factura válidos.")
        return resultado

    # rut del primer bloque para el resumen externo (compatibilidad)
    resultado.partner_rut = bloques[0].get("rut") or None
    resultado.partner_nombre = bloques[0].get("nombre") or None

    for bloque in bloques:
        move_id = bloque["move_id"]; lineas = bloque["lineas"]
        if not lineas and not bloque["invalidas"]:
            continue
        print(f"  Factura move_id={move_id} | {len(lineas)} línea(s)")
        fa = _aplicar_bloque(odoo_client, move_id, lineas, dry_run,
                             invalidas=bloque["invalidas"])
        resultado.facturas.append(fa)

        rut = (bloque.get("rut") or "").strip()
        nombre = (bloque.get("nombre") or "").strip()
        if dir_memo and rut and not dry_run and not fa.errores:
            for linea in fa.lineas:
                clave = CODIGO_A_CLAVE.get(linea.cuenta_final_codigo, "")
                if not clave and linea.cuenta_final_codigo != CODIGO_CATCHALL:
                    continue
                if linea.fue_corregida:
                    registrar_correccion(rut, nombre, linea.glosa, clave or "COMISION_GRANDES_CUENTAS",
                                         aprobado_por, dir_memo)
                else:
                    registrar_aprobacion(rut, nombre, linea.glosa, clave or "COMISION_GRANDES_CUENTAS",
                                         aprobado_por, dir_memo)
    return resultado


def aplicar_directo(odoo_client, resultado_clasificacion, dry_run=True) -> FacturaAplicada:
    factura = resultado_clasificacion.factura
    move_id = factura.move_id
    print(f"  [{factura.name}] {factura.partner_nombre} | ${factura.monto_total:,.0f}")

    lineas_ap = []
    for linea_cl in resultado_clasificacion.lineas:
        odoo_id = linea_cl.cuenta_odoo_id
        cambio = (odoo_id != 0 and odoo_id != CUENTA_CATCHALL_ID)
        lineas_ap.append(LineaAplicada(
            line_id=linea_cl.line_id, glosa=linea_cl.glosa,
            cuenta_final_codigo=linea_cl.cuenta_codigo,
            cuenta_final_odoo_id=odoo_id, fue_corregida=False,
            cambio_cuenta=cambio))

    return _aplicar_bloque(odoo_client, move_id, lineas_ap, dry_run)
