"""
Genera UN SOLO Excel con todas las facturas a distribuir.
Hoja "Propuesta": fila separadora por factura + filas de líneas.
El aplicador detecta filas separadoras por el marcador SEP_MARKER en col A.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from .clasificador import ResultadoClasificacion, CUENTAS_DESTINO

VERDE_DARK = "1B5E20"; VERDE_CLARO = "E8F5E9"; AMARILLO = "FFF9C4"
ROJO_CLARO = "FFEBEE"; GRIS_HEADER = "37474F"; GRIS_SEP = "546E7A"
GRIS_CLARO = "ECEFF1"; BLANCO = "FFFFFF"
SEP_MARKER = "▼FACTURA"

def _fill(c): return PatternFill("solid", fgColor=c)
def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, val=None, bold=False, fondo=BLANCO, color_txt="000000",
          size=10, wrap=True, align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=color_txt, size=size)
    c.fill = _fill(fondo)
    c.alignment = Alignment(
        horizontal="right" if align == "right" else ("center" if align == "center" else "left"),
        vertical="center", wrap_text=wrap)
    c.border = _thin()
    if num_fmt:
        c.number_format = num_fmt
    return c

COLS = [(4,"#"),(13,"Línea ID\n(Odoo)"),(48,"Glosa original"),
        (14,"Cta actual\ncódigo"),(26,"Cta actual\nnombre"),
        (14,"Cta propuesta\ncódigo"),(28,"Cta propuesta\nnombre"),
        (14,"Monto neto\n(CLP)"),(11,"Confianza\nIA"),(12,"Método"),
        (38,"Razón IA"),(12,"APROBADO\n(SI / NO)"),(14,"Cta correcta\n(si NO)")]
N_COLS = len(COLS)
FILA_HEADERS = 1; FILA_INSTRUC = 2; FILA_DATOS = 4


def generar_excel_aprobacion(
    resultados,
    directorio_output: str = "data/distribucion",
    nombre_base: str = None,
) -> Path:
    if isinstance(resultados, ResultadoClasificacion):
        resultados = [resultados]
    Path(directorio_output).mkdir(parents=True, exist_ok=True)

    if nombre_base:
        nombre = nombre_base
    elif len(resultados) == 1:
        r = resultados[0]
        nombre = (f"distribucion_{r.factura.partner_rut.replace('-','')}"
                  f"_{r.factura.folio}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    else:
        nombre = f"distribucion_{len(resultados)}facturas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    ruta = Path(directorio_output) / nombre
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Propuesta"; ws.sheet_view.showGridLines = False

    for j, (ancho, _) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = ancho

    for j, (_, titulo) in enumerate(COLS, start=1):
        _cell(ws, FILA_HEADERS, j, titulo, bold=True, fondo=GRIS_HEADER,
              color_txt="FFFFFF", align="center")
    ws.row_dimensions[FILA_HEADERS].height = 40

    ws.merge_cells(f"A{FILA_INSTRUC}:{get_column_letter(N_COLS)}{FILA_INSTRUC}")
    c = ws.cell(row=FILA_INSTRUC, column=1,
                value="INSTRUCCIÓN: columna L → SI (aprobado) o NO (rechazado). "
                      "Si NO, indicar código de cuenta correcta en columna M. "
                      "Responder este correo con el archivo adjunto.")
    c.font = Font(bold=True, size=10, color="B71C1C"); c.fill = _fill("FFF9C4")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[FILA_INSTRUC].height = 22
    ws.row_dimensions[3].height = 4

    fila_actual = FILA_DATOS; total_global = 0.0

    for resultado in resultados:
        factura = resultado.factura
        lineas_catchall = list(resultado.lineas)

        ws.merge_cells(f"C{fila_actual}:F{fila_actual}")
        ws.merge_cells(f"G{fila_actual}:K{fila_actual}")
        _cell(ws, fila_actual, 1, SEP_MARKER, bold=True, fondo=GRIS_SEP, color_txt=GRIS_SEP, size=8)
        _cell(ws, fila_actual, 2, factura.move_id, bold=True, fondo=GRIS_SEP, color_txt="B0BEC5", size=9)
        _cell(ws, fila_actual, 3, f"  {factura.partner_nombre}",
              bold=True, fondo=GRIS_SEP, color_txt="FFFFFF", size=11)
        _cell(ws, fila_actual, 7,
              f"  FAC {factura.folio}  ·  {factura.fecha}  ·  ${factura.monto_total:,.0f} CLP",
              fondo=GRIS_SEP, color_txt="B0BEC5", size=10)
        for col in [4,5,6,8,9,10,11,12,13]:
            _cell(ws, fila_actual, col, fondo=GRIS_SEP)
        ws.row_dimensions[fila_actual].height = 22; fila_actual += 1

        for i, linea in enumerate(lineas_catchall, start=1):
            es_auto = linea.auto_aplicado
            fondo_linea = (VERDE_CLARO if es_auto else
                           (AMARILLO if linea.confianza >= 0.70 else ROJO_CLARO))
            _cell(ws, fila_actual, 1, i, align="center")
            _cell(ws, fila_actual, 2, linea.line_id, align="center")
            _cell(ws, fila_actual, 3, linea.glosa, fondo=fondo_linea)
            _cell(ws, fila_actual, 4, "42410104", align="center")
            _cell(ws, fila_actual, 5, "COMISION GRANDES CUENTAS")
            _cell(ws, fila_actual, 6, linea.cuenta_codigo, bold=True, fondo=fondo_linea, align="center")
            _cell(ws, fila_actual, 7, linea.cuenta_nombre, bold=True, fondo=fondo_linea)
            _cell(ws, fila_actual, 8, linea.monto_neto, num_fmt="#,##0", align="right")
            _cell(ws, fila_actual, 9, f"{linea.confianza:.0%}", align="center", fondo=fondo_linea)
            _cell(ws, fila_actual, 10, linea.metodo, align="center")
            _cell(ws, fila_actual, 11, linea.razon)
            aprobado_default = "SI" if (es_auto or linea.confianza >= 0.80) else ""
            c_ap = _cell(ws, fila_actual, 12, aprobado_default, bold=True, align="center",
                         fondo=VERDE_CLARO if aprobado_default == "SI" else AMARILLO)
            c_ap.font = Font(bold=True, size=12,
                              color="1B5E20" if aprobado_default == "SI" else "B71C1C")
            _cell(ws, fila_actual, 13, fondo=GRIS_CLARO)
            ws.row_dimensions[fila_actual].height = 26
            total_global += linea.monto_neto; fila_actual += 1

        # Líneas ya asignadas a su cuenta correcta — informativas (sin acción).
        # Col A vacía + APROBADO vacía → el aplicador las ignora (no las toca).
        ids_catchall = {l.line_id for l in lineas_catchall}
        GRIS_TXT = "78909C"
        for linea in factura.lineas:
            if linea.line_id in ids_catchall:
                continue
            _cell(ws, fila_actual, 1, fondo=GRIS_CLARO)
            _cell(ws, fila_actual, 2, linea.line_id, align="center", fondo=GRIS_CLARO, color_txt=GRIS_TXT, size=9)
            _cell(ws, fila_actual, 3, linea.glosa, fondo=GRIS_CLARO, color_txt=GRIS_TXT)
            _cell(ws, fila_actual, 4, linea.cuenta_actual_codigo, align="center", fondo=GRIS_CLARO, color_txt=GRIS_TXT)
            _cell(ws, fila_actual, 5, linea.cuenta_actual_nombre, fondo=GRIS_CLARO, color_txt=GRIS_TXT)
            _cell(ws, fila_actual, 6, "—", align="center", fondo=GRIS_CLARO, color_txt=GRIS_TXT)
            _cell(ws, fila_actual, 7, "Ya asignada — sin acción", fondo=GRIS_CLARO, color_txt=GRIS_TXT)
            _cell(ws, fila_actual, 8, linea.monto_neto, num_fmt="#,##0", align="right", fondo=GRIS_CLARO, color_txt=GRIS_TXT)
            _cell(ws, fila_actual, 9, "—", align="center", fondo=GRIS_CLARO, color_txt=GRIS_TXT)
            _cell(ws, fila_actual, 10, fondo=GRIS_CLARO)
            _cell(ws, fila_actual, 11, fondo=GRIS_CLARO)
            _cell(ws, fila_actual, 12, fondo=GRIS_CLARO)
            _cell(ws, fila_actual, 13, fondo=GRIS_CLARO)
            ws.row_dimensions[fila_actual].height = 22
            fila_actual += 1

        fila_actual += 1

    ws.merge_cells(f"A{fila_actual}:G{fila_actual}")
    _cell(ws, fila_actual, 1, f"TOTAL GLOBAL — {len(resultados)} factura(s)",
          bold=True, fondo=GRIS_HEADER, color_txt="FFFFFF", align="right")
    _cell(ws, fila_actual, 8, total_global, bold=True, fondo=GRIS_HEADER,
          color_txt="FFFFFF", num_fmt="#,##0")
    for col in range(9, N_COLS+1):
        _cell(ws, fila_actual, col, fondo=GRIS_HEADER)
    ws.row_dimensions[fila_actual].height = 22
    ws.freeze_panes = f"A{FILA_DATOS}"

    # Hoja instrucciones
    ws2 = wb.create_sheet("Instrucciones"); ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 28; ws2.column_dimensions["B"].width = 62
    instrucciones = [
        ("CÓMO COMPLETAR ESTE ARCHIVO",""),("",""),
        ("1. Revisa cada fila","Cada fila con color es una línea que necesita redistribución."),
        ("2. Columna APROBADO (L)","SI = de acuerdo con la cuenta propuesta."),
        ("   Si NO","Escribe NO y completa columna M con el código correcto."),
        ("3. Guarda y responde","Responde este correo con el archivo adjunto."),
        ("Filas grises","Líneas YA asignadas a su cuenta correcta (ENVÍOS/MARKETING). "
                        "Son informativas para que el total de la factura cuadre — no requieren acción."),
        ("",""),("CUENTAS VÁLIDAS:",""),
    ]
    for i, (lbl, val) in enumerate(instrucciones, start=1):
        c1 = ws2.cell(row=i, column=1, value=lbl)
        c2 = ws2.cell(row=i, column=2, value=val)
        c1.font = Font(bold=(i==1 or bool(lbl and not lbl.startswith(" "))), size=10)
        c2.font = Font(size=10)
    fila_c = len(instrucciones) + 1
    for col, txt in [(1,"Código"),(2,"Nombre")]:
        c = ws2.cell(row=fila_c, column=col, value=txt)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = _fill(GRIS_HEADER)
    fila_c += 1
    for clave, info in CUENTAS_DESTINO.items():
        ws2.cell(row=fila_c, column=1, value=info["codigo"]).font = Font(bold=True, size=10)
        ws2.cell(row=fila_c, column=2, value=info["nombre"]).font = Font(size=10)
        fila_c += 1

    wb.save(ruta)
    return ruta
