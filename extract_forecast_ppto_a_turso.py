#!/usr/bin/env python3
"""
Extrae los valores "Venta PPTO MMM26" del FORECAST FINAL SKU 26-27 V2.xlsx
y los sube a la tabla Turso `planif_forecast_manual`.

Cada columna del XLSX (cabecera fila 3) tipo " Venta PPTO ENE26", " Venta PPTO FEB26"
... " Venta PPTO ENE27" se convierte en N filas (sku, mes_iso, unidades) y se
upserta a Turso.

Idempotente (UPSERT por sku+mes). Saltea celdas vacías o 0.
"""
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')
PROJECT_ROOT = Path(__file__).parent
EXCEL_PATH = PROJECT_ROOT / 'data' / 'planificacion' / 'FORECAST FINAL SKU 26-27 V2.xlsx'

# Cargar .env si existe (CLI local; en GH Actions las env vars vienen del runner)
env = PROJECT_ROOT / '.env'
if env.exists():
    for line in env.read_text(encoding='utf-8').splitlines():
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

import libsql_client

MESES_ES_A_NUM = {'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
                   'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12}


def _parse_header_ppto(h: str) -> str | None:
    """' Venta PPTO ENE26' → '2026-01'. Devuelve None si no parsea."""
    if not h:
        return None
    s = str(h).strip().upper()
    if not s.startswith('VENTA PPTO'):
        return None
    # tail = 'ENE26' o 'ENE27' o similar
    tail = s.replace('VENTA PPTO', '').strip()
    if len(tail) < 5:
        return None
    mmm = tail[:3]
    yy = tail[3:]
    if mmm not in MESES_ES_A_NUM or not yy.isdigit():
        return None
    mes = MESES_ES_A_NUM[mmm]
    anio = 2000 + int(yy)
    return f'{anio:04d}-{mes:02d}'


def _new_client():
    return libsql_client.create_client_sync(
        url=os.environ['LIBSQL_URL'],
        auth_token=os.environ['LIBSQL_AUTH_TOKEN'],
    )


def exec_retry(sql, args=None, max_retries=4, base_wait=5, label=''):
    last = None
    for a in range(1, max_retries + 1):
        c = None
        try:
            c = _new_client()
            rs = c.execute(sql, args) if args is not None else c.execute(sql)
            c.close()
            return rs
        except Exception as e:
            last = e
            if c:
                try: c.close()
                except: pass
            if a < max_retries:
                w = base_wait * (2 ** (a - 1))
                print(f"  {label} retry {a} ({type(e).__name__}) en {w}s...", flush=True)
                time.sleep(w)
    raise last


def crear_tabla():
    exec_retry("""CREATE TABLE IF NOT EXISTS planif_forecast_manual (
        sku TEXT,
        mes TEXT,
        unidades REAL,
        fuente TEXT DEFAULT 'manual',
        ts_actualizado TEXT,
        PRIMARY KEY (sku, mes)
    )""", label='[create]')


def extract_xlsx_to_long() -> pd.DataFrame:
    """Lee el XLSX y devuelve DataFrame long [sku, mes, unidades]."""
    print(f"[1] Abriendo XLSX (254 MB, ~30s)...", flush=True)
    t0 = time.time()
    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    print(f"    OK en {time.time() - t0:.1f}s", flush=True)

    ws = wb['FCST BASE SKU MACRO']
    rows_iter = ws.iter_rows(values_only=True)

    # Fila 1, 2 (parámetros/fórmulas) → skip
    next(rows_iter); next(rows_iter)
    # Fila 3 = headers
    hdr = next(rows_iter)
    # Detectar índices: SKU está en col 4 (índice base-0)
    idx_sku = 4
    # Detectar cols PPTO
    ppto_cols = []
    for idx, h in enumerate(hdr):
        mes_iso = _parse_header_ppto(h)
        if mes_iso:
            ppto_cols.append((idx, mes_iso, str(h).strip()))
    print(f"[2] Detectadas {len(ppto_cols)} columnas PPTO:", flush=True)
    for _, mes_iso, lbl in ppto_cols:
        print(f"      {mes_iso}  ←  '{lbl}'", flush=True)

    # Iterar filas (data desde fila 4)
    long_rows = []
    n_filas = 0
    for row in rows_iter:
        n_filas += 1
        sku = row[idx_sku] if idx_sku < len(row) else None
        if sku is None:
            continue
        sku_str = str(sku).strip()
        if not sku_str or sku_str.lower() in ('nan', 'none', ''):
            continue
        for col_idx, mes_iso, _ in ppto_cols:
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None:
                continue
            try:
                v_f = float(v)
            except Exception:
                continue
            if v_f == 0:
                continue  # no guardar ceros (ahorra filas Turso)
            long_rows.append({'sku': sku_str, 'mes': mes_iso, 'unidades': v_f})
    wb.close()
    print(f"[3] Filas data leídas: {n_filas:,}. Long rows válidos (no-cero): {len(long_rows):,}", flush=True)

    df = pd.DataFrame(long_rows)
    if df.empty:
        return df
    # Stats
    print(f"[4] Resumen por mes:", flush=True)
    g = df.groupby('mes').agg(n_skus=('sku', 'nunique'),
                              total_uds=('unidades', 'sum')).reset_index().sort_values('mes')
    for _, r in g.iterrows():
        print(f"      {r['mes']}: {r['n_skus']:>4} SKUs, total {r['total_uds']:>10,.0f} uds", flush=True)
    return df


def upsert_to_turso(df: pd.DataFrame, fuente: str = 'xlsx_fcst_final'):
    if df.empty:
        print("[upsert] Sin filas. Skip.", flush=True)
        return

    ts = datetime.now().isoformat(timespec='seconds')
    print(f"[5] UPSERT {len(df):,} filas a planif_forecast_manual (fuente={fuente})...", flush=True)

    cols = 'sku, mes, unidades, fuente, ts_actualizado'
    ph = '(?,?,?,?,?)'
    inserted = 0
    batch = 100
    rows_list = df.to_dict(orient='records')
    for i in range(0, len(rows_list), batch):
        chunk = rows_list[i:i + batch]
        placeholders = ','.join([ph] * len(chunk))
        sql = (f"INSERT INTO planif_forecast_manual ({cols}) VALUES {placeholders} "
               f"ON CONFLICT (sku, mes) DO UPDATE SET unidades=excluded.unidades, "
               f"fuente=excluded.fuente, ts_actualizado=excluded.ts_actualizado")
        flat = []
        for r in chunk:
            flat.extend([str(r['sku']), str(r['mes']), float(r['unidades']), fuente, ts])
        try:
            rs = exec_retry(sql, flat, label=f'  [b{i // batch + 1}]')
            inserted += rs.rows_affected
        except Exception as e:
            print(f"    batch {i // batch + 1} FAILED: {e}", flush=True)
    print(f"    Insertados/actualizados: {inserted:,}", flush=True)


def main():
    print(f"=== EXTRACT FORECAST PPTO XLSX → TURSO — {datetime.now().isoformat()} ===\n", flush=True)
    if not EXCEL_PATH.exists():
        print(f"[ERROR] No existe el XLSX: {EXCEL_PATH}", flush=True)
        return 1

    for v in ('LIBSQL_URL', 'LIBSQL_AUTH_TOKEN'):
        if not os.environ.get(v):
            print(f"[ERROR] env var {v} no seteada", flush=True)
            return 1

    crear_tabla()
    df = extract_xlsx_to_long()
    if df.empty:
        print("Nada para subir.", flush=True)
        return 0
    upsert_to_turso(df)

    # Verificación
    rs = exec_retry("SELECT COUNT(*) FROM planif_forecast_manual", label='[ver]')
    print(f"\n[OK] Total filas en planif_forecast_manual: {rs.rows[0][0]:,}", flush=True)
    return 0


if __name__ == '__main__':
    sys.exit(main())
