"""Procesa un embarque COMEX end-to-end en 1 comando.

Hace todo el flujo: precosteo → maestra → mail al equipo → RFQ en Odoo.

Uso:
    python procesar_embarque.py 0320
    python procesar_embarque.py 0330 --flete 2800        # override flete USD
    python procesar_embarque.py 0424 --skip-mail         # no enviar mail
    python procesar_embarque.py 0429 --confirm-po        # confirmar PO en Odoo (no draft)
    python procesar_embarque.py 0422 --skip-po           # no crear PO

Auto-detecta:
  - Puerto: desde nombre del PI (SZ/NB/XI/AIR)
  - PI + PL: en data/comex/embarques/<emb>/
  - Tarifas: data/comex/embarques/Tarifas_Base_COMEX-<emb>.xlsx
  - Maestra: el backup más reciente en data/comex/

Pre-requisito: PI/PL ya descargados (agente lo hace automático) y Tarifa generada.
"""
import argparse
import base64
import os
import re
import subprocess
import sys
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
PROJECT_ROOT = Path(__file__).parent

DESTINATARIOS_MAIL = 'felipe@unionx.cl, sebastian@unionx.cl, gerardo@unionx.cl, operaciones@unionx.cl'


def _detectar_pi_pl_puerto(emb: str) -> tuple[Path, Path, str]:
    """Encuentra PI, PL y puerto para el embarque."""
    emb_dir = PROJECT_ROOT / 'data' / 'comex' / 'embarques' / emb
    if not emb_dir.exists():
        raise FileNotFoundError(f"Carpeta {emb_dir} no existe. Descarga el PI/PL desde Gmail primero.")

    pi, pl, puerto = None, None, None
    for f in emb_dir.iterdir():
        name = f.name.upper()
        if '40HQ' in name or 'DHL' in name or 'AIR' in name:
            pi = f
            # Detectar puerto: SZ, NB, XI, AIR
            for code in ('SZ', 'NB', 'XI', 'AIR', 'DHL'):
                if f' {code} ' in f' {name} ' or f' {code}.' in name:
                    puerto = code if code != 'DHL' else 'AIR'
                    break
        elif 'PL.XLSX' in name or 'PL ' in name:
            pl = f
    if not pi:
        raise FileNotFoundError(f"No encontré PI principal (40HQ/DHL/AIR) en {emb_dir}")
    if not pl:
        raise FileNotFoundError(f"No encontré PL en {emb_dir}")
    return pi, pl, puerto or 'SZ'


def _detectar_tarifa(emb: str) -> Path:
    p = PROJECT_ROOT / 'data' / 'comex' / 'embarques' / f'Tarifas_Base_COMEX-{emb}.xlsx'
    if not p.exists():
        raise FileNotFoundError(f"Tarifa {p.name} no existe. Genera con generar_tarifas_embarque.py.")
    return p


def _detectar_maestra_reciente() -> Path | None:
    """Maestra Importaciones más reciente en data/comex/."""
    candidatos = list((PROJECT_ROOT / 'data' / 'comex').glob('Maestra Importaciones*.xlsx'))
    if not candidatos:
        return None
    return max(candidatos, key=lambda p: p.stat().st_mtime)


def _actualizar_flete_tarifa(tarifa_path: Path, flete_usd: float):
    """Reemplaza el flete USD en la tarifa."""
    import openpyxl
    wb = openpyxl.load_workbook(str(tarifa_path))
    ws = wb['Tarifas']
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).lower().startswith('flete total'):
                # Valor está en columna siguiente
                ws.cell(cell.row, cell.column + 1).value = flete_usd
                wb.save(str(tarifa_path))
                print(f"  [tarifa] flete actualizado a USD {flete_usd}")
                return
    print(f"  [WARN] no encontré 'Flete Total' en {tarifa_path.name}")


def _run_costear(emb: str, pi: Path, pl: Path, tarifa: Path, maestra: Path | None, out_dir: Path) -> int:
    """Invoca _REACTIVAR_NUEVO_PC/costear_embarque.py."""
    args = [
        sys.executable, str(PROJECT_ROOT / '_REACTIVAR_NUEVO_PC' / 'costear_embarque.py'),
        '--pi', str(pi),
        '--pl', str(pl),
        '--tarifas', str(tarifa),
        '--out', str(out_dir),
    ]
    if maestra and maestra.exists():
        args += ['--maestra', str(maestra)]
    print(f"[1] Pre-costeando 26TP{emb}...")
    r = subprocess.run(args, env={**os.environ, 'PYTHONIOENCODING': 'utf-8'})
    return r.returncode


def _enviar_mail_analisis(emb: str, out_dir: Path) -> str | None:
    """Manda el HTML generado + Excel adjunto al equipo."""
    sys.path.insert(0, str(PROJECT_ROOT / 'agente-comex' / 'src'))
    from gmail_client import GmailClient

    html = (out_dir / f'email_26TP{emb}.html').read_text(encoding='utf-8')
    subject = (out_dir / f'email_26TP{emb}_subject.txt').read_text(encoding='utf-8').strip()
    precosteo = out_dir / f'Pre-costeo_x_CBM_26TP{emb}.xlsx'

    msg = MIMEMultipart('mixed')
    msg['to'] = DESTINATARIOS_MAIL
    msg['subject'] = subject

    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText('Análisis adjunto.', 'plain', 'utf-8'))
    alt.attach(MIMEText(html, 'html', 'utf-8'))
    msg.attach(alt)

    with open(precosteo, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{precosteo.name}"')
    msg.attach(part)

    g = GmailClient()
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    sent = g.service.users().messages().send(userId='me', body={'raw': raw}).execute()
    print(f"[2] Mail equipo enviado | id={sent['id']} | {subject}")
    return sent['id']


def _crear_po_odoo(emb: str, out_dir: Path, confirmar: bool) -> int:
    """Invoca cargar_po_comex_odoo.py."""
    precosteo = out_dir / f'Pre-costeo_x_CBM_26TP{emb}.xlsx'
    args = [
        sys.executable, str(PROJECT_ROOT / 'cargar_po_comex_odoo.py'),
        '--precosteo', str(precosteo),
    ]
    if confirmar:
        args.append('--confirm')
    print(f"[3] Creando RFQ Odoo para 26TP{emb}{' (confirmando)' if confirmar else ' (draft)'}...")
    r = subprocess.run(args, env={**os.environ, 'PYTHONIOENCODING': 'utf-8'})
    return r.returncode


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('embarque', help='Número embarque, ej "0330"')
    parser.add_argument('--flete', type=float, help='Override flete USD (actualiza Tarifa antes de costear)')
    parser.add_argument('--skip-mail', action='store_true', help='No enviar mail al equipo')
    parser.add_argument('--skip-po', action='store_true', help='No crear PO en Odoo')
    parser.add_argument('--confirm-po', action='store_true', help='Confirmar PO (state=purchase). Default: draft')
    parser.add_argument('--puerto', help='Override puerto detectado (SZ/NB/XI/AIR)')
    args = parser.parse_args()

    emb = args.embarque.zfill(4)
    out_dir = PROJECT_ROOT / 'agente-comex' / 'data' / 'output' / f'26TP{emb}'
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"{'='*70}\n  PROCESAR 26TP{emb}\n{'='*70}")
    pi, pl, puerto = _detectar_pi_pl_puerto(emb)
    puerto = args.puerto or puerto
    tarifa = _detectar_tarifa(emb)
    maestra = _detectar_maestra_reciente()

    print(f"  PI:      {pi.name}")
    print(f"  PL:      {pl.name}")
    print(f"  Puerto:  {puerto}")
    print(f"  Tarifa:  {tarifa.name}")
    print(f"  Maestra: {maestra.name if maestra else '(no encontrada, sin comparativo histórico)'}")

    if args.flete:
        _actualizar_flete_tarifa(tarifa, args.flete)

    rc = _run_costear(emb, pi, pl, tarifa, maestra, out_dir)
    if rc != 0:
        print(f"\n[ERROR] costear_embarque falló (rc={rc}). Aborto.")
        return rc

    if not args.skip_mail:
        try:
            _enviar_mail_analisis(emb, out_dir)
        except Exception as e:
            print(f"[WARN] mail falló: {type(e).__name__}: {e}")

    if not args.skip_po:
        rc_po = _crear_po_odoo(emb, out_dir, args.confirm_po)
        if rc_po != 0:
            print(f"[WARN] PO Odoo falló (rc={rc_po})")

    print(f"\n{'='*70}\n  26TP{emb} PROCESADO\n{'='*70}")
    print(f"  Outputs en: {out_dir}")
    return 0


if __name__ == '__main__':
    sys.exit(main() or 0)
