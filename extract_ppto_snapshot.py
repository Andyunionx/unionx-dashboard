"""
Extrae metas de venta neta (CLP) por canal y por marca del archivo PPTO 2026.
Genera parquets en data/planificacion/snapshots/.

Uso: python extract_ppto_snapshot.py
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import pandas as pd
from pathlib import Path

PPTO_PATH = Path(r"C:\Users\felip\Desktop\UNIONX\PPTO 2026\Metas oficiales 2SEM OFICIAL Detallado.xlsx")
SHEET_NAME = "PPTO MARCA 2026"
OUT_DIR = Path(__file__).parent / "data" / "planificacion" / "snapshots"


def _month_col_map(header_row):
    """Returns {col_idx: 'YYYY-MM'} for columns with '2026.XX' header."""
    return {
        i: f"2026-{str(v).split('.')[1]}"
        for i, v in enumerate(header_row)
        if v is not None and str(v).startswith("2026.")
    }


def _first_section_month_col_map(header_row):
    """Only the FIRST block of 12 '2026.XX' columns (venta meta section).
    The marca sheet has multiple horizontal sections (venta meta, contrib, ppto compra, etc.),
    all sharing the same '2026.01'-'2026.12' headers. Only the first block contains venta meta."""
    result = {}
    for i, v in enumerate(header_row):
        if v is not None and str(v).startswith("2026."):
            result[i] = f"2026-{str(v).split('.')[1]}"
            if len(result) == 12:
                break
    return result


def extract_canal(rows):
    """Canal metas: rows where col[0] is canal name and col[1]=='Venta Neta Total'."""
    # Header: first row where col[0]=='CANAL'
    header_idx = next((i for i, r in enumerate(rows) if r[0] == 'CANAL'), None)
    if header_idx is None:
        raise ValueError("Sección CANAL no encontrada")

    month_map = _month_col_map(rows[header_idx])
    if not month_map:
        raise ValueError("Sin columnas de meses en header de canal")

    records = []
    for r in rows[header_idx + 1: header_idx + 20]:  # at most 20 rows after header
        canal = r[0]
        if not canal or 'Total' in str(canal):
            break
        if r[1] != 'Venta Neta Total':
            continue
        for col_idx, mes in month_map.items():
            val = r[col_idx]
            if val is not None:
                try:
                    records.append({'canal': str(canal), 'mes': mes, 'meta_venta_neta': float(val)})
                except (ValueError, TypeError):
                    pass

    return pd.DataFrame(records)


def _marca_header_idx(rows):
    for i, r in enumerate(rows):
        c1 = str(r[1]).strip() if r[1] is not None else ''
        c2 = str(r[2]).strip() if r[2] is not None else ''
        if c1 == 'Tipo canal' and c2 == 'Marca':
            return i
    # Fallback
    for i, r in enumerate(rows):
        if str(r[2]).strip() == 'Marca' and r[3] is not None and str(r[3]).startswith('2026.'):
            return i
    raise ValueError("Header de marcas no encontrado")


def extract_marca(rows):
    """Marca totals: rows where col[0]=='' AND col[1]=='' AND col[2] is marca name.
    Also extracts Bandú and T-Care from their per-canal rows (new brands from H2-2026,
    summed across canals). 'Marcas Flash' is renamed 'Dynamo Tools' since that's the
    individual brand it represents."""
    header_idx = _marca_header_idx(rows)

    # Use ONLY the first 12-month block (venta meta section).
    # Subsequent sections in the same row contain contrib, ppto-compra, etc.
    month_map = _first_section_month_col_map(rows[header_idx])
    if not month_map:
        raise ValueError("Sin columnas de meses en header de marcas")

    # ── Brand-total rows (c0='', c1='') ───────────────────────────────────
    SKIP = {'', None, 'total', 'grand total', 'corportativo otros', 'corporativo otros'}
    records = []
    for r in rows[header_idx + 1:]:
        c0 = (r[0] or '').strip()
        c1 = (str(r[1]) if r[1] is not None else '').strip()
        c2 = r[2]
        if c0 != '' or c1 not in ('', 'None'):
            continue
        if not c2 or str(c2).strip().lower() in SKIP:
            continue
        marca = str(c2).strip()
        if 'Total' in marca or 'TOTAL' in marca:
            continue
        # "Marcas Flash" brand-total = Dynamo Tools canal sum; rename for clarity
        if marca == 'Marcas Flash':
            marca = 'Dynamo Tools'
        found_any = False
        for col_idx, mes in month_map.items():
            val = r[col_idx]
            if val is not None and isinstance(val, (int, float)) and float(val) > 0:
                found_any = True
                records.append({'marca': marca, 'mes': mes, 'meta_venta_neta': float(val)})
        if not found_any and len(records) > 0:
            break

    # ── Per-canal rows for Bandú and T-Care (new brands, H2-2026) ─────────
    # These brands don't have a brand-total row; must sum across canals.
    NEW_BRANDS = {'Bandú', 'T-Care'}
    for r in rows[header_idx + 1:]:
        c0 = (r[0] or '').strip()
        c1 = (str(r[1]) if r[1] is not None else '').strip()
        c2 = str(r[2]).strip() if r[2] is not None else ''
        if c0 != '' or c1 == '':
            continue  # keep only canal-specific rows (c0='', c1=canal)
        if c2 not in NEW_BRANDS:
            continue
        for col_idx, mes in month_map.items():
            val = r[col_idx]
            if val is not None and isinstance(val, (int, float)) and float(val) > 0:
                records.append({'marca': c2, 'mes': mes, 'meta_venta_neta': float(val)})

    df = pd.DataFrame(records)
    if df.empty:
        return df

    # Sum per-canal contributions for Bandú / T-Care across canals
    df = df.groupby(['marca', 'mes'], as_index=False)['meta_venta_neta'].sum()
    return df


def main():
    if not PPTO_PATH.exists():
        print(f"ERROR: No se encontró {PPTO_PATH}")
        return

    print(f"Leyendo {PPTO_PATH}...")
    wb = openpyxl.load_workbook(str(PPTO_PATH), read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Hoja '{SHEET_NAME}' no encontrada. Disponibles: {wb.sheetnames}")
        return

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    print(f"  {len(rows)} filas leídas de la hoja '{SHEET_NAME}'")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── Canal ─────────────────────────────────────────────────────────────
    try:
        df_canal = extract_canal(rows)
        out = OUT_DIR / 'planif_ppto_canal.parquet'
        df_canal.to_parquet(out, index=False)
        print(f"\n✅ Canal: {len(df_canal)} registros → {out}")
        pivot = df_canal.pivot_table(index='canal', values='meta_venta_neta', aggfunc='sum')
        pivot['total_anual_$M'] = (pivot['meta_venta_neta'] / 1e6).round(1)
        print(pivot[['total_anual_$M']].sort_values('total_anual_$M', ascending=False).to_string())
    except Exception as e:
        print(f"❌ ERROR extrayendo canal: {e}")

    # ── Marca ──────────────────────────────────────────────────────────────
    try:
        df_marca = extract_marca(rows)
        df_marca = df_marca.drop_duplicates(subset=['marca', 'mes']).reset_index(drop=True)

        # Normalize: scale marca metas per month so their total matches the canal total.
        # The "supuesto" section sums to a different total than the official canal budget.
        if not df_canal.empty:
            canal_monthly = df_canal.groupby('mes')['meta_venta_neta'].sum()
            marca_monthly = df_marca.groupby('mes')['meta_venta_neta'].sum()
            def _normalize_row(row):
                mes = row['mes']
                c_tot = canal_monthly.get(mes, 0)
                m_tot = marca_monthly.get(mes, 0)
                if m_tot > 0 and c_tot > 0:
                    row['meta_venta_neta'] *= c_tot / m_tot
                return row
            df_marca = df_marca.apply(_normalize_row, axis=1)
            print("\nNormalización marca→canal aplicada:")
            for mes in sorted(canal_monthly.index):
                orig = marca_monthly.get(mes, 0) / 1e6
                norm = canal_monthly.get(mes, 0) / 1e6
                print(f"  {mes}: {orig:.1f}M → {norm:.1f}M (factor {norm/orig:.3f})" if orig > 0 else f"  {mes}: sin datos marca")

        out = OUT_DIR / 'planif_ppto_marca.parquet'
        df_marca.to_parquet(out, index=False)
        print(f"\n✅ Marca: {len(df_marca)} registros → {out}")
        pivot = df_marca.pivot_table(index='marca', columns='mes', values='meta_venta_neta', aggfunc='sum')
        jul_col = '2026-07'
        if jul_col in pivot.columns:
            print(f"\nJulio 2026 por marca ($M):")
            print((pivot[jul_col] / 1e6).round(1).sort_values(ascending=False).to_string())
            print(f"TOTAL Jul: {pivot[jul_col].sum()/1e6:.1f}M")
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"❌ ERROR extrayendo marca: {e}")

    print("\nDone!")


if __name__ == '__main__':
    main()
