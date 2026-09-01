# -*- coding: utf-8 -*-
"""Pulso Reposición Fulfillment v2 — sugerido semanal de carga a bodegas full.

v2 = especificación de la mesa de trabajo 11-08-2026 (mail Nicole 11-08 17:21):
  1. BASE DE PRODUCTOS: hoja "Maestra" del drive de PRICING (1gVJmFCR19...).
  2. VENTA/UME: RAW filtrado a canales DIGITALES (tipo_negocio Marketplace +
     Páginas propias + Fidelización), promedio L6W ISO sin semanas 23/41.
  3. COBERTURA (restricción de envío): stock CA1/Stock ÷ demanda semanal digital
     ×4,33, donde la demanda del SKU = venta directa + venta de packs que lo
     contienen (hoja "Packs" de Pricing). Regla Nicolás: cuidar ~1 mes de CA1
     (umbral por canal en hoja Reglas del panel de Nicole).
  4. STOCK FULL LIVE: feed seller center (Martín) + Walmart Odoo + **tránsito
     a fulls**: pickings internos PENDIENTES hacia bodegas BF* (los picks de
     Gerardo) se suman al stock del canal destino. [Pendiente: stock Fala desde
     la vista no-descarga (scraper de Claudia) cuando exista el feed.]
  5. Blacklist/LARGE/UME Manual/Reglas/UME MIN: siguen en el panel de Nicole
     (drive original, se actualiza los viernes) — se leen en cada corrida.

Semántica sugerido (v1.2, definiciones Nicole 03-08): objetivo = max(UME×cob
canal/large, UME MIN, UME Manual-directo); repo = múltiplo de 2; anulada por
Blacklist o Restricción (cobertura CA1 < umbral canal).

Uso: python pulso_reposicion_fulfillment.py [--no-download] [--no-live]
Salida: data/outputs/Pulso_Reposicion_Fulfillment_<fecha>.xlsx
"""
import os, sys, math, json, argparse, datetime
from pathlib import Path
import pandas as pd
import numpy as np

sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).parent
PANEL_ID = "1eyxR9KsCgDswZWuSi6DuLq6GqoomuvCc"   # panel Nicole (Reglas/BL/Large/Manual/MIN/Dim)
PRICING_ID = "1gVJmFCR19KbYkZfds7fH-62rJ0Zpt32P"  # drive Pricing (Maestra + Packs)
LOCAL_PANEL = ROOT / "data/outputs/NICOLE_fulfillment_sugeridos_UME.xlsx"
LOCAL_PRICING = ROOT / "data/outputs/PRICING_drive.xlsx"
TEMPLATE_DIN = ROOT / "data/templates/pulso_reposicion_template.xlsx"
DETALLE = ROOT / "data/stock/detalle.parquet"
HIST = ROOT / "data/historico/ventas_historico.parquet"
MES = ROOT / "data/historico/ventas_mes_actual.parquet"
OUT_DIR = ROOT / "data/outputs"
CANALES = ["Mercado Libre", "Falabella", "Paris", "Walmart"]
BOD2CANAL = {"BFML": "Mercado Libre", "BFFa": "Falabella", "BFP": "Paris",
             "BFR": "Ripley", "BFW": "Walmart", "BFE": "Falabella"}
DIGITAL = {"Marketplace", "Páginas propias", "Páginas Propias", "Fidelización"}
SEM_EVENTO = {23, 41}
AZUL = "4884FC"


def _drive_creds():
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    tok_env = os.environ.get("DRIVE_OAUTH_TOKEN_JSON", "")
    if tok_env:
        info = json.loads(tok_env)
        creds = Credentials.from_authorized_user_info(info, info.get("scopes"))
    else:
        creds = Credentials.from_authorized_user_file(str(ROOT / "drive_oauth_token.json"))
    if not creds.valid:
        creds.refresh(Request())
    return creds


def descargar(fid: str, destino: Path) -> None:
    import requests
    creds = _drive_creds()
    H = {"Authorization": f"Bearer {creds.token}"}
    r = requests.get(f"https://www.googleapis.com/drive/v3/files/{fid}?alt=media", headers=H, timeout=180)
    if not r.ok:
        r = requests.get(f"https://www.googleapis.com/drive/v3/files/{fid}/export"
                         f"?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         headers=H, timeout=180)
    r.raise_for_status()
    destino.write_bytes(r.content)
    print(f"[drive] {destino.name} refrescado ({len(r.content):,} bytes)")


def mult2(x: float) -> int:
    return max(0, int(math.floor(x / 2 + 0.5) * 2))


def semanas_l6w(hoy: datetime.date) -> list[int]:
    w = hoy.isocalendar()[1]
    return [s for s in range(w - 9, w) if s not in SEM_EVENTO][-6:]


def cargar_parametros():
    """Panel Nicole: Reglas, UME MIN, UME Manual, Blacklist, Large, Dimensiones."""
    xl = pd.ExcelFile(LOCAL_PANEL)
    reglas = xl.parse("Reglas").iloc[:5]
    reglas.columns = ["Canal", "cob", "cob_large", "restr"]
    reglas = reglas.set_index("Canal")

    umin = xl.parse("Ume MIN", header=0)
    t_tipo = {str(k).strip(): v for k, v in zip(umin.iloc[:, 0], pd.to_numeric(umin.iloc[:, 1], errors="coerce")) if pd.notna(v)}
    t_marca = {str(k).strip(): v for k, v in zip(umin.iloc[:, 3], pd.to_numeric(umin.iloc[:, 4], errors="coerce")) if pd.notna(v)}

    bl = xl.parse("BLACKLIST")
    bl.columns = [c.strip() for c in bl.columns]
    blset = {(str(r["Canal"]).strip(), str(r["SKU"]).strip())
             for _, r in bl.iterrows() if pd.to_numeric(r.get("Blacklist"), errors="coerce") == 1}
    lg = xl.parse("LARGE")
    lg.columns = [c.strip() for c in lg.columns]
    lgset = {(str(r["Canal"]).strip(), str(r["SKU"]).strip())
             for _, r in lg.iterrows() if pd.to_numeric(r.get("Large"), errors="coerce") == 1}

    ume_tab = xl.parse("UME")
    ume_tab["UME Manual"] = pd.to_numeric(ume_tab["UME Manual"], errors="coerce")
    manual_map = {(str(r["Sku"]).strip(), str(r["Canal"]).strip()): float(r["UME Manual"])
                  for _, r in ume_tab.dropna(subset=["UME Manual"]).iterrows()}
    bl_panel = {(str(r["Sku"]).strip(), str(r["Canal"]).strip()): int(r["Blacklist"])
                for _, r in ume_tab.iterrows() if pd.notna(pd.to_numeric(r["Blacklist"], errors="coerce"))}

    dim = xl.parse("Dimensiones")
    dim.columns = [str(c).strip() for c in dim.columns]
    dim["SKU"] = dim["SKU"].astype(str).str.strip()
    dim_m3 = dict(zip(dim["SKU"], pd.to_numeric(dim["Dim m3"], errors="coerce")))
    dim_kg = dict(zip(dim["SKU"], pd.to_numeric(dim["Peso"], errors="coerce")))
    return reglas, t_tipo, t_marca, blset, lgset, manual_map, bl_panel, dim_m3, dim_kg, ume_tab


def cargar_pricing():
    """Drive Pricing: Maestra (base de productos) + Packs (componentes)."""
    xl = pd.ExcelFile(LOCAL_PRICING)
    prod = xl.parse("Maestra")
    prod.columns = [str(c).strip() for c in prod.columns]
    prod = prod.dropna(subset=["Sku"]).drop_duplicates("Sku")
    prod["Sku"] = prod["Sku"].astype(str).str.strip()
    packs = xl.parse("Packs")
    packs.columns = [str(c).strip() for c in packs.columns]
    packs["Pack"] = packs["Pack"].astype(str).str.strip()
    packs["SKU"] = packs["SKU"].astype(str).str.strip()
    # componentes por pack (cantidad = nº de filas repetidas del mismo SKU)
    comp = packs.groupby(["Pack", "SKU"]).size().rename("qty").reset_index()
    print(f"[pricing] Maestra: {len(prod)} SKUs | Packs: {comp['Pack'].nunique()} packs / {len(comp)} componentes")
    return prod, comp


def cargar_venta_digital(sem: list[int]) -> pd.DataFrame:
    """Unidades por SKU x canal x semana ISO, SOLO canales digitales."""
    anio = datetime.date.today().year
    corte = datetime.date.today().replace(day=1).isoformat()
    cols = ["sku", "canal", "fecha_venta", "cantidad", "tipo_negocio"]
    h = pd.read_parquet(HIST, columns=cols)
    h = h[h["fecha_venta"].astype(str) < corte]
    m = pd.read_parquet(MES, columns=cols)
    m = m[m["fecha_venta"].astype(str) >= corte]
    v = pd.concat([h, m], ignore_index=True)
    v = v[v["tipo_negocio"].isin(DIGITAL)]
    v["fecha_venta"] = pd.to_datetime(v["fecha_venta"])
    v = v[v["fecha_venta"].dt.year == anio]
    v["week"] = v["fecha_venta"].dt.isocalendar().week.astype(int)
    v = v[v["week"].isin(sem)]
    v["sku"] = v["sku"].astype(str).str.strip()
    v["cantidad"] = pd.to_numeric(v["cantidad"], errors="coerce").fillna(0)
    return v


# Uds mínimas en CA1 para surfacer un SKU no cargado en la Maestra (excluye
# muestras de 1 unidad y saldos residuales; los productos nuevos reales llegan
# por cientos). Los descontinuados no entran: ya están en la Maestra (In/Out=out).
UMBRAL_CA1_NUEVO = 12


def _txt(v):
    """str limpio, tratando NaN/'nan' como vacío (para no pisar con basura)."""
    s = "" if v is None else str(v).strip()
    return "" if s.lower() == "nan" else s


def complementar_base_nuevos(prod, venta, full_map, ca1=None):
    """Complementa la base (hoja Maestra) con SKU NO presentes en ella que tengan:
    venta digital L6W, stock en algún marketplace, o stock en bodega CA1 (productos
    recién recibidos que pricing aún no carga en la Maestra). Así los productos
    nuevos no quedan fuera del sugerido sin depender de que alguien los cargue al
    doc. Enriquece marca/categoría/producto desde el detalle de stock y el RAW
    (Odoo). Marca las filas con _nuevo=True para distinguirlas en el reporte."""
    maestra = set(prod["Sku"].astype(str).str.strip())
    con_venta = set(venta["sku"].astype(str).str.strip())
    con_stock = {sku for (_c, sku) in full_map.keys()}

    # SKU con stock relevante en CA1 (excluye MUESTRA y saldos < umbral). Se guarda
    # producto/marca/categoría del detalle para enriquecer los que no están en el RAW.
    con_ca1, ca1_info = set(), {}
    if ca1 is not None and len(ca1):
        c = ca1.copy()
        c["_sku"] = c["SKU"].astype(str).str.strip()
        prod_col = c["Producto"].astype(str) if "Producto" in c.columns else pd.Series("", index=c.index)
        c = c[~prod_col.str.contains("MUESTRA", case=False, na=False)]
        g = c.groupby("_sku")["Disponible"].sum()
        con_ca1 = set(g[g >= UMBRAL_CA1_NUEVO].index)
        info = c.drop_duplicates("_sku").set_index("_sku")
        for s in con_ca1:
            r = info.loc[s] if s in info.index else None
            ca1_info[s] = (
                _txt(r["Producto"]) if r is not None and "Producto" in info.columns else "",
                _txt(r["Marca"]) if r is not None and "Marca" in info.columns else "",
                _txt(r["Categoria"]) if r is not None and "Categoria" in info.columns else "",
            )

    nuevos = sorted({s for s in (con_venta | con_stock | con_ca1) - maestra
                     if s and s.lower() != "nan"})
    prod = prod.copy()
    prod["_nuevo"] = False
    if not nuevos:
        return prod, []
    rcols = ["sku", "producto", "marca", "categoria_hijo"]
    enr = pd.concat([pd.read_parquet(MES, columns=rcols),
                     pd.read_parquet(HIST, columns=rcols)], ignore_index=True)
    enr["sku"] = enr["sku"].astype(str).str.strip()
    enr = enr[enr["sku"].isin(nuevos)].drop_duplicates("sku").set_index("sku")
    rows = []
    for s in nuevos:
        e = enr.loc[s] if s in enr.index else None
        ci = ca1_info.get(s, ("", "", ""))
        # RAW primero; si no hay dato (típico de nuevo sin venta), cae al detalle CA1
        rows.append({
            "Sku": s, "In/Out": "in", "_nuevo": True, "Cob": None,
            "Producto": (_txt(e["producto"]) if e is not None else "") or ci[0],
            "Marca": (_txt(e["marca"]) if e is not None else "") or ci[1],
            "Categoría": (_txt(e["categoria_hijo"]) if e is not None else "") or ci[2],
            "Tipo de producto": "", "Pack": "",
        })
    add = pd.DataFrame(rows)
    for c in prod.columns:
        if c not in add.columns:
            add[c] = ""
    prod2 = pd.concat([prod, add[prod.columns]], ignore_index=True)
    return prod2, nuevos


FALA_FBF = ROOT / "data/stock/fala_fbf_live.parquet"
WFS_LIVE = ROOT / "data/stock/walmart_wfs_live.parquet"


def _override_walmart(full_df):
    """Reemplaza el stock de Walmart por el export WFS del Seller Center
    (ingest_walmart_wfs, `Available units`) si el parquet existe. Walmart no tiene
    feed de Martín ni API → el WFS es la única fuente real (vs Odoo BFW, que
    descuadra bidireccionalmente). Ver [[pulso_reposicion_fulfillment]] sem 34."""
    if not WFS_LIVE.exists():
        return full_df
    try:
        f = pd.read_parquet(WFS_LIVE)[["canal", "sku", "qty"]].copy()
        f["qty"] = pd.to_numeric(f["qty"], errors="coerce").fillna(0).clip(lower=0)
        base = full_df[full_df["canal"] != "Walmart"]
        out = pd.concat([base, f], ignore_index=True)
        return out.groupby(["canal", "sku"], as_index=False)["qty"].sum()
    except Exception as e:
        print(f"[wfs][WARN] override Walmart falló: {type(e).__name__}: {e}", flush=True)
        return full_df


def _transito_walmart_wfs():
    """Inbound del WFS = tránsito real de Walmart hacia el Full (reemplaza el
    tránsito de Odoo para este canal). {(canal, sku): uds}."""
    if not WFS_LIVE.exists():
        return {}
    try:
        f = pd.read_parquet(WFS_LIVE)
        f["inbound"] = pd.to_numeric(f["inbound"], errors="coerce").fillna(0).clip(lower=0)
        f = f[f["inbound"] > 0]
        return {("Walmart", r.sku): float(r.inbound) for r in f.itertuples(index=False)}
    except Exception as e:
        print(f"[wfs][WARN] tránsito Walmart falló: {type(e).__name__}: {e}", flush=True)
        return {}


def _override_fala(full_df):
    """Reemplaza el stock de Falabella por el feed FBF real (scrape_fala_fbf, API
    del Seller Center) si el parquet existe. Más fiable que el Odoo BF* / Martín."""
    if not FALA_FBF.exists():
        return full_df
    try:
        f = pd.read_parquet(FALA_FBF)[["canal", "sku", "qty"]].copy()
        # El Seller Center devuelve availableStock NEGATIVO en SKU con overstock/
        # reservas (55 SKU, -211 uds al 14-ago). Un stock disponible negativo no es
        # real → se piso a 0. Si no, live_q<0 infla el sugerido (repo=objetivo-live_q
        # = objetivo+|neg|) en vez de dar 0. (Reporte Claudia, sem 34.)
        f["qty"] = pd.to_numeric(f["qty"], errors="coerce").fillna(0).clip(lower=0)
        base = full_df[full_df["canal"] != "Falabella"]
        out = pd.concat([base, f], ignore_index=True)
        return out.groupby(["canal", "sku"], as_index=False)["qty"].sum()
    except Exception as e:
        print(f"[fbf][WARN] override Falabella falló: {type(e).__name__}: {e}", flush=True)
        return full_df


def stock_full_live(no_live: bool):
    det = pd.read_parquet(DETALLE)
    bcode = det["Bodega"].astype(str).str.split("/").str[0]
    odoo = det[bcode.isin(BOD2CANAL)].copy()
    odoo["canal"] = bcode[bcode.isin(BOD2CANAL)].map(BOD2CANAL)
    odoo["sku"] = odoo["SKU"].astype(str).str.strip()
    odoo_g = odoo.groupby(["canal", "sku"], as_index=False)["Disponible"].sum().rename(columns={"Disponible": "qty"})
    fuente, cruce = "Odoo (fallback)", None
    if not no_live:
        try:
            import fulfillment_live as fl
            try:
                live = fl.cargar_live()
                fuente = "Seller center (feed Martín, fresco)"
            except Exception as e:
                live = pd.read_parquet(fl.LIVE_PARQUET)
                fuente = f"Seller center (parquet local; feed no disponible: {type(e).__name__})"
            cruce = fl.generar_cruce(live)
            canales_live = set(live["canal"].unique())
            base = pd.concat([live[["canal", "sku", "qty"]],
                              odoo_g[~odoo_g["canal"].isin(canales_live)]], ignore_index=True)
            full = base.groupby(["canal", "sku"], as_index=False)["qty"].sum()
            if FALA_FBF.exists():
                fuente += " · Falabella=FBF (API Seller Center)"
            if WFS_LIVE.exists():
                fuente += " · Walmart=WFS (export Seller Center)"
            return _override_walmart(_override_fala(full)), fuente, cruce
        except Exception as e:
            print(f"[live][WARN] {type(e).__name__}: {e} -> Odoo BF*")
    if FALA_FBF.exists():
        fuente += " · Falabella=FBF (API Seller Center)"
    if WFS_LIVE.exists():
        fuente += " · Walmart=WFS (export Seller Center)"
    return _override_walmart(_override_fala(odoo_g)), fuente, cruce


def transito_a_fulls():
    """Picks internos PENDIENTES hacia bodegas BF* (reposiciones de Gerardo en
    camino que el seller center aún no refleja) → sumar al stock del canal."""
    import xmlrpc.client, time
    cfg = json.load(open(ROOT / "odoo/odoo_config.json"))["produccion"]
    pw = os.environ.get("ANDRES_ODOO_PASSWORD", "") or (ROOT / "odoo/.odoo_pass").read_text().strip()
    uid = xmlrpc.client.ServerProxy(f"{cfg['url']}/xmlrpc/2/common").authenticate(cfg["db_name"], cfg["username"], pw, {})
    def rpc(model, method, args, kw=None):
        for i in range(3):
            try:
                return xmlrpc.client.ServerProxy(f"{cfg['url']}/xmlrpc/2/object").execute_kw(
                    cfg["db_name"], uid, pw, model, method, args, kw or {})
            except Exception:
                if i == 2:
                    raise
                time.sleep(5)
    mv = rpc("stock.move", "search_read",
             [[("state", "in", ["confirmed", "assigned", "waiting", "partially_available"]),
               ("location_dest_id.complete_name", "like", "BF")]],
             {"fields": ["product_id", "product_uom_qty", "location_dest_id"]})
    filas = []
    pids = list({m["product_id"][0] for m in mv if m["product_id"]})
    codes = {}
    for i in range(0, len(pids), 300):
        for p in rpc("product.product", "read", [pids[i:i+300]], {"fields": ["default_code"]}):
            codes[p["id"]] = str(p["default_code"] or "").strip()
    for m in mv:
        if not m["product_id"]:
            continue
        bod = str(m["location_dest_id"][1]).split("/")[0]
        canal = BOD2CANAL.get(bod)
        sku = codes.get(m["product_id"][0], "")
        if canal and sku:
            filas.append({"canal": canal, "sku": sku, "qty": m["product_uom_qty"]})
    df = pd.DataFrame(filas)
    tot = df["qty"].sum() if len(df) else 0
    print(f"[transito-fulls] {len(df)} líneas / {tot:,.0f} uds en camino a bodegas BF*")
    return df.groupby(["canal", "sku"])["qty"].sum().to_dict() if len(df) else {}


def construir(no_download=False, no_live=False) -> Path:
    hoy = datetime.date.today()
    sem = semanas_l6w(hoy)
    print(f"[pulso v2] {hoy} — semana ISO {hoy.isocalendar()[1]} — L6W: {sem}")
    if not no_download:
        for fid, dest in [(PANEL_ID, LOCAL_PANEL), (PRICING_ID, LOCAL_PRICING)]:
            try:
                descargar(fid, dest)
            except Exception as e:
                print(f"[drive][WARN] {dest.name}: {e} -> copia local")

    (reglas, t_tipo, t_marca, blset, lgset, manual_map, bl_panel,
     dim_m3, dim_kg, ume_nicole) = cargar_parametros()
    prod, comp = cargar_pricing()
    venta = cargar_venta_digital(sem)
    full, fuente_live, cruce = stock_full_live(no_live)
    full_map = {(r.canal, r.sku): r.qty for r in full.itertuples(index=False)}

    det = pd.read_parquet(DETALLE)
    ca1 = det[det["Bodega"].astype(str).str.startswith("CA1")]
    stock_ca1 = ca1.groupby(ca1["SKU"].astype(str).str.strip())["Disponible"].sum()

    # Complementar la base con productos nuevos no presentes en la Maestra: con venta
    # digital, stock en marketplace, o stock en bodega CA1 (productos recién recibidos
    # que pricing aún no carga) → no depende de que se carguen a mano en el doc.
    prod, sku_nuevos = complementar_base_nuevos(prod, venta, full_map, ca1)
    if sku_nuevos:
        print(f"[complemento] +{len(sku_nuevos)} SKU con venta/stock no presentes en la Maestra")
    try:
        tr_fulls = transito_a_fulls()
    except Exception as e:
        print(f"[transito-fulls][WARN] {type(e).__name__}: {e}")
        tr_fulls = {}
    # Walmart: el tránsito real es el `Inbound units` del WFS (Seller Center), no
    # los picks de Odoo → se reemplazan las entradas de Walmart por el WFS.
    wfs_tr = _transito_walmart_wfs()
    if wfs_tr:
        tr_fulls = {k: v for k, v in tr_fulls.items() if k[0] != "Walmart"}
        tr_fulls.update(wfs_tr)
        print(f"[transito] Walmart desde WFS: {sum(wfs_tr.values()):,.0f} uds en {len(wfs_tr)} SKU")

    # venta semanal por sku x canal (para UME) y demanda digital total con packs
    v_canal = venta.groupby(["sku", "canal"])["cantidad"].sum() / len(sem)
    v_sku = venta.groupby("sku")["cantidad"].sum() / len(sem)
    demanda = v_sku.copy()
    for r in comp.itertuples(index=False):
        vp = v_sku.get(r.Pack, 0)
        if vp:
            demanda[r.SKU] = demanda.get(r.SKU, 0) + vp * r.qty
    print(f"[demanda] SKUs con venta digital: {len(v_sku)} | con aporte de packs: "
          f"{sum(1 for r in comp.itertuples(index=False) if v_sku.get(r.Pack,0))}")

    filas = []
    for p in prod.itertuples(index=False):
        d = dict(zip(prod.columns, p))
        sku = str(d.get("Sku")).strip()
        es_nuevo = bool(d.get("_nuevo", False))  # SKU no presente en la Maestra
        tipo = str(d.get("Tipo de producto", "") or "").strip()
        marca = str(d.get("Marca", "") or "").strip()
        umin = (t_tipo.get(tipo, 2) or 0) * (t_marca.get(marca, 0) or 0)
        st_ca1 = float(stock_ca1.get(sku, 0))
        dem = float(demanda.get(sku, 0))
        cob_ca1 = st_ca1 / (dem * 4.33) if dem > 0 else np.inf
        # Cobertura OBJETIVO por SKU (Maestra Pricing, col 'Cob', en meses) y estado In/Out.
        # Confirmado Andrés 13-ago: Cob = meses objetivo (reemplaza cob por canal); out = no reponer.
        cob_obj_meses = pd.to_numeric(d.get("Cob"), errors="coerce")
        es_out = str(d.get("In/Out", "")).strip().lower() == "out"
        canal_rows = []
        for canal in CANALES:
            ume = mult2(float(v_canal.get((sku, canal), 0)))
            manual = manual_map.get((sku, canal))
            large = 1 if (canal, sku) in lgset else 0
            black = bl_panel.get((sku, canal), 1 if (canal, sku) in blset else 0)
            # Objetivo por cobertura por CANAL (hoja Reglas). OJO: la col 'Cob' de la
            # Maestra resultó ser cobertura ACTUAL (llega a 158 meses), NO un objetivo,
            # así que NO se usa como target (inflaba el sugerido 17x). Se muestra como
            # informativa. Pendiente confirmar con comercial la fuente real del objetivo.
            cob = reglas.loc[canal, "cob_large"] if large else reglas.loc[canal, "cob"]
            candidatos = {"Venta": ume * cob, "Piso mínimo": umin, "Manual": manual or 0}
            origen = max(candidatos, key=candidatos.get)
            objetivo = candidatos[origen]
            # Stock disponible en full + tránsito hacia la bodega BF*. Se exponen por
            # separado (para las pestañas Stock / Tránsito del reporte). Piso a 0: un
            # stock negativo del canal (overstock/reservas) no debe inflar el objetivo.
            full_q = max(0.0, float(full_map.get((canal, sku), 0)))
            transito_q = max(0.0, float(tr_fulls.get((canal, sku), 0)))
            live_q = full_q + transito_q
            restr = 1 if cob_ca1 < reglas.loc[canal, "restr"] else 0
            repo = 0 if (black or restr or es_out) else mult2(max(0, objetivo - live_q))
            canal_rows.append({"canal": canal, "ume": ume, "manual": manual, "umin": umin,
                               "black": black, "restr": restr, "large": large, "origen": origen,
                               "objetivo": objetivo, "live_q": live_q, "repo": repo,
                               "full_q": full_q, "transito_q": transito_q, "cob": float(cob)})
        # ── Tope a CA1 (Claudia 24-ago): la suma del sugerido de los 4 canales por SKU
        # no puede superar el stock DISPONIBLE en CA1 (no se puede enviar lo que no hay).
        # Si excede, se reparte CA1 proporcional al sugerido de cada canal y se piso a
        # par (mult2 baja) para no volver a exceder.
        _tot = sum(r["repo"] for r in canal_rows)
        _capado = _tot > st_ca1 and _tot > 0
        if _capado:
            _factor = st_ca1 / _tot
            for r in canal_rows:
                r["repo"] = int(math.floor(r["repo"] * _factor / 2) * 2)
        for r in canal_rows:
            repo = r["repo"]
            filas.append({
                "Categoría": d.get("Categoría", ""), "Tipo de producto": tipo, "Marca": marca,
                "Pack": d.get("Pack", ""), "In/Out": d.get("In/Out", ""), "Sku": sku,
                "Producto": d.get("Producto", ""), "Canal": r["canal"],
                "UME": r["ume"], "UME MIN": r["umin"], "UME Manual": r["manual"],
                "Blacklist": r["black"], "Restricción": r["restr"], "Large": r["large"],
                "Origen del objetivo": r["origen"] if repo > 0 else "",
                "Cobertura canal": r["cob"],
                "Stock objetivo full": r["objetivo"],
                "Stock Full": r["full_q"], "Tránsito": r["transito_q"],
                "Stock Full Live+tránsito": r["live_q"], "Reposición": repo,
                "Dim m3": round((dim_m3.get(sku) or 0) * repo, 4),
                "Dim peso": round((dim_kg.get(sku) or 0) * repo, 2),
                "Stock CA1": st_ca1, "Demanda digital sem (c/packs)": round(dem, 2),
                "Cobertura CA1 (meses)": round(cob_ca1, 2) if np.isfinite(cob_ca1) else "",
                "Cob Maestra (info)": round(cob_obj_meses, 2) if pd.notna(cob_obj_meses) else "",
                "En Maestra": "No (nuevo)" if es_nuevo else "Sí",
                "Tope CA1": "Sí" if _capado else "",
            })
    df = pd.DataFrame(filas)

    act = df[df["Reposición"] > 0]
    res = act.groupby("Canal").agg(SKUs=("Sku", "nunique"), Unidades=("Reposición", "sum"),
                                   m3=("Dim m3", "sum"), kg=("Dim peso", "sum")).reset_index()
    def costo_envio(m3):
        if m3 <= 0: return 0
        pallets = math.ceil(m3 / 2) * 5500
        return pallets if m3 < 2.8 else (80000 + pallets if m3 <= 7 else 120000 + pallets)
    res["Costo envío est."] = res["m3"].apply(costo_envio)
    res["Tramo"] = res["m3"].apply(lambda x: "<2,8 m3 (solo pallets)" if x < 2.8
                                   else ("2,8-7 m3 ($80k)" if x <= 7 else ">12 m3 ($120k)" if x > 12 else "7-12 m3"))

    fn = OUT_DIR / f"Pulso_Reposicion_Fulfillment_{hoy:%Y%m%d}.xlsx"
    from openpyxl.styles import Font, PatternFill
    with pd.ExcelWriter(fn, engine="openpyxl") as w:
        res.to_excel(w, sheet_name="Resumen canal", index=False)
        df.sort_values(["Canal", "Reposición"], ascending=[True, False]).to_excel(w, sheet_name="UME v2", index=False)
        if cruce is not None:
            cruce["resumen"].to_excel(w, sheet_name="Cuadratura", index=False)
            cruce["detalle"].head(3000).to_excel(w, sheet_name="Cuadratura detalle", index=False)
        stats = pd.DataFrame([
            ["Corrida", f"{hoy} (semana ISO {hoy.isocalendar()[1]}) — MOTOR v2 (mesa 11-08)"],
            ["Semanas venta (L6W)", str(sem)],
            ["Venta", "RAW, SOLO canales digitales: Marketplace + Páginas propias + Fidelización"],
            ["Base de productos", "Drive PRICING, hoja Maestra (refrescada en cada corrida). In/Out='out' NO repone."],
            ["Objetivo de stock full", "Cobertura por CANAL (hoja Reglas) × venta digital semanal. La col 'Cob' de "
             "la Maestra es cobertura ACTUAL (no objetivo) -> se muestra informativa, no se usa como target."],
            ["Cobertura (restricción)", "Stock CA1/Stock ÷ demanda digital semanal ×4,33; demanda = venta SKU + packs "
             "que lo contienen (hoja Packs de Pricing). Umbral por canal = hoja Reglas."],
            ["Stock full", f"{fuente_live} + TRÁNSITO a fulls (picks internos pendientes hacia BF*)"],
            ["Blacklist/Large/Manual/MIN", "Panel de Nicole (drive original, actualización de los viernes)"],
            ["Stock Falabella", "FBF real desde la API del Seller Center (scrape_fala_fbf, endpoint "
             "/fby/v2/inbound-shipments/products) — reemplaza el Odoo BF*/feed. Si no hay parquet fresco, cae al fallback."],
        ], columns=["Item", "Detalle"])
        stats.to_excel(w, sheet_name="Metodología", index=False)
        for ws in w.book.worksheets:
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=AZUL)
            ws.freeze_panes = "A2"
    tot = int(act["Reposición"].sum())
    print(f"[pulso v2] {fn.name} | {len(act)} SKU-canal | {tot:,} uds | "
          f"restricción activa en {int((df['Restricción']==1).sum())} filas")
    return fn


def construir_excel_dinamico(flat_path=None) -> bytes:
    """Reporte con TABLA DINÁMICA NATIVA (hoja 'Dinámica': Canal→Marca→SKU; valores
    Reposición/m³/Stock CA1/Demanda) sobre la sábana 'UME v2'. Inyecta en la plantilla
    (armada 1 vez con Excel COM) + zip-surgery (refreshOnLoad) — el runtime NO usa Excel.
    Copia además Resumen canal / Cuadratura / Metodología como hojas planas."""
    import io as _io, re as _re, zipfile as _zip
    import openpyxl
    from openpyxl.utils import get_column_letter
    if flat_path is None:
        flat_path = construir()
    xls = pd.ExcelFile(flat_path)
    sabana = xls.parse("UME v2")
    ncols = len(sabana.columns)

    def _clean(row):
        return [None if (isinstance(v, float) and pd.isna(v)) else v for v in row]

    wb = openpyxl.load_workbook(TEMPLATE_DIN)
    ws = wb["Datos"]
    ws.delete_rows(1, ws.max_row)
    ws.append(list(sabana.columns))
    for row in sabana.itertuples(index=False):
        ws.append(_clean(row))
    for sn in xls.sheet_names:
        if sn == "UME v2":
            continue
        d = xls.parse(sn)
        w = wb.create_sheet(sn[:31])
        w.append([str(c) for c in d.columns])
        for row in d.itertuples(index=False):
            w.append(_clean(row))
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)

    ref = f"A1:{get_column_letter(ncols)}{len(sabana) + 1}"
    zin = _zip.ZipFile(buf, "r"); out = _io.BytesIO()
    with _zip.ZipFile(out, "w", _zip.ZIP_DEFLATED) as zout:
        for it in zin.infolist():
            d = zin.read(it.filename)
            if it.filename == "xl/pivotCache/pivotCacheDefinition1.xml":
                x = d.decode("utf-8")
                x = _re.sub(r'(<worksheetSource ref=")[^"]+(")', r"\g<1>" + ref + r"\g<2>", x)
                if "refreshOnLoad" not in x.split(">")[0]:
                    x = _re.sub(r"(<pivotCacheDefinition )", r'\g<1>refreshOnLoad="1" ', x, count=1)
                else:
                    x = _re.sub(r'refreshOnLoad="[^"]*"', 'refreshOnLoad="1"', x)
                d = x.encode("utf-8")
            zout.writestr(it, d)
    zin.close()
    return out.getvalue()


def construir_excel_5pestanas(flat_path=None) -> bytes:
    """Reporte en 5 pestañas con FÓRMULAS (Andrés/Claudia 25-ago). NO cambia el cálculo
    (Python arma los componentes en la sábana 'UME v2'); cambia el FORMATO:
      1. Propuesta Reposición — por FÓRMULA: tope_CA1(MAX(0, Máximo − Stock − Tránsito))
      2. Stock inicial marketplace (valores)
      3. Tránsito marketplace (valores)
      4. CA1 disponible + Máximos — UME manual EDITABLE → Máximo = fórmula
      5. Dinámica — la tabla dinámica nativa de siempre
    Al editar la UME manual (pestaña 4) recalcula el Máximo y la Propuesta en vivo."""
    import io as _io, re as _re, zipfile as _zip
    import openpyxl
    from openpyxl.utils import get_column_letter as L
    from openpyxl.styles import Font, PatternFill
    if flat_path is None:
        flat_path = construir()
    xls = pd.ExcelFile(flat_path)
    sabana = xls.parse("UME v2")
    ncols = len(sabana.columns)

    s = sabana.copy()
    for c in ["Stock Full", "Tránsito", "UME", "UME MIN", "UME Manual", "Cobertura canal",
              "Stock CA1", "Reposición"]:
        s[c] = pd.to_numeric(s.get(c), errors="coerce")
    s["_hab"] = ((pd.to_numeric(s["Blacklist"], errors="coerce").fillna(0) == 0)
                 & (pd.to_numeric(s["Restricción"], errors="coerce").fillna(0) == 0)
                 & (s["In/Out"].astype(str).str.strip().str.lower() != "out")).astype(int)
    orden = s.groupby("Sku")["Reposición"].sum().sort_values(ascending=False).index.tolist()
    meta = s.groupby("Sku").agg(Producto=("Producto", "first"), Marca=("Marca", "first"),
                                CA1=("Stock CA1", "first")).reindex(orden)
    rowmap = {(r["Sku"], r["Canal"]): r for _, r in s.iterrows()}  # lookup O(1)

    def val(sku, canal, col):
        r = rowmap.get((sku, canal))
        return float(r[col]) if r is not None and pd.notna(r[col]) else 0.0

    def man(sku, canal):
        r = rowmap.get((sku, canal))
        v = r["UME Manual"] if r is not None else None
        return float(v) if pd.notna(v) else None

    def hab(sku, canal):
        r = rowmap.get((sku, canal))
        return int(r["_hab"]) if r is not None else 1

    wb = openpyxl.load_workbook(TEMPLATE_DIN)

    def _clean(row):
        return [None if (isinstance(v, float) and pd.isna(v)) else v for v in row]
    wd = wb["Datos"]; wd.delete_rows(1, wd.max_row); wd.append(list(sabana.columns))
    for row in sabana.itertuples(index=False):
        wd.append(_clean(row))
    for sn in xls.sheet_names:
        if sn == "UME v2":
            continue
        dd = xls.parse(sn); w = wb.create_sheet(sn[:31]); w.append([str(c) for c in dd.columns])
        for row in dd.itertuples(index=False):
            w.append(_clean(row))

    BOLD = Font(bold=True); HFILL = PatternFill("solid", fgColor="DDEBF7")
    EDIT = PatternFill("solid", fgColor="FFF2CC")

    def _hdr(ws, headers):
        for j, h in enumerate(headers, 1):
            c = ws.cell(1, j, h); c.font = BOLD; c.fill = HFILL
        ws.freeze_panes = "D2"

    # 2. Stock inicial MKT / 3. Tránsito MKT (valores)
    for nombre, col in [("2. Stock inicial MKT", "Stock Full"), ("3. Transito MKT", "Tránsito")]:
        ws = wb.create_sheet(nombre)
        _hdr(ws, ["SKU", "Producto", "Marca"] + CANALES)
        for i, sku in enumerate(orden, 2):
            ws.cell(i, 1, sku); ws.cell(i, 2, meta.loc[sku, "Producto"]); ws.cell(i, 3, meta.loc[sku, "Marca"])
            for j, canal in enumerate(CANALES, 4):
                ws.cell(i, j, val(sku, canal, col))

    # 4. CA1 y Máximos (UME manual editable → Máximo fórmula)
    cm = wb.create_sheet("4. CA1 y Maximos")
    hdr4 = ["SKU", "Producto", "Marca", "CA1 disp."]
    for canal in CANALES:
        hdr4 += [f"UME {canal}", f"Cob {canal}", f"UME MIN {canal}", f"UME Manual {canal}",
                 f"Habilit {canal}", f"Máximo {canal}"]
    _hdr(cm, hdr4)
    maxcol, habcol = {}, {}
    for k, canal in enumerate(CANALES):
        base = 5 + k * 6
        maxcol[canal] = base + 5; habcol[canal] = base + 4
    for i, sku in enumerate(orden, 2):
        cm.cell(i, 1, sku); cm.cell(i, 2, meta.loc[sku, "Producto"]); cm.cell(i, 3, meta.loc[sku, "Marca"])
        cm.cell(i, 4, float(meta.loc[sku, "CA1"]) if pd.notna(meta.loc[sku, "CA1"]) else 0)
        for k, canal in enumerate(CANALES):
            base = 5 + k * 6
            cm.cell(i, base, val(sku, canal, "UME"))
            cm.cell(i, base + 1, val(sku, canal, "Cobertura canal"))
            cm.cell(i, base + 2, val(sku, canal, "UME MIN"))
            mc = cm.cell(i, base + 3, man(sku, canal)); mc.fill = EDIT
            cm.cell(i, base + 4, hab(sku, canal))
            uL, cL, mnL, maL = L(base), L(base + 1), L(base + 2), L(base + 3)
            cm.cell(i, base + 5, f"=MAX({uL}{i}*{cL}{i},{mnL}{i},N({maL}{i}))")

    # 1. Propuesta Reposición (fórmula, con tope CA1) — al inicio
    pr = wb.create_sheet("1. Propuesta Reposicion", 0)
    _hdr(pr, ["SKU", "Producto", "Marca"] + CANALES + ["", "raw ML", "raw Fala", "raw Paris", "raw Walmart", "Suma raw"])
    STK, TRA, CMq = "'2. Stock inicial MKT'", "'3. Transito MKT'", "'4. CA1 y Maximos'"
    for i, sku in enumerate(orden, 2):
        pr.cell(i, 1, sku); pr.cell(i, 2, meta.loc[sku, "Producto"]); pr.cell(i, 3, meta.loc[sku, "Marca"])
        for k, canal in enumerate(CANALES):
            can = L(4 + k)  # D..G
            mx, hb = L(maxcol[canal]), L(habcol[canal])
            pr.cell(i, 9 + k, f"={CMq}!{hb}{i}*MROUND(MAX(0,{CMq}!{mx}{i}-{STK}!{can}{i}-{TRA}!{can}{i}),2)")
        pr.cell(i, 13, f"=SUM(I{i}:L{i})")
        for k, canal in enumerate(CANALES):
            rc = L(9 + k)
            pr.cell(i, 4 + k, f"=IF($M{i}>{CMq}!$D{i},FLOOR({rc}{i}*{CMq}!$D{i}/$M{i},2),{rc}{i})")
    for col in "IJKLM":  # ocultar helpers
        pr.column_dimensions[col].hidden = True

    # orden de pestañas: 1..4 + Dinámica + Datos + resto
    din = next((w for w in wb.worksheets if "din" in w.title.lower()), None)
    pref = ["1. Propuesta Reposicion", "2. Stock inicial MKT", "3. Transito MKT", "4. CA1 y Maximos"]
    if din:
        pref.append(din.title)
    pref.append("Datos")
    wb._sheets.sort(key=lambda w: pref.index(w.title) if w.title in pref else 90)

    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    ref = f"A1:{L(ncols)}{len(sabana) + 1}"
    zin = _zip.ZipFile(buf, "r"); out = _io.BytesIO()
    with _zip.ZipFile(out, "w", _zip.ZIP_DEFLATED) as zout:
        for it in zin.infolist():
            d = zin.read(it.filename)
            if it.filename == "xl/pivotCache/pivotCacheDefinition1.xml":
                x = d.decode("utf-8")
                x = _re.sub(r'(<worksheetSource ref=")[^"]+(")', r"\g<1>" + ref + r"\g<2>", x)
                if "refreshOnLoad" not in x.split(">")[0]:
                    x = _re.sub(r"(<pivotCacheDefinition )", r'\g<1>refreshOnLoad="1" ', x, count=1)
                else:
                    x = _re.sub(r'refreshOnLoad="[^"]*"', 'refreshOnLoad="1"', x)
                d = x.encode("utf-8")
            zout.writestr(it, d)
    zin.close()
    return out.getvalue()


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-download", action="store_true")
    ap.add_argument("--no-live", action="store_true")
    a = ap.parse_args()
    construir(no_download=a.no_download, no_live=a.no_live)
