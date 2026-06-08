"""
Pestaña "Cargar ventas offline" del dashboard.
Permite subir Excel con formato RAW y agregarlo/sustituirlo en Turso.

Casos de uso:
- SAWA (mes faltante)
- Walmart Fulfillment (carga manual)
- CMR (mientras no haya integración Drive)
- Cualquier carga ad-hoc donde Odoo no la registra
"""
import io
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import libsql_client


# Mapeo formato RAW (Excel) -> columna DB
RAW_TO_DB = {
    'Tipo Movimiento': 'tipo_movimiento',
    'Bodega': 'bodega',
    'Documento': 'documento',
    'Fecha Documento': 'fecha_documento',
    'Pedido': 'pedido',
    'Estado Pedido': 'estado_pedido',
    'Tipo Despacho': 'tipo_despacho',
    'SKU': 'sku',
    'Canal': 'canal',
    'Fecha Venta': 'fecha_venta',
    'Hora Venta': 'hora_venta',
    'Producto': 'producto',
    'Categoría macro': 'categoria_macro',
    'Categoría padre': 'categoria_padre',
    'Categoría hijo': 'categoria_hijo',
    'Categoría comercial': 'categoria_comercial',
    'Estado SKU': 'estado_sku',
    'Pack': 'pack',
    'Marca': 'marca',
    'Proveedor': 'proveedor',
    'Tipo Marca': 'tipo_marca',
    'Tipo Compra': 'tipo_compra',
    'Tipo Negocio': 'tipo_negocio',
    'KAM': 'kam',
    'Estado Canal': 'estado_canal',
    'Año venta': 'anio_venta',
    'Mes venta': 'mes_venta',
    'Semana venta': 'semana_venta',
    'Día semana': 'dia_semana',
    'Hora venta': 'hora_venta_num',
    'Cantidad': 'cantidad',
    'Venta bruta': 'venta_bruta',
    'Venta Neta': 'venta_neta',
    'Costo Unitario': 'costo_unitario',
    'Costo Total': 'costo_total',
    'Margen Front': 'margen_front',
    'Comision %': 'comision_pct',
    'Comisión': 'comision',
    'Logística': 'logistica',
    'Marketing': 'marketing',
    'Mg final': 'margen_final',
}
COLS_OBLIGATORIAS = ['Pedido', 'Fecha Venta', 'SKU', 'Cantidad', 'Venta bruta']


def _get_turso_client():
    return libsql_client.create_client_sync(
        url=os.environ.get('LIBSQL_URL', ''),
        auth_token=os.environ.get('LIBSQL_AUTH_TOKEN', ''),
    )


def turso_delete_pedidos(pedidos: list[str]) -> int:
    """Borra filas de ventas con pedido en la lista. Devuelve total filas afectadas."""
    if not pedidos:
        return 0
    client = _get_turso_client()
    total = 0
    try:
        for i in range(0, len(pedidos), 200):
            chunk = pedidos[i:i + 200]
            placeholders = ','.join(['?'] * len(chunk))
            sql = f"DELETE FROM ventas WHERE pedido IN ({placeholders})"
            res = client.execute(sql, [str(p) for p in chunk])
            total += getattr(res, 'rows_affected', 0) or 0
    finally:
        client.close()
    return total


def turso_batch_insert(rows, db_cols):
    """Inserta rows (lista de tuplas) en ventas con multi-row INSERT."""
    if not rows:
        return 0
    cols_csv = ','.join(db_cols)
    n_cols = len(db_cols)
    n_total = 0
    chunk = 200
    placeholders_per_row = '(' + ','.join(['?'] * n_cols) + ')'
    client = _get_turso_client()
    try:
        for i in range(0, len(rows), chunk):
            batch = rows[i:i + chunk]
            all_ph = ','.join([placeholders_per_row] * len(batch))
            sql = f"INSERT INTO ventas ({cols_csv}) VALUES {all_ph}"
            flat = []
            for r in batch:
                flat.extend(list(r))
            client.execute(sql, flat)
            n_total += len(batch)
    finally:
        client.close()
    return n_total


def turso_metadata_insert(fuente: str, n_filas: int, fmin, fmax, tipo: str = 'manual_offline'):
    """Inserta entrada en metadata_cargas."""
    client = _get_turso_client()
    try:
        client.execute(
            "INSERT INTO metadata_cargas (fecha_carga, fuente, filas_cargadas, fecha_min_datos, fecha_max_datos, tipo) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            [datetime.now().isoformat(), fuente, n_filas, str(fmin) if fmin else None, str(fmax) if fmax else None, tipo],
        )
    finally:
        client.close()


def render_carga_offline_tab():
    """Renderiza la pestaña de carga offline en Streamlit."""
    st.title("📤 Cargar ventas offline")
    st.caption("Subir Excel con ventas que NO entran por Odoo (SAWA, Walmart Fulfillment, CMR, etc.)")

    with st.expander("ℹ️ Formato esperado", expanded=False):
        st.markdown("""
        - **Excel (.xlsx)** con headers en la primera fila
        - **Columnas obligatorias**: `Pedido`, `Fecha Venta`, `SKU`, `Cantidad`, `Venta bruta`
        - **Columnas opcionales**: cualquiera de las 40 columnas RAW estándar
        - Las columnas faltantes se completan con valores por defecto (NULL o lo que indiques abajo)
        - Si marcas "Pisar pedidos existentes", se borran filas con esos números de pedido y se reinsertan
        """)

    # Botón descarga template
    template_bytes = _generar_template()
    st.download_button(
        label="📥 Descargar template Excel (con ejemplo)",
        data=template_bytes,
        file_name="Template_carga_offline_UnionX.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Excel con las 40 columnas RAW + 2 filas de ejemplo. Las obligatorias están marcadas en la fila de ayuda.",
    )

    col1, col2 = st.columns(2)
    with col1:
        canal_default = st.selectbox(
            "Canal asociado *",
            options=["CMR", "Sawa", "Walmart Fulfillment", "Otro (escribir)"],
            help="Se aplica si la columna Canal del Excel está vacía",
        )
        if canal_default == "Otro (escribir)":
            canal_default = st.text_input("Nombre del canal", value="")

    with col2:
        tipo_negocio_default = st.selectbox(
            "Tipo Negocio *",
            options=["Fidelización", "Marketplace", "Distribución", "B2B", "Otro (escribir)"],
            help="Se aplica si la columna Tipo Negocio del Excel está vacía",
        )
        if tipo_negocio_default == "Otro (escribir)":
            tipo_negocio_default = st.text_input("Nombre tipo negocio", value="")

    pisar = st.checkbox(
        "Pisar pedidos existentes con mismo número de pedido (DELETE + INSERT)",
        value=True,
        help="Útil para CMR: borra los pedidos Shopify con $0 y los reemplaza con tu data real",
    )

    file = st.file_uploader("📎 Subir archivo Excel", type=['xlsx', 'xls'])

    if file is None:
        st.info("Esperando archivo...")
        return

    # Leer Excel
    try:
        df_raw = pd.read_excel(file)
    except Exception as e:
        st.error(f"Error leyendo Excel: {e}")
        return

    st.success(f"✅ {len(df_raw):,} filas leídas, {len(df_raw.columns)} columnas")

    # Validar columnas obligatorias
    faltantes = [c for c in COLS_OBLIGATORIAS if c not in df_raw.columns]
    if faltantes:
        st.error(f"❌ Faltan columnas obligatorias: {faltantes}")
        st.write("**Columnas detectadas:**", list(df_raw.columns))
        return

    # Preview
    st.subheader("Preview (primeras 20 filas)")
    st.dataframe(df_raw.head(20), width='stretch')

    # Resumen
    total_venta = pd.to_numeric(df_raw['Venta bruta'], errors='coerce').fillna(0).sum()
    n_pedidos = df_raw['Pedido'].nunique()
    fecha_min = df_raw['Fecha Venta'].min()
    fecha_max = df_raw['Fecha Venta'].max()

    col1, col2, col3 = st.columns(3)
    col1.metric("Filas", f"{len(df_raw):,}")
    col2.metric("Pedidos únicos", f"{n_pedidos:,}")
    col3.metric("Venta total", f"${int(total_venta):,}".replace(",", "."))
    st.caption(f"📅 Rango fechas: {fecha_min} → {fecha_max}")

    # Confirmar carga
    if st.button("🚀 Cargar a Turso", type="primary"):
        with st.spinner("Procesando..."):
            try:
                _do_load(df_raw, canal_default, tipo_negocio_default, pisar)
            except Exception as e:
                st.error(f"❌ Error: {e}")
                import traceback
                st.code(traceback.format_exc())


def _do_load(df_raw: pd.DataFrame, canal_def: str, tipo_negocio_def: str, pisar: bool):
    """Procesa el DataFrame y lo carga a Turso."""
    # Renombrar a cols DB
    df = df_raw.rename(columns={k: v for k, v in RAW_TO_DB.items() if k in df_raw.columns}).copy()

    # Aplicar defaults
    if 'canal' not in df.columns or df['canal'].isna().all():
        df['canal'] = canal_def
    else:
        df['canal'] = df['canal'].fillna(canal_def)

    if 'tipo_negocio' not in df.columns or df['tipo_negocio'].isna().all():
        df['tipo_negocio'] = tipo_negocio_def
    else:
        df['tipo_negocio'] = df['tipo_negocio'].fillna(tipo_negocio_def)

    # Tipo Movimiento default = Venta
    if 'tipo_movimiento' not in df.columns:
        df['tipo_movimiento'] = 'Venta'
    else:
        df['tipo_movimiento'] = df['tipo_movimiento'].fillna('Venta')

    # Convertir fechas a string ISO YYYY-MM-DD
    for col in ['fecha_documento', 'fecha_venta']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')

    # Asegurar que todas las columnas de la tabla ventas existan en df
    db_cols = list(RAW_TO_DB.values())
    for col in db_cols:
        if col not in df.columns:
            df[col] = None

    # Reemplazar NaN por None
    df = df[db_cols].where(pd.notna(df[db_cols]), None)

    pedidos_unicos = list(set(p for p in df['pedido'].dropna().unique() if p))
    n_pedidos = len(pedidos_unicos)

    # Pisar si aplica
    if pisar and pedidos_unicos:
        n_borradas = turso_delete_pedidos(pedidos_unicos)
        st.write(f"🗑️ Borradas {n_borradas:,} filas previas con esos {n_pedidos} pedidos")

    # Insertar
    rows = list(df.itertuples(index=False, name=None))
    n_ins = turso_batch_insert(rows, db_cols)

    # Registrar en metadata_cargas
    fecha_min = df['fecha_venta'].min()
    fecha_max = df['fecha_venta'].max()
    fuente = f"Carga manual offline ({canal_def})"
    turso_metadata_insert(fuente, n_ins, fecha_min, fecha_max, tipo='manual_offline')

    st.success(f"✅ {n_ins:,} filas insertadas en Turso (canal: {canal_def}, línea: {tipo_negocio_def})")
    st.info("Refrescá el dashboard (botón ⚡ Forzar sync) o esperá 5 min para que aparezca la data nueva en los KPIs.")
    st.cache_resource.clear()
    st.cache_data.clear()


def _generar_template() -> bytes:
    """Genera Excel template con las 40 columnas + 2 filas de ejemplo."""
    cols_raw = list(RAW_TO_DB.keys())

    # Fila 1: ejemplo CMR (con Pedido cruzando con Shopify)
    ejemplo_cmr = {
        'Tipo Movimiento': 'Venta',
        'Bodega': 'Bodega Principal',
        'Documento': 'CMR-001234',
        'Fecha Documento': '2026-04-15',
        'Pedido': '#SH123456',
        'Estado Pedido': 'Pagado',
        'Tipo Despacho': 'Domicilio',
        'SKU': 'ABC-001',
        'Canal': 'CMR',
        'Fecha Venta': '2026-04-15',
        'Hora Venta': '14:30',
        'Producto': 'Pala Lhotse',
        'Categoría macro': 'Outdoor',
        'Categoría padre': 'Snow',
        'Categoría hijo': 'Palas',
        'Categoría comercial': 'Standard',
        'Estado SKU': 'Activo',
        'Pack': '1',
        'Marca': 'Lhotse',
        'Proveedor': 'Lhotse Chile',
        'Tipo Marca': 'Propia',
        'Tipo Compra': 'Importación',
        'Tipo Negocio': 'Fidelización',
        'KAM': 'Clau',
        'Estado Canal': 'Activo',
        'Año venta': 2026,
        'Mes venta': 4,
        'Semana venta': 16,
        'Día semana': 'Lunes',
        'Hora venta': 14,
        'Cantidad': 1,
        'Venta bruta': 49990,
        'Venta Neta': 42008,
        'Costo Unitario': 18000,
        'Costo Total': 18000,
        'Margen Front': 24008,
        'Comision %': 0,
        'Comisión': 0,
        'Logística': 2500,
        'Marketing': 0,
        'Mg final': 21508,
    }

    # Fila 2: ejemplo SAWA (carga simple, mínimo de columnas)
    ejemplo_sawa = {c: None for c in cols_raw}
    ejemplo_sawa.update({
        'Tipo Movimiento': 'Venta',
        'Documento': 'SAWA-04-001',
        'Pedido': 'SAWA-001',
        'SKU': 'XYZ-099',
        'Canal': 'Sawa',
        'Fecha Venta': '2026-04-10',
        'Producto': 'Producto ejemplo',
        'Marca': 'UnionX',
        'Tipo Negocio': 'Fidelización',
        'Año venta': 2026,
        'Mes venta': 4,
        'Cantidad': 1,
        'Venta bruta': 19990,
        'Venta Neta': 16798,
    })

    df = pd.DataFrame([ejemplo_cmr, ejemplo_sawa], columns=cols_raw)

    # Generar Excel con styling: marcar columnas obligatorias
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Template')
        ws = writer.sheets['Template']
        # Marcar headers obligatorios en amarillo
        from openpyxl.styles import PatternFill, Font
        amarillo = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
        bold = Font(bold=True)
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold
            if col_name in COLS_OBLIGATORIAS:
                cell.fill = amarillo
        # Ancho columnas auto
        for col_idx in range(1, len(df.columns) + 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f'A{chr(64 + col_idx - 26)}'].width = 18

    return buf.getvalue()
