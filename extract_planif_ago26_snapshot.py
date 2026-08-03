"""
Extrae datos de 4 hojas desde analisis_planificacion_AGO26 v2 subida.xlsx y guarda parquets.

Hojas procesadas:
  - Critico x Marca         → planif_critico_marca_snapshot.parquet
  - Sobrestock x SKU Padre  → planif_sobrestock_snapshot.parquet
  - Tránsitos por Embarque  → planif_transitos_snapshot.parquet
  - Nuevos en Tránsito      → planif_nuevos_transito_snapshot.parquet

Uso: python extract_planif_ago26_snapshot.py
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import pandas as pd
from pathlib import Path

EXCEL_PATH = Path(r"C:\Users\felip\Desktop\UNIONX\FORECAST FINAL SKU\Analisis Planificacion\analisis_planificacion_AGO26 v2 subida.xlsx")
OUT_DIR    = Path(__file__).parent / "data" / "planificacion" / "snapshots"
OUT_DIR.mkdir(parents=True, exist_ok=True)

print(f"Leyendo {EXCEL_PATH.name} ...")
wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)


# ══════════════════════════════════════════════════════════════════════════════
# 1. Critico x Marca
# ══════════════════════════════════════════════════════════════════════════════
def _extract_critico():
    ws = wb['Critico x Marca']
    rows = list(ws.iter_rows(values_only=True))
    # R0 = title, R1 = headers (Marca|SKUs|Cob.Prom|SinStock|E:SKIP|F:SKIP|StockHoyCST|VentaCST|DetalleLlegadas)
    records = []
    for r in rows[2:]:
        if not r[0]:
            continue
        marca = str(r[0]).strip()
        is_total = (marca == 'TOTAL')
        records.append({
            'marca':           marca,
            'skus':            int(r[1] or 0),
            'cob_prom':        float(r[2] or 0),
            'sin_stock':       int(r[3] or 0),
            # cols 4 (Stock Hoy Unidades) y 5 (Venta PPTO Unid) OMITIDOS
            'stock_hoy_cst':   float(r[6] or 0),
            'venta_cst_ago26': float(r[7] or 0),
            'detalle_llegadas': str(r[8]).strip() if r[8] else '',
            'is_total':        is_total,
        })
    df = pd.DataFrame(records)
    out = OUT_DIR / 'planif_critico_marca_snapshot.parquet'
    df.to_parquet(str(out), index=False)
    print(f"  ✅ critico_marca: {len(df)} filas → {out.name}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 2. Sobrestock x SKU Padre
# ══════════════════════════════════════════════════════════════════════════════
def _clean_nombre(s):
    for ch in ('▶', '▸', '▹', '↳'):
        s = s.replace(ch, '')
    return s.strip()


def _extract_sobrestock():
    ws = wb['Sobrestock x SKU Padre']
    rows = list(ws.iter_rows(values_only=True))
    # R0 = title, R1 = headers
    records = []
    cur_marca = ''
    cur_cat_padre = ''
    cur_cat_hijo = ''
    for r in rows[2:]:
        col0 = r[0]
        if col0 is None:
            continue
        col0_str = str(col0).strip()
        if not col0_str:
            continue
        # Detect level from prefix
        if col0_str.startswith('▶'):
            nivel = 1
        elif '▸' in col0_str:
            nivel = 2
        elif '▹' in col0_str:
            nivel = 3
        elif '↳' in col0_str:
            nivel = 4
        else:
            nivel = 1

        nombre_clean = _clean_nombre(col0_str)

        if nivel == 1:
            cur_marca     = nombre_clean
            cur_cat_padre = ''
            cur_cat_hijo  = ''
        elif nivel == 2:
            cur_cat_padre = nombre_clean
            cur_cat_hijo  = ''
        elif nivel == 3:
            cur_cat_hijo  = nombre_clean

        records.append({
            'nombre':               str(col0),
            'nombre_clean':         nombre_clean,
            'marca_parent':         cur_marca,
            'cat_padre_parent':     cur_cat_padre,
            'cat_hijo_parent':      cur_cat_hijo if nivel == 4 else '',
            'descripcion':          str(r[1]).strip() if r[1] else '',
            'skus':                 r[2],
            'cobert_act':           float(r[3]) if r[3] is not None else None,
            'meses_exceso':         float(r[4]) if r[4] is not None else None,
            'stock_cst':            float(r[5]) if r[5] is not None else None,
            'venta_cst_prom':       float(r[6]) if r[6] is not None else None,
            'stock_optimo':         float(r[7]) if r[7] is not None else None,
            'capital_inmovilizado': float(r[8]) if r[8] is not None else None,
            'tiene_llegadas':       str(r[9]).strip() if r[9] else '',
            'nivel':                nivel,
        })
    df = pd.DataFrame(records)
    out = OUT_DIR / 'planif_sobrestock_snapshot.parquet'
    df.to_parquet(str(out), index=False)
    print(f"  ✅ sobrestock: {len(df)} filas → {out.name}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 3. Tránsitos por Embarque
# ══════════════════════════════════════════════════════════════════════════════
def _extract_transitos():
    ws = wb['Tránsitos por Embarque']
    rows = list(ws.iter_rows(values_only=True))
    # R0 = title, R1 = headers
    # PI rows:  col[0] = 'XXPI', col[4] = marcas (string), col[5] = int skus_distintos
    # SKU rows: col[0] starts with '  ↳  '
    records = []
    current_pi = None
    for r in rows[2:]:
        col0 = r[0]
        if col0 is None:
            continue
        col0_str = str(col0).strip()
        if not col0_str:
            continue

        if col0_str.startswith('↳') or str(r[0]).startswith('  ↳'):
            # SKU row
            sku_raw = col0_str.lstrip('↳ ').strip()
            records.append({
                'row_type':      'sku',
                'pi_embarque':   current_pi,
                'eta_o_desc':    str(r[1]).strip() if r[1] else '',
                'eta_bodega':    str(r[2]).strip() if r[2] else '',
                'mes_llegada':   str(r[3]).strip() if r[3] else '',
                'marcas':        '',
                'skus_distintos': None,
                'criticos':      str(r[6]).strip() if r[6] else '',
                'inquietos':     str(r[7]).strip() if r[7] else '',
                'unidades':      int(r[8]) if r[8] is not None else 0,
                'valor_usd':     float(r[9]) if r[9] is not None else 0.0,
                'nivel_riesgo':  '',
                'sku':           sku_raw,
            })
        else:
            # PI row
            current_pi = col0_str
            records.append({
                'row_type':      'pi',
                'pi_embarque':   col0_str,
                'eta_o_desc':    str(r[1]).strip() if r[1] else '',
                'eta_bodega':    str(r[2]).strip() if r[2] else '',
                'mes_llegada':   str(r[3]).strip() if r[3] else '',
                'marcas':        str(r[4]).strip() if r[4] else '',
                'skus_distintos': int(r[5]) if r[5] is not None else 0,
                'criticos':      str(r[6]).strip() if r[6] is not None else '',
                'inquietos':     str(r[7]).strip() if r[7] is not None else '',
                'unidades':      int(r[8]) if r[8] is not None else 0,
                'valor_usd':     float(r[9]) if r[9] is not None else 0.0,
                'nivel_riesgo':  str(r[10]).strip() if r[10] else '',
                'sku':           '',
            })
    df = pd.DataFrame(records)
    out = OUT_DIR / 'planif_transitos_snapshot.parquet'
    df.to_parquet(str(out), index=False)
    print(f"  ✅ transitos: {len(df)} filas (PI={len(df[df.row_type=='pi'])}, SKU={len(df[df.row_type=='sku'])}) → {out.name}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 4. Nuevos en Tránsito
# ══════════════════════════════════════════════════════════════════════════════
def _extract_nuevos():
    ws = wb['Nuevos en Tránsito']
    rows = list(ws.iter_rows(values_only=True))
    # R0 = title, R1 = headers
    # Group headers: col[0] starts with '  ▶  Llegada XXX'
    # TOTAL row: col[0] == 'TOTAL'
    records = []
    current_grupo = ''
    for r in rows[2:]:
        col0 = r[0]
        col0_str = str(col0).strip() if col0 is not None else ''
        if not col0_str:
            continue
        if col0_str.startswith('TOTAL'):
            continue
        if '▶' in col0_str:
            # Group header
            current_grupo = col0_str.replace('▶', '').strip()
            continue
        # Data row
        records.append({
            'grupo':           current_grupo,
            'sku':             col0_str,
            'descripcion':     str(r[1]).strip() if r[1] else '',
            'marca':           str(r[2]).strip() if r[2] else '',
            'mes_llegada':     str(r[3]).strip() if r[3] else '',
            'fecha_eta_bodega': str(r[4]).strip() if r[4] else '',
            'cantidad':        int(r[5]) if r[5] is not None else 0,
        })
    df = pd.DataFrame(records)
    out = OUT_DIR / 'planif_nuevos_transito_snapshot.parquet'
    df.to_parquet(str(out), index=False)
    print(f"  ✅ nuevos_transito: {len(df)} filas → {out.name}")
    return df


if __name__ == '__main__':
    _extract_critico()
    _extract_sobrestock()
    _extract_transitos()
    _extract_nuevos()
    print("\n✅ Todos los parquets generados correctamente.")
