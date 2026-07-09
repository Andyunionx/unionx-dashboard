# -*- coding: utf-8 -*-
"""Patch v18 -> v19: costeo tercerización refinado (etiquetado 100% B2B + extras).
EIT ahorro negociado +$64,2M (antes 79,7). Full-stack +$349M."""
import sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')
BASE = r"G:\Mi unidad\TRABAJO\RESPALDO\OPERACIONES\UNION X - IA"
wb = openpyxl.load_workbook(BASE + r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v18.xlsx")
ws = wb['8. 2027 y Palancas']
for mr in list(ws.merged_cells.ranges):
    if mr.min_row >= 37: ws.unmerge_cells(str(mr))
FB=Font(bold=True); GRN=Font(color='1E7A1E',bold=True); IT=Font(size=10,italic=True,color='C00000'); WH=Font(bold=True,color='FFFFFF')
fgrn=PatternFill('solid',fgColor='E3F4E4'); fsub=PatternFill('solid',fgColor='D6E4F0'); fhdr=PatternFill('solid',fgColor='0F2340'); blank=PatternFill()
TH=Border(*[Side(style='thin',color='D9D9D9')]*4); RIG=Alignment(horizontal='right'); CEN=Alignment(horizontal='center'); LFT=Alignment(horizontal='left',wrap_text=True)
INT=144; RNO=-169; DEU=1729; DA=18; PAT=625; OT=31.7

# ===== EIT detail block (rows 37-47) — tabla refinada =====
for r in range(37,48):
    for c in range(1,6): ws.cell(r,c).value=None; ws.cell(r,c).fill=blank; ws.cell(r,c).border=Border()
ws.cell(37,1,'PALANCA EIT — Tercerizar operación logística (cambio de bodega, inicia 01/01/2027)').font=FB
hdr=['Concepto (MM)','2027 tarifa INICIAL','2027 NEGOCIADA 08-jul','2028']
for j,h in enumerate(hdr):
    c=ws.cell(38,1+j); c.value=h; c.font=WH; c.fill=fhdr; c.alignment=(LFT if j==0 else CEN); c.border=TH
eit=[('EIT tarifas (con etiquetado 100% B2B)',365.7,331.0,'~381',False),
     ('Extras cotización (rotulación + arriendo pallet + insumos $10M)',10.2,21.8,'~23,6',False),
     ('EIT FINAL',375.9,352.8,'—',True),
     ('Tercerizar final (+ residual con leyes)',495.8,472.7,'—',False),
     ('AHORRO FINAL (con leyes)','+41,1','+64,2','—',True)]
for i,(n,a,b,c28,bold) in enumerate(eit):
    r=39+i
    ws.cell(r,1,n).alignment=LFT; ws.cell(r,1).border=TH
    for cc,val in ((2,a),(3,b),(4,c28)):
        cel=ws.cell(r,cc); cel.value=val; cel.alignment=RIG; cel.border=TH
        if isinstance(val,(int,float)): cel.number_format='#,##0.0'
    if bold:
        for cc in range(1,5): ws.cell(r,cc).font=FB
        if n.startswith('AHORRO'):
            for cc in range(1,5): ws.cell(r,cc).fill=fgrn
            ws.cell(r,2).font=GRN; ws.cell(r,3).font=GRN
        else:
            for cc in range(1,5): ws.cell(r,cc).fill=fsub
ws.cell(44,1,'Recomendación: AVANZAR con TARIFAS NEGOCIADAS 08-jul (+$64,2M/2027 con leyes) sobre operar interno $536,9M. Ya incluye etiquetado 100% B2B + extras (rotulación, arriendo pallet, insumos $10M). Tarifa inicial daba +$41,1M.').font=IT
ws.cell(45,1,'El finiquito de los operarios (one-time $31,7M) se reconoce en 2027 (inicio 01/01/2027). Ahorro pleno desde el año 1.').font=IT

# ===== SUMATORIA (EIT 64) =====
steps=[('Base digital $6.500M',None,20,6500,0),
       ('+ B2B UnionX $300M (contrib. 35%)',105,125,6800,0),
       ('+ Reducción gerencial · SOLO Nicole (nov-26)',55,180,6800,0),
       ('+ Marketing · Michela [ejecutada]',46,226,6800,0),
       ('+ Ecommerce · Ignacia [ejecutada]',33,259,6800,0),
       ('+ Tercerizar EIT (tarifas negociadas +$64,2M)',64,323,6800,OT),
       ('+ Ruteros Trade Mkt (dic-26)',26,349,6800,OT)]
for i,(nom,ah,util,vta,ot) in enumerate(steps):
    r=70+i; ebit=util-RNO; ebitda=ebit+DA
    vals=[nom,ah,ebit,ebitda,util,ebit/vta,util/vta,round(ebit/INT,2),round(DEU/ebitda,2),round(DEU/(PAT+util-ot),2)]
    for j,v in enumerate(vals):
        c=ws.cell(r,1+j); c.border=TH; c.font=Font()
        if j==0: c.value=v; c.alignment=LFT; c.font=FB if i in(0,6) else Font()
        elif v is None: c.value='—'; c.alignment=CEN
        elif j in(5,6): c.value=v; c.number_format='0.0%'; c.alignment=RIG
        elif j in(7,8,9): c.value=v; c.number_format='0.00"x"'; c.alignment=RIG
        else: c.value=v; c.number_format='#,##0;[Red]-#,##0'; c.alignment=RIG
    fill = fsub if i in(0,1) else (fgrn if i==6 else blank)
    for cc in range(1,11): ws.cell(r,cc).fill=fill
    if i in(0,6): ws.cell(r,5).font=GRN
ws.cell(77,1,'Erich NO sale: gerencial = solo Nicole ($55M). 5 palancas +$224M -> full-stack +$349M normalizado, cob 3,6x, D/EBITDA 3,2x, D/Patrimonio 1,8x. EIT negociado +$64,2M (etiquetado 100% B2B + extras incluidos). Finiquito bodega $31,7M one-time 2027. 2026 queda en +$7M.').font=IT

# ===== VISTA por palanca: EIT aporte 64, total 224 =====
ws.cell(83,3,'+64'); ws.cell(83,3).alignment=RIG
ws.cell(83,6,'Tarifas negociadas 08-jul (+$64,2M): incluye etiquetado 100% B2B + extras. Finiquito $31,7M one-time 2027 (cambio inicia 01/01/2027).')
ws.cell(85,3,'+224'); ws.cell(85,3).alignment=RIG

# ===== Puente patrimonio: util 2027 349, patrimonio 942 =====
ws.cell(104,2,365); ws.cell(104,1,'(+) Utilidad 2027 normalizada (full-stack)')
ws.cell(104,2,349)
ws.cell(106,2,942)
ws.cell(108,2,round(DEU/942,2))

# ===== Hoja 7 =====
r7=wb['7. Ratios 2023-2027']
r7.cell(16,1,'Con las 5 palancas (full-stack +$349M normalizado): EBIT% 7,6% · Cobertura 3,6x · Deuda/EBITDA 3,2x · Deuda/Patrimonio 1,8x. EIT negociado +$64,2M. Finiquito bodega $31,7M one-time 2027; 2026 queda en +$7M.').font=IT

# ===== Hoja 1 =====
w1=wb['1. Resumen Ejecutivo']
for r in range(1,80):
    v=str(w1.cell(r,1).value or '')
    if v.startswith('2027'):
        w1.cell(r,1,'2027: base digital $6.500M (+$20M) + B2B $300M (+$105M) = +$125M. Las 5 palancas (+$224M; gerencial solo Nicole $55M — Erich NO sale; EIT/cambio de bodega negociado +$64,2M con etiquetado 100% B2B + extras, inicia 01/01/2027) llevan el 2027 a +$349M normalizado, cob 3,6x, D/EBITDA 3,2x, D/Patrimonio 1,8x. Finiquito de bodega ($31,7M) one-time 2027. 2026 POSITIVO +$7M.')
        break

OUT=BASE+r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v19.xlsx"
wb.save(OUT)
print('Guardado v19 | EIT +64,2 | full-stack +349 | D/Pat 1,83x')
