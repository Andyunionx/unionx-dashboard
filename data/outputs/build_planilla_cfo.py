# -*- coding: utf-8 -*-
"""Planilla CFO UnionX — v2. Tendencia 2023-2027, vertical/horizontal, H1 real, gasto trend."""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')
BASE=r"G:\Mi unidad\TRABAJO\RESPALDO\OPERACIONES\UNION X - IA"
AZ='4884FC'; AZD='1F3A8A'; SUBF='D6E4F0'
FH=Font(bold=True,color='FFFFFF',size=10); FT=Font(bold=True,size=13,color='1F3A8A')
FB=Font(bold=True); FI=Font(italic=True,size=9,color='808080')
fill_h=PatternFill('solid',fgColor=AZ); fill_sub=PatternFill('solid',fgColor=SUBF)
fill_yel=PatternFill('solid',fgColor='FFF2CC'); fill_red=PatternFill('solid',fgColor='FCE4E4'); fill_grn=PatternFill('solid',fgColor='E3F4E4')
RED=Font(color='C00000'); GRN=Font(color='1E7A1E',bold=True)
THIN=Border(*[Side(style='thin',color='D9D9D9')]*4)
RIG=Alignment(horizontal='right'); LEF=Alignment(horizontal='left'); CEN=Alignment(horizontal='center')
YRS=[2023,2024,2025,2026,2027]

def hdr(ws,row,cols,widths=None):
    for j,c in enumerate(cols):
        cell=ws.cell(row,1+j,c); cell.font=FH; cell.fill=fill_h; cell.border=THIN
        cell.alignment=LEF if j==0 else CEN
    if widths:
        for j,w in enumerate(widths): ws.column_dimensions[get_column_letter(1+j)].width=w
def n(cell,v,dec=0,bold=False,pct=False,x=False):
    if v is None: cell.value='—'; cell.alignment=CEN; cell.border=THIN; return
    cell.value=v; cell.alignment=RIG; cell.border=THIN
    cell.number_format=('0.0"x"' if x else ('0.0%' if pct else ('#,##0.'+('0'*dec) if dec else '#,##0')+';[Red]-#,##0'+('.'+'0'*dec if dec else '')))
    if bold: cell.font=FB
def title(ws,t,sub=None):
    ws['A1']=t; ws['A1'].font=FT
    if sub: ws['A2']=sub; ws['A2'].font=FI

# datos P&L 2023-2027
PL={'Ingresos':[4939,5978,6164,6586,7200],'Margen de Contribución':[1500,1690,1773,1782,1950],
 'GAV':[-928,-1353,-1565,-1594,-1608],'EBIT':[572,336,208,188,342],'EBITDA':[600,336,224,212,360],
 'Resultado No Operacional':[-85,-317,-83,-200,-169],'Utilidad del Ejercicio':[487,19,124,-11,173]}
INT=[115,155,176,157,144]; DEUDA=[1279,1942,1954,1936,1729]
wb=Workbook()

# ===== HOJA 1: RESUMEN =====
ws=wb.active; ws.title='1. Resumen Ejecutivo'
title(ws,'UnionX — Análisis CFO · Tendencia 2023-2027 (MM CLP)','2023-2025 real (P&L). 2026 base planilla Planificación Financiera (incluye salidas). 2027 proyección Operaciones $7.200M + palancas.')
hdr(ws,4,['P&L']+[str(y) for y in YRS],[32,11,11,11,11,11])
r=5
for lab,vals in PL.items():
    ws.cell(r,1,lab).border=THIN
    for j,v in enumerate(vals): n(ws.cell(r,2+j),v,bold=lab in('Margen de Contribución','EBIT','EBITDA','Utilidad del Ejercicio'))
    if lab in('EBIT','EBITDA','Utilidad del Ejercicio','Margen de Contribución'):
        for col in range(1,7): ws.cell(r,col).fill=fill_yel
        ws.cell(r,1).font=FB
    if lab=='Utilidad del Ejercicio': ws.cell(r,5).font=RED; ws.cell(r,6).font=GRN
    r+=1
r+=1
ws.cell(r,1,'Indicadores').font=FB
hdr(ws,r,['Indicadores']+[str(y) for y in YRS]); r+=1
cob=[round(PL['EBIT'][i]/INT[i],2) for i in range(5)]
de=[round(DEUDA[i]/PL['EBITDA'][i],2) for i in range(5)]
for lab,vals,fmt in [('Cobertura de intereses (x)',cob,'x'),('Deuda / EBITDA (x)',de,'x'),
    ('EBIT %',[PL['EBIT'][i]/PL['Ingresos'][i] for i in range(5)],'%'),
    ('Utilidad neta %',[PL['Utilidad del Ejercicio'][i]/PL['Ingresos'][i] for i in range(5)],'%')]:
    ws.cell(r,1,lab).border=THIN
    for j,v in enumerate(vals): n(ws.cell(r,2+j),v,dec=2,x=(fmt=='x'),pct=(fmt=='%'))
    r+=1
r+=1
ws.cell(r,1,'Resultado No Operacional 2026 — desglose (¿cómo se calcula?)').font=FB; r+=1
for lab,v in [('Otros ingresos (fluctuación positiva / forwards)',132),('Intereses y reajustes',-157),
   ('Fluctuación de cambios (spot USD)',-125),('Gastos bancarios',-25),('Otros',-24),('= RNO 2026',-199)]:
    ws.cell(r,1,lab).border=THIN; n(ws.cell(r,2),v,bold='RNO' in lab); r+=1
r+=1
for m in ['TENDENCIA MACRO: EBIT% cayó de 11,6% (2023) a 2,9% (2026); cobertura de 5,0x a 1,2x; Deuda/EBITDA de 2,1x a 9,1x. Deterioro estructural de 4 años.',
 'B2B: el problema es MARGEN DIRECTO (Corporativo vende bajo costo, Distribución 24%), NO volumen (H1 real 94% de meta). El Sheet KAM subestimaba el B2B.',
 'Reajuste salarial 2027 (+5%) casi neutraliza el ahorro de las salidas: GAV sube +$14M pese a desvinculaciones. La base subió +7% en 2026 (dato real).',
 '2027: utilidad +$173M, cobertura 2,4x, Deuda/EBITDA 4,8x. El desapalancamiento viene de la amortización comercial ($210M) + EBITDA creciente — el destock NO es factible y se descartó.']:
    c=ws.cell(r,1,m); c.font=Font(size=10); c.alignment=Alignment(wrap_text=True,vertical='top'); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); ws.row_dimensions[r].height=30; r+=1

# ===== HOJA 2: VERTICAL / HORIZONTAL / TENDENCIA =====
ws=wb.create_sheet('2. Vertical-Horizontal')
title(ws,'Análisis Vertical, Horizontal y Tendencia 2023-2027 (MM CLP)','Vertical = % sobre venta. Horizontal = variación año contra año.')
hdr(ws,4,['ANÁLISIS VERTICAL (% venta)']+[str(y) for y in YRS],[32,11,11,11,11,11]); r=5
for lab in ['Margen de Contribución','GAV','EBIT','EBITDA','Utilidad del Ejercicio']:
    ws.cell(r,1,lab).border=THIN
    for j in range(5): n(ws.cell(r,2+j),PL[lab][j]/PL['Ingresos'][j],pct=True)
    r+=1
r+=1
hdr(ws,r,['ANÁLISIS HORIZONTAL (Δ% año/año)','24v23','25v24','26v25','27v26','']); r+=1
for lab in ['Ingresos','Margen de Contribución','GAV','EBIT','Utilidad del Ejercicio']:
    ws.cell(r,1,lab).border=THIN
    for j in range(1,5):
        prev=PL[lab][j-1]; cur=PL[lab][j]
        g=(cur/prev-1) if prev not in(0,None) and (prev>0)==(cur>0) else None
        n(ws.cell(r,1+j),g,pct=True) if g is not None else n(ws.cell(r,1+j),None)
    ws.cell(r,6).border=THIN
    r+=1
r+=1
hdr(ws,r,['TENDENCIA (absoluto)']+[str(y) for y in YRS]); r+=1
for lab in ['Ingresos','Margen de Contribución','EBIT','EBITDA','Utilidad del Ejercicio']:
    ws.cell(r,1,lab).border=THIN
    for j in range(5): n(ws.cell(r,2+j),PL[lab][j],bold=lab=='EBIT')
    r+=1

# ===== HOJA 3: LN VENTA Y CONTRIBUCIÓN =====
ws=wb.create_sheet('3. LN Venta y Contribución')
title(ws,'Línea de Negocio — Venta y Contribución: Meta vs Resultado (MM CLP)','Meta = PPTO original. Resultado/Proyección = FCST. H1 (ene-jun) y H2 (jul-dic). Fuente: planilla FCST/PPTO VENTAS por LN.')
# datos: LN -> (V_meta, V_res, C_meta, C_res) para H1 y H2
H1={'Marketplace':(1929,2044,487,517),'Distribución':(562,350,176,120),'Fidelización':(398,250,132,89),'Páginas Web':(272,239,69,59),'Corporativo':(118,55,47,23)}
H2={'Marketplace':(2587,2649,656,680),'Distribución':(778,363,264,123),'Fidelización':(391,154,128,48),'Páginas Web':(334,339,83,84),'Corporativo':(140,160,47,50)}
orden=['Marketplace','Distribución','Fidelización','Páginas Web','Corporativo']
def bloque(ws,r,titulo,data,res_lbl):
    ws.cell(r,1,titulo).font=FB; r+=1
    hdr(ws,r,['Línea','Venta Meta','Venta '+res_lbl,'%V','Contrib Meta','Contrib '+res_lbl,'%C'],[16,11,11,7,11,11,7]); r+=1
    for ln in orden:
        vm,vr,cm,cr=data[ln]
        ws.cell(r,1,ln).border=THIN
        n(ws.cell(r,2),vm); n(ws.cell(r,3),vr); n(ws.cell(r,4),vr/vm,pct=True); n(ws.cell(r,5),cm); n(ws.cell(r,6),cr); n(ws.cell(r,7),cr/cm,pct=True)
        if cr/cm<0.6: ws.cell(r,7).fill=fill_red
        elif cr/cm>=1: ws.cell(r,7).fill=fill_grn
        r+=1
    tvm=sum(d[0] for d in data.values());tvr=sum(d[1] for d in data.values());tcm=sum(d[2] for d in data.values());tcr=sum(d[3] for d in data.values())
    ws.cell(r,1,'TOTAL').font=FB
    n(ws.cell(r,2),tvm,bold=True);n(ws.cell(r,3),tvr,bold=True);n(ws.cell(r,4),tvr/tvm,pct=True,bold=True);n(ws.cell(r,5),tcm,bold=True);n(ws.cell(r,6),tcr,bold=True);n(ws.cell(r,7),tcr/tcm,pct=True,bold=True)
    for col in range(1,8): ws.cell(r,col).fill=fill_sub
    return r+2
r=bloque(ws,4,'1° SEMESTRE — Meta (PPTO) vs Resultado REAL',H1,'Real')
r=bloque(ws,r,'2° SEMESTRE — Meta vs Proyección (FCST)',H2,'FCST')
# Doble ajuste FY
ws.cell(r,1,'DOBLE AJUSTE A LA BAJA — FY: PPTO original → Ajuste 1 → Ajuste 2 (FCST)').font=FB; r+=1
hdr(ws,r,['Línea','V PPTO','V Aj.1','V FCST','C PPTO','C Aj.1','C FCST'],None); r+=1
da={'Marketplace':(4517,4732,4679,1144,1144,1182),'Distribución':(1340,819,814,439,439,279),'Fidelización':(789,485,473,260,260,169),'Páginas Web':(607,617,586,152,152,145),'Corporativo':(258,217,218,94,94,90)}
for ln in orden:
    x=da[ln]; ws.cell(r,1,ln).border=THIN
    for j,v in enumerate(x): n(ws.cell(r,2+j),v)
    r+=1
ws.cell(r,1,'TOTAL').font=FB
for j,v in enumerate([7510,6872,6770,2089,2089,1864]): n(ws.cell(r,2+j),v,bold=True)
for col in range(1,8): ws.cell(r,col).fill=fill_sub
r+=2
ws.cell(r,1,'H1 = REAL (ventas_historico ene-may + jun); contribución real = venta × margen LN (total $808M ≈ MC del P&L $816M). H2 = FCST. El B2B real está corto: Distribución 62% de venta, Corporativo 47%. El doble ajuste: Aj.1 recortó venta B2B, Aj.2 recortó el margen.').font=FI
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); ws.row_dimensions[r].height=40

# ===== HOJA 4: GASTO TREND =====
ws=wb.create_sheet('4. Gasto Tendencia')
title(ws,'Gasto (GAV) por cuenta — Tendencia 2023-2026 (MM CLP)','Por cuenta contable (pyl). El desglose por centro de costo solo existe 2025-2026 (ver hoja aparte si se requiere).')
hdr(ws,4,['Cuenta']+[str(y) for y in [2023,2024,2025,2026]]+['Δ 26v23'],[36,10,10,10,10,10]); r=5
gav=[('Sueldos, honorarios y leyes',-705,-1065,-1151,-1175),('Oficina y arriendos',-86,-116,-183,-201),
 ('Marketing Branding',0,-21,-37,-46),('Suscripciones y publicaciones',-19,-28,-46,-28),
 ('Movilización/transporte/colación',-13,-31,-48,-22),('Cuenta I+D Productos',0,0,0,-19),
 ('Depreciación',-28,0,-16,-18),('Seguros',-12,-13,-15,-17),('Asesorías varias',-12,-13,-26,-4),
 ('Asesoría contable',-20,-18,-10,-1),('Servicios computacionales',-20,-23,-6,0),
 ('Comisión Transbank/Pago Fácil',-6,-11,-17,0),('TOTAL GAV',-928,-1353,-1565,-1594)]
for row in gav:
    ws.cell(r,1,row[0]).border=THIN
    for j,v in enumerate(row[1:]): n(ws.cell(r,2+j),v,bold='TOTAL' in row[0])
    n(ws.cell(r,6),row[4]-row[1],bold='TOTAL' in row[0])
    if 'TOTAL' in row[0]:
        for col in range(1,7): ws.cell(r,col).fill=fill_sub
    r+=1
r+=1
ws.cell(r,1,'CRECIMIENTO DEL GAV 2025 → 2026: +$29M (+1,9%) — principales movimientos por cuenta (MM)').font=FB; r+=1
hdr(ws,r,['Suben','Δ','Bajan','Δ']); r+=1
movs=[('Sueldos/honorarios/leyes',24,'Movilización/transporte',-26),('Cuenta I+D Productos',19,'Asesorías varias',-22),
 ('Oficina y arriendos',18,'Suscripciones',-18),('Marketing Branding',9,'Comisión Transbank',-17)]
for su,sv,ba,bv in movs:
    ws.cell(r,1,su).border=THIN; n(ws.cell(r,2),sv); ws.cell(r,3,ba).border=THIN; n(ws.cell(r,4),bv); r+=1
r+=1
ws.cell(r,1,'El GAV 2025→2026 queda casi plano (+1,9%): las alzas de nómina (+$24M) y las nuevas cuentas (I+D +$19M) se compensan con recortes en movilización, asesorías, suscripciones y medios de pago. La expansión fuerte del gasto fue 2023-2024 (nómina +51%), ya en régimen.').font=Font(size=10,color='808080',italic=True)
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); ws.row_dimensions[r].height=42
r+=2
ws.cell(r,1,'REMUNERACIONES 2025→2026: por qué queda casi plano (+$24M) — la compensación').font=FB; r+=1
hdr(ws,r,['Concepto','MM','',''],None); r+=1
for nom,v in [('Sueldos 2025',-1151),('(+) Reajuste base ~+7% (efecto 2026)',-56),('(−) Salidas 5 IA + Tamara (efecto parcial año)',32),('(+/−) Neto reclasificaciones/otros',0),('Sueldos 2026',-1175)]:
    ws.cell(r,1,nom).border=THIN; n(ws.cell(r,2),v,bold=('Sueldos' in nom and '20' in nom))
    if 'Sueldos' in nom and '20' in nom:
        for col in range(1,5): ws.cell(r,col).fill=fill_sub
    r+=1
r+=1
ws.cell(r,1,'El alza de base (reajuste +7%, ~+$56M) se compensa casi entero con el ahorro de las salidas (~−$32M) → nómina 2026 casi plana. En 2027 cambia: las salidas rinden full-year pero el reajuste 5% pesa más → GAV sube +$14M (ver hoja 8).').font=Font(size=10,italic=True,color='808080')
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); ws.row_dimensions[r].height=40

# ===== HOJA 5: SUELDOS =====
ws=wb.create_sheet('5. Estructura Sueldos')
title(ws,'Estructura de Remuneraciones por Estrato — sueldo base contractual Jun-2026 (MM CLP)','Fuente: planilla Buk (jun-2026), sueldo base por persona → agrupado a estrato (sin nombres). Anualizado ×12.')
hdr(ws,4,['Estrato / Cargo','N pers.','Base $/mes','Base $/año','% base'],[26,9,12,12,10]); r=5
estr=[('Gerencia',5,31.18,374,0.409),('Jefatura',6,12.69,152,0.167),('Operativo/Admin',16,10.39,125,0.137),
 ('Diseño/Mkt',5,6.06,73,0.080),('KAM',3,5.84,70,0.077),('Subgerencia',1,4.31,52,0.057),
 ('Planner',1,3.35,40,0.044),('Analista',3,2.30,28,0.031)]
for nom,npe,mes,anual,p in estr:
    ws.cell(r,1,nom).border=THIN; n(ws.cell(r,2),npe); n(ws.cell(r,3),mes,dec=1); n(ws.cell(r,4),anual); n(ws.cell(r,5),p,pct=True)
    if nom in('Gerencia','Subgerencia'):
        for col in range(1,6): ws.cell(r,col).fill=fill_yel
    r+=1
ws.cell(r,1,'TOTAL base',).font=FB; n(ws.cell(r,2),40,bold=True); n(ws.cell(r,3),76.1,dec=1,bold=True); n(ws.cell(r,4),914,bold=True); n(ws.cell(r,5),1.0,pct=True,bold=True)
for col in range(1,6): ws.cell(r,col).fill=fill_sub
r+=2
ws.cell(r,1,'Gerencia + Subgerencia = 6 personas = 46,6% del sueldo base. La capa alta concentra casi la mitad de la nómina.').font=Font(size=10,bold=True,color='C00000'); r+=2
ws.cell(r,1,'EVOLUCIÓN SUELDO BASE Dic-2025 → Jun-2026 (persona a persona, cruce por RUT)').font=FB; r+=1
for t in ['· Base total +7,2% (33 de 34 subieron; ninguno bajó). Piso de reajuste ~2,8% para todos.',
 '· Alzas sobre el piso: Subgerencia +15%, Diseño +19-20%, KAM +15%, Jefatura Mkt/Planner +13%, varios operarios +11-17%.',
 '· Gerencia: Andrés +6,5%, Sebastián +5,6%, Nicolás promovido de Comercial a Gerencia Distribución (nueva posición gerencial en 2026).',
 '· Lectura CFO: en año de pérdida hubo alza de base generalizada de +7% con saltos de 15-20% — es lo que sostiene el costo de nómina, no las salidas.']:
    c=ws.cell(r,1,t); c.font=Font(size=10); c.alignment=Alignment(wrap_text=True,vertical='top'); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); ws.row_dimensions[r].height=26; r+=1
r+=1
ws.cell(r,1,'Nota: tabla en sueldo BASE. El costo total (base + gratificación + asignaciones + leyes) usado en el GAV ≈ $1.130M/año.').font=FI

# ===== HOJA 6: CAPITAL DE TRABAJO =====
ws=wb.create_sheet('6. Capital de Trabajo')
title(ws,'Capital de Trabajo — con fórmulas y rangos esperados (MM CLP)')
hdr(ws,4,['Existencias por tipo','2025','2026','Δ'],[36,11,11,10]); r=5
for nom,a,b in [('Bodega (IW)',1177,1428),('Tránsito (IT)',491,375),('Materia prima',73,73),('TOTAL',1741,1876)]:
    ws.cell(r,1,nom).border=THIN; n(ws.cell(r,2),a); n(ws.cell(r,3),b); n(ws.cell(r,4),b-a,bold=nom=='TOTAL')
    if nom=='TOTAL':
        for col in range(1,5): ws.cell(r,col).fill=fill_sub
    r+=1
r+=1
hdr(ws,r,['Indicador','2025','2026','Fórmula','Rango esperado']); r+=1
kt=[('Días de inventario','—','—','Existencias / (Costo/365)',''),
 ('  Meses de inventario bodega',4.4,5.5,'Bodega IW / (Costo/12)','~3 meses'),
 ('Días de cobro (DSO)',45,44,'CxC / Venta × 365','30-45 días'),
 ('Días de pago (DPO)',28,57,'CxP / Costo × 365','45-60 días'),
 ('Ciclo de conversión de caja',217,208,'Días inv + DSO − DPO','< 120 días ideal'),
 ('Razón corriente',2.44,2.10,'Act. Corriente / Pas. Corriente','> 1,5'),
 ('Cobertura de intereses',1.18,1.20,'EBIT / Intereses','> 2,0x'),
 ('Deuda / EBITDA',8.72,9.13,'Deuda financiera / EBITDA','< 3-4x')]
for nom,a,b,f,rango in kt:
    ws.cell(r,1,nom).border=THIN
    isx='Cobertura' in nom or 'EBITDA' in nom or 'corriente' in nom
    dec=2 if isx else (1 if 'Meses' in nom else 0)
    n(ws.cell(r,2),a if a!='—' else None,dec=dec); n(ws.cell(r,3),b if b!='—' else None,dec=dec)
    ws.cell(r,4,f).font=FI; ws.cell(r,4).border=THIN; ws.cell(r,5,rango).alignment=CEN; ws.cell(r,5).border=THIN
    if 'Deuda / EBITDA' in nom: ws.cell(r,3).font=RED
    r+=1
ws.column_dimensions['D'].width=30; ws.column_dimensions['E'].width=16

# ===== HOJA 7: RATIOS 2023-2027 =====
ws=wb.create_sheet('7. Ratios 2023-2027')
title(ws,'Ratios de Rentabilidad y Riesgo Financiero — 2023 a 2027 (MM CLP)')
hdr(ws,4,['Indicador']+[str(y) for y in YRS]+['Ref.'],[32,10,10,10,10,10,12]); r=5
mc_pct=[PL['Margen de Contribución'][i]/PL['Ingresos'][i] for i in range(5)]
ebit_pct=[PL['EBIT'][i]/PL['Ingresos'][i] for i in range(5)]
ebitda_pct=[PL['EBITDA'][i]/PL['Ingresos'][i] for i in range(5)]
un_pct=[PL['Utilidad del Ejercicio'][i]/PL['Ingresos'][i] for i in range(5)]
rows=[('Margen contribución %',mc_pct,'%',''),('EBIT %',ebit_pct,'%',''),('EBITDA %',ebitda_pct,'%',''),
 ('Utilidad neta %',un_pct,'%','>0'),('__',None,None,''),
 ('Cobertura de intereses (x)',cob,'x','>2x'),('Deuda / EBITDA (x)',de,'x','<3-4x'),
 ('Deuda financiera',DEUDA,'n',''),('Intereses',[-x for x in INT],'n','')]
for nom,vals,fmt,ref in rows:
    if nom=='__': r+=1; continue
    ws.cell(r,1,nom).border=THIN
    for j in range(5): n(ws.cell(r,2+j),vals[j],dec=2 if fmt=='x' else 0,pct=(fmt=='%'),x=(fmt=='x'))
    ws.cell(r,7,ref).alignment=CEN; ws.cell(r,7).border=THIN
    if 'Deuda / EBITDA' in nom: ws.cell(r,5).font=RED; ws.cell(r,6).font=GRN
    if 'Cobertura' in nom: ws.cell(r,5).font=RED; ws.cell(r,6).font=GRN
    r+=1

# ===== HOJA 8: 2027 (A DESARROLLAR) =====
ws=wb.create_sheet('8. 2027 y Palancas')
title(ws,'GAV 2027 bottom-up + P&L 2027 + Palancas (MM CLP) — BASE, escenarios a desarrollar')
hdr(ws,4,['GAV 2027 bottom-up','Base','Reajuste','2027'],[34,11,11,11]); r=5
for nom,b,rj,f in [('Remuneraciones (ex-bonos)',1131,'+5%',1188),('Arriendos',176,'+4%',184),('Suscripciones (Odoo/Yuju 50%)',28,'',28),('Marketing branding',36,'',36),('Productos',20,'',20),('Otros CC (oficina/honor/deprec/seg/ases)',89,'',89),('Lumpy (repr/benef/movil/mant/imp)',64,'',64),('GAV 2027 TOTAL',1544,'+64',1608)]:
    ws.cell(r,1,nom).border=THIN; n(ws.cell(r,2),b); ws.cell(r,3,rj).alignment=CEN; ws.cell(r,3).border=THIN; n(ws.cell(r,4),f,bold='TOTAL' in nom)
    if 'TOTAL' in nom:
        for col in range(1,5): ws.cell(r,col).fill=fill_sub
    r+=1
r+=1
hdr(ws,r,['P&L 2027','Valor','','']); r+=1
for nom,v in [('Venta (Operaciones $7.200M)',7200),('Margen Contribución (27%)',1950),('GAV',-1608),('EBIT',342),('EBITDA',360),('Intereses (COMEX $119 + comercial $25)',-144),('Gastos bancarios',-25),('RNO',-169),('Utilidad',173)]:
    ws.cell(r,1,nom).border=THIN; n(ws.cell(r,2),v,bold=nom in('EBIT','EBITDA','Utilidad'))
    if nom=='Utilidad': ws.cell(r,2).font=GRN
    r+=1
r+=1
ws.cell(r,1,'PALANCAS').font=FB; r+=1
for t in ['1. Desapalancamiento REAL: amortización comercial $403M→$194M (paga $210M en 2027). COMEX plano $1.533M (renovable). Deuda $1.939M→$1.729M.','2. El otro motor es el EBITDA creciente ($212M→$360M) → Deuda/EBITDA 9,1x→4,8x, cobertura 1,2x→2,4x. Depende de que la venta $7.200M se cumpla.','3. Destock de bodega NO factible (descartado) — el saneamiento NO depende de liberar inventario.','4. B2B: corregir margen directo (Corporativo bajo costo), no el volumen.','5. Escenarios de sensibilidad: ver pestaña 9. Pendiente: forward/cobertura cambiaria (sesión aparte).']:
    c=ws.cell(r,1,t); c.font=Font(size=10); c.alignment=Alignment(wrap_text=True,vertical='top'); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4); ws.row_dimensions[r].height=26; r+=1
r+=1
ws.cell(r,1,'RECOMENDACIONES CFO').font=Font(bold=True,color='C00000'); r+=1
for t in ['2026 (inmediato): llevar Productos ($20M) y Marketing branding ($36M) AL MÍNIMO · NO aumentar deuda (ya en $1,9B) · revisar estructura: hay doble gerencia y subgerencia.',
 '2027 (estructural, con la venta $7.200M): rediseñar la cúpula — estructura SIN Co-Founder (Erich) + SIN 1 gerente → recorte de ~$10M/mes = −$120M/año en gerencias.',
 'Impacto del recorte gerencial 2027: EBIT $342M→$462M · Utilidad $173M→$293M · Cobertura 2,4x→3,2x · Deuda/EBITDA 4,8x→3,6x (ver escenario en pestaña 9).']:
    c=ws.cell(r,1,t); c.font=Font(size=10); c.alignment=Alignment(wrap_text=True,vertical='top'); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4); ws.row_dimensions[r].height=28; r+=1

# ===== HOJA 9: ESCENARIOS DE SENSIBILIDAD =====
ws=wb.create_sheet('9. Escenarios Sensibilidad')
title(ws,'Escenarios de Sensibilidad 2027 (MM CLP)','Base: venta $7.200M · MC 27,1% · GAV fijo $1.608M · intereses $144M. Shock a la venta → MC flexa al margen; GAV e intereses fijos.')
VB=7200.0; MCP=1950/VB; GAV27=1608; INT27=144; BANC=25; DA27=18; DEUDA27=1729
def esc(s):
    v=VB*(1+s); mc=v*MCP; ebit=mc-GAV27; ebitda=ebit+DA27; util=ebit-INT27-BANC
    return v,mc,ebit,ebitda,util,(ebit/INT27),(DEUDA27/ebitda if ebitda>0 else None)
hdr(ws,4,['Escenario venta','Venta','MC','EBIT','EBITDA','Utilidad','Cobertura','Deuda/EBITDA'],[16,10,10,9,9,10,10,12]); r=5
for s,lab in [(0.10,'+10%'),(0.05,'+5%'),(0.0,'Base'),(-0.05,'−5%'),(-0.089,'Break-even'),(-0.10,'−10%'),(-0.15,'−15%'),(-0.20,'−20%')]:
    v,mc,ebit,ebitda,util,cobx,dex=esc(s)
    ws.cell(r,1,lab).border=THIN; n(ws.cell(r,2),v); n(ws.cell(r,3),mc); n(ws.cell(r,4),ebit,bold=True); n(ws.cell(r,5),ebitda); n(ws.cell(r,6),util,bold=True)
    n(ws.cell(r,7),cobx,dec=2,x=True);
    if dex: n(ws.cell(r,8),dex,dec=1,x=True)
    else: ws.cell(r,8,'—').alignment=CEN; ws.cell(r,8).border=THIN
    if s==0:
        for col in range(1,9): ws.cell(r,col).fill=fill_yel
    if util<0: ws.cell(r,6).font=RED
    elif util>0 and s!=0: ws.cell(r,6).font=GRN
    r+=1
r+=1
ws.cell(r,1,'PUNTO DE QUIEBRE: la utilidad 2027 llega a $0 con una caída de venta de −8,9% (~−$640M). Bajo eso, pérdida. Con −10% la cobertura cae a 1,0x (no alcanza a pagar intereses).').font=Font(size=10,bold=True,color='C00000')
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8); ws.row_dimensions[r].height=40; r+=2
# tornado por LN
ws.cell(r,1,'SENSIBILIDAD POR LÍNEA — impacto en EBIT de una caída de −10% en cada LN (tornado)').font=FB; r+=1
hdr(ws,r,['Línea','Venta','MC%','ΔEBIT si −10%','',''],[16,10,8,14]); r+=1
lns=[('Marketplace',4679,0.253),('Distribución',814,0.343),('Fidelización',473,0.357),('Páginas Web',586,0.247),('Corporativo',218,0.413)]
for nom,v,mcp in sorted(lns,key=lambda x:-x[1]*0.10*x[2]):
    imp=-v*0.10*mcp
    ws.cell(r,1,nom).border=THIN; n(ws.cell(r,2),v); n(ws.cell(r,3),mcp,pct=True); n(ws.cell(r,4),imp,bold=True)
    if nom=='Marketplace': ws.cell(r,4).fill=fill_red
    r+=1
ws.cell(r,1,'Marketplace es la LN crítica: un −10% ahí resta ~$118M de EBIT (más que todas las demás juntas). Blindar Marketplace es la prioridad de defensa.').font=Font(size=10,italic=True)
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); ws.row_dimensions[r].height=28; r+=2
ws.cell(r,1,'ESCENARIO PALANCA: recorte gerencial 2027 (sin Co-Founder + sin 1 gerente = −$120M/año)').font=Font(bold=True,color='1E7A1E'); r+=1
hdr(ws,r,['Métrica 2027','Base','Con recorte −$120M','']); r+=1
for nom,b,c in [('GAV',-1608,-1488),('EBIT',342,462),('EBITDA',360,480),('Utilidad',173,293)]:
    ws.cell(r,1,nom).border=THIN; n(ws.cell(r,2),b); n(ws.cell(r,3),c,bold=True); ws.cell(r,4).border=THIN
    if nom=='Utilidad': ws.cell(r,3).font=GRN
    r+=1
for nom,b,c in [('Cobertura (x)',2.375,3.21),('Deuda/EBITDA (x)',4.80,3.60)]:
    ws.cell(r,1,nom).border=THIN; n(ws.cell(r,2),b,dec=2,x=True); n(ws.cell(r,3),c,dec=2,x=True,bold=True); ws.cell(r,4).border=THIN; r+=1
ws.cell(r,1,'El recorte gerencial es la palanca más potente de 2027: +$120M directo al EBIT lleva la utilidad a $293M y el Deuda/EBITDA a 3,6x. Es la recomendación estructural central.').font=Font(size=10,bold=True,color='1E7A1E')
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); ws.row_dimensions[r].height=28

for s in wb.worksheets:
    s.sheet_view.showGridLines=False; s.row_dimensions[1].height=20
OUT=BASE+r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v5.xlsx"
try: wb.save(OUT)
except PermissionError:
    OUT=BASE+r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v5b.xlsx"; wb.save(OUT)
print('Guardado:',OUT); print('Hojas:',[s.title for s in wb.worksheets])
