# -*- coding: utf-8 -*-
"""Patch v15 -> v16:
1. Erich NO sale -> gerencial = SOLO Nicole ($55M). Full-stack +$365M.
2. Sumatoria: columna Aporte por palanca (ya existe col Ahorro; se refuerza).
3+4. Puente patrimonio + indemnizaciones 2026 (explica D/Pat).
5. Hoja 7: 2027 = base; nota palancas. + columna 2027e Palancas en resumen.
"""
import sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')
BASE = r"G:\Mi unidad\TRABAJO\RESPALDO\OPERACIONES\UNION X - IA"
wb = openpyxl.load_workbook(BASE + r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v15.xlsx")
ws = wb['8. 2027 y Palancas']
for mr in list(ws.merged_cells.ranges):
    if mr.min_row >= 68: ws.unmerge_cells(str(mr))
FB=Font(bold=True); GRN=Font(color='1E7A1E',bold=True); IT=Font(size=10,italic=True,color='C00000'); WH=Font(bold=True,color='FFFFFF')
fgrn=PatternFill('solid',fgColor='E3F4E4'); fsub=PatternFill('solid',fgColor='D6E4F0'); fhdr=PatternFill('solid',fgColor='0F2340'); blank=PatternFill()
TH=Border(*[Side(style='thin',color='D9D9D9')]*4); RIG=Alignment(horizontal='right'); CEN=Alignment(horizontal='center'); LFT=Alignment(horizontal='left',wrap_text=True)
INT=144; RNO=-169; DEU=1729; DA=18; PAT26=625

# ===== SUMATORIA rows 70-76 (Nicole 55 en vez de 120) =====
steps=[('Base digital $6.500M',None,20,6500),
       ('+ B2B UnionX $300M (contrib. 35%)',105,125,6800),
       ('+ Reducción gerencial · SOLO Nicole (nov-26)',55,180,6800),
       ('+ Marketing · Michela [ejecutada]',46,226,6800),
       ('+ Ecommerce · Ignacia [ejecutada]',33,259,6800),
       ('+ Tercerizar EIT (tarifas negociadas)',80,339,6800),
       ('+ Ruteros Trade Mkt (dic-26)',26,365,6800)]
for i,(nom,ah,util,vta) in enumerate(steps):
    r=70+i; ebit=util-RNO; ebitda=ebit+DA
    vals=[nom,ah,ebit,ebitda,util,ebit/vta,util/vta,round(ebit/INT,2),round(DEU/ebitda,2),round(DEU/(PAT26+util),2)]
    for j,v in enumerate(vals):
        c=ws.cell(r,1+j); c.border=TH; c.font=Font()
        if j==0: c.value=v; c.alignment=LFT; c.font=FB if i in(0,6) else Font()
        elif v is None: c.value='—'; c.alignment=CEN
        elif j in(5,6): c.value=v; c.number_format='0.0%'; c.alignment=RIG
        elif j in(7,8,9): c.value=v; c.number_format='0.00"x"'; c.alignment=RIG
        else: c.value=v; c.number_format='#,##0;[Red]-#,##0'; c.alignment=RIG
    fill = fsub if i in(0,1) else (fgrn if i==6 else blank)
    for cc in range(1,11): ws.cell(r,cc).fill=fill
    if i in(0,6): ws.cell(r,5).font=GRN
ws.cell(69,2,'Aporte')  # renombrar col Ahorro -> Aporte (monto que inyecta cada palanca)
ws.cell(77,1,'Erich NO sale: la liberación gerencial es SOLO Nicole ($55M). 5 palancas suman +$240M -> full-stack +$365M, cob 3,7x, D/EBITDA 3,1x, D/Patrimonio 1,7x. La columna Aporte muestra lo que inyecta cada palanca.').font=IT

# ===== VISTA POR PALANCA (Nicole sola) =====
ws.cell(80,1,'1 · Reducción gerencial · SOLO Nicole (nov-26)')
ws.cell(80,2,'1 (Erich NO sale)'); ws.cell(80,3,'+55'); ws.cell(80,4,'−21,3'); ws.cell(80,5,'−13')
ws.cell(80,6,'Nicole Contreras, Subgcia Comercial, 5,2 años. Erich se mantiene -> la palanca gerencial es solo Nicole ($55M costo cargado).')
for j in range(6): ws.cell(80,1+j).alignment=(RIG if j in(2,3,4) else LFT); ws.cell(80,1+j).border=TH
ws.cell(85,3,'+240'); ws.cell(85,3).alignment=RIG
ws.cell(85,6,'Indemn. 2026 $54,9M + EIT $31,7M (2027). 2026 pasa de +$19M a ~+$7M. Full-stack +$365M.')

# ===== PUENTE PATRIMONIO + INDEMNIZACIONES 2026 (nuevo, rows 96-108) =====
ws.cell(96,1,'PUENTE PATRIMONIO E INDEMNIZACIONES 2026 (efecto en Deuda/Patrimonio)').font=FB
hb=['Concepto','Monto']
for j,h in enumerate(hb):
    c=ws.cell(97,1+j); c.value=h; c.font=WH; c.fill=fhdr; c.alignment=(LFT if j==0 else CEN); c.border=TH
bridge=[('Patrimonio 2026-cierre (resultado base +$19M)',637,False),
        ('(−) Indemnizaciones exits 2026 (Michela+Ignacia+Nicole+Ruteros)',-54.9,False),
        ('(+) Ahorro de sueldo en 2026 (año parcial)',43.0,False),
        ('(=) Efecto neto exits en 2026',-11.9,True),
        ('Utilidad 2026 ajustada: +$19M → +$7M','',False),
        ('Patrimonio 2026-cierre AJUSTADO',625,True),
        ('(+) Utilidad 2027 (full-stack con palancas)',365,False),
        ('(=) Patrimonio 2027-cierre',990,True),
        ('Deuda financiera 2027-cierre (tras amortizar $210M)',1729,False),
        ('Deuda / Patrimonio 2027 (full-stack)',round(DEU/990,2),True)]
for i,(n,v,bold) in enumerate(bridge):
    r=98+i
    ws.cell(r,1,n).alignment=LFT; ws.cell(r,1).border=TH
    c=ws.cell(r,2); c.border=TH; c.alignment=RIG
    if v!='':
        c.value=v
        c.number_format=('0.00"x"' if 'Patrimonio 2027 (full' in n else '#,##0;[Red]-#,##0')
    if bold:
        ws.cell(r,1).font=FB; ws.cell(r,2).font=FB
        for cc in (1,2): ws.cell(r,cc).fill=fsub
ws.cell(109,1,'PUNTO 3: la D/Pat baja de 3,0x (2026) a ~2,3x en la base y a 1,7x con palancas por dos efectos reales: la deuda amortiza $210M (1.936→1.729) y el patrimonio crece con el resultado. Las indemnizaciones ($54,9M) netean solo −$12M en 2026 porque el ahorro de sueldo del año las compensa, por eso no golpean el patrimonio.').font=IT

# ===== Escenario gerencial block (48-56) -> Nicole =====
ws.cell(48,1,'PALANCA GERENCIAL: salida de Nicole (Subgerencia Comercial) — Erich NO sale').font=FB
ws.cell(49,3,'Con salida Nicole (−$55M)')
EBIT_B=294
ger=[('GAV',-1608,-(1608-55)),('EBIT',EBIT_B,EBIT_B+55),('EBITDA',EBIT_B+DA,EBIT_B+55+DA),
     ('Utilidad',125,180),('Cobertura (x)',round(EBIT_B/INT,2),round((EBIT_B+55)/INT,2)),
     ('Deuda/EBITDA (x)',round(DEU/(EBIT_B+DA),2),round(DEU/(EBIT_B+55+DA),2))]
for i,(n,a,b) in enumerate(ger):
    r=50+i; ws.cell(r,1,n); ws.cell(r,2,a); ws.cell(r,3,b)
    for cc in (2,3):
        c=ws.cell(r,cc); c.alignment=RIG
        c.number_format=('0.00"x"' if 'x' in n else '#,##0;[Red]-#,##0')
ws.cell(56,1,'Base = digital $6.500M + B2B $300M (+$125M). Erich se mantiene; la única liberación gerencial confirmada es Nicole (nov-26, costo cargado ~$55M, indem $21,3M).').font=IT

# ===== narrativa =====
ws.cell(28,1,'3. Las 5 palancas (+$240M; gerencial = solo Nicole $55M, EIT negociado $80M) llevan el 2027 a +$365M, cobertura 3,7x, Deuda/EBITDA 3,1x, D/Patrimonio 1,7x.')
ws.cell(30,1,'5. Erich NO sale: la reducción gerencial es solo la salida de Nicole (Subgerencia Comercial, nov-26).')

# ===== HOJA 7 Ratios: 2027 col6 = base+B2B, + nota palancas =====
r7=wb['7. Ratios 2023-2027']
r7.cell(4,6,'2027e base')
upd={5:0.280,6:294/6800,7:312/6800,8:125/6800,10:2.04,11:5.54,12:1729,13:-144,14:round(DEU/750,2)}
for row,val in upd.items():
    c=r7.cell(row,6); c.value=val
r7.cell(16,1,'Con las 5 palancas (full-stack +$365M): EBIT% 7,9% · Cobertura 3,7x · Deuda/EBITDA 3,1x · Deuda/Patrimonio 1,7x. Ver hoja 8.').font=IT

# ===== HOJA 1 =====
w1=wb['1. Resumen Ejecutivo']
for r in range(1,80):
    v=str(w1.cell(r,1).value or '')
    if v.startswith('2027'):
        w1.cell(r,1,'2027: base digital $6.500M (+$20M) + B2B $300M (+$105M) = +$125M. Las 5 palancas (+$240M; gerencial = solo Nicole $55M — Erich NO sale — Michela+Ignacia ejecutadas, EIT negociado $80M, ruteros) llevan el 2027 a +$365M, cobertura 3,7x, Deuda/EBITDA 3,1x, D/Patrimonio 1,7x. 2026: +$19M → ~+$7M por indemnizaciones ($54,9M, una vez).')
        break

OUT=BASE+r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v16.xlsx"
wb.save(OUT)
print('Guardado v16 | full-stack +365 | D/Pat base 2,31x full-stack 1,75x')
