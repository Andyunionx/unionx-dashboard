# -*- coding: utf-8 -*-
"""Patch v16 -> v17: sumar indemnización CAMBIO DE BODEGA (tercerización EIT, ~$31,7M) al 2026.
2026 +$19M -> -$25M. Patrimonio 2026 $593M. D/Pat full-stack 1,8x."""
import sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')
BASE = r"G:\Mi unidad\TRABAJO\RESPALDO\OPERACIONES\UNION X - IA"
wb = openpyxl.load_workbook(BASE + r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v16.xlsx")
ws = wb['8. 2027 y Palancas']
for mr in list(ws.merged_cells.ranges):
    if mr.min_row >= 96: ws.unmerge_cells(str(mr))
FB=Font(bold=True); IT=Font(size=10,italic=True,color='C00000'); WH=Font(bold=True,color='FFFFFF')
fsub=PatternFill('solid',fgColor='D6E4F0'); fhdr=PatternFill('solid',fgColor='0F2340'); blank=PatternFill()
TH=Border(*[Side(style='thin',color='D9D9D9')]*4); RIG=Alignment(horizontal='right'); CEN=Alignment(horizontal='center'); LFT=Alignment(horizontal='left',wrap_text=True)
DEU=1729; PAT=593

# ===== Sumatoria: recomputar col J (D/Pat) con PAT=593 =====
utils={70:20,71:125,72:180,73:226,74:259,75:339,76:365}
for r,u in utils.items():
    c=ws.cell(r,10); c.value=round(DEU/(PAT+u),2); c.number_format='0.00"x"'; c.alignment=RIG
ws.cell(77,1,'Erich NO sale: gerencial = solo Nicole ($55M). 5 palancas +$240M -> full-stack +$365M, cob 3,7x, D/EBITDA 3,1x, D/Patrimonio 1,8x. El 2026 absorbe $86,6M de indemnizaciones (incl. cambio de bodega) -> baja a -$25M; se paga una vez.').font=IT

# ===== Vista por palanca: EIT efecto 2026 = -31,7 (cambio de bodega); total =====
ws.cell(83,1,'4 · Tercerización EIT · CAMBIO DE BODEGA (tarifas negociadas)')
ws.cell(83,5,'−31,7')
ws.cell(83,6,'Finiquito de operarios de bodega al tercerizar (cambio de bodega). Indemnización $31,7M en 2026; ahorro $80M pleno 2027.')
ws.cell(85,5,'−44 (2026 → −$25M)')

# ===== PUENTE PATRIMONIO (reescribir 96-110) con cambio de bodega =====
for r in range(96,112):
    for c in range(1,3): ws.cell(r,c).value=None; ws.cell(r,c).fill=blank; ws.cell(r,c).border=Border()
ws.cell(96,1,'PUENTE PATRIMONIO E INDEMNIZACIONES 2026 (efecto en Deuda/Patrimonio)').font=FB
for j,h in enumerate(['Concepto','Monto']):
    c=ws.cell(97,1+j); c.value=h; c.font=WH; c.fill=fhdr; c.alignment=(LFT if j==0 else CEN); c.border=TH
bridge=[('Patrimonio 2026-cierre (resultado base +$19M)',637,False,'#,##0'),
        ('(−) Indemnización personas (Michela+Ignacia+Nicole+Ruteros)',-54.9,False,'#,##0.0;[Red]-#,##0.0'),
        ('(−) Indemnización CAMBIO DE BODEGA (tercerización EIT)',-31.7,False,'#,##0.0;[Red]-#,##0.0'),
        ('(+) Ahorro de sueldo en 2026 (año parcial)',43.0,False,'#,##0.0'),
        ('(=) Efecto neto exits en 2026',-43.6,True,'#,##0.0;[Red]-#,##0.0'),
        ('Utilidad 2026 ajustada: +$19M → −$25M','',False,None),
        ('Patrimonio 2026-cierre AJUSTADO',593,True,'#,##0'),
        ('(+) Utilidad 2027 (full-stack con palancas)',365,False,'#,##0'),
        ('(=) Patrimonio 2027-cierre',958,True,'#,##0'),
        ('Deuda financiera 2027-cierre (tras amortizar $210M)',1729,False,'#,##0'),
        ('Deuda / Patrimonio 2027 (full-stack)',round(DEU/958,2),True,'0.00"x"')]
for i,(n,v,bold,fmt) in enumerate(bridge):
    r=98+i
    ws.cell(r,1,n).alignment=LFT; ws.cell(r,1).border=TH
    c=ws.cell(r,2); c.border=TH; c.alignment=RIG
    if v!='': c.value=v; c.number_format=fmt
    if bold:
        ws.cell(r,1).font=FB; ws.cell(r,2).font=FB
        for cc in (1,2): ws.cell(r,cc).fill=fsub
ws.cell(110,1,'El 2026 se lleva TODA la indemnización (personas + cambio de bodega = $86,6M): baja de +$19M a −$25M. Es un golpe one-time; el 2027 queda limpio y el patrimonio se recupera con el resultado. La D/Pat 2027 (full-stack) queda en 1,8x pese a ello.').font=IT

# ===== Hoja 7 Ratios: D/Pat 2027 base =====
r7=wb['7. Ratios 2023-2027']
r7.cell(14,6,round(DEU/(PAT+125),2))
r7.cell(16,1,'Con las 5 palancas (full-stack +$365M): EBIT% 7,9% · Cobertura 3,7x · Deuda/EBITDA 3,1x · Deuda/Patrimonio 1,8x. El 2026 absorbe $86,6M de indemnizaciones (incl. cambio de bodega) -> -$25M one-time.').font=IT

# ===== Hoja 1 =====
w1=wb['1. Resumen Ejecutivo']
for r in range(1,80):
    v=str(w1.cell(r,1).value or '')
    if v.startswith('2027'):
        w1.cell(r,1,'2027: base digital $6.500M (+$20M) + B2B $300M (+$105M) = +$125M. Las 5 palancas (+$240M; gerencial solo Nicole $55M — Erich NO sale; EIT/cambio de bodega negociado $80M; Michela+Ignacia ejecutadas, ruteros) llevan el 2027 a +$365M, cob 3,7x, D/EBITDA 3,1x, D/Patrimonio 1,8x. El 2026 absorbe $86,6M de indemnizaciones (personas + cambio de bodega) -> +$19M pasa a -$25M one-time.')
        break

OUT=BASE+r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v17.xlsx"
wb.save(OUT)
print('Guardado v17 | 2026 -25 | Pat 593 | D/Pat full-stack 1,80x')
