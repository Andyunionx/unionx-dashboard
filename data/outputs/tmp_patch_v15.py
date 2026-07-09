# -*- coding: utf-8 -*-
"""Patch v14 -> v15: EIT tarifas NEGOCIADAS (+$91M 2027) en vez de básico (+$44M)."""
import sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')
BASE = r"G:\Mi unidad\TRABAJO\RESPALDO\OPERACIONES\UNION X - IA"
wb = openpyxl.load_workbook(BASE + r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v14.xlsx")
ws = wb['8. 2027 y Palancas']
FB=Font(bold=True); GRN=Font(color='1E7A1E',bold=True); IT=Font(size=10,italic=True,color='C00000')
fgrn=PatternFill('solid',fgColor='E3F4E4'); fsub=PatternFill('solid',fgColor='D6E4F0'); blank=PatternFill()
TH=Border(*[Side(style='thin',color='D9D9D9')]*4); RIG=Alignment(horizontal='right'); CEN=Alignment(horizontal='center'); LFT=Alignment(horizontal='left',wrap_text=True)
INT=144; RNO=-169; DEU=1729; PATR=637; DA=18; V=6800; UT_B=125

# ==== SUMATORIA rows 70-76 (EIT step +44 -> +91) ====
steps=[('Base digital $6.500M',None,UT_D:=20,6500),
       ('+ B2B UnionX $300M (contrib. 35%)',105,UT_B,6800),
       ('+ Reducción rango gerencial (incl. Nicole)',120,245,6800),
       ('+ Marketing · Michela [ejecutada]',46,291,6800),
       ('+ Ecommerce · Ignacia [ejecutada]',33,324,6800),
       ('+ Tercerizar EIT (tarifas negociadas)',80,404,6800),
       ('+ Ruteros Trade Mkt (dic-26)',26,430,6800)]
for i,(nom,ah,util,vta) in enumerate(steps):
    r=70+i; ebit=util-RNO; ebitda=ebit+DA
    vals=[nom,ah,ebit,ebitda,util,ebit/vta,util/vta,round(ebit/INT,2),round(DEU/ebitda,2),round(DEU/(PATR+util),2)]
    for j,v in enumerate(vals):
        c=ws.cell(r,1+j); c.border=TH; c.font=Font()
        if j==0: c.value=v; c.alignment=LFT; c.font=FB if i in(0,6) else Font()
        elif v is None: c.value='—'; c.alignment=CEN
        elif j in(5,6): c.value=v; c.number_format='0.0%'; c.alignment=RIG
        elif j in(7,8,9): c.value=v; c.number_format='0.00"x"'; c.alignment=RIG
        else: c.value=v; c.number_format='#,##0;[Red]-#,##0'; c.alignment=RIG
    fill = fsub if i in(0,1) else (fgrn if i==6 else blank)
    for cc in range(1,11): ws.cell(r,cc).fill=fill
    if i in(0,6): ws.cell(r,5).font=GRN
ws.cell(77,1,'Base digital +$20M; con B2B $300M +$125M. Las 5 palancas (+$305M, EIT en tarifas negociadas reales +$80M) llevan el full-stack a +$430M, cobertura 4,2x, Deuda/EBITDA 2,8x, D/Patrimonio 1,6x.').font=IT

# ==== VISTA POR PALANCA: EIT row 83 y TOTAL row 85 ====
ws.cell(83,1,'4 · Tercerizar EIT (tarifas negociadas, 2027)'); ws.cell(83,2,'10 (9 operarios + chofer)')
ws.cell(83,3,'+80'); ws.cell(83,4,'−31,7'); ws.cell(83,5,'0')
ws.cell(83,6,'Tarifas reales negociadas 08-jul (ingreso 230→188, prep B2C 659→588, pos 8.780→7.897). Ahorro $79,7M vs operar $537M. Indemn. one-time en 2027')
for j in range(6): ws.cell(83,1+j).alignment=(RIG if j in(2,3,4) else LFT); ws.cell(83,1+j).border=TH
ws.cell(85,3,'+305'); ws.cell(85,3).alignment=RIG
ws.cell(85,6,'Indemn. 2026 $54,9M + EIT $31,7M (2027). 2026 pasa de +$19M a ~+$9M. EIT en tarifas negociadas reales (+$80M); reducir supervisión de bodega suma adicional')

# ==== EIT detail block: marcar negociado como el elegido (row 44-45) ====
ws.cell(44,1,'Recomendación: AVANZAR con tercerización. Plan CFO usa TARIFAS NEGOCIADAS reales 08-jul (+$79,7M/2027, redondeo $80M). Tarifa inicial daba +$43,9M.')
ws.cell(45,1,'Impacto 2027 (tarifas negociadas reales): +$79,7M de ahorro (EIT $337,3M + residual $119,9M vs operar $536,9M). Indemnización one-time $31,7M; ahorro pleno desde el año 1.')

# ==== narrativa row 28 ====
ws.cell(28,1,'3. Las palancas de estructura (+$305M, EIT en tarifas negociadas reales +$80M) llevan el 2027 a +$430M, cobertura 4,2x, Deuda/EBITDA 2,8x.')

# ==== HOJA 1 ====
w1=wb['1. Resumen Ejecutivo']
for r in range(1,80):
    v=str(w1.cell(r,1).value or '')
    if v.startswith('2027'):
        w1.cell(r,1,'2027: base digital $6.500M neto (+$20M) + B2B $300M (+$105M) = +$125M. Las 5 palancas (+$305M; EIT en tarifas negociadas reales +$80M) llevan el 2027 a +$430M, cobertura 4,2x, Deuda/EBITDA 2,8x. 2026 pasa de +$19M a ~+$9M por indemnizaciones (una vez).')
        break

OUT=BASE+r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v15.xlsx"
wb.save(OUT)
print('Guardado:',OUT,'| full-stack +441')
