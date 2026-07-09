# -*- coding: utf-8 -*-
"""Patch v17 -> v18: indemnización CAMBIO DE BODEGA se reconoce en 2027 (inicio 01/01/2027),
no en 2026. 2026 vuelve a +$7M. 2027 one-time -$31,7M (reportada ~$333M). D/Pat full-stack 1,8x."""
import sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')
BASE = r"G:\Mi unidad\TRABAJO\RESPALDO\OPERACIONES\UNION X - IA"
wb = openpyxl.load_workbook(BASE + r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v17.xlsx")
ws = wb['8. 2027 y Palancas']
for mr in list(ws.merged_cells.ranges):
    if mr.min_row >= 96: ws.unmerge_cells(str(mr))
FB=Font(bold=True); IT=Font(size=10,italic=True,color='C00000'); WH=Font(bold=True,color='FFFFFF')
fsub=PatternFill('solid',fgColor='D6E4F0'); fhdr=PatternFill('solid',fgColor='0F2340'); blank=PatternFill()
TH=Border(*[Side(style='thin',color='D9D9D9')]*4); RIG=Alignment(horizontal='right'); CEN=Alignment(horizontal='center'); LFT=Alignment(horizontal='left',wrap_text=True)
DEU=1729; PAT=625

# ===== Sumatoria D/Pat (bodega one-time en 2027) =====
dpat={70:2.68,71:2.31,72:2.15,73:2.03,74:1.96,75:1.85,76:1.80}
for r,v in dpat.items():
    c=ws.cell(r,10); c.value=v; c.number_format='0.00"x"'; c.alignment=RIG
ws.cell(77,1,'Erich NO sale: gerencial = solo Nicole ($55M). 5 palancas +$240M -> full-stack +$365M normalizado, cob 3,7x, D/EBITDA 3,1x, D/Patrimonio 1,8x. El finiquito del cambio de bodega ($31,7M) se reconoce en 2027 (one-time; reportada ~$333M). 2026 queda en +$7M.').font=IT

# ===== Vista por palanca: EIT efecto 2026 -> 0 (bodega es 2027) =====
ws.cell(83,1,'4 · Tercerización EIT · cambio de bodega (01/01/2027)')
ws.cell(83,5,'0')
ws.cell(83,6,'Cambio de bodega inicia 01/01/2027: finiquito $31,7M es one-time de 2027 (no 2026). Ahorro $80M pleno 2027.')
ws.cell(85,5,'−10 (2026 → +$7M)')

# ===== Puente patrimonio (bodega en 2027) =====
for r in range(96,112):
    for c in range(1,3): ws.cell(r,c).value=None; ws.cell(r,c).fill=blank; ws.cell(r,c).border=Border()
ws.cell(96,1,'PUENTE PATRIMONIO E INDEMNIZACIONES (efecto en Deuda/Patrimonio)').font=FB
for j,h in enumerate(['Concepto','Monto']):
    c=ws.cell(97,1+j); c.value=h; c.font=WH; c.fill=fhdr; c.alignment=(LFT if j==0 else CEN); c.border=TH
bridge=[('Patrimonio 2026-cierre (resultado base +$19M)',637,False,'#,##0'),
        ('(−) Indemnización personas 2026 (Michela+Ignacia+Nicole+Ruteros)',-54.9,False,'#,##0.0;[Red]-#,##0.0'),
        ('(+) Ahorro de sueldo en 2026 (año parcial)',43.0,False,'#,##0.0'),
        ('(=) Efecto neto exits 2026',-11.9,True,'#,##0.0;[Red]-#,##0.0'),
        ('Utilidad 2026: +$19M → +$7M (positivo)','',False,None),
        ('Patrimonio 2026-cierre ajustado',625,True,'#,##0'),
        ('(+) Utilidad 2027 normalizada (full-stack)',365,False,'#,##0'),
        ('(−) One-time cambio de bodega (finiquito, 2027)',-31.7,False,'#,##0.0;[Red]-#,##0.0'),
        ('(=) Patrimonio 2027-cierre',958,True,'#,##0'),
        ('Deuda financiera 2027-cierre (amortiza $210M)',1729,False,'#,##0'),
        ('Deuda / Patrimonio 2027 (full-stack)',round(DEU/958,2),True,'0.00"x"')]
for i,(n,v,bold,fmt) in enumerate(bridge):
    r=98+i
    ws.cell(r,1,n).alignment=LFT; ws.cell(r,1).border=TH
    c=ws.cell(r,2); c.border=TH; c.alignment=RIG
    if v!='': c.value=v; c.number_format=fmt
    if bold:
        ws.cell(r,1).font=FB; ws.cell(r,2).font=FB
        for cc in (1,2): ws.cell(r,cc).fill=fsub
ws.cell(110,1,'CRITERIO: el cambio de bodega inicia el 01/01/2027, por lo que su finiquito ($31,7M) se reconoce en 2027 (evento y caja 2027; correlación con el ahorro que también parte en 2027). Así el 2026 queda POSITIVO (+$7M). Condición: comunicar el aviso al equipo en enero 2027, no antes del cierre 2026.').font=IT

# ===== Hoja 7 =====
r7=wb['7. Ratios 2023-2027']
r7.cell(14,6,round(DEU/(PAT+125),2))  # 2027e base D/Pat 2,31
r7.cell(16,1,'Con las 5 palancas (full-stack +$365M normalizado): EBIT% 7,9% · Cobertura 3,7x · Deuda/EBITDA 3,1x · Deuda/Patrimonio 1,8x. El cambio de bodega ($31,7M) es one-time de 2027; el 2026 queda en +$7M.').font=IT

# ===== Hoja 1 =====
w1=wb['1. Resumen Ejecutivo']
for r in range(1,80):
    v=str(w1.cell(r,1).value or '')
    if v.startswith('2027'):
        w1.cell(r,1,'2027: base digital $6.500M (+$20M) + B2B $300M (+$105M) = +$125M. Las 5 palancas (+$240M; gerencial solo Nicole $55M — Erich NO sale; EIT/cambio de bodega negociado, inicia 01/01/2027) llevan el 2027 a +$365M normalizado, cob 3,7x, D/EBITDA 3,1x, D/Patrimonio 1,8x. El finiquito de bodega ($31,7M) es one-time de 2027. 2026 queda POSITIVO en +$7M (tras $54,9M de indemnizaciones de personas).')
        break

OUT=BASE+r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v18.xlsx"
wb.save(OUT)
print('Guardado v18 | 2026 +7 | bodega one-time 2027 | D/Pat full-stack 1,80x')
