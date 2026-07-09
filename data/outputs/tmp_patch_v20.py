# -*- coding: utf-8 -*-
"""Patch v19 -> v20: la VENTA 2027 son $6.500M e INCLUYEN B2B ($300M).
Digital $6.200M + B2B $300M = $6.500M. Ya no se suma B2B aparte.
Base utilidad +$42M; full-stack +$266M."""
import sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')
BASE = r"G:\Mi unidad\TRABAJO\RESPALDO\OPERACIONES\UNION X - IA"
wb = openpyxl.load_workbook(BASE + r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v19.xlsx")
ws = wb['8. 2027 y Palancas']
for mr in list(ws.merged_cells.ranges):
    if mr.min_row >= 14: ws.unmerge_cells(str(mr))
FB=Font(bold=True); GRN=Font(color='1E7A1E',bold=True); RED=Font(color='C00000'); IT=Font(size=10,italic=True,color='C00000'); WH=Font(bold=True,color='FFFFFF')
fgrn=PatternFill('solid',fgColor='E3F4E4'); fsub=PatternFill('solid',fgColor='D6E4F0'); fhdr=PatternFill('solid',fgColor='0F2340'); blank=PatternFill()
TH=Border(*[Side(style='thin',color='D9D9D9')]*4); RIG=Alignment(horizontal='right'); CEN=Alignment(horizontal='center'); LFT=Alignment(horizontal='left',wrap_text=True)
INT=144; RNO=-169; DEU=1729; DA=18; PAT=625; GAV=1608; OT=31.7
# venta
TOT=6500; B2B=300; DIG=TOT-B2B
sh={'Marketplace':80.7,'Fidelización':9.9,'Páginas Web':9.4}; ts=sum(sh.values())
mcp={'Marketplace':0.27,'Fidelización':0.33,'Páginas Web':0.275}
ventas={k:round(DIG*sh[k]/ts) for k in sh}; mcl={k:round(ventas[k]*mcp[k]) for k in sh}
MC_DIG=sum(mcl.values()); MC_B2B=round(B2B*0.35); MC=MC_DIG+MC_B2B
EBIT=MC-GAV; UT=EBIT+RNO
print('MC',MC,'EBIT',EBIT,'UT',UT)

# ===== P&L block (14-23) single column =====
ws.cell(14,1,'P&L 2027 (MM CLP)'); ws.cell(14,2,'Valor')
for c in (1,2): ws.cell(14,c).font=WH; ws.cell(14,c).fill=fhdr; ws.cell(14,c).alignment=(LFT if c==1 else CEN)
pl=[('Venta neta 2027 (incluye B2B UnionX $300M)',TOT,False),
    ('Margen de contribución (28,0%)',MC,False),
    ('GAV',-GAV,False),('EBIT',EBIT,True),('EBITDA',EBIT+DA,False),
    ('Intereses (COMEX $119 + comercial $25)',-INT,False),
    ('Gastos bancarios',-25,False),('RNO',RNO,False),
    ('Utilidad del ejercicio',UT,True)]
for i,(n,v,bold) in enumerate(pl):
    r=15+i
    ws.cell(r,1,n).alignment=LFT; ws.cell(r,1).font=FB if bold else Font()
    c=ws.cell(r,2); c.value=v; c.number_format='#,##0;[Red]-#,##0'; c.alignment=RIG
    c.font=FB if bold else Font()
    if n.startswith('Utilidad'): c.font=GRN if v>0 else RED
    ws.cell(r,3).value=None  # limpiar 2a columna vieja
ws.cell(14,3).value=None

# ===== MC por línea (88-95): 4 líneas incl B2B =====
ws.cell(88,1,'MARGEN DE CONTRIBUCIÓN POR LÍNEA — venta 2027 $6.500M (incluye B2B)').font=FB
for j,h in enumerate(['Línea','Share','Venta','MC %','MC $']):
    c=ws.cell(89,1+j); c.value=h; c.font=WH; c.fill=fhdr; c.alignment=(LFT if j==0 else CEN); c.border=TH
lines=[('Marketplace',ventas['Marketplace']/TOT,ventas['Marketplace'],0.27,mcl['Marketplace'],False),
       ('Páginas Web',ventas['Páginas Web']/TOT,ventas['Páginas Web'],0.275,mcl['Páginas Web'],False),
       ('Fidelización',ventas['Fidelización']/TOT,ventas['Fidelización'],0.33,mcl['Fidelización'],False),
       ('B2B UnionX',B2B/TOT,B2B,0.35,MC_B2B,False),
       ('Total 2027',1.0,TOT,MC/TOT,MC,True)]
for i,(n,s,v,m,mc,tot) in enumerate(lines):
    r=90+i
    ws.cell(r,1,n).font=FB if tot else Font(); ws.cell(r,1).alignment=LFT
    ws.cell(r,2,s).number_format='0.0%'; ws.cell(r,2).alignment=CEN
    ws.cell(r,3,v).number_format='#,##0'; ws.cell(r,3).alignment=RIG
    ws.cell(r,4,m).number_format='0.0%'; ws.cell(r,4).alignment=RIG
    ws.cell(r,5,mc).number_format='#,##0'; ws.cell(r,5).alignment=RIG
    for cc in range(1,6):
        ws.cell(r,cc).border=TH
        if tot: ws.cell(r,cc).fill=fsub; ws.cell(r,cc).font=FB
ws.cell(95,1,'Los $6.500M incluyen el B2B UnionX ($300M, 4,6% del mix). Fidelización (33%) y B2B (35%) son las líneas más rentables; Marketplace (27%) mueve el volumen.').font=IT

# ===== SUMATORIA (70-76): base +42, sin paso +B2B =====
steps=[('Base 2027 $6.500M (incluye B2B)',None,UT,0),
       ('+ Reducción gerencial · SOLO Nicole (nov-26)',55,UT+55,0),
       ('+ Marketing · Michela [ejecutada]',46,UT+101,0),
       ('+ Ecommerce · Ignacia [ejecutada]',33,UT+134,0),
       ('+ Tercerizar EIT (tarifas negociadas +$64,2M)',64,UT+198,OT),
       ('+ Ruteros Trade Mkt (dic-26)',26,UT+224,OT)]
for i,(nom,ah,util,ot) in enumerate(steps):
    r=70+i; ebit=util-RNO; ebitda=ebit+DA
    vals=[nom,ah,ebit,ebitda,util,ebit/TOT,util/TOT,round(ebit/INT,2),round(DEU/ebitda,2),round(DEU/(PAT+util-ot),2)]
    for j,v in enumerate(vals):
        c=ws.cell(r,1+j); c.border=TH; c.font=Font()
        if j==0: c.value=v; c.alignment=LFT; c.font=FB if i in(0,5) else Font()
        elif v is None: c.value='—'; c.alignment=CEN
        elif j in(5,6): c.value=v; c.number_format='0.0%'; c.alignment=RIG
        elif j in(7,8,9): c.value=v; c.number_format='0.00"x"'; c.alignment=RIG
        else: c.value=v; c.number_format='#,##0;[Red]-#,##0'; c.alignment=RIG
    fill = fsub if i==0 else (fgrn if i==5 else blank)
    for cc in range(1,11): ws.cell(r,cc).fill=fill
    if i in(0,5): ws.cell(r,5).font=GRN
# limpiar fila 76 (antes 7ma)
for cc in range(1,11): ws.cell(76,cc).value=None; ws.cell(76,cc).fill=blank; ws.cell(76,cc).border=Border()
ws.cell(77,1,'La venta 2027 son $6.500M (incluyen B2B). Base +$42M. Las 5 palancas (+$224M) -> full-stack +$266M, cob 3,0x, D/EBITDA 3,8x, D/Patrimonio 2,0x. Finiquito bodega $31,7M one-time 2027; 2026 queda en +$7M.').font=IT

# ===== Escenario gerencial (48-56) base actualizada =====
ws.cell(49,3,'Con salida Nicole (−$55M)')
ger=[('GAV',-GAV,-(GAV-55)),('EBIT',EBIT,EBIT+55),('EBITDA',EBIT+DA,EBIT+55+DA),
     ('Utilidad',UT,UT+55),('Cobertura (x)',round(EBIT/INT,2),round((EBIT+55)/INT,2)),
     ('Deuda/EBITDA (x)',round(DEU/(EBIT+DA),2),round(DEU/(EBIT+55+DA),2))]
for i,(n,a,b) in enumerate(ger):
    r=50+i; ws.cell(r,1,n); ws.cell(r,2,a); ws.cell(r,3,b)
    for cc in (2,3):
        c=ws.cell(r,cc); c.alignment=RIG; c.number_format=('0.00"x"' if 'x' in n else '#,##0;[Red]-#,##0')
ws.cell(56,1,'Base = venta $6.500M (incluye B2B), utilidad +$42M. Erich se mantiene; la única liberación gerencial es Nicole (nov-26, ~$55M, indem $21,3M).').font=IT

# ===== Puente patrimonio: util 266, patrimonio 859 =====
ws.cell(104,1,'(+) Utilidad 2027 normalizada (full-stack)'); ws.cell(104,2,266)
ws.cell(106,2,859); ws.cell(108,2,round(DEU/859,2))

# ===== Hoja 7 =====
r7=wb['7. Ratios 2023-2027']
r7.cell(4,6,'2027e base')
for row,val in {5:0.280,6:EBIT/TOT,7:(EBIT+DA)/TOT,8:UT/TOT,10:round(EBIT/INT,2),11:round(DEU/(EBIT+DA),2),12:1729,13:-144,14:round(DEU/(PAT+UT),2)}.items():
    r7.cell(row,6).value=val
r7.cell(16,1,'Venta 2027 $6.500M (incluye B2B). Con las 5 palancas (full-stack +$266M): EBIT% 6,7% · Cobertura 3,0x · Deuda/EBITDA 3,8x · Deuda/Patrimonio 2,0x. Finiquito bodega $31,7M one-time 2027; 2026 +$7M.').font=IT

# ===== Hoja 1 =====
w1=wb['1. Resumen Ejecutivo']
for r in range(1,80):
    v=str(w1.cell(r,1).value or '')
    if v.startswith('2027'):
        w1.cell(r,1,'2027: venta $6.500M (incluye B2B UnionX $300M; digital $6.200M). MC 28,0% -> base utilidad +$42M. Las 5 palancas (+$224M; gerencial solo Nicole $55M, EIT/cambio de bodega negociado +$64,2M, Michela+Ignacia ejecutadas, ruteros) llevan el 2027 a +$266M, cob 3,0x, D/EBITDA 3,8x, D/Patrimonio 2,0x. Finiquito bodega $31,7M one-time 2027. 2026 POSITIVO +$7M.')
        break

OUT=BASE+r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v20.xlsx"
wb.save(OUT)
print('Guardado v20 | base +42 | full-stack +266 | D/Pat 2,01x')
