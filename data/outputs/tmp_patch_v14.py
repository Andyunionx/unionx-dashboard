# -*- coding: utf-8 -*-
"""Patch v13 -> v14:
- Margenes confirmados por usuario: MP 27% / Web 27,5% / Fidel 33% / B2B 35%
- Base digital 6.500 neto (share real 2026) + B2B 300 en la base
- Palancas: Michela e Ignacia EJECUTADAS jul-26 (finiquitos reales); gerencial $120M rango (Nicole confirmada, sin nombrar a Erich como ejecutado); EIT; Ruteros
"""
import sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')
BASE = r"G:\Mi unidad\TRABAJO\RESPALDO\OPERACIONES\UNION X - IA"
wb = openpyxl.load_workbook(BASE + r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v13.xlsx")
ws = wb['8. 2027 y Palancas']
# desmergear todo lo que voy a reescribir (filas >=14)
for mr in list(ws.merged_cells.ranges):
    if mr.min_row >= 14:
        ws.unmerge_cells(str(mr))

FB=Font(bold=True); GRN=Font(color='1E7A1E',bold=True); RED=Font(color='C00000'); IT=Font(size=10,italic=True,color='C00000')
fgrn=PatternFill('solid',fgColor='E3F4E4'); fsub=PatternFill('solid',fgColor='D6E4F0'); fhdr=PatternFill('solid',fgColor='0F2340'); blank=PatternFill()
WH=Font(bold=True,color='FFFFFF')
TH=Border(*[Side(style='thin',color='D9D9D9')]*4); RIG=Alignment(horizontal='right'); CEN=Alignment(horizontal='center'); LFT=Alignment(horizontal='left',wrap_text=True)

# ===== Constantes =====
INT=144; BANC=25; RNO=-(INT+BANC); DEU=1729; PATR=637; DA=18; GAV=1608
# Digital
sh={'Marketplace':80.7,'Fidelización':9.9,'Páginas Web':9.4}; tot=sum(sh.values())
mcpct={'Marketplace':0.27,'Fidelización':0.33,'Páginas Web':0.275}
ventas={k:round(6500*v/tot) for k,v in sh.items()}
mc_line={k:round(ventas[k]*mcpct[k]) for k in sh}
MC_DIG=sum(mc_line.values())                 # ~1797
EBIT_D=MC_DIG-GAV; UT_D=EBIT_D+RNO           # 189 / 20
B2B_V=300; B2B_MC=round(B2B_V*0.35)          # 105
MC_TOT=MC_DIG+B2B_MC; EBIT_B=EBIT_D+B2B_MC; UT_B=UT_D+B2B_MC   # 1902 / 294 / 125
print('MC_DIG',MC_DIG,'EBIT_D',EBIT_D,'UT_D',UT_D,'| EBIT_B',EBIT_B,'UT_B',UT_B)

def kset(r, vals, fills=None, bold_idx=(), num_last=True):
    for j,v in enumerate(vals):
        c=ws.cell(r,1+j); c.border=TH; c.font=FB if j in bold_idx else Font()
        if v is None: c.value=None; continue
        c.value=v
        if isinstance(v,str): c.alignment=LFT if j==0 else CEN
        else:
            c.alignment=RIG
    if fills:
        for cc in range(1,11): ws.cell(r,cc).fill=fills

# ============ P&L 2027 (rows 14-23) — 2 columnas base digital / +B2B ============
ws.cell(14,1,'P&L 2027 (MM CLP)'); ws.cell(14,2,'Base digital'); ws.cell(14,3,'+ B2B $300M')
for c in (1,2,3): ws.cell(14,c).font=WH; ws.cell(14,c).fill=fhdr; ws.cell(14,c).alignment=(LFT if c==1 else CEN)
rows=[('Venta neta',6500,6800),
      ('Margen de contribución',MC_DIG,MC_TOT),
      ('GAV',-GAV,-GAV),
      ('EBIT',EBIT_D,EBIT_B),
      ('EBITDA',EBIT_D+DA,EBIT_B+DA),
      ('Intereses (COMEX $119 + comercial $25)',-INT,-INT),
      ('Gastos bancarios',-BANC,-BANC),
      ('RNO',RNO,RNO),
      ('Utilidad del ejercicio',UT_D,UT_B)]
for i,(n,a,b) in enumerate(rows):
    r=15+i
    ws.cell(r,1,n).font=FB if n in('EBIT','Utilidad del ejercicio') else Font()
    ws.cell(r,1).alignment=LFT
    for cc,val in ((2,a),(3,b)):
        c=ws.cell(r,cc); c.value=val; c.number_format='#,##0;[Red]-#,##0'; c.alignment=RIG
        c.font=FB if n in('EBIT','Utilidad del ejercicio') else Font()
        if n=='Utilidad del ejercicio': c.font=GRN if val>0 else RED
    # limpiar col D si tenía algo
    ws.cell(r,4).value=None
# limpiar la celda D4-D12? no (es GAV bottom-up). Solo aseguro D15-23 limpio ya hecho.

# ============ narrativa PALANCAS + RECOs (25-35) ============
ws.cell(15,1,'Venta digital neta (ML 80,7% + Web 9,4% + Fid 9,9%)').alignment=LFT
ws.cell(26,1,'1. Base digital VIABLE: con márgenes reales por línea (MP 27% / Web 27,5% / Fid 33%) la operación digital ya deja utilidad +$20M.')
ws.cell(27,1,'2. B2B $300M cae casi directo a EBIT (contrib. 35% = +$105M): lleva la base a +$125M sin nuevas palancas.')
ws.cell(28,1,'3. Las palancas de estructura (+$269M) llevan el 2027 a +$394M, cobertura 3,9x, Deuda/EBITDA 3,0x.')
ws.cell(29,1,'4. Michela (Marketing) e Ignacia (Ecommerce) ya salieron en julio: su ahorro 2027 es run-rate realizado; indemnizaciones pegan en 2026.')
ws.cell(30,1,'5. Reducción del rango gerencial $120M: Nicole (nov-26) es el primer paso confirmado; el resto según decisión.')
for r in range(26,31): ws.cell(r,1).font=Font(size=10)

# ============ Escenario gerencial (48-56) reformulado ============
ws.cell(48,1,'ESCENARIO PALANCA: reducción del rango gerencial −$120M (incl. salida de Nicole nov-26)').font=FB
ws.cell(49,3,'Con reducción −$120M')
ger=[('GAV',-GAV,-(GAV-120)),('EBIT',EBIT_B,EBIT_B+120),('EBITDA',EBIT_B+DA,EBIT_B+120+DA),
     ('Utilidad',UT_B,UT_B+120),('Cobertura (x)',round((EBIT_B)/INT,2),round((EBIT_B+120)/INT,2)),
     ('Deuda/EBITDA (x)',round(DEU/(EBIT_B+DA),2),round(DEU/(EBIT_B+120+DA),2))]
for i,(n,a,b) in enumerate(ger):
    r=50+i; ws.cell(r,1,n); ws.cell(r,2,a); ws.cell(r,3,b)
    for cc in (2,3):
        c=ws.cell(r,cc); c.alignment=RIG
        c.number_format=('0.00"x"' if 'x' in n else '#,##0;[Red]-#,##0')
ws.cell(56,1,'Base = digital $6.500M + B2B $300M. El $120M es un rango de reducción de estructura gerencial; Nicole (nov-26) es el primer paso confirmado.').font=IT

# ============ Michela + Ignacia (58-66) — EJECUTADAS ============
ws.cell(58,1,'SALIDAS EJECUTADAS JULIO 2026 — Michela Rossi (Marketing) e Ignacia Sáez (Ecommerce)').font=FB
ws.cell(59,1,'Concepto'); ws.cell(59,2,'Michela (16/07)'); ws.cell(59,3,'Ignacia (31/07)')
for c in (1,2,3): ws.cell(59,c).font=FB; ws.cell(59,c).fill=fsub
mi=[('Antigüedad','2 años','4 años'),
    ('Finiquito real (una vez, 2026)','−12,9','−14,9'),
    ('Ahorro remuneración 2026 (ago-dic)','+21','+14'),
    ('Efecto neto 2026','+8','−1'),
    ('Ahorro 2027 (año completo)','+46','+33')]
for i,(n,a,b) in enumerate(mi):
    r=60+i; ws.cell(r,1,n); ws.cell(r,2,a); ws.cell(r,3,b)
    for cc in (2,3): ws.cell(r,cc).alignment=RIG
ws.cell(65,1,'Finiquitos reales (Buk): Michela art.161 $12,9M (vac+aviso+2 años, tope 90 UF). Ignacia 4 años ~$14,9M. Ahorro 2027 = costo cargado anual.').font=IT
ws.cell(66,1,None)

# ============ SUMATORIA (68-77) ============
ws.cell(68,1,'SUMATORIA DE PALANCAS 2027 — desde la base digital+B2B hasta el full-stack').font=FB
hdr=['Escenario acumulado','Ahorro','EBIT','EBITDA','Utilidad','EBIT %','Utilidad %','Cobertura','Deuda/EBITDA','Deuda/Patrim']
for j,h in enumerate(hdr):
    c=ws.cell(69,1+j); c.value=h; c.font=WH; c.fill=fhdr; c.alignment=(LFT if j==0 else CEN); c.border=TH
steps=[('Base digital $6.500M',None,UT_D,6500),
       ('+ B2B UnionX $300M (contrib. 35%)',B2B_MC,UT_B,6800),
       ('+ Reducción rango gerencial (incl. Nicole)',120,UT_B+120,6800),
       ('+ Marketing · Michela [ejecutada]',46,UT_B+166,6800),
       ('+ Ecommerce · Ignacia [ejecutada]',33,UT_B+199,6800),
       ('+ Tercerizar EIT (básico)',44,UT_B+243,6800),
       ('+ Ruteros Trade Mkt (dic-26)',26,UT_B+269,6800)]
for i,(nom,ah,util,vta) in enumerate(steps):
    r=70+i
    ebit=util-RNO; ebitda=ebit+DA
    vals=[nom, ah, ebit, ebitda, util, ebit/vta, util/vta, round(ebit/INT,2), round(DEU/ebitda,2), round(DEU/(PATR+util),2)]
    for j,v in enumerate(vals):
        c=ws.cell(r,1+j); c.border=TH; c.font=Font()
        if j==0: c.value=v; c.alignment=LFT; c.font=FB if i in(0,6) else Font()
        elif v is None: c.value='—'; c.alignment=CEN
        elif j in(5,6): c.value=v; c.number_format='0.0%'; c.alignment=RIG
        elif j in(7,8,9): c.value=v; c.number_format='0.00"x"'; c.alignment=RIG
        else: c.value=v; c.number_format='#,##0;[Red]-#,##0'; c.alignment=RIG
    fill = fsub if i in(0,1) else (fgrn if i==6 else blank)
    for cc in range(1,11): ws.cell(r,cc).fill=fill
    if i==0: ws.cell(r,5).font=(GRN if UT_D>0 else RED)
    if i==6: ws.cell(r,5).font=GRN
ws.cell(77,1,'Base digital ya deja +$20M; con B2B $300M +$125M. Las 5 palancas (+$269M) llevan el full-stack a +$394M, cobertura 3,9x, Deuda/EBITDA 3,0x, D/Patrimonio 1,7x.').font=IT

# ============ VISTA POR PALANCA (78-85) ============
ws.cell(78,1,'VISTA POR PALANCA — ahorro 2027, indemnización y efecto en 2026').font=FB
vh=['Palanca','Alcance','Ahorro 2027','Indemnización','Efecto 2026','Nota']
for j,h in enumerate(vh):
    c=ws.cell(79,1+j); c.value=h; c.font=WH; c.fill=fhdr; c.alignment=(LFT if j==0 else CEN); c.border=TH
det=[('1 · Reducción rango gerencial','$120M rango (Nicole confirmada nov-26)','+120','−21,3','−13','Nicole 5,2 años. El resto del rango según ejecución (no comprometido)'),
     ('2 · Marketing · Michela [ejec. 16/07]','1','+46','−12,9','+8','2 años; finiquito real $12,9M (tope 90 UF)'),
     ('3 · Ecommerce · Ignacia [ejec. 31/07]','1','+33','−14,9','−1','4 años de servicio'),
     ('4 · Tercerizar EIT (básico, 2027)','10 (9 operarios + chofer)','+44','−31,7','0','Indemn. consume ~72% del ahorro año 1; pleno desde 2028'),
     ('5 · Ruteros Trade Mkt (dic-26)','2','+26','−5,8','−4','~1,5 años c/u'),
     ('TOTAL','15','+269','−86,6','−10 (2026 → +$9M)','Indemn. 2026 $54,9M + EIT $31,7M (2027). 2026 pasa de +$19M a ~+$9M')]
for i,row in enumerate(det):
    r=80+i
    for j in range(6):
        c=ws.cell(r,1+j); c.border=TH; c.value=row[j]
        c.alignment=RIG if j in(2,3,4) else LFT
        c.font=FB if row[0]=='TOTAL' else Font()
        c.fill=fsub if row[0]=='TOTAL' else blank

# ============ MC POR LINEA CONFIRMADO (88-95) ============
ws.cell(88,1,'MARGEN DE CONTRIBUCIÓN POR LÍNEA — confirmado (Análisis de Contribución)').font=FB
mh=['Línea','Share digital','Venta 2027','MC %','MC $']
for j,h in enumerate(mh):
    c=ws.cell(89,1+j); c.value=h; c.font=WH; c.fill=fhdr; c.alignment=(LFT if j==0 else CEN); c.border=TH
lines=[('Marketplace',sh['Marketplace']/tot,ventas['Marketplace'],0.27,mc_line['Marketplace']),
       ('Páginas Web',sh['Páginas Web']/tot,ventas['Páginas Web'],0.275,mc_line['Páginas Web']),
       ('Fidelización',sh['Fidelización']/tot,ventas['Fidelización'],0.33,mc_line['Fidelización']),
       ('Total digital',1.0,6500,MC_DIG/6500,MC_DIG),
       ('B2B UnionX','—',300,0.35,B2B_MC)]
for i,(n,s,v,m,mc) in enumerate(lines):
    r=90+i; totrow=(n=='Total digital')
    ws.cell(r,1,n).font=FB if totrow else Font(); ws.cell(r,1).alignment=LFT
    cs=ws.cell(r,2); cs.value=(s if isinstance(s,str) else s); cs.alignment=CEN
    if not isinstance(s,str): cs.number_format='0.0%'
    ws.cell(r,3,v).number_format='#,##0'; ws.cell(r,3).alignment=RIG
    ws.cell(r,4,m).number_format='0.0%'; ws.cell(r,4).alignment=RIG
    ws.cell(r,5,mc).number_format='#,##0'; ws.cell(r,5).alignment=RIG
    for cc in range(1,6):
        ws.cell(r,cc).border=TH
        if totrow: ws.cell(r,cc).fill=fsub; ws.cell(r,cc).font=FB
ws.cell(95,1,'Fidelización (33%) es la línea más rentable; Marketplace (27%) la de mayor volumen. B2B (35%) casi sin comisión de marketplace.').font=IT

# ============ HOJA 1 ============
w1=wb['1. Resumen Ejecutivo']
for r in range(1,80):
    v=str(w1.cell(r,1).value or '')
    if v.startswith('2027 BASE DIGITAL') or v.startswith('2027 BASE'):
        w1.cell(r,1,'2027: la base digital $6.500M neto (márgenes reales por línea) ya deja utilidad +$20M. Con B2B $300M (+$105M) sube a +$125M. Las 5 palancas de estructura (+$269M) llevan el 2027 a +$394M, cobertura 3,9x, Deuda/EBITDA 3,0x. 2026 pasa de +$19M a ~+$9M por indemnizaciones (una vez).')

OUT=BASE+r"\data\outputs\Analisis_CFO_UnionX_2026-2027_v14.xlsx"
wb.save(OUT)
print('Guardado:',OUT)
print(f'FULL-STACK 2027: +{UT_B+269} | Base digital {UT_D} | +B2B {UT_B}')
