"""Genera 'Reporte Ventas Empresa LIVE.xlsx' actualizando la hoja Raw de la
plantilla dinámica (diseño de Nicolás) con datos frescos del parquet.

La plantilla tiene 2 hojas — "Resumen TD" (la dinámica con apertura de gastos:
jerarquía Mes>Week>Tipo Negocio>Canal>Marca>Cat.padre>Cat.hijo>Producto>SKU,
20 medidas y 14 campos calculados en %) y "Raw" (la fuente). Todo el diseño
del pivot vive en la propia plantilla, así que acá solo:
  1. Cargamos la plantilla como solo-lectura (no se modifica).
  2. Reemplazamos la hoja Raw (sheet2) con datos frescos del parquet (2025-2026).
  3. Vaciamos pivotCacheRecords1 + refreshOnLoad=1 -> Excel refresca al abrir.
  4. Ajustamos el ref del worksheetSource al total real de filas.
El resto de las partes se copian byte-a-byte.

Salida: data/outputs/Reporte Ventas Empresa LIVE.xlsx
"""
import os
import re
import sys
import tempfile
import zipfile
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

# Columnas de ID largo (16+ dígitos, ej. Pedido/pack-id de ML, SKU barcode). Excel
# y el pivot truncan a 15 dígitos significativos si las tratan como número → se
# fuerzan a TEXTO ('@') en la hoja Raw para que el dígito 16 no se pierda.
COLS_ID_TEXTO = {'Pedido', 'SKU', 'Documento'}

PROJECT_ROOT = Path(__file__).resolve().parent
# Plantilla = "dinámica con apertura de gastos" que diseñó Nicolás (jul-2026):
# 2 hojas (Resumen TD + Raw), pivot con jerarquía Mes>..>SKU, 20 medidas y 14
# campos calculados en % (Crec, Share, Marg Front/Final, Gasto Com/Log/Mkt).
# Toda la personalización vive en el pivot; la fuente Raw es la que generamos.
ORIGINAL = PROJECT_ROOT / 'data' / 'planillas' / 'Reporte Ventas Empresa DINAMICA.xlsx'
# Carpeta Drive "Reporte Automático RAW" (id fijo, compartida con OAuth user)
DRIVE_CARPETA_ID = '18RKgdGwWGM8tEGqltcrruCPB-LaTwZxx'
NOMBRE_LIVE = 'Reporte Ventas Empresa LIVE.xlsx'
# Plantilla dinámica en Drive. Gitignored por tamano (124 MB) -> en GitHub
# Actions no existe en el checkout, se baja de aca. file_id permanente.
PLANTILLA_DRIVE_ID = '1V7c-IFbqBqGmVHc4qkkmAbccqMQDpNqB'
ANIO_TY = 2026  # Ano TY (This Year). LY = ANIO_TY - 1.
# Output local: en GH Actions usa /tmp, local usa data/outputs/
if os.environ.get('GITHUB_ACTIONS') == 'true':
    OUTPUT = Path('/tmp') / NOMBRE_LIVE
else:
    OUTPUT = PROJECT_ROOT / 'data' / 'outputs' / NOMBRE_LIVE

# 56 columnas en el ORDEN EXACTO del original (validado contra archivo)
RAW_COLUMNS = [
    'Fecha Comp', 'Tipo Movimiento', 'Bodega', 'Documento', 'Fecha Documento',
    'Pedido', 'Estado Pedido', 'Tipo Despacho', 'SKU', 'Canal',
    'Fecha Venta', 'Hora Venta', 'Producto', 'Categoría macro',
    'Categoría padre', 'Categoría hijo', 'Categoría comercial', 'Estado SKU',
    'Pack', 'Marca', 'Proveedor', 'Tipo Marca', 'Tipo Compra', 'Tipo Negocio',
    'KAM', 'Estado Canal', 'Año venta', 'Mes venta', 'Semana venta',
    'Día semana', 'Hora venta', 'Cantidad TY', 'Venta bruta TY',
    'Costo Unitario TY', 'Costo Total TY', 'Margen Front TY', 'Comision % TY',
    'Comisión TY', 'Logística TY', 'Marketing TY', 'Cantidad LY',
    'Venta bruta LY', 'Costo Unitario LY', 'Costo Total LY', 'Margen Front LY',
    'Comision % LY', 'Comisión LY', 'Logística LY', 'Marketing LY',
    'Año', 'Mes', 'Semana', 'Fecha Compara', 'Week', 'Obs', None,
]

# Mapeo columna XLSX -> nombre en parquet
PARQUET_MAP = {
    'Tipo Movimiento': 'tipo_movimiento',
    'Bodega': 'bodega',
    'Documento': 'documento',
    'Pedido': 'pedido',
    'Estado Pedido': 'estado_pedido',
    'Tipo Despacho': 'tipo_despacho',
    'SKU': 'sku',
    'Canal': 'canal',
    'Fecha Venta': 'fecha_venta',
    'Producto': 'producto',
    'Categoría macro': 'categoria_macro',
    'Categoría padre': 'categoria_padre',
    'Categoría hijo': 'categoria_hijo',
    'Categoría comercial': 'categoria_comercial',
    'Estado SKU': 'estado_sku',
    'Pack': 'pack',
    'Marca': 'marca',
    'Proveedor': 'proveedor',
    'Tipo Marca': 'tipo_marca',
    'Tipo Compra': 'tipo_compra',
    'Tipo Negocio': 'tipo_negocio',
    'KAM': 'kam',
    'Estado Canal': 'estado_canal',
}


def cargar_raw_parquet() -> pd.DataFrame:
    """Lee parquet (hist + mes_actual), filtra a 2025-2026, prepara 56 cols.

    CRITICAL: mes_actual se filtra a >= CUTOFF_HISTORICO. El historico tiene
    foto fija que ya incluye 31-may y 1-jun; sin el filtro esos dias se
    cuentan DOS veces (bug $647M vs $538M en pivot LIVE, 11-jun-2026).
    """
    cutoff = '2026-06-02'
    try:
        shared_path = PROJECT_ROOT / 'views' / 'shared.py'
        if shared_path.exists():
            for line in shared_path.read_text(encoding='utf-8').splitlines():
                if line.strip().startswith('CUTOFF_HISTORICO'):
                    val = line.split('=', 1)[1].strip().split('#')[0].strip().strip("'\"")
                    if len(val) == 10 and val[4] == '-':
                        cutoff = val
                    break
    except Exception:
        pass

    h = pd.read_parquet(PROJECT_ROOT / 'data' / 'historico' / 'ventas_historico.parquet')
    m = pd.read_parquet(PROJECT_ROOT / 'data' / 'historico' / 'ventas_mes_actual.parquet')
    m_fv = pd.to_datetime(m['fecha_venta'], errors='coerce').dt.strftime('%Y-%m-%d')
    m = m[m_fv >= cutoff].copy()
    cols = [c for c in m.columns if c in h.columns]
    df = pd.concat([h[cols], m[cols]], ignore_index=True)

    df['fv'] = pd.to_datetime(df['fecha_venta'], errors='coerce')
    df['anio'] = df['fv'].dt.year
    df = df[df['anio'].isin([2025, 2026])].copy()

    # Construir DataFrame con las 56 columnas del Excel
    out = pd.DataFrame(index=df.index)

    # Columna 1: Fecha Comp para alinear LY a TY (sin meter fechas en TY+1)
    # - Filas LY (2025): fecha_venta + 364 dias -> 2026 (alinea con TY)
    # - Filas TY (2026): fecha_venta misma (sin alterar, ya esta en TY)
    fc = df['fv'].copy()
    es_ly = df['anio'] < ANIO_TY
    fc.loc[es_ly] = fc.loc[es_ly] + pd.Timedelta(days=364)
    out['Fecha Comp'] = fc.dt.strftime('%Y-%m-%d')

    # Columnas mapeadas directo del parquet
    for excel_col, parq_col in PARQUET_MAP.items():
        if parq_col in df.columns:
            out[excel_col] = df[parq_col].astype('object').fillna('')
        else:
            out[excel_col] = ''

    # Fecha Documento: tomar fecha_venta como fallback
    out['Fecha Documento'] = df['fv'].dt.strftime('%Y-%m-%d')
    # Hora Venta como texto
    if 'hora_venta' in df.columns:
        out['Hora Venta'] = df['hora_venta'].astype('object').fillna('')
    else:
        out['Hora Venta'] = ''

    # Año/Mes/Semana/Día venta
    out['Año venta'] = df['fv'].dt.year.astype('Int64')
    out['Mes venta'] = df['fv'].dt.month.astype('Int64')
    out['Semana venta'] = df['fv'].dt.isocalendar().week.astype('Int64')
    out['Día semana'] = df['fv'].dt.weekday + 1  # lunes=1 ... domingo=7
    out['Hora venta'] = df.get('hora_venta_num', pd.Series([0]*len(df))).fillna(0)

    # Metricas TY (solo 2026) y LY (solo 2025)
    es_ty = df['anio'] == 2026
    for col_xlsx, col_parq, suf in [
        ('Cantidad', 'cantidad', None),
        ('Venta bruta', 'venta_bruta', None),
        ('Costo Unitario', 'costo_unitario', None),
        ('Costo Total', 'costo_total', None),
        ('Margen Front', 'margen_front', None),
        ('Comision %', 'comision_pct', None),
        ('Comisión', 'comision', None),
        ('Logística', 'logistica', None),
        ('Marketing', 'marketing', None),
    ]:
        val = pd.to_numeric(df.get(col_parq, 0), errors='coerce').fillna(0)
        out[f'{col_xlsx} TY'] = val.where(es_ty, 0)
        out[f'{col_xlsx} LY'] = val.where(~es_ty, 0)

    # Columnas 50-54: Año, Mes, Semana, Fecha Compara, Week, Obs
    out['Año'] = df['fv'].dt.year.astype('Int64')
    out['Mes'] = df['fv'].dt.month.astype('Int64')
    out['Semana'] = df['fv'].dt.isocalendar().week.astype('Int64')
    out['Fecha Compara'] = out['Fecha Comp']  # mismo valor
    out['Week'] = df['fv'].dt.isocalendar().week.astype('Int64')
    out['Obs'] = ''

    # Asegurar orden y agregar columna 56 None
    cols_final = [c for c in RAW_COLUMNS if c is not None]
    out = out[cols_final]
    return out


def generar_raw_xml(df: pd.DataFrame, tmp_dir: Path) -> bytes:
    """Genera sheet3.xml (Raw) con los datos del DataFrame usando openpyxl,
    luego extrae el XML del archivo generado."""
    tmp_xlsx = tmp_dir / '_raw_tmp.xlsx'
    wb = Workbook(write_only=True)
    ws = wb.create_sheet('Raw')
    # Headers
    ws.append([c if c is not None else '' for c in RAW_COLUMNS])
    # Datos en chunks para no consumir tanta RAM
    n_rows = len(df)
    print(f'   [generar_raw_xml] {n_rows:,} filas a serializar...')
    # Índices (0-based) de las columnas de ID que van como TEXTO en RAW_COLUMNS.
    id_idx = {i for i, c in enumerate(RAW_COLUMNS) if c in COLS_ID_TEXTO}
    # Generador (no df.values.tolist()) para no duplicar 440k filas en RAM.
    for r in df.itertuples(index=False, name=None):
        # Agregar None para la columna 56
        fila = list(r) + [None]
        for i in id_idx:
            v = fila[i]
            if v not in (None, ''):
                celda = WriteOnlyCell(ws, value=str(v))
                celda.number_format = '@'  # TEXTO → Excel no trunca a 15 dígitos
                fila[i] = celda
        ws.append(fila)
    del df
    import gc
    gc.collect()
    wb.save(tmp_xlsx)
    # Extraer sheet3.xml (que ahora se llama sheet1.xml en este wb chico)
    with zipfile.ZipFile(tmp_xlsx) as z:
        for nm in z.namelist():
            if nm.startswith('xl/worksheets/sheet'):
                return z.read(nm), n_rows
    raise RuntimeError('no se encontro sheet en archivo temporal')


def _asegurar_plantilla():
    """Si la plantilla original no existe (caso GitHub Actions, porque pesa
    172 MB y esta gitignored), la baja de Drive por su file_id. En local no
    hace nada (la plantilla ya esta en disco)."""
    if ORIGINAL.exists():
        return
    print(f'[plantilla] no existe local -> bajando de Drive ({PLANTILLA_DRIVE_ID})...', flush=True)
    from drive_user_helpers import descargar_archivo
    descargar_archivo(PLANTILLA_DRIVE_ID, ORIGINAL)


def construir_live():
    """Genera el LIVE a partir de la plantilla dinámica de Nicolás.

    La plantilla (2 hojas: Resumen TD + Raw) ya trae TODO el diseño del pivot
    (jerarquía de filas, 20 medidas, campos calculados en %). Acá solo:
      - reemplazamos la hoja Raw (sheet2) con datos frescos del parquet,
      - vaciamos pivotCacheRecords1 + refreshOnLoad=1 → Excel refresca al abrir,
      - ajustamos el ref del worksheetSource al total real de filas (si no, la
        caché quedaría corta y no cargarían los primeros meses).
    El resto de las partes se copian byte-a-byte.
    """
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    _asegurar_plantilla()

    print('[1/4] Cargando datos del parquet...')
    print('[2/4] Generando XML de la nueva Raw...')
    # No retener el df fuera de generar_raw_xml: asi su unica referencia es el
    # parametro y se libera (del+gc) antes del zip-surgery, evitando OOM.
    with tempfile.TemporaryDirectory() as tmp:
        new_raw_xml, n_raw = generar_raw_xml(cargar_raw_parquet(), Path(tmp))
    print(f'      Raw XML: {len(new_raw_xml)/1024/1024:.1f} MB ({n_raw:,} filas)')

    print('[3/4] Construyendo archivo LIVE...')

    # En la plantilla de Nicolás: sheet1 = "Resumen TD" (pivot), sheet2 = "Raw".
    RAW_SHEET = 'xl/worksheets/sheet2.xml'
    CACHE_DEF = 'xl/pivotCache/pivotCacheDefinition1.xml'
    CACHE_REC = 'xl/pivotCache/pivotCacheRecords1.xml'

    empty_records = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                     '<pivotCacheRecords xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                     'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                     'count="0"/>')

    if OUTPUT.exists():
        OUTPUT.unlink()

    with zipfile.ZipFile(ORIGINAL, 'r') as src, \
         zipfile.ZipFile(OUTPUT, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as dst:
        for name in src.namelist():
            if name == RAW_SHEET:
                # Hoja Raw = datos frescos del parquet.
                dst.writestr(name, new_raw_xml)
            elif name == CACHE_REC:
                # Records vacío -> con refreshOnLoad Excel regenera desde la Raw.
                dst.writestr(name, empty_records)
            elif name == CACHE_DEF:
                cd = src.read(name).decode('utf-8')
                # Ajustar el ref al total real de filas (la plantilla trae un ref
                # hardcodeado A1:BC448288 que no cubre todas las filas → si no se
                # arregla, no cargan los primeros meses).
                cd = re.sub(r'(<worksheetSource ref="A1:[A-Z]+)\d+"', rf'\g<1>{n_raw + 1}"', cd, count=1)
                # refreshOnLoad=1 (refresca al abrir) + recordCount=0 (cache stale).
                if 'refreshOnLoad' in cd:
                    cd = re.sub(r'refreshOnLoad="\d"', 'refreshOnLoad="1"', cd)
                else:
                    cd = cd.replace('<pivotCacheDefinition ',
                                    '<pivotCacheDefinition refreshOnLoad="1" ', 1)
                cd = re.sub(r'recordCount="\d+"', 'recordCount="0"', cd)
                dst.writestr(name, cd)
            else:
                # Todo lo demás (Resumen TD, pivotTable1, styles, sharedStrings,
                # workbook, rels, content-types) tal cual: es el diseño de Nicolás.
                with src.open(name) as fin, dst.open(name, 'w', force_zip64=True) as fout:
                    while True:
                        chunk = fin.read(1024 * 1024)
                        if not chunk:
                            break
                        fout.write(chunk)

    print('[4/4] Validando ZIP...')
    with zipfile.ZipFile(OUTPUT) as z:
        bad = z.testzip()
        if bad:
            raise RuntimeError(f'ZIP invalido: {bad}')

    size_mb = OUTPUT.stat().st_size / (1024 * 1024)
    print(f'\n[OK] {OUTPUT.name}: {size_mb:.1f} MB')


def subir_a_drive():
    """Sube el archivo OUTPUT al Drive del user via OAuth user."""
    from drive_user_helpers import subir_o_actualizar
    file_id, link = subir_o_actualizar(
        OUTPUT, DRIVE_CARPETA_ID, NOMBRE_LIVE, hacer_publico=True
    )
    print(f'\n[drive] file_id: {file_id}')
    print(f'[drive] link:    {link}')
    return file_id, link


if __name__ == '__main__':
    construir_live()
    # Solo subir a Drive si estamos en GH Actions o si se pide explicito
    if os.environ.get('GITHUB_ACTIONS') == 'true' or '--upload' in sys.argv:
        subir_a_drive()
