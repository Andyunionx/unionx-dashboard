#!/usr/bin/env python3
"""
Extractor del archivo maestro de planificación financiera UnionX.

Lee data/planillas/Planificación Financiera 2026.xlsx (58 hojas, mantenido
manualmente por Andrés) y genera parquets ligeros que alimentan los dashboards
de la app Finanzas — sin tocar el archivo original.

Hojas procesadas:
  - P&L                        → pyl_mensual.parquet (96 meses 2019-2026, todas las líneas)
  - Ppto 2026                  → ppto_2026.parquet (48 meses 2023-2026 por CC)
  - Resumen YTD                → resumen_ytd.parquet (12 filas, YTD vs Ppto vs YoY)
  - KT                         → kt.parquet (capital trabajo + meses inventario/CxC/CxP)
  - Deuda financiera           → deuda.parquet (saldos + intereses + cronograma)
  - Metas 2026                 → metas_2026.parquet (Venta/Contrib/GAV/EBIT/EBITDA mensual)
  - Fcst EERR                  → fcst_eerr.parquet (forecast EERR full)
  - Dashboard Data             → dashboard_data.parquet (KPIs pre-cocinados)
  - Análisis Financiero YTD    → analisis_financiero_ytd.parquet
  - Análisis Comparativo       → analisis_comparativo.parquet

Output: data/finanzas/*.parquet + resumen_general.json (overview rápido)

Cron sugerido: post-modificación del Excel (manual o cada 6h por si se actualizó).
"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

PROJECT_ROOT = Path(__file__).parent
EXCEL_FILE = PROJECT_ROOT / "data" / "planillas" / "Planificación Financiera 2026.xlsx"
OUT_DIR = PROJECT_ROOT / "data" / "finanzas"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def _abrir():
    if not EXCEL_FILE.exists():
        sys.exit(f"[ERROR] No existe {EXCEL_FILE}")
    return openpyxl.load_workbook(str(EXCEL_FILE), read_only=True, data_only=True)


# ============================================================
# HELPERS
# ============================================================
def _es_fecha(v):
    return hasattr(v, "strftime") and not isinstance(v, str)


def _encontrar_fila_fechas(ws, max_filas: int = 15, min_fechas: int = 5):
    """Encuentra la primera fila con cabecera de meses (≥ min_fechas fechas)."""
    for trow in range(1, max_filas + 1):
        rdata = list(ws.iter_rows(values_only=True, min_row=trow, max_row=trow))[0]
        if sum(1 for v in rdata if _es_fecha(v)) >= min_fechas:
            return trow, rdata
    return None, None


def _columnas_fecha(rdata) -> list[tuple[int, datetime]]:
    """[(col_index, fecha)] para las columnas con fecha."""
    return [(j, v) for j, v in enumerate(rdata) if _es_fecha(v)]


# ============================================================
# EXTRACTORES POR HOJA
# ============================================================
def extract_pyl(wb) -> pd.DataFrame:
    """P&L histórico 2019-2026 (96 meses, todas las líneas EERR).

    Estructura del archivo:
      F3 col 6-101: fechas 2019-01 a 2026-12
      F4+: filas con labels en cols B, C o D (jerárquico) y valores en cols fecha

    Devuelve formato LARGO: (fecha, seccion, linea, valor, año, mes, codigo_cc).
    """
    ws = wb["P&L"]
    fila_fechas, rdata = _encontrar_fila_fechas(ws)
    if fila_fechas is None:
        return pd.DataFrame()
    cols_fecha = _columnas_fecha(rdata)

    rows = []
    seccion_actual = ""
    for i, row in enumerate(ws.iter_rows(values_only=True, min_row=fila_fechas + 1, max_row=170)):
        # Identificar label: prioridad col D, luego C, luego B
        label_b = (row[1] if len(row) > 1 else "") or ""
        label_c = (row[2] if len(row) > 2 else "") or ""
        label_d = (row[3] if len(row) > 3 else "") or ""
        label_b = str(label_b).strip()
        label_c = str(label_c).strip()
        label_d = str(label_d).strip()

        # Si col B tiene texto, es título de sección (nivel superior)
        if label_b and not label_d:
            seccion_actual = label_b
            continue
        # Línea con valor en col D
        linea = label_d or label_c or label_b
        if not linea:
            continue

        for col_idx, fecha in cols_fecha:
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            if not isinstance(v, (int, float)):
                continue
            rows.append({
                "fecha": fecha,
                "year": fecha.year,
                "month": fecha.month,
                "seccion": seccion_actual or "Otros",
                "linea": linea,
                "valor": float(v),
            })

    df = pd.DataFrame(rows)
    df["fecha"] = pd.to_datetime(df["fecha"])
    return df


def extract_ppto_2026(wb) -> pd.DataFrame:
    """Ppto 2026 con códigos de cuenta contable (centros de costo).

    Estructura:
      F6 cols 54-101: fechas 2023-01 a 2026-12
      F8+: filas con código contable en col B, nombre en col C
    """
    ws = wb["Ppto 2026"]
    fila_fechas, rdata = _encontrar_fila_fechas(ws, max_filas=12)
    if fila_fechas is None:
        return pd.DataFrame()
    cols_fecha = _columnas_fecha(rdata)

    rows = []
    seccion_actual = ""
    for i, row in enumerate(ws.iter_rows(values_only=True,
                                          min_row=fila_fechas + 1,
                                          max_row=140)):
        codigo = (row[1] if len(row) > 1 else "") or ""
        nombre = (row[2] if len(row) > 2 else "") or ""
        codigo = str(codigo).strip()
        nombre = str(nombre).strip()
        if not nombre:
            continue
        # Detectar sección por código vacío + nombre con mayúscula
        if not codigo and nombre.isupper():
            seccion_actual = nombre
            # No skipear: puede ser un total con valores
        for col_idx, fecha in cols_fecha:
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None:
                continue
            if not isinstance(v, (int, float)):
                continue
            rows.append({
                "fecha": fecha,
                "year": fecha.year,
                "month": fecha.month,
                "codigo_cc": codigo,
                "linea": nombre,
                "seccion": seccion_actual,
                "valor_ppto": float(v),
            })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["fecha"] = pd.to_datetime(df["fecha"])
    return df


def extract_resumen_ytd(wb) -> pd.DataFrame:
    """Hoja Resumen YTD: tabla con YTD real vs Ppto vs YoY + ratios financieros."""
    ws = wb["Resumen YTD"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=70)):
        label = (row[0] if row else "") or ""
        label = str(label).strip()
        if not label:
            continue
        ytd = row[1] if len(row) > 1 else None
        ytd_ppto = row[2] if len(row) > 2 else None
        yoy = row[3] if len(row) > 3 else None
        pct_ppto = row[5] if len(row) > 5 else None
        pct_yoy = row[6] if len(row) > 6 else None
        tot_ppto = row[8] if len(row) > 8 else None
        tot_yoy = row[9] if len(row) > 9 else None
        # Solo si al menos un numérico
        nums = [v for v in (ytd, ytd_ppto, yoy) if isinstance(v, (int, float))]
        if not nums:
            continue
        rows.append({
            "concepto": label,
            "ytd_2026": ytd,
            "ytd_ppto": ytd_ppto,
            "ytd_2025": yoy,
            "var_pct_ppto": pct_ppto,
            "var_pct_yoy": pct_yoy,
            "var_abs_ppto": tot_ppto,
            "var_abs_yoy": tot_yoy,
        })
    return pd.DataFrame(rows)


def extract_kt(wb) -> pd.DataFrame:
    """KT: capital de trabajo con líneas (Existencias, CxC, CxP, Total)."""
    ws = wb["KT"]
    fila_fechas, rdata = _encontrar_fila_fechas(ws, max_filas=10)
    if fila_fechas is None:
        return pd.DataFrame()
    cols_fecha = _columnas_fecha(rdata)

    rows = []
    seccion = ""
    for i, row in enumerate(ws.iter_rows(values_only=True,
                                          min_row=fila_fechas + 1, max_row=89)):
        b = (row[1] if len(row) > 1 else "") or ""
        c = (row[2] if len(row) > 2 else "") or ""
        d = (row[3] if len(row) > 3 else "") or ""
        b = str(b).strip()
        c = str(c).strip()
        d = str(d).strip()

        if c and not d:
            seccion = c
            continue
        linea = d or c or b
        if not linea:
            continue
        for col_idx, fecha in cols_fecha:
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None or not isinstance(v, (int, float)):
                continue
            rows.append({
                "fecha": fecha,
                "year": fecha.year,
                "month": fecha.month,
                "seccion": seccion,
                "linea": linea,
                "valor": float(v),
            })
    df = pd.DataFrame(rows)
    if not df.empty:
        df["fecha"] = pd.to_datetime(df["fecha"])
    return df


def extract_deuda(wb) -> pd.DataFrame:
    """Deuda Financiera: saldos + intereses + cronograma."""
    ws = wb["Deuda financiera"]
    fila_fechas, rdata = _encontrar_fila_fechas(ws, max_filas=10)
    if fila_fechas is None:
        return pd.DataFrame()
    cols_fecha = _columnas_fecha(rdata)

    rows = []
    seccion = ""
    for i, row in enumerate(ws.iter_rows(values_only=True,
                                          min_row=fila_fechas + 1, max_row=90)):
        c = (row[2] if len(row) > 2 else "") or ""
        d = (row[3] if len(row) > 3 else "") or ""
        c = str(c).strip()
        d = str(d).strip()
        if c and not d:
            seccion = c
            continue
        linea = d or c
        if not linea:
            continue
        for col_idx, fecha in cols_fecha:
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None or not isinstance(v, (int, float)):
                continue
            rows.append({
                "fecha": fecha,
                "year": fecha.year,
                "month": fecha.month,
                "seccion": seccion,
                "linea": linea,
                "valor": float(v),
            })
    df = pd.DataFrame(rows)
    if not df.empty:
        df["fecha"] = pd.to_datetime(df["fecha"])
    return df


def extract_metas_2026(wb) -> pd.DataFrame:
    """Metas 2026: Venta / Contribución / GAV / EBIT / EBITDA con
    Meta · Resultado · Var · Var% · Resultado2025 · Variación · Variación%.
    """
    ws = wb["Metas 2026"]
    rows = []
    current_kpi = None
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=47)):
        first = row[0] if row else None
        # Es header con fechas?
        if first and isinstance(first, str) and any(_es_fecha(v) for v in row[1:14]):
            current_kpi = first.strip()
            headers = [v for v in row[1:14]]  # 12 meses
            continue
        # Si la fila tiene un label (Meta, Resultado, Var, etc.) → registro
        if first and isinstance(first, str) and current_kpi and headers:
            tipo = first.strip()
            for col_idx, fecha in enumerate(headers):
                if not _es_fecha(fecha):
                    continue
                v = row[col_idx + 1] if len(row) > col_idx + 1 else None
                if v is None or not isinstance(v, (int, float)):
                    continue
                rows.append({
                    "kpi": current_kpi,
                    "tipo": tipo,
                    "fecha": fecha,
                    "year": fecha.year,
                    "month": fecha.month,
                    "valor": float(v),
                })
    df = pd.DataFrame(rows)
    if not df.empty:
        df["fecha"] = pd.to_datetime(df["fecha"])
    return df


def extract_fcst_eerr(wb) -> pd.DataFrame:
    """Fcst EERR: forecast línea por línea, mismo formato que P&L."""
    if "Fcst EERR" not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb["Fcst EERR"]
    fila_fechas, rdata = _encontrar_fila_fechas(ws, max_filas=15)
    if fila_fechas is None:
        return pd.DataFrame()
    cols_fecha = _columnas_fecha(rdata)

    rows = []
    seccion_actual = ""
    for i, row in enumerate(ws.iter_rows(values_only=True,
                                          min_row=fila_fechas + 1, max_row=139)):
        b = (row[1] if len(row) > 1 else "") or ""
        c = (row[2] if len(row) > 2 else "") or ""
        d = (row[3] if len(row) > 3 else "") or ""
        b = str(b).strip()
        c = str(c).strip()
        d = str(d).strip()
        if b and not d:
            seccion_actual = b
            continue
        linea = d or c or b
        if not linea:
            continue
        for col_idx, fecha in cols_fecha:
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None or not isinstance(v, (int, float)):
                continue
            rows.append({
                "fecha": fecha,
                "year": fecha.year,
                "month": fecha.month,
                "seccion": seccion_actual,
                "linea": linea,
                "valor_fcst": float(v),
            })
    df = pd.DataFrame(rows)
    if not df.empty:
        df["fecha"] = pd.to_datetime(df["fecha"])
    return df


def extract_dashboard_data(wb) -> pd.DataFrame:
    """Hoja Dashboard Data: KPIs mensuales pre-cocinados (Venta/Contrib/GAV/EBIT
    × Meta/Resultado/Var)."""
    if "Dashboard Data" not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb["Dashboard Data"]
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=48)):
        if not row or not row[0]:
            continue
        if i == 0:
            headers = list(row)
            continue
        if not headers:
            continue
        label = str(row[0]).strip()
        for j, v in enumerate(row[1:13]):  # 12 meses
            if v is None or not isinstance(v, (int, float)):
                continue
            mes = headers[j + 1] if j + 1 < len(headers) else None
            rows.append({
                "concepto": label,
                "mes_label": str(mes) if mes else "",
                "mes_num": j + 1,
                "valor": float(v),
            })
    return pd.DataFrame(rows)


def extract_analisis_financiero(wb) -> pd.DataFrame:
    """Análisis Financiero YTD Abr-26 (o último disponible)."""
    sheet = None
    for s in wb.sheetnames:
        if "lisis Financiero YTD" in s:
            sheet = s
            break
    if sheet is None and "Análisis Financiero 2026" in wb.sheetnames:
        sheet = "Análisis Financiero 2026"
    if sheet is None:
        for s in wb.sheetnames:
            if "lisis Financiero" in s and "2026" in s:
                sheet = s
                break
    if sheet is None:
        return pd.DataFrame()
    ws = wb[sheet]
    rows = []
    seccion_actual = ""
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=96)):
        a = (row[0] if row else "") or ""
        b = (row[1] if len(row) > 1 else "") or ""
        c = (row[2] if len(row) > 2 else "") or ""
        d = (row[3] if len(row) > 3 else "") or ""
        e = (row[4] if len(row) > 4 else "") or ""
        a = str(a).strip()
        if not a:
            continue
        # Detectar secciones (texto sin números en col B-E)
        if not any(isinstance(v, (int, float)) for v in (b, c, d, e)):
            seccion_actual = a
            continue
        rows.append({
            "seccion": seccion_actual,
            "concepto": a,
            "ytd_2025": b if isinstance(b, (int, float)) else None,
            "ytd_2026": c if isinstance(c, (int, float)) else None,
            "var_abs": d if isinstance(d, (int, float)) else None,
            "var_pct": e if isinstance(e, (int, float)) else None,
            "nota": row[6] if len(row) > 6 else "",
        })
    return pd.DataFrame(rows)


# ============================================================
# MAIN
# ============================================================
def main():
    print(f"=== Extract Finanzas Planificación — {datetime.now().isoformat()} ===\n", flush=True)
    print(f"Leyendo: {EXCEL_FILE}", flush=True)
    print(f"Última mod: {datetime.fromtimestamp(EXCEL_FILE.stat().st_mtime).isoformat()}\n", flush=True)

    wb = _abrir()

    extractors = [
        ("pyl_mensual", extract_pyl, "P&L histórico mensual (2019-2026)"),
        ("ppto_2026", extract_ppto_2026, "Ppto 2026 por código contable"),
        ("resumen_ytd", extract_resumen_ytd, "Resumen YTD vs Ppto vs YoY"),
        ("kt", extract_kt, "Capital de Trabajo"),
        ("deuda", extract_deuda, "Deuda financiera"),
        ("metas_2026", extract_metas_2026, "Metas 2026 mensuales"),
        ("fcst_eerr", extract_fcst_eerr, "Forecast EERR"),
        ("dashboard_data", extract_dashboard_data, "Dashboard Data pre-cocinada"),
        ("analisis_financiero", extract_analisis_financiero, "Análisis Financiero YTD"),
    ]

    resumen = {
        "generado_en": datetime.now().isoformat(),
        "archivo_origen": EXCEL_FILE.name,
        "archivo_modificado": datetime.fromtimestamp(EXCEL_FILE.stat().st_mtime).isoformat(),
        "hojas_procesadas": {},
    }

    for name, fn, desc in extractors:
        print(f"  → {name} ({desc})", flush=True)
        try:
            df = fn(wb)
            out = OUT_DIR / f"{name}.parquet"
            df.to_parquet(out, index=False)
            resumen["hojas_procesadas"][name] = {
                "filas": len(df),
                "columnas": list(df.columns),
                "archivo": str(out.relative_to(PROJECT_ROOT)),
                "ok": True,
            }
            print(f"     OK: {len(df):,} filas → {out.name}", flush=True)
        except Exception as e:
            print(f"     ERROR: {type(e).__name__}: {e}", flush=True)
            resumen["hojas_procesadas"][name] = {"ok": False, "error": str(e)}

    wb.close()

    # Resumen general
    resumen_path = OUT_DIR / "resumen_general.json"
    with open(resumen_path, "w", encoding="utf-8") as f:
        json.dump(resumen, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n→ Resumen: {resumen_path.relative_to(PROJECT_ROOT)}", flush=True)
    print(f"\nOK", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
