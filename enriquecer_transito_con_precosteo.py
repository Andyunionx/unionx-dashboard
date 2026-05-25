"""Enriquece data/comex/transito.parquet con los costos internados CLP
de los Pre-costeos locales (agente-comex/data/output/26TPXXXX/).

Para cada (PI, SKU) del parquet, busca el Excel Pre-costeo_x_CBM_<PI>.xlsx y
copia 'Costo Internado Total (CLP)' a la columna costo_ingreso_clp del parquet.

No toca el Sheet de Martín. Idempotente: puede correr N veces.

Uso:
    python enriquecer_transito_con_precosteo.py             # enriquece y guarda
    python enriquecer_transito_con_precosteo.py --dry-run   # solo reporta, no escribe
"""
import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')
PROJECT_ROOT = Path(__file__).parent
PARQUET = PROJECT_ROOT / 'data' / 'comex' / 'transito.parquet'
OUTPUT_DIR = PROJECT_ROOT / 'agente-comex' / 'data' / 'output'

COL_DESTINO = 'costo_ingreso_clp'
COL_FUENTE = 'fuente_costo'
HEADER_COSTO_TOTAL = 'Costo Internado Total (CLP)'
HEADER_COSTO_UNIT = 'Costo Internado Unit (CLP)'


def _leer_precosteo(pi_full: str) -> dict[str, dict]:
    """Lee Pre-costeo_x_CBM_<pi>.xlsx y retorna {sku: {total_clp, unit_clp}}.

    pi_full: 26TP0320PI → busca carpeta 26TP0320.
    """
    emb = pi_full.replace('PI', '').strip()
    excel = OUTPUT_DIR / emb / f'Pre-costeo_x_CBM_{emb}.xlsx'
    if not excel.exists():
        return {}

    wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
    if 'Productos' not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb['Productos']
    rows = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        wb.close()
        return {}

    try:
        idx_sku = headers.index('SKU')
        idx_total = headers.index(HEADER_COSTO_TOTAL)
        idx_unit = headers.index(HEADER_COSTO_UNIT)
    except ValueError as e:
        print(f"  [WARN] {excel.name}: header faltante: {e}")
        wb.close()
        return {}

    out: dict[str, dict] = {}
    for row in rows:
        if not row or row[idx_sku] is None:
            continue
        sku = str(row[idx_sku]).strip()
        if not sku:
            continue
        try:
            total = float(row[idx_total]) if row[idx_total] not in (None, '') else None
            unit = float(row[idx_unit]) if row[idx_unit] not in (None, '') else None
        except (ValueError, TypeError):
            continue
        out[sku] = {'total_clp': total, 'unit_clp': unit}
    wb.close()
    return out


def _to_num(s) -> float | None:
    """Limpia número estilo Chile/EU desde string del parquet."""
    if s is None or s == '' or (isinstance(s, float) and pd.isna(s)):
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip().replace('$', '').replace(' ', '').replace('\xa0', '')
    if not s:
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def enriquecer(dry_run: bool = False) -> dict:
    if not PARQUET.exists():
        print(f"[ERROR] No existe {PARQUET}. Corre extract_comex_transito.py primero.")
        return {'ok': False}

    df = pd.read_parquet(PARQUET)
    pis_disponibles = sorted(df['pi'].dropna().unique())
    print(f"[INFO] transito.parquet: {len(df)} filas, {len(pis_disponibles)} PIs distintos")

    # Asegurar columna destino numérica (la actual es object con strings vacíos)
    df[COL_DESTINO] = df[COL_DESTINO].apply(_to_num).astype('float64')
    if COL_FUENTE not in df.columns:
        df[COL_FUENTE] = pd.NA

    # Por PI: leer precosteo si existe
    estadisticas = {
        'pis_con_precosteo': [],
        'pis_sin_precosteo': [],
        'skus_enriquecidos': 0,
        'skus_sin_match_en_precosteo': 0,
        'skus_ya_tenian_costo': 0,
    }

    for pi in pis_disponibles:
        precosteo = _leer_precosteo(pi)
        mask_pi = df['pi'] == pi
        n_filas_pi = mask_pi.sum()

        if not precosteo:
            estadisticas['pis_sin_precosteo'].append(pi)
            print(f"  {pi}: sin precosteo local ({n_filas_pi} filas pendientes)")
            continue

        estadisticas['pis_con_precosteo'].append(pi)
        n_enriq = 0
        n_sin_match = 0
        n_ya_tenia = 0

        for idx in df[mask_pi].index:
            sku = str(df.at[idx, 'sku']).strip()
            costo_existente = df.at[idx, COL_DESTINO]
            fuente_existente = df.at[idx, COL_FUENTE] if COL_FUENTE in df.columns else None
            datos = precosteo.get(sku)
            if not datos or datos.get('total_clp') is None:
                n_sin_match += 1
                continue
            # Si el costo vino del Sheet Martín (fuente NaN), no tocar — es fuente externa
            # autoritativa. Si vino de un precosteo previo (puede tener flete viejo),
            # SÍ sobrescribir con el último precosteo en disco.
            costo_existente_real = pd.notna(costo_existente) and costo_existente > 0
            fuente_es_local = pd.notna(fuente_existente) and fuente_existente == 'precosteo_local'
            if costo_existente_real and not fuente_es_local:
                n_ya_tenia += 1
                continue
            df.at[idx, COL_DESTINO] = datos['total_clp']
            df.at[idx, COL_FUENTE] = 'precosteo_local'
            n_enriq += 1

        estadisticas['skus_enriquecidos'] += n_enriq
        estadisticas['skus_sin_match_en_precosteo'] += n_sin_match
        estadisticas['skus_ya_tenian_costo'] += n_ya_tenia
        print(f"  {pi}: precosteo OK | enriquecidos={n_enriq} | sin_match_sku={n_sin_match} | ya_tenia={n_ya_tenia} | total_filas={n_filas_pi}")

    print()
    print("=== RESUMEN ===")
    print(f"  PIs con precosteo:    {len(estadisticas['pis_con_precosteo'])} → {estadisticas['pis_con_precosteo']}")
    print(f"  PIs sin precosteo:    {len(estadisticas['pis_sin_precosteo'])} → {estadisticas['pis_sin_precosteo']}")
    print(f"  SKUs enriquecidos:    {estadisticas['skus_enriquecidos']}")
    print(f"  SKUs sin match:       {estadisticas['skus_sin_match_en_precosteo']}")
    print(f"  SKUs ya tenían costo: {estadisticas['skus_ya_tenian_costo']}")

    n_con_costo = df[COL_DESTINO].notna().sum()
    print(f"  Total filas con costo_ingreso_clp: {n_con_costo} / {len(df)} ({100*n_con_costo/len(df):.1f}%)")

    if dry_run:
        print("\n[DRY-RUN] no se escribió el parquet")
    else:
        df.to_parquet(PARQUET, index=False)
        print(f"\n[OK] parquet actualizado: {PARQUET}")

    estadisticas['ok'] = True
    return estadisticas


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true', help='Reporta sin escribir')
    args = parser.parse_args()
    enriquecer(dry_run=args.dry_run)


if __name__ == '__main__':
    main()
