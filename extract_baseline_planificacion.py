"""Extrae el baseline de Planificación desde FCST FINAL SKU 26-27 V2.xlsx.

Genera 3 datasets y los sube a Turso:
  - planif_master_sku            (universo de SKUs activos + cols A-K del Excel)
  - planif_stock_baseline_20260511 (foto STOCK al 11/05 10:00)
  - planif_transito_baseline     (foto BASE TRANSITOS)

Idempotente: DELETE+INSERT por snapshot_date.
"""
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')
PROJECT_ROOT = Path(__file__).parent
EXCEL_PATH = PROJECT_ROOT / 'data' / 'planificacion' / 'FORECAST FINAL SKU 26-27 V2.xlsx'
OUT_DIR = PROJECT_ROOT / 'data' / 'planificacion'

SNAPSHOT_DATE = '2026-05-11'  # 11/05 10:00 — foto baseline definida por user

env = PROJECT_ROOT / '.env'
for line in env.read_text(encoding='utf-8').splitlines():
    if '=' in line and not line.startswith('#'):
        k, v = line.split('=', 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


# ============================================================
# Lectura del Excel master (3 hojas)
# ============================================================
def leer_excel_baseline():
    """Lee las 3 hojas relevantes. read_only=True para soportar archivo abierto."""
    from openpyxl import load_workbook

    print(f"[1] Abriendo Excel (puede tardar 2-4 min, archivo 254MB)...")
    t0 = time.time()
    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    print(f"    Abierto en {time.time()-t0:.1f}s. Hojas disponibles: {len(wb.sheetnames)}")

    # ----- FCST BASE SKU MACRO (cols A-K = 11 cols, headers en fila 3) -----
    print(f"\n[2] Leyendo FCST BASE SKU MACRO cols A-K...")
    ws = wb['FCST BASE SKU MACRO']
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 3:
            headers = list(row[:11])  # solo cols A-K
        elif i > 3:
            # Skip filas vacías (SKU/ID nulos)
            if row[0] is None and row[4] is None:
                continue
            rows.append(list(row[:11]))
    df_master = pd.DataFrame(rows, columns=headers)
    print(f"    {len(df_master):,} SKUs en FCST BASE SKU MACRO. Headers: {headers}")

    # ----- STOCK (foto 11/05 10:00) -----
    print(f"\n[3] Leyendo STOCK...")
    ws = wb['STOCK']
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            headers = [h for h in row if h is not None]
            n_cols = len(headers)
        else:
            if row[0] is None:
                continue
            rows.append(list(row[:n_cols]))
    df_stock = pd.DataFrame(rows, columns=headers)
    print(f"    {len(df_stock):,} SKUs en STOCK. Headers: {headers}")

    # ----- BASE TRANSITOS -----
    print(f"\n[4] Leyendo BASE TRANSITOS...")
    ws = wb['BASE TRANSITOS']
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            headers = [h for h in row if h is not None]
            n_cols = len(headers)
        else:
            # Solo filas con SKU
            if row[0] is None:
                continue
            rows.append(list(row[:n_cols]))
    df_transito = pd.DataFrame(rows, columns=headers)
    print(f"    {len(df_transito):,} filas en BASE TRANSITOS. Headers: {headers}")

    wb.close()
    return df_master, df_stock, df_transito


# ============================================================
# Augment master con TODOS los SKUs (los activos del Excel + universo dim_productos)
# ============================================================
def augment_master_con_universo(df_master, exec_retry):
    """Trae lista de SKUs de Turso (ventas + stock) y agrega los que no están en Excel."""
    print(f"\n[5] Consultando universo de SKUs desde Turso...")
    rs = exec_retry(
        "SELECT DISTINCT sku FROM ventas WHERE sku IS NOT NULL AND sku != ''",
        label='[skus-ventas]'
    )
    skus_ventas = {str(r[0]).strip() for r in rs.rows}
    print(f"    Universo SKUs en ventas: {len(skus_ventas):,}")

    # Limpiar SKU en master
    df_master['Sku_str'] = df_master['Sku'].astype(str).str.strip()
    sku_excel = set(df_master['Sku_str'])

    nuevos = skus_ventas - sku_excel
    print(f"    SKUs en ventas no presentes en master Excel: {len(nuevos):,}")

    if nuevos:
        df_nuevos = pd.DataFrame({
            'ID': [None] * len(nuevos),
            'Marca': [None] * len(nuevos),
            'Categoria Padre': [None] * len(nuevos),
            'Categoria Hijo': [None] * len(nuevos),
            'Sku': list(nuevos),
            'Descripcion': [None] * len(nuevos),
            'TOTAL': [None] * len(nuevos),
            'Categoria Producto': [None] * len(nuevos),
            '% Proyeccion Vta': [None] * len(nuevos),
            'Ranking Comercial': [None] * len(nuevos),
            'Stock HOY': [None] * len(nuevos),
            'Sku_str': list(nuevos),
        })
        df_master = pd.concat([df_master, df_nuevos], ignore_index=True)
        print(f"    Master agregado: {len(df_master):,} SKUs totales")

    return df_master.drop(columns='Sku_str')


# ============================================================
# Turso: client + retry helper
# ============================================================
def _new_client():
    import libsql_client
    return libsql_client.create_client_sync(
        url=os.environ['LIBSQL_URL'],
        auth_token=os.environ['LIBSQL_AUTH_TOKEN'],
    )


def exec_retry(sql, args=None, max_retries=4, base_wait=5, label=''):
    last = None
    for a in range(1, max_retries + 1):
        c = None
        try:
            c = _new_client()
            rs = c.execute(sql, args) if args is not None else c.execute(sql)
            c.close()
            return rs
        except Exception as e:
            last = e
            if c:
                try: c.close()
                except: pass
            if a < max_retries:
                w = base_wait * (2 ** (a - 1))
                print(f"  {label} retry {a} ({type(e).__name__}) en {w}s...")
                time.sleep(w)
    raise last


# ============================================================
# Schema Turso (CREATE TABLE IF NOT EXISTS)
# ============================================================
def crear_tablas():
    print(f"\n[6] Creando tablas Turso si no existen...")
    for sql, label in [
        ("""CREATE TABLE IF NOT EXISTS planif_master_sku (
            sku TEXT PRIMARY KEY,
            id_categoria TEXT, marca TEXT,
            categoria_padre TEXT, categoria_hijo TEXT,
            descripcion TEXT, total TEXT,
            categoria_producto TEXT,
            pct_proyeccion_vta REAL, ranking_comercial REAL,
            stock_hoy REAL,
            ts_actualizado TEXT
        )""", "create-master"),
        ("""CREATE TABLE IF NOT EXISTS planif_stock_baseline (
            sku TEXT, snapshot_date TEXT, marca TEXT, producto TEXT,
            stock_total REAL, total_full REAL, bodega_principal REAL,
            full_meli REAL, full_fala REAL, full_paris REAL, full_ripley REAL,
            tiendas REAL, reserva REAL,
            transito_full_fala REAL, transito_full_meli REAL,
            costo REAL, valoracion REAL,
            ts_carga TEXT,
            PRIMARY KEY (sku, snapshot_date)
        )""", "create-stock"),
        ("""CREATE TABLE IF NOT EXISTS planif_transito_baseline (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            snapshot_date TEXT,
            sku TEXT, variante TEXT, pi TEXT, status TEXT,
            tipo_transporte TEXT, nro_pedido TEXT,
            cantidad REAL, costo_uni_usd REAL, gift_box_envio REAL,
            costo_ingreso_clp REAL,
            fecha_embarque TEXT, fecha_eta_chile TEXT, fecha_eta_bodega TEXT,
            mes REAL, stock_actual REAL,
            tipo_categoria TEXT, valor_usd_total REAL, marca TEXT,
            ts_carga TEXT
        )""", "create-transito"),
    ]:
        exec_retry(sql, label=f'  [{label}]')
        print(f"    OK {label}")


# ============================================================
# Inserts (idempotentes: DELETE por snapshot_date + INSERT)
# ============================================================
def _to_float(v):
    import math
    if v is None: return None
    if isinstance(v, float) and math.isnan(v): return None
    if isinstance(v, str):
        v = v.replace(',', '.').strip()
        if not v or v.lower() in ('nan', 'nat', 'none'): return None
        try:
            f = float(v)
            return None if math.isnan(f) else f
        except: return None
    try:
        f = float(v)
        return None if math.isnan(f) else f
    except: return None


def _to_str(v):
    """str() con NaN/NaT/None safety."""
    import math
    if v is None: return None
    if isinstance(v, float) and math.isnan(v): return None
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'nat', 'none'): return None
    return s


def upload_master(df, snapshot_date):
    print(f"\n[7] Upload master ({len(df):,} SKUs) → planif_master_sku...")
    exec_retry("DELETE FROM planif_master_sku", label='[del-master]')

    rows = []
    ts = datetime.now().isoformat(timespec='seconds')
    for _, r in df.iterrows():
        sku = _to_str(r.get('Sku'))
        if not sku:
            continue
        rows.append((
            sku,
            _to_str(r.get('ID')),
            _to_str(r.get('Marca')),
            _to_str(r.get('Categoria Padre')),
            _to_str(r.get('Categoria Hijo')),
            _to_str(r.get('Descripcion')),
            _to_str(r.get('TOTAL')),
            _to_str(r.get('Categoria Producto')),
            _to_float(r.get('% Proyeccion Vta')),
            _to_float(r.get('Ranking Comercial')),
            _to_float(r.get('Stock HOY')),
            ts,
        ))

    cols = ('sku, id_categoria, marca, categoria_padre, categoria_hijo, '
            'descripcion, total, categoria_producto, pct_proyeccion_vta, '
            'ranking_comercial, stock_hoy, ts_actualizado')
    placeholders = '(' + ','.join('?' * 12) + ')'
    inserted = 0
    failed = []
    batch = 50
    for i in range(0, len(rows), batch):
        chunk = rows[i:i + batch]
        sql_batch = f"INSERT INTO planif_master_sku ({cols}) VALUES " + ','.join([placeholders] * len(chunk))
        flat = [x for r in chunk for x in r]
        try:
            rs = exec_retry(sql_batch, flat, label=f'  [master b{i//batch+1}]', max_retries=2)
            inserted += rs.rows_affected
        except Exception as e:
            print(f"    Batch {i//batch+1} falló — caigo row-by-row para identificar...")
            for j, r in enumerate(chunk):
                try:
                    sql_one = f"INSERT INTO planif_master_sku ({cols}) VALUES " + placeholders
                    rs = exec_retry(sql_one, list(r), label=f'    [r{i+j}]', max_retries=1, base_wait=1)
                    inserted += rs.rows_affected
                except Exception as ee:
                    print(f"    ✗ Row {i+j} FAIL: sku={r[0]!r} types={[type(x).__name__ for x in r]}")
                    print(f"      Values: {r}")
                    print(f"      Err: {type(ee).__name__}: {str(ee)[:100]}")
                    failed.append((i+j, r))
    print(f"    Insertados: {inserted:,} / Fallidos: {len(failed)}")


def upload_stock(df, snapshot_date):
    print(f"\n[8] Upload stock baseline ({len(df):,} SKUs, snapshot={snapshot_date})...")
    exec_retry(
        "DELETE FROM planif_stock_baseline WHERE snapshot_date = ?",
        [snapshot_date], label='[del-stock]'
    )

    rows = []
    ts = datetime.now().isoformat(timespec='seconds')
    for _, r in df.iterrows():
        sku = _to_str(r.get('SKU'))
        if not sku:
            continue
        rows.append((
            sku, snapshot_date,
            _to_str(r.get('Marca')),
            _to_str(r.get('Producto')),
            _to_float(r.get('Stock total')),
            _to_float(r.get('Total Full')),
            _to_float(r.get('Bodega Principal')),
            _to_float(r.get('Full Meli')),
            _to_float(r.get('Full Fala')),
            _to_float(r.get('Full Paris')),
            _to_float(r.get('Full Ripley')),
            _to_float(r.get('Tiendas')),
            _to_float(r.get('Reserva')),
            _to_float(r.get('Tránsito Full Fala')),
            _to_float(r.get('Transito Full Meli')),
            _to_float(r.get('Costo')),
            _to_float(r.get('Valoracion')),
            ts,
        ))

    cols = ('sku, snapshot_date, marca, producto, stock_total, total_full, '
            'bodega_principal, full_meli, full_fala, full_paris, full_ripley, '
            'tiendas, reserva, transito_full_fala, transito_full_meli, '
            'costo, valoracion, ts_carga')
    placeholders = '(' + ','.join('?' * 18) + ')'
    inserted = 0
    batch = 100
    for i in range(0, len(rows), batch):
        chunk = rows[i:i + batch]
        sql = f"INSERT INTO planif_stock_baseline ({cols}) VALUES " + ','.join([placeholders] * len(chunk))
        flat = [x for r in chunk for x in r]
        rs = exec_retry(sql, flat, label=f'  [stock b{i//batch+1}]')
        inserted += rs.rows_affected
    print(f"    Insertados: {inserted:,}")


def upload_transito(df, snapshot_date):
    print(f"\n[9] Upload tránsito baseline ({len(df):,} filas, snapshot={snapshot_date})...")
    exec_retry(
        "DELETE FROM planif_transito_baseline WHERE snapshot_date = ?",
        [snapshot_date], label='[del-transito]'
    )

    rows = []
    ts = datetime.now().isoformat(timespec='seconds')

    def _date_str(v):
        if v is None: return None
        if hasattr(v, 'isoformat'):
            return v.isoformat()[:10]
        return str(v)[:10] if v else None

    for _, r in df.iterrows():
        sku = _to_str(r.get('SKU'))
        if not sku:
            continue
        rows.append((
            snapshot_date, sku,
            _to_str(r.get('Variante')),
            _to_str(r.get('PI')),
            _to_str(r.get('STATUS')),
            _to_str(r.get('Tipo de Transporte (Aéreo o Barco)')),
            _to_str(r.get('NRO PEDIDO')),
            _to_float(r.get('Cantidad')),
            _to_float(r.get('Costo Uni USD')),
            _to_float(r.get('GIFT BOX + Envio')),
            _to_float(r.get('COSTO INGRESO CLP CHILE')),
            _date_str(r.get('Fecha de Embarque')),
            _date_str(r.get('Fecha ETA CHILE')),
            _date_str(r.get('Fecha ETA bodega')),
            _to_float(r.get('MES')),
            _to_float(r.get('STOCK ACTUAL')),
            _to_str(r.get('Tipo Categoria')),
            _to_float(r.get('Valor USD TOTAL')),
            _to_str(r.get('MARCA')),
            ts,
        ))

    cols = ('snapshot_date, sku, variante, pi, status, tipo_transporte, nro_pedido, '
            'cantidad, costo_uni_usd, gift_box_envio, costo_ingreso_clp, '
            'fecha_embarque, fecha_eta_chile, fecha_eta_bodega, mes, stock_actual, '
            'tipo_categoria, valor_usd_total, marca, ts_carga')
    placeholders = '(' + ','.join('?' * 20) + ')'
    inserted = 0
    batch = 100
    for i in range(0, len(rows), batch):
        chunk = rows[i:i + batch]
        sql = f"INSERT INTO planif_transito_baseline ({cols}) VALUES " + ','.join([placeholders] * len(chunk))
        flat = [x for r in chunk for x in r]
        rs = exec_retry(sql, flat, label=f'  [transito b{i//batch+1}]')
        inserted += rs.rows_affected
    print(f"    Insertados: {inserted:,}")


# ============================================================
# Main
# ============================================================
def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument('--from-parquet', action='store_true',
                    help='Saltar lectura del Excel — usar parquets de data/planificacion/')
    args = ap.parse_args()

    print(f"=== BASELINE PLANIFICACIÓN — snapshot {SNAPSHOT_DATE} ===\n")

    if args.from_parquet:
        print("[1] Cargando desde parquets locales (skip Excel)...")
        df_master = pd.read_parquet(OUT_DIR / 'baseline_master_sku_excel.parquet')
        df_stock = pd.read_parquet(OUT_DIR / f'baseline_stock_{SNAPSHOT_DATE}.parquet')
        df_transito = pd.read_parquet(OUT_DIR / f'baseline_transito_{SNAPSHOT_DATE}.parquet')
        print(f"    Master: {len(df_master):,} | Stock: {len(df_stock):,} | Tránsito: {len(df_transito):,}")
    else:
        if not EXCEL_PATH.exists():
            print(f"[ERROR] No existe: {EXCEL_PATH}")
            return 1
        df_master, df_stock, df_transito = leer_excel_baseline()

    if not args.from_parquet:
        # Normalizar tipos mixtos (int+str) a str para evitar pyarrow errors
        for df in (df_master, df_stock, df_transito):
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].astype('object').where(df[col].notna(), None)
                    df[col] = df[col].apply(lambda v: str(v) if v is not None and not isinstance(v, (int, float, bool)) else v)
                    tipos = {type(v).__name__ for v in df[col].dropna()}
                    if len(tipos) > 1:
                        df[col] = df[col].apply(lambda v: str(v) if v is not None else None)

        # Backup parquets
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        df_master.to_parquet(OUT_DIR / f'baseline_master_sku_excel.parquet', index=False)
        df_stock.to_parquet(OUT_DIR / f'baseline_stock_{SNAPSHOT_DATE}.parquet', index=False)
        df_transito.to_parquet(OUT_DIR / f'baseline_transito_{SNAPSHOT_DATE}.parquet', index=False)
        print(f"\n[backup] Parquets locales generados en {OUT_DIR}")

    # Augment master con todos los SKUs (Excel + ventas)
    df_master = augment_master_con_universo(df_master, exec_retry)
    df_master.to_parquet(OUT_DIR / 'baseline_master_sku_full.parquet', index=False)

    # Subir a Turso
    crear_tablas()
    upload_master(df_master, SNAPSHOT_DATE)
    upload_stock(df_stock, SNAPSHOT_DATE)
    upload_transito(df_transito, SNAPSHOT_DATE)

    # Validación
    print(f"\n=== VERIFICACIÓN ===")
    for tbl, label in [
        ('planif_master_sku', 'Master SKU'),
        ('planif_stock_baseline', f'Stock @ {SNAPSHOT_DATE}'),
        ('planif_transito_baseline', f'Tránsito @ {SNAPSHOT_DATE}'),
    ]:
        cond = f"WHERE snapshot_date = '{SNAPSHOT_DATE}'" if 'baseline' in tbl else ''
        rs = exec_retry(f"SELECT COUNT(*) FROM {tbl} {cond}", label=f'[ver-{tbl}]')
        print(f"  {label}: {rs.rows[0][0]:,} filas en Turso")

    print(f"\n[OK] Baseline cargado.")
    return 0


if __name__ == '__main__':
    sys.exit(main())
