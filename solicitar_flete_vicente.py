"""Solicita cotización de flete marítimo a Vicente Seimex.

Genera UN mail consolidado HTML con tabla de N embarques + adjunta los PLs.
Reemplaza el paso manual de redactar drafts en Gmail.

Formato (mismo que usa Andrés manualmente, validado):
  - Tabla: Embarque | Puerto Origen | Container | Volumen estimado | ETA tentativa Chile
  - Adjuntos: PLs de cada embarque
  - CTA: "antes del lunes <fecha>" para incorporar al pre-costeo
  - To: vicente@seimex.cl  |  CC: operaciones@unionx.cl
  - SIN línea de "cualquier gasto adicional"

Uso:
    python solicitar_flete_vicente.py 0330 0422 0423 0424
    python solicitar_flete_vicente.py 0330 --dry-run    # genera HTML, no envía
    python solicitar_flete_vicente.py --auto            # detecta PIs con flete placeholder
"""
import argparse
import base64
import re
import sys
from datetime import date, datetime, timedelta
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
PROJECT_ROOT = Path(__file__).parent
EMB_DIR = PROJECT_ROOT / 'data' / 'comex' / 'embarques'

DESTINATARIO = 'vicente@seimex.cl'
CC = 'operaciones@unionx.cl'

PUERTO_NOMBRE = {
    'SZ': 'SZ Shenzhen',
    'NB': 'NB Ningbo',
    'XI': 'XI Xiamen',
    'AIR': 'AIR (DHL)',
}
# Días de tránsito típicos puerto → bodega Chile (heurística para ETA tentativa)
TRANSITO_DIAS = {'SZ': 40, 'NB': 42, 'XI': 45, 'AIR': 7}


def _detectar_puerto(pi_filename: str) -> str:
    name = pi_filename.upper()
    for code in ('SZ', 'NB', 'XI', 'AIR', 'DHL'):
        if re.search(rf'(?<![A-Z]){code}(?![A-Z])', name):
            return 'AIR' if code == 'DHL' else code
    return 'SZ'


def _leer_cbm_total(pl_path: Path) -> float:
    """Lee CBM total del Packing List.

    Prefiere la fila TOTAL del PL (Model vacío + CBM populado). Si no existe,
    suma todas las filas de data.
    """
    wb = openpyxl.load_workbook(pl_path, data_only=True, read_only=True)
    ws = wb.active
    idx_cbm = None
    idx_model = None
    total_row_val = None
    suma_data = 0.0
    in_data = False
    for row in ws.iter_rows(values_only=True):
        if idx_cbm is None:
            for j, v in enumerate(row or ()):
                if v and 'total' in str(v).lower() and 'cbm' in str(v).lower():
                    idx_cbm = j
                if v and str(v).strip().lower() == 'model':
                    idx_model = j
            if idx_cbm is not None:
                in_data = True
            continue
        if not row or idx_cbm >= len(row):
            continue
        cbm = row[idx_cbm]
        if not isinstance(cbm, (int, float)):
            continue
        model = row[idx_model] if idx_model is not None and idx_model < len(row) else None
        if model is None or str(model).strip().lower() in ('', 'total'):
            # Fila TOTAL: usar directo y salir
            total_row_val = float(cbm)
            break
        suma_data += float(cbm)
    wb.close()
    return total_row_val if total_row_val is not None else suma_data


def _leer_eta_tarifa(emb: str) -> date | None:
    """Lee ETA estimada de la tarifa si existe."""
    p = EMB_DIR / f'Tarifas_Base_COMEX-{emb}.xlsx'
    if not p.exists():
        return None
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    ws = wb.active
    eta = None
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        for i, v in enumerate(row):
            if v and 'eta' in str(v).lower() and i + 1 < len(row):
                cand = row[i + 1]
                if isinstance(cand, datetime):
                    eta = cand.date()
                elif isinstance(cand, date):
                    eta = cand
                break
        if eta:
            break
    wb.close()
    return eta


def _proximo_lunes() -> date:
    hoy = date.today()
    return hoy + timedelta(days=(7 - hoy.weekday()) % 7 or 7)


def _detectar_pi_pl(emb: str) -> tuple[Path, Path]:
    base = EMB_DIR / emb
    if not base.exists():
        raise FileNotFoundError(f'No existe {base}')
    pl = next((f for f in base.iterdir() if 'PL' in f.name.upper() and f.name.endswith('.xlsx')), None)
    pi = next((f for f in base.iterdir() if 'PL' not in f.name.upper() and f.name.endswith('.xlsx')), None)
    if not pi or not pl:
        raise FileNotFoundError(f'PI o PL faltante en {base}')
    return pi, pl


def _detectar_container(pi_filename: str) -> str:
    name = pi_filename.upper()
    for code in ('40HQ', '40HC', '20GP', '20STD', 'LCL', 'AIR', 'DHL'):
        if code in name:
            return 'AIR (DHL)' if code in ('AIR', 'DHL') else code
    return '40HQ'


def _datos_embarque(emb: str) -> dict:
    pi, pl = _detectar_pi_pl(emb)
    puerto = _detectar_puerto(pi.name)
    container = _detectar_container(pi.name)
    cbm = _leer_cbm_total(pl)
    eta = _leer_eta_tarifa(emb) or (date.today() + timedelta(days=TRANSITO_DIAS.get(puerto, 40)))
    return {
        'emb': f'26TP{emb}',
        'puerto': PUERTO_NOMBRE.get(puerto, puerto),
        'container': container,
        'cbm': cbm,
        'eta': eta,
        'pi': pi,
        'pl': pl,
    }


def _generar_html(embarques: list[dict], deadline: date) -> str:
    filas = ''
    for d in embarques:
        filas += (
            f'<tr>'
            f'<td style="border:1px solid #888;padding:6px 10px">{d["emb"]}</td>'
            f'<td style="border:1px solid #888;padding:6px 10px">{d["puerto"]}</td>'
            f'<td style="border:1px solid #888;padding:6px 10px">{d["container"]}</td>'
            f'<td style="border:1px solid #888;padding:6px 10px">~{d["cbm"]:.0f} CBM</td>'
            f'<td style="border:1px solid #888;padding:6px 10px">{d["eta"].strftime("%Y-%m-%d")}</td>'
            f'</tr>'
        )
    fecha_es = deadline.strftime('%d-%b').lower().replace('.', '')
    n = len(embarques)
    plural = 'embarques' if n > 1 else 'embarque'
    intro_n = f'{n} embarques' if n > 1 else 'un embarque'
    adj_plural = 'los Packing List de los embarques' if n > 1 else 'el Packing List del embarque'

    return f"""<div style="font-family:Arial,sans-serif;font-size:14px;color:#111">
<p>Hola Vicente,</p>
<p>Te paso {intro_n} que tenemos en pipeline desde Shenzhen Topwill Electronic. Necesitamos tu cotización de flete marítimo para cada uno:</p>
<table style="border-collapse:collapse;font-size:13px">
<thead>
<tr style="background:#1F4E78;color:#fff">
<th style="border:1px solid #888;padding:6px 10px">Embarque</th>
<th style="border:1px solid #888;padding:6px 10px">Puerto Origen</th>
<th style="border:1px solid #888;padding:6px 10px">Container</th>
<th style="border:1px solid #888;padding:6px 10px">Volumen estimado</th>
<th style="border:1px solid #888;padding:6px 10px">ETA tentativa Chile</th>
</tr>
</thead>
<tbody>
{filas}
</tbody>
</table>
<p>Te adjunto {adj_plural} para que dimensiones el booking. Idealmente nos avisás antes del lunes {fecha_es} para incorporar los costos al pre-costeo y RFQ en Odoo.</p>
<p>Cualquier consulta, quedo atento.</p>
<p>Saludos,<br>
Andrés Browne<br>
Gerente Finanzas + Supply Chain<br>
UnionX</p>
</div>"""


def _enviar(asunto: str, html: str, pls: list[Path]) -> str:
    sys.path.insert(0, str(PROJECT_ROOT / 'agente-comex' / 'src'))
    from gmail_client import GmailClient

    msg = MIMEMultipart('mixed')
    msg['to'] = DESTINATARIO
    msg['cc'] = CC
    msg['subject'] = asunto

    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText('Análisis adjunto.', 'plain', 'utf-8'))
    alt.attach(MIMEText(html, 'html', 'utf-8'))
    msg.attach(alt)

    for pl in pls:
        with open(pl, 'rb') as f:
            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{pl.name}"')
        msg.attach(part)

    g = GmailClient()
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    sent = g.service.users().messages().send(userId='me', body={'raw': raw}).execute()
    return sent['id']


def _auto_detectar_pendientes() -> list[str]:
    """Devuelve los embarques con tarifa placeholder (flete = 2800 USD, valor por defecto)."""
    pendientes = []
    for tarifa in EMB_DIR.glob('Tarifas_Base_COMEX-*.xlsx'):
        m = re.search(r'COMEX-(\d{4})', tarifa.name)
        if not m:
            continue
        emb = m.group(1)
        # Detectar flete: si es 2800 (placeholder) asumimos pendiente
        wb = openpyxl.load_workbook(tarifa, data_only=True, read_only=True)
        ws = wb.active
        flete = None
        for row in ws.iter_rows(values_only=True):
            for i, v in enumerate(row or ()):
                if v and 'flete total' in str(v).lower() and i + 1 < len(row):
                    flete = row[i + 1]
                    break
            if flete is not None:
                break
        wb.close()
        if flete == 2800:
            pendientes.append(emb)
    return sorted(pendientes)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('embarques', nargs='*', help='Códigos de embarque (ej 0330 0422)')
    parser.add_argument('--auto', action='store_true', help='Auto-detectar pendientes (flete = 2800)')
    parser.add_argument('--dry-run', action='store_true', help='Generar HTML, no enviar')
    args = parser.parse_args()

    if args.auto:
        embs = _auto_detectar_pendientes()
        if not embs:
            print('[i] Sin embarques con flete pendiente (placeholder 2800).')
            return 0
        print(f'[auto] Detectados pendientes: {embs}')
    else:
        embs = [e.zfill(4) for e in args.embarques]
        if not embs:
            parser.error('Indica al menos un embarque (o usa --auto)')

    datos = []
    for emb in embs:
        d = _datos_embarque(emb)
        datos.append(d)
        print(f'  {d["emb"]}: {d["puerto"]} | {d["container"]} | ~{d["cbm"]:.1f} CBM | ETA {d["eta"]}')

    deadline = _proximo_lunes()
    asunto_embs = ' + '.join(d['emb'].replace('26TP', '') for d in datos)
    asunto = f'Cotización Flete Marítimo - Embarques 26TP {asunto_embs} (40HQ)'
    if len(datos) == 1:
        asunto = f'Cotización Flete Marítimo - Embarque {datos[0]["emb"]} ({datos[0]["puerto"]} → Valparaíso) {datos[0]["container"]}'

    html = _generar_html(datos, deadline)

    if args.dry_run:
        preview_path = PROJECT_ROOT / 'agente-comex' / 'data' / 'output' / 'preview_mail_vicente.html'
        preview_path.parent.mkdir(parents=True, exist_ok=True)
        preview_path.write_text(html, encoding='utf-8')
        print(f'\n[DRY-RUN] HTML guardado en {preview_path}')
        print(f'  Asunto: {asunto}')
        print(f'  To: {DESTINATARIO} | Cc: {CC}')
        print(f'  Adjuntos: {[d["pl"].name for d in datos]}')
        return 0

    pls = [d['pl'] for d in datos]
    msg_id = _enviar(asunto, html, pls)
    print(f'\n[OK] Mail enviado | message_id={msg_id} | to={DESTINATARIO} | cc={CC}')
    print(f'  Asunto: {asunto}')
    return 0


if __name__ == '__main__':
    sys.exit(main() or 0)
