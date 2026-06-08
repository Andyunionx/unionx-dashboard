"""Genera Excel "Resumen TD" replicando la pestaña del archivo
"Reporte Ventas Empresa 2026 VS 2025 RAW.xlsx".

Salida: data/outputs/Reporte Ventas Empresa 2026 vs 2025 - YYYY-MM-DD.xlsx

Logica validada:
  - Fuente: ventas_historico.parquet + ventas_mes_actual.parquet
  - Agrupa por (año, mes) sin filtrar tipo_movimiento
  - Vta $ y Cantidad: suma directa
  - %MF y %MFin: dividen sobre venta_NETA (no bruta)
  - Filtros aplicados estaticos: Proveedor/Bodega/Cat Comercial/Tipo Marca = (Todas)
"""
from pathlib import Path
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent
ANIO_TY = 2026
ANIO_LY = 2025


def cargar_data():
    hist = pd.read_parquet(PROJECT_ROOT / 'data' / 'historico' / 'ventas_historico.parquet')
    mes = pd.read_parquet(PROJECT_ROOT / 'data' / 'historico' / 'ventas_mes_actual.parquet')
    cols = [c for c in mes.columns if c in hist.columns]
    df = pd.concat([hist[cols], mes], ignore_index=True)
    df['fv'] = pd.to_datetime(df['fecha_venta'], errors='coerce')
    df['anio'] = df['fv'].dt.year
    df['mes'] = df['fv'].dt.month
    df = df[df['anio'].isin([ANIO_LY, ANIO_TY])].copy()
    for c in ['cantidad', 'venta_bruta', 'venta_neta', 'margen_front',
              'comision', 'logistica', 'marketing']:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    return df


def calcular_resumen(df: pd.DataFrame) -> dict:
    agg = df.groupby(['anio', 'mes']).agg(
        qty=('cantidad', 'sum'),
        vb=('venta_bruta', 'sum'),
        vn=('venta_neta', 'sum'),
        mf=('margen_front', 'sum'),
        com=('comision', 'sum'),
        log=('logistica', 'sum'),
        mkt=('marketing', 'sum'),
    ).reset_index()
    agg['mfin'] = agg['mf'] - agg['com'] - agg['log'] - agg['mkt']

    out = {m: {} for m in range(1, 13)}
    for _, r in agg.iterrows():
        suf = 'ty' if int(r['anio']) == ANIO_TY else 'ly'
        m = int(r['mes'])
        out[m][f'qty_{suf}'] = float(r['qty'])
        out[m][f'vta_{suf}'] = float(r['vb'])
        out[m][f'vneta_{suf}'] = float(r['vn'])
        out[m][f'mf_{suf}'] = float(r['mf'])
        out[m][f'mfin_{suf}'] = float(r['mfin'])

    # Totales TY/LY (para share)
    tot_ty = sum(out[m].get('vta_ty', 0) for m in out)
    tot_ly = sum(out[m].get('vta_ly', 0) for m in out)
    return {'meses': out, 'tot_ty': tot_ty, 'tot_ly': tot_ly}


def _fmt_num(v):
    return v if v else 0


def generar_excel(resumen: dict, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen TD'

    GRAY = PatternFill('solid', fgColor='D9D9D9')
    HEADER_FILL = PatternFill('solid', fgColor='305496')
    HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    LABEL_FONT = Font(name='Calibri', size=10, bold=True)
    NORMAL_FONT = Font(name='Calibri', size=10)
    TOTAL_FILL = PatternFill('solid', fgColor='F2F2F2')
    TOTAL_FONT = Font(name='Calibri', size=10, bold=True)
    THIN = Side(style='thin', color='B7B7B7')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Filtros (filas 1-6 imitan layout original)
    filtros = [
        ('Proveedor', '(Todas)'),
        ('Bodega', '(Todas)'),
        ('Categoría comercial', '(Todas)'),
        ('Tipo Marca', '(Todas)'),
        ('Fecha Compara', '(Todas)'),
        ('Canal', '(Todas)'),
    ]
    for i, (k, v) in enumerate(filtros, start=1):
        ws.cell(row=i, column=2, value=k).font = LABEL_FONT
        ws.cell(row=i, column=3, value=v).font = NORMAL_FONT

    # Titulo "Valores" sobre la tabla
    ws.cell(row=8, column=6, value='Valores').font = LABEL_FONT

    # Headers fila 9 (idem a la dinamica)
    headers = [
        'Mes',
        ' Venta U TY', ' Venta U LY',
        'Vta $ TY', ' Vta $ LY',
        ' Dif Vta $', ' Crec Vta %',
        ' Share $ TY', ' Share $ LY',
        ' Ticket Prom TY', ' Ticket Prom LY',
        ' Marg Front TY', ' Margen Final TY',
        ' Marg Front LY', ' Margen Final LY',
    ]
    HEAD_ROW = 9
    for j, h in enumerate(headers, start=2):
        c = ws.cell(row=HEAD_ROW, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[HEAD_ROW].height = 32

    # Cuerpo: 12 meses + total
    meses_data = resumen['meses']
    tot_ty = resumen['tot_ty']
    tot_ly = resumen['tot_ly']

    fmt_int = '#,##0'
    fmt_money = '#,##0'
    fmt_pct = '0.00%'

    cur_row = HEAD_ROW + 1
    sum_qty_ty = sum_qty_ly = 0
    sum_mf_ty = sum_mf_ly = sum_mfin_ty = sum_mfin_ly = 0
    sum_vn_ty = sum_vn_ly = 0

    for mes in range(1, 13):
        d = meses_data.get(mes, {})
        qty_ty = d.get('qty_ty', 0)
        qty_ly = d.get('qty_ly', 0)
        vta_ty = d.get('vta_ty', 0)
        vta_ly = d.get('vta_ly', 0)
        vn_ty = d.get('vneta_ty', 0)
        vn_ly = d.get('vneta_ly', 0)
        mf_ty = d.get('mf_ty', 0)
        mf_ly = d.get('mf_ly', 0)
        mfin_ty = d.get('mfin_ty', 0)
        mfin_ly = d.get('mfin_ly', 0)

        sum_qty_ty += qty_ty; sum_qty_ly += qty_ly
        sum_mf_ty += mf_ty;   sum_mf_ly += mf_ly
        sum_mfin_ty += mfin_ty; sum_mfin_ly += mfin_ly
        sum_vn_ty += vn_ty;   sum_vn_ly += vn_ly

        dif = vta_ty - vta_ly
        crec = (vta_ty / vta_ly - 1) if vta_ly else 0
        share_ty = (vta_ty / tot_ty) if tot_ty else 0
        share_ly = (vta_ly / tot_ly) if tot_ly else 0
        ticket_ty = (vta_ty / qty_ty) if qty_ty else 0
        ticket_ly = (vta_ly / qty_ly) if qty_ly else 0
        pmf_ty = (mf_ty / vn_ty) if vn_ty else 0
        pmfin_ty = (mfin_ty / vn_ty) if vn_ty else 0
        pmf_ly = (mf_ly / vn_ly) if vn_ly else 0
        pmfin_ly = (mfin_ly / vn_ly) if vn_ly else 0

        valores = [
            mes,
            qty_ty, qty_ly,
            vta_ty, vta_ly,
            dif, crec,
            share_ty, share_ly,
            ticket_ty, ticket_ly,
            pmf_ty, pmfin_ty,
            pmf_ly, pmfin_ly,
        ]
        formatos = [
            '0',
            fmt_int, fmt_int,
            fmt_money, fmt_money,
            fmt_money, fmt_pct,
            fmt_pct, fmt_pct,
            fmt_money, fmt_money,
            fmt_pct, fmt_pct,
            fmt_pct, fmt_pct,
        ]
        for j, (val, fmt) in enumerate(zip(valores, formatos), start=2):
            c = ws.cell(row=cur_row, column=j, value=val)
            c.number_format = fmt
            c.font = NORMAL_FONT
            c.border = BORDER
            c.alignment = Alignment(horizontal='right')
        cur_row += 1

    # Total general
    tot_dif = tot_ty - tot_ly
    tot_crec = (tot_ty / tot_ly - 1) if tot_ly else 0
    tot_share_ty = 1.0 if tot_ty else 0
    tot_share_ly = 1.0 if tot_ly else 0
    tot_ticket_ty = (tot_ty / sum_qty_ty) if sum_qty_ty else 0
    tot_ticket_ly = (tot_ly / sum_qty_ly) if sum_qty_ly else 0
    tot_pmf_ty = (sum_mf_ty / sum_vn_ty) if sum_vn_ty else 0
    tot_pmfin_ty = (sum_mfin_ty / sum_vn_ty) if sum_vn_ty else 0
    tot_pmf_ly = (sum_mf_ly / sum_vn_ly) if sum_vn_ly else 0
    tot_pmfin_ly = (sum_mfin_ly / sum_vn_ly) if sum_vn_ly else 0

    tot_vals = [
        'Total general',
        sum_qty_ty, sum_qty_ly,
        tot_ty, tot_ly,
        tot_dif, tot_crec,
        tot_share_ty, tot_share_ly,
        tot_ticket_ty, tot_ticket_ly,
        tot_pmf_ty, tot_pmfin_ty,
        tot_pmf_ly, tot_pmfin_ly,
    ]
    tot_fmts = ['@', fmt_int, fmt_int, fmt_money, fmt_money,
                fmt_money, fmt_pct, fmt_pct, fmt_pct,
                fmt_money, fmt_money, fmt_pct, fmt_pct, fmt_pct, fmt_pct]
    for j, (val, fmt) in enumerate(zip(tot_vals, tot_fmts), start=2):
        c = ws.cell(row=cur_row, column=j, value=val)
        c.number_format = fmt
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal='right' if j > 2 else 'left')

    # Ancho columnas
    widths = [3, 16, 12, 12, 16, 16, 16, 12, 12, 12, 16, 16, 14, 16, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes en headers
    ws.freeze_panes = 'C10'

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f'[OK] {out_path} ({out_path.stat().st_size/1024:.1f} KB)')


def main():
    df = cargar_data()
    print(f'[info] {len(df):,} filas leidas 2025-2026')
    resumen = calcular_resumen(df)
    fecha = date.today().isoformat()
    out = PROJECT_ROOT / 'data' / 'outputs' / f'Reporte Ventas Empresa 2026 vs 2025 - {fecha}.xlsx'
    generar_excel(resumen, out)

    # Imprimir snapshot
    print()
    print(f'{"Mes":>4} | {"Vta TY":>10} | {"Vta LY":>10} | {"Crec":>6} | {"%MF TY":>7}')
    for mes in range(1, 13):
        d = resumen['meses'].get(mes, {})
        vty = d.get('vta_ty', 0); vly = d.get('vta_ly', 0)
        vn = d.get('vneta_ty', 0); mf = d.get('mf_ty', 0)
        crec = (vty / vly - 1) * 100 if vly else 0
        pmf = (mf / vn * 100) if vn else 0
        print(f'{mes:>4} | ${vty/1e6:>8,.1f}M | ${vly/1e6:>8,.1f}M | {crec:>+5.1f}% | {pmf:>6.2f}%')
    print(f'{"TOT":>4} | ${resumen["tot_ty"]/1e6:>8,.1f}M | ${resumen["tot_ly"]/1e6:>8,.1f}M')


if __name__ == '__main__':
    main()
