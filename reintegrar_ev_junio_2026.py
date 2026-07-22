# -*- coding: utf-8 -*-
"""
Re-integra El Volcán JUNIO 2026 al histórico, cuadrado con el cierre de Trinidad.

Modelo El Volcán (consignación, retiene 30% de comisión). Por cada línea de la
hoja "Detalle Ventas" del archivo de facturación de Trinidad:

    venta_neta   = (Costo Total − NC) / 0,70     # (Costo Total−NC) = 70% de la venta
    venta_bruta  = venta_neta × 1,19
    comision     = venta_neta × 0,30              # comisión de VENTA, por SKU
    costo_unit   = standard_price de Odoo (por SKU)
    costo_total  = costo_unit × unidades
    margen_front = venta_neta − costo_total
    margen_final = margen_front − comision

Totales objetivo (cierre Trinidad): venta_neta $14.155.228 · comisión $4.246.568 ·
costo Odoo ~$5.679.509 · 516 uds.

Reemplaza SOLO las filas canal='El Volcan' de junio 2026 (deja intacto 2025 y el
resto del histórico). Uso:  python reintegrar_ev_junio_2026.py [--apply]
"""
import argparse
import os
import shutil
import sys
import xmlrpc.client
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent
HIST = ROOT / "data" / "historico" / "ventas_historico.parquet"
FILE = ROOT / "data" / "manuales" / "Facturacion Simplit Junio 2026 (Trinidad).xlsx"
COMISION_PCT = 30.0
IVA = 1.19

ODOO = dict(url="https://unionxb2b.odoo.com", db="bmya-innovatek-sh-prd-6981800",
            user="andres@grupoeter.cl")


def _load_env():
    for p in (".env", "eerr-finanzas/.env"):
        f = ROOT / p
        if f.exists():
            for ln in f.read_text(encoding="utf-8").splitlines():
                if "=" in ln and not ln.strip().startswith("#"):
                    k, v = ln.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def costos_odoo(skus):
    """{sku: standard_price} buscando por default_code y barcode."""
    _load_env()
    pwd = os.environ.get("ANDRES_ODOO_PASSWORD")
    if not pwd:
        sys.exit("[ERROR] falta ANDRES_ODOO_PASSWORD")
    uid = xmlrpc.client.ServerProxy(f"{ODOO['url']}/xmlrpc/2/common").authenticate(
        ODOO["db"], ODOO["user"], pwd, {})
    models = xmlrpc.client.ServerProxy(f"{ODOO['url']}/xmlrpc/2/object")
    recs = models.execute_kw(ODOO["db"], uid, pwd, "product.product", "search_read",
        [["|", ["default_code", "in", skus], ["barcode", "in", skus]]],
        {"fields": ["default_code", "barcode", "standard_price"]})
    out = {}
    for r in recs:
        for key in (r.get("default_code"), r.get("barcode")):
            if key and str(key) in skus and str(key) not in out:
                out[str(key)] = r["standard_price"]
    return out


def leer_detalle():
    """Devuelve las líneas de la hoja 'Detalle Ventas' (una por venta)."""
    wb = openpyxl.load_workbook(FILE, data_only=True)
    ws = wb["Detalle Ventas"]
    hdr = [c.value for c in ws[1]]
    ix = {h: i for i, h in enumerate(hdr)}
    filas = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in r):
            continue
        sku = r[ix["SKU SELLER"]]
        fecha = r[ix["Fecha"]]
        if sku is None or fecha is None:       # salta la fila "Total general"
            continue
        filas.append(dict(
            sku=str(sku).strip(), fecha=pd.to_datetime(fecha),
            producto=r[ix["Producto"]], marca=r[ix["Marca"]],
            unidades=float(r[ix["Unidades"]] or 0),
            costo_total_ev=float(r[ix["Costo Total"]] or 0),
            nc=float(r[ix["NC"]] or 0),
        ))
    return filas


DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


def construir_filas(detalle, costos, attrs_por_sku, plantilla):
    filas = []
    for d in detalle:
        neta = (d["costo_total_ev"] - d["nc"]) / 0.70
        bruta = neta * IVA
        com = neta * COMISION_PCT / 100.0
        cu = float(costos.get(d["sku"], 0.0))
        ct = cu * d["unidades"]
        f = dict(plantilla)  # atributos estructurales constantes
        f.update(attrs_por_sku.get(d["sku"], {}))  # producto/categorías/marca por SKU
        f.update(dict(
            sku=d["sku"], fecha_venta=d["fecha"].strftime("%Y-%m-%d"),
            anio_venta=2026, mes_venta=6,
            semana_venta=int(d["fecha"].isocalendar().week),
            dia_semana=DIAS[d["fecha"].weekday()],
            cantidad=d["unidades"], venta_bruta=bruta, venta_neta=neta,
            costo_unitario=cu, costo_total=ct,
            comision_pct=COMISION_PCT, comision=com, logistica=0.0, marketing=0.0,
            margen_front=neta - ct, margen_final=(neta - ct) - com,
        ))
        filas.append(f)
    return pd.DataFrame(filas)


def main(apply=False):
    df = pd.read_parquet(HIST)
    fv = pd.to_datetime(df["fecha_venta"], errors="coerce")
    es_ev = df["canal"].astype(str).str.strip().str.lower().isin(["el volcan", "el volcán"])
    mask26 = es_ev & (fv.dt.year == 2026) & (fv.dt.month == 6)
    ev26 = df[mask26]
    print(f"El Volcán junio 2026 actual: {len(ev26)} filas · "
          f"venta_neta ${ev26['venta_neta'].sum():,.0f} · comisión ${ev26['comision'].sum():,.0f}")

    detalle = leer_detalle()
    print(f"Detalle Trinidad: {len(detalle)} líneas · {sum(d['unidades'] for d in detalle):.0f} uds")

    # atributos descriptivos por SKU tomados de las filas actuales (ya pisados por Matriz)
    attr_cols = ["producto", "categoria_macro", "categoria_padre", "categoria_hijo",
                 "categoria_comercial", "estado_sku", "pack", "marca", "proveedor",
                 "tipo_marca", "hora_venta", "hora_venta_num"]
    attrs = {}
    for _, r in ev26.iterrows():
        attrs.setdefault(str(r["sku"]), {c: r[c] for c in attr_cols if c in ev26.columns})

    # plantilla estructural (constante para el canal)
    plantilla = {c: "" for c in df.columns}
    plantilla.update(dict(
        tipo_movimiento="Venta", bodega="Bodega El Volcán", documento="",
        fecha_documento="", pedido="", estado_pedido="Ok", tipo_despacho="Seller",
        canal="El Volcan", proveedor="UnionX", tipo_compra="Importación",
        tipo_negocio="Marketplace", kam="Trini", estado_canal="In",
        tipo_marca="Propia", es_despacho=False, pedido_marketplace="", yuju_pack_id="",
    ))

    skus = list({d["sku"] for d in detalle})
    costos = costos_odoo(skus)
    faltan = [s for s in skus if s not in costos]
    if faltan:
        print(f"[WARN] {len(faltan)} SKUs sin costo en Odoo: {faltan[:5]}")

    nuevas = construir_filas(detalle, costos, attrs, plantilla)
    nuevas = nuevas[df.columns]  # mismo orden de columnas

    print("\n=== NUEVO El Volcán junio 2026 ===")
    for c in ["cantidad", "venta_bruta", "venta_neta", "costo_total", "comision",
              "margen_front", "margen_final"]:
        print(f"  Σ {c}: ${nuevas[c].sum():,.0f}")
    print(f"  filas: {len(nuevas)}")

    df_out = pd.concat([df[~mask26], nuevas], ignore_index=True)
    print(f"\nHistórico: {len(df)} → {len(df_out)} filas "
          f"(−{int(mask26.sum())} viejas +{len(nuevas)} nuevas)")

    if apply:
        bak = HIST.with_suffix(".parquet.bak_ev26")
        if not bak.exists():
            shutil.copy2(HIST, bak)
            print(f"backup: {bak.name}")
        df_out.to_parquet(HIST, index=False)
        print("✓ APLICADO")
    else:
        print("→ dry-run (usa --apply para escribir)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    main(apply=ap.parse_args().apply)
