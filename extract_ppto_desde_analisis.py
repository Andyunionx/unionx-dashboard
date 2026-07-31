"""
Extrae metas de venta neta por CANAL y MARCA desde analisis_planificacion_JUL26.xlsx
(fuente correcta: incluye redistribución de Mattel).
Para meses agosto-diciembre completa con los valores del PPTO 2SEM.

Uso: python extract_ppto_desde_analisis.py
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import pandas as pd
from pathlib import Path

ANALISIS_PATH = Path(__file__).parent / 'data' / 'planificacion' / 'analisis_planificacion_JUL26.xlsx'
PPTO_2SEM_PATH = Path(r"C:\Users\felip\Desktop\UNIONX\PPTO 2026\Metas oficiales 2SEM OFICIAL.xlsx")
OUT_DIR = Path(__file__).parent / 'data' / 'planificacion' / 'snapshots'

# Mapping nombre referencia → nombre canónico parquet
CANAL_NOMBRE_MAP = {
    'Corporativo': 'Corporativo',
    'Distribución': 'Distribución',
    'UnionX B2B': 'UnionX B2B',
    'Fidelización': 'Fidelización',
    'Marketplace': 'Marketplace',
    'P.Web': 'P.Web',
    'Tiendas Propias': 'Tiendas Propias',
}
MARCA_NOMBRE_MAP = {
    'LEVO': 'Levo',
    'Lhotse': 'Lhotse',
    'Simplit': 'Simplit',
    'XROAD': 'Xroad',
    'Marca Flash': 'Marcas Flash',
    'Prov. Nacionales': 'Prov. Nacionales',
    'Mattel': 'Mattel',
    'Purito': 'Purito',
}
MONTH_LABEL_MAP = {
    'Ene': '2026-01', 'Feb': '2026-02', 'Mar': '2026-03',
    'Abr': '2026-04', 'May': '2026-05', 'Jun': '2026-06',
    'Jul': '2026-07',
}


def _read_comp_table(ws, dim_col_name, nombre_map):
    """Lee la tabla META del Comp. sheet de analisis.
    Returns list of (dim, mes_str, value)."""
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next((i for i, r in enumerate(rows) if r[0] == dim_col_name), None)
    if hdr_idx is None:
        raise ValueError(f'No se encontró encabezado {dim_col_name}')
    hdr_row = rows[hdr_idx]
    # Columnas META: posición donde el header dice Ene/Feb/Mar... (cada 3 cols)
    meta_cols = {j: MONTH_LABEL_MAP[str(v)] for j, v in enumerate(hdr_row[1:], 1)
                 if v is not None and str(v) in MONTH_LABEL_MAP}
    records = []
    for r in rows[hdr_idx + 2:]:  # skip header + META/REAL/VAR row
        dim_raw = r[0]
        if not dim_raw:
            continue
        if 'TOTAL' in str(dim_raw).upper():
            break
        dim = nombre_map.get(str(dim_raw).strip())
        if dim is None:
            continue
        for j, mes in meta_cols.items():
            val = r[j]
            if val is not None and isinstance(val, (int, float)) and val > 0:
                records.append({'dim': dim, 'mes': mes, 'meta_venta_neta': float(val)})
    return records


def _read_ppto_2sem(ws_rows, dim_col_header, meses_2sem, nombre_map=None):
    """Lee canal/marca metas de PPTO MARCA 2026 para los meses especificados."""
    from extract_ppto_snapshot import _month_col_map  # reuse existing helper

    hdr_idx = next((i for i, r in enumerate(ws_rows) if r[0] == dim_col_header), None)
    if hdr_idx is None:
        return []
    month_map = _month_col_map(ws_rows[hdr_idx])
    # Filtrar solo meses 2SEM
    month_map = {j: m for j, m in month_map.items() if m in meses_2sem}
    if not month_map:
        return []
    records = []
    for r in ws_rows[hdr_idx + 1: hdr_idx + 20]:
        dim_raw = r[0]
        if not dim_raw or 'Total' in str(dim_raw):
            break
        if r[1] != 'Venta Neta Total':
            continue
        if nombre_map:
            dim = nombre_map.get(str(dim_raw).strip())
            if dim is None:
                continue
        else:
            dim = str(dim_raw).strip()
        for j, mes in month_map.items():
            val = r[j]
            if val is not None and isinstance(val, (int, float)) and float(val) > 0:
                records.append({'dim': dim, 'mes': mes, 'meta_venta_neta': float(val)})
    return records


def main():
    if not ANALISIS_PATH.exists():
        print(f"ERROR: No se encontró {ANALISIS_PATH}")
        return
    if not PPTO_2SEM_PATH.exists():
        print(f"ERROR: No se encontró {PPTO_2SEM_PATH}")
        return

    print(f"Leyendo analisis desde {ANALISIS_PATH}...")
    wb_a = openpyxl.load_workbook(str(ANALISIS_PATH), read_only=True, data_only=True)

    print(f"Leyendo PPTO 2SEM desde {PPTO_2SEM_PATH}...")
    wb_p = openpyxl.load_workbook(str(PPTO_2SEM_PATH), read_only=True, data_only=True)
    ws_ppto = wb_p['PPTO MARCA 2026']
    rows_ppto = list(ws_ppto.iter_rows(values_only=True))

    MESES_2SEM = ['2026-08', '2026-09', '2026-10', '2026-11', '2026-12']

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── CANAL ──────────────────────────────────────────────────────────────
    recs_canal_1sem = _read_comp_table(wb_a['Comp. Canales'], 'CANAL', CANAL_NOMBRE_MAP)
    recs_canal_2sem = _read_ppto_2sem(rows_ppto, 'CANAL', MESES_2SEM)
    # Para 2SEM, filtrar solo los 7 canales canónicos
    recs_canal_2sem = [r for r in recs_canal_2sem if r['dim'] in set(CANAL_NOMBRE_MAP.values())]

    df_canal = pd.DataFrame(
        [{'canal': r['dim'], 'mes': r['mes'], 'meta_venta_neta': r['meta_venta_neta']}
         for r in recs_canal_1sem + recs_canal_2sem]
    ).drop_duplicates(subset=['canal', 'mes'])

    out_c = OUT_DIR / 'planif_ppto_canal.parquet'
    df_canal.to_parquet(out_c, index=False)
    print(f"\n✅ Canal: {len(df_canal)} registros → {out_c}")
    piv_c = df_canal.pivot_table(index='canal', columns='mes', values='meta_venta_neta', aggfunc='sum')
    piv_c_m = (piv_c / 1e6).round(1)
    print(piv_c_m.to_string())

    # ── MARCA ──────────────────────────────────────────────────────────────
    recs_marca_1sem = _read_comp_table(wb_a['Comp. Marcas'], 'MARCA', MARCA_NOMBRE_MAP)
    recs_marca_2sem = _read_ppto_2sem(rows_ppto, 'Tipo canal', MESES_2SEM)  # fallback

    # Para 2SEM de marcas: leer directo de PPTO MARCA 2026 (extractor original)
    from extract_ppto_snapshot import extract_marca
    try:
        df_marca_2sem_full = extract_marca(rows_ppto)
        df_marca_2sem_full = df_marca_2sem_full[df_marca_2sem_full['mes'].isin(MESES_2SEM)]
    except Exception as e:
        print(f"Advertencia extrayendo marcas 2SEM: {e}")
        df_marca_2sem_full = pd.DataFrame(columns=['marca', 'mes', 'meta_venta_neta'])

    df_marca_1sem = pd.DataFrame(
        [{'marca': r['dim'], 'mes': r['mes'], 'meta_venta_neta': r['meta_venta_neta']}
         for r in recs_marca_1sem]
    )

    df_marca = pd.concat([df_marca_1sem, df_marca_2sem_full], ignore_index=True)
    df_marca = df_marca.drop_duplicates(subset=['marca', 'mes'])

    out_m = OUT_DIR / 'planif_ppto_marca.parquet'
    df_marca.to_parquet(out_m, index=False)
    print(f"\n✅ Marca: {len(df_marca)} registros → {out_m}")
    piv_m = df_marca.pivot_table(index='marca', columns='mes', values='meta_venta_neta', aggfunc='sum')
    piv_m_m = (piv_m / 1e6).round(1)
    print(piv_m_m[['2026-07']].sort_values('2026-07', ascending=False).to_string())

    print("\nDone!")


if __name__ == '__main__':
    main()
