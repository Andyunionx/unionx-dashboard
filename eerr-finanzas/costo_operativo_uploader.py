"""
Modulo Costo Operativo - parser, persistencia y analisis.

Estructura del archivo de carga (Excel template):
  Hoja "Detalle":
    Columna A: Categoria (Fijo / Variable)
    Columna B: Concepto (ej: "Sueldos Operaciones", "Arriendo Megacentro", etc)
    Columnas C-N: 12 meses (Ene 2026, Feb 2026, ..., Dic 2026)
    Columna O: Total YTD (formula)
    Columna P: Notas

  Hoja "Headcount" (opcional):
    Columna A: Cargo / Persona
    Columna B-N: meses
    Columna O: total

Funciones:
  generar_template(year)            - Crea template Excel descargable
  parsear_archivo(file_obj, year)   - Parsea archivo subido por usuario
  guardar(data, year)               - Persiste en data/kpis_manuales/costo_operativo_<year>.json
  cargar(year)                      - Carga datos persistidos
  analizar(data, ventas_mensuales)  - Devuelve KPIs derivados + comparacion mercado + recomendaciones
"""
import io
import json
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data" / "kpis_manuales"

MESES_ES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

# Plantilla con conceptos sugeridos (el usuario puede agregar/quitar filas)
PLANTILLA_CONCEPTOS = [
    # (categoria, concepto, nota)
    ("Fijo",     "Sueldos Operaciones (planta + bodega)", "Solo planilla equipo bodega + COMEX"),
    ("Fijo",     "Sueldos liderazgo Operaciones",         "Jefatura, supervisores"),
    ("Fijo",     "Honorarios Operaciones",                 "Boletas equipo ops"),
    ("Fijo",     "Arriendo bodega Megacentro",             "Solo arriendo Megacentro"),
    ("Fijo",     "Servicios Megacentro (luz/agua/internet)", ""),
    ("Fijo",     "Seguridad/vigilancia bodega",            ""),
    ("Fijo",     "Seguros mercaderia + bodega",            ""),
    ("Fijo",     "Capacitacion equipo ops",                ""),
    ("Fijo",     "Depreciacion equipos bodega",            "Racks, montacargas, sistema"),
    ("Fijo",     "Software operacional (WMS, ERP modulo)", "Si aplica"),
    ("Variable", "Flete despacho (couriers)",              "Costo total courier B2C+B2B"),
    ("Variable", "Insumos packing (cajas, cinta, etiquet)", ""),
    ("Variable", "Combustible y peajes",                   ""),
    ("Variable", "Movilizacion equipo bodega",             ""),
    ("Variable", "Colacion equipo bodega",                 ""),
    ("Variable", "Comision marketplaces (envio + servicio)", "Falabella, ML, Ripley"),
    ("Variable", "Comision medios de pago",                "Transbank, Pago Facil"),
    ("Variable", "Logistica inversa (devoluciones)",       "Courier retorno + reacondicionamiento"),
    ("Variable", "Otros operativos",                        "Todo lo que no entra arriba"),
]

# Benchmarks de mercado (del Plan Estrategico UnionX 2026-2028)
BENCHMARKS_MERCADO = {
    "costo_pedido_e_com_USD":        {"min": 3, "max": 7, "fuente": "USD/pedido pick+pack+despacho - Plan UnionX p.40"},
    "costo_logistico_pct_venta":     {"min": 0.08, "max": 0.14, "fuente": "8-14% venta segun ticket - Plan UnionX p.50"},
    "ratio_variable_total":          {"min": 0.50, "max": 0.65, "fuente": "Empresas e-com saludables: 50-65% del costo es variable"},
    "sueldos_pct_venta":             {"min": 0.10, "max": 0.18, "fuente": "Retail importador chileno"},
    "arriendo_pct_venta":            {"min": 0.02, "max": 0.05, "fuente": "Bodega + oficina sobre venta"},
}


def _path_persistencia(year: int) -> Path:
    return DATA_DIR / f"costo_operativo_{year}.json"


def cargar(year: int) -> dict:
    """Lee datos persistidos. Devuelve dict con conceptos + valores mensuales."""
    p = _path_persistencia(year)
    if not p.exists():
        return {}
    try:
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def guardar(data: dict, year: int) -> bool:
    """Persiste data en JSON."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    p = _path_persistencia(year)
    payload = dict(data)
    payload["_meta"] = {
        "ts": datetime.now().isoformat(),
        "year": year,
    }
    with open(p, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    return True


def generar_template(year: int) -> bytes:
    """Genera Excel template para descargar. Devuelve bytes del xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle"

    border = Side(style="thin", color="CCCCCC")
    cell_border = Border(top=border, bottom=border, left=border, right=border)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    fijo_fill = PatternFill("solid", fgColor="FEF3C7")
    var_fill = PatternFill("solid", fgColor="DBEAFE")
    header_font = Font(bold=True, color="FFFFFF")

    # Headers
    headers = ["Categoria", "Concepto"] + MESES_ES + ["Total YTD", "Notas"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border

    # Conceptos sugeridos
    for r, (cat, concepto, nota) in enumerate(PLANTILLA_CONCEPTOS, 2):
        ws.cell(row=r, column=1, value=cat).fill = fijo_fill if cat == "Fijo" else var_fill
        ws.cell(row=r, column=2, value=concepto)
        # 12 meses vacios (col 3-14)
        for col in range(3, 15):
            cell = ws.cell(row=r, column=col, value=None)
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '#,##0'
        # Total YTD = formula sum
        ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").number_format = '#,##0'
        ws.cell(row=r, column=16, value=nota)
        # bordes
        for col in range(1, 17):
            ws.cell(row=r, column=col).border = cell_border

    # Anchos
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    for col_letter in "CDEFGHIJKLMN":
        ws.column_dimensions[col_letter].width = 12
    ws.column_dimensions["O"].width = 14
    ws.column_dimensions["P"].width = 30

    # Instrucciones
    last_row = len(PLANTILLA_CONCEPTOS) + 4
    instr = ws.cell(row=last_row, column=1, value="INSTRUCCIONES:")
    instr.font = Font(bold=True, color="1F4E79")
    notas = [
        f"1) Completar valores en MILES de pesos (M$). Ejemplo: 5,000 = $5.000.000",
        f"2) Si una cuenta no aplica, dejar vacio o 0",
        f"3) Podes agregar nuevas filas con tu propio concepto",
        f"4) Categoria debe ser 'Fijo' o 'Variable'",
        f"5) Subir el archivo en Dashboard > Fulfillment > Tab 'Costo Operativo Total'",
    ]
    for i, n in enumerate(notas, 1):
        ws.cell(row=last_row + i, column=1, value=n).font = Font(italic=True, color="64748B")

    # Hoja Headcount opcional
    ws_hc = wb.create_sheet("Headcount")
    headers_hc = ["Cargo / Persona", "Categoria"] + MESES_ES + ["Promedio"]
    for col, h in enumerate(headers_hc, 1):
        c = ws_hc.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = cell_border
    ws_hc.column_dimensions["A"].width = 30
    ws_hc.column_dimensions["B"].width = 14
    for col_letter in "CDEFGHIJKLMN":
        ws_hc.column_dimensions[col_letter].width = 8

    # Guardar a bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def parsear_archivo(file_obj, year: int) -> dict:
    """Lee Excel subido y devuelve estructura normalizada.

    Returns:
      {
        "conceptos": [
          {"categoria": "Fijo", "concepto": "...", "valores": {1: 5000, 2: 5200, ...}, "total": 60000, "nota": "..."},
          ...
        ],
        "year": year,
        "headcount": [...]  # si la hoja existe
      }
    """
    from openpyxl import load_workbook
    wb = load_workbook(file_obj, data_only=True)

    # Hoja Detalle
    if "Detalle" not in wb.sheetnames:
        raise ValueError("El archivo no tiene hoja 'Detalle'. Descarga el template y volve a intentar.")
    ws = wb["Detalle"]

    conceptos = []
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value
        concepto = ws.cell(row=r, column=2).value
        if not concepto or not cat:
            continue
        cat_norm = str(cat).strip().capitalize()
        if cat_norm not in ("Fijo", "Variable"):
            continue

        valores = {}
        total = 0
        for mes_idx in range(12):
            v = ws.cell(row=r, column=3 + mes_idx).value
            if isinstance(v, (int, float)) and v != 0:
                valores[mes_idx + 1] = float(v)
                total += float(v)

        nota = ws.cell(row=r, column=16).value or ""

        if total > 0:  # solo agregar si tiene datos
            conceptos.append({
                "categoria": cat_norm,
                "concepto": str(concepto).strip(),
                "valores": valores,
                "total": total,
                "nota": str(nota).strip(),
            })

    # Hoja Headcount opcional
    headcount = []
    if "Headcount" in wb.sheetnames:
        ws_hc = wb["Headcount"]
        for r in range(2, ws_hc.max_row + 1):
            cargo = ws_hc.cell(row=r, column=1).value
            categoria = ws_hc.cell(row=r, column=2).value
            if not cargo:
                continue
            valores = {}
            for mes_idx in range(12):
                v = ws_hc.cell(row=r, column=3 + mes_idx).value
                if isinstance(v, (int, float)) and v > 0:
                    valores[mes_idx + 1] = float(v)
            if valores:
                headcount.append({
                    "cargo": str(cargo).strip(),
                    "categoria": str(categoria).strip() if categoria else "",
                    "valores": valores,
                })

    wb.close()

    return {
        "conceptos": conceptos,
        "headcount": headcount,
        "year": year,
        "fecha_carga": datetime.now().isoformat(),
    }


def analizar(data: dict, ventas_mensuales: dict = None) -> dict:
    """Analiza los datos cargados y genera KPIs + comparacion mercado + recomendaciones.

    Args:
      data: salida de parsear_archivo o cargar
      ventas_mensuales: dict {mes (1-12): venta_mes} para cruzar con costos

    Returns:
      dict con metricas, benchmarks, gap, recomendaciones
    """
    conceptos = data.get("conceptos", [])
    if not conceptos:
        return {"error": "Sin datos para analizar"}

    # Totales por categoria
    fijos = [c for c in conceptos if c["categoria"] == "Fijo"]
    variables = [c for c in conceptos if c["categoria"] == "Variable"]

    fijos_total = sum(c["total"] for c in fijos)
    variables_total = sum(c["total"] for c in variables)
    total_costo = fijos_total + variables_total

    ratio_var = variables_total / total_costo if total_costo else 0

    # Mes a mes
    mes_a_mes = {m: {"fijo": 0, "variable": 0, "total": 0} for m in range(1, 13)}
    for c in conceptos:
        cat = c["categoria"].lower()
        for m, v in c["valores"].items():
            mes_a_mes[m][cat] = mes_a_mes[m].get(cat, 0) + v
            mes_a_mes[m]["total"] = mes_a_mes[m].get("total", 0) + v

    # Cruce con ventas (si tenemos)
    costo_vs_venta = None
    if ventas_mensuales:
        venta_total = sum(ventas_mensuales.values())
        if venta_total > 0:
            costo_vs_venta = total_costo / venta_total

    # ============= COMPARACION CON BENCHMARKS =============
    comparacion = []

    # Ratio Variable / Total
    bm = BENCHMARKS_MERCADO["ratio_variable_total"]
    estado = "🟢" if bm["min"] <= ratio_var <= bm["max"] else "🟡" if abs(ratio_var - (bm["min"]+bm["max"])/2) < 0.15 else "🔴"
    comparacion.append({
        "kpi": "% Variable / Total",
        "valor": f"{ratio_var*100:.1f}%",
        "benchmark": f"{bm['min']*100:.0f}-{bm['max']*100:.0f}%",
        "estado": estado,
        "interpretacion": (
            "Estructura saludable: tenés flexibilidad para reducir costos en momentos malos." if estado == "🟢"
            else "Estructura muy fija: vulnerable a caídas de venta." if ratio_var < bm["min"]
            else "Estructura muy variable: poca palanca operativa, costos de cada venta son altos."
        ),
        "fuente": bm["fuente"],
    })

    # Costo logistico % venta (si hay ventas)
    if costo_vs_venta is not None:
        bm = BENCHMARKS_MERCADO["costo_logistico_pct_venta"]
        estado = "🟢" if bm["min"] <= costo_vs_venta <= bm["max"] else ("🟡" if costo_vs_venta < bm["min"] * 1.5 else "🔴")
        comparacion.append({
            "kpi": "Costo Operativo / Venta",
            "valor": f"{costo_vs_venta*100:.1f}%",
            "benchmark": f"{bm['min']*100:.0f}-{bm['max']*100:.0f}%",
            "estado": estado,
            "interpretacion": (
                "En rango competitivo." if estado == "🟢"
                else "Por encima del rango: revisar negociacion couriers y costos fijos."
                if costo_vs_venta > bm["max"] else "Por debajo del rango — verificar que esten todas las cuentas incluidas."
            ),
            "fuente": bm["fuente"],
        })

    # Top 3 conceptos por costo
    top_conceptos = sorted(conceptos, key=lambda c: c["total"], reverse=True)[:5]

    # ============= RECOMENDACIONES =============
    recomendaciones = []

    # 1. Si ratio variable bajo, recomendar foco en convertir fijos
    if ratio_var < BENCHMARKS_MERCADO["ratio_variable_total"]["min"]:
        recomendaciones.append({
            "prioridad": "🟡 Media",
            "area": "Estructura de costos",
            "mensaje": f"Solo {ratio_var*100:.0f}% del costo es variable. Buscar oportunidades de tercerizacion o pago variable.",
            "accion": "Negociar pago variable con couriers (por pedido, no fijo) · Outsourcing de funciones no core · Revisar contratos arriendo flexibles",
        })

    # 2. Concentracion en pocos conceptos
    if top_conceptos:
        top_3_pct = sum(c["total"] for c in top_conceptos[:3]) / total_costo if total_costo else 0
        if top_3_pct > 0.65:
            recomendaciones.append({
                "prioridad": "🔴 Alta",
                "area": "Concentracion",
                "mensaje": f"Top 3 conceptos = {top_3_pct*100:.0f}% del costo total. Foco en optimizar: {', '.join(c['concepto'] for c in top_conceptos[:3])}.",
                "accion": "Negociacion con proveedores de las 3 cuentas top · Benchmarks especificos · Buscar alternativas",
            })

    # 3. Costo / venta alto
    if costo_vs_venta and costo_vs_venta > BENCHMARKS_MERCADO["costo_logistico_pct_venta"]["max"]:
        recomendaciones.append({
            "prioridad": "🔴 Alta",
            "area": "Eficiencia operativa",
            "mensaje": f"Costo Operativo {costo_vs_venta*100:.1f}% de venta vs benchmark max {BENCHMARKS_MERCADO['costo_logistico_pct_venta']['max']*100:.0f}%.",
            "accion": "Revisar mix de couriers · Optimizar slotting bodega · Aumentar AOV para mejorar ratio · Liquidar slow movers que ocupan espacio",
        })

    # 4. Tendencia mes a mes (si hay >=3 meses con datos)
    meses_con_datos = sorted([m for m, v in mes_a_mes.items() if v["total"] > 0])
    if len(meses_con_datos) >= 3:
        # Comparar ultimo mes vs promedio anteriores
        ultimo = meses_con_datos[-1]
        anteriores = meses_con_datos[:-1]
        promedio_ant = sum(mes_a_mes[m]["total"] for m in anteriores) / len(anteriores)
        ultimo_v = mes_a_mes[ultimo]["total"]
        if promedio_ant and (ultimo_v - promedio_ant) / promedio_ant > 0.10:
            recomendaciones.append({
                "prioridad": "🟡 Media",
                "area": "Tendencia",
                "mensaje": f"Costo de {MESES_ES[ultimo-1]} subio {((ultimo_v-promedio_ant)/promedio_ant)*100:+.1f}% vs promedio meses anteriores.",
                "accion": "Identificar concepto que provoca el alza · Validar si es estacional o estructural",
            })

    return {
        "fijos_total": fijos_total,
        "variables_total": variables_total,
        "total_costo": total_costo,
        "ratio_variable": ratio_var,
        "costo_vs_venta": costo_vs_venta,
        "mes_a_mes": mes_a_mes,
        "top_conceptos": top_conceptos,
        "comparacion": comparacion,
        "recomendaciones": recomendaciones,
        "n_conceptos": len(conceptos),
        "error": None,
    }
