"""
PASO 3: Crear Reporte de Rentabilidad por Canal - FEBRERO 2026
Con datos validados del sheet "Análisis Resultados"
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows

class CrearReporteRentabilidad:
    """Crea reporte Excel con datos de rentabilidad validados"""

    def __init__(self):
        self.ruta_analisis = Path("../data/planillas/Análisis Contribución 2026 V02.02.xlsx")
        self.ruta_salida = Path("data/outputs/Reporte_Rentabilidad_Febrero_2026_VALIDADO.xlsx")
        self.ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    def extraer_datos(self):
        """Extrae datos de febrero 2026 por canal"""
        df = pd.read_excel(self.ruta_analisis, sheet_name='Análisis Resultados', header=0)

        # Filtrar febrero 2026
        df_febrero = df[(df['AÑO'] == 2026) & (df['Mes'] == 2.0)].copy()

        # Agrupar por canal y sumar
        columnas_suma = {
            'Costo Venta': 'Costo',
            'Margen Directo': 'Margen Directo',
            'Comisión Venta': 'Comisión Venta',
            'Comisión Envío': 'Comisión Envío',
            'Marketing': 'Marketing',
            'Total Comisiones': 'Total Comisiones',
            'Contribución': 'Contribución'
        }

        resumen = []
        for canal in sorted(df_febrero['Canal'].unique()):
            if pd.isna(canal):
                continue

            datos_canal = df_febrero[df_febrero['Canal'] == canal]

            fila = {'Canal': canal}

            # Sumar datos
            for col_origen, col_nuevo in columnas_suma.items():
                fila[col_nuevo] = pd.to_numeric(datos_canal[col_origen], errors='coerce').sum()

            # Calcular ratios
            margen_directo = fila['Margen Directo']
            contribucion = fila['Contribución']

            # Buscar venta para calcular %
            venta = pd.to_numeric(datos_canal['Venta'], errors='coerce').sum()
            fila['Venta'] = venta

            if venta > 0:
                fila['% Margen Directo'] = (margen_directo / venta * 100)
                fila['% Contribución'] = (contribucion / venta * 100)
            else:
                fila['% Margen Directo'] = 0
                fila['% Contribución'] = 0

            resumen.append(fila)

        df_resumen = pd.DataFrame(resumen)

        # Ordenar por Contribución descendente
        df_resumen = df_resumen.sort_values('Contribución', ascending=False).reset_index(drop=True)

        return df_resumen

    def crear_excel(self, df_resumen):
        """Crea archivo Excel formateado"""

        wb = Workbook()

        # =====================================================================
        # SHEET 1: RESUMEN EJECUTIVO
        # =====================================================================
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Ejecutivo"

        # Estilos
        estilo_titulo = Font(size=16, bold=True, color="FFFFFF")
        estilo_subtitulo = Font(size=12, bold=True, color="FFFFFF")
        estilo_encabezado = Font(size=11, bold=True, color="FFFFFF")
        fill_titulo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_subtitulo = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        fill_encabezado = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        fill_total = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws_resumen.merge_cells('A1:H1')
        ws_resumen['A1'] = "REPORTE DE RENTABILIDAD POR CANAL - FEBRERO 2026"
        ws_resumen['A1'].font = estilo_titulo
        ws_resumen['A1'].fill = fill_titulo
        ws_resumen['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_resumen.row_dimensions[1].height = 25

        # Subtítulo
        ws_resumen.merge_cells('A2:H2')
        ws_resumen['A2'] = f"Datos validados | {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_resumen['A2'].font = Font(size=10, italic=True)
        ws_resumen['A2'].alignment = Alignment(horizontal="center")

        # Totales de resumen
        fila = 4
        ws_resumen['A4'] = "TOTALES FEBRERO 2026"
        ws_resumen['A4'].font = estilo_subtitulo
        ws_resumen['A4'].fill = fill_subtitulo
        ws_resumen.merge_cells('A4:B4')

        fila = 5
        totales_items = [
            ("Venta Neta", df_resumen['Venta'].sum()),
            ("Costo", df_resumen['Costo'].sum()),
            ("Margen Directo", df_resumen['Margen Directo'].sum()),
            ("Comisión Venta", df_resumen['Comisión Venta'].sum()),
            ("Comisión Envío", df_resumen['Comisión Envío'].sum()),
            ("Marketing", df_resumen['Marketing'].sum()),
            ("Total Comisiones", df_resumen['Total Comisiones'].sum()),
            ("CONTRIBUCIÓN", df_resumen['Contribución'].sum()),
        ]

        for concepto, valor in totales_items:
            ws_resumen[f'A{fila}'] = concepto
            ws_resumen[f'B{fila}'] = valor
            ws_resumen[f'B{fila}'].number_format = '#,##0'

            if concepto == "CONTRIBUCIÓN":
                ws_resumen[f'A{fila}'].font = Font(bold=True, size=12)
                ws_resumen[f'B{fila}'].font = Font(bold=True, size=12)
                ws_resumen[f'A{fila}'].fill = fill_total
                ws_resumen[f'B{fila}'].fill = fill_total

            fila += 1

        # =====================================================================
        # SHEET 2: DETALLE POR CANAL
        # =====================================================================
        ws_detalle = wb.create_sheet("Detalle por Canal")

        # Encabezados
        headers = ['Canal', 'Venta', 'Costo', 'Margen Directo', '% M.Dir',
                   'Comisión Venta', 'Comisión Envío', 'Marketing',
                   'Total Comisiones', 'Contribución', '% Contribución']

        for col, header in enumerate(headers, 1):
            celda = ws_detalle.cell(row=1, column=col)
            celda.value = header
            celda.font = estilo_encabezado
            celda.fill = fill_encabezado
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = border

        # Datos
        for row_idx, (_, row) in enumerate(df_resumen.iterrows(), 2):
            ws_detalle.cell(row=row_idx, column=1).value = row['Canal']
            ws_detalle.cell(row=row_idx, column=1).border = border

            # Columnas numéricas
            columnas_numeros = [
                (2, 'Venta'),
                (3, 'Costo'),
                (4, 'Margen Directo'),
                (5, '% Margen Directo'),
                (6, 'Comisión Venta'),
                (7, 'Comisión Envío'),
                (8, 'Marketing'),
                (9, 'Total Comisiones'),
                (10, 'Contribución'),
                (11, '% Contribución'),
            ]

            for col_idx, col_name in columnas_numeros:
                celda = ws_detalle.cell(row=row_idx, column=col_idx)
                valor = row[col_name]

                if pd.isna(valor):
                    valor = 0

                if '%' in col_name:
                    celda.value = valor
                    celda.number_format = '0.00"%"'
                else:
                    celda.value = valor
                    celda.number_format = '#,##0'

                celda.border = border
                celda.alignment = Alignment(horizontal="right")

        # Fila total
        total_row = len(df_resumen) + 2
        ws_detalle[f'A{total_row}'] = "TOTAL"
        ws_detalle[f'A{total_row}'].font = Font(bold=True)
        ws_detalle[f'A{total_row}'].fill = fill_total

        for col in range(2, len(headers) + 1):
            celda = ws_detalle.cell(row=total_row, column=col)
            celda.value = f"=SUM({chr(64+col)}2:{chr(64+col)}{total_row-1})"
            celda.font = Font(bold=True)
            celda.fill = fill_total
            celda.border = border

            # Formato
            if col in [5, 11]:  # Columnas %
                celda.number_format = '0.00"%"'
            else:
                celda.number_format = '#,##0'

        # Ancho de columnas
        ws_detalle.column_dimensions['A'].width = 30
        for col in range(2, len(headers) + 1):
            ws_detalle.column_dimensions[chr(64 + col)].width = 14

        # =====================================================================
        # SHEET 3: TOP 10 CANALES
        # =====================================================================
        ws_top = wb.create_sheet("Top 10 Canales")

        top_10 = df_resumen.head(10).copy()
        top_10.insert(0, 'Ranking', range(1, len(top_10) + 1))

        ws_top['A1'] = "TOP 10 CANALES POR CONTRIBUCIÓN"
        ws_top['A1'].font = Font(size=12, bold=True)

        for row_idx, (_, row) in enumerate(top_10.iterrows(), 2):
            ws_top.cell(row=row_idx, column=1).value = row['Ranking']
            ws_top.cell(row=row_idx, column=2).value = row['Canal']
            ws_top.cell(row=row_idx, column=3).value = row['Contribución']
            ws_top.cell(row=row_idx, column=3).number_format = '#,##0'
            ws_top.cell(row=row_idx, column=4).value = row['% Contribución']
            ws_top.cell(row=row_idx, column=4).number_format = '0.00"%"'

        ws_top.column_dimensions['A'].width = 8
        ws_top.column_dimensions['B'].width = 30
        ws_top.column_dimensions['C'].width = 15
        ws_top.column_dimensions['D'].width = 15

        # Ancho de columnas en resumen
        ws_resumen.column_dimensions['A'].width = 30
        ws_resumen.column_dimensions['B'].width = 20

        # Guardar
        wb.save(self.ruta_salida)
        print(f"\n[OK] Reporte generado: {self.ruta_salida.name}")

        return str(self.ruta_salida)

    def ejecutar(self) -> str:
        """Ejecuta la creación del reporte"""
        print("\n" + "="*80)
        print(" PASO 3: CREAR REPORTE DE RENTABILIDAD")
        print("="*80)

        try:
            print(f"\n[Extrayendo datos...]")
            df_resumen = self.extraer_datos()

            print(f"[OK] {len(df_resumen)} canales extraídos")
            print(f"\nPrimeros 5 canales:")
            print(df_resumen[['Canal', 'Venta', 'Contribución', '% Contribución']].head().to_string())

            print(f"\n[Creando Excel...]")
            ruta = self.crear_excel(df_resumen)

            return ruta

        except Exception as e:
            print(f"[ERROR] {e}")
            import traceback
            traceback.print_exc()
            return ""

# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    creador = CrearReporteRentabilidad()
    ruta = creador.ejecutar()

    if ruta:
        print(f"\n" + "="*80)
        print(" REPORTE COMPLETADO")
        print("="*80)
        print(f"\n[ARCHIVO] {ruta}")
        print(f"\n[SHEETS]")
        print(f"  1. Resumen Ejecutivo")
        print(f"  2. Detalle por Canal")
        print(f"  3. Top 10 Canales")
        print(f"\nProximo paso: PASO 4 - Validar números del reporte")
    else:
        print(f"\n[ERROR] No se pudo crear el reporte")
