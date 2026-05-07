"""
Procesa EERR real adaptando al formato real del archivo
Extrae de la hoja "DETALLE" o "DETALLE GASTOS"
"""

import openpyxl
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent))

from eerr_classifier import EERRClassifier, EERRExporter, FilaEERR
from integracion_distribucion_comisiones import DistribuidorComisiones

def procesar_eerr_detalle(ruta_archivo, mes="Febrero"):
    """
    Lee la hoja DETALLE y extrae datos para clasificar
    """
    print("\n" + "="*70)
    print(f"PROCESANDO EERR - {mes}")
    print("="*70)

    wb = openpyxl.load_workbook(ruta_archivo, data_only=True)

    # Buscar hoja DETALLE o DETALLE GASTOS
    hoja_target = None
    for nombre in ["DETALLE", "DETALLE GASTOS"]:
        if nombre in wb.sheetnames:
            hoja_target = nombre
            break

    if not hoja_target:
        print(f"\nError: No se encontró hoja DETALLE o DETALLE GASTOS")
        print(f"Hojas disponibles: {wb.sheetnames}")
        return

    ws = wb[hoja_target]
    print(f"\nLeyendo hoja: {hoja_target}")
    print(f"Dimensiones: {ws.max_row} x {ws.max_column}")

    # Estructura conocida de la hoja DETALLE
    # Basada en análisis del archivo real
    encabezado_row = 2
    col_codigo = 8      # Código contable
    col_monto = 7       # M$
    col_ln = 1          # Línea de Negocio
    col_glosa = 14      # Glosa
    col_contra = 21     # Contraparte
    datos_inicio = 3    # Los datos comienzan en fila 3

    print(f"\nEstructura de DETALLE:")
    print(f"  Encabezado: Fila {encabezado_row}")
    print(f"  Datos: desde fila {datos_inicio}")
    print(f"  Código (col {col_codigo}), Monto (col {col_monto})")
    print(f"  LN (col {col_ln}), Glosa (col {col_glosa}), Contra (col {col_contra})")

    # Extraer datos
    datos = []
    for row_idx in range(datos_inicio, ws.max_row + 1):
        try:
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            # Validar que row tiene suficientes columnas
            if len(row) < col_contra:
                continue

            codigo_val = row[col_codigo - 1]  # Col 8
            monto_val = row[col_monto - 1]   # Col 7

            # Skip si no hay código o monto
            if not codigo_val or monto_val is None:
                continue

            codigo = str(codigo_val).strip()
            if codigo == "None" or not codigo:
                continue

            # Convertir monto (ya está en miles, no dividir)
            try:
                monto = float(monto_val)
            except:
                continue

            # Otros campos
            glosa = str(row[col_glosa - 1]) if col_glosa and col_glosa <= len(row) else ""
            contra = str(row[col_contra - 1]) if col_contra and col_contra <= len(row) else ""
            ln_existente = str(row[col_ln - 1]) if col_ln and col_ln <= len(row) else ""

            # Limpiar valores
            glosa = glosa.strip() if glosa not in ["None", "0", "False"] else ""
            contra = contra.strip() if contra not in ["None", "0", "False"] else ""
            ln_existente = ln_existente.strip() if ln_existente not in ["None", "0", "False"] else ""

            datos.append({
                "codigo": codigo,
                "glosa": glosa,
                "contraparte": contra,
                "cuenta": codigo,
                "saldo": monto,
                "ln_existente": ln_existente
            })

        except Exception as e:
            continue

    print(f"\nExtraídos: {len(datos)} movimientos")

    if len(datos) == 0:
        print("Error: No se extrajeron datos")
        return

    # Clasificar
    print(f"\nClasificando...")
    classifier = EERRClassifier()
    filas_procesadas, stats = classifier.procesar_eerr(datos)

    print(f"\nEstadísticas de clasificación:")
    for key, val in stats.items():
        print(f"  {key.replace('_', ' ').title():.<30} {val}")

    # Distribuir por canal
    print(f"\nDistribuyendo por canal...")
    distribuidor = DistribuidorComisiones()
    distribucion = distribuidor.procesar_eerr(filas_procesadas, mes)

    print(f"\nDistribución por canal:")
    for canal, resumen in distribucion["resumen"].items():
        if resumen["cantidad"] > 0:
            print(f"  {canal:.<30} {resumen['cantidad']:>3} mov | M$ {resumen['monto_total']:>10.2f}")

    # Exportar
    base_name = Path(ruta_archivo).stem
    print(f"\nExportando...")
    EERRExporter.a_json(filas_procesadas, f"{base_name}_CLASIFICADO.json")
    EERRExporter.a_excel(filas_procesadas, f"{base_name}_CLASIFICADO.xlsx", stats)
    DistribuidorComisiones.a_json_skill(distribucion, f"{base_name}_DISTRIBUCION_CANALES.json")
    DistribuidorComisiones.generar_reporte_html(distribucion, f"{base_name}_REPORTE_CANALES.html")

    print(f"\n" + "="*70)
    print("PROCESO COMPLETADO")
    print("="*70)
    print(f"\nArchivos generados:")
    print(f"  OK {base_name}_CLASIFICADO.json")
    print(f"  OK {base_name}_CLASIFICADO.xlsx")
    print(f"  OK {base_name}_DISTRIBUCION_CANALES.json  <- Para skill")
    print(f"  OK {base_name}_REPORTE_CANALES.html")

    return filas_procesadas, distribucion


if __name__ == "__main__":
    ruta = "Junior Revenue/EERR/02 EE.RR Febrero 2026.xlsx"
    procesar_eerr_detalle(ruta, "Febrero")
