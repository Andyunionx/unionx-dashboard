"""
REPORTE 2: KPIs Operacionales (Incluye COMEX)
Automatizado cada lunes 9 AM (o diario como dashboard)

Entrada:
  - Odoo: Inventario, pedidos, despachos (via API o export)
  - Datos COMEX: Importaciones en trnsito, costos, ETAs (desde agente COMEX)
  - Histrico de mtricas (rotacin, fulfillment, lead times)

Salida:
  - Dashboard interactivo (Excel o Google Data Studio)
  - Email ejecutivo con KPIs clave
  - Alertas si se detectan anomalas operacionales
"""

import json
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference


class GeneradorReporte2:
    """Genera Reporte 2: KPIs Operacionales"""

    def __init__(self, odoo_data_path: str, comex_data_path: Optional[str] = None):
        """
        Args:
            odoo_data_path: Ruta a datos exportados de Odoo (JSON)
            comex_data_path: Ruta a datos COMEX (JSON con maestra por CC)
        """
        self.odoo_data = self._cargar_odoo(odoo_data_path)
        self.comex_data = self._cargar_comex(comex_data_path)
        self.hoy = datetime.now()

    def _cargar_odoo(self, path: str) -> Dict:
        """Carga datos de Odoo"""
        if not Path(path).exists():
            return {}

        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error cargando Odoo: {e}")
            return {}

    def _cargar_comex(self, path: Optional[str]) -> Dict:
        """Carga datos COMEX con maestra por centro de costo"""
        if not path or not Path(path).exists():
            return {}

        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error cargando COMEX: {e}")
            return {}

    def generar(self) -> str:
        """
        Genera el reporte KPIs completo

        Returns:
            Ruta al archivo Excel generado
        """

        # 1. Extraer KPIs de Inventario
        kpis_inventario = self._extraer_kpis_inventario()

        # 2. Extraer KPIs de Despachos/Fulfillment
        kpis_fulfillment = self._extraer_kpis_fulfillment()

        # 3. Extraer KPIs COMEX (con maestra por CC)
        kpis_comex = self._extraer_kpis_comex()

        # 4. Calcular mtricas derivadas
        metricas_adicionales = self._calcular_metricas_adicionales(kpis_inventario, kpis_fulfillment)

        # 5. Generar Excel
        ruta_salida = self._generar_excel(
            kpis_inventario,
            kpis_fulfillment,
            kpis_comex,
            metricas_adicionales
        )

        return ruta_salida

    def _extraer_kpis_inventario(self) -> Dict:
        """Extrae KPIs de inventario desde Odoo"""

        # ESTRUCTURA ESPERADA:
        # odoo_data['inventory'] = {
        #   'total_unidades': int,
        #   'valor_stock': float,
        #   'ocupacion_pct': float,
        #   'items_bajo_minimo': [...]
        # }

        inventario = self.odoo_data.get('inventory', {})

        return {
            'total_unidades': inventario.get('total_unidades', 0),
            'valor_stock': inventario.get('valor_stock', 0),
            'ocupacion_pct': inventario.get('ocupacion_pct', 0),
            'items_bajo_minimo': len(inventario.get('items_bajo_minimo', [])),
            'rotacion_promedio': inventario.get('rotacion_promedio', 0),
            'sku_total': inventario.get('sku_total', 0),
            'sku_activos': inventario.get('sku_activos', 0),
        }

    def _extraer_kpis_fulfillment(self) -> Dict:
        """Extrae KPIs de despachos/fulfillment"""

        # ESTRUCTURA ESPERADA:
        # odoo_data['fulfillment'] = {
        #   'pedidos_pendientes': int,
        #   'pedidos_despachados_hoy': int,
        #   'pedidos_ontime_pct': float,
        #   'tiempo_promedio_fulfillment_dias': float
        # }

        fulfillment = self.odoo_data.get('fulfillment', {})

        return {
            'pedidos_pendientes': fulfillment.get('pedidos_pendientes', 0),
            'pedidos_despachados_hoy': fulfillment.get('pedidos_despachados_hoy', 0),
            'tasa_ontime_pct': fulfillment.get('pedidos_ontime_pct', 95),
            'tiempo_promedio_dias': fulfillment.get('tiempo_promedio_fulfillment_dias', 2),
            'ordenes_atrasadas': fulfillment.get('ordenes_atrasadas', 0),
        }

    def _extraer_kpis_comex(self) -> Dict:
        """Extrae KPIs COMEX con maestra por centro de costo"""

        # ESTRUCTURA: maestra COMEX con histrico por CC
        # comex_data['importaciones_activas'] = [
        #   {
        #     'id': 'IMP-001',
        #     'proveedor': 'Steven',
        #     'cc': 'DISTRIBUCION',
        #     'status': 'en_transito',
        #     'eta_original': '2026-04-15',
        #     'eta_actual': '2026-04-20',
        #     'dias_retraso': 5,
        #     'costo': 5000,
        #     'costeo_cn': 3000,
        #     'flete': 2000,
        #     'margen_importacion_pct': 18
        #   }
        # ]

        importaciones = self.comex_data.get('importaciones_activas', [])

        # Agrupar por centro de costo para maestra
        por_cc = {}
        for imp in importaciones:
            cc = imp.get('cc', 'OTROS')
            if cc not in por_cc:
                por_cc[cc] = {
                    'cantidad': 0,
                    'valor_total': 0,
                    'costo_promedio': 0,
                    'retrasos_promedio_dias': 0,
                    'margen_promedio': 0,
                }

            por_cc[cc]['cantidad'] += 1
            por_cc[cc]['valor_total'] += imp.get('costo', 0)
            por_cc[cc]['retrasos_promedio_dias'] += imp.get('dias_retraso', 0)
            por_cc[cc]['margen_promedio'] += imp.get('margen_importacion_pct', 0)

        # Calcular promedios
        for cc in por_cc:
            cantidad = por_cc[cc]['cantidad']
            if cantidad > 0:
                por_cc[cc]['costo_promedio'] = por_cc[cc]['valor_total'] / cantidad
                por_cc[cc]['retrasos_promedio_dias'] /= cantidad
                por_cc[cc]['margen_promedio'] /= cantidad

        return {
            'importaciones_activas': len(importaciones),
            'en_transito': len([i for i in importaciones if i.get('status') == 'en_transito']),
            'retrasos_detectados': len([i for i in importaciones if i.get('dias_retraso', 0) > 0]),
            'valor_en_transito': sum(i.get('costo', 0) for i in importaciones if i.get('status') == 'en_transito'),
            'maestra_por_cc': por_cc,
        }

    def _calcular_metricas_adicionales(self, inv: Dict, fulfill: Dict) -> Dict:
        """Calcula mtricas derivadas"""

        # ESTADO de alertas
        alertas = []

        if inv['ocupacion_pct'] > 90:
            alertas.append(" Almacn ocupado > 90%")

        if inv['items_bajo_minimo'] > 0:
            alertas.append(f" {inv['items_bajo_minimo']} SKUs bajo stock mnimo")

        if fulfill['tasa_ontime_pct'] < 95:
            alertas.append(f" Fulfillment {fulfill['tasa_ontime_pct']:.1f}% (meta: 98%)")

        if fulfill['ordenes_atrasadas'] > 0:
            alertas.append(f" {fulfill['ordenes_atrasadas']} rdenes atrasadas")

        return {
            'alertas': alertas,
            'salud_operacional': self._calcular_salud(inv, fulfill),
            'fecha_calculo': self.hoy.isoformat(),
        }

    def _calcular_salud(self, inv: Dict, fulfill: Dict) -> str:
        """Score simple de salud operacional"""

        score = 100

        if inv['ocupacion_pct'] > 90:
            score -= 15
        elif inv['ocupacion_pct'] > 80:
            score -= 5

        if inv['items_bajo_minimo'] > 5:
            score -= 10
        elif inv['items_bajo_minimo'] > 0:
            score -= 5

        if fulfill['tasa_ontime_pct'] < 95:
            score -= 20
        elif fulfill['tasa_ontime_pct'] < 98:
            score -= 5

        if fulfill['ordenes_atrasadas'] > 5:
            score -= 15

        return f"{max(0, score)}/100"

    def _generar_excel(self, inv: Dict, fulfill: Dict, comex: Dict, metricas: Dict) -> str:
        """Genera Excel con todas las secciones"""

        ruta_salida = f"data/outputs/Reporte_KPIs_{datetime.now().strftime('%Y%m%d')}.xlsx"
        Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "KPIs"

        # HEADER
        self._agregar_header(ws, metricas)

        # SECCIN 1: Inventario
        self._agregar_seccion_inventario(ws, inv)

        # SECCIN 2: Fulfillment
        self._agregar_seccion_fulfillment(ws, fulfill)

        # SECCIN 3: COMEX (con maestra por CC)
        self._agregar_seccion_comex(ws, comex)

        # SECCIN 4: Alertas
        self._agregar_seccion_alertas(ws, metricas)

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        # Guardar
        wb.save(ruta_salida)
        print(f" Reporte KPIs generado: {ruta_salida}")

        return ruta_salida

    def _agregar_header(self, ws, metricas: Dict):
        """Agrega header del reporte"""

        ws['A1'] = "REPORTE 2: KPIs OPERACIONALES"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)

        ws['A3'] = f"Salud Operacional: {metricas['salud_operacional']}"
        ws['A3'].font = Font(bold=True, size=11)

    def _agregar_seccion_inventario(self, ws, inv: Dict):
        """Agrega KPIs de inventario"""

        fila = 5
        ws[f'A{fila}'] = "1. INVENTARIO & ALMACN"
        ws[f'A{fila}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{fila}'].fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
        ws.merge_cells(f'A{fila}:C{fila}')

        fila += 1
        datos_inventario = [
            ("Total Unidades", f"{inv['total_unidades']:,}", "unid"),
            ("Valor Stock", f"${inv['valor_stock']:,.0f}", "CLP"),
            ("Ocupacin Almacn", f"{inv['ocupacion_pct']:.1f}%", "% capacidad"),
            ("Rotacin Promedio", f"{inv['rotacion_promedio']:.2f}", "veces/mes"),
            ("SKUs Activos", f"{inv['sku_activos']} de {inv['sku_total']}", "items"),
            ("Items Bajo Mnimo", f"{inv['items_bajo_minimo']}", "SKUs " if inv['items_bajo_minimo'] > 0 else "OK"),
        ]

        for metrica, valor, unidad in datos_inventario:
            ws.cell(row=fila, column=1, value=metrica).font = Font(bold=True)
            ws.cell(row=fila, column=2, value=valor)
            ws.cell(row=fila, column=3, value=unidad).font = Font(italic=True, size=9)
            fila += 1

    def _agregar_seccion_fulfillment(self, ws, fulfill: Dict):
        """Agrega KPIs de fulfillment"""

        fila = ws.max_row + 3
        ws[f'A{fila}'] = "2. DESPACHOS & FULFILLMENT"
        ws[f'A{fila}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{fila}'].fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
        ws.merge_cells(f'A{fila}:C{fila}')

        fila += 1
        datos_fulfillment = [
            ("Tasa On-Time", f"{fulfill['tasa_ontime_pct']:.1f}%", "Meta: 98%"),
            ("Tiempo Promedio", f"{fulfill['tiempo_promedio_dias']:.1f}", "das"),
            ("Pedidos Pendientes", f"{fulfill['pedidos_pendientes']}", "hoy"),
            ("Despachados Hoy", f"{fulfill['pedidos_despachados_hoy']}", "pedidos"),
            ("rdenes Atrasadas", f"{fulfill['ordenes_atrasadas']}", "rdenes " if fulfill['ordenes_atrasadas'] > 0 else "OK"),
        ]

        for metrica, valor, nota in datos_fulfillment:
            ws.cell(row=fila, column=1, value=metrica).font = Font(bold=True)
            ws.cell(row=fila, column=2, value=valor)
            ws.cell(row=fila, column=3, value=nota).font = Font(italic=True, size=9)
            fila += 1

    def _agregar_seccion_comex(self, ws, comex: Dict):
        """Agrega KPIs COMEX con maestra por CC"""

        fila = ws.max_row + 3
        ws[f'A{fila}'] = "3. COMEX & IMPORTACIONES"
        ws[f'A{fila}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{fila}'].fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
        ws.merge_cells(f'A{fila}:C{fila}')

        # Resumen general
        fila += 1
        datos_comex = [
            ("Importaciones Activas", f"{comex['importaciones_activas']}", "total"),
            ("En Trnsito", f"{comex['en_transito']}", "importaciones"),
            ("Retrasos Detectados", f"{comex['retrasos_detectados']}", "importaciones "),
            ("Valor en Trnsito", f"${comex['valor_en_transito']:,.0f}", "CLP"),
        ]

        for metrica, valor, nota in datos_comex:
            ws.cell(row=fila, column=1, value=metrica).font = Font(bold=True)
            ws.cell(row=fila, column=2, value=valor)
            ws.cell(row=fila, column=3, value=nota).font = Font(italic=True, size=9)
            fila += 1

        # Maestra por Centro de Costo
        fila += 2
        ws[f'A{fila}'] = "Maestra COMEX por Centro de Costo"
        ws[f'A{fila}'].font = Font(bold=True, size=10)

        fila += 1
        headers_cc = ['Centro Costo', 'Importaciones', 'Valor Promedio', 'Retraso Promedio', 'Margen %']
        for col, header in enumerate(headers_cc, 1):
            cell = ws.cell(row=fila, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill(start_color="757575", end_color="757575", fill_type="solid")

        fila += 1
        for cc, datos_cc in comex.get('maestra_por_cc', {}).items():
            ws.cell(row=fila, column=1, value=cc)
            ws.cell(row=fila, column=2, value=datos_cc['cantidad'])
            ws.cell(row=fila, column=3, value=datos_cc['costo_promedio']).number_format = '$#,##0'
            ws.cell(row=fila, column=4, value=datos_cc['retrasos_promedio_dias']).number_format = '0.0'
            ws.cell(row=fila, column=5, value=datos_cc['margen_promedio']/100).number_format = '0.0%'
            fila += 1

    def _agregar_seccion_alertas(self, ws, metricas: Dict):
        """Agrega seccin de alertas"""

        fila = ws.max_row + 3
        ws[f'A{fila}'] = "4. ALERTAS"
        ws[f'A{fila}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{fila}'].fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        ws.merge_cells(f'A{fila}:C{fila}')

        fila += 1
        alertas = metricas.get('alertas', [])

        if not alertas:
            ws.cell(row=fila, column=1, value=" Sin alertas - Todo OK")
            ws.cell(row=fila, column=1).font = Font(italic=True, color="2E7D32")
        else:
            for alerta in alertas:
                ws.cell(row=fila, column=1, value=alerta)
                ws.cell(row=fila, column=1).font = Font(color="D32F2F")
                fila += 1


# ============================================================================
# SCRIPT EJECUTABLE (lunes 9 AM o diario)
# ============================================================================

if __name__ == "__main__":
    # Rutas
    ruta_odoo = "data/outputs/odoo_export_20260401.json"  # TODO: reemplazar con ruta real
    ruta_comex = "data/outputs/comex_maestra_cc.json"

    # Generar
    generador = GeneradorReporte2(ruta_odoo, ruta_comex)
    ruta_salida = generador.generar()

    print(f"\n Reporte 2 completo: {ruta_salida}")
