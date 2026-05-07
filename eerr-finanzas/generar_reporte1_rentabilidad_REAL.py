"""
REPORTE 1: RENTABILIDAD (DATOS REALES)
Lee desde "Análisis Contribución 2026" y "Planificación Financiera"
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class GeneradorReporte1Real:
    """Genera Reporte 1 con datos reales"""

    def __init__(self):
        self.ruta_analisis = Path("../data/planillas/Análisis Contribución 2026 V02.02.xlsx")
        self.ruta_planificacion = Path("../data/planillas/Planificación Financiera.xlsx")
        self.ruta_presupuesto = Path("../data/planillas/Presupuesto_Febrero_2026.xlsx")

    def _leer_resumen_ytd(self):
        """Lee datos reales de Planificación Financiera - Resumen YTD"""
        try:
            df = pd.read_excel(self.ruta_planificacion, sheet_name='Resumen YTD', header=None)

            # Estructura: Fila 2 = títulos, columnas con datos reales
            datos = {}
            for idx, row in df.iterrows():
                concepto = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""

                if concepto == "Ingresos por Ventas":
                    datos['ventas_ytd'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                elif concepto == "Costos Directos":
                    datos['costo_directo'] = abs(float(row.iloc[1])) if pd.notna(row.iloc[1]) else 0
                elif concepto == "Margen Frontal":
                    datos['margen_frontal'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                elif concepto == "Margen Contribución":
                    datos['margen_contribucion'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                elif "Gastos de Administración" in concepto or "Gastos Operacionales" in concepto:
                    datos['gastos_operacionales'] = abs(float(row.iloc[1])) if pd.notna(row.iloc[1]) else 0

            return datos

        except Exception as e:
            print(f"Error leyendo Resumen YTD: {e}")
            return {}

    def _leer_presupuesto(self):
        """Lee presupuesto para comparativo"""
        try:
            df = pd.read_excel(self.ruta_presupuesto, header=0)

            presupuesto = {
                'ventas': 0,
                'costo': 0,
                'gastos': 0
            }

            for idx, row in df.iterrows():
                concepto = str(row.iloc[0]).lower()

                if 'venta' in concepto:
                    presupuesto['ventas'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                elif 'costo' in concepto:
                    presupuesto['costo'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                elif 'gasto' in concepto:
                    presupuesto['gastos'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0

            return presupuesto

        except Exception as e:
            print(f"Error leyendo presupuesto: {e}")
            return {}

    def _calcular_margenes(self, datos):
        """Calcula los 3 márgenes"""
        ventas = datos.get('ventas_ytd', 0)

        if ventas == 0:
            return {
                'margen_directo': 0,
                'margen_contribucion': 0,
                'margen_operacional': 0
            }

        margen_directo = (datos.get('margen_frontal', 0) / ventas * 100) if ventas else 0
        margen_contrib = (datos.get('margen_contribucion', 0) / ventas * 100) if ventas else 0
        margen_operacional = margen_contrib - (datos.get('gastos_operacionales', 0) / ventas * 100 if ventas else 0)

        return {
            'margen_directo': round(margen_directo, 2),
            'margen_contribucion': round(margen_contrib, 2),
            'margen_operacional': round(margen_operacional, 2)
        }

    def generar(self) -> str:
        """Genera reporte Excel con datos reales"""

        # Leer datos reales
        datos_reales = self._leer_resumen_ytd()
        presupuesto = self._leer_presupuesto()
        margenes = self._calcular_margenes(datos_reales)

        if not datos_reales:
            print("[ERROR] No se pudieron leer datos reales")
            return ""

        # Crear Excel
        ruta_salida = f"data/outputs/Reporte_Rentabilidad_{datetime.now().strftime('%Y%m%d')}.xlsx"
        Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Rentabilidad"

        # HEADER
        ws['A1'] = "REPORTE 1: RENTABILIDAD"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        ws['A2'] = f"Datos Reales - Febrero 2026 | {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)

        # SECCION 1: RENTABILIDAD
        fila = 4
        ws[f'A{fila}'] = "SECCION 1: RENTABILIDAD (3 MARGENES)"
        ws[f'A{fila}'].font = Font(bold=True, size=11)

        fila += 1
        headers = ['Concepto', 'Valor %', 'Estado']
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col)
            celda.value = header
            celda.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            celda.font = Font(bold=True)

        # Datos de márgenes
        fila += 1
        margenes_list = [
            ("Margen Directo", margenes['margen_directo']),
            ("Margen Contribucion", margenes['margen_contribucion']),
            ("Margen Operacional", margenes['margen_operacional'])
        ]

        for concepto, valor in margenes_list:
            ws.cell(row=fila, column=1).value = concepto
            ws.cell(row=fila, column=2).value = valor
            ws.cell(row=fila, column=2).number_format = '0.00"%"'

            # Color según umbral
            if valor < 27:
                ws.cell(row=fila, column=3).value = "CRITICO"
                ws.cell(row=fila, column=3).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws.cell(row=fila, column=3).font = Font(color="FFFFFF")
            elif valor < 30:
                ws.cell(row=fila, column=3).value = "ALERTA"
                ws.cell(row=fila, column=3).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            else:
                ws.cell(row=fila, column=3).value = "OK"
                ws.cell(row=fila, column=3).fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                ws.cell(row=fila, column=3).font = Font(color="FFFFFF")

            fila += 1

        # SECCION 2: RESUMEN P&L
        fila += 2
        ws[f'A{fila}'] = "SECCION 2: RESUMEN P&L REAL"
        ws[f'A{fila}'].font = Font(bold=True, size=11)

        fila += 1
        headers = ['Concepto', 'Valor']
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col)
            celda.value = header
            celda.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            celda.font = Font(bold=True)

        fila += 1
        pyl_items = [
            ("Ventas", datos_reales.get('ventas_ytd', 0)),
            ("Costo Directo", -datos_reales.get('costo_directo', 0)),
            ("Margen Frontal", datos_reales.get('margen_frontal', 0)),
            ("Otros Costos", -datos_reales.get('gastos_operacionales', 0)),
            ("Margen Contribucion", datos_reales.get('margen_contribucion', 0))
        ]

        for concepto, valor in pyl_items:
            ws.cell(row=fila, column=1).value = concepto
            ws.cell(row=fila, column=2).value = valor
            ws.cell(row=fila, column=2).number_format = '#,##0'
            fila += 1

        # Ajustar ancho
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # Guardar
        wb.save(ruta_salida)
        print(f"[OK] Reporte Rentabilidad generado: {ruta_salida}")

        return ruta_salida

# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    print("\n" + "="*70)
    print("GENERADOR DE REPORTE 1: RENTABILIDAD (DATOS REALES)")
    print("="*70)

    generador = GeneradorReporte1Real()
    ruta = generador.generar()

    if ruta:
        print(f"\n[LISTO] Reporte generado: {ruta}")
    else:
        print(f"\n[ERROR] No se pudo generar reporte")
