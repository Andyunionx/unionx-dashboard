"""
Reporte de Stock Mejorado — Generado desde Odoo API
Consulta stock.quant, product.product, stock.location y sale.order.line
para generar un Excel con dashboard, detalle por bodega, fulfillment, alertas.

Uso:
    python eerr-finanzas/reporte_stock_odoo.py
"""
import xmlrpc.client
import json
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONFIG_PATH = os.path.join(BASE_DIR, "odoo", "odoo_config.json")
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "outputs")

# Estilos
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUB_FONT = Font(name="Calibri", size=10, bold=True)
DATA_FONT = Font(name="Calibri", size=9)
TOTAL_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="Calibri", size=9, color="C00000", bold=True)
ORANGE_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="Calibri", size=9, color="006100")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
KPI_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
KPI_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
KPI_VAL_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
BORDER = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

FULFILLMENT_KEYWORDS = ["BFML", "BFP", "BFR", "BFW", "Fulfillment", "fulfillment"]
MARKETING_KEYWORDS = ["Mk", "Marketing", "BMPE", "BMPN", "BMPVS"]
PV_OUTLET_KEYWORDS = ["BPV", "Post Venta", "Outlet", "Bo"]


def write_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = BORDER


def write_row(ws, row, values, fill=None, font=None):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = font or DATA_FONT
        c.border = BORDER
        if fill:
            c.fill = fill
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            c.number_format = '#,##0'
            c.alignment = Alignment(horizontal='right')


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# ODOO CONNECTION
# ============================================================
class OdooClient:
    def __init__(self):
        with open(CONFIG_PATH) as f:
            cfg = json.load(f)["produccion"]
        self.url = cfg["url"]
        self.db = cfg["db_name"]
        self.user = cfg["username"]
        self.pwd = cfg["password"]
        self.uid = None
        self.models = None

    def connect(self):
        print(f"Conectando a {self.url}...")
        common = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/common")
        self.uid = common.authenticate(self.db, self.user, self.pwd, {})
        self.models = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/object")
        print(f"  Conectado. UID={self.uid}")

    def search_read(self, model, domain, fields, limit=0, order=""):
        kwargs = {"fields": fields}
        if limit:
            kwargs["limit"] = limit
        if order:
            kwargs["order"] = order
        return self.models.execute_kw(self.db, self.uid, self.pwd, model, "search_read", [domain], kwargs)

    def search_read_batch(self, model, domain, fields, batch_size=500):
        """Lee en batches para datasets grandes."""
        all_ids = self.models.execute_kw(self.db, self.uid, self.pwd, model, "search", [domain])
        results = []
        for i in range(0, len(all_ids), batch_size):
            batch = all_ids[i:i + batch_size]
            records = self.models.execute_kw(self.db, self.uid, self.pwd, model, "read", [batch], {"fields": fields})
            results.extend(records)
            print(f"    {model}: {len(results)}/{len(all_ids)}...")
        return results


# ============================================================
# DATA EXTRACTION
# ============================================================
def extract_data(odoo):
    data = {}

    # 1. Ubicaciones internas
    print("\n1. Extrayendo ubicaciones...")
    locs = odoo.search_read("stock.location",
                            [["usage", "=", "internal"]],
                            ["id", "complete_name", "name", "location_id"])
    data["locations"] = {l["id"]: l for l in locs}
    print(f"  {len(locs)} ubicaciones internas")

    # 2. Stock quants (qty > 0)
    print("\n2. Extrayendo stock.quant...")
    quants = odoo.search_read_batch("stock.quant",
                                     [["location_id.usage", "=", "internal"], ["quantity", ">", 0]],
                                     ["location_id", "product_id", "product_categ_id", "quantity",
                                      "reserved_quantity", "available_quantity", "value",
                                      "storage_category_id", "in_date"])
    data["quants"] = quants
    print(f"  {len(quants)} quants con stock > 0")

    # 3. Productos (todos los almacenables)
    print("\n3. Extrayendo productos...")
    products = odoo.search_read_batch("product.product",
                                       [["is_storable", "=", True], ["active", "=", True]],
                                       ["id", "name", "default_code", "categ_id", "brand_id",
                                        "avg_cost", "standard_price", "list_price", "total_value",
                                        "qty_available", "free_qty", "incoming_qty", "outgoing_qty",
                                        "uom_id", "active"])
    data["products"] = {p["id"]: p for p in products}
    print(f"  {len(products)} productos almacenables")

    # 4. Ventas ultimos 30 dias (para rotacion)
    print("\n4. Extrayendo ventas ultimos 30 dias...")
    fecha_30d = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    sales = odoo.search_read_batch("sale.order.line",
                                    [["order_id.date_order", ">=", fecha_30d],
                                     ["order_id.state", "in", ["sale", "done"]]],
                                    ["product_id", "product_uom_qty"])
    # Agregar por producto
    ventas_30d = defaultdict(float)
    for s in sales:
        pid = s["product_id"][0] if s["product_id"] else None
        if pid:
            ventas_30d[pid] += s["product_uom_qty"]
    data["ventas_30d"] = ventas_30d
    print(f"  {len(sales)} lineas de venta, {len(ventas_30d)} productos con venta")

    return data


# ============================================================
# DATA PROCESSING
# ============================================================
def process_data(data):
    locations = data["locations"]
    products = data["products"]
    ventas_30d = data["ventas_30d"]

    def get_parent_loc(loc_id):
        loc = locations.get(loc_id)
        if not loc:
            return "Desconocida"
        full = loc["complete_name"]
        parts = full.split("/")
        return "/".join(parts[:2]).strip() if len(parts) >= 2 else full

    def get_child_loc(loc_id):
        loc = locations.get(loc_id)
        return loc["complete_name"] if loc else "Desconocida"

    def classify_location(complete_name):
        cn = complete_name.upper()
        if any(k.upper() in cn for k in FULFILLMENT_KEYWORDS):
            return "Fulfillment"
        if any(k.upper() in cn for k in MARKETING_KEYWORDS):
            return "Marketing"
        if any(k.upper() in cn for k in PV_OUTLET_KEYWORDS):
            return "PV/Outlet"
        return "Planner"

    # Build stock detail rows
    stock_detail = []
    stock_by_sku = defaultdict(lambda: {
        "product_name": "", "sku": "", "category": "", "brand": "", "uom": "",
        "qty_available": 0, "qty_reserved": 0, "qty_free": 0,
        "cost_unit": 0, "value": 0, "venta_30d": 0, "locations": []
    })

    for q in data["quants"]:
        pid = q["product_id"][0] if q["product_id"] else None
        prod = products.get(pid, {})
        loc_id = q["location_id"][0] if q["location_id"] else None
        loc_name = get_child_loc(loc_id)
        parent_loc = get_parent_loc(loc_id)
        loc_type = classify_location(loc_name)

        qty = q.get("quantity", 0) or 0
        reserved = q.get("reserved_quantity", 0) or 0
        available = q.get("available_quantity", 0) or 0
        value = q.get("value", 0) or 0

        cost_unit = prod.get("avg_cost", 0) or prod.get("standard_price", 0) or 0
        if qty > 0 and value == 0:
            value = qty * cost_unit

        sku = prod.get("default_code", "") or ""
        product_name = prod.get("name", q["product_id"][1] if q["product_id"] else "?")
        category = prod.get("categ_id", [0, ""])[1] if prod.get("categ_id") else (q.get("product_categ_id", [0, ""])[1] if q.get("product_categ_id") else "")
        brand = prod.get("brand_id", [0, ""])[1] if prod.get("brand_id") else ""
        uom = prod.get("uom_id", [0, ""])[1] if prod.get("uom_id") else ""

        row = {
            "product_id": pid,
            "product_name": product_name,
            "sku": sku,
            "category": category,
            "brand": brand,
            "uom": uom,
            "parent_loc": parent_loc,
            "child_loc": loc_name,
            "loc_type": loc_type,
            "qty": qty,
            "reserved": reserved,
            "available": available,
            "cost_unit": cost_unit,
            "value": value,
        }
        stock_detail.append(row)

        # Aggregate by SKU
        key = pid or 0
        agg = stock_by_sku[key]
        agg["product_name"] = product_name
        agg["sku"] = sku
        agg["category"] = category
        agg["brand"] = brand
        agg["uom"] = uom
        agg["qty_available"] += qty
        agg["qty_reserved"] += reserved
        agg["qty_free"] += available
        agg["cost_unit"] = cost_unit
        agg["value"] += value
        agg["venta_30d"] = ventas_30d.get(key, 0)
        agg["locations"].append(parent_loc)

    # Calculate dias stock and estado
    for pid, agg in stock_by_sku.items():
        v30 = agg["venta_30d"]
        if v30 > 0:
            venta_diaria = v30 / 30
            agg["dias_stock"] = round(agg["qty_available"] / venta_diaria, 0)
        else:
            agg["dias_stock"] = 999 if agg["qty_available"] > 0 else 0

        # Semaforo: target = 3 meses (90 dias) de inventario
        if agg["qty_available"] == 0 and v30 > 0:
            agg["estado"] = "QUIEBRE"
        elif agg["dias_stock"] < 30 and v30 > 0:
            agg["estado"] = "CRITICO (<30d)"
        elif agg["dias_stock"] < 90 and v30 > 0:
            agg["estado"] = "BAJO (30-89d)"
        elif agg["dias_stock"] <= 180 and v30 > 0:
            agg["estado"] = "OPTIMO (90-180d)"
        elif agg["dias_stock"] > 180 and v30 > 0:
            agg["estado"] = "SOBRESTOCK (>180d)"
        else:
            agg["estado"] = "SIN VENTA"

    # Aggregate by bodega
    stock_by_bodega = defaultdict(lambda: {"qty": 0, "value": 0, "skus": set()})
    for row in stock_detail:
        b = row["parent_loc"]
        stock_by_bodega[b]["qty"] += row["qty"]
        stock_by_bodega[b]["value"] += row["value"]
        stock_by_bodega[b]["skus"].add(row["product_id"])

    return {
        "detail": stock_detail,
        "by_sku": dict(stock_by_sku),
        "by_bodega": dict(stock_by_bodega),
        "ventas_30d": ventas_30d,
    }


# ============================================================
# EXCEL GENERATION
# ============================================================
def generate_excel(processed, output_path):
    wb = Workbook()
    detail = processed["detail"]
    by_sku = processed["by_sku"]
    by_bodega = processed["by_bodega"]

    total_valor = sum(a["value"] for a in by_sku.values())
    total_qty = sum(a["qty_available"] for a in by_sku.values())
    total_skus = len(by_sku)

    # ============================================================
    # HOJA 1: DASHBOARD
    # ============================================================
    ws1 = wb.active
    ws1.title = "Dashboard"
    ws1.sheet_properties.tabColor = "C00000"

    ws1.merge_cells('A1:H1')
    c = ws1.cell(row=1, column=1, value=f"REPORTE DE STOCK — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = Font(name="Calibri", size=16, bold=True, color="1F4E79")

    # KPIs
    kpis = [
        ("Valor Total Inventario", f"${total_valor:,.0f}"),
        ("Total SKUs con Stock", f"{total_skus:,}"),
        ("Total Unidades", f"{total_qty:,.0f}"),
        ("Costo Promedio/SKU", f"${total_valor / total_skus:,.0f}" if total_skus else "$0"),
    ]

    quiebres = sum(1 for a in by_sku.values() if a["estado"] == "QUIEBRE")
    criticos = sum(1 for a in by_sku.values() if a["estado"] == "CRITICO")
    sobrestock = sum(1 for a in by_sku.values() if a["estado"] == "SOBRESTOCK")

    kpis.extend([
        ("SKUs en Quiebre", str(quiebres)),
        ("SKUs Criticos (<7d)", str(criticos)),
        ("SKUs Sobrestock (>90d)", str(sobrestock)),
        ("SKUs con Venta 30d", str(sum(1 for a in by_sku.values() if a["venta_30d"] > 0))),
    ])

    for i, (label, val) in enumerate(kpis):
        col = (i % 4) * 2 + 1
        row = 3 + (i // 4) * 2
        c = ws1.cell(row=row, column=col, value=label)
        c.fill = KPI_FILL
        c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.border = BORDER
        c = ws1.cell(row=row, column=col + 1, value=val)
        c.font = KPI_VAL_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal='center')

    # Resumen por bodega
    row_start = 8
    ws1.cell(row=row_start, column=1, value="RESUMEN POR BODEGA").font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    write_header(ws1, row_start + 1, ["Bodega", "SKUs", "Unidades", "Valor ($)", "% Valor"])
    r = row_start + 2
    for bodega, d in sorted(by_bodega.items(), key=lambda x: -x[1]["value"]):
        pct = d["value"] / total_valor if total_valor else 0
        write_row(ws1, r, [bodega, len(d["skus"]), round(d["qty"]), round(d["value"]), round(pct * 100, 1)])
        ws1.cell(row=r, column=5).number_format = '0.0"%"'
        r += 1
    write_row(ws1, r, ["TOTAL", total_skus, round(total_qty), round(total_valor), 100.0], fill=TOTAL_FILL, font=TOTAL_FONT)

    # Top 20 SKUs por valor
    r += 2
    ws1.cell(row=r, column=1, value="TOP 20 SKUs POR VALOR DE INVENTARIO").font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    r += 1
    write_header(ws1, r, ["SKU", "Producto", "Categoria", "Qty", "Valor ($)", "Venta 30d", "Dias Stock", "Estado"])
    r += 1
    top_valor = sorted(by_sku.items(), key=lambda x: -x[1]["value"])[:20]
    for pid, a in top_valor:
        estado = a["estado"]
        fill = RED_FILL if estado == "QUIEBRE" else (ORANGE_FILL if estado in ("CRITICO", "BAJO") else (YELLOW_FILL if estado == "SOBRESTOCK" else None))
        write_row(ws1, r, [a["sku"], a["product_name"][:40], a["category"], round(a["qty_available"]),
                           round(a["value"]), round(a["venta_30d"]), a["dias_stock"] if a["dias_stock"] < 999 else "Sin venta", estado], fill=fill)
        r += 1

    # Top 20 por rotacion
    r += 1
    ws1.cell(row=r, column=1, value="TOP 20 SKUs POR ROTACION (mas vendidos)").font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    r += 1
    write_header(ws1, r, ["SKU", "Producto", "Venta 30d", "Stock", "Dias Stock", "Valor ($)", "Estado"])
    r += 1
    top_rot = sorted(by_sku.items(), key=lambda x: -x[1]["venta_30d"])[:20]
    for pid, a in top_rot:
        if a["venta_30d"] <= 0:
            continue
        estado = a["estado"]
        fill = RED_FILL if estado == "QUIEBRE" else (ORANGE_FILL if estado in ("CRITICO", "BAJO") else None)
        write_row(ws1, r, [a["sku"], a["product_name"][:40], round(a["venta_30d"]),
                           round(a["qty_available"]), a["dias_stock"] if a["dias_stock"] < 999 else "Sin venta",
                           round(a["value"]), estado], fill=fill)
        r += 1

    set_col_widths(ws1, [30, 18, 30, 14, 16, 14, 14, 14])

    # ============================================================
    # HOJA 2: STOCK TOTAL EMPRESA
    # ============================================================
    ws2 = wb.create_sheet("Stock Total Empresa")
    ws2.sheet_properties.tabColor = "1F4E79"

    ws2.merge_cells('A1:N1')
    ws2.cell(row=1, column=1, value="STOCK TOTAL EMPRESA — Consolidado por SKU").font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

    h2 = ["SKU", "Producto", "Categoria", "Marca", "UdM", "Qty Disponible", "Qty Reservada",
          "Qty Libre", "Costo Unitario", "Valor Stock", "Venta 30d", "Dias Stock", "Estado", "Bodegas"]
    write_header(ws2, 3, h2)

    r2 = 4
    for pid, a in sorted(by_sku.items(), key=lambda x: -x[1]["value"]):
        estado = a["estado"]
        fill = RED_FILL if estado == "QUIEBRE" else (ORANGE_FILL if estado in ("CRITICO", "BAJO") else (YELLOW_FILL if estado == "SOBRESTOCK" else (GREEN_FILL if estado == "OK" else None)))
        bodegas = ", ".join(sorted(set(a["locations"])))
        dias = a["dias_stock"] if a["dias_stock"] < 999 else "Sin venta"
        write_row(ws2, r2, [
            a["sku"], a["product_name"][:50], a["category"], a["brand"], a["uom"],
            round(a["qty_available"]), round(a["qty_reserved"]), round(a["qty_free"]),
            round(a["cost_unit"]), round(a["value"]),
            round(a["venta_30d"]), dias, estado, bodegas[:60]
        ], fill=fill)
        r2 += 1

    write_row(ws2, r2, ["TOTAL", "", "", "", "", round(total_qty), "", "",
                         "", round(total_valor), round(sum(a["venta_30d"] for a in by_sku.values())),
                         "", "", ""], fill=TOTAL_FILL, font=TOTAL_FONT)

    set_col_widths(ws2, [16, 42, 28, 18, 8, 14, 14, 12, 14, 16, 12, 12, 14, 40])
    ws2.auto_filter.ref = f"A3:N{r2}"
    ws2.freeze_panes = "A4"

    # ============================================================
    # HOJA 3: STOCK POR BODEGA (detalle)
    # ============================================================
    ws3 = wb.create_sheet("Stock por Bodega")
    ws3.sheet_properties.tabColor = "548235"

    ws3.merge_cells('A1:J1')
    ws3.cell(row=1, column=1, value="STOCK POR BODEGA — Detalle por ubicacion").font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

    h3 = ["Ubicacion Padre", "Ubicacion Hija", "SKU", "Producto", "Categoria",
          "Qty", "Reservada", "Disponible", "Costo Unit", "Valor"]
    write_header(ws3, 3, h3)

    r3 = 4
    for row in sorted(detail, key=lambda x: (x["parent_loc"], -x["value"])):
        write_row(ws3, r3, [
            row["parent_loc"], row["child_loc"], row["sku"],
            row["product_name"][:45], row["category"],
            round(row["qty"]), round(row["reserved"]), round(row["available"]),
            round(row["cost_unit"]), round(row["value"])
        ])
        r3 += 1

    set_col_widths(ws3, [22, 38, 16, 40, 28, 10, 10, 12, 14, 14])
    ws3.auto_filter.ref = f"A3:J{r3 - 1}"
    ws3.freeze_panes = "A4"

    # ============================================================
    # HOJA 4: STOCK FULFILLMENT
    # ============================================================
    ws4 = wb.create_sheet("Stock Fulfillment")
    ws4.sheet_properties.tabColor = "BF8F00"

    ws4.merge_cells('A1:J1')
    ws4.cell(row=1, column=1, value="STOCK FULFILLMENT — ML, Falabella, Paris, Ripley").font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

    write_header(ws4, 3, h3)
    r4 = 4
    for row in sorted(detail, key=lambda x: (x["parent_loc"], -x["value"])):
        if row["loc_type"] != "Fulfillment":
            continue
        write_row(ws4, r4, [
            row["parent_loc"], row["child_loc"], row["sku"],
            row["product_name"][:45], row["category"],
            round(row["qty"]), round(row["reserved"]), round(row["available"]),
            round(row["cost_unit"]), round(row["value"])
        ])
        r4 += 1

    set_col_widths(ws4, [22, 38, 16, 40, 28, 10, 10, 12, 14, 14])
    ws4.auto_filter.ref = f"A3:J{r4 - 1}"
    ws4.freeze_panes = "A4"

    # ============================================================
    # HOJA 5: MARKETING + PV + OUTLET
    # ============================================================
    ws5 = wb.create_sheet("Mktg PV Outlet")
    ws5.sheet_properties.tabColor = "7030A0"

    ws5.merge_cells('A1:K1')
    ws5.cell(row=1, column=1, value="STOCK MARKETING + POST VENTA + OUTLET").font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

    h5 = h3 + ["Tipo"]
    write_header(ws5, 3, h5)
    r5 = 4
    for row in sorted(detail, key=lambda x: (x["loc_type"], x["parent_loc"], -x["value"])):
        if row["loc_type"] not in ("Marketing", "PV/Outlet"):
            continue
        write_row(ws5, r5, [
            row["parent_loc"], row["child_loc"], row["sku"],
            row["product_name"][:45], row["category"],
            round(row["qty"]), round(row["reserved"]), round(row["available"]),
            round(row["cost_unit"]), round(row["value"]), row["loc_type"]
        ])
        r5 += 1

    set_col_widths(ws5, [22, 38, 16, 40, 28, 10, 10, 12, 14, 14, 12])
    ws5.auto_filter.ref = f"A3:K{r5 - 1}"
    ws5.freeze_panes = "A4"

    # ============================================================
    # HOJA 6: ALERTAS
    # ============================================================
    ws6 = wb.create_sheet("Alertas Stock")
    ws6.sheet_properties.tabColor = "FF0000"

    ws6.merge_cells('A1:I1')
    ws6.cell(row=1, column=1, value="ALERTAS DE STOCK").font = Font(name="Calibri", size=14, bold=True, color="C00000")

    # Quiebres
    r6 = 3
    ws6.cell(row=r6, column=1, value="QUIEBRES — Stock 0 con venta ultimos 30 dias").font = Font(name="Calibri", size=11, bold=True, color="C00000")
    r6 += 1
    write_header(ws6, r6, ["SKU", "Producto", "Categoria", "Marca", "Venta 30d", "Ultimo Stock", "Valor Perdido Est"])
    r6 += 1
    quiebre_list = [(pid, a) for pid, a in by_sku.items() if a["estado"] == "QUIEBRE"]
    quiebre_list.sort(key=lambda x: -x[1]["venta_30d"])
    for pid, a in quiebre_list:
        valor_perdido = a["venta_30d"] * a["cost_unit"]
        write_row(ws6, r6, [a["sku"], a["product_name"][:40], a["category"], a["brand"],
                            round(a["venta_30d"]), 0, round(valor_perdido)], fill=RED_FILL, font=RED_FONT)
        r6 += 1

    # Sobrestock
    r6 += 1
    ws6.cell(row=r6, column=1, value="SOBRESTOCK — Mas de 90 dias de cobertura").font = Font(name="Calibri", size=11, bold=True, color="BF8F00")
    r6 += 1
    write_header(ws6, r6, ["SKU", "Producto", "Categoria", "Qty Stock", "Venta 30d", "Dias Stock", "Valor Inmovilizado"])
    r6 += 1
    sobre_list = [(pid, a) for pid, a in by_sku.items() if a["estado"] == "SOBRESTOCK" and a["venta_30d"] > 0]
    sobre_list.sort(key=lambda x: -x[1]["dias_stock"])
    for pid, a in sobre_list[:50]:
        write_row(ws6, r6, [a["sku"], a["product_name"][:40], a["category"],
                            round(a["qty_available"]), round(a["venta_30d"]),
                            a["dias_stock"], round(a["value"])], fill=YELLOW_FILL)
        r6 += 1

    # Sin movimiento
    r6 += 1
    ws6.cell(row=r6, column=1, value="SIN MOVIMIENTO — Stock > 0 sin venta en 30 dias").font = Font(name="Calibri", size=11, bold=True, color="666666")
    r6 += 1
    write_header(ws6, r6, ["SKU", "Producto", "Categoria", "Qty Stock", "Valor", "Bodegas"])
    r6 += 1
    sin_mov = [(pid, a) for pid, a in by_sku.items() if a["venta_30d"] == 0 and a["qty_available"] > 0]
    sin_mov.sort(key=lambda x: -x[1]["value"])
    for pid, a in sin_mov[:100]:
        bodegas = ", ".join(sorted(set(a["locations"])))
        write_row(ws6, r6, [a["sku"], a["product_name"][:40], a["category"],
                            round(a["qty_available"]), round(a["value"]), bodegas[:50]])
        r6 += 1

    set_col_widths(ws6, [16, 40, 28, 14, 14, 14, 18, 14, 40])

    # Guardar
    wb.save(output_path)
    return wb


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("REPORTE DE STOCK MEJORADO — Desde Odoo API")
    print("=" * 60)

    odoo = OdooClient()
    odoo.connect()

    data = extract_data(odoo)
    processed = process_data(data)

    today = datetime.now().strftime("%Y%m%d")
    output_path = os.path.join(OUTPUT_DIR, f"Reporte_Stock_{today}.xlsx")
    generate_excel(processed, output_path)

    by_sku = processed["by_sku"]
    total_valor = sum(a["value"] for a in by_sku.values())
    total_qty = sum(a["qty_available"] for a in by_sku.values())
    quiebres = sum(1 for a in by_sku.values() if a["estado"] == "QUIEBRE")

    print(f"\n{'=' * 60}")
    print(f"Reporte guardado: {output_path}")
    print(f"  SKUs:      {len(by_sku):,}")
    print(f"  Unidades:  {total_qty:,.0f}")
    print(f"  Valor:     ${total_valor:,.0f}")
    print(f"  Quiebres:  {quiebres}")
    print(f"  Bodegas:   {len(processed['by_bodega'])}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
