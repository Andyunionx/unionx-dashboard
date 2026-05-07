"""
PASO 3a ALTERNATIVA: Usa Raw ventas Y.xlsx en lugar de conectar a Odoo

Flujo:
1. Lee Raw ventas Y.xlsx (sheet RAW)
2. Filtra febrero 2026
3. Agrupa por período/canal/negocio/KAM
4. Convierte a Excel con formato
5. Inyecta en "Análisis Resultado"

Ventaja: No requiere credenciales Odoo
Uso: python paso3a_desde_excel.py
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def ejecutar():
    """Ejecuta PASO 3a desde Excel"""

    print("\n" + "="*100)
    print(" PASO 3a ALTERNATIVA: Desde Excel (Raw ventas Y.xlsx)")
    print("="*100)

    # PASO 1: Leer RAW
    ruta_raw = Path("../datos_entrada/Raw ventas Y.xlsx")

    if not ruta_raw.exists():
        print(f"\n[ERROR] No existe: {ruta_raw}")
        print(f"\nAsegúrate de que Raw ventas Y.xlsx esté en: datos_entrada/")
        print(f"Ruta esperada: {ruta_raw.resolve()}")
        return False

    print(f"\n[Leyendo] {ruta_raw.name}")
    try:
        df_raw = pd.read_excel(ruta_raw, sheet_name='RAW', header=0)
        print(f"[OK] {len(df_raw):,} filas cargadas")
    except Exception as e:
        print(f"[ERROR] No se pudo leer: {e}")
        return False

    # PASO 2: Filtrar febrero
    print(f"\n[Filtrando] Febrero 2026")
    df_febrero = df_raw[(df_raw['Año venta'] == 2026) & (df_raw['Mes venta'] == 2)].copy()
    print(f"[OK] {len(df_febrero):,} filas de febrero 2026")

    # PASO 3: Agrupar
    print(f"\n[Agrupando] Por período/canal/negocio/kam")

    df_agrupado = df_febrero.groupby(
        ['Año venta', 'Mes venta', 'Canal', 'Tipo Negocio', 'KAM'],
        as_index=False
    ).agg({
        'Venta bruta': 'sum',
        'Costo Total': 'sum',
        'Margen Front': 'sum',
        'Cantidad': 'sum'
    })

    # Renombrar
    df_agrupado = df_agrupado.rename(columns={
        'Año venta': 'AÑO',
        'Mes venta': 'Mes',
        'Venta bruta': 'Venta',
        'Costo Total': 'Costo Venta',
        'Margen Front': 'Margen Directo',
    })

    print(f"[OK] {len(df_agrupado)} filas agrupadas")

    # Mostrar resumen
    print(f"\n[RESUMEN] Totales febrero 2026")
    print(f"  Venta total: ${df_agrupado['Venta'].sum():,.0f}")
    print(f"  Costo total: ${df_agrupado['Costo Venta'].sum():,.0f}")
    print(f"  Margen directo: ${df_agrupado['Margen Directo'].sum():,.0f}")

    # PASO 4: Guardar como Excel
    ruta_output = Path("data/outputs/raw_agregado_febrero_2026.xlsx")
    ruta_output.parent.mkdir(parents=True, exist_ok=True)

    print(f"\n[Convirtiendo] A Excel...")

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "RAW Agregado"

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Encabezados
    for col_idx, col_name in enumerate(df_agrupado.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_idx, (_, row) in enumerate(df_agrupado.iterrows(), 2):
        for col_idx, col_name in enumerate(df_agrupado.columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row[col_name]

            # Formato numérico
            if col_name in ['Venta', 'Costo Venta', 'Margen Directo', 'Cantidad']:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Ajustar ancho columnas
    for col_idx, col_name in enumerate(df_agrupado.columns, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    # Guardar
    wb.save(ruta_output)
    print(f"[OK] {ruta_output}")

    # PASO 5: Inyectar en Análisis Resultado
    print(f"\n[Inyectando] en Análisis Resultado...")

    try:
        from inyectar_raw_analisis_resultado import InyectarRawAnalisis

        # Cambiar ruta de input a la versión Excel
        inyector = InyectarRawAnalisis()
        inyector.ruta_raw = ruta_output  # Usar Excel convertido

        # Leer datos para inyección
        df_analisis_actual = inyector.leer_analisis()
        if df_analisis_actual is None:
            return False

        # Preparar inyección
        df_inyeccion = inyector.preparar_inyeccion(df_agrupado)

        # Inyectar
        exito = inyector.inyectar_en_excel(df_analisis_actual, df_inyeccion)

        if exito:
            print(f"\n{'='*100}")
            print(" PASO 3a COMPLETADO")
            print(f"{'='*100}")
            print(f"""
Extracción: Raw ventas Y.xlsx
Período: Febrero 2026
Destino: Análisis Contribución > Análisis Resultados
Estado: Inyectado exitosamente

Archivos generados:
  - {ruta_output.name}

Siguiente: PASO 3b - Mapear EERR + Skill distribución
""")

        return exito

    except Exception as e:
        print(f"\n[ERROR] No se pudo inyectar: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    exito = ejecutar()

    if not exito:
        print("\n[INSTRUCCIONES MANUALES]")
        print("  1. Abre: Análisis Contribución 2026 V02.02.xlsx")
        print("  2. Ve a: Análisis Resultados")
        print("  3. Copia los datos desde raw_agregado_febrero_2026.xlsx")
        print("  4. Pega al final del sheet (sin borrar histórico)")
        print("  5. Verifica que las tablas dinámicas se actualicen")
