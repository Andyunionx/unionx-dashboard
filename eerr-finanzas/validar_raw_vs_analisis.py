"""
Valida que el RAW agregado coincida EXACTAMENTE con "Análisis Resultado"
Y convierte CSV a Excel
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def validar_coincidencias():
    """Compara RAW agregado vs Análisis Resultado para febrero 2026"""

    print("\n" + "="*100)
    print(" VALIDACION: RAW AGREGADO vs ANÁLISIS RESULTADO (Febrero 2026)")
    print("="*100)

    # Leer RAW agregado
    ruta_csv = Path("data/outputs/raw_agregado_febrero_2026.csv")
    df_raw = pd.read_csv(ruta_csv)

    print(f"\n[RAW AGREGADO] {len(df_raw)} combinaciones de Canal/Negocio/KAM")

    # Leer Análisis Resultado
    ruta_analisis = Path("../data/planillas/Análisis Contribución 2026 V02.02.xlsx")
    df_analisis = pd.read_excel(ruta_analisis, sheet_name='Análisis Resultados', header=0)

    # Filtrar por Febrero 2026
    df_analisis_febrero = df_analisis[
        (df_analisis['AÑO'] == 2026) & (df_analisis['Mes'] == 2.0)
    ].copy()

    print(f"[ANÁLISIS RESULTADO] {len(df_analisis_febrero)} filas de febrero 2026")

    # Comparar por CANAL
    print(f"\n[COMPARACION] Totales por CANAL\n")
    print(f"{'Canal':<30} | {'RAW Venta':>15} | {'Análisis Venta':>15} | {'Coincide':>10}")
    print("-" * 75)

    coincidencias = 0
    total_canales = 0

    for canal in sorted(df_raw['Canal'].unique()):
        venta_raw = df_raw[df_raw['Canal'] == canal]['Venta'].sum()
        venta_analisis = df_analisis_febrero[df_analisis_febrero['Canal'] == canal]['Venta'].sum()

        diferencia = abs(venta_raw - venta_analisis)
        coincide = diferencia < 1  # Permitir diferencia menor a 1 por redondeo

        if coincide:
            coincidencias += 1
        total_canales += 1

        marca = "[OK]" if coincide else "[ERROR]"
        print(f"{canal:<30} | ${venta_raw:>14,.0f} | ${venta_analisis:>14,.0f} | {marca}")

    print(f"\n[RESULTADO] {coincidencias}/{total_canales} canales coinciden")

    # Totales generales
    print(f"\n[TOTALES FEBRERO 2026]")
    print(f"\n{'Métrica':<25} | {'RAW':>18} | {'Análisis':>18} | {'Diferencia':>18}")
    print("-" * 85)

    metricas = {
        'Venta': 'Venta',
        'Costo': 'Costo Venta',
        'Margen Directo': 'Margen Directo'
    }

    for nombre, col in metricas.items():
        val_raw = df_raw[col].sum() if col in df_raw.columns else 0
        col_analisis = col if col in df_analisis.columns else None

        if col_analisis:
            val_analisis = df_analisis_febrero[col_analisis].sum()
            diferencia = val_raw - val_analisis

            print(f"{nombre:<25} | ${val_raw:>17,.0f} | ${val_analisis:>17,.0f} | ${diferencia:>17,.0f}")

    return df_raw

def convertir_a_excel(df):
    """Convierte el CSV a Excel"""

    print(f"\n[Convirtiendo a Excel...]")

    ruta_excel = Path("data/outputs/raw_agregado_febrero_2026.xlsx")

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "RAW Agregado"

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Encabezados
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row[col_name]

            # Formato numérico para números
            if col_name in ['Venta', 'Costo Venta', 'Margen Directo', 'Cantidad']:
                cell.number_format = '#,##0.00'

            cell.alignment = Alignment(horizontal="right" if isinstance(row[col_name], (int, float)) else "left")

    # Ajustar ancho de columnas
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    # Guardar
    wb.save(ruta_excel)
    print(f"[OK] Guardado: {ruta_excel}")

    return ruta_excel

# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    df = validar_coincidencias()
    ruta = convertir_a_excel(df)

    print(f"\n" + "="*100)
    print(f" LISTO PARA INYECTAR")
    print("="*100)
    print(f"\nArchivo Excel: {ruta}")
    print(f"\nProximo paso: Inyectar en 'Análisis Resultado'")
