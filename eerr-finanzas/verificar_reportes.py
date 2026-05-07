"""
Verifica el contenido de los 3 reportes generados
Muestra estructura, datos y validaciones
"""

import openpyxl
from pathlib import Path
import json

def verificar_excel(ruta: str, nombre: str):
    """Inspecciona un archivo Excel"""
    print(f"\n{'='*70}")
    print(f" {nombre}")
    print(f"{'='*70}")

    try:
        ruta_path = Path(ruta)
        if not ruta_path.exists():
            print(f"[ERROR] No existe: {ruta}")
            return False

        # Info del archivo
        tamaño_mb = ruta_path.stat().st_size / 1024 / 1024
        print(f"\n[OK] Archivo: {ruta_path.name}")
        print(f"     Tamaño: {tamaño_mb:.2f} MB")

        # Cargar workbook
        wb = openpyxl.load_workbook(ruta)
        print(f"\n[Sheets] Total: {len(wb.sheetnames)}")

        for sheet_name in wb.sheetnames[:3]:  # Primeros 3 sheets
            ws = wb[sheet_name]
            print(f"\n     Sheet: {sheet_name}")
            print(f"     Dimensiones: {ws.dimensions}")
            print(f"     Filas: {ws.max_row}, Columnas: {ws.max_column}")

            # Primeras filas
            print(f"     Primeros datos:")
            for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
                datos = [str(v)[:30] if v else "" for v in row[:5]]
                print(f"       Fila {i}: {datos}")

        # Verificar celdas con valores
        total_celdas = 0
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                total_celdas += sum(1 for cell in row if cell.value is not None)

        print(f"\n[Datos] Celdas con valores: {total_celdas}")

        if total_celdas > 10:
            print(f"[OK] Reporte contiene datos validos")
            return True
        else:
            print(f"[AVISO] Reporte parece estar vacio o con pocos datos")
            return False

    except Exception as e:
        print(f"[ERROR] {e}")
        return False

def verificar_json(ruta: str, nombre: str):
    """Inspecciona archivo JSON"""
    print(f"\n{'='*70}")
    print(f" {nombre}")
    print(f"{'='*70}")

    try:
        ruta_path = Path(ruta)
        if not ruta_path.exists():
            print(f"[ERROR] No existe: {ruta}")
            return False

        with open(ruta, 'r', encoding='utf-8') as f:
            datos = json.load(f)

        tamaño_kb = ruta_path.stat().st_size / 1024
        print(f"\n[OK] Archivo: {ruta_path.name}")
        print(f"     Tamaño: {tamaño_kb:.1f} KB")

        if isinstance(datos, dict):
            print(f"\n[Estructura] Claves principales:")
            for clave in list(datos.keys())[:10]:
                valor = datos[clave]
                if isinstance(valor, list):
                    print(f"     - {clave}: Lista con {len(valor)} items")
                elif isinstance(valor, dict):
                    print(f"     - {clave}: Objeto con {len(valor)} propiedades")
                else:
                    print(f"     - {clave}: {type(valor).__name__}")

            # Mostrar primeros items si es lista
            if any(isinstance(v, list) for v in datos.values()):
                for clave, valor in datos.items():
                    if isinstance(valor, list) and valor:
                        print(f"\n[Datos] {clave} (primeros 2):")
                        for item in valor[:2]:
                            print(f"     {item}")
                        break

        print(f"\n[OK] JSON valido y con estructura")
        return True

    except Exception as e:
        print(f"[ERROR] {e}")
        return False

def main():
    print("\n" + "="*70)
    print("VERIFICACION DE REPORTES GENERADOS")
    print("="*70)

    base_path = Path("data/outputs")

    # Reportes Excel
    reportes = [
        ("data/outputs/Reporte_Rentabilidad_20260401.xlsx", "REPORTE 1: RENTABILIDAD"),
        ("data/outputs/Reporte_KPIs_20260401.xlsx", "REPORTE 2: KPIs OPERACIONALES"),
        ("data/outputs/Reporte_Planificacion_20260401.xlsx", "REPORTE 3: PLANIFICACION FINANCIERA"),
    ]

    resultados = {}
    for ruta, nombre in reportes:
        resultados[nombre] = verificar_excel(ruta, nombre)

    # Alertas JSON
    resultados["ALERTAS"] = verificar_json("data/outputs/alertas_tiempo_real.json", "SISTEMA DE ALERTAS")

    # Resumen ejecutivo
    print(f"\n{'='*70}")
    print(" RESUMEN EJECUTIVO HTML")
    print(f"{'='*70}")
    ruta_html = Path("data/outputs/Resumen_Semanal_20260401.html")
    if ruta_html.exists():
        tamaño = ruta_html.stat().st_size / 1024
        print(f"\n[OK] Archivo: {ruta_html.name}")
        print(f"     Tamaño: {tamaño:.1f} KB")
        print(f"     Estado: Listo para enviar por email al CEO")

    # Resumen final
    print(f"\n{'='*70}")
    print("VERIFICACION FINAL")
    print(f"{'='*70}")

    ok_count = sum(1 for v in resultados.values() if v)
    total_count = len(resultados)

    print(f"\nReportes validados: {ok_count}/{total_count}")
    for nombre, estado in resultados.items():
        status = "OK" if estado else "ERROR"
        print(f"  [{status}] {nombre}")

    if ok_count == total_count:
        print(f"\n[OK] Todos los reportes se generaron correctamente")
        print(f"\nPasos siguientes:")
        print(f"  1. Revisar contenido en Excel (abrir cada archivo)")
        print(f"  2. Validar que numeros tienen sentido con datos reales")
        print(f"  3. Configurar automatizacion (Task Scheduler)")
        return True
    else:
        print(f"\n[AVISO] Algunos reportes tienen problemas")
        return False

if __name__ == "__main__":
    exito = main()
