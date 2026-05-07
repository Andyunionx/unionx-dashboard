"""
Inspecciona estructura de archivos fuente para entender qué datos tienen
"""

import openpyxl
import pandas as pd
from pathlib import Path

def inspeccionar_analisis_contribucion():
    """Inspecciona Análisis Contribución 2026"""
    print("\n" + "="*70)
    print(" ANALISIS CONTRIBUCION 2026 - ESTRUCTURA PARA RENTABILIDAD")
    print("="*70)

    ruta = Path("../data/planillas/Análisis Contribución 2026 V02.02.xlsx")

    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        print(f"\n[OK] Archivo: {ruta.name} (15 MB)")
        print(f"[Sheets] Total: {len(wb.sheetnames)}")

        # Listar primeros sheets
        print(f"\nPrimeros 10 sheets:")
        for i, sheet_name in enumerate(wb.sheetnames[:10], 1):
            ws = wb[sheet_name]
            print(f"  {i}. {sheet_name:40} (Filas: {ws.max_row}, Cols: {ws.max_column})")

            # Mostrar primeras 3 filas del primero
            if i == 1:
                print(f"\n     Datos del primer sheet:")
                for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
                    datos = [str(v)[:25] if v else "" for v in row[:6]]
                    print(f"       Fila {row_idx}: {datos}")

    except Exception as e:
        print(f"[ERROR] {e}")

def inspeccionar_planificacion():
    """Inspecciona Planificación Financiera que cargó"""
    print("\n" + "="*70)
    print(" PLANIFICACION FINANCIERA - ESTRUCTURA")
    print("="*70)

    ruta = Path("../data/planillas/Planificación Financiera.xlsx")

    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        print(f"\n[OK] Archivo: {ruta.name} (1.4 MB)")
        print(f"[Sheets] Total: {len(wb.sheetnames)}")

        # Listar sheets
        print(f"\nSheets disponibles:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            print(f"  {i}. {sheet_name:40} (Filas: {ws.max_row}, Cols: {ws.max_column})")

        # Inspeccionar cada sheet en detalle
        for sheet_name in wb.sheetnames[:5]:
            ws = wb[sheet_name]
            print(f"\n[Sheet] {sheet_name}")
            print(f"        Dimensiones: {ws.dimensions}")

            # Primeras 10 filas
            print(f"        Datos (primeras 10 filas):")
            for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
                # Mostrar solo columnas no vacías
                datos = [str(v)[:20] if v else "" for v in row[:8]]
                datos_filtrados = [d for d in datos if d]
                if datos_filtrados:
                    print(f"          Fila {row_idx}: {datos}")

    except Exception as e:
        print(f"[ERROR] {e}")

def inspeccionar_con_pandas():
    """Inspecciona usando Pandas para mejor legibilidad"""
    print("\n" + "="*70)
    print(" LECTURA CON PANDAS")
    print("="*70)

    # Planificación Financiera
    ruta_plan = Path("../data/planillas/Planificación Financiera.xlsx")

    try:
        # Leer todos los sheets
        excel_file = pd.ExcelFile(ruta_plan)
        print(f"\n[Planificación Financiera] Sheets: {excel_file.sheet_names}")

        # Leer primero
        if excel_file.sheet_names:
            primer_sheet = excel_file.sheet_names[0]
            df = pd.read_excel(ruta_plan, sheet_name=primer_sheet, header=None)

            print(f"\n[Sheet] {primer_sheet}")
            print(f"        Dimensiones: {df.shape}")
            print(f"\n        Primeras 15 filas:")
            print(df.head(15).to_string())

    except Exception as e:
        print(f"[ERROR] {e}")

# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    print("\n" + "="*70)
    print("INSPECCION DE ARCHIVOS FUENTE")
    print("="*70)

    inspeccionar_analisis_contribucion()
    inspeccionar_planificacion()
    inspeccionar_con_pandas()

    print("\n" + "="*70)
    print("INSPECCION COMPLETADA")
    print("="*70)
