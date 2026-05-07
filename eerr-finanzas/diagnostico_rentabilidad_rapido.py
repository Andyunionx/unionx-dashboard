"""
DIAGNOSTICO RAPIDO: Lee solo los primeros datos del sheet
Sin cargar todo el archivo en memoria
"""

import openpyxl
from pathlib import Path

def diagnosticar_rapido():
    """Inspecciona rápidamente sin cargar todo"""
    print("\n" + "="*80)
    print(" DIAGNOSTICO RAPIDO: ANALISIS RESULTADO")
    print("="*80)

    ruta = Path("../data/planillas/Análisis Contribución 2026 V02.02.xlsx")

    try:
        # Cargar solo los nombres de sheets
        wb = openpyxl.load_workbook(ruta, data_only=False)

        print(f"\n[OK] Archivo: {ruta.name}")
        print(f"[Sheets totales] {len(wb.sheetnames)}")

        # Buscar sheet "Análisis Resultado"
        sheet_analisis = None
        for sheet_name in wb.sheetnames:
            if "análisis resultado" in sheet_name.lower() or "analisis resultado" in sheet_name.lower():
                sheet_analisis = sheet_name
                print(f"\n[ENCONTRADO] Sheet: '{sheet_analisis}'")
                break

        if not sheet_analisis:
            print(f"\n[AVISO] Sheet 'Análisis Resultado' no encontrado")
            print(f"\nSheets disponibles ({len(wb.sheetnames)}):")
            for i, sheet in enumerate(wb.sheetnames, 1):
                print(f"  {i:2}. {sheet}")
            return False

        # Leer solo el sheet encontrado
        ws = wb[sheet_analisis]

        print(f"\n[DIMENSIONES] {ws.max_row} filas × {ws.max_column} columnas")
        print(f"[RANGO] {ws.dimensions}")

        # Leer primeras filas (solo estructura, sin datos)
        print(f"\n[ESTRUCTURA] Primeras 15 filas:")
        print(f"{'Fila':<5} | {('Contenido (primeras 8 columnas)'):<70}")
        print("-" * 80)

        for row_idx in range(1, min(16, ws.max_row + 1)):
            valores = []
            for col_idx in range(1, min(9, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    valores.append(str(val)[:15])
                else:
                    valores.append("")

            contenido = " | ".join(valores)
            print(f"{row_idx:<5} | {contenido}")

        # Buscar dónde están los encabezados
        print(f"\n[ENCABEZADOS] Buscando palabras clave...")
        encontrados = []

        for row_idx in range(1, min(50, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    row_values.append(str(cell.value).lower())

            row_str = " ".join(row_values)

            palabras_clave = ["venta", "costo", "margen", "comision", "canal", "linea", "negocio"]
            coincidencias = [p for p in palabras_clave if p in row_str]

            if coincidencias:
                encontrados.append((row_idx, coincidencias))

        if encontrados:
            for row_idx, palabras in encontrados:
                valores = []
                for col_idx in range(1, min(10, ws.max_column + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val is not None:
                        valores.append(str(val)[:12])
                    else:
                        valores.append("")

                print(f"\n  Fila {row_idx} [{', '.join(palabras)}]:")
                print(f"    {' | '.join(valores)}")

        # Mostrar información del sheet
        print(f"\n[INFORMACION]")
        print(f"  Filas con datos: ~{ws.max_row}")
        print(f"  Columnas: ~{ws.max_column}")

        return True

    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()
        return False

# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    exito = diagnosticar_rapido()

    if exito:
        print("\n" + "="*80)
        print(" PROXIMO PASO:")
        print("="*80)
        print("""
Ahora que vemos la estructura, tenemos que:

1. Confirmar qué fila es el encabezado exacto
2. Ver qué columnas hay (Venta, Costo, Margen, Comisiones, etc.)
3. Ver cómo se organiza POR CANAL DE VENTA
4. Buscar datos de FEBRERO 2026 específicamente
5. Validar los números con lo que viste en el archivo manualmente
        """)
