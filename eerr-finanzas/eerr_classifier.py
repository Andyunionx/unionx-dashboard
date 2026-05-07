"""
EERR Classifier & P&L Distribution System
Extrae reglas del HTML de PL_Manager y automatiza clasificación de ingresos/egresos
"""

import json
import re
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional, Tuple
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ============================================================================
# REGLAS DE CLASIFICACIÓN (88 reglas del HTML)
# ============================================================================

REGLAS = [
    {"codigo":"41410101","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"COSTO VENTA","area2":"UNIONX","sa":"UNIONX","ca":"UNIONX"},
    {"codigo":"41410109","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"COSTO VENTA","area2":"UNIONX","sa":"UNIONX","ca":"UNIONX"},
    {"codigo":"42410103","campo":"GLOSA","keyword":"SUBTEL","ln":"UNION X","cc":"PRODUCTOS","area2":"COMERCIAL","sa":"PRODUCTOS","ca":"CERTIFICACIONES SEC"},
    {"codigo":"42410103","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"INSUMOS","area2":"GRUPO ETER","sa":"GRUPO ETER","ca":"OTROS"},
    {"codigo":"42410203","campo":"GLOSA","keyword":"SIMPLIT","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"REVERSSO"},
    {"codigo":"42410203","campo":"GLOSA","keyword":"LHOTSE","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"REVERSSO"},
    {"codigo":"42410203","campo":"GLOSA","keyword":"MELOLLEVO","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"REVERSSO"},
    {"codigo":"42410203","campo":"GLOSA","keyword":"DEVOLUCION","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"REVERSSO"},
    {"codigo":"42410203","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"FLETES Y GASTOS DE ENVIO","area2":"UNION X","sa":"UNION X","ca":""},
    {"codigo":"42410401","campo":"GLOSA","keyword":"FACEBK","ln":"UNION X","cc":"MARKETING DIGITAL","area2":"UNION X","sa":"UNION X","ca":""},
    {"codigo":"42410401","campo":"GLOSA","keyword":"GOOGLE ADS","ln":"UNION X","cc":"MARKETING DIGITAL","area2":"UNION X","sa":"UNION X","ca":""},
    {"codigo":"42410401","campo":"GLOSA","keyword":"DLOCAL","ln":"UNION X","cc":"MARKETING DIGITAL","area2":"UNION X","sa":"UNION X","ca":""},
    {"codigo":"42410401","campo":"GLOSA","keyword":"KLAVIYO","ln":"UNION X","cc":"MARKETING DIGITAL","area2":"UNION X","sa":"UNION X","ca":""},
    {"codigo":"42410401","campo":"GLOSA","keyword":"PUBLICIDAD","ln":"UNION X","cc":"MARKETING DIGITAL","area2":"UNION X","sa":"UNION X","ca":""},
    {"codigo":"42410401","campo":"GLOSA","keyword":"CHATWOOT","ln":"GRUPO ETER","cc":"SUSCRIPCIÓN Y PUBLICACIONES","area2":"GRUPO ETER","sa":"LOGISTICA","ca":"CHATWOOT"},
    {"codigo":"42410401","campo":"GLOSA","keyword":"GOOGLE WORKSPACE","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"CORREO - GSUITE GMAIL"},
    {"codigo":"43420101","campo":"GLOSA","keyword":"GRUPO ETER CONTABILIDAD","ln":"GRUPO ETER","cc":"REMUNERACIONES","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"CONTABILIDAD","ca":"CONTABILIDAD"},
    {"codigo":"43420101","campo":"GLOSA","keyword":"GRUPO ETER GERENCIA","ln":"GRUPO ETER","cc":"REMUNERACIONES","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"FINANZAS Y ADMINISTRACIÓN"},
    {"codigo":"43420101","campo":"GLOSA","keyword":"GRUPO ETER OPERACIONES","ln":"GRUPO ETER","cc":"REMUNERACIONES","area2":"OPERACIONES","sa":"LOGISTICA","ca":"LOGISTICA"},
    {"codigo":"43420101","campo":"GLOSA","keyword":"MELOLLEVO","ln":"UNION X","cc":"REMUNERACIONES","area2":"COMERCIAL","sa":"E-COMMERCE","ca":"E-COMMERCE"},
    {"codigo":"43420101","campo":"GLOSA","keyword":"PAGINAS WEB","ln":"UNION X","cc":"REMUNERACIONES","area2":"COMERCIAL","sa":"PAGINAS WEB","ca":"PAGINAS WEB"},
    {"codigo":"43420101","campo":"GLOSA","keyword":"KAM","ln":"UNION X","cc":"REMUNERACIONES","area2":"COMERCIAL","sa":"DISTRIBUCION","ca":"DISTRIBUCION"},
    {"codigo":"43420101","campo":"NINGUNO","keyword":"","ln":"__SPLIT__","cc":"remuneraciones","area2":"","sa":"","ca":""},
    {"codigo":"43420102","campo":"GLOSA","keyword":"FOTOS","ln":"__SPLIT__","cc":"honorarios","area2":"","sa":"","ca":""},
    {"codigo":"43420102","campo":"GLOSA","keyword":"BRANDING","ln":"__SPLIT__","cc":"honorarios","area2":"","sa":"","ca":""},
    {"codigo":"43420102","campo":"NINGUNO","keyword":"","ln":"__SPLIT__","cc":"honorarios","area2":"","sa":"","ca":""},
    {"codigo":"43420103","campo":"GLOSA","keyword":"GRUPO ETER CONTABILIDAD","ln":"GRUPO ETER","cc":"REMUNERACIONES","area2":"GRUPO ETER","sa":"LEYES SOCIALES","ca":"FINANZAS Y ADMINISTRACIÓN"},
    {"codigo":"43420104","campo":"GLOSA","keyword":"RINDE FONDOS","ln":"__SPLIT__","cc":"rendiciones","area2":"","sa":"","ca":""},
    {"codigo":"43420104","campo":"CONTRAPARTE","keyword":"JOSE TOBAR","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"DISTRIBUCIÓN AGUA"},
    {"codigo":"43420105","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"MANTENCION ACTIVOS","area2":"OPERACIONES","sa":"LOGISTICA","ca":"VEHICULOS Y EQ. TRANSPORTE"},
    {"codigo":"43420111","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"ASESORIAS","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"CONTABILIDAD","ca":"JT ASESORÍAS"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"ADOBE","ln":"UNION X","cc":"PRODUCTOS","area2":"COMERCIAL","sa":"SUSCRIPCIÓN Y PUBLICACIONES","ca":"ADOBE"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"CHATGPT","ln":"GRUPO ETER","cc":"SUSCRIPCIÓN Y PUBLICACIONES","area2":"OPERACIONES","sa":"LOGISTICA","ca":"CHATGPT"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"CLAUDE","ln":"GRUPO ETER","cc":"SUSCRIPCIÓN Y PUBLICACIONES","area2":"GRUPO ETER","sa":"GRUPO ETER","ca":"CLAUDE.AI"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"ODOO","ln":"GRUPO ETER","cc":"SUSCRIPCIÓN Y PUBLICACIONES","area2":"OPERACIONES","sa":"OPERACIONES","ca":"ODOO"},
    {"codigo":"44410190","campo":"NINGUNO","keyword":"","ln":"NO SE CONSIDERA","cc":"","area2":"","sa":"","ca":""},
    {"codigo":"44410302","campo":"NINGUNO","keyword":"","ln":"NO SE CONSIDERA","cc":"","area2":"","sa":"","ca":""},
    # 45 nuevas reglas basadas en análisis de Febrero 2026
    {"codigo":"42410104","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"COMISIÓN GRANDES CUENTAS","area2":"COMERCIAL","sa":"DISTRIBUCION","ca":"COMISIONES"},
    {"codigo":"42410104","campo":"GLOSA","keyword":"MERCADOLIBRE","ln":"UNION X","cc":"COMISIÓN GRANDES CUENTAS","area2":"COMERCIAL","sa":"DISTRIBUCION","ca":"COMISIONES"},
    {"codigo":"42410104","campo":"GLOSA","keyword":"COMERCIAL ECCSA","ln":"UNION X","cc":"COMISIÓN GRANDES CUENTAS","area2":"COMERCIAL","sa":"DISTRIBUCION","ca":"COMISIONES"},
    {"codigo":"42410201","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"FLETES Y GASTOS DE ENVIO","area2":"COMERCIAL","sa":"DISTRIBUCION","ca":"FLETES"},
    {"codigo":"42410201","campo":"GLOSA","keyword":"LOGISTICA","ln":"UNION X","cc":"FLETES Y GASTOS DE ENVIO","area2":"COMERCIAL","sa":"DISTRIBUCION","ca":"FLETES"},
    {"codigo":"42410201","campo":"GLOSA","keyword":"MKP","ln":"UNION X","cc":"FLETES Y GASTOS DE ENVIO","area2":"COMERCIAL","sa":"DISTRIBUCION","ca":"FLETES"},
    {"codigo":"44410305","campo":"NINGUNO","keyword":"","ln":"NO SE CONSIDERA","cc":"","area2":"","sa":"","ca":""},
    {"codigo":"44410306","campo":"NINGUNO","keyword":"","ln":"NO SE CONSIDERA","cc":"","area2":"","sa":"","ca":""},
    {"codigo":"44410307","campo":"NINGUNO","keyword":"","ln":"NO SE CONSIDERA","cc":"","area2":"","sa":"","ca":""},
    {"codigo":"43420114","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"ARRIENDOS","area2":"OPERACIONES","sa":"LOGISTICA","ca":"BODEGAS"},
    {"codigo":"43420114","campo":"GLOSA","keyword":"ARRIENDO","ln":"GRUPO ETER","cc":"ARRIENDOS","area2":"OPERACIONES","sa":"LOGISTICA","ca":"BODEGAS"},
    {"codigo":"43420114","campo":"CONTRAPARTE","keyword":"RENTCO","ln":"GRUPO ETER","cc":"ARRIENDOS","area2":"OPERACIONES","sa":"LOGISTICA","ca":"BODEGAS"},
    {"codigo":"43420103","campo":"GLOSA","keyword":"MELOLLEVO COMERCIAL","ln":"UNION X","cc":"REMUNERACIONES","area2":"COMERCIAL","sa":"E-COMMERCE","ca":"E-COMMERCE"},
    {"codigo":"43420103","campo":"GLOSA","keyword":"GRUPO ETER GERENCIA","ln":"GRUPO ETER","cc":"REMUNERACIONES","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"FINANZAS Y ADMINISTRACIÓN"},
    {"codigo":"43420103","campo":"GLOSA","keyword":"GRUPO ETER OPERACIONES","ln":"GRUPO ETER","cc":"REMUNERACIONES","area2":"OPERACIONES","sa":"OPERACIONES","ca":"OPERACIONES"},
    {"codigo":"43420108","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"REMUNERACIONES","area2":"GRUPO ETER","sa":"POSTVENTA","ca":"INDEMNIZACIONES"},
    {"codigo":"43420108","campo":"GLOSA","keyword":"FINIQUITO","ln":"GRUPO ETER","cc":"REMUNERACIONES","area2":"GRUPO ETER","sa":"POSTVENTA","ca":"INDEMNIZACIONES"},
    {"codigo":"43420109","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"SEGUROS","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"SEGUROS"},
    {"codigo":"43420109","campo":"GLOSA","keyword":"SEGURO","ln":"GRUPO ETER","cc":"SEGUROS","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"SEGUROS"},
    {"codigo":"43420109","campo":"CONTRAPARTE","keyword":"REALE CHILE","ln":"GRUPO ETER","cc":"SEGUROS","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"SEGUROS"},
    {"codigo":"43420115","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"TELEFONO E INTERNET"},
    {"codigo":"43420115","campo":"GLOSA","keyword":"TELEFONO","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"TELEFONO E INTERNET"},
    {"codigo":"43420115","campo":"GLOSA","keyword":"INTERNET","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"TELEFONO E INTERNET"},
    {"codigo":"43420117","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"SUSCRIPCION Y PUBLICACIONES","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"BUK"},
    {"codigo":"43420117","campo":"GLOSA","keyword":"NUBOX","ln":"GRUPO ETER","cc":"SUSCRIPCION Y PUBLICACIONES","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"NUBOX"},
    {"codigo":"43420118","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"ARRIENDOS","area2":"OPERACIONES","sa":"LOGISTICA","ca":"SEGURIDAD OFICINA"},
    {"codigo":"43420118","campo":"GLOSA","keyword":"MONITOREO","ln":"GRUPO ETER","cc":"ARRIENDOS","area2":"OPERACIONES","sa":"LOGISTICA","ca":"SEGURIDAD OFICINA"},
    {"codigo":"43420107","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"MOVILIZACION TRANSPORTE Y COLACION","area2":"OPERACIONES","sa":"LOGISTICA","ca":"AUTOPISTAS"},
    {"codigo":"43420107","campo":"GLOSA","keyword":"PEAJE","ln":"GRUPO ETER","cc":"MOVILIZACION TRANSPORTE Y COLACION","area2":"OPERACIONES","sa":"LOGISTICA","ca":"AUTOPISTAS"},
    {"codigo":"43420104","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"REVERSSO"},
    {"codigo":"43420104","campo":"CONTRAPARTE","keyword":"PRISA","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"PRISA"},
    {"codigo":"43420110","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"MEDIOS DE PAGOS","area2":"COMERCIAL","sa":"PAGINAS WEB","ca":"MEDIOS DE PAGOS"},
    {"codigo":"43420110","campo":"GLOSA","keyword":"MERCADOPAGO","ln":"UNION X","cc":"MEDIOS DE PAGOS","area2":"COMERCIAL","sa":"PAGINAS WEB","ca":"MEDIOS DE PAGOS"},
    {"codigo":"43420110","campo":"GLOSA","keyword":"TRANSBANK","ln":"UNION X","cc":"MEDIOS DE PAGOS","area2":"COMERCIAL","sa":"PAGINAS WEB","ca":"MEDIOS DE PAGOS"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"WAALAXY","ln":"UNION X","cc":"SUSCRIPCION Y PUBLICACIONES","area2":"COMERCIAL","sa":"UNIONX","ca":"WAALAXY"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"APIFY","ln":"UNION X","cc":"SUSCRIPCION Y PUBLICACIONES","area2":"OPERACIONES","sa":"UNIONX","ca":"APIFY"},
    {"codigo":"43410124","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"REPRESENTACION Y VIATICOS","area2":"COMERCIAL","sa":"DISTRIBUCION","ca":"TRANSPORTE"},
    {"codigo":"43410124","campo":"GLOSA","keyword":"VEHICLE RENTAL","ln":"UNION X","cc":"REPRESENTACION Y VIATICOS","area2":"COMERCIAL","sa":"DISTRIBUCION","ca":"TRANSPORTE"},
    {"codigo":"42420120","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"MOVILIZACION TRANSPORTE Y COLACION","area2":"OPERACIONES","sa":"LOGISTICA","ca":"MOVILIZACION"},
    {"codigo":"42420120","campo":"GLOSA","keyword":"DIESEL","ln":"GRUPO ETER","cc":"MOVILIZACION TRANSPORTE Y COLACION","area2":"OPERACIONES","sa":"LOGISTICA","ca":"MOVILIZACION"},
    {"codigo":"43420130","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"DEPRECIACION","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"GRUPO ETER","ca":"DEPRECIACION"},
    {"codigo":"43420141","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"SERVICIOS BASICOS"},
    {"codigo":"43420137","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"ARRIENDOS","area2":"COMERCIAL","sa":"E-COMMERCE","ca":"GGCC LAS CONDES"},
    {"codigo":"43420136","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"SUSCRIPCION Y PUBLICACIONES","area2":"COMERCIAL","sa":"UNIONX","ca":"MEIKIFY"},
    {"codigo":"43420126","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"REMUNERACIONES","area2":"COMERCIAL","sa":"PAGINAS WEB","ca":"CATALOGO"},
    {"codigo":"43420134","campo":"GLOSA","keyword":"ECOMMERCE","ln":"UNION X","cc":"PRODUCTOS","area2":"COMERCIAL","sa":"PRODUCTOS","ca":"HONORARIOS"},
    {"codigo":"43420134","campo":"GLOSA","keyword":"LETRERO","ln":"UNION X","cc":"PRODUCTOS","area2":"COMERCIAL","sa":"HONORARIOS","ca":"SESION CREACIONES VIDEOS & FOTOS"},
    {"codigo":"43510204","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"AJUSTE SENCILLO","area2":"GRUPO ETER","sa":"GRUPO ETER","ca":"GRUPO ETER"},
    {"codigo":"44430105","campo":"NINGUNO","keyword":"","ln":"NO SE CONSIDERA","cc":"","area2":"","sa":"","ca":""},
    # Reglas adicionales para códigos residuales (30 movimientos restantes)
    {"codigo":"42410401","campo":"NINGUNO","keyword":"","ln":"UNION X","cc":"MARKETING DIGITAL","area2":"UNION X","sa":"UNION X","ca":""},
    {"codigo":"43420103","campo":"GLOSA","keyword":"PAGINAS WEB","ln":"UNION X","cc":"REMUNERACIONES","area2":"COMERCIAL","sa":"PAGINAS WEB","ca":"PAGINAS WEB"},
    {"codigo":"43420103","campo":"GLOSA","keyword":"UNION X COMERCIAL","ln":"UNION X","cc":"REMUNERACIONES","area2":"COMERCIAL","sa":"COMERCIAL","ca":"COMERCIAL"},
    {"codigo":"43420103","campo":"GLOSA","keyword":"UNION X PRODUCTOS","ln":"UNION X","cc":"REMUNERACIONES","area2":"COMERCIAL","sa":"PRODUCTOS","ca":"PRODUCTOS"},
    {"codigo":"43420139","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"GASTOS FINANCIEROS"},
    {"codigo":"43420134","campo":"GLOSA","keyword":"RINDE","ln":"UNION X","cc":"PRODUCTOS","area2":"COMERCIAL","sa":"PRODUCTOS","ca":"HONORARIOS"},
    {"codigo":"43420134","campo":"GLOSA","keyword":"MARKETING","ln":"UNION X","cc":"PRODUCTOS","area2":"COMERCIAL","sa":"PRODUCTOS","ca":"HONORARIOS"},
    {"codigo":"43420134","campo":"GLOSA","keyword":"CONTENIDO","ln":"UNION X","cc":"PRODUCTOS","area2":"COMERCIAL","sa":"PRODUCTOS","ca":"HONORARIOS"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"PADDLE","ln":"UNION X","cc":"SUSCRIPCION Y PUBLICACIONES","area2":"COMERCIAL","sa":"UNIONX","ca":"PADDLE"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"MAILCHIMP","ln":"UNION X","cc":"SUSCRIPCION Y PUBLICACIONES","area2":"COMERCIAL","sa":"UNIONX","ca":"MAILCHIMP"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"WI-FI","ln":"GRUPO ETER","cc":"GASTOS OFICINA Y SERVICIOS","area2":"OPERACIONES","sa":"OPERACIONES","ca":"REDES"},
    {"codigo":"43420133","campo":"GLOSA","keyword":"UBER","ln":"GRUPO ETER","cc":"MOVILIZACION TRANSPORTE Y COLACION","area2":"OPERACIONES","sa":"LOGISTICA","ca":"TRANSPORTE"},
    {"codigo":"43420133","campo":"GLOSA","keyword":"UNIMARC","ln":"GRUPO ETER","cc":"MOVILIZACION TRANSPORTE Y COLACION","area2":"OPERACIONES","sa":"OPERACIONES","ca":"COLACION"},
    {"codigo":"43420133","campo":"GLOSA","keyword":"CAPCUT","ln":"UNION X","cc":"SUSCRIPCION Y PUBLICACIONES","area2":"COMERCIAL","sa":"PRODUCTOS","ca":"CAPCUT"},
    {"codigo":"44410304","campo":"NINGUNO","keyword":"","ln":"NO SE CONSIDERA","cc":"","area2":"","sa":"","ca":""},
    {"codigo":"44410304","campo":"GLOSA","keyword":"INTERESES","ln":"NO SE CONSIDERA","cc":"","area2":"","sa":"","ca":""},
    {"codigo":"44410304","campo":"GLOSA","keyword":"IMPUESTOS","ln":"NO SE CONSIDERA","cc":"","area2":"","sa":"","ca":""},
    {"codigo":"44410304","campo":"GLOSA","keyword":"COMISION","ln":"NO SE CONSIDERA","cc":"","area2":"","sa":"","ca":""},
    {"codigo":"43420106","campo":"NINGUNO","keyword":"","ln":"GRUPO ETER","cc":"AJUSTE SENCILLO","area2":"GRUPO ETER","sa":"GRUPO ETER","ca":"AJUSTE"},
    # Reglas finales para últimos 6 movimientos residuales
    {"codigo":"43420103","campo":"GLOSA","keyword":"GERENCIA","ln":"UNION X","cc":"REMUNERACIONES","area2":"COMERCIAL","sa":"GERENCIA","ca":"GERENCIA"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"CONTAB","ln":"GRUPO ETER","cc":"SUSCRIPCION Y PUBLICACIONES","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"CONTABILIDAD"},
    {"codigo":"43420116","campo":"GLOSA","keyword":"TOKU","ln":"GRUPO ETER","cc":"SUSCRIPCION Y PUBLICACIONES","area2":"FINANZAS Y ADMINISTRACIÓN","sa":"FINANZAS Y ADMINISTRACIÓN","ca":"CONTABILIDAD"},
    {"codigo":"43420133","campo":"GLOSA","keyword":"VUELO","ln":"GRUPO ETER","cc":"MOVILIZACION TRANSPORTE Y COLACION","area2":"OPERACIONES","sa":"LOGISTICA","ca":"TRANSPORTE"},
    {"codigo":"43420133","campo":"GLOSA","keyword":"MOVIMIENTO TC","ln":"GRUPO ETER","cc":"MOVILIZACION TRANSPORTE Y COLACION","area2":"OPERACIONES","sa":"OPERACIONES","ca":"TRANSPORTE"},
    {"codigo":"43420134","campo":"GLOSA","keyword":"DESPACHO","ln":"UNION X","cc":"PRODUCTOS","area2":"COMERCIAL","sa":"PRODUCTOS","ca":"HONORARIOS"},
]

SPLIT_ACCOUNTS = [
    {"id":"rendiciones", "titulo":"Rendiciones / Fondos por rendir", "mode":"amount", "cuenta":"43420133", "codigo":"43420133,43420104"},
    {"id":"remuneraciones", "titulo":"Remuneraciones (sin glosa)", "mode":"rows", "cuenta":"43420101", "codigo":"43420101"},
    {"id":"honorarios", "titulo":"Honorarios profesionales", "mode":"rows", "cuenta":"43420102", "codigo":"43420102"},
    {"id":"mktgbranding", "titulo":"Marketing Branding", "mode":"amount", "cuenta":"43420134", "codigo":"43420134"},
]

FCST_COSTO_CODES = {'41410101','41410109','42410104','42410203','42410201','43420110','42410401'}

# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class Clasificacion:
    codigo: str
    ln: str
    cc: str
    area2: str
    sa: str
    ca: str
    metodo: str = "auto"  # auto, manual, split

    def to_dict(self):
        return asdict(self)

@dataclass
class FilaEERR:
    num_fila: int
    codigo: str
    glosa: str
    contraparte: str
    cuenta: str
    saldo: float  # en miles
    clasificacion: Optional[Clasificacion] = None

    def to_dict(self):
        return {
            "num_fila": self.num_fila,
            "codigo": self.codigo,
            "glosa": self.glosa,
            "contraparte": self.contraparte,
            "cuenta": self.cuenta,
            "saldo": self.saldo,
            "clasificacion": self.clasificacion.to_dict() if self.clasificacion else None,
        }

# ============================================================================
# CLASIFICADOR
# ============================================================================

class EERRClassifier:
    def __init__(self):
        self.reglas = REGLAS

    def _normalizar(self, texto: str) -> str:
        """Normaliza texto para búsqueda"""
        return str(texto or "").upper().replace("Á","A").replace("É","E").replace("Í","I").replace("Ó","O").replace("Ú","U").strip()

    def clasificar_fila(self, codigo: str, glosa: str, contraparte: str = "") -> Clasificacion:
        """
        Aplica reglas de clasificación en orden:
        1. Busca por código
        2. Si campo=GLOSA, busca keyword en glosa normalizada
        3. Si campo=CONTRAPARTE, busca keyword en contraparte normalizada
        4. Si campo=NINGUNO, aplica directamente
        """
        codigo = str(codigo or "").strip()
        glosa_norm = self._normalizar(glosa)
        contra_norm = self._normalizar(contraparte)

        # Buscar todas las reglas que coincidan con el código
        reglas_codigo = [r for r in self.reglas if r["codigo"] == codigo]

        for regla in reglas_codigo:
            if regla["campo"] == "NINGUNO" or not regla.get("keyword"):
                # Aplicar directamente
                return Clasificacion(
                    codigo=codigo,
                    ln=regla["ln"],
                    cc=regla["cc"],
                    area2=regla["area2"],
                    sa=regla["sa"],
                    ca=regla["ca"],
                    metodo="auto"
                )
            elif regla["campo"] == "GLOSA" and regla.get("keyword"):
                if self._normalizar(regla["keyword"]) in glosa_norm:
                    return Clasificacion(
                        codigo=codigo,
                        ln=regla["ln"],
                        cc=regla["cc"],
                        area2=regla["area2"],
                        sa=regla["sa"],
                        ca=regla["ca"],
                        metodo="auto"
                    )
            elif regla["campo"] == "CONTRAPARTE" and regla.get("keyword"):
                if self._normalizar(regla["keyword"]) in contra_norm:
                    return Clasificacion(
                        codigo=codigo,
                        ln=regla["ln"],
                        cc=regla["cc"],
                        area2=regla["area2"],
                        sa=regla["sa"],
                        ca=regla["ca"],
                        metodo="auto"
                    )

        # Sin clasificación
        return Clasificacion(codigo=codigo, ln="", cc="", area2="", sa="", ca="", metodo="auto")

    def procesar_eerr(self, filas: List[Dict]) -> Tuple[List[FilaEERR], Dict]:
        """
        Procesa un listado de filas EERR
        Retorna: (filas procesadas, estadísticas)
        """
        resultado = []
        stats = {"ge": 0, "ux": 0, "nc": 0, "sin_clasificar": 0, "split": 0}

        for i, fila in enumerate(filas, 1):
            cls = self.clasificar_fila(
                codigo=fila.get("codigo") or fila.get("CÓDIGO"),
                glosa=fila.get("glosa") or fila.get("GLOSA"),
                contraparte=fila.get("contraparte") or fila.get("CONTRAPARTE"),
            )

            fila_eerr = FilaEERR(
                num_fila=i,
                codigo=cls.codigo,
                glosa=fila.get("glosa") or fila.get("GLOSA") or "",
                contraparte=fila.get("contraparte") or fila.get("CONTRAPARTE") or "",
                cuenta=fila.get("cuenta") or fila.get("CUENTA") or "",
                saldo=float(fila.get("saldo") or fila.get("SALDO") or 0),
                clasificacion=cls,
            )

            resultado.append(fila_eerr)

            # Estadísticas
            if not cls.ln or cls.ln == "":
                stats["sin_clasificar"] += 1
            elif cls.ln == "__SPLIT__":
                stats["split"] += 1
            elif cls.ln == "GRUPO ETER":
                stats["ge"] += 1
            elif cls.ln == "UNION X":
                stats["ux"] += 1
            elif cls.ln == "NO SE CONSIDERA":
                stats["nc"] += 1

        return resultado, stats

# ============================================================================
# EXPORTADORES
# ============================================================================

class EERRExporter:
    @staticmethod
    def a_json(filas: List[FilaEERR], output_path: str):
        """Exporta a JSON"""
        datos = [f.to_dict() for f in filas]
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        print(f"OK Exportado a {output_path}")

    @staticmethod
    def a_excel(filas: List[FilaEERR], output_path: str, stats: Dict = None):
        """Exporta a Excel con formato"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EERR Clasificado"

        # Encabezados
        headers = ["#", "Código", "Glosa", "Contraparte", "Cuenta", "Saldo M$",
                   "Línea Negocio", "Centro Costos", "Área Empresa", "Sub-área", "Cuenta Analítica"]
        ws.append(headers)

        # Formato encabezado
        header_fill = PatternFill(start_color="0F2140", end_color="0F2140", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Datos
        for fila in filas:
            cls = fila.clasificacion
            ws.append([
                fila.num_fila,
                fila.codigo,
                fila.glosa[:50] if fila.glosa else "",
                fila.contraparte[:30] if fila.contraparte else "",
                fila.cuenta,
                fila.saldo,
                cls.ln if cls else "",
                cls.cc if cls else "",
                cls.area2 if cls else "",
                cls.sa if cls else "",
                cls.ca if cls else "",
            ])

            # Color de fila según clasificación
            if cls.ln == "GRUPO ETER":
                color = "E8F5EE"  # Verde claro
            elif cls.ln == "UNION X":
                color = "FFFBEB"  # Amarillo claro
            elif cls.ln == "NO SE CONSIDERA":
                color = "FFF7ED"  # Naranja claro
            elif not cls.ln:
                color = "FEF2F2"  # Rojo claro
            else:
                color = "FFFFFF"

            row_num = ws.max_row
            for cell in ws.iter_rows(min_row=row_num, max_row=row_num):
                for c in cell:
                    c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Ajuste de ancho
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 20

        # Estadísticas en hoja adicional
        if stats:
            ws_stats = wb.create_sheet("Estadísticas")
            ws_stats.append(["Métrica", "Cantidad"])
            for key, val in stats.items():
                ws_stats.append([key.replace("_", " ").title(), val])
            ws_stats.column_dimensions['A'].width = 25
            ws_stats.column_dimensions['B'].width = 12

        wb.save(output_path)
        print(f"OK Exportado a {output_path}")

# ============================================================================
# EJEMPLO DE USO
# ============================================================================

if __name__ == "__main__":
    # Simulación de datos EERR
    datos_eerr = [
        {
            "codigo": "41410101",
            "glosa": "Venta Recíbelo",
            "contraparte": "RECIBELO",
            "cuenta": "Ingresos",
            "saldo": 1500.5
        },
        {
            "codigo": "43420101",
            "glosa": "Sueldo GRUPO ETER CONTABILIDAD",
            "contraparte": "",
            "cuenta": "Remuneraciones",
            "saldo": -250.0
        },
        {
            "codigo": "43420116",
            "glosa": "Suscripción CLAUDE",
            "contraparte": "ANTHROPIC",
            "cuenta": "Servicios",
            "saldo": -15.0
        },
        {
            "codigo": "42410401",
            "glosa": "Publicidad GOOGLE ADS",
            "contraparte": "GOOGLE",
            "cuenta": "Marketing",
            "saldo": -300.0
        },
        {
            "codigo": "44410190",
            "glosa": "Descuento especial",
            "contraparte": "",
            "cuenta": "Ajustes",
            "saldo": -50.0
        },
        {
            "codigo": "43420102",
            "glosa": "Trabajo fotos estudio BRANDING",
            "contraparte": "FOTOGRAFICO",
            "cuenta": "Honorarios",
            "saldo": -120.0
        },
    ]

    print("="*70)
    print("EERR CLASSIFIER - DEMO")
    print("="*70)

    classifier = EERRClassifier()
    filas_procesadas, stats = classifier.procesar_eerr(datos_eerr)

    print("\n📊 ESTADÍSTICAS:")
    for key, val in stats.items():
        print(f"  {key.replace('_', ' ').title():.<30} {val}")

    print("\n📋 RESULTADO DE CLASIFICACIÓN:")
    for fila in filas_procesadas:
        cls = fila.clasificacion
        print(f"\n  Fila {fila.num_fila}: {fila.codigo} | {fila.glosa[:40]}")
        print(f"    → {cls.ln} | {cls.cc} | {cls.ca}")

    # Exportar
    exporter = EERRExporter()
    exporter.a_json(filas_procesadas, "eerr_clasificado.json")
    exporter.a_excel(filas_procesadas, "eerr_clasificado.xlsx", stats)

    print("\n✓ Proceso completado")
