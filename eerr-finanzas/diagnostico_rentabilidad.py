"""
DIAGNOSTICO PASO 1: Entender la estructura de "Análisis Resultado"
Inspecciona el sheet madre del archivo Análisis Contribución
"""

import pandas as pd
import openpyxl
from pathlib import Path

def diagnosticar_analisis_resultado():
    """Inspecciona el sheet 'Análisis Resultado'"""
    print("\n" + "="*80)
    print(" DIAGNOSTICO: ANALISIS RESULTADO (Sheet Madre)")
    print("="*80)

    ruta = Path("../data/planillas/Análisis Contribución 2026 V02.02.xlsx")

    try:
        # Leer con openpyxl para ver estructura exacta
        wb = openpyxl.load_workbook(ruta, data_only=True)

        print(f"\n[OK] Archivo: {ruta.name} (15 MB)")
        print(f"[Sheets totales] {len(wb.sheetnames)}")

        # Buscar sheet "Análisis Resultado" o similar
        sheet_analisis = None
        for sheet_name in wb.sheetnames:
            if "análisis resultado" in sheet_name.lower() or "analisis resultado" in sheet_name.lower():
                sheet_analisis = sheet_name
                break

        if not sheet_analisis:
            print(f"\n[AVISO] No encontró sheet 'Análisis Resultado'")
            print(f"Sheets disponibles:")
            for i, sheet in enumerate(wb.sheetnames[:10], 1):
                print(f"  {i}. {sheet}")
            return

        ws = wb[sheet_analisis]
        print(f"\n[Sheet encontrado] {sheet_analisis}")
        print(f"[Dimensiones] {ws.max_row} filas x {ws.max_column} columnas")
        print(f"[Rango] {ws.dimensions}")

        # Leer con Pandas para entender estructura
        df = pd.read_excel(ruta, sheet_name=sheet_analisis, header=None)

        print(f"\n[ESTRUCTURA] Primeras 20 filas y 10 columnas:")
        print(df.iloc[:20, :10].to_string())

        # Encontrar encabezados
        print(f"\n[ENCABEZADOS] Buscando fila de títulos...")
        for idx, row in df.iterrows():
            row_values = [str(v).lower() if pd.notna(v) else "" for v in row]

            # Buscar palabras clave
            if any(kw in " ".join(row_values) for kw in ["venta", "costo", "margen", "comision", "canal"]):
                print(f"\n  Fila {idx} PUEDE ser encabezado:")
                print(f"  {list(row)[:12]}")

        # Estadísticas
        print(f"\n[ESTADISTICAS]")
        print(f"  Filas no vacías: {df.notna().any(axis=1).sum()}")
        print(f"  Columnas no vacías: {df.notna().any(axis=0).sum()}")

        # Buscar datos de FEBRERO
        print(f"\n[BUSQUEDA] Datos de FEBRERO 2026...")
        for idx, row in df.iterrows():
            row_str = " ".join([str(v) for v in row if pd.notna(v)])
            if "febrero" in row_str.lower() or "2026" in row_str.lower():
                print(f"  Fila {idx}: {row_str[:100]}")

    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()

# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    diagnosticar_analisis_resultado()

    print("\n" + "="*80)
    print(" PROXIMOS PASOS:")
    print("="*80)
    print("""
1. Confirmar la estructura del sheet 'Análisis Resultado'
2. Identificar exactamente qué columnas hay (venta, costo, margen, comisiones)
3. Ver formato de fechas (mes/año)
4. Validar números de FEBRERO 2026
5. Luego: Ver cómo se estructura por CANAL DE VENTA
6. Luego: Mapear dónde vienen los datos de cada columna
    """)
