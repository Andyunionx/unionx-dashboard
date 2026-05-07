"""
PASO 4: Inyecta RAW agregado en "Análisis Resultado"
- Lee RAW procesado (desde Odoo o Excel)
- Agrega NUEVAS filas sin borrar histórico
- Formatea e inyecta en Análisis Contribución 2026.xlsx
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime

class InyectarRawAnalisis:
    """Inyecta datos RAW en sheet 'Análisis Resultado'"""

    def __init__(self):
        self.ruta_analisis = Path("../data/planillas/Análisis Contribución 2026 V02.02.xlsx")
        self.ruta_raw = Path("data/outputs/raw_desde_odoo_febrero_2026.csv")  # O raw_agregado_febrero_2026.csv
        self.sheet_destino = "Análisis Resultados"

        print(f"\n{'='*100}")
        print(" PASO 4: INYECTAR RAW EN ANÁLISIS RESULTADO")
        print(f"{'='*100}")

        print(f"\nArchivos:")
        print(f"  Análisis: {self.ruta_analisis}")
        print(f"  RAW input: {self.ruta_raw}")
        print(f"  Sheet destino: {self.sheet_destino}")

    def leer_raw(self) -> pd.DataFrame:
        """Lee el CSV del RAW agregado"""
        print(f"\n[Leyendo] RAW agregado...")

        # Intentar Odoo primero, luego Excel
        if self.ruta_raw.exists():
            df = pd.read_csv(self.ruta_raw)
        else:
            # Alternativa: usar Excel original convertido
            ruta_alt = Path("data/outputs/raw_agregado_febrero_2026.csv")
            if ruta_alt.exists():
                df = pd.read_csv(ruta_alt)
            else:
                print(f"[ERROR] No existe RAW agregado")
                return None

        print(f"[OK] {len(df)} filas leídas")
        return df

    def leer_analisis(self) -> pd.DataFrame:
        """Lee el sheet actual de Análisis Resultado"""
        print(f"\n[Leyendo] Análisis Resultado actual...")

        if not self.ruta_analisis.exists():
            print(f"[ERROR] No existe: {self.ruta_analisis}")
            return None

        df = pd.read_excel(self.ruta_analisis, sheet_name=self.sheet_destino, header=0)

        print(f"[OK] {len(df)} filas históricas")
        return df

    def preparar_inyeccion(self, df_raw: pd.DataFrame) -> pd.DataFrame:
        """Prepara el RAW para inyectar en Análisis Resultado"""
        print(f"\n[Preparando] Datos para inyección...")

        # Seleccionar y reordenar columnas para coincidir con Análisis Resultado
        # Estructura esperada en Análisis Resultado:
        # AÑO | Negocio | Canal | KAM | Mes | Trimestre | Venta | Costo Venta | Margen Directo | ...

        columnas_esperadas = [
            'AÑO', 'Mes', 'Canal', 'Tipo Negocio', 'KAM',
            'Venta', 'Costo Venta', 'Margen Directo', 'Cantidad'
        ]

        # Validar que existan
        cols_presentes = [c for c in columnas_esperadas if c in df_raw.columns]
        if len(cols_presentes) < len(columnas_esperadas):
            faltantes = [c for c in columnas_esperadas if c not in df_raw.columns]
            print(f"[AVISO] Faltan columnas: {faltantes}")

        # Crear DataFrame para inyectar
        df_inyeccion = df_raw[cols_presentes].copy()

        # Renombrar "Tipo Negocio" -> "Negocio" si es necesario
        if 'Tipo Negocio' in df_inyeccion.columns and 'Negocio' not in df_inyeccion.columns:
            df_inyeccion = df_inyeccion.rename(columns={'Tipo Negocio': 'Negocio'})

        # Agregar columnas adicionales que Análisis Resultado espera
        df_inyeccion['Trimestre'] = df_inyeccion['Mes'].apply(lambda m: (m - 1) // 3 + 1)
        df_inyeccion['Costo Venta'] = df_inyeccion.get('Costo Venta', 0)
        df_inyeccion['Margen Directo'] = df_inyeccion.get('Margen Directo', 0)
        df_inyeccion['Comisión Venta'] = df_inyeccion.get('Comisión', 0)
        df_inyeccion['Comisión Envío'] = 0
        df_inyeccion['Marketing'] = df_inyeccion.get('Marketing', 0)
        df_inyeccion['Total Comisiones'] = df_inyeccion.get('Comisión', 0)

        # Calcular Contribución
        df_inyeccion['Contribución'] = (
            df_inyeccion['Margen Directo'] -
            df_inyeccion['Total Comisiones'] -
            df_inyeccion['Marketing']
        )

        # Reordenar columnas para coincidir con Análisis Resultado
        cols_orden = [
            'AÑO', 'Negocio', 'Canal', 'KAM', 'Mes', 'Trimestre',
            'Venta', 'Costo Venta', 'Margen Directo', 'Comisión Venta',
            'Comisión Envío', 'Marketing', 'Total Comisiones', 'Contribución'
        ]

        # Solo incluir columnas que existan
        cols_final = [c for c in cols_orden if c in df_inyeccion.columns]
        df_inyeccion = df_inyeccion[cols_final]

        print(f"[OK] {len(df_inyeccion)} filas preparadas para inyectar")
        return df_inyeccion

    def inyectar_en_excel(self, df_actual: pd.DataFrame, df_nueva: pd.DataFrame) -> bool:
        """Inyecta datos nuevos en el Excel sin borrar histórico"""
        print(f"\n[Inyectando] en Excel...")

        try:
            # Cargar workbook
            wb = load_workbook(self.ruta_analisis)
            ws = wb[self.sheet_destino]

            # Encontrar última fila con datos
            ultima_fila_actual = ws.max_row

            print(f"  Última fila actual: {ultima_fila_actual}")
            print(f"  Nuevas filas a agregar: {len(df_nueva)}")

            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Agregar datos nuevos
            fila_inicio = ultima_fila_actual + 1

            for idx, (_, row) in enumerate(df_nueva.iterrows()):
                fila_actual = fila_inicio + idx

                for col_idx, col_name in enumerate(df_nueva.columns, 1):
                    cell = ws.cell(row=fila_actual, column=col_idx)
                    cell.value = row[col_name]
                    cell.border = border

                    # Formato numérico para números
                    if col_name in ['Venta', 'Costo Venta', 'Margen Directo', 'Comisión Venta',
                                    'Comisión Envío', 'Marketing', 'Total Comisiones', 'Contribución']:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    elif col_name in ['AÑO', 'Mes', 'Trimestre']:
                        cell.number_format = '0'
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(horizontal="left")

            # Guardar
            ruta_backup = self.ruta_analisis.parent / f"Análisis Contribución 2026 BACKUP {datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Crear backup
            import shutil
            shutil.copy(self.ruta_analisis, ruta_backup)
            print(f"  [OK] Backup creado: {ruta_backup.name}")

            # Guardar cambios
            wb.save(self.ruta_analisis)
            print(f"  [OK] Datos inyectados en: {self.ruta_analisis}")

            return True

        except Exception as e:
            print(f"[ERROR] No se pudo inyectar: {e}")
            import traceback
            traceback.print_exc()
            return False

    def validar_inyeccion(self, df_anterior: pd.DataFrame, df_nueva: pd.DataFrame) -> bool:
        """Valida que la inyección fue exitosa"""
        print(f"\n[Validando] Inyección...")

        # Leer nuevamente el archivo para confirmar
        df_despues = pd.read_excel(self.ruta_analisis, sheet_name=self.sheet_destino, header=0)

        filas_nuevas_esperadas = len(df_anterior) + len(df_nueva)
        filas_nuevas_reales = len(df_despues)

        print(f"  Filas históricas: {len(df_anterior)}")
        print(f"  Filas nuevas agregadas: {len(df_nueva)}")
        print(f"  Total esperado: {filas_nuevas_esperadas}")
        print(f"  Total actual: {filas_nuevas_reales}")

        if filas_nuevas_reales >= filas_nuevas_esperadas:
            print(f"[OK] Inyección validada exitosamente")
            return True
        else:
            print(f"[ERROR] Faltan filas. Verifica la inyección.")
            return False

    def ejecutar(self):
        """Ejecuta inyección completa"""

        # Leer datos
        df_raw = self.leer_raw()
        if df_raw is None or df_raw.empty:
            return False

        df_analisis_actual = self.leer_analisis()
        if df_analisis_actual is None:
            return False

        # Preparar inyección
        df_inyeccion = self.preparar_inyeccion(df_raw)
        if df_inyeccion is None or df_inyeccion.empty:
            return False

        # Inyectar
        exito = self.inyectar_en_excel(df_analisis_actual, df_inyeccion)

        if exito:
            # Validar
            self.validar_inyeccion(df_analisis_actual, df_inyeccion)

            print(f"\n{'='*100}")
            print(" INYECCION COMPLETADA")
            print(f"{'='*100}")
            print(f"\n[PROXIMO PASO] Las tablas dinámicas en 'Análisis Contribución' se actualizarán automáticamente")
            print(f"              con los nuevos datos de febrero 2026")

        return exito


# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    inyector = InyectarRawAnalisis()
    exito = inyector.ejecutar()

    if not exito:
        print("\n[PASOS MANUALES]")
        print("  1. Abre: Análisis Contribución 2026 V02.02.xlsx")
        print("  2. Ve a sheet: 'Análisis Resultados'")
        print("  3. Copia los datos preparados (CSV) al final del sheet")
        print("  4. Verifica que las tablas dinámicas se actualicen automáticamente")
