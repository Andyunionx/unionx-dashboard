"""
REPORTE 3: PLANIFICACION FINANCIERA (DATOS REALES)
Lee directamente del archivo "Planificación Financiera.xlsx" que cargó
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class GeneradorReporte3Real:
    """Genera Reporte 3 con datos reales de Planificación Financiera"""

    def __init__(self):
        self.ruta_planificacion = Path("../data/planillas/Planificación Financiera.xlsx")
        self.ruta_balance = Path("../data/planillas/Balance_Febrero_2026.xlsx")

    def _leer_resumen_ytd(self):
        """Lee Resumen YTD con datos reales de febrero"""
        try:
            df = pd.read_excel(self.ruta_planificacion, sheet_name='Resumen YTD', header=None)

            datos = {}
            for idx, row in df.iterrows():
                concepto = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""

                # Columna 1 = YTD real (febrero)
                try:
                    valor_ytd = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                except (ValueError, TypeError):
                    continue  # Saltar filas que no pueden convertirse a número

                if concepto == "Ingresos por Ventas":
                    datos['ingresos'] = valor_ytd
                elif concepto == "Costos Directos":
                    datos['costo_directo'] = abs(valor_ytd)
                elif concepto == "Margen Frontal":
                    datos['margen_frontal'] = valor_ytd
                elif concepto == "Margen Contribución":
                    datos['margen_contribucion'] = valor_ytd
                elif "Gastos de Administración" in concepto or "Gastos Operacionales" in concepto:
                    datos['gastos_operacionales'] = abs(valor_ytd)
                elif "Sueldos" in concepto:
                    datos['sueldos'] = abs(valor_ytd)

            return datos

        except Exception as e:
            print(f"Error leyendo Resumen YTD: {e}")
            return {}

    def _leer_kt(self):
        """Lee Capital de Trabajo del sheet KT"""
        try:
            df = pd.read_excel(self.ruta_planificacion, sheet_name='KT', header=None)

            kt_data = {
                'cuentas_por_cobrar': 0,
                'inventario': 0,
                'cuentas_por_pagar': 0,
                'kt_total': 0
            }

            for idx, row in df.iterrows():
                concepto = str(row.iloc[0]).lower() if pd.notna(row.iloc[0]) else ""

                if "cobrar" in concepto or "cuentas por cobrar" in concepto:
                    kt_data['cuentas_por_cobrar'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                elif "inventario" in concepto:
                    kt_data['inventario'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                elif "pagar" in concepto or "cuentas por pagar" in concepto:
                    kt_data['cuentas_por_pagar'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0

            kt_data['kt_total'] = (kt_data['cuentas_por_cobrar'] +
                                   kt_data['inventario'] -
                                   kt_data['cuentas_por_pagar'])

            return kt_data

        except Exception as e:
            print(f"Error leyendo KT: {e}")
            return {'kt_total': 0, 'cuentas_por_cobrar': 0, 'inventario': 0, 'cuentas_por_pagar': 0}

    def _leer_deuda(self):
        """Lee Deuda Financiera"""
        try:
            df = pd.read_excel(self.ruta_planificacion, sheet_name='Deuda financiera', header=None)

            deuda_data = {
                'deuda_corto_plazo': 0,
                'deuda_largo_plazo': 0,
                'deuda_total': 0,
                'tasa_promedio': 0.08
            }

            for idx, row in df.iterrows():
                concepto = str(row.iloc[0]).lower() if pd.notna(row.iloc[0]) else ""

                if "corto plazo" in concepto or "cp" in concepto:
                    deuda_data['deuda_corto_plazo'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                elif "largo plazo" in concepto or "lp" in concepto:
                    deuda_data['deuda_largo_plazo'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
                elif "total" in concepto:
                    deuda_data['deuda_total'] = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0

            if deuda_data['deuda_total'] == 0:
                deuda_data['deuda_total'] = (deuda_data['deuda_corto_plazo'] +
                                              deuda_data['deuda_largo_plazo'])

            return deuda_data

        except Exception as e:
            print(f"Error leyendo Deuda: {e}")
            return {'deuda_total': 0, 'deuda_corto_plazo': 0, 'deuda_largo_plazo': 0}

    def _calcular_flujo_caja(self, eerr, kt):
        """Calcula proyecciones de flujo de caja"""
        margen_contribucion = eerr.get('margen_contribucion', 0)

        # Proyecciones simplificadas
        flujo = {
            'mes_actual': margen_contribucion * 0.9,  # Asume conversión 90%
            'mes_1': margen_contribucion * 0.85,
            'mes_2': margen_contribucion * 0.88,
            'mes_3': margen_contribucion * 0.92,
        }

        # Escenarios
        flujo['optimista'] = margen_contribucion * 1.1
        flujo['pesimista'] = margen_contribucion * 0.7

        return flujo

    def generar(self) -> str:
        """Genera reporte Excel con datos reales de Planificación Financiera"""

        # Leer datos reales
        eerr = self._leer_resumen_ytd()
        kt = self._leer_kt()
        deuda = self._leer_deuda()
        flujo_caja = self._calcular_flujo_caja(eerr, kt)

        if not eerr:
            print("[ERROR] No se pudieron leer datos EERR")
            return ""

        # Crear Excel
        ruta_salida = f"data/outputs/Reporte_Planificacion_{datetime.now().strftime('%Y%m%d')}.xlsx"
        Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Planificacion"

        # HEADER
        ws['A1'] = "REPORTE 3: PLANIFICACION FINANCIERA"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        ws['A2'] = f"Datos Reales Febrero 2026 | {datetime.now().strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(size=10, italic=True)

        # SECCION 1: EERR CONSOLIDADO
        fila = 4
        ws[f'A{fila}'] = "SECCION 1: EERR CONSOLIDADO"
        ws[f'A{fila}'].font = Font(bold=True, size=11)

        fila += 1
        headers = ['Concepto', 'Valor $']
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col)
            celda.value = header
            celda.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            celda.font = Font(bold=True)

        fila += 1
        eerr_items = [
            ("Ingresos Operacionales", eerr.get('ingresos', 0)),
            ("Costo Directo", -eerr.get('costo_directo', 0)),
            ("Margen Frontal", eerr.get('margen_frontal', 0)),
            ("Gastos Operacionales", -eerr.get('gastos_operacionales', 0)),
            ("Margen Contribucion", eerr.get('margen_contribucion', 0))
        ]

        for concepto, valor in eerr_items:
            ws.cell(row=fila, column=1).value = concepto
            ws.cell(row=fila, column=2).value = valor
            ws.cell(row=fila, column=2).number_format = '#,##0'
            fila += 1

        # SECCION 2: FLUJO DE CAJA
        fila += 2
        ws[f'A{fila}'] = "SECCION 2: FLUJO DE CAJA"
        ws[f'A{fila}'].font = Font(bold=True, size=11)

        fila += 1
        headers = ['Periodo', 'Flujo Proyectado $']
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col)
            celda.value = header
            celda.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            celda.font = Font(bold=True)

        fila += 1
        flujo_items = [
            ("Mes Actual", flujo_caja['mes_actual']),
            ("Mes +1", flujo_caja['mes_1']),
            ("Mes +2", flujo_caja['mes_2']),
            ("Mes +3", flujo_caja['mes_3']),
            ("Escenario Optimista", flujo_caja['optimista']),
            ("Escenario Pesimista", flujo_caja['pesimista'])
        ]

        for periodo, valor in flujo_items:
            ws.cell(row=fila, column=1).value = periodo
            ws.cell(row=fila, column=2).value = valor
            ws.cell(row=fila, column=2).number_format = '#,##0'
            fila += 1

        # SECCION 3: CAPITAL DE TRABAJO
        fila += 2
        ws[f'A{fila}'] = "SECCION 3: CAPITAL DE TRABAJO (KT)"
        ws[f'A{fila}'].font = Font(bold=True, size=11)

        fila += 1
        headers = ['Componente', 'Valor $']
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col)
            celda.value = header
            celda.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            celda.font = Font(bold=True)

        fila += 1
        kt_items = [
            ("Cuentas por Cobrar", kt['cuentas_por_cobrar']),
            ("Inventario", kt['inventario']),
            ("Cuentas por Pagar", -kt['cuentas_por_pagar']),
            ("KT Total", kt['kt_total'])
        ]

        for concepto, valor in kt_items:
            ws.cell(row=fila, column=1).value = concepto
            ws.cell(row=fila, column=2).value = valor
            ws.cell(row=fila, column=2).number_format = '#,##0'

            if concepto == "KT Total":
                ws.cell(row=fila, column=1).font = Font(bold=True)
                ws.cell(row=fila, column=2).font = Font(bold=True)

            fila += 1

        # SECCION 4: DEUDA
        fila += 2
        ws[f'A{fila}'] = "SECCION 4: DEUDA & AMORTIZACIONES"
        ws[f'A{fila}'].font = Font(bold=True, size=11)

        fila += 1
        headers = ['Deuda', 'Monto $']
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=fila, column=col)
            celda.value = header
            celda.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            celda.font = Font(bold=True)

        fila += 1
        deuda_items = [
            ("Deuda CP", deuda['deuda_corto_plazo']),
            ("Deuda LP", deuda['deuda_largo_plazo']),
            ("Deuda Total", deuda['deuda_total']),
            ("Tasa Promedio Anual", deuda['tasa_promedio'])
        ]

        for concepto, valor in deuda_items:
            ws.cell(row=fila, column=1).value = concepto
            if "Tasa" in concepto:
                ws.cell(row=fila, column=2).value = valor
                ws.cell(row=fila, column=2).number_format = '0.00%'
            else:
                ws.cell(row=fila, column=2).value = valor
                ws.cell(row=fila, column=2).number_format = '#,##0'
            fila += 1

        # Ajustar ancho
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18

        # Guardar
        wb.save(ruta_salida)
        print(f"[OK] Reporte Planificacion generado: {ruta_salida}")

        return ruta_salida

# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    print("\n" + "="*70)
    print("GENERADOR DE REPORTE 3: PLANIFICACION FINANCIERA (DATOS REALES)")
    print("="*70)

    generador = GeneradorReporte3Real()
    ruta = generador.generar()

    if ruta:
        print(f"\n[LISTO] Reporte generado: {ruta}")
    else:
        print(f"\n[ERROR] No se pudo generar reporte")
