"""
REPORTE 3: Planificacin Financiera (Flujo Caja, KT, Deuda, Proyecciones)
Automatizado cada lunes 9 AM

Entrada (Arquitectura de 2 Excels):
  EXCEL #1: "Planificacin Financiera"  PRINCIPAL para Reporte 3
   Proyecciones de flujo de caja futuro
   EERR (descargado manual de Odoo, pero pronto automtico)
   Sueldos (Excel desde BUK)
   Honorarios (Excel con distribucin manual por CC)
   Rendiciones (Excel con distribucin manual por CC)
   Deuda (Excel balance, pronto Odoo)
   Balance
   Ajustes forecast

  EXCEL #2: "P&L Comparativo"  SECUNDARIO
   Comparar presupuesto vs gasto real (anlisis de diferencias)

Salida:
  - Excel consolidado con flujo caja, KT, deuda, proyecciones
  - Email ejecutivo a CEO
  - Alertas si flujo proyectado es negativo
"""

import json
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class GeneradorReporte3:
    """Genera Reporte 3: Planificacin Financiera"""

    def __init__(
        self,
        ruta_planificacion: str,
        ruta_eerr: str,
        ruta_sueldos: Optional[str] = None,
        ruta_deuda: Optional[str] = None,
        ruta_p_l_comparativo: Optional[str] = None
    ):
        """
        Args:
            ruta_planificacion: Excel "Planificacin Financiera" (PRINCIPAL)
            ruta_eerr: EERR desde Odoo (JSON o Excel)
            ruta_sueldos: Excel sueldos desde BUK
            ruta_deuda: Excel balance con deuda
            ruta_p_l_comparativo: Excel "P&L Comparativo" (SECUNDARIO)
        """
        self.ruta_planificacion = ruta_planificacion
        self.ruta_eerr = ruta_eerr
        self.ruta_sueldos = ruta_sueldos
        self.ruta_deuda = ruta_deuda
        self.ruta_p_l_comparativo = ruta_p_l_comparativo

        self.planificacion = self._cargar_planificacion()
        self.eerr = self._cargar_eerr()
        self.sueldos = self._cargar_sueldos()
        self.deuda = self._cargar_deuda()
        self.p_l_comp = self._cargar_p_l_comparativo()

        self.mes_actual = datetime.now().month
        self.ao_actual = datetime.now().year
        self.hoy = datetime.now()

    def _cargar_planificacion(self) -> Dict:
        """Carga Excel de Planificacin Financiera (PRINCIPAL)"""
        if not Path(self.ruta_planificacion).exists():
            return {}

        try:
            df = pd.read_excel(self.ruta_planificacion, sheet_name='Planificacin')
            return df.to_dict('list')
        except Exception as e:
            print(f"Error cargando Planificacin: {e}")
            return {}

    def _cargar_eerr(self) -> Dict:
        """Carga EERR desde Odoo o JSON"""
        if not Path(self.ruta_eerr).exists():
            return {}

        try:
            # Intentar JSON primero
            if self.ruta_eerr.endswith('.json'):
                with open(self.ruta_eerr, 'r', encoding='utf-8') as f:
                    return json.load(f)
            # Si no, intentar Excel
            else:
                df = pd.read_excel(self.ruta_eerr)
                return df.to_dict('list')
        except Exception as e:
            print(f"Error cargando EERR: {e}")
            return {}

    def _cargar_sueldos(self) -> Dict:
        """Carga Excel sueldos desde BUK/contador"""
        if not self.ruta_sueldos or not Path(self.ruta_sueldos).exists():
            return {}

        try:
            df = pd.read_excel(self.ruta_sueldos)
            return df.set_index('Empleado').to_dict('index')
        except Exception as e:
            print(f"Error cargando sueldos: {e}")
            return {}

    def _cargar_deuda(self) -> Dict:
        """Carga deuda desde Excel balance"""
        if not self.ruta_deuda or not Path(self.ruta_deuda).exists():
            return {}

        try:
            df = pd.read_excel(self.ruta_deuda, sheet_name='Deuda')
            return df.to_dict('list')
        except Exception as e:
            print(f"Error cargando deuda: {e}")
            return {}

    def _cargar_p_l_comparativo(self) -> Dict:
        """Carga P&L Comparativo (SECUNDARIO, para anlista)"""
        if not self.ruta_p_l_comparativo or not Path(self.ruta_p_l_comparativo).exists():
            return {}

        try:
            df = pd.read_excel(self.ruta_p_l_comparativo)
            return df.to_dict('list')
        except Exception as e:
            print(f"Error cargando P&L Comparativo: {e}")
            return {}

    def generar(self) -> str:
        """
        Genera el reporte Planificacin Financiera completo

        Returns:
            Ruta al archivo Excel generado
        """

        # 1. Consolidar ingresos y egresos
        consolidacion = self._consolidar_eerr_sueldos_gastos()

        # 2. Calcular flujo de caja histrico y proyectado
        flujo_caja = self._calcular_flujo_caja(consolidacion)

        # 3. Calcular KT (Capital de Trabajo)
        kt = self._calcular_kt(consolidacion)

        # 4. Proyectar deuda y amortizaciones
        proyeccion_deuda = self._proyectar_deuda()

        # 5. Generar Excel
        ruta_salida = self._generar_excel(
            consolidacion,
            flujo_caja,
            kt,
            proyeccion_deuda
        )

        # 6. Validar alertas (flujo negativo)
        self._validar_alertas(flujo_caja)

        return ruta_salida

    def _consolidar_eerr_sueldos_gastos(self) -> Dict:
        """Consolida EERR + Sueldos + Honorarios + Rendiciones"""

        # ESTRUCTURA ESPERADA:
        # - EERR con ingresos y costos
        # - Sueldos con distribucin por empleado
        # - Honorarios con distribucin por CC (MANUAL)
        # - Rendiciones con distribucin por CC (MANUAL)

        consolidacion = {
            'mes': f"{self.mes_actual}/{self.ao_actual}",
            'ingresos_operacionales': 0,
            'costo_venta': 0,
            'margen_bruto': 0,
            'gastos_operacionales': 0,
            'remuneraciones': 0,
            'otros_gastos': 0,
            'ebit': 0,
            'gastos_financieros': 0,
            'impuestos': 0,
            'utilidad_neta': 0,
            'detalles_por_cc': {},
        }

        # Extraer del EERR
        if self.eerr:
            consolidacion['ingresos_operacionales'] = sum(
                float(v.get('monto', 0))
                for v in self.eerr.get('ingresos', [])
            )
            consolidacion['costo_venta'] = sum(
                float(v.get('monto', 0))
                for v in self.eerr.get('costos', [])
            )

        # Calcular margen bruto
        consolidacion['margen_bruto'] = (
            consolidacion['ingresos_operacionales'] - consolidacion['costo_venta']
        )

        # Agregar sueldos
        consolidacion['remuneraciones'] = sum(
            float(v.get('monto', 0)) for v in self.sueldos.values()
        )

        # Agregar otros gastos (desde EERR)
        consolidacion['otros_gastos'] = sum(
            float(v.get('monto', 0))
            for v in self.eerr.get('gastos_operacionales', [])
        )

        # Calcular EBIT
        consolidacion['ebit'] = (
            consolidacion['margen_bruto'] - consolidacion['remuneraciones'] - consolidacion['otros_gastos']
        )

        # Gastos financieros y impuestos (del EERR)
        consolidacion['gastos_financieros'] = sum(
            float(v.get('monto', 0))
            for v in self.eerr.get('gastos_financieros', [])
        )
        consolidacion['impuestos'] = sum(
            float(v.get('monto', 0))
            for v in self.eerr.get('impuestos', [])
        )

        # Utilidad neta
        consolidacion['utilidad_neta'] = (
            consolidacion['ebit'] - consolidacion['gastos_financieros'] - consolidacion['impuestos']
        )

        return consolidacion

    def _calcular_flujo_caja(self, consolidacion: Dict) -> Dict:
        """Calcula flujo de caja histrico y proyecciones"""

        # CLCULO SIMPLE:
        # Flujo = Utilidad Neta + Depreciacin - Cambio en CT - Pago de deuda

        flujo_operativo = consolidacion['utilidad_neta']

        # Proyecciones simplificadas (3 escenarios)
        flujo_proyectado = {
            'mes_actual': flujo_operativo,
            'proximo_mes': flujo_operativo * 0.95,  # Asumir pequea cada
            'mes_3': flujo_operativo * 0.98,
            'escenario_optimista': flujo_operativo * 1.1,
            'escenario_pesimista': flujo_operativo * 0.85,
        }

        # Saldo acumulado
        saldo_hoy = 50000  # TODO: obtener de estado de cuenta real
        saldo_proyectado = {}

        saldo_acumulado = saldo_hoy
        for periodo in ['mes_actual', 'proximo_mes', 'mes_3']:
            saldo_acumulado += flujo_proyectado[periodo]
            saldo_proyectado[periodo] = saldo_acumulado

        flujo_proyectado['saldo_proyectado'] = saldo_proyectado

        return flujo_proyectado

    def _calcular_kt(self, consolidacion: Dict) -> Dict:
        """Calcula Capital de Trabajo (activos circulantes - pasivos circulantes)"""

        # ESTRUCTURA:
        # KT = Cuentas por cobrar + Inventario - Cuentas por pagar

        # VALORES EJEMPLO (TODO: obtener de balance)
        cuentas_por_cobrar = 120000
        inventario = 80000
        cuentas_por_pagar = 95000

        kt = cuentas_por_cobrar + inventario - cuentas_por_pagar

        # KT necesario (estimado como % de ingresos)
        kt_necesario = consolidacion['ingresos_operacionales'] * 0.15  # 15% de ingresos

        return {
            'cuentas_por_cobrar': cuentas_por_cobrar,
            'inventario': inventario,
            'cuentas_por_pagar': cuentas_por_pagar,
            'kt_actual': kt,
            'kt_necesario': kt_necesario,
            'kt_exceso_deficit': kt - kt_necesario,
        }

    def _proyectar_deuda(self) -> Dict:
        """Proyecta deuda y amortizaciones"""

        # ESTRUCTURA ESPERADA DE EXCEL DEUDA:
        # - Prstamo 1: monto, tasa, plazo
        # - Prstamo 2: ...

        # VALORES EJEMPLO (TODO: obtener de Excel deuda)
        deuda_total_actual = 500000
        tasa_promedio_anual = 0.08  # 8%

        # Proyectar prximos 12 meses
        proyecciones = {}
        saldo = deuda_total_actual

        for mes in range(1, 13):
            interes_mes = saldo * (tasa_promedio_anual / 12)
            amortizacion_mes = 5000  # Pago fijo aproximado
            saldo -= amortizacion_mes

            proyecciones[f'mes_{mes}'] = {
                'saldo_inicial': saldo + amortizacion_mes,
                'interes': interes_mes,
                'amortizacion': amortizacion_mes,
                'saldo_final': max(0, saldo),
            }

        return {
            'deuda_total_actual': deuda_total_actual,
            'tasa_promedio_anual': tasa_promedio_anual,
            'proyecciones_12_meses': proyecciones,
        }

    def _generar_excel(
        self,
        consolidacion: Dict,
        flujo: Dict,
        kt: Dict,
        deuda: Dict
    ) -> str:
        """Genera Excel con todas las secciones"""

        ruta_salida = f"data/outputs/Reporte_Planificacion_{datetime.now().strftime('%Y%m%d')}.xlsx"
        Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Planificacin"

        # HEADER
        self._agregar_header(ws)

        # SECCIN 1: EERR Consolidado
        self._agregar_seccion_eerr(ws, consolidacion)

        # SECCIN 2: Flujo de Caja
        self._agregar_seccion_flujo(ws, flujo)

        # SECCIN 3: KT
        self._agregar_seccion_kt(ws, kt)

        # SECCIN 4: Deuda y Proyecciones
        self._agregar_seccion_deuda(ws, deuda)

        # Guardar
        wb.save(ruta_salida)
        print(f" Reporte Planificacin generado: {ruta_salida}")

        return ruta_salida

    def _agregar_header(self, ws):
        """Agrega header"""
        ws['A1'] = "REPORTE 3: PLANIFICACIN FINANCIERA"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')} | Perodo: {self.mes_actual}/{self.ao_actual}"
        ws['A2'].font = Font(size=10, italic=True)

    def _agregar_seccion_eerr(self, ws, consolidacion: Dict):
        """Agrega EERR consolidado"""

        fila = 4
        ws[f'A{fila}'] = "1. EERR CONSOLIDADO"
        ws[f'A{fila}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{fila}'].fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
        ws.merge_cells(f'A{fila}:B{fila}')

        fila += 1
        datos_eerr = [
            ("Ingresos Operacionales", consolidacion['ingresos_operacionales']),
            ("Costo de Venta", consolidacion['costo_venta']),
            ("Margen Bruto", consolidacion['margen_bruto']),
            ("Remuneraciones", consolidacion['remuneraciones']),
            ("Otros Gastos", consolidacion['otros_gastos']),
            ("EBIT", consolidacion['ebit']),
            ("Gastos Financieros", consolidacion['gastos_financieros']),
            ("Impuestos", consolidacion['impuestos']),
            ("Utilidad Neta", consolidacion['utilidad_neta']),
        ]

        for concepto, valor in datos_eerr:
            ws.cell(row=fila, column=1, value=concepto).font = Font(bold=True)
            ws.cell(row=fila, column=2, value=valor).number_format = '$#,##0'
            fila += 1

    def _agregar_seccion_flujo(self, ws, flujo: Dict):
        """Agrega proyeccin de flujo de caja"""

        fila = ws.max_row + 3
        ws[f'A{fila}'] = "2. FLUJO DE CAJA Y PROYECCIONES"
        ws[f'A{fila}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{fila}'].fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
        ws.merge_cells(f'A{fila}:B{fila}')

        fila += 1
        datos_flujo = [
            ("Flujo Operativo (Mes Actual)", flujo['mes_actual']),
            ("Proyeccin Prximo Mes", flujo['proximo_mes']),
            ("Proyeccin Mes +3", flujo['mes_3']),
            ("", None),
            ("Escenario Optimista", flujo['escenario_optimista']),
            ("Escenario Pesimista", flujo['escenario_pesimista']),
        ]

        for concepto, valor in datos_flujo:
            if valor is None:
                fila += 1
                continue

            ws.cell(row=fila, column=1, value=concepto).font = Font(bold=True if valor > 0 else None)
            ws.cell(row=fila, column=2, value=valor).number_format = '$#,##0'

            # Colorear si es negativo
            if valor < 0:
                ws.cell(row=fila, column=2).font = Font(color="D32F2F", bold=True)

            fila += 1

        # Saldos proyectados
        fila += 1
        ws.cell(row=fila, column=1, value="SALDOS PROYECTADOS").font = Font(bold=True)
        fila += 1

        for periodo, saldo in flujo.get('saldo_proyectado', {}).items():
            ws.cell(row=fila, column=1, value=f"Saldo {periodo}")
            ws.cell(row=fila, column=2, value=saldo).number_format = '$#,##0'

            # Alerta si es negativo
            if saldo < 0:
                ws.cell(row=fila, column=2).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

            fila += 1

    def _agregar_seccion_kt(self, ws, kt: Dict):
        """Agrega Capital de Trabajo"""

        fila = ws.max_row + 3
        ws[f'A{fila}'] = "3. CAPITAL DE TRABAJO (KT)"
        ws[f'A{fila}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{fila}'].fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
        ws.merge_cells(f'A{fila}:B{fila}')

        fila += 1
        datos_kt = [
            ("Cuentas por Cobrar", kt['cuentas_por_cobrar']),
            ("Inventario", kt['inventario']),
            ("Cuentas por Pagar", -kt['cuentas_por_pagar']),
            ("KT ACTUAL", kt['kt_actual']),
            ("KT Necesario", kt['kt_necesario']),
            ("KT Exceso / (Dficit)", kt['kt_exceso_deficit']),
        ]

        for concepto, valor in datos_kt:
            ws.cell(row=fila, column=1, value=concepto).font = Font(bold=True)
            ws.cell(row=fila, column=2, value=valor).number_format = '$#,##0'
            fila += 1

    def _agregar_seccion_deuda(self, ws, deuda: Dict):
        """Agrega deuda y proyecciones"""

        fila = ws.max_row + 3
        ws[f'A{fila}'] = "4. DEUDA Y AMORTIZACIONES"
        ws[f'A{fila}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{fila}'].fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
        ws.merge_cells(f'A{fila}:B{fila}')

        fila += 1
        ws.cell(row=fila, column=1, value="Deuda Total Actual").font = Font(bold=True)
        ws.cell(row=fila, column=2, value=deuda['deuda_total_actual']).number_format = '$#,##0'

        fila += 1
        ws.cell(row=fila, column=1, value="Tasa Promedio Anual").font = Font(bold=True)
        ws.cell(row=fila, column=2, value=deuda['tasa_promedio_anual']).number_format = '0.0%'

        fila += 2
        ws.cell(row=fila, column=1, value="Proyecciones 12 Meses (primeros 3 meses)").font = Font(italic=True, size=9)

        fila += 1
        headers = ['Perodo', 'Saldo Inicial', 'Inters', 'Amortizacin', 'Saldo Final']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=fila, column=col, value=header)
            cell.font = Font(bold=True, size=9)

        fila += 1
        # Mostrar solo primeros 3 meses por brevedad
        for mes in range(1, 4):
            proy = deuda['proyecciones_12_meses'].get(f'mes_{mes}', {})
            ws.cell(row=fila, column=1, value=f"Mes {mes}")
            ws.cell(row=fila, column=2, value=proy.get('saldo_inicial', 0)).number_format = '$#,##0'
            ws.cell(row=fila, column=3, value=proy.get('interes', 0)).number_format = '$#,##0'
            ws.cell(row=fila, column=4, value=proy.get('amortizacion', 0)).number_format = '$#,##0'
            ws.cell(row=fila, column=5, value=proy.get('saldo_final', 0)).number_format = '$#,##0'
            fila += 1

    def _validar_alertas(self, flujo: Dict):
        """Valida alertas (flujo negativo)"""

        alertas = []

        # Verificar si algn saldo proyectado es negativo
        for periodo, saldo in flujo.get('saldo_proyectado', {}).items():
            if saldo < 0:
                alertas.append(f" ALERTA: Flujo negativo proyectado en {periodo} (${saldo:,.0f})")

        if alertas:
            print("\n ALERTAS DETECTADAS:")
            for alerta in alertas:
                print(f"  {alerta}")

        return alertas


# ============================================================================
# SCRIPT EJECUTABLE (lunes 9 AM)
# ============================================================================

if __name__ == "__main__":
    # Rutas
    ruta_planificacion = "data/planillas/Planificacin Financiera.xlsx"
    ruta_eerr = "data/outputs/02 EE.RR Febrero 2026_CLASIFICADO.json"
    ruta_sueldos = "data/planillas/Sueldos_Febrero_2026.xlsx"
    ruta_deuda = "data/planillas/Balance_Febrero_2026.xlsx"

    # Generar
    generador = GeneradorReporte3(
        ruta_planificacion,
        ruta_eerr,
        ruta_sueldos,
        ruta_deuda
    )
    ruta_salida = generador.generar()

    print(f"\n Reporte 3 completo: {ruta_salida}")
