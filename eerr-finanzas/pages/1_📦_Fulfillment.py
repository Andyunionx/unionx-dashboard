"""
Pagina Fulfillment — KPIs operacionales del area Fulfillment del Plan UnionX 2026-2028.

Estructura:
  Tab 1: Resumen — cards principales
  Tab 2: Inventario (Odoo automatico) — DIO ABC, slow movers, stockouts SKUs A
  Tab 3: Operacion bodega (manual + roadmap) — OFR, OCT, Pick Accuracy, Inventory Accuracy
  Tab 4: Costo / productividad (mixto)
  Tab 5: Carga manual + plantillas
"""
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import plotly.graph_objects as go

HERE = Path(__file__).resolve().parent
PARENT = HERE.parent
PROJECT_ROOT = PARENT.parent
sys.path.insert(0, str(PARENT))
sys.path.insert(0, str(PROJECT_ROOT))

from auth_helper import require_login, get_user_roles, has_any_role  # noqa: E402

require_login()
roles = get_user_roles(st.session_state.get("username", ""))
can_upload = has_any_role(roles, ["admin", "uploader"])

st.set_page_config(page_title="Fulfillment - UnionX", page_icon="📦", layout="wide")
st.title("📦 Fulfillment")
st.caption("KPIs del area Fulfillment del Plan Estrategico UnionX 2026-2028. "
           "Verde = automatico desde Odoo · Amarillo = manual con uploader · Naranja = esperando WMS H2.")

MANUAL_KPIS_FILE = PROJECT_ROOT / "data" / "kpis_manuales" / "fulfillment.json"


@st.cache_data(ttl=900, show_spinner="Consultando Odoo (puede tomar 30-60s la primera vez)…")
def _cargar_kpis_odoo():
    sys.path.insert(0, str(PARENT))
    from kpis_odoo import (
        kpi_abc_inventario, kpi_slow_movers, kpi_stockouts_skus_a, get_odoo_client
    )
    odoo = get_odoo_client()
    if odoo is None:
        return {"error": "Odoo no disponible (verificar ANDRES_ODOO_PASSWORD env var)"}
    return {
        "abc": kpi_abc_inventario(odoo),
        "slow_movers": kpi_slow_movers(odoo, dias=180),
        "stockouts": kpi_stockouts_skus_a(odoo),
        "error": None,
    }


def _cargar_kpis_manuales():
    """Lee el JSON de KPIs manuales (OFR, OCT, accuracy, etc.)."""
    if not MANUAL_KPIS_FILE.exists():
        return {}
    try:
        with open(MANUAL_KPIS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _guardar_kpi_manual(mes_anio: str, kpi_key: str, valor: float):
    """Persiste un KPI manual en el JSON."""
    MANUAL_KPIS_FILE.parent.mkdir(parents=True, exist_ok=True)
    data = _cargar_kpis_manuales()
    if mes_anio not in data:
        data[mes_anio] = {}
    data[mes_anio][kpi_key] = {"valor": valor, "ts": datetime.now().isoformat()}
    with open(MANUAL_KPIS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _ultimo_valor_manual(kpi_key: str):
    """Devuelve (mes, valor) del último input manual."""
    data = _cargar_kpis_manuales()
    if not data:
        return None, None
    mes_sorted = sorted(data.keys(), reverse=True)
    for m in mes_sorted:
        if kpi_key in data[m]:
            return m, data[m][kpi_key]["valor"]
    return None, None


# ============================================================================
# Helper: cargar costo operativo total desde P&L de Planificacion
# ============================================================================
COSTO_OP_CONFIG = PROJECT_ROOT / "config" / "costo_operativo.yaml"


def _cargar_config_costo():
    """Lee config/costo_operativo.yaml. Devuelve dict componentes."""
    try:
        import yaml
        with open(COSTO_OP_CONFIG, encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
        return cfg.get("componentes", {})
    except Exception:
        return {}


def _guardar_config_costo(componentes_dict):
    """Persiste cambios en factor de cada componente."""
    try:
        import yaml
        # Leer estructura completa actual
        with open(COSTO_OP_CONFIG, encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
        # Actualizar solo factors
        for key, factor in componentes_dict.items():
            if key in cfg.get("componentes", {}):
                cfg["componentes"][key]["factor"] = float(factor)
        with open(COSTO_OP_CONFIG, "w", encoding="utf-8") as f:
            yaml.safe_dump(cfg, f, allow_unicode=True, sort_keys=False, default_flow_style=False)
        return True
    except Exception as e:
        return str(e)


@st.cache_data(ttl=900)
def _cargar_costo_operativo_pl(year: int = None, _config_hash: str = ""):
    """Calcula el costo operativo total desde P&L aplicando los factores del yaml.

    El parametro _config_hash sirve para invalidar cache cuando cambian los factors.

    Returns dict con: fijos_total, variables_total, total, detalle, mes_a_mes
    """
    from openpyxl import load_workbook
    import sys as _sys
    _proj = str(Path(__file__).resolve().parent.parent.parent)
    if _proj not in _sys.path:
        _sys.path.insert(0, _proj)
    import shared_paths as _sp

    if year is None:
        year = datetime.now().year

    try:
        wb = load_workbook(_sp.PLANIFICACION_FINANCIERA, read_only=True, data_only=True)
        ws = wb["P&L"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return {"error": str(e)[:120]}

    if not rows or len(rows) < 4:
        return {"error": "P&L sin datos"}

    fechas_row = rows[2]
    cols_year = []
    for i, v in enumerate(fechas_row):
        if hasattr(v, 'year') and v.year == year:
            cols_year.append((i, v))
    if not cols_year:
        return {"error": f"Sin meses para {year}"}

    config_componentes = _cargar_config_costo()
    if not config_componentes:
        return {"error": "config/costo_operativo.yaml vacio o no encontrado"}

    detalle = []
    fijos_total = 0
    variables_total = 0
    mes_a_mes = {f.month: {"fijo": 0, "variable": 0, "total": 0} for _, f in cols_year}

    for key, comp in config_componentes.items():
        idx = comp.get("row_idx")
        factor = float(comp.get("factor", 1.0))
        cat = comp.get("categoria", "variable").lower()
        label = comp.get("label", key)
        if idx is None or idx >= len(rows):
            continue
        suma_bruta = 0
        suma_aplicada = 0
        valores_mes = {}
        for col_idx, fecha in cols_year:
            v = rows[idx][col_idx] if col_idx < len(rows[idx]) else None
            if isinstance(v, (int, float)):
                v_abs = abs(v)
                v_aplicado = v_abs * factor
                suma_bruta += v_abs
                suma_aplicada += v_aplicado
                valores_mes[fecha.month] = v_aplicado
                mes_a_mes[fecha.month][cat] = mes_a_mes[fecha.month].get(cat, 0) + v_aplicado
                mes_a_mes[fecha.month]["total"] = mes_a_mes[fecha.month].get("total", 0) + v_aplicado
        detalle.append({
            "key": key,
            "Categoria": "Fijo" if cat == "fijo" else "Variable",
            "Concepto": label,
            "Total Cuenta YTD": suma_bruta,
            "Factor": factor,
            "Atribuido al Costo Op.": suma_aplicada,
            "Mensual": valores_mes,
        })
        if cat == "fijo":
            fijos_total += suma_aplicada
        else:
            variables_total += suma_aplicada

    return {
        "year": year,
        "fijos_total": fijos_total,
        "variables_total": variables_total,
        "total": fijos_total + variables_total,
        "detalle": detalle,
        "mes_a_mes": mes_a_mes,
        "error": None,
    }


# ============================================================================
# TABS
# ============================================================================
tabs = st.tabs([
    "📊 Resumen",
    "📦 Stock en vivo (Odoo)",
    "🔤 ABC / Slow movers",
    "🎯 Operación bodega",
    "💰 Costo Operativo Total",
    "📥 Carga manual" + ("" if can_upload else " 🔒"),
])

# Cargar datos Odoo (cached)
odoo_data = _cargar_kpis_odoo()

# ----------------------------------------------------------------------------
# Tab 1: Resumen
# ----------------------------------------------------------------------------
with tabs[0]:
    if odoo_data.get("error"):
        st.warning(f"⚠️ Odoo: {odoo_data['error']}")
    else:
        abc = odoo_data["abc"].get("valor") or []
        slow = odoo_data["slow_movers"].get("valor") or []
        stockouts = odoo_data["stockouts"].get("valor")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            n_skus_a = sum(1 for s in abc if s.get("clase") == "A")
            st.metric("📦 SKUs Clase A", f"{n_skus_a}", help="Top 80% del valor de venta")
        with c2:
            so_color = "🟢" if stockouts is not None and stockouts <= 0.03 else ("🟡" if stockouts is not None and stockouts <= 0.10 else "🔴")
            st.metric(f"{so_color} Stockouts SKUs A", f"{stockouts*100:.1f}%" if stockouts is not None else "—",
                      help="Meta: ≤3% según plan")
        with c3:
            st.metric("🐢 Slow movers", f"{len(slow)}", help="SKUs sin venta en últimos 180 días con stock > 0")
        with c4:
            valor_slow = sum(s.get("stock", 0) for s in slow)
            st.metric("$ Slow movers", f"{valor_slow:,.0f} u", help="Unidades inmovilizadas")

    st.divider()

    # KPIs manuales (último valor cargado)
    st.markdown("#### KPIs operacionales (último valor manual)")
    manuales_keys = [
        ("ofr", "🎯 OFR (Order Fulfillment Rate)", "≥97%", "%"),
        ("oct_b2c", "⏱️ Order Cycle Time B2C", "≤24h", "h"),
        ("pick_accuracy", "✅ Pick & Pack Accuracy", "≥99.5%", "%"),
        ("inventory_accuracy", "📊 Inventory Accuracy", "≥98%", "%"),
        ("productividad_picking", "👷 Productividad picking", "60-120 líneas/h", "líneas/h"),
        ("costo_pedido", "💵 Costo / Pedido", "↓15% YoY", "$"),
    ]
    cols_m = st.columns(3)
    for i, (key, label, meta, unit) in enumerate(manuales_keys):
        col = cols_m[i % 3]
        mes, val = _ultimo_valor_manual(key)
        col.metric(
            label,
            f"{val:.1f} {unit}" if val is not None else "—",
            help=f"Meta: {meta}. {f'Cargado en {mes}' if mes else 'Sin datos manuales aún. Carga en tab 5.'}"
        )

# ----------------------------------------------------------------------------
# Tab 2 NUEVO: Stock en vivo (delega a stock_dashboard.py via runpy)
# ----------------------------------------------------------------------------
with tabs[1]:
    st.markdown("### 📦 Stock en vivo desde Odoo")
    st.caption("Datos en tiempo real de Odoo. Si tarda, dale ~30s — primera consulta carga 600+ SKUs.")
    # Marcar como embebido para evitar conflicto de set_page_config en stock_dashboard.py
    st.session_state["_embedded_context"] = True
    try:
        import runpy
        _stock_path = str(PARENT / "stock_dashboard.py")
        runpy.run_path(_stock_path, run_name="__stock_in_fulfillment__")
    except Exception as _e_stock:
        st.error(f"No se pudo cargar stock_dashboard: {_e_stock}")
        st.info("Alternativa: http://localhost:8503 (próximamente — dashboard Stock standalone).")
    finally:
        # Liberar el flag para no afectar otras paginas
        st.session_state["_embedded_context"] = False

# ----------------------------------------------------------------------------
# Tab 3: ABC / Slow movers (era Inventario Odoo)
# ----------------------------------------------------------------------------
with tabs[2]:
    if odoo_data.get("error"):
        st.error(odoo_data["error"])
    else:
        abc = odoo_data["abc"].get("valor") or []
        slow = odoo_data["slow_movers"].get("valor") or []

        st.markdown("### 📊 Clasificación ABC (últimos 365 días)")
        if abc:
            df_abc = pd.DataFrame(abc)
            counts = df_abc["clase"].value_counts().to_dict()
            ventas_clase = df_abc.groupby("clase")["venta"].sum().to_dict()

            c1, c2, c3 = st.columns(3)
            for i, clase in enumerate(["A", "B", "C"]):
                col = [c1, c2, c3][i]
                n = counts.get(clase, 0)
                v = ventas_clase.get(clase, 0)
                col.metric(
                    f"Clase {clase}",
                    f"{n} SKUs",
                    delta=f"${v/1e6:,.0f}M venta YTD",
                    help=f"Clase {clase}: {'top 80% valor' if clase=='A' else 'siguiente 15%' if clase=='B' else 'bottom 5%'}"
                )

            st.divider()

            st.markdown("**Top 30 SKUs (Clase A)**")
            df_top = df_abc[df_abc["clase"] == "A"].head(30).copy()
            df_top["venta"] = df_top["venta"].apply(lambda v: f"${v/1e6:,.1f}M")
            df_top["pct_acumulado"] = df_top["pct_acumulado"].apply(lambda v: f"{v*100:.1f}%")
            df_top["qty_vendida"] = df_top["qty_vendida"].apply(lambda v: f"{v:,.0f}")
            st.dataframe(
                df_top[["nombre", "qty_vendida", "venta", "pct_acumulado", "clase"]],
                use_container_width=True, hide_index=True
            )

        st.divider()

        st.markdown("### 🐢 Slow Movers (sin venta en 180 días)")
        if slow:
            df_slow = pd.DataFrame(slow[:50]).copy()
            df_slow["stock"] = df_slow["stock"].apply(lambda v: f"{v:,.0f} u")
            st.dataframe(df_slow[["nombre", "stock"]], use_container_width=True, hide_index=True)
            st.caption(f"Mostrando top 50 de {len(slow)} slow movers totales.")
        else:
            st.success("✅ No hay slow movers detectados.")

# ----------------------------------------------------------------------------
# Tab 4: Operación bodega (manual + roadmap)
# ----------------------------------------------------------------------------
with tabs[3]:
    st.markdown("### 🎯 Operación de bodega — KPIs del Plan Estratégico")
    st.caption(
        "Estos KPIs requieren WMS o capa de medición sobre el sistema actual. "
        "**Hoy: input manual** · **Roadmap H2:** integración WMS para tiempo real."
    )

    operacion_kpis = [
        ("ofr", "🎯 OFR (Order Fulfillment Rate)", "% pedidos completos en SLA", "≥ 97%"),
        ("oct_b2c", "⏱️ Order Cycle Time B2C", "Tiempo orden→courier", "≤ 24h"),
        ("oct_b2b", "⏱️ Order Cycle Time B2B", "Tiempo orden→courier", "≤ 48h"),
        ("pick_accuracy", "✅ Pick & Pack Accuracy", "% pedidos sin error", "≥ 99.5%"),
        ("inventory_accuracy", "📊 Inventory Accuracy", "Stock sistema vs físico (cycle counting)", "≥ 98%"),
        ("tiempo_recepcion", "📥 Tiempo recepción contenedor", "Llegada→disponible", "≤ 48h"),
    ]

    for key, label, desc, meta in operacion_kpis:
        with st.container(border=True):
            c1, c2, c3 = st.columns([2, 1, 1])
            with c1:
                st.markdown(f"**{label}**")
                st.caption(desc)
            with c2:
                mes, val = _ultimo_valor_manual(key)
                if val is not None:
                    st.metric("Último", f"{val:.1f}", help=f"Cargado: {mes}")
                else:
                    st.markdown("**Sin datos** 🟠")
                    st.caption("→ Tab 5 para cargar")
            with c3:
                st.markdown(f"**Meta:** {meta}")
                st.caption("Roadmap H2: WMS")

# ----------------------------------------------------------------------------
# Tab 5: Costo Operativo Total (NUEVO) + Productividad
# ----------------------------------------------------------------------------
with tabs[4]:
    st.markdown("### 💰 Costo Operativo Total")

    # Selector de año + selector de fuente
    col_year, col_source = st.columns([1, 2])
    with col_year:
        year_costo = st.selectbox("Año", [2026, 2025, 2024], index=0, key="costo_year")
    with col_source:
        # Si hay archivo cargado por terceros, ofrecer elegir
        from costo_operativo_uploader import cargar as _cargar_co_detalle
        data_detalle = _cargar_co_detalle(year_costo)
        tiene_detalle = bool(data_detalle.get("conceptos"))
        opciones_fuente = ["📊 P&L con factores (estimado)"]
        if tiene_detalle:
            opciones_fuente.append("📋 Carga detallada por terceros (real)")
        fuente_sel = st.radio("Fuente de cálculo", opciones_fuente, horizontal=True, key="costo_fuente")

    # =================================================================
    # Sub-tabs dentro del Tab 5
    # =================================================================
    sub_tabs = st.tabs([
        "📊 Análisis del costo",
        "📥 Carga por terceros",
        "⚙️ Configuración (P&L factores)",
    ])

    # =================================================================
    # SUB-TAB 1: Análisis del costo (la vista principal)
    # =================================================================
    with sub_tabs[0]:
        usar_detalle = (fuente_sel.startswith("📋"))

        if usar_detalle and tiene_detalle:
            # ---- Modo: archivo cargado por terceros ----
            from costo_operativo_uploader import analizar as _analizar_co

            # Intentar cruzar con ventas mensuales de Odoo (si hay)
            ventas_m_odoo = {}
            try:
                from kpis_odoo import get_odoo_client
                odoo = get_odoo_client()
                if odoo:
                    desde = f"{year_costo}-01-01"
                    hasta = f"{year_costo}-12-31"
                    ventas = odoo.search_read(
                        'sale.order',
                        [('date_order', '>=', desde), ('date_order', '<=', hasta),
                         ('state', 'in', ['sale', 'done'])],
                        ['date_order', 'amount_total'], limit=200000,
                    )
                    for o in ventas:
                        try:
                            d = datetime.strptime(o['date_order'][:10], '%Y-%m-%d')
                            ventas_m_odoo[d.month] = ventas_m_odoo.get(d.month, 0) + (o.get('amount_total', 0) / 1e3)
                        except Exception:
                            pass
            except Exception:
                pass

            analisis = _analizar_co(data_detalle, ventas_mensuales=ventas_m_odoo)
            if analisis.get("error"):
                st.warning(analisis["error"])
            else:
                fijos = analisis["fijos_total"]
                variables = analisis["variables_total"]
                total = analisis["total_costo"]
                ratio = analisis["ratio_variable"]

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Costo Fijo YTD", f"{fijos/1e3:,.0f} MM$")
                c2.metric("Costo Variable YTD", f"{variables/1e3:,.0f} MM$")
                c3.metric("Costo Operativo Total", f"{total/1e3:,.0f} MM$")
                c4.metric("% Variable / Total", f"{ratio*100:.1f}%")

                if analisis.get("costo_vs_venta") is not None:
                    cv = analisis["costo_vs_venta"]
                    cv_color = "🟢" if 0.08 <= cv <= 0.14 else ("🟡" if 0.14 < cv <= 0.20 else "🔴")
                    st.metric(f"{cv_color} Costo Operativo / Venta", f"{cv*100:.1f}%",
                              help="Benchmark Plan UnionX: 8-14% según ticket. Cruce con ventas Odoo.")

                st.divider()

                # Comparacion con benchmarks de mercado
                st.markdown("#### 📊 Comparación vs benchmarks de mercado")
                if analisis["comparacion"]:
                    rows_bm = []
                    for c in analisis["comparacion"]:
                        rows_bm.append({
                            "KPI": c["kpi"],
                            "Tu valor": c["valor"],
                            "Benchmark mercado": c["benchmark"],
                            "Estado": c["estado"],
                            "Interpretación": c["interpretacion"],
                        })
                    st.dataframe(pd.DataFrame(rows_bm), use_container_width=True, hide_index=True)

                st.divider()

                # Recomendaciones automaticas
                st.markdown("#### 💡 Recomendaciones automáticas (heurísticas)")
                if not analisis["recomendaciones"]:
                    st.success("✅ Sin alertas críticas. Estructura de costos en rangos esperados.")
                else:
                    for r in analisis["recomendaciones"]:
                        with st.container(border=True):
                            sev_col, msg_col = st.columns([1, 4])
                            with sev_col:
                                st.markdown(f"**{r['prioridad']}**")
                                st.caption(r["area"])
                            with msg_col:
                                st.markdown(f"**{r['mensaje']}**")
                                st.markdown(f"*Acción sugerida:* {r['accion']}")

                st.divider()

                # Top conceptos
                st.markdown("#### 🏆 Top 5 conceptos por monto")
                if analisis["top_conceptos"]:
                    rows_top = []
                    for c in analisis["top_conceptos"]:
                        rows_top.append({
                            "Categoría": c["categoria"],
                            "Concepto": c["concepto"],
                            "Total YTD": f"{c['total']/1e3:,.0f} MM$",
                            "% del total": f"{c['total']/total*100:.1f}%" if total else "—",
                        })
                    st.dataframe(pd.DataFrame(rows_top), use_container_width=True, hide_index=True)

                # Mes a mes apilado
                st.markdown("#### 📈 Evolución mensual (Fijo vs Variable)")
                mes_a_mes = analisis["mes_a_mes"]
                meses_con_datos = sorted([m for m, v in mes_a_mes.items() if v["total"] > 0])
                if meses_con_datos:
                    meses_es_local = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
                    x_labels = [meses_es_local[m-1] for m in meses_con_datos]
                    fijos_y = [mes_a_mes[m].get("fijo", 0)/1e3 for m in meses_con_datos]
                    vars_y = [mes_a_mes[m].get("variable", 0)/1e3 for m in meses_con_datos]
                    fig = go.Figure()
                    fig.add_trace(go.Bar(name='Fijos', x=x_labels, y=fijos_y, marker_color='#F59E0B'))
                    fig.add_trace(go.Bar(name='Variables', x=x_labels, y=vars_y, marker_color='#3B82F6'))
                    fig.update_layout(barmode='stack', height=320, margin=dict(l=20, r=20, t=20, b=20),
                        yaxis_title="MM$", hovermode='x unified',
                        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
                    st.plotly_chart(fig, use_container_width=True)

                meta = data_detalle.get("_meta", {})
                if meta:
                    st.caption(f"📅 Datos cargados: {meta.get('ts','')[:16]} · {analisis['n_conceptos']} conceptos.")

        else:
            # ---- Modo: estimado desde P&L con factores ----
            st.caption(
                "Modo **estimado**: tomamos cuentas del P&L de Planificación y aplicamos factores de atribución "
                "(% del total que corresponde a Operaciones). Para datos exactos, cargá un archivo en Tab 'Carga por terceros'."
            )

            # Modo estimado P&L
            cfg_actual = _cargar_config_costo()
            cfg_hash = "|".join(f"{k}:{v.get('factor', 1.0)}" for k, v in sorted(cfg_actual.items()))
            costo_data = _cargar_costo_operativo_pl(year_costo, _config_hash=cfg_hash)
            if costo_data.get("error"):
                st.warning(f"⚠️ {costo_data['error']}")
            else:
                c1, c2, c3, c4 = st.columns(4)
                fijos = costo_data["fijos_total"]
                variables = costo_data["variables_total"]
                total = costo_data["total"]
                ratio = (variables / total) if total else 0
                c1.metric("Costo Fijo YTD", f"{fijos/1e3:,.0f} MM$")
                c2.metric("Costo Variable YTD", f"{variables/1e3:,.0f} MM$")
                c3.metric("Costo Operativo Total", f"{total/1e3:,.0f} MM$")
                c4.metric("% Variable / Total", f"{ratio*100:.1f}%")
                st.divider()
                st.markdown("#### Detalle por concepto — YTD")
                rows_costo = []
                for d in costo_data["detalle"]:
                    rows_costo.append({
                        "Categoría": d["Categoria"],
                        "Concepto": d["Concepto"],
                        "Total cuenta": f"{d['Total Cuenta YTD']/1e3:,.0f} MM$" if d["Total Cuenta YTD"] else "—",
                        "Factor": f"{d['Factor']*100:.0f}%",
                        "Atribuido a Op.": f"{d['Atribuido al Costo Op.']/1e3:,.0f} MM$" if d["Atribuido al Costo Op."] else "—",
                        "% del total": f"{d['Atribuido al Costo Op.']/total*100:.1f}%" if total else "—",
                    })

                def highlight_cat(row):
                    if row["Categoría"] == "Fijo":
                        return ['background-color: #FEF3C7'] * len(row)
                    return ['background-color: #DBEAFE'] * len(row)

                df_costo = pd.DataFrame(rows_costo)
                st.dataframe(df_costo.style.apply(highlight_cat, axis=1), use_container_width=True, hide_index=True)

                st.markdown("#### Evolución mensual (Fijo vs Variable)")
                meses_es = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
                mes_a_mes = costo_data["mes_a_mes"]
                meses_con_datos = sorted(mes_a_mes.keys())
                if meses_con_datos:
                    x_labels = [meses_es[m-1] for m in meses_con_datos]
                    fijos_y = [mes_a_mes[m].get("fijo", 0)/1e3 for m in meses_con_datos]
                    vars_y = [mes_a_mes[m].get("variable", 0)/1e3 for m in meses_con_datos]
                    fig = go.Figure()
                    fig.add_trace(go.Bar(name='Fijos', x=x_labels, y=fijos_y, marker_color='#F59E0B'))
                    fig.add_trace(go.Bar(name='Variables', x=x_labels, y=vars_y, marker_color='#3B82F6'))
                    fig.update_layout(barmode='stack', height=320, margin=dict(l=20, r=20, t=20, b=20),
                        yaxis_title="MM$", hovermode='x unified',
                        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
                    st.plotly_chart(fig, use_container_width=True)

                st.info(
                    "💡 Para mayor exactitud, pedile al equipo de RRHH/Contabilidad que cargue el archivo "
                    "detallado en la pestaña **'Carga por terceros'**. Eso desbloquea: análisis vs benchmark, "
                    "recomendaciones automáticas, top conceptos."
                )

    # =================================================================
    # SUB-TAB 2: Carga por terceros (uploader + template)
    # =================================================================
    with sub_tabs[1]:
        st.markdown("#### 📥 Carga de archivo detallado por terceros")
        st.caption(
            "Para que **RRHH, Contabilidad u otro tercero** suban el detalle real de costos operativos del mes. "
            "Una vez cargado, el dashboard analiza vs benchmarks de mercado y sugiere mejoras."
        )

        col_a, col_b = st.columns([1, 1])
        with col_a:
            st.markdown("##### 1️⃣ Descargar template")
            try:
                from costo_operativo_uploader import generar_template
                template_bytes = generar_template(year_costo)
                st.download_button(
                    f"📥 Descargar template Costo Operativo {year_costo}",
                    data=template_bytes,
                    file_name=f"costo_operativo_{year_costo}_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                st.caption(
                    "El template incluye conceptos sugeridos (Sueldos Operaciones, Arriendo Megacentro, "
                    "Flete despacho, Insumos packing, etc.). Podés agregar nuevas filas si necesitas."
                )
            except Exception as e:
                st.error(f"No se pudo generar template: {e}")

        with col_b:
            st.markdown("##### 2️⃣ Subir archivo completado")
            if not can_upload:
                st.info("🔒 Solo admin/uploader pueden subir archivos. Compartí el template y subir vos cuando lo recibas.")
            else:
                up = st.file_uploader(
                    f"Excel con costos {year_costo}",
                    type=["xlsx", "xlsm"],
                    key=f"upload_costo_op_{year_costo}",
                )
                if up is not None:
                    try:
                        from costo_operativo_uploader import parsear_archivo, guardar
                        data_parseada = parsear_archivo(up, year_costo)
                        n = len(data_parseada.get("conceptos", []))
                        if n == 0:
                            st.error("❌ El archivo no tiene datos válidos. Verificá categorías ('Fijo'/'Variable') y valores numéricos.")
                        else:
                            st.success(f"✅ {n} conceptos detectados. Guardando…")
                            guardar(data_parseada, year_costo)
                            st.cache_data.clear()
                            st.info("Recargá la pestaña 'Análisis del costo' para ver el resultado.")
                    except Exception as e:
                        st.error(f"❌ Error al procesar: {e}")

        # Mostrar info del último archivo cargado
        if tiene_detalle:
            st.divider()
            st.markdown("##### 📊 Último archivo cargado")
            meta = data_detalle.get("_meta", {})
            n = len(data_detalle.get("conceptos", []))
            st.success(
                f"📅 Cargado el {meta.get('ts','')[:16]} · {n} conceptos · año {meta.get('year', year_costo)}"
            )
            with st.expander("Ver conceptos cargados", expanded=False):
                conceptos_view = []
                for c in data_detalle.get("conceptos", []):
                    conceptos_view.append({
                        "Categoría": c["categoria"],
                        "Concepto": c["concepto"],
                        "Total YTD": f"{c['total']/1e3:,.0f} MM$",
                        "Notas": c.get("nota", ""),
                    })
                st.dataframe(pd.DataFrame(conceptos_view), use_container_width=True, hide_index=True)

    # =================================================================
    # SUB-TAB 3: Configuración (factores P&L estimado)
    # =================================================================
    with sub_tabs[2]:
        st.markdown("#### ⚙️ Configuración de factores (modo estimado P&L)")
        st.caption(
            "Aplica solo al modo estimado. Cada cuenta del P&L se multiplica por un factor "
            "(% atribuible a Operaciones). Si tenés el archivo cargado por terceros, esto pierde relevancia."
        )

        if not can_upload:
            st.info("🔒 Solo admin/uploader pueden modificar factores.")
        else:
            cfg = _cargar_config_costo()
            with st.form("config_costo_op_v2"):
                nuevos_factores = {}
                for key, comp in cfg.items():
                    cat_emoji = "🟡 Fijo" if comp.get("categoria") == "fijo" else "🔵 Variable"
                    label = comp.get("label", key)
                    factor_actual = float(comp.get("factor", 1.0))
                    nota = comp.get("nota", "")
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        st.markdown(f"**{cat_emoji} · {label}**")
                        st.caption(nota)
                    with col2:
                        nuevos_factores[key] = st.slider(
                            "Factor",
                            min_value=0.0, max_value=1.0,
                            value=factor_actual, step=0.05,
                            key=f"factor_v2_{key}",
                            label_visibility="collapsed", format="%.0f%%",
                        )
                guardar_btn = st.form_submit_button("💾 Guardar nuevos factores", type="primary")
                if guardar_btn:
                    res = _guardar_config_costo(nuevos_factores)
                    if res is True:
                        st.success("✅ Factores guardados. Recargando…")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(f"❌ Error: {res}")

        st.divider()
        st.markdown("##### 🛣️ Roadmap de automatización del Costo Operativo")
        st.markdown("""
- **🟢 Hoy (corto plazo):** Modo estimado P&L con factores configurables · Carga por terceros con template Excel.
- **🟡 Mediano plazo (3-6 meses):** Análisis automatizado de los archivos cargados con benchmarks de mercado y recomendaciones (✅ implementado).
- **🟠 Mediano-largo (6-18 meses):** Migrar contabilidad a Odoo con **cuentas analíticas** o **centros de costo** ('Operaciones', 'Comercial', 'Admin'). Una vez migrado, el costo operativo se calcula 100% automático sin necesidad de uploaders.
- **🔵 Largo (18-24m+):** Modelo predictivo: forecast de costo operativo basado en estacionalidad y volumen de venta proyectado.
        """)

    st.divider()

    st.markdown("### ⚙️ Productividad y costos unitarios (manual)")
    st.caption("Estos KPIs requieren WMS o medición manual.")

    costo_kpis = [
        ("costo_pedido", "💵 Costo de fulfillment / pedido",
         "Calculable: Σ gasto bodega P&L / Σ pedidos Odoo", "↓ 10-15% YoY"),
        ("productividad_picking", "👷 Productividad picking",
         "Líneas pickeadas / hora-persona", "60-120 líneas/h B2C"),
        ("tiempo_recepcion_h", "📥 Tiempo recepción contenedor",
         "Horas desde llegada a disponible", "≤ 48 hrs"),
    ]

    for key, label, desc, meta in costo_kpis:
        with st.container(border=True):
            c1, c2 = st.columns([3, 1])
            with c1:
                st.markdown(f"**{label}**")
                st.caption(desc)
                mes, val = _ultimo_valor_manual(key)
                if val is not None:
                    st.metric("Último valor", f"{val:.2f}", help=f"Cargado: {mes}")
            with c2:
                st.markdown(f"**Meta:** {meta}")

    st.info(
        "📊 **Costo / pedido automatizable**: cuando se conecte el sincronizador Odoo y "
        "se identifique la cuenta contable de gasto bodega en P&L, esto pasa a 🟢 automático."
    )

# ----------------------------------------------------------------------------
# Tab 6: Carga manual + plantilla
# ----------------------------------------------------------------------------
with tabs[5]:
    if not can_upload:
        st.info("🔒 Sección solo para admin/uploader. Ingresa con cuenta autorizada.")
        st.stop()

    st.markdown("### 📥 Carga manual de KPIs operacionales")
    st.caption("Carga el valor mensual de cada KPI. Se persiste en `data/kpis_manuales/fulfillment.json`.")

    # Selector de mes
    col_mes, _ = st.columns([1, 3])
    with col_mes:
        mes_sel = st.selectbox(
            "Mes / Año",
            [f"{datetime.now().year}-{m:02d}" for m in range(1, 13)],
            index=datetime.now().month - 1,
        )

    todos_kpis = [
        ("ofr", "OFR (%) — Order Fulfillment Rate"),
        ("oct_b2c", "Order Cycle Time B2C (horas)"),
        ("oct_b2b", "Order Cycle Time B2B (horas)"),
        ("pick_accuracy", "Pick & Pack Accuracy (%)"),
        ("inventory_accuracy", "Inventory Accuracy (%)"),
        ("productividad_picking", "Productividad picking (líneas/hora)"),
        ("costo_pedido", "Costo de fulfillment / pedido ($)"),
        ("tiempo_recepcion", "Tiempo recepción contenedor (horas)"),
    ]

    st.markdown("**Ingresá los valores que tengas (los demás dejá en blanco)**")

    valores_form = {}
    cols_form = st.columns(2)
    for i, (key, label) in enumerate(todos_kpis):
        col = cols_form[i % 2]
        # Pre-cargar último valor si existe para este mes
        data_existente = _cargar_kpis_manuales().get(mes_sel, {}).get(key, {}).get("valor")
        valores_form[key] = col.number_input(
            label, value=float(data_existente) if data_existente is not None else 0.0,
            step=0.01, key=f"input_{key}_{mes_sel}", format="%.2f"
        )

    if st.button("💾 Guardar todos", type="primary"):
        try:
            guardados = 0
            for key, val in valores_form.items():
                if val and val > 0:
                    _guardar_kpi_manual(mes_sel, key, val)
                    guardados += 1
            st.success(f"✅ {guardados} KPI(s) guardado(s) para {mes_sel}.")
            st.cache_data.clear()
        except Exception as e:
            st.error(f"❌ Error: {e}")

    st.divider()

    # Tabla de histórico cargado
    st.markdown("### 📊 Histórico cargado")
    data_all = _cargar_kpis_manuales()
    if data_all:
        rows_hist = []
        for mes_anio in sorted(data_all.keys(), reverse=True):
            for key, info in data_all[mes_anio].items():
                rows_hist.append({
                    "Mes": mes_anio,
                    "KPI": key,
                    "Valor": info.get("valor"),
                    "Cargado": info.get("ts", "")[:16],
                })
        st.dataframe(pd.DataFrame(rows_hist), use_container_width=True, hide_index=True)
    else:
        st.info("Sin datos cargados todavía. Cargá el primer mes arriba.")

    st.divider()
    st.caption(
        "🟠 **Roadmap H2 (6-18 meses):** integrar WMS para tener OFR/OCT/Accuracy en tiempo real. "
        "Una vez integrado, esta tab pasará a ser solo de auditoría histórica."
    )
