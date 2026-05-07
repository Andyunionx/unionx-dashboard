"""
Pagina Planificacion Financiera - vistas + carga mensual de EERR/Balance.

Source of truth: data/planillas/Planificación Financiera.xlsx (29 hojas).
Estructura visualizada:
  Tab 1 Resumen YTD   - hoja "Resumen YTD"
  Tab 2 P&L           - hoja "P&L" + "Ppto 2026"
  Tab 3 EEFF/Balance  - hoja "EEFF" + "Ref Balances"
  Tab 4 Analisis      - "Analisis Financiero 2026" + "Comparativo"
  Tab 5 KT/Deuda/PP&E - hojas correspondientes
  Tab 6 CARGA MENSUAL - solo admin/uploader
"""
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px

HERE = Path(__file__).resolve().parent
PARENT = HERE.parent
PROJECT_ROOT = PARENT.parent
sys.path.insert(0, str(PARENT))
sys.path.insert(0, str(PROJECT_ROOT))

from auth_helper import require_login  # noqa: E402
import shared_paths as sp  # noqa: E402

require_login()

st.set_page_config(page_title="Planificacion - UnionX", page_icon="📋", layout="wide")
st.title("📋 Planificación Financiera")
st.caption(f"Fuente: `{sp.PLANIFICACION_FINANCIERA.name}` · "
           f"Última modificación: {datetime.fromtimestamp(sp.PLANIFICACION_FINANCIERA.stat().st_mtime).strftime('%d/%m/%Y %H:%M') if sp.PLANIFICACION_FINANCIERA.exists() else 'N/A'}")

# -------------------------------------------------------------------- helpers
@st.cache_data(ttl=300)
def cargar_hoja(path_str: str, sheet_name: str) -> pd.DataFrame:
    """Lee una hoja del Excel maestro (cached 5min)."""
    return pd.read_excel(path_str, sheet_name=sheet_name, engine="openpyxl")


@st.cache_data(ttl=300)
def cargar_dashboard_data(path_str: str) -> dict:
    """Lee la hoja Dashboard Data y devuelve estructura parseada.

    Layout de la hoja:
      filas 1-16: bloque mensual (Meta/Resultado/Var% para 5 metricas)
      filas 18-23: KPI por trimestre (Q1-Q4 + Total) Var% acumulado
    """
    from openpyxl import load_workbook
    wb = load_workbook(path_str, read_only=True, data_only=True)
    if "Dashboard Data" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Dashboard Data"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    meses = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

    # Bloque mensual: filas 2-16 (idx 1-15)
    metricas = ["Venta", "Contrib.", "GAV", "EBIT", "Utilidad"]
    mensual = {}
    for i, m in enumerate(metricas):
        base = 1 + i*3  # filas 2,5,8,11,14 (Meta), 3,6,9,12,15 (Resultado), 4,7,10,13,16 (Var%)
        try:
            mensual[m] = {
                "Meta":      list(rows[base][1:13]),
                "Resultado": list(rows[base+1][1:13]),
                "Var%":      list(rows[base+2][1:13]),
            }
        except Exception:
            pass

    # Bloque trimestral: filas 19-23 (idx 18-22), columnas B-F = Q1,Q2,Q3,Q4,Total
    metricas_q = ["Venta", "Contribucion", "GAV", "EBIT", "Utilidad"]
    trimestral = {}
    for i, m in enumerate(metricas_q):
        try:
            trimestral[m] = {
                "Q1":    rows[18+i][1],
                "Q2":    rows[18+i][2],
                "Q3":    rows[18+i][3],
                "Q4":    rows[18+i][4],
                "Total": rows[18+i][5],
            }
        except Exception:
            pass

    return {"meses": meses, "mensual": mensual, "trimestral": trimestral}


def _fmt_pct(v):
    if v is None:
        return "—"
    try:
        return f"{float(v)*100:+.1f}%"
    except Exception:
        return str(v)


def _fmt_num(v):
    if v is None:
        return "—"
    try:
        return f"{float(v):,.1f}"
    except Exception:
        return str(v)


def _color_var(v):
    """Devuelve emoji semaforo segun Var%."""
    if v is None:
        return "⚪"
    try:
        x = float(v)
        if x >= -0.05:
            return "🟢"
        if x >= -0.15:
            return "🟡"
        return "🔴"
    except Exception:
        return "⚪"


@st.cache_data(ttl=300)
def cargar_metas_2026(path_str: str) -> dict:
    """Lee la hoja Metas 2026 y devuelve dict {metric: {Meta, Resultado, Var, Var%, Resultado2025, Variacion, VariacionPct}}.

    Layout: cada metrica ocupa 9 filas (etiqueta + 8 filas de datos).
    Bloques en filas: 1-9 (Venta), 10-18 (Contrib), 19-27 (GAV), 28-36 (EBIT), 37-45 (Utilidad).
    """
    from openpyxl import load_workbook
    wb = load_workbook(path_str, read_only=True, data_only=True)
    if "Metas 2026" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Metas 2026"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    metricas = ["Venta", "Contribucion", "GAV", "EBIT", "Utilidad"]
    out = {}
    for i, m in enumerate(metricas):
        base = i * 9
        try:
            out[m] = {
                "Meta":           list(rows[base+1][1:13]),
                "Resultado":      list(rows[base+2][1:13]),
                "Var":            list(rows[base+3][1:13]),
                "Var%":           list(rows[base+4][1:13]),
                "Resultado2025":  list(rows[base+5][1:13]),
                "Variacion":      list(rows[base+6][1:13]),
                "VariacionPct":   list(rows[base+7][1:13]),
            }
        except Exception:
            pass
    return out


@st.cache_data(ttl=300)
def cargar_analisis_financiero(path_str: str) -> dict:
    """Lee 'Análisis Financiero 2026'. Devuelve secciones: flujo_caja, deuda, ratios_deuda, kpis."""
    from openpyxl import load_workbook
    wb = load_workbook(path_str, read_only=True, data_only=True)
    sname = "Análisis Financiero 2026"
    if sname not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    def parse_table(start_row, end_row):
        """Parse tabla con header en start_row-1 y datos hasta end_row."""
        result = []
        for i in range(start_row - 1, min(end_row, len(rows))):
            row = rows[i]
            if not row or not row[0]:
                continue
            result.append({
                "concepto": str(row[0]).strip() if row[0] else "",
                "v_2025":   row[1] if len(row) > 1 else None,
                "v_2026":   row[2] if len(row) > 2 else None,
                "var_abs":  row[3] if len(row) > 3 else None,
                "var_pct":  row[4] if len(row) > 4 else None,
                "nota":     row[6] if len(row) > 6 else None,
            })
        return result

    return {
        "flujo_caja":   parse_table(7, 20),    # rows 7-19
        "deuda":        parse_table(25, 32),   # rows 25-31
        "ratios_deuda": parse_table(35, 38),   # rows 35-37
        "kpis":         parse_table(45, 70),   # rows 45-65 aprox (saltea headers de seccion)
    }


def _detectar_columna_inicio(rows: list) -> tuple[int, list]:
    """Encuentra la fila de fechas y la primera columna numerica. Devuelve (col_start, fechas).

    Recorre filas 1-15 buscando la primera con muchas fechas seguidas.
    """
    for r_idx in range(min(15, len(rows))):
        row = rows[r_idx]
        if not row:
            continue
        fechas = [(i, v) for i, v in enumerate(row) if hasattr(v, 'year')]
        if len(fechas) >= 12:
            col_start = fechas[0][0]
            return col_start, [v for i, v in fechas]
    return 0, []


@st.cache_data(ttl=300)
def cargar_kt_resumen(path_str: str) -> dict:
    """Lee hoja KT correctamente: detecta col donde empiezan los datos via headers de fecha."""
    from openpyxl import load_workbook
    wb = load_workbook(path_str, read_only=True, data_only=True)
    if "KT" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["KT"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    col_start, fechas = _detectar_columna_inicio(rows)
    if not fechas:
        return {}

    def serie(idx):
        """Devuelve list de (fecha, valor) para la fila idx desde col_start."""
        if idx >= len(rows):
            return []
        out = []
        for i, fecha in enumerate(fechas):
            col = col_start + i
            if col < len(rows[idx]):
                v = rows[idx][col]
                if isinstance(v, (int, float)):
                    out.append((fecha, v))
        return out

    def label(idx, col=3):
        if idx >= len(rows) or not rows[idx]:
            return ""
        return str(rows[idx][col]).strip() if rows[idx][col] else ""

    return {
        "fechas": fechas,
        "Existencias":          {"label": label(8), "serie": serie(8)},
        "CxC Comerciales":      {"label": label(9), "serie": serie(9)},
        "Otros Act Cor":        {"label": label(10), "serie": serie(10)},
        "Total Act Corrientes": {"label": label(11), "serie": serie(11)},
        "CxP Comerciales":      {"label": label(14), "serie": serie(14)},
        "Otros Pas Cor":        {"label": label(15), "serie": serie(15)},
        "Total Pas Corrientes": {"label": label(16), "serie": serie(16)},
        "KT Neto":              {"label": label(18), "serie": serie(18)},
        "Cambios en KT":        {"label": label(19), "serie": serie(19)},
        "Meses de Inventario":  {"label": label(22), "serie": serie(22)},
        "Meses de CxC":         {"label": label(24), "serie": serie(24)},
    }


@st.cache_data(ttl=300)
def cargar_deuda_resumen(path_str: str) -> dict:
    """Lee hoja Deuda financiera correctamente."""
    from openpyxl import load_workbook
    wb = load_workbook(path_str, read_only=True, data_only=True)
    if "Deuda financiera" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Deuda financiera"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    col_start, fechas = _detectar_columna_inicio(rows)
    if not fechas:
        return {}

    def serie(idx):
        if idx >= len(rows):
            return []
        out = []
        for i, fecha in enumerate(fechas):
            col = col_start + i
            if col < len(rows[idx]):
                v = rows[idx][col]
                if isinstance(v, (int, float)):
                    out.append((fecha, v))
        return out

    def label(idx, col=3):
        if idx >= len(rows) or not rows[idx]:
            return ""
        return str(rows[idx][col]).strip() if rows[idx][col] else ""

    return {
        "fechas": fechas,
        "Saldo inicial":   {"label": label(7), "serie": serie(7)},
        "Emision":         {"label": label(8), "serie": serie(8)},
        "Amortizacion":    {"label": label(9), "serie": serie(9)},
        "Saldo final":     {"label": label(10), "serie": serie(10)},
        "Balance prom":    {"label": label(15), "serie": serie(15)},
        "Tasa interes":    {"label": label(16), "serie": serie(16)},
        "Gasto intereses": {"label": label(17), "serie": serie(17)},
    }


@st.cache_data(ttl=300)
def cargar_pl_proyectado(path_str: str, year_filter: int = None) -> dict:
    """Lee hoja P&L y devuelve estructura tabular limpia con conceptos vs meses.

    Args:
        year_filter: si se da, solo devuelve meses de ese año (ej: 2026)
    """
    from openpyxl import load_workbook
    wb = load_workbook(path_str, read_only=True, data_only=True)
    if "P&L" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["P&L"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    col_start, fechas = _detectar_columna_inicio(rows)
    if not fechas:
        return {}

    # Filtrar por año si corresponde
    indices_meses = list(range(len(fechas)))
    if year_filter:
        indices_meses = [i for i, f in enumerate(fechas) if hasattr(f, 'year') and f.year == year_filter]

    fechas_filtered = [fechas[i] for i in indices_meses]

    def get_label(row):
        # La etiqueta puede estar en cols B, D, E (cols 1, 3, 4)
        for col in (3, 1, 2, 4):
            if col < len(row) and row[col]:
                return str(row[col]).strip()
        return ""

    # Filas clave del P&L (basadas en inspeccion):
    # 4=Ingresos, 7=Costo Directo, 9=Margen Frontal, 10=% Margen Frontal,
    # 13=Comisión y Envio, 14=Flete/insumos, 15=Marketing,
    # 16=Total Otros Costos Excedentes, 18=Margen Contribución
    # 22=Sueldos, 23=Oficina, 24=Movilización, 25=Comisión Transb...
    filas_pl = [4, 7, 9, 10, 13, 14, 15, 16, 18, 22, 23, 24, 25]
    out_rows = []
    for idx in filas_pl:
        if idx >= len(rows):
            continue
        row = rows[idx]
        if not row:
            continue
        label_concepto = get_label(row)
        if not label_concepto:
            continue
        valores = []
        for i in indices_meses:
            col = col_start + i
            if col < len(row):
                v = row[col]
                valores.append(v if isinstance(v, (int, float)) else None)
            else:
                valores.append(None)
        out_rows.append({"concepto": label_concepto, "valores": valores})

    return {"fechas": fechas_filtered, "rows": out_rows}


@st.cache_data(ttl=300)
def calcular_indicadores_ytd(path_str: str, mes_actual: int) -> list[dict]:
    """Calcula indicadores financieros con datos REALES YTD (no del Fcst).

    Devuelve lista de dicts con: kpi, ytd_value, anualizado_value, benchmark, estado.
    Se basa en:
      - Metas 2026 (para Venta, Contrib, EBIT, Utilidad reales)
      - KT (para Activos/Pasivos Corrientes)
      - Deuda (para Deuda Financiera)
    """
    metas = cargar_metas_2026(path_str)
    kt = cargar_kt_resumen(path_str)
    deu = cargar_deuda_resumen(path_str)

    if not metas:
        return []

    meses_ytd = max(1, mes_actual - 1)

    def sum_ytd(metric, key="Resultado"):
        return sum(v for v in metas.get(metric, {}).get(key, [])[:meses_ytd] if isinstance(v,(int,float)))

    def ultimo_serie(d, k):
        s = d.get(k, {}).get("serie", []) if d else []
        return s[-1][1] if s else None

    venta_ytd = sum_ytd("Venta")
    contrib_ytd = sum_ytd("Contribucion")
    gav_ytd = sum_ytd("GAV")
    ebit_ytd = sum_ytd("EBIT")
    util_ytd = sum_ytd("Utilidad")

    # Anualizar (extrapolacion simple)
    fact = 12 / meses_ytd if meses_ytd else 1
    venta_anual = venta_ytd * fact
    contrib_anual = contrib_ytd * fact
    ebit_anual = ebit_ytd * fact
    util_anual = util_ytd * fact

    # Stocks (último valor de cada serie)
    act_cor = ultimo_serie(kt, "Total Act Corrientes")
    pas_cor = abs(ultimo_serie(kt, "Total Pas Corrientes")) if ultimo_serie(kt, "Total Pas Corrientes") else None
    existencias = ultimo_serie(kt, "Existencias")
    deuda_total = abs(ultimo_serie(deu, "Saldo final")) if ultimo_serie(deu, "Saldo final") else None

    indicadores = []

    def agregar(grupo, kpi, ytd, anual, bench, estado_evaluator):
        ytd_estado = estado_evaluator(ytd) if ytd is not None else "—"
        anual_estado = estado_evaluator(anual) if anual is not None else "—"
        indicadores.append({
            "Grupo": grupo,
            "KPI": kpi,
            "YTD (real)": _fmt_kpi_value(ytd),
            "🚦 YTD": ytd_estado,
            "Anualizado proy.": _fmt_kpi_value(anual),
            "🚦 Anual": anual_estado,
            "Benchmark": bench,
        })

    # Liquidez (no cambia con anualizacion - son stocks)
    if act_cor and pas_cor:
        rc = act_cor / pas_cor
        agregar("LIQUIDEZ", "Razón Corriente", rc, rc, "1.5 - 2.0x",
                lambda v: "🟢" if 1.5 <= v <= 3 else ("🟡" if 1.0 <= v < 1.5 or 3 < v <= 5 else "🔴"))
        if existencias:
            pa = (act_cor - existencias) / pas_cor
            agregar("LIQUIDEZ", "Prueba Ácida", pa, pa, ">1.0x",
                    lambda v: "🟢" if v >= 1.0 else "🔴")

    # Rentabilidad — comparar YTD vs Anualizado
    if venta_ytd:
        margen_op_ytd = ebit_ytd / venta_ytd
        margen_op_anual = ebit_anual / venta_anual if venta_anual else None
        agregar("RENTABILIDAD", "Margen Operacional", margen_op_ytd, margen_op_anual, ">5%",
                lambda v: "🟢" if v >= 0.05 else ("🟡" if v >= 0 else "🔴"))

        margen_contrib_ytd = contrib_ytd / venta_ytd
        agregar("RENTABILIDAD", "Margen Contribución", margen_contrib_ytd, margen_contrib_ytd, ">27%",
                lambda v: "🟢" if v >= 0.27 else ("🟡" if v >= 0.20 else "🔴"))

    # Apalancamiento (stocks - igual YTD que anualizado)
    if deuda_total and venta_anual:
        d_ebitda_ytd = deuda_total / (ebit_ytd*fact + 1)  # aproximacion: EBIT como proxy de EBITDA
        agregar("APALANCAMIENTO", "Deuda / EBIT anualizado", d_ebitda_ytd, d_ebitda_ytd, "<4.0x",
                lambda v: "🟢" if 0 < v <= 4 else ("🟡" if 4 < v <= 6 else "🔴"))

    return indicadores


def _fmt_kpi_value(v):
    if v is None:
        return "—"
    try:
        x = float(v)
        if abs(x) < 1:  # ratio porcentual
            return f"{x*100:+.1f}%"
        return f"{x:.2f}x"
    except Exception:
        return str(v)


def generar_propuestas_eerr(metas: dict, mes_actual: int) -> list[dict]:
    """Heuristicas simples sobre los datos de Metas 2026 para sugerir acciones.

    Devuelve lista de dicts con: severidad, area, mensaje, accion_sugerida.
    """
    propuestas = []
    if not metas:
        return propuestas

    meses_es = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    meses_ytd = max(1, mes_actual - 1)

    def safe_sum(arr):
        return sum(v for v in arr[:meses_ytd] if isinstance(v, (int, float)))

    def safe_pct_avg(arr):
        vals = [v for v in arr[:meses_ytd] if isinstance(v, (int, float))]
        return sum(vals) / len(vals) if vals else None

    # 1. Var% Venta YTD
    var_venta = safe_pct_avg(metas.get("Venta", {}).get("Var%", []))
    if var_venta is not None and var_venta < -0.10:
        propuestas.append({
            "severidad": "🔴 ALTA",
            "area": "Ventas",
            "mensaje": f"Venta YTD está {var_venta*100:+.1f}% bajo meta",
            "accion": "Revisar pipeline comercial · Analizar canales con peor performance · Considerar campañas de aceleración",
        })
    elif var_venta is not None and var_venta < -0.05:
        propuestas.append({
            "severidad": "🟡 MEDIA",
            "area": "Ventas",
            "mensaje": f"Venta YTD ligeramente bajo meta ({var_venta*100:+.1f}%)",
            "accion": "Monitorear tendencia · Validar forecast actualizado",
        })

    # 2. Margen contribución
    venta_real = safe_sum(metas.get("Venta", {}).get("Resultado", []))
    contrib_real = safe_sum(metas.get("Contribucion", {}).get("Resultado", []))
    if venta_real and contrib_real:
        margen_contrib = contrib_real / venta_real
        if margen_contrib < 0.27:
            propuestas.append({
                "severidad": "🔴 ALTA",
                "area": "Contribución",
                "mensaje": f"Margen de contribución {margen_contrib*100:.1f}% bajo umbral 27%",
                "accion": "Revisar pricing · Renegociar costos directos · Validar mix de productos",
            })

    # 3. GAV vs venta
    gav_real = safe_sum(metas.get("GAV", {}).get("Resultado", []))
    if venta_real and gav_real:
        gav_ratio = abs(gav_real) / venta_real
        if gav_ratio > 0.30:
            propuestas.append({
                "severidad": "🟡 MEDIA",
                "area": "Gastos",
                "mensaje": f"GAV representa {gav_ratio*100:.1f}% de venta (umbral 30%)",
                "accion": "Auditar gastos administrativos · Identificar partidas no esenciales",
            })

    # 4. EBIT
    ebit_real = safe_sum(metas.get("EBIT", {}).get("Resultado", []))
    if venta_real and ebit_real is not None:
        ebit_margin = ebit_real / venta_real
        if ebit_margin < 0:
            propuestas.append({
                "severidad": "🔴 ALTA",
                "area": "EBIT",
                "mensaje": f"EBIT margin negativo ({ebit_margin*100:+.1f}%) acumulado YTD",
                "accion": "Plan de recuperación urgente · Revisar estructura de costos · Acelerar venta o reducir gastos",
            })
        elif ebit_margin < 0.05:
            propuestas.append({
                "severidad": "🟡 MEDIA",
                "area": "EBIT",
                "mensaje": f"EBIT margin {ebit_margin*100:.1f}% por debajo benchmark 5%",
                "accion": "Mejoras de eficiencia operacional",
            })

    # 5. YoY Venta (vs 2025)
    var_yoy = safe_pct_avg(metas.get("Venta", {}).get("VariacionPct", []))
    if var_yoy is not None:
        if var_yoy > 0.10:
            propuestas.append({
                "severidad": "🟢 POSITIVA",
                "area": "YoY",
                "mensaje": f"Crecimiento Venta YoY +{var_yoy*100:.1f}% (vs 2025)",
                "accion": "Mantener inversión en canales que aceleran · Capturar momentum",
            })
        elif var_yoy < -0.05:
            propuestas.append({
                "severidad": "🔴 ALTA",
                "area": "YoY",
                "mensaje": f"Decrecimiento Venta YoY {var_yoy*100:+.1f}% (vs 2025)",
                "accion": "Análisis comparativo profundo · Identificar canales con caída · Plan de recuperación",
            })

    return propuestas


def tabla_simple(df: pd.DataFrame, titulo: str = ""):
    if titulo:
        st.markdown(f"**{titulo}**")
    if df is None or df.empty:
        st.info("Sin datos.")
        return
    # Limpiar filas/columnas totalmente vacias
    df_clean = df.dropna(how="all").dropna(how="all", axis=1)
    st.dataframe(df_clean, use_container_width=True, hide_index=True)


def detectar_mes_anio(filename: str) -> tuple[int, int] | None:
    """Detecta mes y año del nombre del archivo. Ej: '02 EE.RR Febrero 2026.xlsx' -> (2, 2026)."""
    meses_map = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
        'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    fn = filename.lower()
    mes = None
    for nombre, num in meses_map.items():
        if nombre in fn:
            mes = num
            break
    # Tambien probar prefijo numerico "01 ", "02 ", ... "12 "
    if mes is None:
        m = re.match(r'^\s*(0[1-9]|1[0-2])\s', fn)
        if m:
            mes = int(m.group(1))
    # Año
    m_anio = re.search(r'\b(20\d{2})\b', fn)
    anio = int(m_anio.group(1)) if m_anio else None
    if mes and anio:
        return (mes, anio)
    return None


# Determinar rol del usuario (heuristica: esta en sesion, o por username)
sess_user = st.session_state.get("username", "")
sess_roles = []
try:
    import yaml as _yaml
    auth_cfg = _yaml.safe_load((PARENT / "auth_config.yaml").read_text(encoding="utf-8"))
    sess_roles = auth_cfg.get("credentials", {}).get("usernames", {}).get(sess_user, {}).get("roles", []) or []
except Exception:
    pass

is_admin = "admin" in sess_roles or sess_user == "andres"
can_upload = is_admin or "uploader" in sess_roles

# ============================================================================
# TABS
# ============================================================================
with st.expander("ℹ️ ¿De qué archivo sale toda la información de esta página?", expanded=False):
    from datetime import datetime as _dt
    _mtime = _dt.fromtimestamp(sp.PLANIFICACION_FINANCIERA.stat().st_mtime).strftime('%d/%m/%Y %H:%M')
    st.markdown(f"""
**Archivo activo (detectado dinámicamente):**
`{sp.PLANIFICACION_FINANCIERA.name}`

**Ruta completa:**
`{sp.PLANIFICACION_FINANCIERA}`

**Última modificación del archivo:** {_mtime}

**Detección automática:** el sistema busca con glob `Planificación Financiera V*.xlsx` en
`G:\\Mi unidad\\TRABAJO\\RESPALDO\\OPERACIONES\\Finanzas\\Empresa\\2026\\Planificación Financiera\\`
y toma el de fecha de modificación más reciente. Cuando guardes V52, V53, V54... se detecta solo.

**Hojas usadas en cada tab:**
- Tab 1 (Resumen EERR + Propuestas): hoja `Metas 2026`
- Tab 2 (vs Presupuesto): hoja `Metas 2026`
- Tab 3 (P&L Proyectado): hojas `P&L`, `Fcst EERR`
- Tab 4 (YoY): hoja `Metas 2026`
- Tab 5 (KT): hoja `KT`
- Tab 6 (Deuda Financiera): hoja `Deuda financiera`
- Tab 7 (EEFF Flujo & Indicadores): hoja `Análisis Financiero 2026` + Odoo (DIO/DSO/DPO/CCC)
    """)

tabs = st.tabs([
    "1. 📊 Resumen EERR + Propuestas",
    "2. 🎯 Resultado vs Presupuesto",
    "3. 🔮 P&L Proyectado",
    "4. 📅 YoY (2025 vs 2026)",
    "5. 💼 KT por cuentas",
    "6. 🏦 Deuda Financiera",
    "7. 💵 EEFF: Flujo & Indicadores",
    "📥 Carga mensual" + ("" if can_upload else " 🔒"),
])

# ----- Tab 1: Resumen EERR + Propuestas ---------------------------------------
with tabs[0]:
    metas = cargar_metas_2026(str(sp.PLANIFICACION_FINANCIERA))
    dd = cargar_dashboard_data(str(sp.PLANIFICACION_FINANCIERA))
    mes_actual = datetime.now().month
    meses_es = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    meses_ytd = max(1, mes_actual - 1)

    st.markdown(f"### 📊 Resumen EERR ({meses_es[0]} → {meses_es[meses_ytd-1]})")
    st.caption(f"Anualización proyectada en base a {meses_ytd} meses cerrados.")

    if metas:
        # KPIs YTD + anualizado proyectado
        cols = st.columns(5)
        for i, (m, alias) in enumerate([("Venta","💰 Venta"),("Contribucion","📈 Contrib."),("GAV","🏢 GAV"),("EBIT","🎯 EBIT"),("Utilidad","💵 Utilidad")]):
            try:
                meta = sum(v for v in metas[m]["Meta"][:meses_ytd] if isinstance(v,(int,float)))
                res  = sum(v for v in metas[m]["Resultado"][:meses_ytd] if isinstance(v,(int,float)))
                # Anualizado proyectado = ritmo YTD * 12 / meses cerrados
                anualizado = (res * 12 / meses_ytd) if meses_ytd else None
                var = (res-meta)/meta if meta else None
                cols[i].metric(
                    f"{_color_var(var)} {alias}",
                    f"{res/1e6:,.1f}M",
                    delta=f"Anualiz: {anualizado/1e6:,.0f}M" if anualizado else None,
                )
            except Exception:
                cols[i].metric(alias, "—")

        st.divider()

        # Propuestas de mejora (heuristicas)
        st.markdown("### 💡 Propuestas de mejora (heurísticas sobre resultado anualizado)")
        propuestas = generar_propuestas_eerr(metas, mes_actual)
        if not propuestas:
            st.success("✅ Sin alertas críticas. Performance dentro de umbrales esperados.")
        else:
            for p in propuestas:
                with st.container(border=True):
                    sev_col, msg_col = st.columns([1, 4])
                    with sev_col:
                        st.markdown(f"**{p['severidad']}**")
                        st.caption(p["area"])
                    with msg_col:
                        st.markdown(f"**{p['mensaje']}**")
                        st.markdown(f"*Acción sugerida:* {p['accion']}")

        st.divider()

        # Tabla de las 5 metricas con resultado anualizado proyectado
        st.markdown("### 📋 Anualización proyectada por métrica")
        rows_anual = []
        for m in metas:
            try:
                meta_total = sum(v for v in metas[m]["Meta"] if isinstance(v,(int,float)))
                res_ytd = sum(v for v in metas[m]["Resultado"][:meses_ytd] if isinstance(v,(int,float)))
                anualizado = res_ytd * 12 / meses_ytd if meses_ytd else 0
                gap = anualizado - meta_total
                gap_pct = gap / meta_total if meta_total else None
                rows_anual.append({
                    "Métrica": m,
                    "Meta Anual": f"{meta_total/1e6:,.1f}M",
                    "YTD": f"{res_ytd/1e6:,.1f}M",
                    "Anualizado proy.": f"{anualizado/1e6:,.1f}M",
                    "Gap vs Meta": f"{gap/1e6:+,.1f}M",
                    "Gap %": _fmt_pct(gap_pct),
                    "🚦": _color_var(gap_pct),
                })
            except Exception:
                pass
        st.dataframe(pd.DataFrame(rows_anual), use_container_width=True, hide_index=True)
    else:
        st.error("No se pudo leer hoja `Metas 2026`. Validá que la planilla `Planificación Financiera` esté actualizada.")

# ----- Tab 2: Resultado vs Presupuesto (Metas 2026) ---------------------------
with tabs[1]:
    metas = cargar_metas_2026(str(sp.PLANIFICACION_FINANCIERA))
    meses = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    if not metas:
        st.error("No se pudo leer hoja `Metas 2026`.")
    else:
        st.markdown("### 🎯 Resultado vs Presupuesto (Metas 2026)")
        st.caption("Fuente: hoja `Metas 2026`. Cada métrica con Meta, Resultado y Var%.")

        # KPIs YTD por metrica
        mes_actual = datetime.now().month
        meses_ytd = max(1, mes_actual - 1)
        cols = st.columns(5)
        for i, (m, alias) in enumerate([("Venta","💰 Venta"),("Contribucion","📈 Contrib."),("GAV","🏢 GAV"),("EBIT","🎯 EBIT"),("Utilidad","💵 Utilidad")]):
            try:
                meta = sum(v for v in metas[m]["Meta"][:meses_ytd] if isinstance(v,(int,float)))
                res  = sum(v for v in metas[m]["Resultado"][:meses_ytd] if isinstance(v,(int,float)))
                var = (res-meta)/meta if meta else None
                cols[i].metric(f"{_color_var(var)} {alias}", f"{res/1e6:,.1f}M", _fmt_pct(var))
            except Exception:
                cols[i].metric(alias, "—")

        st.divider()

        # Selector de metrica + grafico Meta vs Resultado vs Var
        col_sel, _ = st.columns([1, 3])
        with col_sel:
            metric_pres = st.selectbox("Métrica", list(metas.keys()), key="ppto_metric")

        meta_vals = [v if isinstance(v,(int,float)) else None for v in metas[metric_pres]["Meta"]]
        res_vals  = [v if isinstance(v,(int,float)) else None for v in metas[metric_pres]["Resultado"]]
        var_vals  = [v if isinstance(v,(int,float)) else None for v in metas[metric_pres]["Var%"]]

        fig = go.Figure()
        fig.add_trace(go.Bar(x=meses, y=meta_vals, name="Meta", marker_color="#94A3B8"))
        fig.add_trace(go.Bar(x=meses, y=res_vals, name="Resultado", marker_color="#1F4E79"))
        fig.update_layout(barmode='group', height=380, hovermode='x unified',
                          margin=dict(l=20,r=20,t=20,b=20),
                          legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        st.plotly_chart(fig, use_container_width=True)

        # Tabla detallada
        rows_t = []
        for i, mes in enumerate(meses):
            rows_t.append({
                "Mes": mes,
                "Meta": _fmt_num((meta_vals[i] or 0)/1e6) + "M" if meta_vals[i] else "—",
                "Resultado": _fmt_num((res_vals[i] or 0)/1e6) + "M" if res_vals[i] else "—",
                "Var%": _fmt_pct(var_vals[i]),
                "🚦": _color_var(var_vals[i]),
            })
        st.dataframe(pd.DataFrame(rows_t), use_container_width=True, hide_index=True)

# ----- Tab 3: P&L Proyectado --------------------------------------------------
with tabs[2]:
    st.markdown("### 🔮 P&L Proyectado")
    st.caption("Fuente: hoja `P&L`. Datos mensuales filtrados por año seleccionado.")

    col_year, _ = st.columns([1, 4])
    with col_year:
        year_pl = st.selectbox("Año", [2026, 2025, 2024, 2023], index=0, key="pl_year")

    pl_data = cargar_pl_proyectado(str(sp.PLANIFICACION_FINANCIERA), year_pl)
    if not pl_data or not pl_data.get("rows"):
        st.warning(f"Sin datos de P&L para {year_pl}. Mostrando hoja completa.")
        try:
            df = cargar_hoja(str(sp.PLANIFICACION_FINANCIERA), "Fcst EERR")
            df = df.dropna(how="all").dropna(how="all", axis=1)
            st.dataframe(df, use_container_width=True, hide_index=True, height=520)
        except Exception as e:
            st.error(f"No se pudo leer: {e}")
    else:
        fechas = pl_data["fechas"]
        rows_pl = pl_data["rows"]

        meses_es_full = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
        col_headers = [meses_es_full[f.month - 1] if hasattr(f, 'month') else str(f) for f in fechas]

        # Construir DataFrame con conceptos como filas, meses como columnas
        data_dict = {"Concepto": [r["concepto"] for r in rows_pl]}
        for j, mes in enumerate(col_headers):
            valores_col = []
            for r in rows_pl:
                v = r["valores"][j] if j < len(r["valores"]) else None
                if v is None:
                    valores_col.append("—")
                elif isinstance(v, (int, float)):
                    if abs(v) < 1:  # asumir ratio porcentual (margen frontal etc)
                        valores_col.append(f"{v*100:.1f}%")
                    else:
                        # P&L viene en miles de CLP — dividir por 1000 para mostrar en MM$
                        valores_col.append(f"{v/1e3:,.0f} MM$" if abs(v) >= 1000 else f"{v:,.0f} M$")
                else:
                    valores_col.append(str(v))
            data_dict[mes] = valores_col

        # Total YTD por fila (sumar columnas con valores numéricos)
        totales_ytd = []
        for r in rows_pl:
            total = sum(v for v in r["valores"] if isinstance(v, (int, float)) and abs(v) >= 1)
            if total == 0:
                totales_ytd.append("—")
            else:
                totales_ytd.append(f"{total/1e3:,.0f} MM$" if abs(total) >= 1000 else f"{total:,.0f} M$")
        data_dict[f"Total {year_pl}"] = totales_ytd

        df_pl = pd.DataFrame(data_dict)

        # Highlight rows clave (Margen Frontal, Margen Contribución, etc.)
        def highlight_key_rows(row):
            key_concepts = ["Ingresos", "Margen Frontal", "Margen Contribución", "Total Otros Costos"]
            if any(kc.lower() in str(row["Concepto"]).lower() for kc in key_concepts):
                return ['background-color: #E8EEF5; font-weight: bold'] * len(row)
            return [''] * len(row)

        st.dataframe(
            df_pl.style.apply(highlight_key_rows, axis=1),
            use_container_width=True,
            hide_index=True,
            height=520,
        )

        st.caption(
            f"Conceptos clave destacados. **MM$** = millones de pesos · **M$** = miles de pesos · márgenes en %. "
            f"Datos del año {year_pl} de la hoja `P&L`."
        )

# ----- Tab 4: YoY (2025 vs 2026) ----------------------------------------------
with tabs[3]:
    metas = cargar_metas_2026(str(sp.PLANIFICACION_FINANCIERA))
    meses = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    if not metas:
        st.error("No se pudo leer hoja `Metas 2026`.")
    else:
        st.markdown("### 📅 Comparación Año contra Año (2025 vs 2026)")
        st.caption("Resultado mes a mes, mismo mes año anterior, variación absoluta y %.")

        # KPIs YoY YTD
        mes_actual = datetime.now().month
        meses_ytd = max(1, mes_actual - 1)
        cols = st.columns(5)
        for i, (m, alias) in enumerate([("Venta","💰 Venta"),("Contribucion","📈 Contrib."),("GAV","🏢 GAV"),("EBIT","🎯 EBIT"),("Utilidad","💵 Utilidad")]):
            try:
                r2026 = sum(v for v in metas[m]["Resultado"][:meses_ytd] if isinstance(v,(int,float)))
                r2025 = sum(v for v in metas[m]["Resultado2025"][:meses_ytd] if isinstance(v,(int,float)))
                var = (r2026-r2025)/r2025 if r2025 else None
                cols[i].metric(f"{_color_var(var)} {alias} YoY", f"{r2026/1e6:,.1f}M", _fmt_pct(var))
            except Exception:
                cols[i].metric(alias, "—")

        st.divider()

        # Selector + grafico
        col_sel, _ = st.columns([1,3])
        with col_sel:
            m_yoy = st.selectbox("Métrica YoY", list(metas.keys()), key="yoy_metric")

        v2025 = [v if isinstance(v,(int,float)) else None for v in metas[m_yoy]["Resultado2025"]]
        v2026 = [v if isinstance(v,(int,float)) else None for v in metas[m_yoy]["Resultado"]]
        var_yoy = [v if isinstance(v,(int,float)) else None for v in metas[m_yoy]["VariacionPct"]]

        fig = go.Figure()
        fig.add_trace(go.Scatter(x=meses, y=v2025, mode='lines+markers', name='2025', line=dict(color='#94A3B8', width=2)))
        fig.add_trace(go.Scatter(x=meses, y=v2026, mode='lines+markers', name='2026', line=dict(color='#1F4E79', width=3)))
        fig.update_layout(height=380, hovermode='x unified', margin=dict(l=20,r=20,t=20,b=20),
                          legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        st.plotly_chart(fig, use_container_width=True)

        # Tabla
        rows_t = []
        for i, mes in enumerate(meses):
            rows_t.append({
                "Mes": mes,
                "2025 (M$)": _fmt_num((v2025[i] or 0)/1e6) if v2025[i] else "—",
                "2026 (M$)": _fmt_num((v2026[i] or 0)/1e6) if v2026[i] else "—",
                "Variación %": _fmt_pct(var_yoy[i]),
                "🚦": _color_var(var_yoy[i]),
            })
        st.dataframe(pd.DataFrame(rows_t), use_container_width=True, hide_index=True)

# ----- Tab 5: KT por cuentas --------------------------------------------------
with tabs[4]:
    st.markdown("### 💼 Capital de Trabajo (resumen por cuentas)")
    st.caption("Fuente: hoja `KT`. Datos en M$ (millones de pesos). Se actualiza al subir EERR/Balance.")
    kt = cargar_kt_resumen(str(sp.PLANIFICACION_FINANCIERA))
    if not kt or not kt.get("fechas"):
        st.error("No se pudo leer hoja `KT` o no tiene datos con fecha.")
    else:
        cuentas_balance = ["Existencias", "CxC Comerciales", "Otros Act Cor", "Total Act Corrientes",
                           "CxP Comerciales", "Otros Pas Cor", "Total Pas Corrientes",
                           "KT Neto", "Cambios en KT"]
        cuentas_indicadores = ["Meses de Inventario", "Meses de CxC"]

        # Helper: obtener (fecha_ultimo, valor_ultimo, fecha_anterior, valor_anterior)
        def latest_two(cuenta):
            s = kt.get(cuenta, {}).get("serie", [])
            if not s:
                return None, None, None, None
            ult_f, ult_v = s[-1]
            pen_f, pen_v = s[-2] if len(s) > 1 else (None, None)
            return ult_f, ult_v, pen_f, pen_v

        # === Cards con últimos valores
        cols = st.columns(4)
        for i, k in enumerate(["Total Act Corrientes", "Total Pas Corrientes", "KT Neto", "Cambios en KT"]):
            f_u, v_u, f_p, v_p = latest_two(k)
            delta = (v_u - v_p) if (v_u is not None and v_p is not None) else None
            cols[i].metric(
                f"{k}",
                f"{v_u/1e3:,.0f} MM$" if v_u is not None else "—",
                delta=f"{delta/1e3:+,.0f} MM$" if delta is not None else None,
                help=f"Último: {f_u.strftime('%b %Y') if f_u else 'N/A'}",
            )

        st.divider()

        # === Tabla detallada con últimos 4 períodos
        st.markdown("#### 📋 Evolución últimos 4 períodos por cuenta (M$)")
        # Tomar últimos 4 fechas que aparezcan en alguna cuenta
        fechas_recientes = []
        for k in cuentas_balance:
            s = kt.get(k, {}).get("serie", [])
            if s:
                fechas_recientes = [f for f, _ in s[-4:]]
                break

        if fechas_recientes:
            mes_es = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
            col_headers = [f"{mes_es[f.month-1]} {f.year}" if hasattr(f,'month') else str(f) for f in fechas_recientes]

            data = {"Cuenta": []}
            for h in col_headers:
                data[h] = []
            data["Δ último"] = []

            for k in cuentas_balance:
                s = kt.get(k, {}).get("serie", [])
                if not s:
                    continue
                # Buscar valor para cada fecha en col_headers
                d_valores = dict(s)
                data["Cuenta"].append(k)
                for j, f in enumerate(fechas_recientes):
                    v = d_valores.get(f)
                    data[col_headers[j]].append(f"{v/1e3:,.0f} MM$" if isinstance(v,(int,float)) else "—")
                # Delta del último vs anterior
                if len(s) > 1 and isinstance(s[-1][1],(int,float)) and isinstance(s[-2][1],(int,float)):
                    delta = (s[-1][1] - s[-2][1]) / 1e6
                    data["Δ último"].append(f"{delta:+,.0f}M")
                else:
                    data["Δ último"].append("—")

            df_kt = pd.DataFrame(data)

            def highlight_totales(row):
                if "Total" in str(row["Cuenta"]) or "KT Neto" in str(row["Cuenta"]):
                    return ['background-color: #E8EEF5; font-weight: bold'] * len(row)
                return [''] * len(row)

            st.dataframe(df_kt.style.apply(highlight_totales, axis=1), use_container_width=True, hide_index=True)

        # === Indicadores de eficiencia
        st.markdown("#### 🎯 Indicadores de eficiencia operativa")
        e1, e2 = st.columns(2)
        for col, k, icon in [(e1, "Meses de Inventario", "📦"), (e2, "Meses de CxC", "💳")]:
            f_u, v_u, _, v_p = latest_two(k)
            delta = (v_u - v_p) if (v_u is not None and v_p is not None) else None
            col.metric(
                f"{icon} {k}",
                f"{v_u:.1f} meses" if v_u is not None else "—",
                delta=f"{delta:+.1f}" if delta is not None else None,
                help=f"Último: {f_u.strftime('%b %Y') if f_u else 'N/A'}",
            )

        # === Gráfico evolución KT Neto
        st.markdown("#### 📈 Evolución KT Neto")
        kt_serie = kt.get("KT Neto", {}).get("serie", [])
        if kt_serie:
            fechas_g = [f for f, _ in kt_serie]
            valores_g = [v for _, v in kt_serie]
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=fechas_g, y=valores_g, mode='lines+markers',
                                     line=dict(color='#1F4E79', width=3),
                                     fill='tozeroy', fillcolor='rgba(31,78,121,0.08)',
                                     name='KT Neto'))
            fig.update_layout(height=320, margin=dict(l=20,r=20,t=20,b=20),
                              xaxis_title="", yaxis_title="KT Neto",
                              showlegend=False, hovermode='x unified')
            st.plotly_chart(fig, use_container_width=True)

# ----- Tab 6: Deuda Financiera ------------------------------------------------
with tabs[5]:
    st.markdown("### 🏦 Deuda Financiera")
    st.caption("Fuente: hoja `Deuda financiera`. Datos en M$ (millones).")
    deu = cargar_deuda_resumen(str(sp.PLANIFICACION_FINANCIERA))
    if not deu or not deu.get("fechas"):
        st.error("No se pudo leer hoja `Deuda financiera`.")
    else:
        def latest_two_d(cuenta):
            s = deu.get(cuenta, {}).get("serie", [])
            if not s:
                return None, None, None, None
            ult_f, ult_v = s[-1]
            pen_f, pen_v = s[-2] if len(s) > 1 else (None, None)
            return ult_f, ult_v, pen_f, pen_v

        # === Cards
        cols = st.columns(4)
        for i, (k, alias, fmt) in enumerate([
            ("Saldo final", "Saldo final", "M"),
            ("Balance prom", "Balance promedio", "M"),
            ("Tasa interes", "Tasa mensual", "%"),
            ("Gasto intereses", "Gasto intereses", "M"),
        ]):
            f_u, v_u, _, v_p = latest_two_d(k)
            if v_u is None:
                cols[i].metric(alias, "—")
                continue
            if fmt == "%":
                val_str = f"{v_u*100:.2f}%"
                delta = f"{(v_u-v_p)*100:+.2f}pp" if v_p is not None else None
            else:
                val_str = f"{v_u/1e3:,.0f} MM$"
                delta = f"{(v_u-v_p)/1e3:+,.0f} MM$" if v_p is not None else None
            cols[i].metric(alias, val_str, delta=delta,
                           help=f"Último: {f_u.strftime('%b %Y') if f_u else 'N/A'}")

        st.divider()

        # === Tabla últimos 4 períodos
        st.markdown("#### 📋 Evolución últimos 4 períodos")
        # Buscar fechas recientes
        fechas_recientes = []
        for k in ["Saldo final"]:
            s = deu.get(k, {}).get("serie", [])
            if s:
                fechas_recientes = [f for f, _ in s[-4:]]
                break

        if fechas_recientes:
            mes_es = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
            col_headers = [f"{mes_es[f.month-1]} {f.year}" if hasattr(f,'month') else str(f) for f in fechas_recientes]

            cuentas_d = ["Saldo inicial", "Emision", "Amortizacion", "Saldo final",
                         "Balance prom", "Tasa interes", "Gasto intereses"]
            data = {"Concepto": []}
            for h in col_headers:
                data[h] = []

            for k in cuentas_d:
                s = deu.get(k, {}).get("serie", [])
                if not s:
                    continue
                d_valores = dict(s)
                data["Concepto"].append(k)
                for j, f in enumerate(fechas_recientes):
                    v = d_valores.get(f)
                    if v is None:
                        data[col_headers[j]].append("—")
                    elif k == "Tasa interes":
                        data[col_headers[j]].append(f"{v*100:.2f}%")
                    else:
                        data[col_headers[j]].append(f"{v/1e3:,.0f} MM$")

            df_d = pd.DataFrame(data)

            def highlight_d(row):
                if "Saldo final" in str(row["Concepto"]):
                    return ['background-color: #FEE2E2; font-weight: bold'] * len(row)
                return [''] * len(row)

            st.dataframe(df_d.style.apply(highlight_d, axis=1), use_container_width=True, hide_index=True)

        # === Gráfico evolución
        saldo_serie = deu.get("Saldo final", {}).get("serie", [])
        if saldo_serie:
            fechas_g = [f for f, _ in saldo_serie]
            valores_g = [v for _, v in saldo_serie]
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=fechas_g, y=valores_g, mode='lines+markers',
                                     line=dict(color='#d32f2f', width=3),
                                     fill='tozeroy', fillcolor='rgba(211,47,47,0.08)',
                                     name='Saldo Deuda'))
            fig.update_layout(height=320, margin=dict(l=20,r=20,t=20,b=20),
                              xaxis_title="", yaxis_title="Saldo Deuda",
                              showlegend=False, hovermode='x unified')
            st.plotly_chart(fig, use_container_width=True)

# ----- Tab 7: EEFF Flujo de Caja + Indicadores -------------------------------
with tabs[6]:
    st.markdown("### 💵 EEFF: Flujo de Caja & Indicadores Clave")
    st.caption("Fuente: hoja `Análisis Financiero 2026`.")
    af = cargar_analisis_financiero(str(sp.PLANIFICACION_FINANCIERA))
    if not af:
        st.error("No se pudo leer 'Análisis Financiero 2026'.")
    else:
        # Sec 1: Flujo de Caja
        with st.expander("1️⃣ Flujo de Caja 2026", expanded=True):
            df = pd.DataFrame(af.get("flujo_caja", []))
            if not df.empty:
                df["v_2025"] = df["v_2025"].apply(lambda v: f"{v:,.0f}" if isinstance(v,(int,float)) else "—")
                df["v_2026"] = df["v_2026"].apply(lambda v: f"{v:,.0f}" if isinstance(v,(int,float)) else "—")
                df["var_pct"] = df["var_pct"].apply(_fmt_pct)
                df.columns = ["Concepto", "2025 (M$)", "2026 (M$)", "Var Abs", "Var %", "Nota"]
                st.dataframe(df, use_container_width=True, hide_index=True)

        # Sec 2: Estructura de Deuda
        col_a, col_b = st.columns([2, 1])
        with col_a:
            with st.expander("2️⃣ Estructura de Deuda", expanded=True):
                df = pd.DataFrame(af.get("deuda", []))
                if not df.empty:
                    df["v_2025"] = df["v_2025"].apply(lambda v: f"{v:,.0f}" if isinstance(v,(int,float)) else "—")
                    df["v_2026"] = df["v_2026"].apply(lambda v: f"{v:,.0f}" if isinstance(v,(int,float)) else "—")
                    df["var_pct"] = df["var_pct"].apply(_fmt_pct)
                    df.columns = ["Concepto", "Dic 2025", "Dic 2026", "Var Abs", "Var %", "Nota"]
                    st.dataframe(df, use_container_width=True, hide_index=True)
        with col_b:
            with st.expander("Ratios Deuda", expanded=True):
                for r in af.get("ratios_deuda", []):
                    val = r.get("v_2026")
                    delta = r.get("var_abs")
                    if isinstance(val,(int,float)):
                        st.metric(r.get("concepto",""), f"{val:.2f}",
                                  delta=f"{delta:+.2f}" if isinstance(delta,(int,float)) else None)

        # Sec 3: KPIs Financieros — DOBLE VISTA: YTD real vs Anualizado proyectado
        with st.expander("3️⃣ KPIs Financieros — YTD (real) vs Anualizado proyectado", expanded=True):
            st.caption(
                "**YTD**: indicadores calculados con datos reales hasta el último mes cerrado — pone el foco en si HAY riesgo HOY. "
                "**Anualizado**: extrapolación al cierre del año en curso — útil para validar si la tendencia llega a meta."
            )

            mes_actual = datetime.now().month
            indicadores_ytd = calcular_indicadores_ytd(str(sp.PLANIFICACION_FINANCIERA), mes_actual)

            if indicadores_ytd:
                df_ind = pd.DataFrame(indicadores_ytd)
                # Reordenar columnas
                df_ind = df_ind[["Grupo", "KPI", "YTD (real)", "🚦 YTD", "Anualizado proy.", "🚦 Anual", "Benchmark"]]

                def highlight_grupo(row):
                    if row["Grupo"] == "LIQUIDEZ":
                        return ['background-color: #DBEAFE'] * len(row)
                    if row["Grupo"] == "RENTABILIDAD":
                        return ['background-color: #DCFCE7'] * len(row)
                    if row["Grupo"] == "APALANCAMIENTO":
                        return ['background-color: #FEE2E2'] * len(row)
                    return [''] * len(row)

                st.dataframe(df_ind.style.apply(highlight_grupo, axis=1), use_container_width=True, hide_index=True)

            st.divider()
            st.markdown("**📚 KPIs originales del Análisis Financiero (referencia anualizada Fcst 2026):**")

            # Tabla original como referencia
            df = pd.DataFrame(af.get("kpis", []))
            if not df.empty:
                def fmt_val(v):
                    if not isinstance(v,(int,float)):
                        return "—"
                    if abs(v) < 1:
                        return f"{v*100:.1f}%"
                    return f"{v:.2f}x"
                df["v_2025"] = df["v_2025"].apply(fmt_val)
                df["v_2026"] = df["v_2026"].apply(fmt_val)
                df["var_pct"] = df["var_pct"].apply(_fmt_pct)
                df.columns = ["KPI", "2025", "2026 Fcst", "Var Abs", "Var %", "Benchmark"]
                st.dataframe(df, use_container_width=True, hide_index=True)

        # ====================================================================
        # 4️⃣ Capital de Trabajo Operativo (DIO/DSO/DPO/CCC desde Odoo)
        # ====================================================================
        with st.expander("4️⃣ Capital de Trabajo Operativo — DIO/DSO/DPO/CCC desde Odoo", expanded=True):
            st.caption("Calculados con datos reales de Odoo (últimos 365 días).")

            @st.cache_data(ttl=900, show_spinner="Consultando Odoo…")
            def _cargar_kt_odoo():
                # Importar kpis_odoo del directorio padre
                import sys as _sys
                _eerr = str(Path(__file__).resolve().parent.parent)
                if _eerr not in _sys.path:
                    _sys.path.insert(0, _eerr)
                from kpis_odoo import kpi_ccc, kpi_morosidad_b2b, get_odoo_client
                odoo = get_odoo_client()
                if odoo is None:
                    return {"error": "Odoo no disponible. Verificar env var ANDRES_ODOO_PASSWORD."}
                ccc = kpi_ccc(odoo)
                mor = kpi_morosidad_b2b(odoo, dias_corte=30)
                return {"ccc": ccc, "morosidad": mor, "error": None}

            kt_odoo = _cargar_kt_odoo()
            if kt_odoo.get("error"):
                st.warning(f"⚠️ {kt_odoo['error']}")
            else:
                ccc_data = kt_odoo.get("ccc", {})
                ccc_val = ccc_data.get("valor")
                comp = ccc_data.get("componentes", {})

                cols = st.columns(4)
                # DIO
                dio_val = comp.get("DIO") if isinstance(comp.get("DIO"), (int, float)) else None
                cols[0].metric(
                    "📦 DIO (Inventario)",
                    f"{dio_val:.0f} días" if dio_val is not None else "—",
                    help="Días de inventario en stock. Meta: 60-90 días"
                )
                # DSO
                dso_val = comp.get("DSO") if isinstance(comp.get("DSO"), (int, float)) else None
                cols[1].metric(
                    "💳 DSO (Cobranza)",
                    f"{dso_val:.0f} días" if dso_val is not None else "—",
                    help="Días promedio cobranza. Meta B2B: ≤45 días"
                )
                # DPO
                dpo_val = comp.get("DPO") if isinstance(comp.get("DPO"), (int, float)) else None
                cols[2].metric(
                    "💸 DPO (Pago a proveedores)",
                    f"{dpo_val:.0f} días" if dpo_val is not None else "—",
                    help="Días promedio de pago. Meta: ≥60 días"
                )
                # CCC
                ccc_color = "🟢" if ccc_val and ccc_val <= 90 else ("🟡" if ccc_val and ccc_val <= 120 else "🔴")
                cols[3].metric(
                    f"{ccc_color} CCC",
                    f"{ccc_val:.0f} días" if ccc_val is not None else "—",
                    help="Cash Conversion Cycle = DIO + DSO − DPO. Meta: ≤90 días"
                )

                if ccc_val is not None and ccc_val > 90:
                    st.error(f"🔴 CCC actual ({ccc_val:.0f} días) supera la meta de 90 días. "
                             f"Foco: {('reducir DIO' if dio_val and dio_val > 90 else 'reducir DSO' if dso_val and dso_val > 45 else 'aumentar DPO')}.")

                # Morosidad
                mor_val = kt_odoo.get("morosidad", {}).get("valor")
                if mor_val is not None:
                    mor_color = "🟢" if mor_val <= 0.05 else ("🟡" if mor_val <= 0.10 else "🔴")
                    st.metric(f"{mor_color} % Morosidad B2B (>30d)",
                              f"{mor_val*100:.1f}%",
                              help="Cartera B2B vencida más de 30 días. Meta: ≤5%")

        # ====================================================================
        # 5️⃣ Forecast accuracy
        # ====================================================================
        with st.expander("5️⃣ Forecast accuracy ingresos — Real vs Presupuesto", expanded=False):
            st.caption("Compara venta real (Metas 2026 → Resultado) vs presupuesto (Meta).")
            metas = cargar_metas_2026(str(sp.PLANIFICACION_FINANCIERA))
            if metas:
                meses = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
                meses_ytd_local = max(1, datetime.now().month - 1)
                metas_v = metas["Venta"]["Meta"][:meses_ytd_local]
                res_v = metas["Venta"]["Resultado"][:meses_ytd_local]
                rows_fa = []
                for i in range(meses_ytd_local):
                    m = metas_v[i] if i < len(metas_v) else None
                    r = res_v[i] if i < len(res_v) else None
                    if isinstance(m, (int, float)) and isinstance(r, (int, float)) and m != 0:
                        acc = abs(r - m) / m
                        rows_fa.append({
                            "Mes": meses[i],
                            "Presupuesto": f"{m/1e6:,.0f}M",
                            "Real": f"{r/1e6:,.0f}M",
                            "Desviación": f"{(r-m)/1e6:+,.0f}M",
                            "Accuracy": f"{(1-acc)*100:.1f}%",
                            "🚦": "🟢" if acc <= 0.10 else ("🟡" if acc <= 0.20 else "🔴"),
                        })
                if rows_fa:
                    st.dataframe(pd.DataFrame(rows_fa), use_container_width=True, hide_index=True)
                    # Accuracy promedio
                    accs = [abs((r-m)/m) for m, r in zip(metas_v, res_v) if isinstance(m,(int,float)) and isinstance(r,(int,float)) and m]
                    if accs:
                        avg_acc = sum(accs) / len(accs)
                        st.metric("Accuracy promedio YTD", f"{(1-avg_acc)*100:.1f}%",
                                  help="Plan: ≥90%. Distancia entre real y presupuesto.")

        # ====================================================================
        # 6️⃣ Cierre contable y operativo (manual)
        # ====================================================================
        with st.expander("6️⃣ Cierre contable y operativo (input manual)", expanded=False):
            st.caption("KPI: días para cerrar EERR mensual. Meta plan: ≤5 días hábiles.")
            st.info(
                "📥 **Próximamente**: Uploader para registrar mes a mes los días que tomó el cierre. "
                "Por ahora se carga manualmente en planilla. Ver `Roadmap H2` para integración con calendario."
            )

# ----- Tab 8: Carga mensual (admin/uploader) ----------------------------------
with tabs[7]:
    if not can_upload:
        st.info("🔒 Esta seccion es solo para administradores y uploaders. Iniciá sesión con permisos adecuados.")
        st.stop()

    st.markdown("### 📥 Carga mensual")
    st.caption(
        "Subi el EERR y/o Balance del mes. El archivo se guarda en `data/eerr/` y "
        "se hace backup automatico de la planilla maestra antes de cualquier inserción."
    )

    sub_col1, sub_col2 = st.columns(2)

    # ====== EERR ======
    with sub_col1:
        st.markdown("#### Estado de Resultados (EERR)")
        eerr_up = st.file_uploader(
            "Subir EERR mensual (.xlsx)",
            type=["xlsx", "xlsm"],
            key="eerr_uploader",
        )
        if eerr_up:
            mes_anio = detectar_mes_anio(eerr_up.name)
            st.success(f"📄 {eerr_up.name} — {eerr_up.size/1024:.0f} KB")
            if mes_anio:
                st.info(f"Detectado: mes {mes_anio[0]}, año {mes_anio[1]}")
            else:
                st.warning("⚠️ No se pudo detectar mes/año del nombre. Usá formato: `02 EE.RR Febrero 2026.xlsx`")

            if st.button("💾 Guardar EERR + backup planilla", type="primary", key="save_eerr"):
                try:
                    # 1. Backup de planilla maestra
                    bk = sp.make_backup(sp.PLANIFICACION_FINANCIERA)
                    if bk:
                        st.info(f"📦 Backup creado: {bk.name}")

                    # 2. Guardar EERR en data/eerr/
                    sp.EERR_DIR.mkdir(parents=True, exist_ok=True)
                    target = sp.EERR_DIR / eerr_up.name
                    target.write_bytes(eerr_up.read())
                    st.success(f"✓ Guardado en {target}")

                    # 3. TODO: insertar en hojas P&L y EEFF
                    st.warning(
                        "📝 **TODO de inserción en planilla maestra**: el writer "
                        "automatico se implementará cuando confirmes el mapping de celdas. "
                        "Por ahora el archivo está guardado en `data/eerr/` y la planilla "
                        "fue respaldada. Andrés puede insertarlo manualmente o se ejecutará "
                        "cuando el writer esté wired."
                    )
                    cargar_hoja.clear()
                except Exception as e:
                    st.error(f"❌ Error: {e}")

    # ====== Balance ======
    with sub_col2:
        st.markdown("#### Balance mensual")
        bal_up = st.file_uploader(
            "Subir Balance mensual (.xlsx)",
            type=["xlsx", "xlsm"],
            key="balance_uploader",
        )
        if bal_up:
            mes_anio = detectar_mes_anio(bal_up.name)
            st.success(f"📄 {bal_up.name} — {bal_up.size/1024:.0f} KB")
            if mes_anio:
                st.info(f"Detectado: mes {mes_anio[0]}, año {mes_anio[1]}")
            else:
                st.warning("⚠️ Convención esperada: `Balance_<Mes>_<Año>.xlsx`")

            if st.button("💾 Guardar Balance + backup planilla", type="primary", key="save_bal"):
                try:
                    bk = sp.make_backup(sp.PLANIFICACION_FINANCIERA)
                    if bk:
                        st.info(f"📦 Backup creado: {bk.name}")

                    sp.EERR_DIR.parent.mkdir(parents=True, exist_ok=True)
                    target = sp.EERR_DIR.parent / "balances" / bal_up.name
                    target.parent.mkdir(parents=True, exist_ok=True)
                    target.write_bytes(bal_up.read())
                    st.success(f"✓ Guardado en {target}")

                    st.warning(
                        "📝 **TODO inserción Balance**: idem EERR. El writer en "
                        "hoja `Ref Balances` se conectará cuando se confirme el mapping."
                    )
                    cargar_hoja.clear()
                except Exception as e:
                    st.error(f"❌ Error: {e}")

    st.divider()

    # ====== Acciones rápidas ======
    st.markdown("### 🔧 Acciones rápidas")
    a1, a2, a3 = st.columns(3)
    with a1:
        if st.button("🔄 Refrescar caché"):
            cargar_hoja.clear()
            st.success("Caché invalidada. Recargá la página.")
    with a2:
        if sp.PLANIFICACION_FINANCIERA.exists():
            with open(sp.PLANIFICACION_FINANCIERA, "rb") as f:
                st.download_button(
                    "⬇️ Descargar planilla actual",
                    f.read(),
                    sp.PLANIFICACION_FINANCIERA.name,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
    with a3:
        # Trigger del preview del email del CEO
        if "show_email_preview" not in st.session_state:
            st.session_state.show_email_preview = False
        if st.button("📤 Enviar Reporte Ejecutivo al CEO"):
            st.session_state.show_email_preview = True

    if st.session_state.get("show_email_preview"):
        from email_preview import preview_y_enviar, construir_html_resumen_reportes
        # Recopilar reportes disponibles en data/outputs/
        adjuntos = []
        for pat in ["Reporte_*.xlsx", "reporte_*.xlsx"]:
            adjuntos.extend(str(p) for p in sp.OUTPUTS_DIR.glob(pat))
        adjuntos = sorted(set(adjuntos))[:3]  # max 3
        cuerpo = construir_html_resumen_reportes(
            {"reporte_1": adjuntos[0] if len(adjuntos)>0 else None,
             "reporte_2": adjuntos[1] if len(adjuntos)>1 else None,
             "reporte_3": adjuntos[2] if len(adjuntos)>2 else None,
             "alertas": []},
            datetime.now(),
        )
        result = preview_y_enviar(
            asunto=f"Reportes Ejecutivos UnionX - {datetime.now().strftime('%d/%m/%Y')}",
            cuerpo_html=cuerpo,
            adjuntos=adjuntos,
            modo="reporte",
            key_prefix="planif_reporte",
        )
        if result:
            st.session_state.show_email_preview = False

    # ====== Historial de backups ======
    st.divider()
    st.markdown("### 📜 Backups recientes (últimos 10)")
    if sp.BACKUPS_DIR.exists():
        backups = sorted(sp.BACKUPS_DIR.glob("Planificación Financiera_*.xlsx"), reverse=True)[:10]
        if backups:
            for b in backups:
                ts = datetime.fromtimestamp(b.stat().st_mtime).strftime("%d/%m/%Y %H:%M")
                st.text(f"  · {b.name}  —  {ts}  ({b.stat().st_size//1024} KB)")
        else:
            st.info("Aún no hay backups.")
    else:
        st.info("Directorio de backups no existe todavía.")
