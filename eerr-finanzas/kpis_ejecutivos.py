"""
Consolidador de KPIs ejecutivos — los 5 pilares estrategicos del Plan UnionX 2026-2028.

Funcion principal:
    get_kpis_pilares(mes_actual) -> dict con 5 pilares

Cada pilar tiene: kpi_central, valor, meta, semaforo, tendencia, fuente.
"""
from datetime import datetime
from pathlib import Path
import sys

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))


def _semaforo(valor, umbral_verde, umbral_amarillo, mayor_es_mejor: bool = True):
    """Devuelve 🟢/🟡/🔴 segun el valor y los umbrales.

    Si mayor_es_mejor=True: verde si valor >= verde, amarillo si entre amarillo y verde, rojo si < amarillo.
    Si mayor_es_mejor=False (ej: CCC, donde menos es mejor): inverso.
    """
    if valor is None:
        return "⚪"
    try:
        v = float(valor)
    except Exception:
        return "⚪"
    if mayor_es_mejor:
        if v >= umbral_verde:
            return "🟢"
        if v >= umbral_amarillo:
            return "🟡"
        return "🔴"
    else:
        if v <= umbral_verde:
            return "🟢"
        if v <= umbral_amarillo:
            return "🟡"
        return "🔴"


def get_kpis_pilares(mes_actual: int = None) -> dict:
    """Devuelve los 5 pilares con sus KPIs centrales.

    Returns:
        dict con keys: rentabilidad, liquidez, crecimiento, eficiencia, marca.
        Cada uno es dict con: nombre, kpi_nombre, valor, meta, semaforo, fuente, error.
    """
    if mes_actual is None:
        mes_actual = datetime.now().month
    meses_ytd = max(1, mes_actual - 1)

    pilares = {}

    # ========================================================================
    # PILAR 1: RENTABILIDAD — EBITDA % YTD (proxy: EBIT/Venta)
    # ========================================================================
    try:
        # Importar helper de planificacion (carga Metas 2026)
        sys.path.insert(0, str(PROJECT_ROOT / "eerr-finanzas"))
        # Cargar Metas 2026 directamente (sin pasar por Streamlit cache)
        from openpyxl import load_workbook
        import shared_paths as sp
        wb = load_workbook(sp.PLANIFICACION_FINANCIERA, read_only=True, data_only=True)
        if "Metas 2026" not in wb.sheetnames:
            wb.close()
            raise ValueError("Hoja Metas 2026 no existe")
        ws = wb["Metas 2026"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # bloques: 0-8 Venta, 9-17 Contrib, 18-26 GAV, 27-35 EBIT, 36-44 Utilidad
        venta_res = [v for v in rows[2][1:13] if isinstance(v, (int, float))][:meses_ytd]
        ebit_res = [v for v in rows[29][1:13] if isinstance(v, (int, float))][:meses_ytd]
        contrib_res = [v for v in rows[11][1:13] if isinstance(v, (int, float))][:meses_ytd]

        venta_ytd = sum(venta_res)
        ebit_ytd = sum(ebit_res)
        contrib_ytd = sum(contrib_res)

        ebitda_pct = ebit_ytd / venta_ytd if venta_ytd else None
        margen_contrib = contrib_ytd / venta_ytd if venta_ytd else None

        pilares["rentabilidad"] = {
            "nombre": "Rentabilidad",
            "icono": "💰",
            "kpi_nombre": "EBIT % YTD",
            "valor": ebitda_pct,
            "valor_fmt": f"{ebitda_pct*100:+.1f}%" if ebitda_pct is not None else "—",
            "meta": "≥ 12%",
            "semaforo": _semaforo(ebitda_pct, 0.12, 0.06, mayor_es_mejor=True),
            "fuente": "Metas 2026 (planilla)",
            "extra": f"Margen Contrib YTD: {margen_contrib*100:+.1f}%" if margen_contrib is not None else "",
            "error": None,
        }

        # ========================================================================
        # PILAR 4: EFICIENCIA — Margen contribucion YTD
        # ========================================================================
        pilares["eficiencia"] = {
            "nombre": "Eficiencia",
            "icono": "⚡",
            "kpi_nombre": "Margen Contribución YTD",
            "valor": margen_contrib,
            "valor_fmt": f"{margen_contrib*100:+.1f}%" if margen_contrib is not None else "—",
            "meta": "≥ 35%",
            "semaforo": _semaforo(margen_contrib, 0.35, 0.27, mayor_es_mejor=True),
            "fuente": "Metas 2026 (planilla)",
            "extra": f"Venta YTD: ${venta_ytd/1e6:,.1f}M",
            "error": None,
        }
    except Exception as e:
        for k in ["rentabilidad", "eficiencia"]:
            pilares.setdefault(k, {
                "nombre": k.title(), "icono": "❓", "kpi_nombre": "—",
                "valor": None, "valor_fmt": "—", "meta": "—", "semaforo": "⚪",
                "fuente": "Metas 2026", "error": str(e)[:120]
            })

    # ========================================================================
    # PILAR 2: LIQUIDEZ — CCC (DIO + DSO − DPO)
    # ========================================================================
    try:
        from kpis_odoo import kpi_ccc, get_odoo_client
        odoo = get_odoo_client()
        if odoo is None:
            raise RuntimeError("Odoo no disponible (verificar ANDRES_ODOO_PASSWORD env var)")
        ccc_data = kpi_ccc(odoo)
        ccc_val = ccc_data.get("valor")
        pilares["liquidez"] = {
            "nombre": "Liquidez",
            "icono": "💧",
            "kpi_nombre": "CCC (días)",
            "valor": ccc_val,
            "valor_fmt": f"{ccc_val:,.0f} días" if ccc_val is not None else "—",
            "meta": "≤ 90 días",
            "semaforo": _semaforo(ccc_val, 90, 120, mayor_es_mejor=False),
            "fuente": "Odoo (DIO + DSO - DPO)",
            "extra": (f"DIO {ccc_data['componentes']['DIO']:.0f} · "
                      f"DSO {ccc_data['componentes']['DSO']:.0f} · "
                      f"DPO {ccc_data['componentes']['DPO']:.0f}")
                     if "componentes" in ccc_data and ccc_val is not None else "",
            "error": ccc_data.get("error"),
        }
    except Exception as e:
        pilares["liquidez"] = {
            "nombre": "Liquidez", "icono": "💧", "kpi_nombre": "CCC (días)",
            "valor": None, "valor_fmt": "—", "meta": "≤ 90 días", "semaforo": "⚪",
            "fuente": "Odoo", "extra": "", "error": str(e)[:120]
        }

    # ========================================================================
    # PILAR 3: CRECIMIENTO — Var % ingresos YoY YTD
    # ========================================================================
    try:
        from kpis_odoo import kpi_yoy_ingresos, get_odoo_client
        odoo = get_odoo_client()
        if odoo is None:
            raise RuntimeError("Odoo no disponible")
        yoy_data = kpi_yoy_ingresos(odoo)
        yoy = yoy_data.get("valor")
        pilares["crecimiento"] = {
            "nombre": "Crecimiento",
            "icono": "📈",
            "kpi_nombre": "Ingresos YoY YTD",
            "valor": yoy,
            "valor_fmt": f"{yoy*100:+.1f}%" if yoy is not None else "—",
            "meta": "≥ 25%",
            "semaforo": _semaforo(yoy, 0.25, 0.10, mayor_es_mejor=True),
            "fuente": "Odoo (sale.order)",
            "extra": "",
            "error": yoy_data.get("error"),
        }
    except Exception as e:
        pilares["crecimiento"] = {
            "nombre": "Crecimiento", "icono": "📈", "kpi_nombre": "Ingresos YoY YTD",
            "valor": None, "valor_fmt": "—", "meta": "≥ 25%", "semaforo": "⚪",
            "fuente": "Odoo", "extra": "", "error": str(e)[:120]
        }

    # ========================================================================
    # PILAR 5: MARCA Y CLIENTE — Tasa de recompra
    # ========================================================================
    try:
        from kpis_odoo import kpi_repeat_customer, get_odoo_client
        odoo = get_odoo_client()
        if odoo is None:
            raise RuntimeError("Odoo no disponible")
        rep_data = kpi_repeat_customer(odoo, dias=180)
        rep = rep_data.get("valor")
        pilares["marca"] = {
            "nombre": "Marca/Cliente",
            "icono": "🎯",
            "kpi_nombre": "Repeat Rate (180d)",
            "valor": rep,
            "valor_fmt": f"{rep*100:+.1f}%" if rep is not None else "—",
            "meta": "≥ 25%",
            "semaforo": _semaforo(rep, 0.25, 0.15, mayor_es_mejor=True),
            "fuente": "Odoo (partners ≥2 órdenes 180d)",
            "extra": "",
            "error": rep_data.get("error"),
        }
    except Exception as e:
        pilares["marca"] = {
            "nombre": "Marca/Cliente", "icono": "🎯", "kpi_nombre": "Repeat Rate",
            "valor": None, "valor_fmt": "—", "meta": "≥ 25%", "semaforo": "⚪",
            "fuente": "Odoo", "extra": "", "error": str(e)[:120]
        }

    return pilares


if __name__ == "__main__":
    # Test manual
    pilares = get_kpis_pilares()
    for k, p in pilares.items():
        print(f"{p.get('icono','')} {p.get('nombre')}: {p.get('valor_fmt')} {p.get('semaforo')} (meta {p.get('meta')})")
        if p.get('error'):
            print(f"   error: {p['error']}")
