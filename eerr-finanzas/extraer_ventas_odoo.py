"""
Extrae datos de ventas directamente de Odoo
Genera archivo Excel compatible con el sistema de reportes

Conecta a: https://unionxb2b.odoo.com
Base: bmya-innovatek-sh-prd-6981800
Usuario: andres@grupoeter.cl
"""

import xmlrpc.client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime
import os

class ExtraerVentasOdoo:
    """Extrae ventas de Odoo y genera Excel"""

    def __init__(self):
        self.url = "https://unionxb2b.odoo.com"
        self.db = "bmya-innovatek-sh-prd-6981800"
        self.usuario = "andres@grupoeter.cl"

        # Obtener password del .env
        try:
            from dotenv import load_dotenv
            load_dotenv("../.env")
            self.password = os.getenv("ANDRES_PASSWORD")
        except:
            self.password = None

        self.ruta_destino = Path("../datos_entrada/GoogleSheet_Ventas_Export.xlsx")

        print(f"\n{'='*70}")
        print("EXTRACTOR DE VENTAS - ODOO")
        print(f"{'='*70}")
        print(f"\nURL: {self.url}")
        print(f"Base: {self.db}")
        print(f"Usuario: {self.usuario}")

    def conectar(self) -> bool:
        """Conecta a Odoo"""
        try:
            common = xmlrpc.client.ServerProxy(f'{self.url}/xmlrpc/2/common')
            self.uid = common.authenticate(self.db, self.usuario, self.password, {})

            if not self.uid:
                print("[ERROR] Autenticación fallida")
                return False

            self.models = xmlrpc.client.ServerProxy(f'{self.url}/xmlrpc/2/object')
            print(f"\n[OK] Conectado a Odoo (UID: {self.uid})")
            return True

        except Exception as e:
            print(f"[ERROR] No se pudo conectar a Odoo: {e}")
            return False

    def extraer_ventas_febrero(self) -> list:
        """Extrae órdenes de venta de febrero 2026"""
        try:
            # Búsqueda de SO en febrero 2026
            domain = [
                ('date_order', '>=', '2026-02-01'),
                ('date_order', '<', '2026-03-01'),
                ('state', 'in', ['sale', 'done'])
            ]

            orders = self.models.execute_kw(
                self.db, self.uid, self.password,
                'sale.order', 'search_read',
                [domain],
                {'fields': ['name', 'date_order', 'partner_id', 'amount_total', 'amount_untaxed']}
            )

            print(f"[OK] Encontradas {len(orders)} órdenes de venta febrero")
            return orders

        except Exception as e:
            print(f"[ERROR] No se pudieron extraer órdenes: {e}")
            return []

    def extraer_lineas_venta(self, order_id) -> list:
        """Extrae líneas de detalle de una orden"""
        try:
            lines = self.models.execute_kw(
                self.db, self.uid, self.password,
                'sale.order.line', 'search_read',
                [('order_id', '=', order_id)],
                {'fields': ['product_id', 'name', 'quantity', 'price_unit', 'price_subtotal']}
            )
            return lines
        except:
            return []

    def extraer_facturas_febrero(self) -> list:
        """Extrae facturas de febrero"""
        try:
            domain = [
                ('invoice_date', '>=', '2026-02-01'),
                ('invoice_date', '<', '2026-03-01'),
                ('state', '=', 'posted'),
                ('move_type', '=', 'out_invoice')
            ]

            invoices = self.models.execute_kw(
                self.db, self.uid, self.password,
                'account.move', 'search_read',
                [domain],
                {'fields': ['name', 'invoice_date', 'partner_id', 'amount_total']}
            )

            print(f"[OK] Encontradas {len(invoices)} facturas febrero")
            return invoices

        except Exception as e:
            print(f"[ERROR] No se pudieron extraer facturas: {e}")
            return []

    def generar_excel(self, orders: list, invoices: list) -> bool:
        """Genera archivo Excel con datos de ventas"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ventas Febrero"

            # Encabezado
            headers = ['Fecha', 'Documento', 'Cliente', 'Concepto', 'Cantidad', 'Valor Unitario', 'Subtotal']
            ws.append(headers)

            # Formato header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Cargar datos - Órdenes
            total_general = 0
            for order in orders[:50]:  # Limitar a primeras 50
                fecha = order['date_order'][:10] if order['date_order'] else ""
                doc = order['name']
                cliente = order['partner_id'][1] if order['partner_id'] else ""

                lines = self.extraer_lineas_venta(order['id'])
                if lines:
                    for line in lines:
                        ws.append([
                            fecha,
                            doc,
                            cliente,
                            line['name'][:50] if line['name'] else "",
                            line['quantity'],
                            line['price_unit'],
                            line['price_subtotal']
                        ])
                        total_general += line['price_subtotal']
                else:
                    ws.append([
                        fecha,
                        doc,
                        cliente,
                        "Venta",
                        "",
                        "",
                        order['amount_subtotal']
                    ])
                    total_general += order['amount_subtotal']

            # Ajustar ancho columnas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12

            # Crear carpeta si no existe
            self.ruta_destino.parent.mkdir(parents=True, exist_ok=True)

            # Guardar
            wb.save(self.ruta_destino)
            print(f"[OK] Excel generado: {self.ruta_destino.name}")
            print(f"     Total registros: {len(orders)} órdenes")
            print(f"     Valor total: ${total_general:,.0f}")
            return True

        except Exception as e:
            print(f"[ERROR] No se pudo generar Excel: {e}")
            return False

    def ejecutar(self):
        """Ejecuta extracción completa"""
        if not self.conectar():
            print("\n[INSTRUCCIONES MANUALES]")
            print("1. Abre: https://unionxb2b.odoo.com")
            print("2. Ve a: Ventas > Órdenes de Venta")
            print("3. Filtra por fecha: Febrero 2026")
            print("4. Descarga como Excel")
            print("5. Coloca en: UNION X - IA/datos_entrada/")
            print("6. Renombra a: GoogleSheet_Ventas_Export.xlsx")
            return False

        # Extraer datos
        orders = self.extraer_ventas_febrero()
        invoices = self.extraer_facturas_febrero()

        if not orders and not invoices:
            print("\n[AVISO] No hay datos de febrero en Odoo")
            return False

        # Generar Excel
        return self.generar_excel(orders, invoices)


# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    extractor = ExtraerVentasOdoo()
    exito = extractor.ejecutar()

    if exito:
        print("\n[LISTO] Datos de ventas extraídos de Odoo")
    else:
        print("\n[ALTERNATIVA] Descarga manual desde Odoo")
        print("URL: https://unionxb2b.odoo.com")
        print("Usuario: andres@grupoeter.cl")
