"""
GENERADOR DE DATOS DE EJEMPLO
Crea todos los archivos necesarios para testear los reportes automáticos
Sin depender de datos reales de Odoo/Excel

Ejecutar:
  python generar_datos_ejemplo.py

Resultado:
  data/planillas/Presupuesto_Febrero_2026.xlsx
  data/outputs/odoo_export_20260401.json
  data/outputs/comex_maestra_cc.json
  data/planillas/Planificación Financiera.xlsx
  data/planillas/Sueldos_Febrero_2026.xlsx
  data/planillas/Balance_Febrero_2026.xlsx
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime


def crear_presupuesto_excel():
    """Crea Presupuesto_Febrero_2026.xlsx"""

    ruta = Path("data/planillas/Presupuesto_Febrero_2026.xlsx")
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    # Headers
    ws['A1'] = "Concepto"
    ws['B1'] = "Presupuesto"
    ws['C1'] = "Real"
    ws['D1'] = "Desvío %"

    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

    # Datos
    datos = [
        ("Ventas", 500000, 475000),
        ("Margen Bruto %", 0.30, 0.28),
        ("Costo Venta", 350000, 342000),
        ("Comisiones", 25000, 28000),
        ("Gastos Operacionales", 50000, 52000),
        ("Remuneraciones", 80000, 80000),
    ]

    fila = 2
    for concepto, presupuesto, real in datos:
        ws.cell(row=fila, column=1, value=concepto)
        ws.cell(row=fila, column=2, value=presupuesto)
        ws.cell(row=fila, column=3, value=real)

        if isinstance(presupuesto, (int, float)) and presupuesto != 0:
            desvio = (real - presupuesto) / presupuesto
            ws.cell(row=fila, column=4, value=desvio).number_format = '0.0%'

        fila += 1

    wb.save(ruta)
    print(f"✅ Presupuesto creado: {ruta}")


def crear_odoo_export_json():
    """Crea odoo_export_20260401.json"""

    ruta = Path("data/outputs/odoo_export_20260401.json")
    ruta.parent.mkdir(parents=True, exist_ok=True)

    datos = {
        "inventory": {
            "total_unidades": 5240,
            "valor_stock": 125000,
            "ocupacion_pct": 0.78,
            "items_bajo_minimo": [
                {"sku": "SKU-001", "stock_actual": 5, "minimo": 10, "pct_restante": 50},
                {"sku": "SKU-045", "stock_actual": 3, "minimo": 8, "pct_restante": 37.5}
            ],
            "rotacion_promedio": 2.15,
            "sku_total": 150,
            "sku_activos": 125,
            "movimiento_semanal": {
                "entradas": 800,
                "salidas": 850,
                "neto": -50
            }
        },
        "fulfillment": {
            "pedidos_pendientes": 12,
            "pedidos_despachados_hoy": 28,
            "pedidos_ontime_pct": 96.5,
            "tiempo_promedio_fulfillment_dias": 2.1,
            "ordenes_atrasadas": 1,
            "tasa_fulfillment_perfecta": 98.2
        },
        "ventas_por_canal": [
            {"canal": "Recíbelo", "ventas": 250000, "margen_contrib": 0.20},
            {"canal": "Blue Express", "ventas": 150000, "margen_contrib": 0.30},
            {"canal": "Grupo Eter", "ventas": 75000, "margen_contrib": 0.28},
        ]
    }

    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(datos, f, indent=2, ensure_ascii=False)

    print(f"✅ Odoo export creado: {ruta}")


def crear_comex_maestra_json():
    """Crea comex_maestra_cc.json"""

    ruta = Path("data/outputs/comex_maestra_cc.json")
    ruta.parent.mkdir(parents=True, exist_ok=True)

    datos = {
        "importaciones_activas": [
            {
                "id": "IMP-001",
                "proveedor": "Steven",
                "cc": "DISTRIBUCION",
                "status": "en_transito",
                "eta_original": "2026-04-15",
                "eta_actual": "2026-04-15",
                "dias_retraso": 0,
                "costo": 8000,
                "costeo_cn": 5000,
                "flete": 3000,
                "margen_importacion_pct": 18,
                "lead_time_promedio": 25
            },
            {
                "id": "IMP-002",
                "proveedor": "Steven",
                "cc": "LOGISTICA",
                "status": "en_transito",
                "eta_original": "2026-04-20",
                "eta_actual": "2026-04-25",
                "dias_retraso": 5,
                "costo": 5000,
                "costeo_cn": 3000,
                "flete": 2000,
                "margen_importacion_pct": 15,
                "lead_time_promedio": 25
            },
            {
                "id": "IMP-003",
                "proveedor": "Steven",
                "cc": "E-COMMERCE",
                "status": "en_puerto",
                "eta_original": "2026-04-10",
                "eta_actual": "2026-04-10",
                "dias_retraso": 0,
                "costo": 6500,
                "costeo_cn": 4000,
                "flete": 2500,
                "margen_importacion_pct": 20,
                "lead_time_promedio": 25
            }
        ]
    }

    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(datos, f, indent=2, ensure_ascii=False)

    print(f"✅ COMEX maestra creado: {ruta}")


def crear_planificacion_financiera_excel():
    """Crea Planificación Financiera.xlsx"""

    ruta = Path("data/planillas/Planificación Financiera.xlsx")
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planificación"

    # Header
    ws['A1'] = "PLANIFICACIÓN FINANCIERA FEBRERO 2026"
    ws['A1'].font = Font(bold=True, size=14)

    fila = 3

    # INGRESOS
    ws[f'A{fila}'] = "INGRESOS OPERACIONALES"
    ws[f'A{fila}'].font = Font(bold=True, color="FFFFFF")
    ws[f'A{fila}'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

    fila += 1
    datos_ingresos = [
        ("Ventas Recíbelo", 250000),
        ("Ventas Blue Express", 150000),
        ("Ventas Grupo Eter", 75000),
        ("Otros Ingresos", 25000),
    ]

    for concepto, monto in datos_ingresos:
        ws.cell(row=fila, column=1, value=concepto)
        ws.cell(row=fila, column=2, value=monto).number_format = '$#,##0'
        fila += 1

    ws.cell(row=fila, column=1, value="TOTAL INGRESOS").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=500000).number_format = '$#,##0'
    fila += 2

    # COSTOS
    ws[f'A{fila}'] = "COSTO DE VENTA"
    ws[f'A{fila}'].font = Font(bold=True, color="FFFFFF")
    ws[f'A{fila}'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

    fila += 1
    ws.cell(row=fila, column=1, value="COGS")
    ws.cell(row=fila, column=2, value=342000).number_format = '$#,##0'
    fila += 1

    ws.cell(row=fila, column=1, value="Margen Bruto").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=158000).number_format = '$#,##0'
    fila += 2

    # REMUNERACIONES
    ws[f'A{fila}'] = "REMUNERACIONES"
    ws[f'A{fila}'].font = Font(bold=True, color="FFFFFF")
    ws[f'A{fila}'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

    fila += 1
    datos_remuneraciones = [
        ("Sueldos", 80000),
        ("Honorarios", 10000),
        ("Rendiciones", 5000),
    ]

    for concepto, monto in datos_remuneraciones:
        ws.cell(row=fila, column=1, value=concepto)
        ws.cell(row=fila, column=2, value=monto).number_format = '$#,##0'
        fila += 1

    ws.cell(row=fila, column=1, value="TOTAL REMUNERACIONES").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=95000).number_format = '$#,##0'
    fila += 2

    # OTROS GASTOS
    ws[f'A{fila}'] = "GASTOS OPERACIONALES"
    ws[f'A{fila}'].font = Font(bold=True, color="FFFFFF")
    ws[f'A{fila}'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

    fila += 1
    datos_gastos = [
        ("Marketing", 20000),
        ("Operacionales", 35000),
        ("Administrativos", 15000),
    ]

    for concepto, monto in datos_gastos:
        ws.cell(row=fila, column=1, value=concepto)
        ws.cell(row=fila, column=2, value=monto).number_format = '$#,##0'
        fila += 1

    ws.cell(row=fila, column=1, value="TOTAL GASTOS").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=70000).number_format = '$#,##0'
    fila += 2

    # EBIT
    ws[f'A{fila}'] = "EBIT"
    ws[f'A{fila}'].font = Font(bold=True, color="2E7D32")
    ws.cell(row=fila, column=2, value=-7000).number_format = '$#,##0'
    ws.cell(row=fila, column=2).font = Font(bold=True)

    wb.save(ruta)
    print(f"✅ Planificación Financiera creado: {ruta}")


def crear_sueldos_excel():
    """Crea Sueldos_Febrero_2026.xlsx"""

    ruta = Path("data/planillas/Sueldos_Febrero_2026.xlsx")
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Headers
    ws['A1'] = "Empleado"
    ws['B1'] = "Monto"
    ws['C1'] = "Centro Costos"

    for cell in ['A1', 'B1', 'C1']:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

    # Datos
    datos = [
        ("Gerente Comercial", 3500, "COMERCIAL"),
        ("Jefe Operaciones", 3200, "LOGISTICA"),
        ("Asistente Logística", 2000, "LOGISTICA"),
        ("Contador", 2500, "FINANZAS"),
        ("Gerente Finanzas (Andrés)", 4000, "FINANZAS"),
        ("Community Manager", 2200, "COMERCIAL"),
        ("Operator Almacén", 1800, "LOGISTICA"),
        ("Operario Almacén", 1800, "LOGISTICA"),
    ]

    fila = 2
    total = 0

    for empleado, monto, cc in datos:
        ws.cell(row=fila, column=1, value=empleado)
        ws.cell(row=fila, column=2, value=monto).number_format = '$#,##0'
        ws.cell(row=fila, column=3, value=cc)
        total += monto
        fila += 1

    ws.cell(row=fila, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=total).number_format = '$#,##0'
    ws.cell(row=fila, column=2).font = Font(bold=True)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20

    wb.save(ruta)
    print(f"✅ Sueldos creado: {ruta}")


def crear_balance_excel():
    """Crea Balance_Febrero_2026.xlsx"""

    ruta = Path("data/planillas/Balance_Febrero_2026.xlsx")
    ruta.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Sheet Deuda
    ws_deuda = wb.active
    ws_deuda.title = "Deuda"

    ws_deuda['A1'] = "Préstamo"
    ws_deuda['B1'] = "Monto"
    ws_deuda['C1'] = "Tasa Anual"
    ws_deuda['D1'] = "Plazo Meses"

    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws_deuda[cell].font = Font(bold=True, color="FFFFFF")
        ws_deuda[cell].fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")

    datos_deuda = [
        ("Banco A - Línea de Crédito", 300000, 0.08, 48),
        ("Leasing Maquinaria", 200000, 0.06, 36),
    ]

    fila = 2
    total_deuda = 0

    for prestamo, monto, tasa, plazo in datos_deuda:
        ws_deuda.cell(row=fila, column=1, value=prestamo)
        ws_deuda.cell(row=fila, column=2, value=monto).number_format = '$#,##0'
        ws_deuda.cell(row=fila, column=3, value=tasa).number_format = '0.0%'
        ws_deuda.cell(row=fila, column=4, value=plazo)
        total_deuda += monto
        fila += 1

    ws_deuda.cell(row=fila, column=1, value="TOTAL DEUDA").font = Font(bold=True)
    ws_deuda.cell(row=fila, column=2, value=total_deuda).number_format = '$#,##0'
    ws_deuda.cell(row=fila, column=2).font = Font(bold=True)

    # Sheet Balance
    ws_balance = wb.create_sheet("Balance")

    ws_balance['A1'] = "Concepto"
    ws_balance['B1'] = "Monto"

    for cell in ['A1', 'B1']:
        ws_balance[cell].font = Font(bold=True, color="FFFFFF")
        ws_balance[cell].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

    datos_balance = [
        ("ACTIVOS", None),
        ("Cuentas por Cobrar", 120000),
        ("Inventario", 80000),
        ("Efectivo", 50000),
        ("TOTAL ACTIVOS", 250000),
        ("", None),
        ("PASIVOS", None),
        ("Cuentas por Pagar", 95000),
        ("Deuda Corto Plazo", 100000),
        ("TOTAL PASIVOS", 195000),
        ("", None),
        ("PATRIMONIO", 55000),
    ]

    fila = 2
    for concepto, monto in datos_balance:
        ws_balance.cell(row=fila, column=1, value=concepto)
        if monto is not None:
            ws_balance.cell(row=fila, column=2, value=monto).number_format = '$#,##0'

        if "TOTAL" in concepto:
            ws_balance.cell(row=fila, column=1).font = Font(bold=True)
            ws_balance.cell(row=fila, column=2).font = Font(bold=True)

        fila += 1

    ws_deuda.column_dimensions['A'].width = 30
    ws_balance.column_dimensions['A'].width = 30

    wb.save(ruta)
    print(f"✅ Balance creado: {ruta}")


# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("GENERANDO DATOS DE EJEMPLO PARA TESTING")
    print("=" * 70)
    print()

    crear_presupuesto_excel()
    crear_odoo_export_json()
    crear_comex_maestra_json()
    crear_planificacion_financiera_excel()
    crear_sueldos_excel()
    crear_balance_excel()

    print()
    print("=" * 70)
    print("✅ TODOS LOS DATOS DE EJEMPLO CREADOS")
    print("=" * 70)
    print()
    print("Ahora puedes ejecutar:")
    print("  python orquestador_reportes.py")
    print()
    print("Los reportes se generarán en: data/outputs/")
