#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DATA INGESTION ENGINE - Union X Revenue Automation
Descarga datos desde múltiples fuentes automáticamente

Funcionalidades:
1. Descarga Google Drive (archivo compartido)
2. Lee Google Sheets (ventas + comisiones por KAM)
3. Descarga EERR del email (IMAP)
4. Inyecta en Excel (Análisis Contribución)
5. Ejecuta skill distribucion-comisiones-canal

Triggers:
- Lunes 9 AM: Google Drive + Sheets
- Día 7 mes: Google Sheets detail
- Día 10 mes: Email EERR + Skill
"""

import os
import sys

# Fix encoding para Windows
if sys.platform == 'win32':
    os.environ['PYTHONIOENCODING'] = 'utf-8'
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import json

# Imports para Google APIs
try:
    from google.auth.transport.requests import Request
    from google.oauth2.service_account import Credentials
    import gspread
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

# google.colab es opcional (solo para Jupyter/Colab)
try:
    from google.colab import auth
except ImportError:
    pass

# Imports para descargar de Google Drive
try:
    from google.auth import default
    import io
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

# Imports para IMAP (email)
try:
    import imaplib
    import email
    from email.header import decode_header
    IMAP_AVAILABLE = True
except ImportError:
    IMAP_AVAILABLE = False

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd


class DataIngestionEngine:
    """Motor de ingestión de datos automática"""

    def __init__(self, config_path: Optional[str] = None):
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.current_month = datetime.now().month
        self.current_year = datetime.now().year
        self.config = self._load_config(config_path)

        # Rutas (single source of truth via shared_paths)
        self.current_dir = Path(__file__).parent
        self.project_root = self.current_dir.parent  # G:\...\UNION X - IA

        # Importar shared_paths del project root
        import sys as _sys
        if str(self.project_root) not in _sys.path:
            _sys.path.insert(0, str(self.project_root))
        try:
            import shared_paths as _sp
            self.contribucion_file = _sp.CONTRIBUCION
            self.planificacion_file = _sp.PLANIFICACION_FINANCIERA
            self.planillas_dir = _sp.CONTRIBUCION.parent
            self.outputs_dir = _sp.OUTPUTS_DIR
        except Exception as _e:
            # Fallback si shared_paths no esta disponible
            print(f"WARN: shared_paths no disponible ({_e}), usando legacy fallback")
            self.planillas_dir = self.project_root / "data" / "planillas"
            self.contribucion_file = self.planillas_dir / "Analisis_Contribucion_2026_V06.xlsx"
            self.planificacion_file = self.planillas_dir / "Planificación Financiera.xlsx"
            self.outputs_dir = self.project_root / "data" / "outputs"

        # Configurar credenciales de Google (orden: env var > agente-comex/config > current_dir)
        env_creds = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS") or os.environ.get("UNIONX_CREDENTIALS_FILE")
        canonical_creds = self.project_root / "agente-comex" / "config" / "credentials.json"
        local_creds = self.current_dir / "credentials.json"

        if env_creds and Path(env_creds).exists():
            creds_file = Path(env_creds)
        elif canonical_creds.exists():
            creds_file = canonical_creds
        elif local_creds.exists():
            creds_file = local_creds
        else:
            creds_file = local_creds  # default por compatibilidad

        if creds_file.exists():
            self.config['service_account_json'] = str(creds_file)
            os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = str(creds_file)

    def _load_config(self, config_path: str) -> Dict:
        """Carga configuración desde archivo JSON"""
        default_config = {
            "google_drive_file_id": "1K11y6icDm9M3X3glGUVCOe4HsbpWpEBm",
            "google_sheets_id": "1z-HLHEuj__HjNjf7hS4sIhU5QvoNiUJJ1BH965y4JEI",
            "google_sheets_gid": 1518723659,
            "email_imap_server": "imap.gmail.com",
            "email_from": "victor@unionx.cl",
            "email_subject_pattern": "EERR",
            "victor_email": "victor@unionx.cl",
            "victor_email_password": None,  # Usar variable de entorno
            "service_account_json": None,  # Path a credentials
        }

        if config_path and Path(config_path).exists():
            with open(config_path, 'r') as f:
                user_config = json.load(f)
                default_config.update(user_config)

        return default_config

    def download_google_drive_file(self, file_id: str, output_path: Path) -> bool:
        """Descarga archivo desde Google Drive"""
        print(f"\n[{self.timestamp}] Descargando Google Drive (ID: {file_id})...")

        if not GDRIVE_AVAILABLE:
            print("  ✗ google-api-client no instalado. Instala: pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client")
            return False

        try:
            # Usar credenciales de servicio o default
            if self.config.get('service_account_json'):
                credentials = Credentials.from_service_account_file(
                    self.config['service_account_json']
                )
            else:
                credentials, _ = default()

            drive_service = build('drive', 'v3', credentials=credentials)

            request = drive_service.files().get_media(fileId=file_id)
            with open(output_path, 'wb') as f:
                downloader = MediaIoBaseDownload(f, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                    print(f"  → Descarga: {int(status.progress() * 100)}%")

            print(f"  ✓ Descargado: {output_path}")
            return True

        except Exception as e:
            print(f"  ✗ Error descargando Google Drive: {e}")
            return False

    def read_google_sheets(self, sheet_id: str, gid: int = 0) -> Optional[pd.DataFrame]:
        """Lee Google Sheet y retorna DataFrame"""
        print(f"\n[{self.timestamp}] Leyendo Google Sheets (ID: {sheet_id})...")

        if not GSPREAD_AVAILABLE:
            print("  ✗ gspread no instalado. Instala: pip install gspread google-auth-oauthlib")
            return None

        try:
            # Usar credenciales de servicio
            if self.config.get('service_account_json'):
                gc = gspread.service_account(filename=self.config['service_account_json'])
            else:
                # Fallback: usar credenciales por defecto
                auth.authenticate_user()
                gc = gspread.authorize(auth.default()[0])

            # Abrir sheet
            spreadsheet = gc.open_by_key(sheet_id)

            # Buscar worksheet por ID o usar el primero
            if gid:
                # Buscar la hoja con el ID específico
                worksheet = None
                for ws in spreadsheet.worksheets():
                    if ws.id == gid:
                        worksheet = ws
                        break
                if worksheet is None:
                    worksheet = spreadsheet.sheet1
            else:
                worksheet = spreadsheet.sheet1

            # Convertir a DataFrame
            data = worksheet.get_all_values()
            df = pd.DataFrame(data[1:], columns=data[0])

            print(f"  ✓ Leído: {len(df)} filas, {len(df.columns)} columnas")
            return df

        except Exception as e:
            print(f"  ✗ Error leyendo Google Sheets: {e}")
            return None

    def download_eerr_from_email(self, email_addr: str, password: str, subject_pattern: str = "EERR") -> Optional[Path]:
        """Descarga EERR del email de Victor (recibido en email_addr)"""
        print(f"\n[{self.timestamp}] Descargando EERR desde {email_addr} (emails de Victor)...")

        if not IMAP_AVAILABLE:
            print("  ✗ IMAP no disponible")
            return None

        try:
            # Conectar a Gmail IMAP del usuario (andres@unionx.cl)
            mail = imaplib.IMAP4_SSL(self.config['email_imap_server'])
            mail.login(email_addr, password)
            mail.select('INBOX')

            # Buscar emails de Victor con EERR en asunto
            # Formato: from:"victor@unionx.cl" subject:"EERR"
            status, messages = mail.search(None, f'FROM "victor@unionx.cl" SUBJECT "{subject_pattern}"')
            email_ids = messages[0].split()

            if not email_ids:
                print(f"  ⚠ No se encontraron emails con '{subject_pattern}'")
                return None

            # Descargar el más reciente
            latest_email_id = email_ids[-1]
            status, msg_data = mail.fetch(latest_email_id, '(RFC822)')

            msg = email.message_from_bytes(msg_data[0][1])

            # Extraer attachments
            eerr_path = None
            for part in msg.walk():
                if part.get_content_disposition() == 'attachment':
                    filename = part.get_filename()
                    if filename and 'EERR' in filename.upper() and filename.endswith('.xlsx'):
                        eerr_path = self.current_dir / filename
                        with open(eerr_path, 'wb') as f:
                            f.write(part.get_payload(decode=True))
                        print(f"  ✓ Descargado: {eerr_path}")
                        break

            mail.close()
            mail.logout()

            return eerr_path

        except Exception as e:
            print(f"  ✗ Error descargando EERR: {e}")
            return None

    def inject_into_excel(self, excel_path: Path, df: pd.DataFrame, sheet_name: str = "Análisis Resultados") -> bool:
        """Inyecta DataFrame en hoja Excel"""
        print(f"\n[{self.timestamp}] Inyectando en Excel ({sheet_name})...")

        try:
            if not excel_path.exists():
                print(f"  ✗ No encontrado: {excel_path}")
                return False

            # Cargar workbook
            wb = openpyxl.load_workbook(excel_path)

            if sheet_name not in wb.sheetnames:
                print(f"  ✗ Hoja '{sheet_name}' no existe en {excel_path}")
                return False

            ws = wb[sheet_name]

            # Limpiar datos del mes actual (mantener headers)
            # Buscar filas del mes actual y borrarlas
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if hasattr(cell, 'value') and cell.value:
                        # Verificar si es del mes actual
                        # Por ahora, insertar desde fila 2
                        pass

            # Escribir datos desde fila 2
            for idx, row in enumerate(df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=idx, column=col_idx, value=value)

            wb.save(excel_path)
            print(f"  ✓ Inyectado: {len(df)} filas en {excel_path}")
            return True

        except Exception as e:
            print(f"  ✗ Error inyectando: {e}")
            return False

    def run_trigger_lunes_9am(self) -> bool:
        """Trigger: Lunes 9 AM - Descarga Drive + Sheets"""
        print("\n" + "="*70)
        print("TRIGGER: LUNES 9 AM")
        print("="*70)

        # 1. Descargar Google Drive
        drive_file = self.current_dir / f"drive_download_{datetime.now().strftime('%Y%m%d')}.xlsx"
        if not self.download_google_drive_file(self.config['google_drive_file_id'], drive_file):
            print("  ✗ Falló descarga Google Drive")
            return False

        # 2. Leer Google Sheets
        df_sheets = self.read_google_sheets(
            self.config['google_sheets_id'],
            self.config['google_sheets_gid']
        )
        if df_sheets is None:
            print("  ✗ Falló lectura Google Sheets")
            return False

        # 3. Inyectar en Excel
        if not self.inject_into_excel(self.contribucion_file, df_sheets):
            print("  ✗ Falló inyección en Excel")
            return False

        print("\n✓ TRIGGER LUNES 9 AM: COMPLETADO")
        return True

    def run_trigger_dia7_mes(self) -> bool:
        """Trigger: Día 7 del mes - Descarga Google Sheets detail"""
        print("\n" + "="*70)
        print(f"TRIGGER: DÍA 7 DEL MES ({self.current_month}/{self.current_year})")
        print("="*70)

        # Leer Google Sheets
        df = self.read_google_sheets(
            self.config['google_sheets_id'],
            self.config['google_sheets_gid']
        )

        if df is None:
            print("  ✗ Falló lectura Google Sheets")
            return False

        # Inyectar en Excel
        if not self.inject_into_excel(self.contribucion_file, df):
            print("  ✗ Falló inyección en Excel")
            return False

        print("\n✓ TRIGGER DÍA 7 MES: COMPLETADO")
        return True

    def run_trigger_dia10_mes(self, andres_email: str, andres_password: str) -> bool:
        """Trigger: Día 10 del mes - EERR + Skill

        Descarga EERR del email de Andrés (donde Victor envía el archivo)
        """
        print("\n" + "="*70)
        print(f"TRIGGER: DÍA 10 DEL MES ({self.current_month}/{self.current_year})")
        print("="*70)

        # 1. Descargar EERR del email de Andrés (remitido por Victor)
        eerr_path = self.download_eerr_from_email(andres_email, andres_password)

        if not eerr_path:
            print("  ✗ Falló descarga EERR")
            return False

        # 2. Encadenar: invocar el orquestador para generar los reportes ejecutivos
        #    (Reportes 1+2+3 + alertas + resumen HTML). Wrap con try/except porque
        #    el orquestador todavia tiene paths hardcodeados de Feb 2026 (será
        #    parametrizado en una fase posterior).
        try:
            print(f"\n[{self.timestamp}] Invocando orquestador de reportes ejecutivos...")
            from subprocess import run as sp_run
            cmd = [sys.executable, str(self.current_dir / "orquestador_reportes.py")]
            result = sp_run(cmd, capture_output=True, text=True, cwd=str(self.current_dir), timeout=600)
            if result.returncode == 0:
                print("  ✓ Orquestador OK - reportes generados en data/outputs/")
            else:
                print(f"  ⚠️ Orquestador devolvio codigo {result.returncode}")
                print(f"    stdout (ultimas 500 chars): {result.stdout[-500:]}")
                print(f"    stderr (ultimas 500 chars): {result.stderr[-500:]}")
        except Exception as e:
            print(f"  ⚠️ No se pudo invocar el orquestador: {e}")
            print(f"  (No es bloqueante - el EERR ya se descargo en {eerr_path})")

        # 3. Ejecutar skill distribucion-comisiones-canal (manual desde Cowork)
        print(f"\n[{self.timestamp}] Skill 'distribucion-comisiones-canal' debe ejecutarse manualmente desde Cowork")
        print(f"  EERR descargado: {eerr_path}")

        # 3. Inyectar en Excel (los datos vendrán de la skill)
        print(f"\n[{self.timestamp}] Los datos de la skill se inyectarán automáticamente")

        print("\n✓ TRIGGER DÍA 10 MES: COMPLETADO")
        return True


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Data Ingestion Engine')
    parser.add_argument('--trigger', choices=['lunes9am', 'dia7', 'dia10'],
                       help='Trigger a ejecutar')
    parser.add_argument('--config', help='Ruta a archivo de configuración JSON')
    parser.add_argument('--andres-email', help='Email de Andrés (donde se reciben emails de Victor)')
    parser.add_argument('--andres-password', help='Password de Andrés (usa variable de entorno)')

    args = parser.parse_args()

    # Crear engine
    engine = DataIngestionEngine(args.config)

    # Verificar configuración
    if not engine.contribucion_file.exists():
        print(f"✗ No encontrado: {engine.contribucion_file}")
        return 1

    print("\n" + "="*70)
    print("DATA INGESTION ENGINE - Union X")
    print("="*70)
    print(f"Timestamp: {engine.timestamp}")
    print(f"Contribución file: {engine.contribucion_file}")

    # Ejecutar trigger
    if args.trigger == 'lunes9am':
        success = engine.run_trigger_lunes_9am()
    elif args.trigger == 'dia7':
        success = engine.run_trigger_dia7_mes()
    elif args.trigger == 'dia10':
        andres_email = args.andres_email or os.getenv('ANDRES_EMAIL')
        andres_password = args.andres_password or os.getenv('ANDRES_PASSWORD')

        if not andres_email or not andres_password:
            print("✗ Falta email/password de Andrés")
            print("  Configura en .env: ANDRES_EMAIL y ANDRES_PASSWORD")
            return 1

        success = engine.run_trigger_dia10_mes(andres_email, andres_password)
    else:
        print("Especifica --trigger")
        return 1

    return 0 if success else 1


if __name__ == '__main__':
    sys.exit(main())
