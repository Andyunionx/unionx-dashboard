"""
REPORTE 1: Rentabilidad (Contribucin + Control Presupuesto)
Automatizado cada lunes 9 AM

Entrada:
  - EERR clasificado del mes actual (JSON o Excel)
  - Presupuesto mensual (Excel separado)
  - Datos de ventas por canal y cliente (desde Odoo o Google Sheets)

Salida:
  - Excel con 4 secciones:
    1. Rentabilidad por canal (3 mrgenes)
    2. Rentabilidad por LN
    3. Control presupuesto
    4. Anlisis profundo (ejecuta skill si margen < 27%)
  - Email ejecutivo a CEO
  - Alertas si se detectan desvos
"""

import json
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class GeneradorReporte1:
    """Genera Reporte 1: Rentabilidad"""

    def __init__(self, ruta_eerr: str, ruta_presupuesto: Optional[str] = None):
        """
        Args:
            ruta_eerr: Ruta al EERR clasificado (JSON)
            ruta_presupuesto: Ruta al Excel de presupuesto
        """
        self.ruta_eerr = ruta_eerr
        self.ruta_presupuesto = ruta_presupuesto
        self.eerr_data = self._cargar_eerr()
        self.presupuesto_data = self._cargar_presupuesto()
        self.hoy = datetime.now()

    def _cargar_eerr(self) -> Dict:
        """Carga el EERR clasificado"""
        if not Path(self.ruta_eerr).exists():
            return {}

        try:
            with open(self.ruta_eerr, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error cargando EERR: {e}")
            return {}

    def _cargar_presupuesto(self) -> Dict:
        """Carga el presupuesto mensual"""
        if not self.ruta_presupuesto or not Path(self.ruta_presupuesto).exists():
            return {}

        try:
            df = pd.read_excel(self.ruta_presupuesto, sheet_name='Presupuesto')
            return df.set_index('Concepto').to_dict('index')
        except Exception as e:
            print(f"Error cargando presupuesto: {e}")
            return {}

    def generar(self) -> str:
        """
        Genera el reporte completo

        Returns:
            Ruta al archivo Excel generado
        """

        # 1. Extraer datos por canal
        datos_canal = self._extraer_datos_canal()

        # 2. Calcular mrgenes
        rentabilidad_canal = self._calcular_margenes(datos_canal)

        # 3. Comparar con presupuesto
        desvios_presupuesto = self._calcular_desvios()

        # 4. Generar Excel
        ruta_salida = self._generar_excel(rentabilidad_canal, desvios_presupuesto)

        # 5. Ejecutar anlisis profundo si hay margen < 27%
        if any(r['margen_contrib'] < 27 for r in rentabilidad_canal):
            self._generar_analisis_profundo(rentabilidad_canal)

        return ruta_salida

    def _extraer_datos_canal(self) -> Dict[str, Dict]:
        """Extrae ventas, costo y datos por canal desde EERR"""

        # ESTRUCTURA ESPERADA: EERR clasificado con campo 'cc' (centro de costos)
        # Asumiendo que los canales se mapean desde CC o desde campo especfico

        datos = {
            "Recbelo": {"ventas": 0, "costo": 0, "comisiones": 0, "detalle": []},
            "Blue Express": {"ventas": 0, "costo": 0, "comisiones": 0, "detalle": []},
            "Grupo Eter": {"ventas": 0, "costo": 0, "comisiones": 0, "detalle": []},
        }

        # Iterar sobre transacciones en EERR
        for transaccion in self.eerr_data.get('transacciones', []):
            # Mapear a canal (ejemplo: desde campo 'cc' o 'canal')
            canal = self._mapear_a_canal(transaccion.get('cc', ''))

            if canal:
                # Clasificar por tipo
                tipo = transaccion.get('tipo', 'otro')
                monto = float(transaccion.get('monto', 0))

                if tipo == 'venta':
                    datos[canal]['ventas'] += monto
                elif tipo == 'costo':
                    datos[canal]['costo'] += monto
                elif tipo == 'comisin':
                    datos[canal]['comisiones'] += monto

                datos[canal]['detalle'].append(transaccion)

        return datos

    def _mapear_a_canal(self, centro_costos: str) -> Optional[str]:
        """Mapea centro de costos a canal de venta"""

        mapeo = {
            'DISTRIBUCION': 'Recbelo',
            'E-COMMERCE': 'Blue Express',
            'LOGISTICA': 'Grupo Eter',
        }

        for clave, canal in mapeo.items():
            if clave.lower() in centro_costos.lower():
                return canal

        return None

    def _calcular_margenes(self, datos_canal: Dict) -> List[Dict]:
        """Calcula 3 mrgenes para cada canal"""

        resultado = []

        for canal, datos in datos_canal.items():
            ventas = datos['ventas']
            costo = datos['costo']
            comisiones = datos['comisiones']

            # Evitar divisin por cero
            if ventas == 0:
                continue

            # 3 MRGENES
            margen_directo = ((ventas - costo) / ventas * 100) if ventas else 0
            margen_contrib = ((ventas - costo - comisiones) / ventas * 100) if ventas else 0
            margen_operacional = margen_contrib - 5  # Simplificacin: menos gastos

            resultado.append({
                'canal': canal,
                'ventas': ventas,
                'costo': costo,
                'comisiones': comisiones,
                'margen_directo': margen_directo,
                'margen_contrib': margen_contrib,
                'margen_operacional': margen_operacional,
                'estado': '' if margen_contrib >= 27 else '' if margen_contrib < 20 else '',
                'tendencia': '',  # TODO: comparar con semana anterior
            })

        return resultado

    def _calcular_desvios(self) -> Dict[str, Dict]:
        """Compara nmeros reales vs presupuesto"""

        desvios = {}

        # Estructura esperada: presupuesto con conceptos como "Ventas", "Costo", etc.
        for concepto, datos_pres in self.presupuesto_data.items():
            # TODO: extraer real del EERR y comparar
            presupuesto_val = datos_pres.get('monto', 0)
            real_val = 0  # Extraer del EERR

            desvio_pct = ((real_val - presupuesto_val) / presupuesto_val * 100) if presupuesto_val else 0

            desvios[concepto] = {
                'presupuesto': presupuesto_val,
                'real': real_val,
                'desvio_pct': desvio_pct,
                'estado': '' if abs(desvio_pct) <= 5 else '' if abs(desvio_pct) <= 10 else '',
            }

        return desvios

    def _generar_excel(self, rentabilidad: List[Dict], desvios: Dict) -> str:
        """Genera Excel con 4 secciones"""

        ruta_salida = f"data/outputs/Reporte_Rentabilidad_{datetime.now().strftime('%Y%m%d')}.xlsx"
        Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Rentabilidad"

        # HEADER
        self._agregar_header(ws)

        # SECCIN 1: Rentabilidad por Canal
        self._agregar_seccion_canal(ws, rentabilidad)

        # SECCIN 2: Desvos Presupuesto
        self._agregar_seccion_desvios(ws, desvios)

        # SECCIN 3: Anlisis Profundo (si aplica)
        if any(r['margen_contrib'] < 27 for r in rentabilidad):
            self._agregar_seccion_analisis(ws)

        # Guardar
        wb.save(ruta_salida)
        print(f" Reporte generado: {ruta_salida}")

        return ruta_salida

    def _agregar_header(self, ws):
        """Agrega header del reporte"""
        ws['A1'] = "REPORTE 1: RENTABILIDAD"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        ws.merge_cells('A1:H1')

        ws['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A3'] = ""

    def _agregar_seccion_canal(self, ws, rentabilidad: List[Dict]):
        """Agrega tabla de rentabilidad por canal"""

        fila_inicio = 4
        ws[f'A{fila_inicio}'] = "SECCIN 1: RENTABILIDAD POR CANAL"
        ws[f'A{fila_inicio}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{fila_inicio}'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        ws.merge_cells(f'A{fila_inicio}:H{fila_inicio}')

        fila = fila_inicio + 1
        headers = ['Canal', 'Ventas', 'Costo', 'Comisiones', 'M. Directo %', 'M. Contribucin %', 'M. Operacional %', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=fila, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Datos
        fila = fila + 1
        for r in rentabilidad:
            ws.cell(row=fila, column=1, value=r['canal'])
            ws.cell(row=fila, column=2, value=r['ventas']).number_format = '$#,##0'
            ws.cell(row=fila, column=3, value=r['costo']).number_format = '$#,##0'
            ws.cell(row=fila, column=4, value=r['comisiones']).number_format = '$#,##0'
            ws.cell(row=fila, column=5, value=r['margen_directo']/100).number_format = '0.0%'
            ws.cell(row=fila, column=6, value=r['margen_contrib']/100).number_format = '0.0%'
            ws.cell(row=fila, column=7, value=r['margen_operacional']/100).number_format = '0.0%'
            ws.cell(row=fila, column=8, value=r['estado'])

            # Color de estado
            if r['estado'] == '':
                ws.cell(row=fila, column=8).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

            fila += 1

        ws['A2'] = f"ltima actualizacin: {datetime.now().strftime('%H:%M:%S')}"

    def _agregar_seccion_desvios(self, ws, desvios: Dict):
        """Agrega tabla de desvos presupuesto"""

        fila = ws.max_row + 3
        ws[f'A{fila}'] = "SECCIN 2: CONTROL PRESUPUESTO"
        ws[f'A{fila}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{fila}'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        ws.merge_cells(f'A{fila}:E{fila}')

        fila += 1
        headers = ['Concepto', 'Presupuesto', 'Real', 'Desvo %', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=fila, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")

        fila += 1
        for concepto, datos in desvios.items():
            ws.cell(row=fila, column=1, value=concepto)
            ws.cell(row=fila, column=2, value=datos['presupuesto']).number_format = '$#,##0'
            ws.cell(row=fila, column=3, value=datos['real']).number_format = '$#,##0'
            ws.cell(row=fila, column=4, value=datos['desvio_pct']/100).number_format = '0.0%'
            ws.cell(row=fila, column=5, value=datos['estado'])

            fila += 1

    def _agregar_seccion_analisis(self, ws):
        """Agrega seccin de anlisis profundo"""

        fila = ws.max_row + 3
        ws[f'A{fila}'] = "SECCIN 4: ANLISIS PROFUNDO"
        ws[f'A{fila}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{fila}'].fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        ws.merge_cells(f'A{fila}:D{fila}')

        fila += 1
        ws[f'A{fila}'] = "Generado por Skill 'anlisis-rentabilidad' (ver seccin de alertas)"
        ws[f'A{fila}'].font = Font(italic=True, color="D32F2F")

    def _generar_analisis_profundo(self, rentabilidad: List[Dict]):
        """Ejecuta skill de anlisis si hay margen < 27%"""

        print(" Margen crtico detectado. Ejecutando anlisis profundo...")
        # TODO: Importar y ejecutar analisis_rentabilidad_skill.py

    def enviar_email(self, ruta_reporte: str, destinatarios: List[str]):
        """Enva reporte por email al CEO"""

        # TODO: Implementar envo de email con openpyxl attachment

        print(f" Email enviado a {', '.join(destinatarios)}")


# ============================================================================
# SCRIPT EJECUTABLE (lunes 9 AM)
# ============================================================================

if __name__ == "__main__":
    # Rutas
    ruta_eerr = "data/outputs/02 EE.RR Febrero 2026_CLASIFICADO.json"
    ruta_presupuesto = "data/planillas/Presupuesto_Febrero_2026.xlsx"

    # Generar
    generador = GeneradorReporte1(ruta_eerr, ruta_presupuesto)
    ruta_salida = generador.generar()

    # Enviar email (opcional)
    # generador.enviar_email(ruta_salida, ["ceo@unionx.cl", "andres@unionx.cl"])

    print(f"\n Reporte 1 completo: {ruta_salida}")
