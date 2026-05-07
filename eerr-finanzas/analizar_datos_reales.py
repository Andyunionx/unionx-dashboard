"""
ANALIZADOR DE DATOS REALES
Inspecciona los archivos Excel existentes y muestra estructura/contenido
Ayuda a entender qu datos hay disponibles

Ejecutar:
  python analizar_datos_reales.py
"""

import openpyxl
from pathlib import Path
import json


def analizar_eerr_febrero():
    """Analiza estructura del EERR Febrero"""

    ruta = Path("../data/eerr/02 EE.RR Febrero 2026.xlsx")

    print("\n" + "="*70)
    print(" EERR FEBRERO 2026 - ESTRUCTURA")
    print("="*70)

    try:
        wb = openpyxl.load_workbook(ruta, read_only=True)
        print(f"\n[OK] Archivo: {ruta}")
        print(f"[Sheets] Disponibles ({len(wb.sheetnames)}):")

        for i, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            print(f"\n  {i}. {sheet_name}")
            print(f"     Tamao: {ws.dimensions}")

            # Mostrar primeras 10 filas
            print(f"     Datos (primeras 5 filas):")
            for fila_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
                print(f"       {fila_idx}: {row[:5]}")  # Primeras 5 columnas

            # Detectar qu tipo de datos hay
            if 'balance' in sheet_name.lower() or 'activo' in sheet_name.lower():
                print(f"     [TIPO] Balance / Activos")
            elif 'ingreso' in sheet_name.lower():
                print(f"     [TIPO] Ingresos")
            elif 'costo' in sheet_name.lower() or 'egreso' in sheet_name.lower():
                print(f"     [TIPO] Costos / Egresos")
            elif 'resumen' in sheet_name.lower():
                print(f"     [TIPO] Resumen / P&L")

    except Exception as e:
        print(f"[ERROR] {e}")


def analizar_analisis_contribucion():
    """Analiza estructura del Anlisis de Contribucin"""

    ruta = Path("../data/planillas/Anlisis Contribucin 2026 V02.02.xlsx")

    print("\n" + "="*70)
    print(" ANLISIS CONTRIBUCIN - ESTRUCTURA")
    print("="*70)

    try:
        wb = openpyxl.load_workbook(ruta, read_only=True)
        print(f"\n[OK] Archivo: {ruta}")
        print(f"[Sheets] Disponibles ({len(wb.sheetnames)}):")

        for i, sheet_name in enumerate(wb.sheetnames[:10], 1):  # Primeros 10
            ws = wb[sheet_name]
            max_row = min(ws.max_row, 5)

            print(f"\n  {i}. {sheet_name}")

            # Leer primeras filas
            datos = []
            for row in ws.iter_rows(max_row=max_row, values_only=True):
                datos.append(row[:5])  # Primeras 5 columnas

            if datos:
                print(f"     Primeras filas: {datos[0]}")

    except Exception as e:
        print(f"[ERROR] {e}")


def analizar_json_clasificado():
    """Analiza JSON EERR clasificado"""

    ruta = Path("../data/outputs/02 EE.RR Febrero 2026_CLASIFICADO.json")

    print("\n" + "="*70)
    print(" JSON EERR CLASIFICADO")
    print("="*70)

    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            datos = json.load(f)

        print(f"\n[OK] Archivo: {ruta}")
        print(f" Claves principales: {list(datos.keys())}")

        # Mostrar estructura
        for clave in list(datos.keys())[:3]:
            valor = datos[clave]
            if isinstance(valor, list):
                print(f"\n  {clave}: {len(valor)} items")
                if valor:
                    print(f"    Ejemplo: {valor[0]}")
            elif isinstance(valor, dict):
                print(f"\n  {clave}:")
                for k, v in list(valor.items())[:3]:
                    print(f"     {k}: {v}")

    except Exception as e:
        print(f"[ERROR] {e}")


def extraer_datos_utiles():
    """Extrae datos tiles de archivos reales"""

    print("\n" + "="*70)
    print(" EXTRAYENDO DATOS TILES")
    print("="*70)

    try:
        # Leer EERR clasificado
        with open("../data/outputs/02 EE.RR Febrero 2026_CLASIFICADO.json", 'r', encoding='utf-8') as f:
            eerr = json.load(f)

        # Estimar ingresos y costos
        ingresos_total = 0
        costos_total = 0

        if 'transacciones' in eerr:
            for t in eerr['transacciones']:
                monto = float(t.get('monto', 0))
                tipo = t.get('tipo', '').lower()

                if 'venta' in tipo or 'ingreso' in tipo:
                    ingresos_total += monto
                elif 'costo' in tipo or 'egreso' in tipo:
                    costos_total += abs(monto)

        print(f"\n[DATOS] Extrados del EERR clasificado:")
        print(f"  Ingresos totales: ${ingresos_total:,.0f}")
        print(f"  Costos totales: ${costos_total:,.0f}")
        print(f"  Margen bruto: ${ingresos_total - costos_total:,.0f} ({(1 - costos_total/ingresos_total if ingresos_total else 0)*100:.1f}%)")

    except Exception as e:
        print(f"[AVISO] No se pudo extraer: {e}")


# ============================================================================
# EJECUTAR
# ============================================================================

if __name__ == "__main__":
    print("\n" + "="*70)
    print("ANALIZADOR DE DATOS REALES - UNION X")
    print("="*70)

    analizar_eerr_febrero()
    analizar_analisis_contribucion()
    analizar_json_clasificado()
    extraer_datos_utiles()

    print("\n" + "="*70)
    print("[OK] ANALISIS COMPLETADO")
    print("="*70)
    print("\nPrximo paso: Revisar la salida y preparar archivos faltantes")
    print("Ver: PREPARAR_DATOS_ENTRADA.md para instrucciones especficas")
