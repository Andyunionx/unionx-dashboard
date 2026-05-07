"""
ORQUESTADOR DE REPORTES AUTOMÁTICOS
Coordina la generación de los 3 reportes semanales + alertas
Se ejecuta cada LUNES 9 AM automáticamente

Flujo:
1. Extrae datos de Odoo, Excel, y archivos JSON
2. Genera Reporte 1: Rentabilidad
3. Genera Reporte 2: KPIs Operacionales
4. Genera Reporte 3: Planificación Financiera
5. Evalúa 10 alertas en tiempo real
6. Envía emails a CEO y equipo
7. Genera resumen ejecutivo HTML
"""

import json
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import logging

# Configurar logging
log_path = Path("../data/outputs/reporte_semanal.log")
log_path.parent.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(str(log_path)),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class OrquestadorReportes:
    """Coordina la generación automática de los 3 reportes"""

    def __init__(self):
        self.fecha_ejecucion = datetime.now()
        self.resultados = {
            'reporte_1': None,
            'reporte_2': None,
            'reporte_3': None,
            'alertas': [],
        }
        self.errores = []

    def ejecutar(self) -> bool:
        """
        Ejecuta la cadena completa de reportes

        Returns:
            True si todo OK, False si hay errores críticos
        """

        logger.info("=" * 70)
        logger.info("INICIANDO GENERACIÓN DE REPORTES SEMANALES")
        logger.info(f"Fecha/Hora: {self.fecha_ejecucion.strftime('%d/%m/%Y %H:%M:%S')}")
        logger.info("=" * 70)

        try:
            # 1. Validar rutas de entrada
            self._validar_archivos_entrada()

            # 2. Generar Reporte 1
            logger.info("\n[1/5] Generando Reporte 1: Rentabilidad...")
            self.resultados['reporte_1'] = self._generar_reporte_1()

            # 3. Generar Reporte 2
            logger.info("\n[2/5] Generando Reporte 2: KPIs Operacionales...")
            self.resultados['reporte_2'] = self._generar_reporte_2()

            # 4. Generar Reporte 3
            logger.info("\n[3/5] Generando Reporte 3: Planificación Financiera...")
            self.resultados['reporte_3'] = self._generar_reporte_3()

            # 5. Evaluar alertas
            logger.info("\n[4/5] Evaluando alertas en tiempo real...")
            self.resultados['alertas'] = self._evaluar_alertas()

            # 6. Generar resumen ejecutivo
            logger.info("\n[5/5] Generando resumen ejecutivo...")
            resumen_path = self._generar_resumen_ejecutivo()

            # 7. Enviar emails
            logger.info("\nEnviando emails...")
            self._enviar_emails()

            logger.info("=" * 70)
            logger.info("✅ REPORTES COMPLETADOS EXITOSAMENTE")
            logger.info("=" * 70)

            return True

        except Exception as e:
            logger.error(f"❌ ERROR: {str(e)}")
            self.errores.append(str(e))
            return False

    def _validar_archivos_entrada(self):
        """Valida que existan los archivos de entrada necesarios"""

        # Source of truth via shared_paths
        try:
            import sys as _sys
            project_root = Path(__file__).resolve().parent.parent
            if str(project_root) not in _sys.path:
                _sys.path.insert(0, str(project_root))
            import shared_paths as _sp
            archivos_requeridos = [
                str(_sp.EERR_DIR / '02 EE.RR Febrero 2026.xlsx'),  # legacy hardcode mes
                str(_sp.CONTRIBUCION),
            ]
        except Exception:
            archivos_requeridos = [
                'data/eerr/02 EE.RR Febrero 2026.xlsx',
                'data/planillas/Analisis_Contribucion_2026_V06.xlsx',
            ]

        faltantes = []
        for archivo in archivos_requeridos:
            if not Path(archivo).exists():
                faltantes.append(archivo)

        if faltantes:
            logger.warning(f"⚠️ Archivos faltantes: {', '.join(faltantes)}")
            logger.info("Se continuará con datos disponibles...")

    def _generar_reporte_1(self) -> str:
        """Genera Reporte 1: Rentabilidad"""

        try:
            from generar_reporte1_rentabilidad import GeneradorReporte1

            ruta_eerr = "data/outputs/02 EE.RR Febrero 2026_CLASIFICADO.json"
            ruta_presupuesto = "data/planillas/Presupuesto_Febrero_2026.xlsx"

            generador = GeneradorReporte1(ruta_eerr, ruta_presupuesto)
            ruta_salida = generador.generar()

            logger.info(f"   ✓ Reporte 1 generado: {ruta_salida}")
            return ruta_salida

        except Exception as e:
            logger.error(f"   ❌ Error en Reporte 1: {str(e)}")
            self.errores.append(f"Reporte 1: {str(e)}")
            return None

    def _generar_reporte_2(self) -> str:
        """Genera Reporte 2: KPIs Operacionales"""

        try:
            from generar_reporte2_kpis import GeneradorReporte2

            ruta_odoo = "data/outputs/odoo_export_20260401.json"
            ruta_comex = "data/outputs/comex_maestra_cc.json"

            generador = GeneradorReporte2(ruta_odoo, ruta_comex)
            ruta_salida = generador.generar()

            logger.info(f"   ✓ Reporte 2 generado: {ruta_salida}")
            return ruta_salida

        except Exception as e:
            logger.error(f"   ❌ Error en Reporte 2: {str(e)}")
            self.errores.append(f"Reporte 2: {str(e)}")
            return None

    def _generar_reporte_3(self) -> str:
        """Genera Reporte 3: Planificación Financiera"""

        try:
            from generar_reporte3_planificacion import GeneradorReporte3

            ruta_planificacion = "data/planillas/Planificación Financiera.xlsx"
            ruta_eerr = "data/outputs/02 EE.RR Febrero 2026_CLASIFICADO.json"
            ruta_sueldos = "data/planillas/Sueldos_Febrero_2026.xlsx"
            ruta_deuda = "data/planillas/Balance_Febrero_2026.xlsx"

            generador = GeneradorReporte3(
                ruta_planificacion,
                ruta_eerr,
                ruta_sueldos,
                ruta_deuda
            )
            ruta_salida = generador.generar()

            logger.info(f"   ✓ Reporte 3 generado: {ruta_salida}")
            return ruta_salida

        except Exception as e:
            logger.error(f"   ❌ Error en Reporte 3: {str(e)}")
            self.errores.append(f"Reporte 3: {str(e)}")
            return None

    def _evaluar_alertas(self) -> List[Dict]:
        """Evalúa los 10 criterios de alerta"""

        try:
            from sistema_alertas_tiempo_real import SistemaAlertas

            sistema = SistemaAlertas()

            # Recopilar datos de los reportes
            # (En producción, extraería datos reales)
            datos_ejemplo = self._preparar_datos_alertas()

            alertas = sistema.evaluar(
                datos_ejemplo['rentabilidad'],
                datos_ejemplo['operaciones'],
                datos_ejemplo['comex'],
                datos_ejemplo['flujo']
            )

            # Enviar alertas
            sistema.enviar_alertas()

            # Exportar JSON
            sistema.exportar_json('data/outputs/alertas_tiempo_real.json')

            logger.info(f"   ✓ {len(alertas)} alertas evaluadas")
            return alertas

        except Exception as e:
            logger.error(f"   ❌ Error en alertas: {str(e)}")
            self.errores.append(f"Alertas: {str(e)}")
            return []

    def _preparar_datos_alertas(self) -> Dict:
        """Prepara datos REALES para evaluacion de alertas A1-A10.

        Cada bloque (rentabilidad, operaciones, comex, flujo) intenta
        cargar de su fuente. Si falla, devuelve dict vacio (la alerta
        no se dispara).
        """
        return {
            'rentabilidad': self._datos_rentabilidad(),
            'operaciones': self._datos_operaciones(),
            'comex': self._datos_comex(),
            'flujo': self._datos_flujo(),
        }

    def _datos_rentabilidad(self) -> Dict:
        """A1, A3, A8, A9: lee canales/desvios/clientes de Analisis Contribucion."""
        try:
            from openpyxl import load_workbook
            project_root = Path(__file__).resolve().parent.parent
            try:
                import sys as _sys
                if str(project_root) not in _sys.path:
                    _sys.path.insert(0, str(project_root))
                import shared_paths as _sp
                excel = _sp.CONTRIBUCION
            except Exception:
                excel = project_root / "data" / "planillas" / "Analisis_Contribucion_2026_V06.xlsx"
            if not excel.exists():
                return {}
            wb = load_workbook(str(excel), data_only=True, read_only=True)
            # Estructura esperada minima — ajustar cuando se conozca el layout exacto
            canales = []
            sheet_names = [s for s in wb.sheetnames if 'canal' in s.lower() or 'contribucion' in s.lower()]
            for sn in sheet_names[:3]:
                ws = wb[sn]
                # Heuristic: buscar filas con 'margen' en col A
                for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
                    if not row or row[0] is None:
                        continue
                    label = str(row[0]).lower()
                    if 'margen' in label and 'contrib' in label:
                        for v in row[1:]:
                            if isinstance(v, (int, float)) and 0 < v < 1:
                                canales.append({'nombre': sn, 'margen_contrib': float(v)})
                                break
                        break
            wb.close()
            return {'canales': canales, 'desvios_presupuesto': {}, 'clientes': []}
        except Exception as e:
            logger.warning(f"   ⚠️ datos_rentabilidad fallo: {e}")
            return {}

    def _datos_operaciones(self) -> Dict:
        """A2, A6, A7: query Odoo + heuristica de stock/ocupacion."""
        try:
            import sys as _sys
            project_root = Path(__file__).resolve().parent.parent
            backend_path = project_root / "finanzas-unionx" / "backend"
            if backend_path.exists() and str(backend_path) not in _sys.path:
                _sys.path.insert(0, str(backend_path))
            from app.core.odoo_client import OdooClient
            from app.config import Config
            odoo = OdooClient(Config.ODOO_URL, Config.ODOO_DB, Config.ODOO_USER, Config.ODOO_PASSWORD)

            # Stock bajo: orderpoints donde qty_available < product_min_qty
            ops = odoo.search_read('stock.warehouse.orderpoint', [],
                                   ['product_id', 'product_min_qty', 'qty_available_now'], limit=500) if hasattr(odoo, 'search_read') else []
            items_bajo = []
            for op in ops:
                actual = op.get('qty_available_now') or 0
                minimo = op.get('product_min_qty') or 0
                if minimo and actual < minimo:
                    sku = op['product_id'][1] if op.get('product_id') else 'SKU'
                    items_bajo.append({'sku': sku, 'stock_actual': actual, 'minimo': minimo})

            return {
                'items_bajo_minimo': items_bajo,
                # Ocupacion y fulfillment requieren datos adicionales aun no disponibles
                'ocupacion_pct': 0.0,
                'tasa_ontime_pct': 100.0,
                'rotacion_promedio': 0.0,
                'rotacion_historica': 0.0,
            }
        except Exception as e:
            logger.warning(f"   ⚠️ datos_operaciones fallo: {e}")
            return {}

    def _datos_comex(self) -> Dict:
        """A4: lee carpetas OHNSO en agente-comex/data/inbox para detectar retrasos."""
        try:
            project_root = Path(__file__).resolve().parent.parent
            inbox = project_root / "agente-comex" / "data" / "inbox"
            if not inbox.exists():
                return {}
            # Heuristica: cualquier carpeta con shipping_plan_ohnso > 30 dias considera "retraso"
            from datetime import timedelta
            cutoff = datetime.now() - timedelta(days=30)
            importaciones = []
            for folder in inbox.iterdir():
                if not folder.is_dir() or 'ohnso' not in folder.name.lower():
                    continue
                try:
                    yyyy, mm, dd = int(folder.name[:4]), int(folder.name[4:6]), int(folder.name[6:8])
                    f_date = datetime(yyyy, mm, dd)
                except Exception:
                    continue
                if f_date < cutoff:
                    importaciones.append({
                        'id': folder.name,
                        'dias_retraso': (datetime.now() - f_date).days,
                        'lead_time_promedio': 30,
                        'eta_original': 'N/A',
                        'eta_actual': 'pendiente',
                    })
            return {'importaciones_activas': importaciones}
        except Exception as e:
            logger.warning(f"   ⚠️ datos_comex fallo: {e}")
            return {}

    def _datos_flujo(self) -> Dict:
        """A10: flujo proyectado (placeholder, requiere modelo de cashflow)."""
        # TODO: integrar con planilla de Planificacion Financiera cuando exista
        return {'saldos_proyectados': {}}

    def _generar_resumen_ejecutivo(self) -> str:
        """Genera resumen HTML ejecutivo"""

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
                .header {{ background: #1a237e; color: white; padding: 20px; border-radius: 5px; }}
                .status {{ display: flex; gap: 20px; margin: 20px 0; }}
                .report {{ background: white; padding: 15px; margin: 10px 0; border-left: 4px solid #1976d2; }}
                .alert {{ background: #fff3e0; padding: 10px; margin: 5px 0; border-left: 4px solid #f57c00; }}
                .ok {{ color: #2e7d32; }}
                .error {{ color: #d32f2f; }}
                table {{ width: 100%; border-collapse: collapse; background: white; }}
                th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }}
                th {{ background: #f5f5f5; font-weight: bold; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>📊 REPORTE SEMANAL AUTOMÁTICO</h1>
                <p>Generado: {self.fecha_ejecucion.strftime('%d de %B de %Y a las %H:%M')}</p>
            </div>

            <div class="status">
                <div class="report">
                    <h3>✅ Reporte 1: Rentabilidad</h3>
                    <p>Estado: {'✓ Completado' if self.resultados['reporte_1'] else '❌ Error'}</p>
                    <p class="{'ok' if self.resultados['reporte_1'] else 'error'}">
                        {self.resultados['reporte_1'] if self.resultados['reporte_1'] else 'No se pudo generar'}
                    </p>
                </div>

                <div class="report">
                    <h3>✅ Reporte 2: KPIs Operacionales</h3>
                    <p>Estado: {'✓ Completado' if self.resultados['reporte_2'] else '❌ Error'}</p>
                    <p class="{'ok' if self.resultados['reporte_2'] else 'error'}">
                        {self.resultados['reporte_2'] if self.resultados['reporte_2'] else 'No se pudo generar'}
                    </p>
                </div>

                <div class="report">
                    <h3>✅ Reporte 3: Planificación Financiera</h3>
                    <p>Estado: {'✓ Completado' if self.resultados['reporte_3'] else '❌ Error'}</p>
                    <p class="{'ok' if self.resultados['reporte_3'] else 'error'}">
                        {self.resultados['reporte_3'] if self.resultados['reporte_3'] else 'No se pudo generar'}
                    </p>
                </div>
            </div>

            <h2>📢 Alertas Detectadas ({len(self.resultados['alertas'])})</h2>
            {''.join(f'<div class="alert">{a}</div>' for a in self.resultados['alertas'][:5])}
            {f'<p>... y {len(self.resultados["alertas"]) - 5} más. Ver archivo alertas_tiempo_real.json</p>' if len(self.resultados['alertas']) > 5 else '<p>Sin alertas críticas</p>'}

            <h2>📋 Próximos Pasos</h2>
            <ol>
                <li>Descargar los 3 reportes desde data/outputs/</li>
                <li>Revisar alertas en tiempo real (archivo JSON)</li>
                <li>Tomar acciones según recomendaciones</li>
                <li>Confirmar ejecución automática para próximo lunes</li>
            </ol>

            <p style="color: #666; font-size: 12px; margin-top: 30px;">
                Reporte generado automáticamente por UNION X - IA.<br>
                Para consultas: andres@unionx.cl
            </p>
        </body>
        </html>
        """

        ruta_salida = f"data/outputs/Resumen_Semanal_{self.fecha_ejecucion.strftime('%Y%m%d')}.html"
        Path(ruta_salida).parent.mkdir(parents=True, exist_ok=True)

        with open(ruta_salida, 'w', encoding='utf-8') as f:
            f.write(html_content)

        logger.info(f"   ✓ Resumen ejecutivo generado: {ruta_salida}")
        return ruta_salida

    def _enviar_emails(self):
        """Envía emails a CEO y equipo (via Gmail API compartido).

        Adjunta los reportes generados (Excel) y el resumen HTML.
        Respeta GMAIL_DRY_RUN=1 para mandar a borradores.
        """
        try:
            # Importar wrapper compartido (esta en project root)
            import sys as _sys
            from pathlib import Path as _Path
            project_root = _Path(__file__).resolve().parent.parent
            if str(project_root) not in _sys.path:
                _sys.path.insert(0, str(project_root))
            from shared_email import enviar_reporte_ceo

            # Recopilar adjuntos disponibles
            adjuntos = []
            for k in ('reporte_1', 'reporte_2', 'reporte_3'):
                p = self.resultados.get(k)
                if p and Path(p).exists():
                    adjuntos.append(p)

            # Cuerpo HTML basico (resumen de resultados)
            cuerpo = f"""
            <h2>Reportes Ejecutivos UnionX</h2>
            <p>Fecha de ejecucion: {self.fecha_ejecucion.strftime('%d/%m/%Y %H:%M')}</p>
            <ul>
              <li>Reporte 1 (Rentabilidad): {'OK' if self.resultados['reporte_1'] else 'FALLO'}</li>
              <li>Reporte 2 (KPIs Operacionales): {'OK' if self.resultados['reporte_2'] else 'FALLO'}</li>
              <li>Reporte 3 (Planificacion Financiera): {'OK' if self.resultados['reporte_3'] else 'FALLO'}</li>
              <li>Alertas evaluadas: {len(self.resultados.get('alertas') or [])}</li>
            </ul>
            <p>Adjuntos: {len(adjuntos)} archivo(s)</p>
            """

            asunto = f"Reportes Ejecutivos UnionX - {self.fecha_ejecucion.strftime('%d/%m/%Y')}"
            msg_id = enviar_reporte_ceo(asunto, cuerpo, adjuntos)
            logger.info(f"   ✓ Email enviado/dratf: {msg_id}")
        except Exception as e:
            logger.error(f"   ❌ Error enviando email: {e}")
            self.errores.append(f"Email: {e}")


# ============================================================================
# CONFIGURACIÓN PARA TASK SCHEDULER (Windows)
# ============================================================================

SETUP_WINDOWS_SCHEDULER = """
REM Crear tarea programada en Windows Task Scheduler
REM Guardar este contenido en: setup_scheduler.bat

REM Para Lunes 9 AM:
schtasks /create /tn "UNION_X_Reportes_Lunes_9am" /tr "python C:\\path\\to\\orquestador_reportes.py" /sc weekly /d MON /st 09:00:00 /f

REM Para ejecutar inmediatamente:
REM schtasks /run /tn "UNION_X_Reportes_Lunes_9am"

REM Para ver tareas:
REM schtasks /query /tn "UNION_X_Reportes_Lunes_9am"

REM Para eliminar:
REM schtasks /delete /tn "UNION_X_Reportes_Lunes_9am" /f
"""


# ============================================================================
# SCRIPT EJECUTABLE
# ============================================================================

if __name__ == "__main__":
    # Crear orquestador
    orquestador = OrquestadorReportes()

    # Ejecutar
    exito = orquestador.ejecutar()

    # Salir con código apropiado
    exit(0 if exito else 1)
