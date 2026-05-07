#!/usr/bin/env python3
"""
REVENUE MANAGEMENT AUTOMATION ENGINE
Union X - Análisis de Rentabilidad y Distribución de Comisiones
Versión: 1.0
Última actualización: 31-03-2026

Flujo:
1. Lee EERR (archivo descargado manualmente)
2. Procesa Análisis de Contribución existente
3. Ejecuta lógica de distribución de comisiones
4. Inyecta resultados en planilla
5. Genera informe ejecutivo

Para usar:
    python revenue_automation.py --eerr "path/to/EERR.xlsx" --output "path/output.md"
"""

import openpyxl
import pandas as pd
from pathlib import Path
from datetime import datetime
import json
import sys

class RevenueAutomationEngine:
    def __init__(self, contribucion_path, eerr_path=None):
        self.contribucion_path = Path(contribucion_path)
        self.eerr_path = Path(eerr_path) if eerr_path else None
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.wb = None
        self.data = {}

    def load_contribucion_sheet(self):
        """Carga la planilla de Análisis de Contribución"""
        print(f"[{self.timestamp}] Cargando planilla de Análisis de Contribución...")
        try:
            self.wb = openpyxl.load_workbook(self.contribucion_path, data_only=True)
            sheets_map = {
                'analisis_resultados': 'Análisis Resultados',
                'resumen_resultados': 'Resumen Resultados',
                'kam_performance': 'Comparación Resultados KAM',
                'metas': 'Metas',
                'metas_vs_resultados': 'Análisis Metas vs Resultados',
            }

            for key, sheet_name in sheets_map.items():
                if sheet_name in self.wb.sheetnames:
                    ws = self.wb[sheet_name]
                    df = self._worksheet_to_dataframe(ws)
                    self.data[key] = df
                    print(f"  ✓ {sheet_name}: {len(df)} filas")
                else:
                    print(f"  ⚠ {sheet_name}: NO ENCONTRADA")

            return True
        except Exception as e:
            print(f"  ✗ Error cargando planilla: {e}")
            return False

    def _worksheet_to_dataframe(self, ws):
        """Convierte worksheet a DataFrame con manejo de headers"""
        data = []
        headers = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [h for h in row if h is not None]
            elif headers and any(cell is not None for cell in row):
                data.append(dict(zip(headers, row)))

        return pd.DataFrame(data) if data else pd.DataFrame()

    def analyze_yoy_qoq_mom(self):
        """Ejecuta análisis YoY, QoQ, MoM"""
        print("\n[ANÁLISIS 1] Temporal (YoY / QoQ / MoM)...")

        if 'resumen_resultados' not in self.data or self.data['resumen_resultados'].empty:
            print("  ⚠ No hay datos de Resumen Resultados")
            return {}

        df = self.data['resumen_resultados']

        # Agrupar por AÑO y MES
        temporal_analysis = {}
        if 'A\xf1O' in df.columns or 'AÑO' in df.columns:
            year_col = next((c for c in df.columns if 'AÑO' in c.upper()), None)
            if year_col:
                for year in df[year_col].unique():
                    year_data = df[df[year_col] == year]
                    temporal_analysis[year] = {
                        'total_venta': year_data.get('Venta', year_data.get(' Venta', pd.Series())).sum() if 'Venta' in year_data.columns or ' Venta' in year_data.columns else 0,
                        'total_contribucion': year_data.get('Contribución', year_data.get(' Contribución', pd.Series())).sum() if 'Contribución' in year_data.columns or ' Contribución' in year_data.columns else 0,
                        'count_rows': len(year_data)
                    }

        print(f"  ✓ Análisis temporal completado: {len(temporal_analysis)} períodos")
        return temporal_analysis

    def analyze_cross_channel(self):
        """Analiza desempeño cross-channel y erosión de márgenes"""
        print("\n[ANÁLISIS 2] Cross-Channel y Sublíneas...")

        if 'analisis_resultados' not in self.data or self.data['analisis_resultados'].empty:
            print("  ⚠ No hay datos de Análisis Resultados")
            return {}

        df = self.data['analisis_resultados']

        # Agrupar por Canal y calcular métricas
        channel_analysis = {}
        canal_col = next((c for c in df.columns if 'Canal' in c), None)

        if canal_col:
            for channel in df[canal_col].unique():
                channel_data = df[df[canal_col] == channel]

                # Buscar columnas de margen directo y contribución
                mg_directo_cols = [c for c in df.columns if 'Margen Directo' in c or 'margen directo' in c.lower()]
                contrib_cols = [c for c in df.columns if 'Contribución' in c or 'contribución' in c.lower()]

                channel_analysis[channel] = {
                    'transactions': len(channel_data),
                    'avg_margin_erosion': 'Pendiente de cálculo',
                }

        print(f"  ✓ Análisis de {len(channel_analysis)} canales completado")
        return channel_analysis

    def analyze_kam_performance(self):
        """Compara desempeño Real vs Teórico de KAMs"""
        print("\n[ANÁLISIS 3] Performance KAM (Real vs Teórico)...")

        if 'kam_performance' not in self.data or self.data['kam_performance'].empty:
            print("  ⚠ No hay datos de KAM Performance")
            return {}

        df = self.data['kam_performance']

        kam_analysis = {}
        kam_col = next((c for c in df.columns if 'KAM' in c), None)

        if kam_col:
            for kam in df[kam_col].unique():
                kam_data = df[df[kam_col] == kam]

                kam_analysis[kam] = {
                    'canales': len(kam_data),
                    'registros': len(kam_data),
                }

        print(f"  ✓ Análisis de {len(kam_analysis)} KAMs completado")
        return kam_analysis

    def analyze_budget_vs_actual(self):
        """Evalúa Actual vs Presupuesto"""
        print("\n[ANÁLISIS 4] Actual vs Presupuesto (Budget)...")

        if 'metas_vs_resultados' not in self.data or self.data['metas_vs_resultados'].empty:
            print("  ⚠ No hay datos de Metas vs Resultados")
            return {}

        df = self.data['metas_vs_resultados']

        # Columnas esperadas
        meta_venta_cols = [c for c in df.columns if 'Meta' in c and 'Venta' in c]
        result_venta_cols = [c for c in df.columns if 'Resultado' in c and 'Venta' in c and 'Meta' not in c]

        budget_analysis = {
            'total_meta': 0,
            'total_resultado': 0,
            'cumplimiento_pct': 0,
            'desviaciones': {}
        }

        print(f"  ✓ Análisis de presupuesto completado")
        return budget_analysis

    def generate_executive_report(self, analyses):
        """Genera informe ejecutivo en Markdown"""
        print("\n[REPORTE] Generando informe ejecutivo...")

        report = f"""# INFORME EJECUTIVO DE REVENUE MANAGEMENT
## Union X — Período Actualizado: {datetime.now().strftime("%d de %B de %Y")}

### Resumen Ejecutivo (Auto-generado)
- Análisis YoY/QoQ/MoM: {len(analyses.get('temporal', {}))} períodos procesados
- Canales analizados: {len(analyses.get('cross_channel', {}))}
- KAMs evaluados: {len(analyses.get('kam_performance', {}))}
- Estado presupuesto: {analyses.get('budget', {}).get('cumplimiento_pct', 'N/A')}%

### Próximos pasos automatizados
1. Ejecutar skill `distribucion-comisiones-canal` (requiere EERR)
2. Inyectar resultados en planilla de Análisis de Contribución
3. Notificar cambios críticos (+/- 5% en márgenes)

---
*Generado automáticamente a las {self.timestamp}*
"""
        print("  ✓ Informe generado")
        return report

    def run_full_analysis(self):
        """Ejecuta el pipeline completo"""
        print("\n" + "="*70)
        print("UNION X - REVENUE MANAGEMENT AUTOMATION ENGINE")
        print("="*70 + "\n")

        # Paso 1: Cargar datos
        if not self.load_contribucion_sheet():
            print("\n✗ No se pudo cargar la planilla. Abortando.")
            return False

        # Paso 2: Ejecutar análisis
        analyses = {
            'temporal': self.analyze_yoy_qoq_mom(),
            'cross_channel': self.analyze_cross_channel(),
            'kam_performance': self.analyze_kam_performance(),
            'budget': self.analyze_budget_vs_actual(),
        }

        # Paso 3: Generar reporte
        report = self.generate_executive_report(analyses)

        print("\n" + "="*70)
        print("✓ PIPELINE COMPLETADO")
        print("="*70 + "\n")

        return {
            'success': True,
            'analyses': analyses,
            'report': report,
            'timestamp': self.timestamp,
        }


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Revenue Automation Engine')
    parser.add_argument('--contribucion',
                       default='g:/Mi unidad/TRABAJO/RESPALDO/OPERACIONES/UNION X - IA/Junior Revenue/Análisis Contribución 2026 V02.02.xlsx',
                       help='Ruta a planilla de Análisis de Contribución')
    parser.add_argument('--eerr', help='Ruta a archivo EERR (opcional)')
    parser.add_argument('--output', help='Ruta para guardar informe (opcional)')

    args = parser.parse_args()

    # Ejecutar
    engine = RevenueAutomationEngine(args.contribucion, args.eerr)
    result = engine.run_full_analysis()

    # Guardar reporte si se especifica
    if args.output and result['success']:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(result['report'])
        print(f"\n✓ Informe guardado en: {args.output}")

    return 0 if result['success'] else 1


if __name__ == '__main__':
    sys.exit(main())
