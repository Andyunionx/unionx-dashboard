"""Notifica a Felipe + equipo cuando hay SKUs del precosteo que NO existen en
maestra Odoo (default_code).

Se invoca:
  1. Automáticamente al final de cargar_po_comex_odoo.py si hay no_encontrados.
  2. Manualmente vía CLI para consolidar varios embarques en un solo mail.

Uso CLI:
    python notificar_skus_faltantes.py 0423 0429 0430        # Consolidado
    python notificar_skus_faltantes.py 0424 --dry-run        # Preview HTML
"""
import argparse
import base64
import os
import sys
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
PROJECT_ROOT = Path(__file__).parent

# Mismos destinatarios que procesar_embarque.py
DESTINATARIOS = 'felipe@unionx.cl, sguzman@grupoeter.cl, gerardo@unionx.cl, operaciones@unionx.cl'


def _cargar_env():
    env = PROJECT_ROOT / '.env'
    if env.exists():
        for line in env.read_text(encoding='utf-8').splitlines():
            if '=' in line and not line.startswith('#'):
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def _conectar_odoo(max_retries: int = 4):
    """Conecta a Odoo con backoff exponencial para resistir 503 transitorios."""
    import time
    import xmlrpc.client
    _cargar_env()
    url = os.environ.get('ODOO_URL', 'https://unionxb2b.odoo.com')
    db = os.environ.get('ODOO_DB', 'bmya-innovatek-sh-prd-6981800')
    user = os.environ.get('ODOO_USER') or 'andres@grupoeter.cl'
    pwd = os.environ.get('ODOO_API_KEY') or os.environ.get('ANDRES_ODOO_PASSWORD')
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
            uid = common.authenticate(db, user, pwd, {})
            models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)
            return uid, models, db, pwd
        except xmlrpc.client.ProtocolError as e:
            last_err = e
            if attempt < max_retries:
                wait = 5 * (2 ** (attempt - 1))
                print(f"  [retry] Odoo {e.errcode} attempt {attempt}/{max_retries}, esperando {wait}s...")
                time.sleep(wait)
    raise last_err


def _leer_skus_precosteo(emb: str) -> list[dict]:
    """Lee los productos del precosteo y devuelve la lista cruda (sin filtrar gastos)."""
    p = PROJECT_ROOT / 'agente-comex' / 'data' / 'output' / f'26TP{emb}' / f'Pre-costeo_x_CBM_26TP{emb}.xlsx'
    if not p.exists():
        return []
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    ws = wb['Productos']
    hdrs = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {h: i for i, h in enumerate(hdrs)}
    skus = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        sku = str(r[idx['SKU']] or '').strip()
        if not sku or sku.lower() in ('none', 'nan'):
            continue
        skus.append({
            'sku': sku,
            'model': r[idx.get('Model', 0)],
            'descripcion': str(r[idx.get('Descripcion', 2)] or '')[:120].replace('\n', ' '),
            'qty': r[idx['Qty']] or 0,
            'precio_usd': r[idx['Price']] or 0,
            'clp_unit': r[idx['Costo Internado Unit (CLP)']] or 0,
        })
    wb.close()
    return skus


def detectar_faltantes(embarques: list[str]) -> dict[str, list[dict]]:
    """Para cada embarque, devuelve la lista de SKUs sin match Odoo."""
    uid, models, db, pwd = _conectar_odoo()
    resultado = {}
    for emb in embarques:
        skus = _leer_skus_precosteo(emb)
        if not skus:
            continue
        codes = [s['sku'] for s in skus]
        found = models.execute_kw(db, uid, pwd, 'product.product', 'search_read',
            [[('default_code', 'in', codes)]], {'fields': ['default_code']})
        found_set = {p['default_code'] for p in found}
        faltantes = [s for s in skus if s['sku'] not in found_set]
        if faltantes:
            resultado[f'26TP{emb}'] = faltantes
    return resultado


def _generar_html(faltantes_por_emb: dict[str, list[dict]]) -> str:
    n_total = sum(len(v) for v in faltantes_por_emb.values())
    embs = sorted(faltantes_por_emb.keys())

    bloques = ''
    for emb in embs:
        filas = ''
        for s in faltantes_por_emb[emb]:
            filas += (
                f'<tr>'
                f'<td style="border:1px solid #ddd;padding:6px 10px"><b>{s["sku"]}</b></td>'
                f'<td style="border:1px solid #ddd;padding:6px 10px">{s["descripcion"]}</td>'
                f'<td style="border:1px solid #ddd;padding:6px 10px;text-align:right">{int(s["qty"])}</td>'
                f'<td style="border:1px solid #ddd;padding:6px 10px;text-align:right">${s["precio_usd"]:.2f}</td>'
                f'<td style="border:1px solid #ddd;padding:6px 10px;text-align:right">${s["clp_unit"]:,.0f}</td>'
                f'</tr>'
            )
        bloques += f"""
<h3 style="margin-top:25px;color:#1F4E78">{emb} — {len(faltantes_por_emb[emb])} SKU(s)</h3>
<table style="border-collapse:collapse;font-size:13px;width:100%">
<thead>
<tr style="background:#c0392b;color:#fff">
<th style="border:1px solid #ddd;padding:6px 10px">SKU</th>
<th style="border:1px solid #ddd;padding:6px 10px">Descripción</th>
<th style="border:1px solid #ddd;padding:6px 10px">Qty</th>
<th style="border:1px solid #ddd;padding:6px 10px">USD/u</th>
<th style="border:1px solid #ddd;padding:6px 10px">CLP/u internado</th>
</tr>
</thead>
<tbody>
{filas}
</tbody>
</table>"""

    return f"""<div style="font-family:Arial,sans-serif;font-size:14px;color:#111">
<h2 style="color:#c0392b">⚠ {n_total} SKU(s) sin maestra Odoo — Requieren creación</h2>
<p>Los siguientes SKUs vienen en el/los precosteo(s) pero <b>no existen como productos en Odoo</b>. NO fueron incluidos en la(s) RFQ automática(s). Es necesario crearlos en Odoo y cargarlos manualmente a la(s) PO en draft para que ingresen al stock al recibir el embarque.</p>
{bloques}
<p style="margin-top:20px;color:#666;font-size:12px">Mail automático del agente COMEX. Si algún SKU es texto sospechoso (ej. un número en columna SKU), revisar el PI original — puede ser un error de parseo.</p>
</div>"""


def enviar_mail(faltantes_por_emb: dict[str, list[dict]]) -> str:
    """Envía el mail consolidado. Devuelve message_id."""
    sys.path.insert(0, str(PROJECT_ROOT / 'agente-comex' / 'src'))
    from gmail_client import GmailClient

    embs = sorted(faltantes_por_emb.keys())
    n_total = sum(len(v) for v in faltantes_por_emb.values())
    asunto_embs = '/'.join(e.replace('26TP', '') for e in embs)
    subject = f'[26TP{asunto_embs}] AVISO — {n_total} SKUs no existen en maestra Odoo, requieren creación'
    html = _generar_html(faltantes_por_emb)

    msg = MIMEMultipart('alternative')
    msg['to'] = DESTINATARIOS
    msg['subject'] = subject
    msg.attach(MIMEText('Detalle en HTML adjunto.', 'plain', 'utf-8'))
    msg.attach(MIMEText(html, 'html', 'utf-8'))

    g = GmailClient()
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    sent = g.service.users().messages().send(userId='me', body={'raw': raw}).execute()
    return sent['id']


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('embarques', nargs='+', help='Códigos embarque (ej 0423 0429)')
    parser.add_argument('--dry-run', action='store_true', help='Preview HTML, no enviar')
    args = parser.parse_args()

    embs = [e.zfill(4) for e in args.embarques]
    print(f'[1] Detectando SKUs faltantes para: {embs}')
    faltantes = detectar_faltantes(embs)

    if not faltantes:
        print('  Sin SKUs faltantes. No mando mail.')
        return 0

    n = sum(len(v) for v in faltantes.values())
    print(f'  {n} SKUs faltantes en {len(faltantes)} embarques:')
    for emb, skus in faltantes.items():
        print(f'    {emb}: {len(skus)} → {[s["sku"] for s in skus]}')

    if args.dry_run:
        preview = PROJECT_ROOT / 'agente-comex' / 'data' / 'output' / 'preview_skus_faltantes.html'
        preview.parent.mkdir(parents=True, exist_ok=True)
        preview.write_text(_generar_html(faltantes), encoding='utf-8')
        print(f'\n[DRY-RUN] HTML guardado: {preview}')
        return 0

    print('\n[2] Enviando mail consolidado...')
    msg_id = enviar_mail(faltantes)
    print(f'  [OK] message_id={msg_id} | to={DESTINATARIOS}')
    return 0


if __name__ == '__main__':
    sys.exit(main() or 0)
