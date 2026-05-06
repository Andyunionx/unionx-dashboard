#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Test simple injection de datos en Excel"""

import os
import sys
from pathlib import Path
from datetime import datetime

# Fix encoding
if sys.platform == 'win32':
    os.environ['PYTHONIOENCODING'] = 'utf-8'
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import pandas as pd

print("\n" + "="*70)
print("TEST INJECTION - Inyectar datos en Excel")
print("="*70 + "\n")

# Paths
current_dir = Path(__file__).parent
excel_file = current_dir / "Análisis Contribución 2026 V02.02.xlsx"
drive_file = current_dir / "drive_download_20260331.xlsx"

print(f"[1] Verificando archivos...")
if not excel_file.exists():
    print(f"  ✗ Excel no encontrado: {excel_file}")
    sys.exit(1)
print(f"  ✓ Excel principal: {excel_file.name}")

if not drive_file.exists():
    print(f"  ✗ Drive file no encontrado: {drive_file}")
    sys.exit(1)
print(f"  ✓ Drive file: {drive_file.name}")

print(f"\n[2] Leyendo datos de Google Drive...")
try:
    df_drive = pd.read_excel(drive_file, sheet_name=0, engine='openpyxl')
    print(f"  ✓ Leído: {len(df_drive)} filas, {len(df_drive.columns)} columnas")
    print(f"  ✓ Primeras columnas: {list(df_drive.columns[:5])}")
except Exception as e:
    print(f"  ✗ Error: {e}")
    sys.exit(1)

print(f"\n[3] Inyectando en Excel...")
try:
    wb = openpyxl.load_workbook(excel_file)

    # Buscar hoja "Análisis Resultados"
    if "Análisis Resultados" in wb.sheetnames:
        ws = wb["Análisis Resultados"]
        print(f"  ✓ Hoja encontrada: 'Análisis Resultados'")
    else:
        ws = wb.active
        print(f"  ✓ Usando hoja activa: '{ws.title}'")

    # Escribir datos desde fila 2
    for idx, row in enumerate(df_drive.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=idx, column=col_idx, value=value)

    print(f"  ✓ Inyectadas {len(df_drive)} filas")

    # Guardar
    wb.save(excel_file)
    print(f"  ✓ Excel guardado: {excel_file.name}")

except Exception as e:
    print(f"  ✗ Error inyectando: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("\n" + "="*70)
print("✅ INYECCIÓN COMPLETADA")
print("="*70)
print(f"""
Resultado:
  ✓ Datos de Google Drive: {len(df_drive)} filas
  ✓ Inyectadas en: Análisis Contribución 2026 V02.02.xlsx
  ✓ Archivo guardado exitosamente

Abre el Excel para ver los datos nuevos.
""")
