"""Fix final Detail_NB: reasignar Embarque (Steven) y fechas usando Container (col 1) como fuente."""
import openpyxl, sys
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
sys.stdout.reconfigure(encoding='utf-8')

P = "data/planillas/Shipping Plan September V.02_actualizado.xlsx"

VERDE  = PatternFill('solid', fgColor='FF92D050')
AZUL_S = PatternFill('solid', fgColor='FF9DC3E6')
ROJO   = PatternFill('solid', fgColor='FFF08080')
GRIS   = PatternFill('solid', fgColor='FFD9D9D9')
AZUL_H = PatternFill('solid', fgColor='FF1F3864')
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
NORMAL = Font(size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TRANSITO = {'SZ': 52, 'NB': 45}
CRD_ETD = 7; PTO_BOD = 7

# Fechas por embarque conocido
EMB_CRD = {
    'SZ-04,40HQ': datetime(2026, 6, 29),
    'SZ-03-08,40HQ': datetime(2026, 7, 5),
    'SZ-01(623)': datetime(2026, 7, 15),
    'SZ-02(623)': datetime(2026, 8, 10),
    'NB-01': datetime(2026, 7, 20),
    'IMP OP 350-26 40HQ': datetime(2026, 6, 29),
    'NB-02': datetime(2026, 8, 15),
    'SZ-05,40HQ': datetime(2026, 8, 15),
}

def fechas(emb):
    crd = EMB_CRD.get(emb, datetime(2026, 8, 15))
    transit = TRANSITO['NB'] if ('NB' in emb.upper() or 'IMP OP' in emb.upper()) else TRANSITO['SZ']
    etd = crd + timedelta(days=CRD_ETD)
    etap = etd + timedelta(days=transit)
    etab = etap + timedelta(days=PTO_BOD)
    return crd, etd, etap, etab

wb = openpyxl.load_workbook(P)
ws = wb['Detail_NB']

print("=== Fix Detail_NB: reasignar por Container ===")
for r in range(2, ws.max_row+1):
    sku = ws.cell(row=r, column=6).value
    if not sku: continue
    # Container está en col 1 (correcto)
    container = ws.cell(row=r, column=1).value
    if not container: continue
    container_str = str(container).strip()

    # Detectar si es embarque válido
    if container_str in EMB_CRD or any(container_str.startswith(x) for x in ['SZ-','NB-','IMP OP']):
        crd, etd, etap, etab = fechas(container_str)
        # Embarque (Steven) = col 14
        c = ws.cell(row=r, column=14, value=container_str)
        c.alignment = CENTER; c.border = BORDER; c.font = NORMAL
        # CRD = col 15
        c = ws.cell(row=r, column=15, value=crd)
        c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL
        # ETD = col 16
        c = ws.cell(row=r, column=16, value=etd)
        c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL
        # ETA Puerto = col 17
        c = ws.cell(row=r, column=17, value=etap)
        c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL
        # ETA Bodega = col 18
        c = ws.cell(row=r, column=18, value=etab)
        c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL

print("  Detail_NB cols corregidas")

# También verificar Detail_SZ por si acaso
ws_sz = wb['Detail_SZ']
print("\n=== Verificar Detail_SZ ===")
count_fix = 0
for r in range(2, ws_sz.max_row+1):
    sku = ws_sz.cell(row=r, column=6).value
    if not sku: continue
    container = ws_sz.cell(row=r, column=1).value
    emb_col = ws_sz.cell(row=r, column=14).value
    if isinstance(emb_col, datetime) and container:
        # Bug: reasignar
        container_str = str(container).strip()
        crd, etd, etap, etab = fechas(container_str)
        ws_sz.cell(row=r, column=14, value=container_str).alignment = CENTER
        ws_sz.cell(row=r, column=15, value=crd).number_format='DD-MM-YYYY'
        ws_sz.cell(row=r, column=16, value=etd).number_format='DD-MM-YYYY'
        ws_sz.cell(row=r, column=17, value=etap).number_format='DD-MM-YYYY'
        ws_sz.cell(row=r, column=18, value=etab).number_format='DD-MM-YYYY'
        count_fix += 1
print(f"  Detail_SZ filas corregidas: {count_fix}")

wb.save(P)
print("\n[OK] Guardado")

# Verificar
import pandas as pd
for sh in ['Detail_SZ', 'Detail_NB', 'Missing']:
    df = pd.read_excel(P, sheet_name=sh)
    df.columns = [str(c).strip() for c in df.columns]
    df_v = df[df['Description'].notna() & (df['Description'].astype(str).str.strip()!='')] if 'Description' in df.columns else df.dropna(how='all')
    print(f"\n{sh}: {len(df_v)} filas")
    if 'Embarque (Steven)' in df_v.columns:
        e = df_v.groupby('Embarque (Steven)').size().reset_index(name='n')
        for _, r in e.iterrows():
            print(f"    {r['Embarque (Steven)']}: {r['n']}")
