"""Precosteo 26TP0609 DHL HÍBRIDO:
  - Gastos: los de Erich (flete USD 709 + comisión Steven 3% + seguro + derechos + honorario)
  - Prorrateo: nuestro (flete por PESO, resto por FOB o equitativo)
"""
import sys
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

OUT = Path("data/comex/embarques/0609/Pre-costeo_26TP0609_HIBRIDO_v2.xlsx")
DOLAR = 894.79
PESO_TOTAL = 44.0  # 3 cajas, 44 kg
N_ITEMS = 14

# === Gastos de ERICH (mejor info disponible)
GASTO_FLETE_DHL      = 709.00     # USD - "Delivery cost by DHL" (puerta a puerta)
COMISION_STEVEN_PCT  = 0.03       # 3% sobre FOB
GASTO_SEGURO         = 18.07      # USD
GASTO_DERECHOS       = 76.34      # USD = CLP 68.308 / 894,79
GASTO_HONORARIO_DHL  = 96.00      # USD = CLP 85.900 / 894,79

# === Items con peso DHL FACTURABLE (max real, volumétrico)
# El Trash Bin sube por volumen (32×28×33 cm = 5,9 kg vol). Suma debe dar 44 kg DHL.
items = [
    # (nombre, model, peso_kg, FOB_USD, qty)
    ('Trash Bin 12L Negro',                'TB12L #1',     5.5,  8.90, 1),
    ('Trash Bin 12L (segunda unidad)',     'TB12L #2',     5.5,  8.90, 1),
    ('Cutlery Tray',                        'H-0259',       0.5,  5.10, 1),
    ('Cafetera 4 tazas Negro',             'TP009B',       2.0,  4.60, 1),
    ('Cafetera 4 tazas Verde',             'TP009G',       2.0,  4.00, 1),
    ('Cafetera 12 tazas Negro',            'TP010B',       4.5,  6.30, 1),
    ('Cafetera 12 tazas Verde',            'TP010G',       4.5,  6.00, 1),
    ('Smartwatch GT1 Apex Negro',          'GT1 Negro',    0.3, 24.70, 1),
    ('Smartwatch GT1 Apex Rosado',         'GT1 Rosado',   0.3, 24.70, 1),
    ('Smartwatch GT1 Apex Naranjo',        'GT1 Naranjo',  0.3, 24.70, 1),
    ('Cobertor carpa cubre SUV XRoad',     'TP166',        3.2,  4.60, 1),
    ('Barra para auto universal 122cm',    'TP407-120',    7.5, 20.35, 1),
    ('Barra para auto universal 129cm',    'TP407-127',    7.9, 21.00, 1),
    ('Gamer controller samples',           'gamer ctrl',   0.5,  4.50, 1),
]

suma_peso = sum(i[2] for i in items)
suma_fob  = sum(i[3]*i[4] for i in items)
print(f"Peso total: {suma_peso} kg (objetivo 44)")
print(f"FOB total:  ${suma_fob:.2f}")
print()

# Recalcular comisión Steven sobre FOB total
comision_steven_usd = suma_fob * COMISION_STEVEN_PCT
total_gastos_usd = (GASTO_FLETE_DHL + comision_steven_usd + GASTO_SEGURO +
                    GASTO_DERECHOS + GASTO_HONORARIO_DHL)

print(f"=== GASTOS TOTALES (USD) ===")
print(f"  Flete DHL (Erich):          ${GASTO_FLETE_DHL:>8.2f}")
print(f"  Comisión Steven 3% (Erich): ${comision_steven_usd:>8.2f}")
print(f"  Seguro:                     ${GASTO_SEGURO:>8.2f}")
print(f"  Derechos aduaneros:         ${GASTO_DERECHOS:>8.2f}")
print(f"  Honorario DHL:              ${GASTO_HONORARIO_DHL:>8.2f}")
print(f"  TOTAL:                      ${total_gastos_usd:>8.2f} = ${total_gastos_usd*DOLAR:,.0f} CLP")
print()

# Prorrateo HÍBRIDO:
#   - Flete DHL → por PESO (DHL cobra por peso/volumen)
#   - Comisión Steven 3% → por FOB (es % del FOB de cada item)
#   - Seguro → por FOB
#   - Derechos aduaneros → por FOB
#   - Honorario DHL → equitativo por item
print(f"{'Model':<14} {'Item':<35} {'Peso':>5} {'FOB':>7} "
      f"{'Flete':>7} {'Comis':>6} {'Seg':>5} {'Der':>5} {'Hon':>5} {'Gastos':>7} {'Costo CLP/u':>13}")
print('-'*120)
resultados = []
total_clp = 0
for name, model, peso, fob, qty in items:
    pct_peso = peso / suma_peso
    pct_fob  = (fob*qty) / suma_fob
    flete    = pct_peso * GASTO_FLETE_DHL
    comision = pct_fob  * comision_steven_usd
    seguro   = pct_fob  * GASTO_SEGURO
    derechos = pct_fob  * GASTO_DERECHOS
    honor    = GASTO_HONORARIO_DHL / N_ITEMS
    gastos   = flete + comision + seguro + derechos + honor
    costo_unit_usd = fob + gastos / qty
    costo_unit_clp = costo_unit_usd * DOLAR
    total_clp += costo_unit_clp * qty
    resultados.append({
        'Model':model,'Item':name,'Peso':peso,'FOB':fob,
        'Flete':flete,'Comision':comision,'Seguro':seguro,
        'Derechos':derechos,'Honorario':honor,'Gastos':gastos,
        'Costo unit USD':costo_unit_usd,'Costo unit CLP':costo_unit_clp,
    })
    print(f"{model:<14} {name[:34]:<35} {peso:>4.1f}kg ${fob:>5.2f} ${flete:>5.2f} "
          f"${comision:>4.2f} ${seguro:>3.2f} ${derechos:>3.2f} ${honor:>3.2f} ${gastos:>5.2f}  "
          f"${costo_unit_clp:>10,.0f}")
print('-'*120)
print(f"{'TOTAL':<50} {suma_peso:>4.1f}kg ${suma_fob:>5.2f} "
      f"${GASTO_FLETE_DHL:>5.2f} ${comision_steven_usd:>4.2f} ${GASTO_SEGURO:>3.2f} "
      f"${GASTO_DERECHOS:>3.2f} ${GASTO_HONORARIO_DHL:>3.2f} ${total_gastos_usd:>5.2f}  "
      f"${total_clp:>10,.0f}")

# === Comparativa contra Erich
print(f"\n{'='*100}")
print(f"COMPARATIVA: Híbrido (gastos Erich + prorrateo por peso) vs Erich (todo FOB)")
print(f"{'='*100}")
erich_unit = {  # CLP/u del costeo Erich
    'TB12L #1':     51754, 'TB12L #2':     51754,
    'H-0259':       29657,
    'TP009B':       26749, 'TP009G':       23260,
    'TP010B':       36635, 'TP010G':       34891,
    'GT1 Negro':   143632, 'GT1 Rosado':  143632, 'GT1 Naranjo': 143632,
    'TP166':        26749,
    'TP407-120':   118337, 'TP407-127':   122117,
    'gamer ctrl':   26168,
}

print(f"{'Model':<14} {'Item':<32} {'Híbrido':>13} {'Erich':>13} {'Δ CLP':>11} {'Δ %':>7}")
print('-'*92)
for r in resultados:
    e = erich_unit.get(r['Model'], 0)
    d = r['Costo unit CLP'] - e
    pct = (d/e)*100 if e else 0
    arrow = '⬆' if d>0 else '⬇' if d<0 else '='
    print(f"{r['Model']:<14} {r['Item'][:31]:<32} ${r['Costo unit CLP']:>10,.0f}  ${e:>10,.0f}  {arrow}{abs(d):>8,.0f} {pct:>+5.1f}%")

print(f"\n{'TOTAL':<47} ${total_clp:>10,.0f}  ${sum(erich_unit.values()):>10,.0f}")

# === Generar Excel
print(f"\n[Generando Excel...]")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Costeo 26TP0609 DHL"
ws.sheet_view.showGridLines = False

AZUL_H  = PatternFill('solid', fgColor='FF1F3864')
AZUL_S  = PatternFill('solid', fgColor='FF1F4E79')
GRIS    = PatternFill('solid', fgColor='FFF2F2F2')
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
BLACK_B = Font(color='FF000000', bold=True, size=11)
NORMAL  = Font(size=10)
CENTER  = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT   = Alignment(horizontal='right', vertical='center')
LEFT    = Alignment(horizontal='left', vertical='center')
thin    = Side(border_style='thin', color='FFB4B4B4')
BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)

# Título
ws.merge_cells('A1:M1')
ws['A1'] = 'COSTEO 26TP0609PI DHL  ·  AWB 7477757895  ·  44 kg / 3 cajas'
ws['A1'].fill = AZUL_H; ws['A1'].font = Font(color='FFFFFFFF', bold=True, size=14)
ws['A1'].alignment = CENTER
ws.row_dimensions[1].height = 28

ws.merge_cells('A2:M2')
ws['A2'] = f'Dólar: $894,79 CLP/USD  |  Lógica: flete por PESO  |  Seguro/Comisión/Derechos por FOB  |  Honorario equitativo'
ws['A2'].fill = AZUL_S; ws['A2'].font = WHITE_B; ws['A2'].alignment = CENTER
ws.row_dimensions[2].height = 20

# Encabezado tabla
HEADERS = ['#','Model','Item','Peso (kg)','FOB USD','Flete USD','Comisión USD',
           'Seguro USD','Derechos USD','Honor. USD','Gastos USD','Costo unit USD','Costo unit CLP']
for i, h in enumerate(HEADERS):
    c = ws.cell(row=4, column=i+1, value=h)
    c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

# Filas de items
for idx, r in enumerate(resultados):
    row = 5 + idx
    vals = [idx+1, r['Model'], r['Item'], r['Peso'], r['FOB'],
            r['Flete'], r['Comision'], r['Seguro'], r['Derechos'],
            r['Honorario'], r['Gastos'], r['Costo unit USD'], r['Costo unit CLP']]
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=j+1, value=v)
        c.font = NORMAL; c.border = BORDER
        if j == 0: c.alignment = CENTER
        elif j in (1,2): c.alignment = LEFT
        else:
            c.alignment = RIGHT
            if j == 3: c.number_format = '0.00" kg"'
            elif j in (4,5,6,7,8,9,10,11): c.number_format = '$#,##0.00'
            elif j == 12:
                c.number_format = '$#,##0'
                c.font = Font(bold=True, size=10)

# Fila total
total_row = 5 + len(resultados)
ws.cell(row=total_row, column=1, value='').fill = GRIS
ws.cell(row=total_row, column=2, value='TOTAL').fill = GRIS
ws.cell(row=total_row, column=2).font = BLACK_B
ws.cell(row=total_row, column=3, value=f'{N_ITEMS} items').fill = GRIS
totals = [PESO_TOTAL, suma_fob, GASTO_FLETE_DHL, comision_steven_usd,
          GASTO_SEGURO, GASTO_DERECHOS, GASTO_HONORARIO_DHL, total_gastos_usd,
          (suma_fob + total_gastos_usd), total_clp]
for j, v in enumerate(totals):
    c = ws.cell(row=total_row, column=4+j, value=v)
    c.fill = GRIS; c.font = BLACK_B; c.alignment = RIGHT; c.border = BORDER
    if j == 0: c.number_format = '0.00" kg"'
    elif j in (1,2,3,4,5,6,7,8): c.number_format = '$#,##0.00'
    elif j == 9: c.number_format = '$#,##0'

# Resumen gastos
sr = total_row + 3
ws.cell(row=sr, column=1, value='RESUMEN GASTOS').font = BLACK_B
ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=5)
ws.cell(row=sr, column=1).fill = AZUL_H
ws.cell(row=sr, column=1).font = WHITE_B
ws.cell(row=sr, column=1).alignment = CENTER

resumen_gastos = [
    ('Flete DHL puerta-a-puerta (Erich)', GASTO_FLETE_DHL),
    ('Comisión Steven 3% sobre FOB',       comision_steven_usd),
    ('Seguro',                              GASTO_SEGURO),
    ('Derechos aduaneros',                  GASTO_DERECHOS),
    ('Honorario desaduanaje DHL',           GASTO_HONORARIO_DHL),
    ('TOTAL GASTOS',                        total_gastos_usd),
    ('',                                    None),
    ('FOB total',                           suma_fob),
    ('Internado total USD',                 suma_fob + total_gastos_usd),
    ('Internado total CLP (× 894,79)',     (suma_fob + total_gastos_usd) * DOLAR),
]
for i, (label, val) in enumerate(resumen_gastos):
    rr = sr+1+i
    ws.cell(row=rr, column=1, value=label).font = NORMAL
    if val is not None:
        c = ws.cell(row=rr, column=2, value=val)
        c.font = BLACK_B if 'TOTAL' in label or 'Internado' in label else NORMAL
        c.alignment = RIGHT
        if 'CLP' in label: c.number_format = '$#,##0'
        else: c.number_format = '$#,##0.00'

# Anchos
for col, w in {'A':4,'B':14,'C':36,'D':10,'E':9,'F':10,'G':10,'H':9,'I':10,'J':9,'K':9,'L':12,'M':13}.items():
    ws.column_dimensions[col].width = w

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"[OK] Excel: {OUT}")
print(f"\nTOTAL HÍBRIDO: ${total_clp:,.0f} CLP")
print(f"TOTAL ERICH:   ${sum(erich_unit.values()):,.0f} CLP")
print(f"Diferencia:    ${total_clp - sum(erich_unit.values()):+,.0f} CLP (debería ser ~0)")
