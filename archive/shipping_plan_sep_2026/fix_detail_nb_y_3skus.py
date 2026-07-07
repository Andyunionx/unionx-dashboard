"""Fix bugs post-reset:
  1. Detail_NB: consolidar 'Embarque' (col 14 original) con 'Embarque (Steven)' (col 16 nueva)
     + limpiar cols duplicadas 19, 20
  2. Mover XRCOOLPORT40-GR, XRCOOLPORTF40-GR, XRCARINF34-CL de Missing (CBM=0 en OHNSO)
"""
import openpyxl, sys, pandas as pd, re
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
sys.stdout.reconfigure(encoding='utf-8')

P = "data/planillas/Shipping Plan September V.02_actualizado.xlsx"
OHNSO = "agente-comex/data/inbox/20260701_0912_shipping_plan_ohnso/OHNSO-Jul.01st.xls"

VERDE  = PatternFill('solid', fgColor='FF92D050')
AZUL_S = PatternFill('solid', fgColor='FF9DC3E6')
ROJO   = PatternFill('solid', fgColor='FFF08080')
GRIS   = PatternFill('solid', fgColor='FFD9D9D9')
AZUL_H = PatternFill('solid', fgColor='FF1F3864')
NARANJA= PatternFill('solid', fgColor='FFFCE4D6')
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
NORMAL = Font(size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.load_workbook(P)

# ============ FIX 1: Detail_NB cols
print("=== FIX 1: Detail_NB — consolidar cols ===")
ws = wb['Detail_NB']

# La estructura correcta debe ser:
# 1 Container | 2 CRD | 3 ETA | 4 Order No | 5 Brand | 6 SKU | 7 Description | 8 Units | 9 CBM | 10 USD Value | 11 Need to pay? | 12 Late? | 13 Cantidad cargada | 14 Embarque (Steven) | 15 CRD | 16 ETD | 17 ETA Puerto | 18 ETA Bodega

# Verificar estado actual
hdr = {c: ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)}
print(f"  Headers antes: {hdr}")

# Para cada fila: preferir v16 (Embarque Steven) sobre v14 si v14 es fórmula "=A2"
for r in range(2, ws.max_row+1):
    v14 = ws.cell(row=r, column=14).value
    v16 = ws.cell(row=r, column=16).value
    # Preferir v16 si tiene valor válido
    if v16 and str(v16).strip():
        ws.cell(row=r, column=14, value=v16)
    elif v14 and str(v14).strip().startswith('='):
        ws.cell(row=r, column=14, value=None)

# Consolidar CRD, ETD, ETA Puerto, ETA Bodega en cols 15-18
# Estado actual: col 15 vacío, col 17 vacío, col 18=ETD, col 19=ETA Puerto, col 20=ETA Bodega
# CRD original está en col 2

for r in range(2, ws.max_row+1):
    # CRD (col 2 original) → col 15
    crd = ws.cell(row=r, column=2).value
    # ETD (col 18) → col 16
    etd = ws.cell(row=r, column=18).value
    # ETA Puerto (col 19) → col 17
    etap = ws.cell(row=r, column=19).value
    # ETA Bodega (col 20) → col 18
    etab = ws.cell(row=r, column=20).value

    if crd:
        c = ws.cell(row=r, column=15, value=crd)
        c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL
    if etd:
        c = ws.cell(row=r, column=16, value=etd)
        c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL
    if etap:
        c = ws.cell(row=r, column=17, value=etap)
        c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL
    if etab:
        c = ws.cell(row=r, column=18, value=etab)
        c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL

# Headers correctos
ws.cell(row=1, column=14, value='Embarque (Steven)').font = WHITE_B
ws.cell(row=1, column=15, value='CRD').font = WHITE_B
ws.cell(row=1, column=16, value='ETD').font = WHITE_B
ws.cell(row=1, column=17, value='ETA Puerto').font = WHITE_B
ws.cell(row=1, column=18, value='ETA Bodega').font = WHITE_B
for c in [14,15,16,17,18]:
    ws.cell(row=1, column=c).fill = AZUL_H
    ws.cell(row=1, column=c).alignment = CENTER
    ws.cell(row=1, column=c).border = BORDER

# Limpiar cols 19, 20 (residuo)
for c in [19, 20]:
    for r in range(1, ws.max_row+1):
        cell = ws.cell(row=r, column=c)
        cell.value = None
        cell.border = Border()
        cell.fill = PatternFill(fill_type=None)

print("  Cols consolidadas ✓")

# ============ FIX 2: mover 3 SKUs CBM=0
print("\n=== FIX 2: mover XRCOOLPORT40-GR / XRCOOLPORTF40-GR / XRCARINF34-CL ===")

# Parse OHNSO para price
def parsear(sheet):
    df = pd.read_excel(OHNSO, sheet_name=sheet, header=None)
    items, seccion, hdr = [], None, None
    for _, row in df.iterrows():
        vals = [v for v in row.values if pd.notna(v)]
        if not vals: continue
        first = str(vals[0]).strip()
        m = re.match(r'^\d+\.(.+)$', first)
        if m and len(vals) <= 2: seccion = m.group(1).strip(); continue
        if first == 'No' and len(vals)>1 and 'Model' in str(vals[1]):
            hdr = [str(v).strip() for v in row.values]; continue
        try: int(first)
        except: continue
        if hdr is None or seccion is None: continue
        it = {col: row.values[j] for j,col in enumerate(hdr) if j<len(row.values)}
        items.append(it)
    return items

ohnso_items = parsear('SHENZHEN') + parsear('NINGBO')
ohnso_price = {}
for it in ohnso_items:
    sku = str(it.get('SKU','') or '').strip()
    if not sku: continue
    p = pd.to_numeric(it.get('Price\n(USD)'), errors='coerce')
    if pd.notna(p) and p > 0: ohnso_price[sku] = float(p)

CRD_NEW = datetime(2026, 8, 15)
def fechas_NB(crd):
    etd = crd + timedelta(days=7)
    etap = etd + timedelta(days=45)
    etab = etap + timedelta(days=7)
    return crd, etd, etap, etab
def fechas_SZ(crd):
    etd = crd + timedelta(days=7)
    etap = etd + timedelta(days=52)
    etab = etap + timedelta(days=7)
    return crd, etd, etap, etab

asignaciones_extras = {
    'XRCOOLPORT40-GR':  {'emb': 'NB-02',       'units': 200, 'brand': 'Xride', 'desc': 'Cooler Eléctrico Portátil 40L 12V Gray Xroad'},
    'XRCOOLPORTF40-GR': {'emb': 'NB-02',       'units': 100, 'brand': 'Xride', 'desc': 'Cooler Freezer Eléctrico Portátil 40L 12V Gray Xroad'},
    'XRCARINF34-CL':    {'emb': 'SZ-05,40HQ',  'units': 20,  'brand': 'Xride', 'desc': 'Carpa Inflable 3-4 Personas 3 Estaciones Xroad'},
}

ws_m = wb['Missing']
ws_sz = wb['Detail_SZ']
ws_nb = wb['Detail_NB']

rows_borrar = []
for r in range(2, ws_m.max_row+1):
    sku = str(ws_m.cell(row=r, column=2).value or '').strip()
    if sku in asignaciones_extras:
        info = asignaciones_extras[sku]
        emb = info['emb']
        crd, etd, etap, etab = fechas_NB(CRD_NEW) if 'NB' in emb else fechas_SZ(CRD_NEW)
        usd = round(info['units'] * ohnso_price.get(sku, 0), 2) if ohnso_price.get(sku) else None
        ws_target = ws_nb if 'NB' in emb.upper() else ws_sz
        r_new = max([rr for rr in range(2, ws_target.max_row+1) if ws_target.cell(row=rr, column=6).value] + [1]) + 1
        ws_target.cell(row=r_new, column=1, value=emb)
        ws_target.cell(row=r_new, column=3, value=etab)
        ws_target.cell(row=r_new, column=4, value='MISSING->OHNSO')
        ws_target.cell(row=r_new, column=5, value=info['brand'])
        ws_target.cell(row=r_new, column=6, value=sku)
        ws_target.cell(row=r_new, column=7, value=info['desc'])
        ws_target.cell(row=r_new, column=8, value=info['units'])
        if usd is not None:
            ws_target.cell(row=r_new, column=10, value=usd)
        ws_target.cell(row=r_new, column=13, value=info['units'])
        ws_target.cell(row=r_new, column=14, value=emb)
        ws_target.cell(row=r_new, column=15, value=crd)
        ws_target.cell(row=r_new, column=16, value=etd)
        ws_target.cell(row=r_new, column=17, value=etap)
        ws_target.cell(row=r_new, column=18, value=etab)
        for c in range(1, 19):
            cell = ws_target.cell(row=r_new, column=c)
            cell.border = BORDER; cell.font = NORMAL
            if c in (3, 15, 16, 17, 18):
                cell.number_format = 'DD-MM-YYYY'; cell.alignment = CENTER
            elif c in (8, 10, 13):
                cell.number_format = '#,##0'; cell.alignment = RIGHT
            else:
                cell.alignment = LEFT if c in (5, 7) else CENTER
        ws_target.cell(row=r_new, column=1).fill = NARANJA
        ws_target.cell(row=r_new, column=6).fill = NARANJA
        ws_target.cell(row=r_new, column=13).fill = VERDE
        rows_borrar.append(r)
        print(f"  {sku} ({info['units']}u) -> {emb}")

for r in sorted(rows_borrar, reverse=True):
    ws_m.delete_rows(r, 1)

# ============ Re-aplicar CF Detail_NB
print("\n=== Re-aplicar CF Detail_NB ===")
ws = wb['Detail_NB']
c_l = 'M'; u_l = 'H'
last = ws.max_row
ws.conditional_formatting._cf_rules = {}
rng = f'{c_l}2:{c_l}{last}'
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2={u_l}2)'], fill=VERDE))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2>{u_l}2)'], fill=AZUL_S))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2<{u_l}2,{c_l}2>0)'], fill=ROJO))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR({c_l}2="",ISBLANK({c_l}2))'], fill=GRIS))
for r in range(2, last+1):
    u = ws.cell(row=r, column=8).value
    c_val = ws.cell(row=r, column=13).value
    cell = ws.cell(row=r, column=13)
    if c_val is None or not isinstance(c_val,(int,float)):
        cell.fill = GRIS
    elif not isinstance(u,(int,float)):
        cell.fill = GRIS
    elif c_val == u: cell.fill = VERDE
    elif c_val > u: cell.fill = AZUL_S
    else: cell.fill = ROJO

wb.save(P)
print("\n[OK] Guardado")

# Verificar
for sh in ['Detail_SZ', 'Detail_NB', 'Missing']:
    df = pd.read_excel(P, sheet_name=sh)
    df.columns = [str(c).strip() for c in df.columns]
    df_v = df[df['Description'].notna() & (df['Description'].astype(str).str.strip()!='')] if 'Description' in df.columns else df.dropna(how='all')
    print(f"  {sh}: {len(df_v)} filas")
    if 'Embarque (Steven)' in df_v.columns:
        e = df_v.groupby('Embarque (Steven)').size().reset_index(name='n')
        for _, r in e.iterrows():
            print(f"    {r['Embarque (Steven)']}: {r['n']}")
