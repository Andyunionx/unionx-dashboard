"""V2: fórmulas dinámicas + CF en Detail_SZ/NB + cruce Late/Pend/Missing + Summary plan ajustado."""
import pandas as pd, sys, re
from collections import defaultdict
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
sys.stdout.reconfigure(encoding='utf-8')

P = "data/planillas/Shipping Plan September V.02_actualizado.xlsx"
SV = "data/comex/Shipping Plan B-Jun.30th.xls"

TRANSITO = {'SZ': 52, 'NB': 45}
CRD_ETD = 7
PTO_BOD = 7

# --- Estilos
AZUL_H = PatternFill('solid', fgColor='FF1F3864')
ROJO   = PatternFill('solid', fgColor='FFE6B8B8')
AZUL_S = PatternFill('solid', fgColor='FFBDD7EE')
VERDE  = PatternFill('solid', fgColor='FFC6EFCE')
GRIS   = PatternFill('solid', fgColor='FFF2F2F2')
AMARI  = PatternFill('solid', fgColor='FFFFFF00')
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
NORMAL = Font(size=10)
BOLD = Font(bold=True, size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Parse Steven
def parsear_steven(sheet):
    df = pd.read_excel(SV, sheet_name=sheet, header=None)
    items, seccion, hdr = [], None, None
    for _, row in df.iterrows():
        vals = [v for v in row.values if pd.notna(v)]
        if not vals: continue
        first = str(vals[0]).strip()
        m = re.match(r'^\d+\.(.+)$', first)
        if m and len(vals) <= 2: seccion = m.group(1).strip(); continue
        if first == 'No' and len(vals)>1 and 'Model' in str(vals[1]):
            hdr = [str(v).strip() for v in row.values]; continue
        try: int(first)
        except: continue
        if hdr is None or seccion is None: continue
        it = {col: row.values[j] for j,col in enumerate(hdr) if j<len(row.values)}
        it['__seccion'] = seccion
        items.append(it)
    return items

steven = parsear_steven('SHENZHEN') + parsear_steven('NINGBO')
idx = defaultdict(list)
for it in steven:
    sku = str(it.get('SKU', '') or '').strip()
    if not sku or sku.lower()=='nan': continue
    qty = pd.to_numeric(it.get('Qty'), errors='coerce')
    if not pd.notna(qty): continue
    finish = it.get('Finish Time')
    idx[sku].append({
        'embarque': it['__seccion'],
        'qty': int(qty),
        'finish': finish if isinstance(finish, datetime) else None,
        'model': str(it.get('Model','')).strip(),
    })

# CRD por embarque
emb_crd = {}
for sku, lst in idx.items():
    for m in lst:
        if m['finish']:
            e = m['embarque']
            if e not in emb_crd or m['finish'] > emb_crd[e]:
                emb_crd[e] = m['finish']
print("CRD por embarque Steven:")
for e, d in emb_crd.items():
    print(f"  {e}: {d.date()}")

def fechas_embarque(emb):
    if emb not in emb_crd: return None,None,None,None
    crd = emb_crd[emb]
    transit = TRANSITO['NB'] if 'NB' in emb.upper() or 'NINGBO' in emb.upper() else TRANSITO['SZ']
    etd = crd + timedelta(days=CRD_ETD)
    etap = etd + timedelta(days=transit)
    etab = etap + timedelta(days=PTO_BOD)
    return crd, etd, etap, etab

# --- Open workbook
wb = openpyxl.load_workbook(P)

# === 1) Detail_SZ + Detail_NB: fórmulas + CF
print("\n=== Detail_SZ + Detail_NB: fórmulas + CF ===")
for sh, n_rows in [('Detail_SZ', 33), ('Detail_NB', 3)]:
    ws = wb[sh]
    # Columnas: A Container B OT C ... F SKU G Description H Units ... M Cantidad cargada N Embarque O CRD P ETD Q ETA Pto R ETA Bod
    # Δ y Estado están en S y T? Verificar
    # max_col=22 antes; busco las headers
    hdr_row = 1
    headers = {ws.cell(row=hdr_row, column=c).value: c for c in range(1, ws.max_column+1)}
    print(f"  [{sh}] headers: {headers}")

    col_units = headers.get('Units')
    col_cant  = headers.get('Cantidad cargada')

    # CF sobre Cantidad cargada (col M=13 normalmente)
    if col_units and col_cant:
        u_letter = openpyxl.utils.get_column_letter(col_units)
        c_letter = openpyxl.utils.get_column_letter(col_cant)
        last = ws.max_row
        rng = f'{c_letter}2:{c_letter}{last}'
        # Limpiar reglas previas en ese rango
        ws.conditional_formatting._cf_rules = {k:v for k,v in ws.conditional_formatting._cf_rules.items() if k.sqref != rng}
        ws.conditional_formatting.add(rng,
            FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2), {c_letter}2={u_letter}2)'], fill=VERDE))
        ws.conditional_formatting.add(rng,
            FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2), {c_letter}2>{u_letter}2)'], fill=AZUL_S))
        ws.conditional_formatting.add(rng,
            FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2), {c_letter}2<{u_letter}2, {c_letter}2>0)'], fill=ROJO))
        ws.conditional_formatting.add(rng,
            FormulaRule(formula=[f'OR({c_letter}2="", ISBLANK({c_letter}2))'], fill=GRIS))
        # Eliminar fills duros previos en cantidad cargada para que el CF mande
        for r in range(2, last+1):
            ws.cell(row=r, column=col_cant).fill = PatternFill(fill_type=None)
        print(f"    CF aplicado a {rng}")

# === 2) Late_Arrivals — agregar columnas con cruce
print("\n=== Late_Arrivals: agregar cruce ===")
ws = wb['Late_Arrivals']
last_col = ws.max_column
# Headers nuevas en cols 9..14
new_headers = ['Cantidad cargada', 'Embarque (Steven)', 'CRD', 'ETD', 'ETA Puerto', 'ETA Bodega']
for j, h in enumerate(new_headers):
    c = ws.cell(row=1, column=last_col+1+j, value=h)
    c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

MATCHES_ESP = {'TCMULTSTY5N1-BG': 'TCMULTSTY5N1'}  # = HD4 Khaki, mismo modelo

for r in range(2, ws.max_row+1):
    sku = str(ws.cell(row=r, column=2).value or '').strip()
    units = pd.to_numeric(ws.cell(row=r, column=5).value, errors='coerce')
    if not sku: continue
    sku_lookup = MATCHES_ESP.get(sku, sku)
    matches = idx.get(sku_lookup, [])
    # Preferir embarques NO Rest items con qty >= units, exact si existe
    cargada = None; embarque = None
    if matches:
        # exact qty match en embarque real
        for m in matches:
            if m['embarque']!='Rest items' and m['qty']==int(units or 0):
                cargada = m['qty']; embarque = m['embarque']; break
        # match >= units en embarque real
        if cargada is None:
            for m in matches:
                if m['embarque']!='Rest items' and m['qty']>=int(units or 0):
                    cargada = m['qty']; embarque = m['embarque']; break
        # any embarque real
        if cargada is None:
            for m in matches:
                if m['embarque']!='Rest items':
                    cargada = m['qty']; embarque = m['embarque']; break
        # Rest items como fallback (suma)
        if cargada is None:
            rest = [m for m in matches if m['embarque']=='Rest items']
            if rest:
                cargada = sum(m['qty'] for m in rest); embarque = 'Rest items'
    # Escribir
    base = last_col
    c1 = ws.cell(row=r, column=base+1, value=cargada);  c1.alignment=RIGHT; c1.border=BORDER; c1.font=NORMAL; c1.number_format='#,##0'
    c2 = ws.cell(row=r, column=base+2, value=embarque); c2.alignment=CENTER; c2.border=BORDER; c2.font=NORMAL
    if embarque:
        crd, etd, etap, etab = fechas_embarque(embarque)
        for j, val in enumerate([crd, etd, etap, etab]):
            c = ws.cell(row=r, column=base+3+j, value=val)
            c.alignment=CENTER; c.border=BORDER; c.font=NORMAL; c.number_format='DD-MM-YYYY'
    else:
        for j in range(4):
            c = ws.cell(row=r, column=base+3+j); c.border=BORDER

# CF Cantidad cargada en Late_Arrivals (col last_col+1) vs Units (col 5)
c_letter = openpyxl.utils.get_column_letter(last_col+1)
u_letter = 'E'
rng = f'{c_letter}2:{c_letter}{ws.max_row}'
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2={u_letter}2)'], fill=VERDE))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2>{u_letter}2)'], fill=AZUL_S))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2<{u_letter}2,{c_letter}2>0)'], fill=ROJO))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR({c_letter}2="",ISBLANK({c_letter}2))'], fill=GRIS))

# Anchos
for j, w in enumerate([14, 22, 12, 12, 12, 12]):
    ws.column_dimensions[openpyxl.utils.get_column_letter(last_col+1+j)].width = w

# === 3) Pend_Confirmation
print("\n=== Pend_Confirmation: agregar cruce ===")
ws = wb['Pend_Confirmation']
last_col = ws.max_column
for j, h in enumerate(new_headers):
    c = ws.cell(row=1, column=last_col+1+j, value=h)
    c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

for r in range(2, ws.max_row+1):
    sku = str(ws.cell(row=r, column=2).value or '').strip()
    units = pd.to_numeric(ws.cell(row=r, column=4).value, errors='coerce')
    if not sku: continue
    matches = idx.get(sku, [])
    cargada = None; embarque = None
    if matches:
        for m in matches:
            if m['embarque']!='Rest items' and m['qty']>=int(units or 0):
                cargada=m['qty']; embarque=m['embarque']; break
        if cargada is None:
            rest = [m for m in matches if m['embarque']=='Rest items']
            if rest:
                cargada = sum(m['qty'] for m in rest); embarque = 'Rest items'
    base = last_col
    c1 = ws.cell(row=r, column=base+1, value=cargada); c1.alignment=RIGHT; c1.border=BORDER; c1.font=NORMAL; c1.number_format='#,##0'
    c2 = ws.cell(row=r, column=base+2, value=embarque); c2.alignment=CENTER; c2.border=BORDER; c2.font=NORMAL
    if embarque:
        crd, etd, etap, etab = fechas_embarque(embarque)
        for j, val in enumerate([crd, etd, etap, etab]):
            c = ws.cell(row=r, column=base+3+j, value=val)
            c.alignment=CENTER; c.border=BORDER; c.font=NORMAL; c.number_format='DD-MM-YYYY'

c_letter = openpyxl.utils.get_column_letter(last_col+1)
u_letter = 'D'
rng = f'{c_letter}2:{c_letter}{ws.max_row}'
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2={u_letter}2)'], fill=VERDE))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2>{u_letter}2)'], fill=AZUL_S))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2<{u_letter}2,{c_letter}2>0)'], fill=ROJO))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR({c_letter}2="",ISBLANK({c_letter}2))'], fill=GRIS))
for j, w in enumerate([14, 22, 12, 12, 12, 12]):
    ws.column_dimensions[openpyxl.utils.get_column_letter(last_col+1+j)].width = w

# === 4) Missing — agregar LHSHNUCO-BK y cruce SÓLO para amarillos
print("\n=== Missing: agregar LHSHNUCO-BK + cruce amarillos ===")
ws = wb['Missing']
last_col = ws.max_column
# Headers
for j, h in enumerate(new_headers):
    c = ws.cell(row=1, column=last_col+1+j, value=h)
    c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

# Detectar filas amarillas existentes
amarillos = []
for r in range(2, ws.max_row+1):
    f = ws.cell(row=r, column=2).fill
    if f and f.start_color and str(f.start_color.rgb)=='FFFFFF00':
        amarillos.append(r)
print(f"Amarillos detectados: filas {amarillos}")

# Asignaciones para los 3 SKUs amarillos (NY=500, BG=300, BK=200; suma=1000=stock Steven)
asignacion = {
    'LHSHNUCO-NY': 500,
    'LHSHNUCO-BG': 300,
    'LHSHNUCO-BK': 200,
}

# Agregar LHSHNUCO-BK como nueva fila al final
new_row = ws.max_row + 1
brand = ws.cell(row=amarillos[0], column=1).value if amarillos else 'Lhotse'
ws.cell(row=new_row, column=1, value=brand)
ws.cell(row=new_row, column=2, value='LHSHNUCO-BK')
ws.cell(row=new_row, column=3, value='Shaker Vaso Proteina Acero Inox NutriCore Negro')
ws.cell(row=new_row, column=4, value=200)
ws.cell(row=new_row, column=5, value='BATCH 617')
# Pintar amarillo cols 1-5
for c in range(1, 6):
    cell = ws.cell(row=new_row, column=c)
    cell.fill = AMARI; cell.font = NORMAL; cell.border = BORDER
    cell.alignment = LEFT if c<=3 or c==5 else RIGHT
print(f"  Agregada fila {new_row}: LHSHNUCO-BK 200u")

# Re-detectar amarillos (incluye nueva)
amarillos = []
for r in range(2, ws.max_row+1):
    f = ws.cell(row=r, column=2).fill
    if f and f.start_color and str(f.start_color.rgb)=='FFFFFF00':
        amarillos.append(r)
print(f"Amarillos finales: filas {amarillos}")

# Llenar cruce para amarillos
for r in amarillos:
    sku = str(ws.cell(row=r, column=2).value or '').strip()
    qty_asig = asignacion.get(sku, 0)
    embarque = 'Rest items'  # los amarillos vienen del stock Rest items 1000 mix
    base = last_col
    c1 = ws.cell(row=r, column=base+1, value=qty_asig); c1.alignment=RIGHT; c1.border=BORDER; c1.font=NORMAL; c1.number_format='#,##0'
    c2 = ws.cell(row=r, column=base+2, value=embarque); c2.alignment=CENTER; c2.border=BORDER; c2.font=NORMAL
    # CRD del item TP30 = 2026-07-10
    crd = datetime(2026, 7, 10)
    etd = crd + timedelta(days=CRD_ETD)
    etap = etd + timedelta(days=TRANSITO['SZ'])
    etab = etap + timedelta(days=PTO_BOD)
    for j, val in enumerate([crd, etd, etap, etab]):
        c = ws.cell(row=r, column=base+3+j, value=val)
        c.alignment=CENTER; c.border=BORDER; c.font=NORMAL; c.number_format='DD-MM-YYYY'

c_letter = openpyxl.utils.get_column_letter(last_col+1)
u_letter = 'D'
rng = f'{c_letter}2:{c_letter}{ws.max_row}'
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2={u_letter}2)'], fill=VERDE))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2>{u_letter}2)'], fill=AZUL_S))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2<{u_letter}2,{c_letter}2>0)'], fill=ROJO))
for j, w in enumerate([14, 22, 12, 12, 12, 12]):
    ws.column_dimensions[openpyxl.utils.get_column_letter(last_col+1+j)].width = w

# === 5) Summary — agregar bloque "PLAN AJUSTADO (Steven B-Jun.30th)"
print("\n=== Summary: bloque plan ajustado ===")
ws = wb['Summary']
start = ws.max_row + 2

c = ws.cell(row=start, column=1, value='PLAN AJUSTADO (Steven Plan B-Jun.30th)')
c.font = Font(bold=True, size=12, color='FF1F3864')
ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=10)

hdr = ['Embarque Steven', 'Origen', 'CRD', 'ETD', 'ETA Puerto', 'ETA Bodega', 'SKUs', 'Uds totales', 'Uds UnionX V02', '¿A tiempo (≤31-ago-26)?']
for j, h in enumerate(hdr):
    cc = ws.cell(row=start+1, column=j+1, value=h)
    cc.fill = AZUL_H; cc.font = WHITE_B; cc.alignment = CENTER; cc.border = BORDER

# Datos por embarque
embarques_ord = ['SZ-04,40HQ', 'SZ-03-08,40HQ', 'SZ-01(623)', 'SZ-02(623)', 'IMP OP 350-26 40HQ', 'NB-01']
DEADLINE = datetime(2026, 8, 31)

# Contar uds UnionX V02 cargadas por embarque (de Detail_SZ + Detail_NB)
ux_x_emb = defaultdict(lambda: {'skus':0, 'uds':0})
for sh in ['Detail_SZ', 'Detail_NB']:
    ws_d = wb[sh]
    for r in range(2, ws_d.max_row+1):
        e = ws_d.cell(row=r, column=14).value
        cant = ws_d.cell(row=r, column=13).value
        if e and isinstance(cant, (int, float)):
            ux_x_emb[e]['skus'] += 1
            ux_x_emb[e]['uds'] += cant

# Uds totales por embarque Steven (todos los items, no solo UnionX)
total_x_emb = defaultdict(lambda: {'skus':0, 'uds':0})
for sku, lst in idx.items():
    for m in lst:
        total_x_emb[m['embarque']]['skus'] += 1
        total_x_emb[m['embarque']]['uds'] += m['qty']

row = start + 2
for e in embarques_ord:
    crd, etd, etap, etab = fechas_embarque(e)
    origen = 'NB' if 'NB' in e.upper() or 'NINGBO' in e.upper() else 'SZ'
    a_tiempo = 'SÍ' if etab and etab <= DEADLINE else 'NO'
    vals = [e, origen, crd, etd, etap, etab,
            total_x_emb[e]['skus'], total_x_emb[e]['uds'],
            ux_x_emb[e]['uds'], a_tiempo]
    for j, v in enumerate(vals):
        cc = ws.cell(row=row, column=j+1, value=v)
        cc.font = NORMAL; cc.border = BORDER
        if j in (2,3,4,5):
            cc.number_format='DD-MM-YYYY'; cc.alignment=CENTER
        elif j in (6,7,8):
            cc.number_format='#,##0'; cc.alignment=RIGHT
        elif j == 9:
            cc.alignment=CENTER; cc.font=Font(bold=True)
            cc.fill = VERDE if v=='SÍ' else ROJO
        else:
            cc.alignment=LEFT if j==0 else CENTER
    row += 1

# Total
tc = ws.cell(row=row, column=1, value='TOTAL'); tc.font=Font(bold=True); tc.fill=GRIS; tc.border=BORDER
ws.cell(row=row, column=7, value=sum(total_x_emb[e]['skus'] for e in embarques_ord)).font = Font(bold=True)
ws.cell(row=row, column=8, value=sum(total_x_emb[e]['uds'] for e in embarques_ord)).font = Font(bold=True)
ws.cell(row=row, column=9, value=sum(ux_x_emb[e]['uds'] for e in embarques_ord)).font = Font(bold=True)
for j in range(1, 11):
    cc = ws.cell(row=row, column=j); cc.fill=GRIS; cc.border=BORDER
    if j in (7,8,9): cc.number_format='#,##0'; cc.alignment=RIGHT

# Rest items aparte
row += 2
c = ws.cell(row=row, column=1, value='STOCK NO ASIGNADO (Rest items)')
c.font = Font(bold=True, size=11, color='FFC65911')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
row += 1
ws.cell(row=row, column=1, value='Rest items').font = NORMAL
ws.cell(row=row, column=7, value=total_x_emb['Rest items']['skus']).font = NORMAL
ws.cell(row=row, column=7).number_format='#,##0'; ws.cell(row=row, column=7).alignment=RIGHT
ws.cell(row=row, column=8, value=total_x_emb['Rest items']['uds']).font = NORMAL
ws.cell(row=row, column=8).number_format='#,##0'; ws.cell(row=row, column=8).alignment=RIGHT
for j in range(1, 11):
    ws.cell(row=row, column=j).border = BORDER

# Anchos summary
for col, w in [('A',26),('B',9),('C',12),('D',12),('E',12),('F',12),('G',8),('H',12),('I',13),('J',22)]:
    ws.column_dimensions[col].width = w

# === Guardar
wb.save(P)
print(f"\n[OK] Guardado: {P}")
print("\nResumen final por embarque (Steven):")
for e in embarques_ord:
    print(f"  {e:<22} skus={total_x_emb[e]['skus']:>3}  uds_total={total_x_emb[e]['uds']:>6,}  uds_UX={ux_x_emb[e]['uds']:>6,}")
print(f"  Rest items             skus={total_x_emb['Rest items']['skus']:>3}  uds_total={total_x_emb['Rest items']['uds']:>6,}")
