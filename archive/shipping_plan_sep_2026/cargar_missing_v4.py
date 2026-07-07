"""V4: cargar SKUs de Missing en embarques actuales (donde entre) o crear nuevos.
Reglas:
  - CBM proporcional según OHNSO Jul.01st
  - Origen: SHENZHEN sheet → SZ / NINGBO sheet → NB
  - Match embarque real prevalece (SMBATSAV9-CR → IMP OP, XRHYDLAV-OR → SZ-04)
  - Rest LHSHNUCO ya en NB-01 (no tocar)
  - Nuevos embarques: SZ-05,40HQ / NB-02 / IMP OP 351-26 40HQ (CRD = MAX Finish Time asignados)
"""
import openpyxl, pandas as pd, sys, re
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

P = "data/planillas/Shipping Plan September V.02_actualizado.xlsx"
OHNSO = "agente-comex/data/inbox/20260701_0912_shipping_plan_ohnso/OHNSO-Jul.01st.xls"

TRANSITO = {'SZ': 52, 'NB': 45}
CRD_ETD = 7
PTO_BOD = 7
CAP_40HQ = 68  # CBM

# Estilos
VERDE  = PatternFill('solid', fgColor='FF92D050')
AZUL_S = PatternFill('solid', fgColor='FF9DC3E6')
ROJO   = PatternFill('solid', fgColor='FFF08080')
GRIS   = PatternFill('solid', fgColor='FFD9D9D9')
AZUL_H = PatternFill('solid', fgColor='FF1F3864')
NARANJA= PatternFill('solid', fgColor='FFFCE4D6')
AMARI  = PatternFill('solid', fgColor='FFFFFF00')
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
NORMAL = Font(size=10)
BOLD = Font(bold=True, size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ============ Parsear OHNSO
def parsear(sheet):
    df = pd.read_excel(OHNSO, sheet_name=sheet, header=None)
    items, seccion, hdr = [], None, None
    for _, row in df.iterrows():
        vals = [v for v in row.values if pd.notna(v)]
        if not vals: continue
        first = str(vals[0]).strip()
        m = re.match(r'^\d+\.(.+)$', first)
        if m and len(vals) <= 2: seccion = m.group(1).strip(); continue
        if first == 'No' and len(vals)>1 and 'Model' in str(vals[1]):
            hdr = [str(v).strip() for v in row.values]; continue
        try: int(first)
        except: continue
        if hdr is None or seccion is None: continue
        it = {col: row.values[j] for j,col in enumerate(hdr) if j<len(row.values)}
        it['__seccion'] = seccion
        it['__sheet'] = sheet
        items.append(it)
    return items

items = parsear('SHENZHEN') + parsear('NINGBO')

idx = defaultdict(list)
for it in items:
    sku = str(it.get('SKU','') or '').strip()
    if not sku or sku.lower()=='nan': continue
    qty = pd.to_numeric(it.get('Qty'), errors='coerce')
    if not pd.notna(qty): continue
    cbm_it = pd.to_numeric(it.get('TOTAL CBM'), errors='coerce')
    cbm_ct = pd.to_numeric(it.get('(CBM)\n/CTN'), errors='coerce')
    qty_ct = pd.to_numeric(it.get("Q'ty/ctn"), errors='coerce')
    finish = it.get('Finish Time')
    idx[sku].append({
        'embarque': it['__seccion'],
        'sheet': it['__sheet'],
        'qty': int(qty),
        'cbm_total': cbm_it if pd.notna(cbm_it) else 0,
        'cbm_ctn': cbm_ct if pd.notna(cbm_ct) else 0,
        'qty_ctn': qty_ct if pd.notna(qty_ct) else 0,
        'finish': finish if isinstance(finish, datetime) else None,
    })

# ============ Ocupación actual
cbm_x_emb = defaultdict(float)
for it in items:
    if it['__seccion']=='Rest items': continue
    cbm = pd.to_numeric(it.get('TOTAL CBM'), errors='coerce')
    if pd.notna(cbm): cbm_x_emb[it['__seccion']] += cbm

# CRD por embarque = MAX Finish Time
emb_crd = {}
for it in items:
    if it['__seccion']=='Rest items': continue
    f = it.get('Finish Time')
    if isinstance(f, datetime):
        e = it['__seccion']
        if e not in emb_crd or f > emb_crd[e]:
            emb_crd[e] = f
    else:
        # "Ready" o "Reday" tratar como fecha ya alcanzada, usar today referencia
        pass

# ============ Cargar workbook
wb = openpyxl.load_workbook(P)
ws = wb['Missing']

# Amarillos ya asignados a NB-01 — no tocar
YA_ASIGNADOS = {'LHSHNUCO-NY','LHSHNUCO-BG','LHSHNUCO-BK'}

# Columnas de Missing: A=Brand B=SKU C=Description D=Missing Units E=Comments F=Cant cargada G=Embarque H=CRD I=ETD J=ETAPto K=ETABod
COL_CANT, COL_EMB, COL_CRD, COL_ETD, COL_ETAP, COL_ETAB = 6, 7, 8, 9, 10, 11

# Recolectar filas de Missing a asignar
items_a_cargar = []
for r in range(2, ws.max_row+1):
    sku = ws.cell(row=r, column=2).value
    desc = ws.cell(row=r, column=3).value
    units = ws.cell(row=r, column=4).value
    if not sku or not units: continue
    sku = str(sku).strip()
    if sku in YA_ASIGNADOS: continue
    if not isinstance(units,(int,float)): continue
    units = int(units)
    ms = idx.get(sku, [])
    if not ms:
        items_a_cargar.append({'row':r,'sku':sku,'units':units,'cbm':None,'origen':None,'match_real':None,'finish':None})
        continue
    m_real = [m for m in ms if m['embarque']!='Rest items']
    m = m_real[0] if m_real else ms[0]
    if m['qty_ctn'] > 0 and m['cbm_ctn'] > 0:
        cbm_needed = (units / m['qty_ctn']) * m['cbm_ctn']
    else:
        cbm_needed = (units / m['qty']) * m['cbm_total'] if m['qty'] > 0 else 0
    origen = 'NB' if m['sheet']=='NINGBO' else 'SZ'
    items_a_cargar.append({
        'row':r,'sku':sku,'units':units,'cbm':cbm_needed,'origen':origen,
        'match_real':m['embarque'] if m['embarque']!='Rest items' else None,
        'finish':m['finish'],
    })

# ============ Bin packing por origen
# Espacios libres (asumo 623 = 68 CBM también, ajustable)
libre = {
    'SZ-04,40HQ': CAP_40HQ - cbm_x_emb['SZ-04,40HQ'],
    'SZ-03-08,40HQ': CAP_40HQ - cbm_x_emb['SZ-03-08,40HQ'],
    'SZ-01(623)': CAP_40HQ - cbm_x_emb['SZ-01(623)'],
    'SZ-02(623)': CAP_40HQ - cbm_x_emb['SZ-02(623)'],
    'IMP OP 350-26 40HQ': CAP_40HQ - cbm_x_emb['IMP OP 350-26 40HQ'],
    'NB-01': CAP_40HQ - cbm_x_emb['NB-01'],
}
print("Espacio libre actuales:")
for e, v in libre.items(): print(f"  {e}: {v:.2f} CBM")

# Nuevos embarques con capacidad completa
nuevos = {}
def alloc(origen, cbm_needed):
    """Retorna nombre embarque nuevo, creándolo si es necesario."""
    if origen=='SZ':
        # SZ-05, SZ-06, ...
        candidatos = [k for k in nuevos if k.startswith('SZ-') and libre.get(k, 0) >= cbm_needed]
        if candidatos: return candidatos[0]
        n = 5 + sum(1 for k in nuevos if k.startswith('SZ-'))
        name = f'SZ-{n:02d},40HQ'
    elif origen=='NB':
        candidatos = [k for k in nuevos if k.startswith('NB-') and libre.get(k, 0) >= cbm_needed]
        if candidatos: return candidatos[0]
        n = 2 + sum(1 for k in nuevos if k.startswith('NB-'))
        name = f'NB-{n:02d}'
    else:  # IMP OP
        n = 351 + sum(1 for k in nuevos if k.startswith('IMP OP'))
        name = f'IMP OP {n}-26 40HQ'
    nuevos[name] = True
    libre[name] = CAP_40HQ
    print(f"  → CREADO {name}")
    return name

# Preferir match real; sino fit-decreasing en embarques del mismo origen (menor libre suficiente)
# Ordenar items por CBM descendente (los grandes primero, mejor bin packing)
def sortkey(it):
    return -(it['cbm'] or 0)
items_ordenados = sorted(items_a_cargar, key=sortkey)

asignaciones = {}   # row → embarque
crd_asignados = defaultdict(list)  # embarque → [finish times]

print("\nAsignaciones:")
for it in items_ordenados:
    r, sku, units, cbm, origen = it['row'], it['sku'], it['units'], it['cbm'] or 0, it['origen']
    fin = it['finish']

    if origen is None:  # sin match OHNSO
        asignaciones[r] = None
        print(f"  R{r} {sku:<20} {units:>4}u — sin match, queda sin embarque")
        continue

    # 1) Match real
    if it['match_real']:
        # ver si tiene espacio
        if libre.get(it['match_real'], 0) >= cbm:
            emb = it['match_real']
        else:
            # crear nuevo del mismo tipo
            if 'IMP OP' in it['match_real']:
                emb = alloc('IMP', cbm)
            elif 'NB' in it['match_real']:
                emb = alloc('NB', cbm)
            else:
                emb = alloc('SZ', cbm)
        libre[emb] -= cbm
        asignaciones[r] = emb
        if fin: crd_asignados[emb].append(fin)
        print(f"  R{r} {sku:<20} {units:>4}u CBM={cbm:>6.2f} → {emb} (match real)")
        continue

    # 2) Fit-decreasing entre embarques del mismo origen
    # Buscar embarques compatibles (SZ o NB)
    compat = [e for e in libre if (
        (origen=='SZ' and (e.startswith('SZ-') or e.startswith('IMP OP')))
        or (origen=='NB' and (e.startswith('NB-') or e.startswith('IMP OP')))
    ) and libre[e] >= cbm]
    if compat:
        # Best-fit: menor libre suficiente
        compat.sort(key=lambda e: libre[e])
        emb = compat[0]
    else:
        emb = alloc(origen, cbm)
    libre[emb] -= cbm
    asignaciones[r] = emb
    if fin: crd_asignados[emb].append(fin)
    print(f"  R{r} {sku:<20} {units:>4}u CBM={cbm:>6.2f} → {emb}")

print(f"\nNuevos embarques creados: {list(nuevos.keys())}")
print("Espacio libre final:")
for e, v in libre.items(): print(f"  {e}: {v:.2f} CBM")

# CRD para nuevos: MAX finish + fallback
CRD_FALLBACK = datetime(2026, 8, 15)
def crd_de(emb):
    if emb in emb_crd: return emb_crd[emb]  # existente
    if emb in crd_asignados and crd_asignados[emb]:
        return max(crd_asignados[emb])
    return CRD_FALLBACK

def fechas(emb):
    crd = crd_de(emb)
    if not crd: crd = CRD_FALLBACK
    transit = TRANSITO['NB'] if 'NB' in emb.upper() else TRANSITO['SZ']
    etd = crd + timedelta(days=CRD_ETD)
    etap = etd + timedelta(days=transit)
    etab = etap + timedelta(days=PTO_BOD)
    return crd, etd, etap, etab

# ============ Actualizar Missing rows
print("\nActualizando hoja Missing...")
for it in items_a_cargar:
    r = it['row']
    emb = asignaciones.get(r)
    if not emb:
        # dejar sin embarque
        for col in [COL_CANT, COL_EMB, COL_CRD, COL_ETD, COL_ETAP, COL_ETAB]:
            c = ws.cell(row=r, column=col)
            if col not in (COL_CANT,) or c.value is None:
                c.border = BORDER
        continue
    crd, etd, etap, etab = fechas(emb)
    # Cantidad cargada = units
    c = ws.cell(row=r, column=COL_CANT, value=it['units']); c.number_format='#,##0'; c.alignment=RIGHT; c.border=BORDER; c.font=NORMAL
    c = ws.cell(row=r, column=COL_EMB, value=emb); c.alignment=CENTER; c.border=BORDER; c.font=NORMAL
    for j, val in enumerate([crd, etd, etap, etab]):
        c = ws.cell(row=r, column=COL_CRD+j, value=val)
        c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL

# CF sobre Cantidad cargada Missing
c_letter = get_column_letter(COL_CANT)
u_letter = 'D'
ws.conditional_formatting._cf_rules = {}
rng = f'{c_letter}2:{c_letter}{ws.max_row}'
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2={u_letter}2)'], fill=VERDE))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2>{u_letter}2)'], fill=AZUL_S))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2<{u_letter}2,{c_letter}2>0)'], fill=ROJO))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR({c_letter}2="",ISBLANK({c_letter}2))'], fill=GRIS))

# Fill duro inicial en Missing
for r in range(2, ws.max_row+1):
    u = ws.cell(row=r, column=4).value
    c_val = ws.cell(row=r, column=COL_CANT).value
    cell = ws.cell(row=r, column=COL_CANT)
    # Preservar amarillo si fila amarilla en SKU
    yellow_row = False
    f_sku = ws.cell(row=r, column=2).fill
    if f_sku and f_sku.start_color and str(f_sku.start_color.rgb)=='FFFFFF00':
        yellow_row = True
    if c_val is None or not isinstance(c_val,(int,float)):
        cell.fill = GRIS
    elif not isinstance(u,(int,float)):
        cell.fill = GRIS
    elif c_val == u:
        cell.fill = VERDE
    elif c_val > u:
        cell.fill = AZUL_S
    else:
        cell.fill = ROJO

# ============ Actualizar Summary — agregar filas para nuevos embarques
print("\nAgregando nuevos embarques a Summary...")
ws_s = wb['Summary']
# Localizar fila TOTAL en bloque PLAN AJUSTADO
row_total = None
for r in range(1, ws_s.max_row+1):
    if ws_s.cell(row=r, column=1).value == 'TOTAL':
        row_total = r; break
if row_total:
    # Insertar filas antes del TOTAL para los nuevos embarques
    n_nuevos = len(nuevos)
    if n_nuevos > 0:
        ws_s.insert_rows(row_total, amount=n_nuevos)
        # Rellenar cada fila nueva
        DEADLINE = datetime(2026, 8, 31)
        for i, emb in enumerate(nuevos.keys()):
            r = row_total + i
            origen = 'NB' if 'NB' in emb.upper() else 'SZ'
            crd, etd, etap, etab = fechas(emb)
            a_tiempo = 'SÍ' if etab and etab <= DEADLINE else 'NO'
            # Fórmulas SUMIFS
            emb_ref = f'A{r}'
            formula_skus = (
                f'=COUNTIFS(Detail_SZ!N:N,{emb_ref})+COUNTIFS(Detail_NB!N:N,{emb_ref})'
                f'+COUNTIFS(Late_Arrivals!J:J,{emb_ref})+COUNTIFS(Pend_Confirmation!G:G,{emb_ref})'
                f'+COUNTIFS(Missing!G:G,{emb_ref})'
            )
            formula_uds = (
                f'=SUMIFS(Detail_SZ!M:M,Detail_SZ!N:N,{emb_ref})+SUMIFS(Detail_NB!M:M,Detail_NB!N:N,{emb_ref})'
                f'+SUMIFS(Late_Arrivals!I:I,Late_Arrivals!J:J,{emb_ref})'
                f'+SUMIFS(Pend_Confirmation!F:F,Pend_Confirmation!G:G,{emb_ref})'
                f'+SUMIFS(Missing!F:F,Missing!G:G,{emb_ref})'
            )
            vals = [emb, origen, crd, etd, etap, etab, formula_skus, formula_uds, None, a_tiempo]
            for j, v in enumerate(vals):
                cc = ws_s.cell(row=r, column=j+1, value=v)
                cc.font = NORMAL; cc.border = BORDER
                # Marcar fila nueva con fondo naranja suave
                cc.fill = NARANJA
                if j in (2,3,4,5): cc.number_format='DD-MM-YYYY'; cc.alignment=CENTER
                elif j in (6,7): cc.number_format='#,##0'; cc.alignment=RIGHT
                elif j == 9:
                    cc.alignment=CENTER; cc.font=Font(bold=True)
                    if v=='SÍ': cc.fill = VERDE
                    else: cc.fill = ROJO
                else: cc.alignment=LEFT if j==0 else CENTER
        # Actualizar fórmulas TOTAL para incluir las nuevas filas
        new_total_row = row_total + n_nuevos
        data_start = row_total - 6  # eran 6 embarques originales
        ws_s.cell(row=new_total_row, column=7, value=f'=SUM(G{data_start}:G{new_total_row-1})')
        ws_s.cell(row=new_total_row, column=8, value=f'=SUM(H{data_start}:H{new_total_row-1})')

# Guardar
wb.save(P)
print(f"\n[OK] Guardado: {P}")

# Resumen final
print("\n=== Resumen final ===")
print(f"SKUs Missing asignados: {sum(1 for r in asignaciones.values() if r)}")
print(f"SKUs Missing sin match (no en OHNSO): {sum(1 for r in asignaciones.values() if r is None)}")
print(f"Nuevos embarques: {list(nuevos.keys())}")
