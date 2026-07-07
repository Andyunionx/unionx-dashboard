"""Cruce v3 con 3 fixes:
  1. NO contar 'Rest items' como plan de carga (excluir del cálculo de sobra)
  2. CONSOLIDAR demanda UnionX por SKU (sumar todas las OTs)
  3. Match por SKU O por (Model + Descripción) para items sin SKU explícito
"""
import pandas as pd, sys, re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')

UX = "data/planillas/Shipping Plan September V.02.xlsx"
SV = "data/comex/Shipping Plan B-Jun.30th.xls"
OUT = "data/comex/Cruce_Shipping_Sep_vs_StevenB_v3.xlsx"

# Embarques VÁLIDOS de Steven (los que son plan de carga real, NO Rest items)
EMBARQUES_VALIDOS = {
    'SZ-04,40HQ', 'SZ-03-08,40HQ', 'SZ-01(623)', 'SZ-02(623)',
    'IMP OP 350-26 40HQ', 'NB-01',
    # También admitir variaciones de escritura
}

def es_embarque_valido(seccion):
    s = seccion.upper().replace(' ', '').replace('\n', '')
    for v in EMBARQUES_VALIDOS:
        vn = v.upper().replace(' ', '').replace('\n', '')
        if s == vn: return True
    # Match flexible
    if 'REST' in s or 'ITEM' in s: return False
    return True   # acepta todo lo que no sea Rest Items

def parsear_steven(sheet):
    df = pd.read_excel(SV, sheet_name=sheet, header=None)
    items = []
    seccion = None; hdr = None
    for i, row in df.iterrows():
        vals = [v for v in row.values if pd.notna(v)]
        if not vals: continue
        first = str(vals[0]).strip()
        m = re.match(r'^\d+\.(.+)$', first)
        if m and len(vals) <= 2:
            seccion = m.group(1).strip(); continue
        if first == 'No' and len(vals)>1 and 'Model' in str(vals[1]):
            hdr = [str(v).strip() for v in row.values]; continue
        try: int(first)
        except: continue
        if hdr is None or seccion is None: continue
        it = {col: row.values[j] for j,col in enumerate(hdr) if j<len(row.values)}
        it['__seccion'] = seccion
        items.append(it)
    return items

steven = parsear_steven('SHENZHEN') + parsear_steven('NINGBO')

# Construir 2 índices:
#   - por SKU exacto
#   - por (Model lower + primeras palabras de Descripcion lower)
def normalize_text(s):
    s = str(s or '').strip().lower()
    s = re.sub(r'\s+', ' ', s)
    return s

steven_por_sku = defaultdict(list)
steven_por_descripcion = []   # [(model_lower, desc_lower_norm, item)]

for it in steven:
    sku = str(it.get('SKU', '') or '').strip()
    qty = pd.to_numeric(it.get('Qty'), errors='coerce')
    if not pd.notna(qty): continue
    seccion = it['__seccion']
    es_valido = es_embarque_valido(seccion)
    model = normalize_text(it.get('Model'))
    desc = normalize_text(it.get('DESCRIPTON'))
    item = {
        'embarque': seccion,
        'es_valido': es_valido,
        'model': str(it.get('Model','')).strip(),
        'desc': str(it.get('DESCRIPTON','')).strip(),
        'qty': int(qty),
        'sku_raw': sku,
    }
    if sku and sku.lower() not in ('nan',''):
        steven_por_sku[sku].append(item)
    steven_por_descripcion.append((model, desc, item))

# UnionX V02 — CONSOLIDAR por SKU
ux_sz = pd.read_excel(UX, sheet_name='Detail_SZ')
ux_nb = pd.read_excel(UX, sheet_name='Detail_NB')
ux = pd.concat([ux_sz, ux_nb], ignore_index=True)
ux['SKU'] = ux['SKU'].astype(str).str.strip()
ux['Units'] = pd.to_numeric(ux['Units'], errors='coerce').fillna(0).astype(int)

# Por cada SKU: sumar units, listar OTs y contenedores
agrupado = ux.groupby('SKU').agg(
    units_total=('Units', 'sum'),
    descripcion=('Description', 'first'),
    brand=('Brand', 'first'),
    contenedores=('Container', lambda x: ' + '.join(f"{c}:{q}" for c, q in zip(x, ux.loc[x.index, 'Units']))),
    ots=('Order No', lambda x: ', '.join(str(o) for o in x.unique())),
).reset_index()

print(f"UnionX V02 — SKUs únicos: {len(agrupado)}")
print(f"UnionX V02 — filas totales (con duplicados por OT): {len(ux)}")
print()

# Cruce
rows = []
for _, r in agrupado.iterrows():
    sku = str(r['SKU']).strip()
    units_ux = int(r['units_total'])
    desc_ux = str(r['descripcion'])[:60]

    # 1) Match por SKU
    matches = steven_por_sku.get(sku, [])

    # 2) Si no match por SKU, intentar por descripción (fuzzy)
    matched_by_desc = False
    if not matches and len(desc_ux) > 5:
        desc_norm = normalize_text(desc_ux)
        # Tomar primeras 4 palabras significativas
        tokens_ux = [t for t in desc_norm.split() if len(t) >= 3][:5]
        for model_s, desc_s, item in steven_por_descripcion:
            if not desc_s: continue
            # Si todos los tokens UX están en desc Steven → match
            tokens_match = sum(1 for t in tokens_ux if t in desc_s)
            if tokens_match >= max(3, len(tokens_ux)*0.7):
                matches.append(item)
                matched_by_desc = True

    # Separar matches válidos (plan de carga) vs Rest items
    matches_validos = [m for m in matches if m['es_valido']]
    matches_rest    = [m for m in matches if not m['es_valido']]
    qty_validos = sum(m['qty'] for m in matches_validos)
    qty_rest    = sum(m['qty'] for m in matches_rest)

    # Detalle por embarque
    detalle_valido = ' | '.join(f"{m['embarque']}: {m['qty']}" for m in matches_validos) if matches_validos else '(ninguno)'
    detalle_rest   = ' | '.join(f"Rest: {m['qty']}" for m in matches_rest) if matches_rest else ''

    # Estado: comparar units_ux contra qty_validos
    if not matches:
        diff = -units_ux
        estado = '❌ NO EN PLAN STEVEN'
    elif not matches_validos:
        diff = -units_ux
        estado = '🔴 SOLO REST (no asignado)'
    else:
        diff = qty_validos - units_ux
        if diff == 0: estado = '✅ OK'
        elif diff > 0: estado = f'🟦 SOBRA {diff}'
        else: estado = f'🔴 FALTA {abs(diff)}'

    rows.append({
        'SKU': sku,
        'Brand': r['brand'],
        'Descripción': desc_ux,
        'OTs': r['ots'],
        'Cont UnionX (Units)': r['contenedores'],
        'Units UnionX (total)': units_ux,
        'Plan Steven (embarques)': detalle_valido,
        'Stock Rest items': detalle_rest if matches_rest else '—',
        'Qty Steven (plan)': qty_validos,
        'Δ (Plan - UX)': diff,
        'Estado': estado,
        'Match por': 'SKU' if not matched_by_desc else 'Descripción',
    })

df_res = pd.DataFrame(rows)
print(f"Total SKUs UnionX (consolidados): {len(df_res)}\n")

# Imprimir
for _, r in df_res.iterrows():
    print(f"{r['SKU']:<18} {r['Brand'][:8]:<8} UX={r['Units UnionX (total)']:>5} → {r['Estado']:<25} "
          f"| Plan: {r['Plan Steven (embarques)']:<55} | Stock: {r['Stock Rest items']}")

# Resumen
print(f"\n{'='*80}")
print("Resumen por estado")
print('='*80)
g = df_res.groupby(df_res['Estado'].str.split().str[0]).agg(
    n=('SKU','count'),
    units_ux=('Units UnionX (total)','sum'),
    plan_steven=('Qty Steven (plan)','sum'),
).reset_index()
print(g.to_string(index=False))

# === Excel
print(f"\nGenerando Excel...")
AZUL_H = PatternFill('solid', fgColor='FF1F3864')
ROJO   = PatternFill('solid', fgColor='FFE6B8B8')
AZUL_S = PatternFill('solid', fgColor='FFBDD7EE')
VERDE  = PatternFill('solid', fgColor='FFC6EFCE')
GRIS   = PatternFill('solid', fgColor='FFF2F2F2')
NARANJA = PatternFill('solid', fgColor='FFFCE4D6')
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
BLACK_B = Font(bold=True, size=11)
NORMAL = Font(size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cruce v3"
ws.sheet_view.showGridLines = False

ws.merge_cells('A1:L1')
ws['A1'] = 'CRUCE V3: Shipping Plan September V.02 vs Plan B Steven (Jun 30th)'
ws['A1'].fill = AZUL_H; ws['A1'].font = Font(color='FFFFFFFF', bold=True, size=14)
ws['A1'].alignment = CENTER
ws.row_dimensions[1].height = 28

ws.merge_cells('A2:L2')
ws['A2'] = (f'{len(df_res)} SKUs únicos UnionX (consolidados desde {len(ux)} filas con duplicados por OT)  |  '
            f'Plan Steven excluye Rest items (= stock no asignado)  |  Match por SKU y por Descripción')
ws['A2'].fill = GRIS; ws['A2'].font = Font(italic=True, size=10); ws['A2'].alignment = CENTER

headers = list(df_res.columns)
for j, h in enumerate(headers):
    c = ws.cell(row=4, column=j+1, value=h)
    c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

for i, (_, r) in enumerate(df_res.iterrows()):
    row = 5+i
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=j+1, value=r[h])
        c.font = NORMAL; c.border = BORDER
        if h in ('Units UnionX (total)','Qty Steven (plan)','Δ (Plan - UX)'):
            c.alignment = RIGHT
            if h == 'Δ (Plan - UX)': c.number_format = '+#,##0;-#,##0;0'
            else: c.number_format = '#,##0'
        elif h == 'Estado':
            c.alignment = CENTER; c.font = BLACK_B
            est = str(r[h])
            if 'NO EN PLAN' in est or 'SOLO REST' in est or 'FALTA' in est: c.fill = ROJO
            elif 'SOBRA' in est: c.fill = AZUL_S
            elif '✅' in est: c.fill = VERDE
        elif h == 'Stock Rest items' and r[h] != '—':
            c.alignment = LEFT
            c.fill = NARANJA
        elif h == 'Match por' and r[h] == 'Descripción':
            c.alignment = CENTER
            c.fill = NARANJA
        else:
            c.alignment = LEFT if h in ('Descripción','Plan Steven (embarques)','Stock Rest items','Cont UnionX (Units)','OTs') else CENTER

widths = {'A':17,'B':9,'C':52,'D':14,'E':16,'F':10,'G':40,'H':22,'I':10,'J':10,'K':22,'L':12}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = 'A5'
wb.save(OUT)
print(f"[OK] Excel: {OUT}")
