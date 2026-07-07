"""Actualiza Shipping Plan September V.02 con datos del Plan B Steven Jun.30th:
  - Columna "Cantidad cargada" (qty real plan Steven, por OT)
  - Columna "Embarque" (en qué embarque viaja según Steven)
  - Columnas CRD / ETD / ETA Puerto / ETA WH (calculadas según embarque)

Reglas:
  - Match por SKU exacto
  - Special case: "HD4 Khaki" = TCMULTSTY5N1-BG
  - Excluir Rest items (no es plan de carga)
  - Asignación greedy por qty exacta cuando hay múltiples matches del mismo SKU
"""
import shutil, sys, re
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

UX_ORIG = Path("data/planillas/Shipping Plan September V.02.xlsx")
SV      = Path("data/comex/Shipping Plan B-Jun.30th.xls")
OUT     = Path("data/planillas/Shipping Plan September V.02_actualizado.xlsx")

HOY = datetime(2026, 6, 30)
TRANSITO = {'SZ': 52, 'NB': 45}
DIAS_CRD_ETD = 7
DIAS_PTO_BODEGA = 7

# Match especial: SKU UnionX -> Model Steven (cuando Steven no usa el SKU)
MATCHES_ESPECIALES = {
    'TCMULTSTY5N1-BG': {'model_steven': 'HD4 Khaki'},
}


def parsear_steven(sheet):
    """Devuelve lista de items con seccion + finish_time."""
    df = pd.read_excel(SV, sheet_name=sheet, header=None)
    items = []
    seccion = None; hdr = None
    for i, row in df.iterrows():
        vals = [v for v in row.values if pd.notna(v)]
        if not vals: continue
        first = str(vals[0]).strip()
        m = re.match(r'^\d+\.(.+)$', first)
        if m and len(vals) <= 2:
            seccion = m.group(1).strip(); continue
        if first == 'No' and len(vals)>1 and 'Model' in str(vals[1]):
            hdr = [str(v).strip() for v in row.values]; continue
        try: int(first)
        except: continue
        if hdr is None or seccion is None: continue
        it = {col: row.values[j] for j,col in enumerate(hdr) if j<len(row.values)}
        it['__seccion'] = seccion
        items.append(it)
    return items


def es_rest(seccion):
    return 'rest' in seccion.lower() or 'item' in seccion.lower() and 'rest' in seccion.lower()


def parsear_finish_time(ft):
    if pd.isna(ft): return HOY  # default = hoy
    if isinstance(ft, datetime): return ft
    s = str(ft).strip().lower()
    if s in ('ready','reday'): return HOY
    try: return pd.to_datetime(s).to_pydatetime()
    except: return HOY


def puerto_de_embarque(embarque):
    e = embarque.upper()
    if e.startswith('SZ-') or 'SHENZHEN' in e: return 'SZ'
    if e.startswith('NB-') or 'NINGBO' in e: return 'NB'
    if 'IMP OP 350' in e: return 'NB'  # IMP OP 350-26 está en Ningbo
    return 'SZ'


def calcular_fechas_embarque(crd):
    """Dado CRD, devolver ETD, ETA Puerto, ETA WH según puerto."""
    return crd  # placeholder, se computa abajo según embarque


# Parsear Steven
steven_sz = parsear_steven('SHENZHEN')
steven_nb = parsear_steven('NINGBO')
steven_todos = steven_sz + steven_nb

# Calcular CRD/ETD/ETA por embarque (max Finish Time del embarque)
embarque_info = {}
for it in steven_todos:
    seccion = it['__seccion']
    if es_rest(seccion): continue
    ft = parsear_finish_time(it.get('Finish Time'))
    if seccion not in embarque_info:
        embarque_info[seccion] = {'crd_max': ft, 'items': []}
    if ft > embarque_info[seccion]['crd_max']:
        embarque_info[seccion]['crd_max'] = ft
    embarque_info[seccion]['items'].append(it)

# Calcular fechas por embarque
for emb, info in embarque_info.items():
    puerto = puerto_de_embarque(emb)
    crd = info['crd_max']
    etd = crd + timedelta(days=DIAS_CRD_ETD)
    eta_pto = etd + timedelta(days=TRANSITO[puerto])
    eta_wh  = eta_pto + timedelta(days=DIAS_PTO_BODEGA)
    info['puerto'] = puerto
    info['CRD'] = crd
    info['ETD'] = etd
    info['ETA_Puerto'] = eta_pto
    info['ETA_WH'] = eta_wh

print("=== FECHAS POR EMBARQUE ===")
print(f"{'Embarque':<22} {'Puerto':<7} {'CRD':<12} {'ETD':<12} {'ETA Pto':<12} {'ETA WH':<12} {'SKUs':>5}")
for emb in sorted(embarque_info.keys()):
    info = embarque_info[emb]
    print(f"{emb:<22} {info['puerto']:<7} {info['CRD']:%d-%m-%Y}  {info['ETD']:%d-%m-%Y}  {info['ETA_Puerto']:%d-%m-%Y}  {info['ETA_WH']:%d-%m-%Y}  {len(info['items']):>5}")

# === Construir índice Steven (solo embarques válidos)
# por SKU
from collections import defaultdict
steven_idx = defaultdict(list)
for emb, info in embarque_info.items():
    for it in info['items']:
        sku = str(it.get('SKU','') or '').strip()
        model = str(it.get('Model','') or '').strip()
        qty = pd.to_numeric(it.get('Qty'), errors='coerce')
        if not pd.notna(qty): continue
        entry = {'embarque': emb, 'model': model, 'qty': int(qty),
                 'CRD': info['CRD'], 'ETD': info['ETD'],
                 'ETA_Puerto': info['ETA_Puerto'], 'ETA_WH': info['ETA_WH'],
                 'sku_raw': sku, 'used': False}
        if sku: steven_idx[sku].append(entry)
        # también indexar por model (para matches especiales)
        if model: steven_idx[f"__model__::{model}"].append(entry)

# === Copiar archivo y abrir
shutil.copy(UX_ORIG, OUT)
wb = openpyxl.load_workbook(OUT)

# Asignar embarque a cada fila UX (greedy: match exacto qty primero)
def buscar_match(sku_ux, qty_ux, model_steven_esperado=None):
    """Busca match Steven para esta fila. Marca used=True."""
    candidatos = steven_idx.get(sku_ux, [])

    # Si tiene match especial por model_steven
    if model_steven_esperado:
        candidatos = candidatos + steven_idx.get(f"__model__::{model_steven_esperado}", [])

    # Filtrar no usados
    libres = [c for c in candidatos if not c['used']]

    if not libres:
        # Permitir reuso si solo hay 1 (caso UX duplicado vs Steven único)
        if candidatos:
            return candidatos[0]
        return None

    # 1) Match exacto qty
    exactos = [c for c in libres if c['qty'] == qty_ux]
    if exactos:
        exactos[0]['used'] = True
        return exactos[0]
    # 2) Match qty >= qty_ux (Steven carga al menos lo pedido)
    suficientes = [c for c in libres if c['qty'] >= qty_ux]
    if suficientes:
        # Tomar el de menor sobra
        suficientes.sort(key=lambda c: c['qty'])
        suficientes[0]['used'] = True
        return suficientes[0]
    # 3) Cualquiera
    libres[0]['used'] = True
    return libres[0]


# Encabezados nuevos
NUEVAS_COLS = ['Cantidad cargada', 'Embarque (Steven)', 'CRD', 'ETD', 'ETA Puerto', 'ETA Bodega']

AZUL_H = PatternFill('solid', fgColor='FF1F3864')
ROJO   = PatternFill('solid', fgColor='FFE6B8B8')
AZUL_S = PatternFill('solid', fgColor='FFBDD7EE')
VERDE  = PatternFill('solid', fgColor='FFC6EFCE')
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
NORMAL = Font(size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right', vertical='center')
LEFT   = Alignment(horizontal='left', vertical='center')
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def actualizar_sheet(ws_name):
    ws = wb[ws_name]
    # Header en fila 1 (col A..L = 12 cols originales)
    # Agregar columnas M..R
    col_inicio = 13   # M
    for j, h in enumerate(NUEVAS_COLS):
        c = ws.cell(row=1, column=col_inicio + j, value=h)
        c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

    # Procesar filas (desde fila 2 en adelante)
    n_filas = ws.max_row
    for row in range(2, n_filas + 1):
        sku = str(ws.cell(row=row, column=6).value or '').strip()   # F = SKU
        units_raw = ws.cell(row=row, column=8).value                # H = Units
        if not sku or units_raw is None: continue
        try: units = int(units_raw)
        except: continue

        # Match especial
        model_steven = MATCHES_ESPECIALES.get(sku, {}).get('model_steven')
        match = buscar_match(sku, units, model_steven)

        if match:
            vals = [match['qty'], match['embarque'], match['CRD'], match['ETD'],
                    match['ETA_Puerto'], match['ETA_WH']]
            for j, v in enumerate(vals):
                c = ws.cell(row=row, column=col_inicio + j, value=v)
                c.font = NORMAL; c.border = BORDER
                if j == 0:
                    c.alignment = RIGHT; c.number_format = '#,##0'
                    # Pintar según comparación qty
                    if match['qty'] == units: c.fill = VERDE
                    elif match['qty'] > units: c.fill = AZUL_S
                    else: c.fill = ROJO
                elif j == 1:
                    c.alignment = CENTER
                else:
                    c.alignment = CENTER; c.number_format = 'DD-MM-YYYY'
        else:
            # Sin match
            c = ws.cell(row=row, column=col_inicio, value=0)
            c.fill = ROJO; c.font = NORMAL; c.alignment = RIGHT
            c = ws.cell(row=row, column=col_inicio + 1, value='❌ NO EN PLAN STEVEN')
            c.fill = ROJO; c.font = NORMAL; c.alignment = CENTER

    # Anchos cols nuevas
    for j, _ in enumerate(NUEVAS_COLS):
        ws.column_dimensions[get_column_letter(col_inicio + j)].width = 16

actualizar_sheet('Detail_SZ')
actualizar_sheet('Detail_NB')

# === Crear hoja resumen "Fechas Cargas"
if 'Fechas Cargas' in wb.sheetnames:
    del wb['Fechas Cargas']
ws_r = wb.create_sheet('Fechas Cargas', 0)
ws_r.sheet_view.showGridLines = False

ws_r.merge_cells('A1:H1')
ws_r['A1'] = 'RESUMEN DE FECHAS DE CARGA — Shipping Plan September V.02 vs Plan B Steven (Jun 30th)'
ws_r['A1'].fill = AZUL_H; ws_r['A1'].font = Font(color='FFFFFFFF', bold=True, size=14)
ws_r['A1'].alignment = CENTER
ws_r.row_dimensions[1].height = 28

ws_r.merge_cells('A2:H2')
ws_r['A2'] = f'Generado: {HOY:%d-%m-%Y}  |  Tránsito Shenzhen 52d, Ningbo 45d  |  CRD→ETD 7d  |  ETA Puerto→Bodega 7d'
ws_r['A2'].font = Font(italic=True, size=10); ws_r['A2'].alignment = CENTER

H = ['Embarque', 'Puerto', 'SKUs', 'Qty total', 'CRD', 'ETD', 'ETA Puerto', 'ETA Bodega']
for j, h in enumerate(H):
    c = ws_r.cell(row=4, column=j+1, value=h)
    c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

# Datos
embarques_ordenados = sorted(embarque_info.keys(), key=lambda e: (embarque_info[e]['ETA_WH'], e))
for i, emb in enumerate(embarques_ordenados):
    info = embarque_info[emb]
    qty_total = sum(pd.to_numeric(it.get('Qty'), errors='coerce') for it in info['items']
                    if pd.notna(pd.to_numeric(it.get('Qty'), errors='coerce')))
    vals = [emb, info['puerto'], len(info['items']), int(qty_total),
            info['CRD'], info['ETD'], info['ETA_Puerto'], info['ETA_WH']]
    for j, v in enumerate(vals):
        c = ws_r.cell(row=5+i, column=j+1, value=v)
        c.font = NORMAL; c.border = BORDER
        if j in (4,5,6,7): c.number_format = 'DD-MM-YYYY'; c.alignment = CENTER
        elif j in (2, 3): c.alignment = RIGHT; c.number_format = '#,##0'
        else: c.alignment = CENTER if j != 0 else LEFT

# Anchos
for col, w in {'A':22,'B':8,'C':7,'D':10,'E':13,'F':13,'G':13,'H':13}.items():
    ws_r.column_dimensions[col].width = w

wb.save(OUT)
print(f"\n[OK] Archivo actualizado: {OUT}")

# Resumen final
print("\nSKUs procesados:")
total_ok = 0; total_sobra = 0; total_no = 0
for ws_name in ['Detail_SZ', 'Detail_NB']:
    ws = wb[ws_name]
    for row in range(2, ws.max_row+1):
        sku = ws.cell(row=row, column=6).value
        units = ws.cell(row=row, column=8).value
        emb_steven = ws.cell(row=row, column=14).value
        qty_steven = ws.cell(row=row, column=13).value
        if sku and units:
            if not emb_steven or 'NO' in str(emb_steven): total_no += 1
            elif qty_steven == units: total_ok += 1
            else: total_sobra += 1
print(f"  OK exacto:        {total_ok}")
print(f"  Con sobra/falta:  {total_sobra}")
print(f"  Sin match:        {total_no}")
