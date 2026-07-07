"""V3 fixes:
  1. Detail_SZ/NB — fill duro inicial (visible desde apertura) + CF con colores más intensos
  2. Summary — fórmulas SUMIFS/COUNTIFS en columnas SKUs/Uds UnionX
  3. Missing amarillos: Embarque = NB-01 (no Rest items)
"""
import openpyxl, sys, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
sys.stdout.reconfigure(encoding='utf-8')

P = "data/planillas/Shipping Plan September V.02_actualizado.xlsx"

# Paletas más intensas
VERDE  = PatternFill('solid', fgColor='FF92D050')   # verde vivo
AZUL_S = PatternFill('solid', fgColor='FF9DC3E6')   # azul sobra
ROJO   = PatternFill('solid', fgColor='FFF08080')   # rojo falta
GRIS   = PatternFill('solid', fgColor='FFD9D9D9')   # gris vacío
AZUL_H = PatternFill('solid', fgColor='FF1F3864')
AMARI  = PatternFill('solid', fgColor='FFFFFF00')
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
NORMAL = Font(size=10)
BOLD = Font(bold=True, size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.load_workbook(P)

# =========================================================
# 1) DETAIL_SZ + DETAIL_NB — fill duro + CF
# =========================================================
print("=== Detail_SZ + Detail_NB: fill duro + CF reforzado ===")
for sh in ['Detail_SZ', 'Detail_NB']:
    ws = wb[sh]
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    col_units = hdr['Units']
    col_cant  = hdr['Cantidad cargada']
    u_l = get_column_letter(col_units)
    c_l = get_column_letter(col_cant)
    last = ws.max_row

    # Limpiar CF previas
    ws.conditional_formatting._cf_rules = {}

    # Aplicar CF nuevas
    rng = f'{c_l}2:{c_l}{last}'
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2={u_l}2)'], fill=VERDE))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2>{u_l}2)'], fill=AZUL_S))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2<{u_l}2,{c_l}2>0)'], fill=ROJO))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR({c_l}2="",ISBLANK({c_l}2))'], fill=GRIS))

    # Fill duro inicial según estado actual
    verde_n = azul_n = rojo_n = gris_n = 0
    for r in range(2, last+1):
        u = ws.cell(row=r, column=col_units).value
        c = ws.cell(row=r, column=col_cant).value
        cell = ws.cell(row=r, column=col_cant)
        if c is None or c == '' or not isinstance(c, (int, float)):
            cell.fill = GRIS; gris_n += 1
        elif not isinstance(u, (int, float)):
            cell.fill = GRIS; gris_n += 1
        elif c == u:
            cell.fill = VERDE; verde_n += 1
        elif c > u:
            cell.fill = AZUL_S; azul_n += 1
        else:
            cell.fill = ROJO; rojo_n += 1
    print(f"  [{sh}] fills iniciales: verde={verde_n} azul={azul_n} rojo={rojo_n} gris={gris_n}")

# =========================================================
# 2) SUMMARY — fórmulas SUMIFS/COUNTIFS
# =========================================================
print("\n=== Summary: fórmulas dinámicas ===")
ws = wb['Summary']

# Localizar bloque "PLAN AJUSTADO" — busco fila con ese texto
row_plan = None
for r in range(1, ws.max_row+1):
    if ws.cell(row=r, column=1).value and 'PLAN AJUSTADO' in str(ws.cell(row=r, column=1).value):
        row_plan = r; break
print(f"  Bloque PLAN AJUSTADO en fila {row_plan}")

# Estructura columnas (según V2): A=Embarque B=Origen C=CRD D=ETD E=ETAPto F=ETABod G=SKUs H=UdsTotales I=UdsUX J=ATiempo
# Header en row_plan+1, datos empiezan row_plan+2
data_start = row_plan + 2

# Cambiar el header G a "SKUs UnionX" y H a "Uds UnionX V02" (elimino "Uds totales" porque es dato externo)
# Actualizar headers
ws.cell(row=row_plan+1, column=7, value='SKUs UnionX V02')
ws.cell(row=row_plan+1, column=8, value='Uds UnionX V02')
# Ocultar columna I (que tenía Uds UnionX V02 antes) — mejor la reemplazo con "¿A tiempo?"
# Simplifico: G=SKUs UnionX, H=Uds UnionX, I=¿A tiempo? (era J)

# Leer los 6 embarques
embarques_ord = ['SZ-04,40HQ', 'SZ-03-08,40HQ', 'SZ-01(623)', 'SZ-02(623)', 'IMP OP 350-26 40HQ', 'NB-01']
DEADLINE = datetime(2026, 8, 31)

for i, e in enumerate(embarques_ord):
    r = data_start + i
    emb_cell_addr = get_column_letter(1) + str(r)  # A{r}
    # Fórmulas SUMIFS / COUNTIFS sobre TODAS las hojas de demanda UnionX
    # Detail_SZ: SKU en F, Cantidad cargada en M, Embarque en N
    # Detail_NB: idem
    # Late_Arrivals: SKU B, Units E, Cargada F (9), Embarque G (10)  → verificar cols reales
    # Pend_Confirmation: SKU B, Units D, Cargada F, Embarque G
    # Missing: SKU B, Units D, Cargada F, Embarque G
    # Fórmula uds UnionX: sumar Cantidad cargada donde Embarque == este embarque
    formula_uds = (
        f'=SUMIFS(Detail_SZ!M:M,Detail_SZ!N:N,{emb_cell_addr})'
        f'+SUMIFS(Detail_NB!M:M,Detail_NB!N:N,{emb_cell_addr})'
        f'+SUMIFS(Late_Arrivals!I:I,Late_Arrivals!J:J,{emb_cell_addr})'
        f'+SUMIFS(Pend_Confirmation!F:F,Pend_Confirmation!G:G,{emb_cell_addr})'
        f'+SUMIFS(Missing!F:F,Missing!G:G,{emb_cell_addr})'
    )
    formula_skus = (
        f'=COUNTIFS(Detail_SZ!N:N,{emb_cell_addr})'
        f'+COUNTIFS(Detail_NB!N:N,{emb_cell_addr})'
        f'+COUNTIFS(Late_Arrivals!J:J,{emb_cell_addr})'
        f'+COUNTIFS(Pend_Confirmation!G:G,{emb_cell_addr})'
        f'+COUNTIFS(Missing!G:G,{emb_cell_addr})'
    )
    ws.cell(row=r, column=7, value=formula_skus)   # G
    ws.cell(row=r, column=8, value=formula_uds)    # H
    # Limpiar antigua col I (uds UnionX que estaba ahí) y mover ¿A tiempo? a col I
    ws.cell(row=r, column=9, value=None)
    # ¿A tiempo? sigue en col J (10)

    # Formato
    ws.cell(row=r, column=7).number_format = '#,##0'; ws.cell(row=r, column=7).alignment = RIGHT
    ws.cell(row=r, column=8).number_format = '#,##0'; ws.cell(row=r, column=8).alignment = RIGHT

# Header col I (limpiarlo — antes decía "Uds UnionX V02", ahora esa info está en H)
ws.cell(row=row_plan+1, column=9, value='')

# Recalcular TOTAL row
total_row = data_start + len(embarques_ord)
if ws.cell(row=total_row, column=1).value == 'TOTAL':
    ws.cell(row=total_row, column=7, value=f'=SUM(G{data_start}:G{data_start+len(embarques_ord)-1})')
    ws.cell(row=total_row, column=8, value=f'=SUM(H{data_start}:H{data_start+len(embarques_ord)-1})')
    ws.cell(row=total_row, column=9, value=None)
    ws.cell(row=total_row, column=7).number_format='#,##0'; ws.cell(row=total_row, column=7).font=Font(bold=True)
    ws.cell(row=total_row, column=8).number_format='#,##0'; ws.cell(row=total_row, column=8).font=Font(bold=True)

print(f"  Fórmulas SUMIFS/COUNTIFS aplicadas en filas {data_start}-{data_start+5}")

# =========================================================
# 3) MISSING amarillos: Embarque = NB-01
# =========================================================
print("\n=== Missing: amarillos LHSHNUCO → NB-01 ===")
ws = wb['Missing']
last_col = ws.max_column
# Columnas cruce: F=Cantidad cargada, G=Embarque, H=CRD, I=ETD, J=ETA Pto, K=ETA Bod
# Verificar headers
hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
print(f"  headers Missing: {hdr}")

# Detectar amarillos
amarillos = []
for r in range(2, ws.max_row+1):
    f = ws.cell(row=r, column=2).fill
    if f and f.start_color and str(f.start_color.rgb)=='FFFFFF00':
        amarillos.append(r)
print(f"  amarillos: {amarillos}")

# Fechas NB-01
CRD_NB = datetime(2026, 7, 20)
ETD_NB = CRD_NB + timedelta(days=7)
ETAPTO_NB = ETD_NB + timedelta(days=45)
ETABOD_NB = ETAPTO_NB + timedelta(days=7)
print(f"  NB-01 fechas: CRD={CRD_NB.date()} ETD={ETD_NB.date()} ETA_Pto={ETAPTO_NB.date()} ETA_Bod={ETABOD_NB.date()}")

# Cambiar embarque + fechas
col_cant = hdr.get('Cantidad cargada')
col_emb  = hdr.get('Embarque (Steven)')
col_crd  = hdr.get('CRD')
col_etd  = hdr.get('ETD')
col_etap = hdr.get('ETA Puerto')
col_etab = hdr.get('ETA Bodega')

for r in amarillos:
    ws.cell(row=r, column=col_emb, value='NB-01').alignment = CENTER
    for col, val in [(col_crd, CRD_NB), (col_etd, ETD_NB), (col_etap, ETAPTO_NB), (col_etab, ETABOD_NB)]:
        c = ws.cell(row=r, column=col, value=val)
        c.number_format = 'DD-MM-YYYY'; c.alignment = CENTER; c.border = BORDER; c.font = NORMAL

# =========================================================
# Guardar
# =========================================================
wb.save(P)
print(f"\n[OK] Guardado: {P}")
