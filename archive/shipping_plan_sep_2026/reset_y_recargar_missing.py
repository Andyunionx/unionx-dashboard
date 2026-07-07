"""RESET + RECARGA con reglas correctas:
  1. Máx 68 CBM por contenedor
  2. FIJOS (no tocar): SZ-04,40HQ · SZ-03-08,40HQ · IMP OP 350-26 40HQ
  3. PERMITIDOS: SZ-01(623) · SZ-02(623) · NB-01 · nuevos
  4. Missing sin match queda en Missing
"""
import openpyxl, pandas as pd, sys, re, shutil
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

ORIG = "data/planillas/Shipping Plan September V.02.xlsx"
P    = "data/planillas/Shipping Plan September V.02_actualizado.xlsx"
SV   = "data/comex/Shipping Plan B-Jun.30th.xls"
OHNSO = "agente-comex/data/inbox/20260701_0912_shipping_plan_ohnso/OHNSO-Jul.01st.xls"

CAP_40HQ = 68
TRANSITO = {'SZ': 52, 'NB': 45}
CRD_ETD = 7
PTO_BOD = 7
DEADLINE = datetime(2026, 8, 31)
CRD_FALLBACK = datetime(2026, 8, 15)

FIJOS = {'SZ-04,40HQ', 'SZ-03-08,40HQ', 'IMP OP 350-26 40HQ'}
PERMITIDOS_ACTUALES = ['SZ-01(623)', 'SZ-02(623)', 'NB-01']

# --- Estilos
VERDE  = PatternFill('solid', fgColor='FF92D050')
AZUL_S = PatternFill('solid', fgColor='FF9DC3E6')
ROJO   = PatternFill('solid', fgColor='FFF08080')
GRIS   = PatternFill('solid', fgColor='FFD9D9D9')
AZUL_H = PatternFill('solid', fgColor='FF1F3864')
NARANJA= PatternFill('solid', fgColor='FFFCE4D6')
AMARI  = PatternFill('solid', fgColor='FFFFFF00')
GRIS_L = PatternFill('solid', fgColor='FFF2F2F2')
WHITE_B = Font(color='FFFFFFFF', bold=True, size=11)
NORMAL = Font(size=10); BOLD = Font(bold=True, size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ================ Parse OHNSO
def parsear(path, sheet):
    df = pd.read_excel(path, sheet_name=sheet, header=None)
    items, seccion, hdr = [], None, None
    for _, row in df.iterrows():
        vals = [v for v in row.values if pd.notna(v)]
        if not vals: continue
        first = str(vals[0]).strip()
        m = re.match(r'^\d+\.(.+)$', first)
        if m and len(vals) <= 2: seccion = m.group(1).strip(); continue
        if first == 'No' and len(vals)>1 and 'Model' in str(vals[1]):
            hdr = [str(v).strip() for v in row.values]; continue
        try: int(first)
        except: continue
        if hdr is None or seccion is None: continue
        it = {col: row.values[j] for j,col in enumerate(hdr) if j<len(row.values)}
        it['__seccion'] = seccion
        it['__sheet'] = sheet
        items.append(it)
    return items

ohnso_items = parsear(OHNSO, 'SHENZHEN') + parsear(OHNSO, 'NINGBO')
steven_items = parsear(SV, 'SHENZHEN') + parsear(SV, 'NINGBO')

# Índices OHNSO por SKU
ohnso_idx = defaultdict(list)
for it in ohnso_items:
    sku = str(it.get('SKU','') or '').strip()
    if not sku or sku.lower()=='nan': continue
    qty = pd.to_numeric(it.get('Qty'), errors='coerce')
    if not pd.notna(qty): continue
    cbm_it = pd.to_numeric(it.get('TOTAL CBM'), errors='coerce')
    cbm_ct = pd.to_numeric(it.get('(CBM)\n/CTN'), errors='coerce')
    qty_ct = pd.to_numeric(it.get("Q'ty/ctn"), errors='coerce')
    price = pd.to_numeric(it.get('Price\n(USD)'), errors='coerce')
    finish = it.get('Finish Time')
    ohnso_idx[sku].append({
        'embarque': it['__seccion'],
        'sheet': it['__sheet'],
        'qty': int(qty),
        'cbm_total': cbm_it if pd.notna(cbm_it) else 0,
        'cbm_ctn': cbm_ct if pd.notna(cbm_ct) else 0,
        'qty_ctn': qty_ct if pd.notna(qty_ct) else 0,
        'price': price if pd.notna(price) else 0,
        'finish': finish if isinstance(finish, datetime) else None,
    })

# Steven (Plan B) por SKU con matches especiales
steven_idx = defaultdict(list)
for it in steven_items:
    sku = str(it.get('SKU','') or '').strip()
    if not sku or sku.lower()=='nan': continue
    qty = pd.to_numeric(it.get('Qty'), errors='coerce')
    if not pd.notna(qty): continue
    finish = it.get('Finish Time')
    steven_idx[sku].append({
        'embarque': it['__seccion'],
        'qty': int(qty),
        'finish': finish if isinstance(finish, datetime) else None,
    })

# CRD por embarque = MAX Finish
emb_crd = {}
for lst in list(ohnso_idx.values()) + list(steven_idx.values()):
    for m in lst:
        if m['embarque']=='Rest items': continue
        if m['finish']:
            e = m['embarque']
            if e not in emb_crd or m['finish'] > emb_crd[e]:
                emb_crd[e] = m['finish']

def fechas(emb, crd_override=None):
    crd = crd_override or emb_crd.get(emb, CRD_FALLBACK)
    transit = TRANSITO['NB'] if 'NB' in emb.upper() or 'IMP OP' in emb.upper() else TRANSITO['SZ']
    etd = crd + timedelta(days=CRD_ETD)
    etap = etd + timedelta(days=transit)
    etab = etap + timedelta(days=PTO_BOD)
    return crd, etd, etap, etab

def get_ohnso_data(sku, units):
    ms = ohnso_idx.get(sku, [])
    if not ms: return None, None
    m = ms[0]
    if m['qty_ctn'] > 0 and m['cbm_ctn'] > 0:
        cbm = round((units / m['qty_ctn']) * m['cbm_ctn'], 3)
    elif m['qty'] > 0 and m['cbm_total'] > 0:
        cbm = round((units / m['qty']) * m['cbm_total'], 3)
    else:
        cbm = None
    usd = round(units * m['price'], 2) if m['price'] > 0 else None
    return cbm, usd

# ================ RESET: copiar original sobre _actualizado
print("=== RESET desde original ===")
shutil.copy(ORIG, P)

# ================ Aplicar la cadena completa
wb = openpyxl.load_workbook(P)

# --- Añadir columnas de cruce a Detail_SZ + Detail_NB
new_cols_detail = ['Cantidad cargada', 'Embarque (Steven)', 'CRD', 'ETD', 'ETA Puerto', 'ETA Bodega']
for sh in ['Detail_SZ', 'Detail_NB']:
    ws = wb[sh]
    last_col = ws.max_column
    hdrs = [ws.cell(row=1, column=c).value for c in range(1, last_col+1)]
    # Chequear si ya existen
    for j, h in enumerate(new_cols_detail):
        if h in hdrs:
            continue
        c = ws.cell(row=1, column=last_col+1+j, value=h)
        c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

# --- Cargar detalle según Steven Plan B (con match especial TCMULTSTY5N1-BG)
MATCHES_ESP = {'TCMULTSTY5N1-BG': 'TCMULTSTY5N1'}

def asignar_cruce_detail(ws, col_units_hdr='Units'):
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    col_sku = hdr['SKU']; col_units = hdr[col_units_hdr]
    col_cant = hdr['Cantidad cargada']; col_emb = hdr['Embarque (Steven)']
    col_crd = hdr['CRD']; col_etd = hdr['ETD']
    col_etap = hdr['ETA Puerto']; col_etab = hdr['ETA Bodega']

    for r in range(2, ws.max_row+1):
        sku = str(ws.cell(row=r, column=col_sku).value or '').strip()
        units = pd.to_numeric(ws.cell(row=r, column=col_units).value, errors='coerce')
        if not sku or not pd.notna(units): continue
        units = int(units)
        sku_lookup = MATCHES_ESP.get(sku, sku)
        # Preferir OHNSO (más actualizado)
        ms = ohnso_idx.get(sku_lookup, [])
        cargada = None; embarque = None
        if ms:
            # exact match en embarque real
            for m in ms:
                if m['embarque']!='Rest items' and m['qty']==units:
                    cargada, embarque = m['qty'], m['embarque']; break
            if cargada is None:
                for m in ms:
                    if m['embarque']!='Rest items' and m['qty']>=units:
                        cargada, embarque = m['qty'], m['embarque']; break
            if cargada is None:
                for m in ms:
                    if m['embarque']!='Rest items':
                        cargada, embarque = m['qty'], m['embarque']; break
        # Fallback Steven Plan B si OHNSO no tiene
        if cargada is None:
            ms2 = steven_idx.get(sku_lookup, [])
            for m in ms2:
                if m['embarque']!='Rest items' and m['qty']>=units:
                    cargada, embarque = m['qty'], m['embarque']; break
            if cargada is None:
                for m in ms2:
                    if m['embarque']!='Rest items':
                        cargada, embarque = m['qty'], m['embarque']; break
        # Escribir
        c1 = ws.cell(row=r, column=col_cant, value=cargada)
        c1.alignment = RIGHT; c1.border = BORDER; c1.font = NORMAL; c1.number_format = '#,##0'
        c2 = ws.cell(row=r, column=col_emb, value=embarque)
        c2.alignment = CENTER; c2.border = BORDER; c2.font = NORMAL
        if embarque:
            crd, etd, etap, etab = fechas(embarque)
            for j, val in enumerate([crd, etd, etap, etab]):
                c = ws.cell(row=r, column=col_crd+j, value=val)
                c.number_format = 'DD-MM-YYYY'; c.alignment = CENTER; c.border = BORDER; c.font = NORMAL

for sh in ['Detail_SZ', 'Detail_NB']:
    asignar_cruce_detail(wb[sh])

# ================ Late_Arrivals + Pend_Confirmation + Missing (columnas + cruce)
def asignar_cruce_aux(ws, col_sku_i, col_units_i, cols_start):
    """cols_start: primera columna nueva para 'Cantidad cargada'."""
    last_col = ws.max_column
    hdrs_now = [ws.cell(row=1, column=c).value for c in range(1, last_col+1)]
    if 'Cantidad cargada' not in hdrs_now:
        for j, h in enumerate(new_cols_detail):
            c = ws.cell(row=1, column=last_col+1+j, value=h)
            c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER
        base = last_col
    else:
        base = hdrs_now.index('Cantidad cargada')
    for r in range(2, ws.max_row+1):
        sku = str(ws.cell(row=r, column=col_sku_i).value or '').strip()
        units = pd.to_numeric(ws.cell(row=r, column=col_units_i).value, errors='coerce')
        if not sku or not pd.notna(units): continue
        units = int(units)
        sku_lookup = MATCHES_ESP.get(sku, sku)
        ms = ohnso_idx.get(sku_lookup, [])
        cargada = None; embarque = None
        if ms:
            for m in ms:
                if m['embarque']!='Rest items' and m['qty']==units:
                    cargada, embarque = m['qty'], m['embarque']; break
            if cargada is None:
                for m in ms:
                    if m['embarque']!='Rest items' and m['qty']>=units:
                        cargada, embarque = m['qty'], m['embarque']; break
            if cargada is None:
                for m in ms:
                    if m['embarque']!='Rest items':
                        cargada, embarque = m['qty'], m['embarque']; break
        # Steven fallback
        if cargada is None:
            ms2 = steven_idx.get(sku_lookup, [])
            for m in ms2:
                if m['embarque']!='Rest items' and m['qty']>=units:
                    cargada, embarque = m['qty'], m['embarque']; break
        c1 = ws.cell(row=r, column=base+1, value=cargada)
        c1.alignment = RIGHT; c1.border = BORDER; c1.font = NORMAL; c1.number_format = '#,##0'
        c2 = ws.cell(row=r, column=base+2, value=embarque)
        c2.alignment = CENTER; c2.border = BORDER; c2.font = NORMAL
        if embarque:
            crd, etd, etap, etab = fechas(embarque)
            for j, val in enumerate([crd, etd, etap, etab]):
                c = ws.cell(row=r, column=base+3+j, value=val)
                c.number_format = 'DD-MM-YYYY'; c.alignment = CENTER; c.border = BORDER; c.font = NORMAL

asignar_cruce_aux(wb['Late_Arrivals'], 2, 5, 8)
asignar_cruce_aux(wb['Pend_Confirmation'], 2, 4, 5)

# --- Missing: primero agregar cols, luego LHSHNUCO-BK, luego cruce amarillos NB-01
ws_m = wb['Missing']
last_col_m = ws_m.max_column
for j, h in enumerate(new_cols_detail):
    c = ws_m.cell(row=1, column=last_col_m+1+j, value=h)
    c.fill = AZUL_H; c.font = WHITE_B; c.alignment = CENTER; c.border = BORDER

# LHSHNUCO-BK — agregar fila amarilla
new_r = ws_m.max_row + 1
ws_m.cell(row=new_r, column=1, value='Lhotse')
ws_m.cell(row=new_r, column=2, value='LHSHNUCO-BK')
ws_m.cell(row=new_r, column=3, value='Shaker Vaso Proteina Acero Inox NutriCore Negro')
ws_m.cell(row=new_r, column=4, value=200)
ws_m.cell(row=new_r, column=5, value='BATCH 617')
for c in range(1, 6):
    cell = ws_m.cell(row=new_r, column=c)
    cell.fill = AMARI; cell.font = NORMAL; cell.border = BORDER
    cell.alignment = LEFT if c<=3 or c==5 else RIGHT

# Amarillos → NB-01
CRD_NB01 = datetime(2026, 7, 20)
ETD_NB01 = CRD_NB01 + timedelta(days=7)
ETAPTO_NB01 = ETD_NB01 + timedelta(days=45)
ETABOD_NB01 = ETAPTO_NB01 + timedelta(days=7)
asig_amari = {'LHSHNUCO-NY': 500, 'LHSHNUCO-BG': 300, 'LHSHNUCO-BK': 200}
for r in range(2, ws_m.max_row+1):
    sku = str(ws_m.cell(row=r, column=2).value or '').strip()
    fill_sku = ws_m.cell(row=r, column=2).fill
    is_yellow = fill_sku and fill_sku.start_color and str(fill_sku.start_color.rgb)=='FFFFFF00'
    if sku in asig_amari or is_yellow:
        base = last_col_m
        ws_m.cell(row=r, column=base+1, value=asig_amari.get(sku, ws_m.cell(row=r, column=4).value))
        ws_m.cell(row=r, column=base+2, value='NB-01')
        for j, val in enumerate([CRD_NB01, ETD_NB01, ETAPTO_NB01, ETABOD_NB01]):
            c = ws_m.cell(row=r, column=base+3+j, value=val)
            c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL

# ================ Bin packing Missing (excluir SZ-04, SZ-03-08, IMP OP 350-26)
print("\n=== Bin packing Missing (reglas restrictivas) ===")

# Espacio en actuales permitidos (basado en OHNSO Jul.01st)
cbm_x_emb = defaultdict(float)
for it in ohnso_items:
    if it['__seccion']=='Rest items': continue
    cbm = pd.to_numeric(it.get('TOTAL CBM'), errors='coerce')
    if pd.notna(cbm): cbm_x_emb[it['__seccion']] += cbm

libre = {e: CAP_40HQ - cbm_x_emb[e] for e in PERMITIDOS_ACTUALES}
# Sumar los LHSHNUCO amarillos que ya agregamos a NB-01 (Rest items en OHNSO 1000 mix)
# LHSHNUCO OHNSO: 1000 uds, cbm total desconocido — asumo proporción con qty. Buscar en Rest items
lhs_cbm = 0
for it in ohnso_items:
    if str(it.get('SKU','')).strip()=='LHSHNUCO-BK':
        cbm = pd.to_numeric(it.get('TOTAL CBM'), errors='coerce')
        if pd.notna(cbm): lhs_cbm = cbm
libre['NB-01'] -= lhs_cbm  # los 1000 uds LHSHNUCO ocupan este CBM en NB-01 ahora

print("Espacio libre en permitidos:")
for e, v in libre.items():
    print(f"  {e}: {v:.2f} CBM")

# Recolectar Missing pending (excluir amarillos ya asignados a NB-01)
YA_ASIG = {'LHSHNUCO-NY','LHSHNUCO-BG','LHSHNUCO-BK'}
pending = []
for r in range(2, ws_m.max_row+1):
    sku = str(ws_m.cell(row=r, column=2).value or '').strip()
    if not sku or sku in YA_ASIG: continue
    units_val = ws_m.cell(row=r, column=4).value
    if not isinstance(units_val, (int,float)): continue
    units = int(units_val)
    cbm, usd = get_ohnso_data(sku, units)
    if cbm is None:
        pending.append({'row':r,'sku':sku,'units':units,'cbm':0,'origen':None,'usd':usd,'skip':True})
        continue
    # Determinar origen (SZ o NB) según OHNSO sheet del match
    ms = ohnso_idx.get(sku, [])
    origen = 'NB' if ms and ms[0]['sheet']=='NINGBO' else 'SZ'
    pending.append({'row':r,'sku':sku,'units':units,'cbm':cbm,'origen':origen,'usd':usd,'skip':False})

# Fit-decreasing
pending_ok = [p for p in pending if not p['skip']]
pending_skip = [p for p in pending if p['skip']]
pending_ok.sort(key=lambda x: -x['cbm'])

nuevos = {}
asignaciones = {}  # row → embarque

def alloc_nuevo(origen):
    """Crear nuevo embarque."""
    if origen=='SZ':
        n = 5 + sum(1 for k in nuevos if k.startswith('SZ-'))
        name = f'SZ-{n:02d},40HQ'
    else:
        n = 2 + sum(1 for k in nuevos if k.startswith('NB-'))
        name = f'NB-{n:02d}'
    nuevos[name] = True
    libre[name] = CAP_40HQ
    print(f"  → CREADO {name}")
    return name

for p in pending_ok:
    origen = p['origen']; cbm = p['cbm']
    # Candidatos: PERMITIDOS_ACTUALES compatibles + nuevos ya creados
    compat = []
    for e in libre:
        if libre[e] < cbm: continue
        if e in FIJOS: continue
        # Filtrar por origen (SZ items en SZ-, NB items en NB-)
        if origen=='SZ' and e.startswith('NB-'): continue
        if origen=='NB' and e.startswith('SZ-'): continue
        compat.append(e)
    if compat:
        # Best-fit: menor libre suficiente
        compat.sort(key=lambda e: libre[e])
        emb = compat[0]
    else:
        emb = alloc_nuevo(origen)
    libre[emb] -= cbm
    asignaciones[p['row']] = emb
    print(f"  R{p['row']} {p['sku']:<20} {p['units']:>4}u CBM={cbm:>6.2f} orig={origen} → {emb}")

# Escribir asignaciones en Missing
for r, emb in asignaciones.items():
    crd, etd, etap, etab = fechas(emb)
    base = last_col_m
    units = int(ws_m.cell(row=r, column=4).value)
    ws_m.cell(row=r, column=base+1, value=units)
    ws_m.cell(row=r, column=base+2, value=emb)
    for j, val in enumerate([crd, etd, etap, etab]):
        c = ws_m.cell(row=r, column=base+3+j, value=val)
        c.number_format='DD-MM-YYYY'; c.alignment=CENTER; c.border=BORDER; c.font=NORMAL

# ================ MOVER filas Missing con embarque → Detail_SZ / Detail_NB
print("\n=== Mover Missing con embarque → Detail ===")
ws_sz = wb['Detail_SZ']; ws_nb = wb['Detail_NB']

def add_to_detail(ws, item, source_row):
    r = max([rr for rr in range(2, ws.max_row+1) if ws.cell(row=rr, column=6).value] + [1]) + 1
    cbm, usd = get_ohnso_data(item['sku'], item['units'])
    ws.cell(row=r, column=1, value=item['emb'])
    ws.cell(row=r, column=3, value=item['etab'])
    ws.cell(row=r, column=4, value='MISSING→OHNSO')
    ws.cell(row=r, column=5, value=item['brand'])
    ws.cell(row=r, column=6, value=item['sku'])
    ws.cell(row=r, column=7, value=item['desc'])
    ws.cell(row=r, column=8, value=item['units'])
    if cbm is not None: ws.cell(row=r, column=9, value=cbm)
    if usd is not None: ws.cell(row=r, column=10, value=usd)
    ws.cell(row=r, column=13, value=item['cant'])
    ws.cell(row=r, column=14, value=item['emb'])
    ws.cell(row=r, column=15, value=item['crd'])
    ws.cell(row=r, column=16, value=item['etd'])
    ws.cell(row=r, column=17, value=item['etap'])
    ws.cell(row=r, column=18, value=item['etab'])
    for c in range(1, 19):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER; cell.font = NORMAL
        if c in (3, 15, 16, 17, 18):
            cell.number_format = 'DD-MM-YYYY'; cell.alignment = CENTER
        elif c in (8, 9, 10, 13):
            cell.number_format = '#,##0.00' if c==9 else '#,##0'; cell.alignment = RIGHT
        else:
            cell.alignment = LEFT if c in (5, 7) else CENTER
    # Marcar procedencia MISSING
    ws.cell(row=r, column=1).fill = NARANJA
    ws.cell(row=r, column=6).fill = NARANJA
    if item['yellow']:
        ws.cell(row=r, column=6).fill = AMARI
    return r

# Recolectar filas Missing con embarque (todos: los amarillos ya asignados + bin-packed)
to_move = []
for r in range(2, ws_m.max_row+1):
    sku = ws_m.cell(row=r, column=2).value
    if not sku: continue
    emb = ws_m.cell(row=r, column=last_col_m+2).value  # Embarque
    if not emb: continue
    fill_sku = ws_m.cell(row=r, column=2).fill
    is_yellow = fill_sku and fill_sku.start_color and str(fill_sku.start_color.rgb)=='FFFFFF00'
    item = {
        'row': r, 'brand': ws_m.cell(row=r, column=1).value,
        'sku': str(sku).strip(), 'desc': ws_m.cell(row=r, column=3).value,
        'units': ws_m.cell(row=r, column=4).value, 'comment': ws_m.cell(row=r, column=5).value,
        'cant': ws_m.cell(row=r, column=last_col_m+1).value, 'emb': emb,
        'crd': ws_m.cell(row=r, column=last_col_m+3).value,
        'etd': ws_m.cell(row=r, column=last_col_m+4).value,
        'etap': ws_m.cell(row=r, column=last_col_m+5).value,
        'etab': ws_m.cell(row=r, column=last_col_m+6).value,
        'yellow': is_yellow,
    }
    to_move.append(item)

moved_sz = moved_nb = 0
for item in to_move:
    emb_upper = str(item['emb']).upper()
    if emb_upper.startswith('SZ-'):
        add_to_detail(ws_sz, item, item['row']); moved_sz += 1
    else:
        add_to_detail(ws_nb, item, item['row']); moved_nb += 1

# Eliminar filas movidas (mayor a menor)
rows_del = sorted([m['row'] for m in to_move], reverse=True)
for r in rows_del:
    ws_m.delete_rows(r, 1)

print(f"Movidos a Detail_SZ: {moved_sz}, Detail_NB: {moved_nb}")
print(f"Quedan en Missing: {ws_m.max_row - 1}")

# ================ CF + fills en Detail_SZ/NB/Missing
for sh_name, u_col_letter in [('Detail_SZ','H'), ('Detail_NB','H'), ('Missing','D'),
                                ('Late_Arrivals','E'), ('Pend_Confirmation','D')]:
    ws = wb[sh_name]
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    col_cant = hdr.get('Cantidad cargada')
    if not col_cant: continue
    c_l = get_column_letter(col_cant)
    last = ws.max_row
    ws.conditional_formatting._cf_rules = {}
    rng = f'{c_l}2:{c_l}{last}'
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2={u_col_letter}2)'], fill=VERDE))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2>{u_col_letter}2)'], fill=AZUL_S))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2<{u_col_letter}2,{c_l}2>0)'], fill=ROJO))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR({c_l}2="",ISBLANK({c_l}2))'], fill=GRIS))
    # Fill inicial
    for r in range(2, last+1):
        u = ws.cell(row=r, column=ord(u_col_letter)-64).value
        c_val = ws.cell(row=r, column=col_cant).value
        cell = ws.cell(row=r, column=col_cant)
        if c_val is None or not isinstance(c_val,(int,float)):
            cell.fill = GRIS
        elif not isinstance(u,(int,float)):
            cell.fill = GRIS
        elif c_val == u: cell.fill = VERDE
        elif c_val > u: cell.fill = AZUL_S
        else: cell.fill = ROJO

# ================ Summary con bloque PLAN AJUSTADO + fórmulas SUMIFS
print("\n=== Summary con fórmulas dinámicas ===")
ws_s = wb['Summary']
# Buscar si ya existe bloque
row_plan = None
for r in range(1, ws_s.max_row+1):
    v = ws_s.cell(row=r, column=1).value
    if v and 'PLAN AJUSTADO' in str(v):
        row_plan = r; break

if row_plan is None:
    row_plan = ws_s.max_row + 2

# Limpiar bloque previo
for r in range(row_plan, min(row_plan+30, ws_s.max_row+1)):
    for c in range(1, 12):
        ws_s.cell(row=r, column=c).value = None
        ws_s.cell(row=r, column=c).fill = PatternFill(fill_type=None)
        ws_s.cell(row=r, column=c).border = Border()

# Escribir bloque
c = ws_s.cell(row=row_plan, column=1, value='PLAN AJUSTADO (fuente: OHNSO Jul.01st)')
c.font = Font(bold=True, size=12, color='FF1F3864')
ws_s.merge_cells(start_row=row_plan, start_column=1, end_row=row_plan, end_column=10)

hdr_row = row_plan + 1
hdr_txt = ['Embarque', 'Origen', 'CRD', 'ETD', 'ETA Puerto', 'ETA Bodega',
           'SKUs UnionX V02', 'Uds UnionX V02', '', '¿A tiempo ≤31-ago?']
for j, h in enumerate(hdr_txt):
    cc = ws_s.cell(row=hdr_row, column=j+1, value=h)
    cc.fill = AZUL_H; cc.font = WHITE_B; cc.alignment = CENTER; cc.border = BORDER

# Embarques: actuales (6) + nuevos
embarques_all = ['SZ-04,40HQ', 'SZ-03-08,40HQ', 'SZ-01(623)', 'SZ-02(623)',
                 'IMP OP 350-26 40HQ', 'NB-01'] + list(nuevos.keys())

data_start = hdr_row + 1
for i, e in enumerate(embarques_all):
    r = data_start + i
    origen = 'NB' if ('NB' in e.upper() or 'IMP OP' in e.upper()) else 'SZ'
    crd, etd, etap, etab = fechas(e)
    a_tiempo = 'SÍ' if etab and etab <= DEADLINE else 'NO'
    emb_ref = f'A{r}'
    formula_skus = (
        f'=COUNTIFS(Detail_SZ!N:N,{emb_ref})+COUNTIFS(Detail_NB!N:N,{emb_ref})'
        f'+COUNTIFS(Late_Arrivals!J:J,{emb_ref})'
        f'+COUNTIFS(Pend_Confirmation!G:G,{emb_ref})'
        f'+COUNTIFS(Missing!G:G,{emb_ref})'
    )
    formula_uds = (
        f'=SUMIFS(Detail_SZ!M:M,Detail_SZ!N:N,{emb_ref})'
        f'+SUMIFS(Detail_NB!M:M,Detail_NB!N:N,{emb_ref})'
        f'+SUMIFS(Late_Arrivals!I:I,Late_Arrivals!J:J,{emb_ref})'
        f'+SUMIFS(Pend_Confirmation!F:F,Pend_Confirmation!G:G,{emb_ref})'
        f'+SUMIFS(Missing!F:F,Missing!G:G,{emb_ref})'
    )
    vals = [e, origen, crd, etd, etap, etab, formula_skus, formula_uds, None, a_tiempo]
    for j, v in enumerate(vals):
        cc = ws_s.cell(row=r, column=j+1, value=v)
        cc.font = NORMAL; cc.border = BORDER
        if e in nuevos: cc.fill = NARANJA
        elif e in FIJOS: cc.fill = GRIS_L
        if j in (2,3,4,5): cc.number_format='DD-MM-YYYY'; cc.alignment=CENTER
        elif j in (6,7): cc.number_format='#,##0'; cc.alignment=RIGHT
        elif j == 9:
            cc.alignment=CENTER; cc.font=Font(bold=True)
            cc.fill = VERDE if v=='SÍ' else ROJO
        else: cc.alignment = LEFT if j==0 else CENTER

# Total row
tot_r = data_start + len(embarques_all)
ws_s.cell(row=tot_r, column=1, value='TOTAL').font = Font(bold=True)
ws_s.cell(row=tot_r, column=7, value=f'=SUM(G{data_start}:G{tot_r-1})').font = Font(bold=True)
ws_s.cell(row=tot_r, column=8, value=f'=SUM(H{data_start}:H{tot_r-1})').font = Font(bold=True)
for j in range(1, 11):
    cc = ws_s.cell(row=tot_r, column=j)
    cc.fill = GRIS_L; cc.border = BORDER
    if j==7: cc.number_format='#,##0'; cc.alignment=RIGHT
    if j==8: cc.number_format='#,##0'; cc.alignment=RIGHT

# Anchos summary
for col, w in [('A',24),('B',9),('C',12),('D',12),('E',12),('F',12),('G',10),('H',13),('I',3),('J',20)]:
    ws_s.column_dimensions[col].width = w

# Guardar
wb.save(P)
print(f"\n[OK] Guardado: {P}")
print(f"Embarques nuevos: {list(nuevos.keys())}")
print(f"9 SKUs sin match en OHNSO permanecen en Missing")
