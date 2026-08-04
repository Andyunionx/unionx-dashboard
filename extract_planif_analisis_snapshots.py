"""Genera los 5 snapshots de la vista Análisis Planificación desde el Excel
mensual COMMITEADO EN EL REPO (no del Desktop de Felipe).

Auto-detecta el Excel del mes más reciente en data/planificacion/
(analisis_planificacion_<MES>26[_vN].xlsx) y produce:
  - planif_cst_flat_snapshot.parquet        (hoja 'CST x Marca')          → Coberturas
  - planif_critico_marca_snapshot.parquet   (hoja 'Critico x Marca')      → Críticos por Marca
  - planif_sobrestock_snapshot.parquet      (hoja 'Sobrestock x SKU Padre')→ Sobrestock
  - planif_transitos_snapshot.parquet       (hoja 'Tránsitos por Embarque')→ Tránsitos
  - planif_nuevos_transito_snapshot.parquet (hoja 'Nuevos en Tránsito')    → Nuevos en Tránsito

Diseñado para correr DIARIO por cron (.github/workflows/sync_planif_analisis.yml):
cuando Felipe commitea un Excel de un mes nuevo, el cron regenera los snapshots
dentro de 1 día. Idempotente: si no hay Excel nuevo, los parquets no cambian.

Reemplaza los scripts manuales extract_cst_flat.py + extract_planif_ago26_snapshot.py
(que leían de C:\\Users\\felip\\Desktop).

Uso: python extract_planif_analisis_snapshots.py
"""
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).parent
PLANIF_DIR = ROOT / "data" / "planificacion"
OUT_DIR = PLANIF_DIR / "snapshots"
OUT_DIR.mkdir(parents=True, exist_ok=True)

MESES = {'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
         'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12}


def detectar_excel_mas_reciente() -> Path:
    """Elige analisis_planificacion_<MES>26[_vN].xlsx del mes más reciente (mayor versión)."""
    pat = re.compile(r'analisis_planificacion_([A-Z]{3})26(?:_v(\d+))?\.xlsx$', re.IGNORECASE)
    mejor = None  # (mes_num, version, path)
    for f in PLANIF_DIR.glob('analisis_planificacion_*26*.xlsx'):
        if f.name.startswith('~$'):
            continue
        m = pat.search(f.name)
        if not m:
            continue
        mes_num = MESES.get(m.group(1).upper())
        if not mes_num:
            continue
        version = int(m.group(2)) if m.group(2) else 0
        key = (mes_num, version)
        if mejor is None or key > mejor[0]:
            mejor = (key, f)
    if mejor is None:
        raise FileNotFoundError(f"No encontré analisis_planificacion_*26*.xlsx en {PLANIF_DIR}")
    return mejor[1]


# --- Sheet 'CST x Marca' → cst_flat (Coberturas) ---
_MESES_CST = {
    '2026-08': {'stk_ini': 3, 'llegadas': 4, 'sp': 5, 'venta': 6, 'cob': 7},
    '2026-09': {'stk_ini': 8, 'llegadas': 9, 'sp': 10, 'venta': 11, 'cob': 12},
    '2026-10': {'stk_ini': 13, 'llegadas': 14, 'sp': 15, 'venta': 16, 'cob': 17},
}
_SUMMARY_ROWS = {'TOTAL PROPIA': '_TOTAL_PROPIA', 'TOTAL EMPRESA': '_TOTAL_EMPRESA'}
_MARCA_NORM = {'Dynamo': 'Dynamo Tools', 'Dynamo TL': 'Dynamo Tools', 'Dinamo Tools': 'Dynamo Tools',
               'Dynamo Tools': 'Dynamo Tools', 'Bandú': 'Bandú', 'Bandu': 'Bandú'}


def _extract_cst(wb):
    ws = wb['CST x Marca']
    records = []
    for r in ws.iter_rows(values_only=True):
        marca_raw = r[0]
        if not marca_raw or not isinstance(marca_raw, str):
            continue
        marca_raw = marca_raw.strip()
        if 'REPORTE' in marca_raw or 'Valores' in marca_raw or marca_raw == 'Marca' or marca_raw == 'PROV. NACIONALES':
            continue
        marca = _SUMMARY_ROWS[marca_raw] if marca_raw in _SUMMARY_ROWS else _MARCA_NORM.get(marca_raw, marca_raw)
        row = {'marca': marca, 'stock_hoy_cst': float(r[1] or 0) / 1e6, 'cobert_act': float(r[2] or 0)}
        for mes, cols in _MESES_CST.items():
            row[f'{mes}_stk_ini'] = float(r[cols['stk_ini']] or 0) / 1e6
            row[f'{mes}_llegadas'] = float(r[cols['llegadas']] or 0) / 1e6
            row[f'{mes}_sp'] = float(r[cols['sp']] or 0) / 1e6
            row[f'{mes}_venta'] = float(r[cols['venta']] or 0) / 1e6
            row[f'{mes}_cob'] = float(r[cols['cob']] or 0)
        records.append(row)
    df = pd.DataFrame(records)
    df.to_parquet(str(OUT_DIR / 'planif_cst_flat_snapshot.parquet'), index=False)
    print(f"  ✅ cst_flat: {len(df)} filas")


def _extract_critico(wb):
    ws = wb['Critico x Marca']
    rows = list(ws.iter_rows(values_only=True))
    records = []
    for r in rows[2:]:
        if not r[0]:
            continue
        marca = str(r[0]).strip()
        records.append({'marca': marca, 'skus': int(r[1] or 0), 'cob_prom': float(r[2] or 0),
                        'sin_stock': int(r[3] or 0), 'stock_hoy_cst': float(r[6] or 0),
                        'venta_cst_ago26': float(r[7] or 0),
                        'detalle_llegadas': str(r[8]).strip() if r[8] else '', 'is_total': marca == 'TOTAL'})
    pd.DataFrame(records).to_parquet(str(OUT_DIR / 'planif_critico_marca_snapshot.parquet'), index=False)
    print(f"  ✅ critico_marca: {len(records)} filas")


def _clean(s):
    for ch in ('▶', '▸', '▹', '↳'):
        s = s.replace(ch, '')
    return s.strip()


def _extract_sobrestock(wb):
    ws = wb['Sobrestock x SKU Padre']
    rows = list(ws.iter_rows(values_only=True))
    records = []
    cur_marca = cur_padre = cur_hijo = ''
    for r in rows[2:]:
        col0 = r[0]
        if col0 is None or not str(col0).strip():
            continue
        s = str(col0).strip()
        nivel = 1 if s.startswith('▶') else (2 if '▸' in s else (3 if '▹' in s else (4 if '↳' in s else 1)))
        nombre = _clean(s)
        if nivel == 1:
            cur_marca, cur_padre, cur_hijo = nombre, '', ''
        elif nivel == 2:
            cur_padre, cur_hijo = nombre, ''
        elif nivel == 3:
            cur_hijo = nombre
        records.append({'nombre': str(col0), 'nombre_clean': nombre, 'marca_parent': cur_marca,
                        'cat_padre_parent': cur_padre, 'cat_hijo_parent': cur_hijo if nivel == 4 else '',
                        'descripcion': str(r[1]).strip() if r[1] else '', 'skus': r[2],
                        'cobert_act': float(r[3]) if r[3] is not None else None,
                        'meses_exceso': float(r[4]) if r[4] is not None else None,
                        'stock_cst': float(r[5]) if r[5] is not None else None,
                        'venta_cst_prom': float(r[6]) if r[6] is not None else None,
                        'stock_optimo': float(r[7]) if r[7] is not None else None,
                        'capital_inmovilizado': float(r[8]) if r[8] is not None else None,
                        'tiene_llegadas': str(r[9]).strip() if r[9] else '', 'nivel': nivel})
    pd.DataFrame(records).to_parquet(str(OUT_DIR / 'planif_sobrestock_snapshot.parquet'), index=False)
    print(f"  ✅ sobrestock: {len(records)} filas")


_TRANSITOS_EXCLUIR = {'26TP0528PI', '26TP0704PI', 'TOTAL'}


def _extract_transitos(wb):
    ws = wb['Tránsitos por Embarque']
    rows = list(ws.iter_rows(values_only=True))
    records = []
    current_pi = None
    skip_pi = False
    for r in rows[2:]:
        col0 = r[0]
        if col0 is None or not str(col0).strip():
            continue
        s = str(col0).strip()
        if s.startswith('↳') or str(r[0]).startswith('  ↳'):
            if skip_pi:
                continue
            records.append({'row_type': 'sku', 'pi_embarque': current_pi,
                            'eta_o_desc': str(r[1]).strip() if r[1] else '', 'eta_bodega': str(r[2]).strip() if r[2] else '',
                            'mes_llegada': str(r[3]).strip() if r[3] else '', 'marcas': '', 'skus_distintos': None,
                            'criticos': str(r[6]).strip() if r[6] else '', 'inquietos': str(r[7]).strip() if r[7] else '',
                            'unidades': int(r[8]) if r[8] is not None else 0, 'valor_usd': float(r[9]) if r[9] is not None else 0.0,
                            'nivel_riesgo': '', 'sku': s.lstrip('↳ ').strip()})
        else:
            current_pi = s
            skip_pi = s in _TRANSITOS_EXCLUIR
            if skip_pi:
                continue
            records.append({'row_type': 'pi', 'pi_embarque': s,
                            'eta_o_desc': str(r[1]).strip() if r[1] else '', 'eta_bodega': str(r[2]).strip() if r[2] else '',
                            'mes_llegada': str(r[3]).strip() if r[3] else '', 'marcas': str(r[4]).strip() if r[4] else '',
                            'skus_distintos': int(r[5]) if r[5] is not None else 0,
                            'criticos': str(r[6]).strip() if r[6] is not None else '', 'inquietos': str(r[7]).strip() if r[7] is not None else '',
                            'unidades': int(r[8]) if r[8] is not None else 0, 'valor_usd': float(r[9]) if r[9] is not None else 0.0,
                            'nivel_riesgo': str(r[10]).strip() if r[10] else '', 'sku': ''})
    df = pd.DataFrame(records)
    df.to_parquet(str(OUT_DIR / 'planif_transitos_snapshot.parquet'), index=False)
    print(f"  ✅ transitos: {len(df)} filas")


def _extract_nuevos(wb):
    ws = wb['Nuevos en Tránsito']
    rows = list(ws.iter_rows(values_only=True))
    records = []
    current_grupo = ''
    for r in rows[2:]:
        col0 = r[0]
        s = str(col0).strip() if col0 is not None else ''
        if not s or s.startswith('TOTAL'):
            continue
        if '▶' in s:
            current_grupo = s.replace('▶', '').strip()
            continue
        records.append({'grupo': current_grupo, 'sku': s, 'descripcion': str(r[1]).strip() if r[1] else '',
                        'marca': str(r[2]).strip() if r[2] else '', 'mes_llegada': str(r[3]).strip() if r[3] else '',
                        'fecha_eta_bodega': str(r[4]).strip() if r[4] else '', 'cantidad': int(r[5]) if r[5] is not None else 0})
    pd.DataFrame(records).to_parquet(str(OUT_DIR / 'planif_nuevos_transito_snapshot.parquet'), index=False)
    print(f"  ✅ nuevos_transito: {len(records)} filas")


def main():
    excel = detectar_excel_mas_reciente()
    print(f"Excel detectado: {excel.name}")
    wb = openpyxl.load_workbook(str(excel), read_only=True, data_only=True)
    _extract_cst(wb)
    _extract_critico(wb)
    _extract_sobrestock(wb)
    _extract_transitos(wb)
    _extract_nuevos(wb)
    print("✅ 5 snapshots generados desde el repo.")


if __name__ == '__main__':
    main()
