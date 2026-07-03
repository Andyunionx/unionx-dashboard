"""Mover filas de Missing con embarque asignado → Detail_SZ o Detail_NB.
  - Embarque SZ-* → Detail_SZ
  - Embarque NB-* o IMP OP → Detail_NB
  - Missing queda solo con los sin embarque (no en OHNSO)
"""
import openpyxl, pandas as pd, sys, re
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

P = "data/planillas/Shipping Plan September V.02_actualizado.xlsx"
OHNSO = "agente-comex/data/inbox/20260701_0912_shipping_plan_ohnso/OHNSO-Jul.01st.xls"

# Estilos
VERDE  = PatternFill('solid', fgColor='FF92D050')
AZUL_S = PatternFill('solid', fgColor='FF9DC3E6')
ROJO   = PatternFill('solid', fgColor='FFF08080')
GRIS   = PatternFill('solid', fgColor='FFD9D9D9')
AMARI  = PatternFill('solid', fgColor='FFFFFF00')
NARANJA= PatternFill('solid', fgColor='FFFCE4D6')
NORMAL = Font(size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='FFB4B4B4')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Parsear OHNSO para CBM/USD/precio
def parsear(sheet):
    df = pd.read_excel(OHNSO, sheet_name=sheet, header=None)
    items, seccion, hdr = [], None, None
    for _, row in df.iterrows():
        vals = [v for v in row.values if pd.notna(v)]
        if not vals: continue
        first = str(vals[0]).strip()
        m = re.match(r'^\d+\.(.+)$', first)
        if m and len(vals) <= 2: seccion = m.group(1).strip(); continue
        if first == 'No' and len(vals)>1 and 'Model' in str(vals[1]):
            hdr = [str(v).strip() for v in row.values]; continue
        try: int(first)
        except: continue
        if hdr is None or seccion is None: continue
        it = {col: row.values[j] for j,col in enumerate(hdr) if j<len(row.values)}
        it['__seccion'] = seccion
        items.append(it)
    return items

ohnso_items = parsear('SHENZHEN') + parsear('NINGBO')
ohnso_idx = defaultdict(list)
for it in ohnso_items:
    sku = str(it.get('SKU','') or '').strip()
    if not sku or sku.lower()=='nan': continue
    ohnso_idx[sku].append(it)

def get_ohnso_data(sku, units):
    """Retorna CBM y USD Value proporcional a units."""
    ms = ohnso_idx.get(sku, [])
    if not ms: return None, None
    m = ms[0]
    qty = pd.to_numeric(m.get('Qty'), errors='coerce')
    price = pd.to_numeric(m.get('Price\n(USD)'), errors='coerce')
    cbm_ctn = pd.to_numeric(m.get('(CBM)\n/CTN'), errors='coerce')
    qty_ctn = pd.to_numeric(m.get("Q'ty/ctn"), errors='coerce')
    cbm_total = pd.to_numeric(m.get('TOTAL CBM'), errors='coerce')
    if pd.notna(qty_ctn) and pd.notna(cbm_ctn) and qty_ctn>0:
        cbm = round((units / qty_ctn) * cbm_ctn, 3)
    elif pd.notna(qty) and pd.notna(cbm_total) and qty>0:
        cbm = round((units / qty) * cbm_total, 3)
    else:
        cbm = None
    usd = round(units * price, 2) if pd.notna(price) else None
    return cbm, usd

# Cargar workbook
wb = openpyxl.load_workbook(P)
ws_m = wb['Missing']
ws_sz = wb['Detail_SZ']
ws_nb = wb['Detail_NB']

# Recolectar filas Missing con embarque asignado
COL_CANT, COL_EMB, COL_CRD, COL_ETD, COL_ETAP, COL_ETAB = 6, 7, 8, 9, 10, 11
mover = []
sin_embarque = []
for r in range(2, ws_m.max_row+1):
    sku = ws_m.cell(row=r, column=2).value
    if not sku: continue
    emb = ws_m.cell(row=r, column=COL_EMB).value
    row_data = {
        'row': r,
        'brand': ws_m.cell(row=r, column=1).value,
        'sku': str(sku).strip(),
        'desc': ws_m.cell(row=r, column=3).value,
        'units': ws_m.cell(row=r, column=4).value,
        'comment': ws_m.cell(row=r, column=5).value,
        'cant': ws_m.cell(row=r, column=COL_CANT).value,
        'emb': emb,
        'crd': ws_m.cell(row=r, column=COL_CRD).value,
        'etd': ws_m.cell(row=r, column=COL_ETD).value,
        'etap': ws_m.cell(row=r, column=COL_ETAP).value,
        'etab': ws_m.cell(row=r, column=COL_ETAB).value,
    }
    # Preservar si fila amarilla
    f_sku = ws_m.cell(row=r, column=2).fill
    row_data['yellow'] = f_sku and f_sku.start_color and str(f_sku.start_color.rgb)=='FFFFFF00'
    if emb:
        mover.append(row_data)
    else:
        sin_embarque.append(row_data)

print(f"Filas Missing a mover: {len(mover)}")
print(f"Filas Missing sin embarque (quedan): {len(sin_embarque)}")

# Agregar a Detail_SZ o Detail_NB
# Header Detail_SZ: A Container B ?? C ETA D Order No E Brand F SKU G Description H Units I CBM J USD Value K Need to pay L Late? M Cantidad cargada N Embarque O CRD P ETD Q ETA Puerto R ETA Bodega
def add_to_detail(ws, item):
    r = ws.max_row + 1
    # Encontrar espacios vacíos si max_row incluye filas basura
    # Mejor: encontrar última fila con SKU
    last_real = 1
    for rr in range(2, ws.max_row+1):
        if ws.cell(row=rr, column=6).value:
            last_real = rr
    r = last_real + 1

    cbm, usd = get_ohnso_data(item['sku'], item['units'])
    # Container = Embarque (para agrupar visualmente)
    ws.cell(row=r, column=1, value=item['emb'])           # A Container
    ws.cell(row=r, column=3, value=item['etab'])          # C ETA (usar ETA Bodega)
    ws.cell(row=r, column=4, value='MISSING→OHNSO')       # D Order No
    ws.cell(row=r, column=5, value=item['brand'])         # E Brand
    ws.cell(row=r, column=6, value=item['sku'])           # F SKU
    ws.cell(row=r, column=7, value=item['desc'])          # G Description
    ws.cell(row=r, column=8, value=item['units'])         # H Units
    if cbm is not None:
        ws.cell(row=r, column=9, value=cbm)               # I CBM
    if usd is not None:
        ws.cell(row=r, column=10, value=usd)              # J USD Value
    # K Need to pay?, L Late? — vacío
    ws.cell(row=r, column=13, value=item['cant'])         # M Cantidad cargada
    ws.cell(row=r, column=14, value=item['emb'])          # N Embarque
    ws.cell(row=r, column=15, value=item['crd'])          # O CRD
    ws.cell(row=r, column=16, value=item['etd'])          # P ETD
    ws.cell(row=r, column=17, value=item['etap'])         # Q ETA Puerto
    ws.cell(row=r, column=18, value=item['etab'])         # R ETA Bodega

    # Formatos
    for c in range(1, 19):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER; cell.font = NORMAL
        if c in (3, 15, 16, 17, 18):
            cell.number_format = 'DD-MM-YYYY'; cell.alignment = CENTER
        elif c in (8, 9, 10, 13):
            cell.number_format = '#,##0.00' if c==9 else '#,##0'
            cell.alignment = RIGHT
        else:
            cell.alignment = LEFT if c in (5, 7) else CENTER

    # Marcar fila con fondo naranja suave para señalar procedencia Missing
    for c in [1, 6]:  # Container y SKU
        ws.cell(row=r, column=c).fill = NARANJA
    # Amarillo original (LHSHNUCO)
    if item['yellow']:
        for c in range(1, 19):
            f_actual = ws.cell(row=r, column=c).fill
            # No pisar Cantidad cargada CF, pero sí SKU
            if c == 6:
                ws.cell(row=r, column=c).fill = AMARI
    return r

# Mover
moved_sz = 0; moved_nb = 0
for item in mover:
    emb = str(item['emb']).upper()
    if emb.startswith('SZ-'):
        add_to_detail(ws_sz, item)
        moved_sz += 1
    else:  # NB-, IMP OP
        add_to_detail(ws_nb, item)
        moved_nb += 1

print(f"\nMovidos a Detail_SZ: {moved_sz}")
print(f"Movidos a Detail_NB: {moved_nb}")

# Eliminar filas de Missing (de mayor a menor para no romper índices)
rows_to_delete = sorted([m['row'] for m in mover], reverse=True)
for r in rows_to_delete:
    ws_m.delete_rows(r, 1)

print(f"Filas eliminadas de Missing: {len(rows_to_delete)}")

# Re-aplicar CF en Detail_SZ y Detail_NB (por si extendieron rango)
for ws in [ws_sz, ws_nb]:
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
    col_units = hdr['Units']
    col_cant  = hdr['Cantidad cargada']
    u_l = get_column_letter(col_units)
    c_l = get_column_letter(col_cant)
    last = ws.max_row

    ws.conditional_formatting._cf_rules = {}
    rng = f'{c_l}2:{c_l}{last}'
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2={u_l}2)'], fill=VERDE))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2>{u_l}2)'], fill=AZUL_S))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_l}2),{c_l}2<{u_l}2,{c_l}2>0)'], fill=ROJO))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR({c_l}2="",ISBLANK({c_l}2))'], fill=GRIS))

    # Fill duro inicial
    for r in range(2, last+1):
        u = ws.cell(row=r, column=col_units).value
        c_val = ws.cell(row=r, column=col_cant).value
        cell = ws.cell(row=r, column=col_cant)
        if c_val is None or not isinstance(c_val,(int,float)):
            cell.fill = GRIS
        elif not isinstance(u,(int,float)):
            cell.fill = GRIS
        elif c_val == u:
            cell.fill = VERDE
        elif c_val > u:
            cell.fill = AZUL_S
        else:
            cell.fill = ROJO

# Re-aplicar CF Missing
ws_m.conditional_formatting._cf_rules = {}
c_letter = get_column_letter(COL_CANT); u_letter = 'D'
rng = f'{c_letter}2:{c_letter}{ws_m.max_row}'
ws_m.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2={u_letter}2)'], fill=VERDE))
ws_m.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2>{u_letter}2)'], fill=AZUL_S))
ws_m.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(ISNUMBER({c_letter}2),{c_letter}2<{u_letter}2,{c_letter}2>0)'], fill=ROJO))
ws_m.conditional_formatting.add(rng, FormulaRule(formula=[f'OR({c_letter}2="",ISBLANK({c_letter}2))'], fill=GRIS))

wb.save(P)
print(f"\n[OK] Guardado: {P}")

# Contar resultado
import pandas as pd
for sh in ['Detail_SZ', 'Detail_NB', 'Missing']:
    df = pd.read_excel(P, sheet_name=sh)
    df.columns = [str(c).strip() for c in df.columns]
    if 'Description' in df.columns:
        n = df[df['Description'].notna() & (df['Description'].astype(str).str.strip()!='')].shape[0]
    else:
        n = len(df)
    print(f"  {sh}: {n} filas con datos")
