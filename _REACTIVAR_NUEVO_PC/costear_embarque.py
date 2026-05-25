#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Costear Embarque - Implementacion local del skill comex-workflow.

USO:
    python costear_embarque.py --pi <PI.xlsx> --pl <PL.xlsx> --tarifas <Tarifas.xlsx>
                                [--maestra <Maestra_Importaciones.xlsx>]
                                [--out <carpeta_salida>]

EJEMPLO:
    python costear_embarque.py ^
        --pi    "C:\\...\\26TP0228PI NB 40HQ.xlsx" ^
        --pl    "C:\\...\\26TP0228PI PL.xlsx" ^
        --tarifas "C:\\...\\Tarifas_Base_COMEX - 0228.xlsx" ^
        --maestra "C:\\Users\\andre\\Desktop\\Maestra_Importaciones.xlsx" ^
        --out    "G:\\Mi unidad\\TRABAJO\\RESPALDO\\OPERACIONES\\UNION X - IA\\agente-comex\\data\\output\\26TP0228"

QUE HACE:
    1. Lee PI, PL y Tarifas. Detecta embarque y puerto.
    2. Valida conceptos. Alerta si hay items no reconocidos.
    3. Aplica formulas COMEX (delivery prorrateado, gift box +3%, comision Steven 3%, 4 CC).
    4. Genera Pre-costeo_x_CBM_<EMBARQUE>.xlsx
    5. Si --maestra: actualiza Maestra, 1.Apertura CC, 4.Matriz SKU, 5.Resumen Variaciones.
    6. Genera email_<EMBARQUE>.html con el analisis (template del skill).
    7. Imprime resumen ejecutivo en consola.

REQUISITOS:
    pip install openpyxl pandas
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

BENCHMARKS = {
    "SZ":  {"nombre": "Shenzhen", "benchmark": 16.0},
    "NB":  {"nombre": "Ningbo",   "benchmark": 18.0},
    "XI":  {"nombre": "Xiamen",   "benchmark": 17.0},
    "AIR": {"nombre": "Aereo",    "benchmark": 25.0},
    "DHL": {"nombre": "Aereo",    "benchmark": 25.0},
}

CONCEPTOS_INLAND_CHINA_KNOWN = {
    "form_f":         ["form f", "ff", "f f", "form-f", "f.f"],
    "local_charge":   ["local charge", "storage", "local-charge",
                        "monitor", "loading fee", "monitor loading",
                        "syntrans", "loading"],
    "long_vehicle":   ["long vehicle", "long vechile",  # typo común en PIs Steven
                        "cleaning custom", "customs cleaning", "cleaing custom",
                        "transport", "vehicle", "vechile",
                        "customs", "cleaing customs"],
    "comision_steven":["steven", "comision", "3%"],
}

PCT_AGENTE_ADUANA = 0.0016  # 0.16% sobre CIF en USD (segun skill)

# ---------------------------------------------------------------------------
# DATACLASSES
# ---------------------------------------------------------------------------

@dataclass
class Producto:
    model: str
    sku: str
    descripcion: str
    qty: float
    price: float
    gift_box_pi: float = 0.0
    gift_box_real: float = 0.0  # gift_box_pi * 1.03
    delivery_total: float = 0.0
    delivery_unitario: float = 0.0
    cbm_total: float = 0.0
    pct_cbm: float = 0.0
    # Costos calculados
    pxq: float = 0.0
    exw_producto: float = 0.0
    inland_china_producto: float = 0.0
    flete_producto: float = 0.0
    cif_producto: float = 0.0
    inland_chile_producto: float = 0.0
    costo_internado_total: float = 0.0
    costo_internado_unit: float = 0.0
    # Comparacion historica
    ultimo_costo: Optional[float] = None
    ultimo_embarque: Optional[str] = None
    variacion_pct: Optional[float] = None
    es_nuevo: bool = False

@dataclass
class GastosInlandChina:
    form_f: float = 0.0
    local_charge: float = 0.0
    long_vehicle: float = 0.0
    comision_steven: float = 0.0  # se recalcula
    detectados: dict = field(default_factory=dict)
    no_reconocidos: list = field(default_factory=list)

@dataclass
class Tarifas:
    puerto: str = "SZ"
    puerto_nombre: str = "Shenzhen"
    dolar: float = 950.0
    fecha_eta: str = ""
    flete_total_usd: float = 0.0
    capacidad_40hq_cbm: float = 68.0
    pct_agente_aduana: float = PCT_AGENTE_ADUANA
    gastos_chile_clp: dict = field(default_factory=dict)
    no_reconocidos: list = field(default_factory=list)

@dataclass
class Embarque:
    numero: str
    puerto: str
    puerto_nombre: str
    productos: list = field(default_factory=list)
    inland_china: GastosInlandChina = field(default_factory=GastosInlandChina)
    tarifas: Tarifas = field(default_factory=Tarifas)
    # Totales del embarque
    total_pxq: float = 0.0
    total_cbm: float = 0.0
    total_unidades: float = 0.0
    cc_exw: float = 0.0
    cc_inland_china: float = 0.0
    cc_flete: float = 0.0
    cc_inland_chile: float = 0.0
    total_cif: float = 0.0
    total_internado_clp: float = 0.0
    sobrecosto_usd: float = 0.0
    sobrecosto_pct: float = 0.0

# ---------------------------------------------------------------------------
# UTILIDADES
# ---------------------------------------------------------------------------

def _to_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", ".").replace("$", "").replace("USD", "").replace("CLP", "")
    s = re.sub(r"[^\d\.\-]", "", s)
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def _norm(s) -> str:
    if s is None:
        return ""
    import unicodedata
    txt = str(s).strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', txt) if unicodedata.category(c) != 'Mn')

def _detect_embarque_y_puerto_from_filename(path: Path) -> tuple[str, str]:
    """Extrae numero embarque y puerto del nombre del archivo (ej: 26TP0228PI NB 40HQ.xlsx)."""
    name = path.stem.upper()
    m = re.search(r"(\d{2}TP\d{4})", name)
    embarque = m.group(1) if m else "DESCONOCIDO"
    puerto = "SZ"
    for code in BENCHMARKS:
        if f" {code} " in f" {name} " or f" {code}." in name:
            puerto = code
            break
    return embarque, puerto

# ---------------------------------------------------------------------------
# LECTURA PI
# ---------------------------------------------------------------------------

def leer_pi(path: Path) -> tuple[list[Producto], GastosInlandChina, str, str]:
    """
    Lee el PI y devuelve productos (con delivery prorrateado), gastos Inland China,
    numero de embarque y puerto detectado.
    """
    print(f"\n[PI] Leyendo {path.name}...")
    embarque, puerto = _detect_embarque_y_puerto_from_filename(path)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Buscar fila de header (donde aparece "Model" o "MODEL")
    header_row_idx = None
    headers = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        for j, cell in enumerate(row):
            if _norm(cell) in ("model", "no", "no.", "n°"):
                header_row_idx = i
                break
        if header_row_idx:
            break

    if header_row_idx is None:
        # Asumir fila 1 si no encontramos
        header_row_idx = 1

    # Mapear columnas (acepta variantes con saltos de línea y unidades, ej. "Price\n(USD)")
    header_row = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    for j, cell in enumerate(header_row):
        n = _norm(cell).replace("\n", " ").strip()
        # Quitar paréntesis con unidades: "price (usd)" -> "price"
        n_base = re.sub(r"\s*\([^)]*\)\s*", "", n).strip()
        if n_base in ("model",):                              headers["model"] = j
        elif n_base in ("descripton", "description"):         headers["descripcion"] = j
        elif n_base in ("qty(pcs)", "qty", "q'ty", "qty pcs"): headers["qty"] = j
        elif n_base.startswith("qty"):                        headers["qty"] = j
        elif n_base in ("price",) or n_base.startswith("price"):  headers["price"] = j
        elif n_base in ("amount",) or n_base.startswith("amount"): headers["amount"] = j
        elif n_base in ("gift box", "giftbox"):               headers["gift_box"] = j
        elif n_base in ("sku",):                              headers["sku"] = j
        elif n_base in ("no", "no.", "n°"):                   headers["no"] = j

    print(f"  Header en fila {header_row_idx}. Columnas detectadas: {list(headers.keys())}")

    # Leer todas las filas tras el header
    productos_raw = []
    inland = GastosInlandChina()

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if not any(c is not None for c in row):
            continue

        descripcion_raw = row[headers.get("descripcion", -1)] if "descripcion" in headers else ""
        descripcion = _norm(descripcion_raw)
        model_raw = row[headers["model"]] if "model" in headers else ""
        model = str(model_raw).strip() if model_raw else ""
        model_norm = _norm(model)
        # Texto unificado para detectar conceptos (Steven a veces pone el concepto en
        # 'Model' y deja 'Description' vacío — ver bug doc del 0429/0430).
        texto_busqueda = f"{descripcion} {model_norm}".strip()
        amount = _to_float(row[headers["amount"]]) if "amount" in headers else 0.0
        if amount == 0 and "price" in headers:
            amount = _to_float(row[headers["price"]])

        # 1) Detectar conceptos Inland China (busca en descripcion + model)
        concepto_inland = None
        for key, patterns in CONCEPTOS_INLAND_CHINA_KNOWN.items():
            if any(p in texto_busqueda for p in patterns):
                concepto_inland = key
                break

        if concepto_inland:
            if concepto_inland == "form_f":
                inland.form_f += amount
            elif concepto_inland == "local_charge":
                inland.local_charge += amount
            elif concepto_inland == "long_vehicle":
                inland.long_vehicle += amount
            elif concepto_inland == "comision_steven":
                inland.comision_steven += amount  # valor del PI; lo recalculamos despues
            etiqueta = (descripcion or model_norm)[:40]
            inland.detectados[etiqueta] = amount
            continue

        # 2) Detectar delivery cost (también busca en model)
        if "delivery" in texto_busqueda:
            productos_raw.append({"_tipo": "delivery", "amount": amount, "row": row_idx})
            continue

        # 3) Producto valido
        qty = _to_float(row[headers["qty"]]) if "qty" in headers else 0.0

        # Salvaguarda: una fila con qty=1, sin SKU, sin descripcion y con un
        # 'model' que NO contiene dígitos es casi seguro un gasto no reconocido
        # (ej. "FF", "Monitor loading fee", "Delivery cost"), NO un producto.
        # Los códigos reales siempre contienen al menos un dígito (TP150-N, AT-12).
        # Se acumula como local_charge para que se prorratee correctamente y
        # ADEMÁS queda registrado en no_reconocidos para visibilidad.
        sku_raw_check = str(row[headers.get("sku", -1)] or "").strip() if "sku" in headers else ""
        model_tiene_digito = bool(re.search(r'\d', model)) if model else False
        if qty == 1 and not sku_raw_check and not descripcion and not model_tiene_digito and amount > 0:
            inland.local_charge += amount
            inland.no_reconocidos.append({"row": row_idx, "descripcion": model, "amount": amount})
            print(f"  [salvaguarda] '{model}' (USD {amount}) tratado como local_charge — agregar a CONCEPTOS_INLAND_CHINA_KNOWN si recurrente")
            continue

        if model and qty > 0:
            producto = {
                "_tipo": "producto",
                "model": model,
                "descripcion": str(row[headers.get("descripcion", -1)] or "").strip() if "descripcion" in headers else "",
                "qty": qty,
                "price": _to_float(row[headers.get("price", -1)]) if "price" in headers else 0.0,
                "amount": amount,
                "gift_box_pi": _to_float(row[headers.get("gift_box", -1)]) if "gift_box" in headers else 0.0,
                "sku": str(row[headers.get("sku", -1)] or "").strip() if "sku" in headers else "",
                "row": row_idx,
            }
            productos_raw.append(producto)
        else:
            # Item raro - no lo conocemos
            if amount > 0 and descripcion and "total" not in descripcion:
                inland.no_reconocidos.append({"row": row_idx, "descripcion": str(descripcion_raw).strip(), "amount": amount})

    # Prorratear delivery: cada "delivery cost" en el PI aplica a TODOS los
    # productos consecutivos anteriores (puede ser un grupo con distintos
    # Models). Antes el código cerraba el grupo al cambiar de Model y dejaba
    # huérfanos los productos previos, concentrando todo el delivery en el
    # último producto → causaba variaciones de +700% (ver SIMBOWSIE-4 en
    # 26TP0430).
    productos: list[Producto] = []
    grupo: list[dict] = []

    for item in productos_raw:
        if item["_tipo"] == "delivery":
            if grupo:
                total_qty = sum(p["qty"] for p in grupo)
                if total_qty > 0:
                    delivery_unit = item["amount"] / total_qty
                    for p in grupo:
                        p["delivery_unitario"] = delivery_unit
                        p["delivery_total"] = delivery_unit * p["qty"]
                grupo.clear()
            continue

        # Producto: acumular hasta que llegue un delivery cost (sin importar
        # si cambia el Model).
        grupo.append(item)

    # Si quedó un grupo abierto sin delivery al final, los productos van con 0.
    for p in grupo:
        p.setdefault("delivery_total", 0.0)
        p.setdefault("delivery_unitario", 0.0)

    # Recolectar todos los productos (los cerrados ya tienen delivery, los abiertos quedaron en grupo)
    for item in productos_raw:
        if item["_tipo"] == "producto":
            productos.append(Producto(
                model=item["model"],
                sku=item["sku"],
                descripcion=item["descripcion"],
                qty=item["qty"],
                price=item["price"],
                gift_box_pi=item["gift_box_pi"],
                gift_box_real=item["gift_box_pi"] * 1.03,
                delivery_total=item.get("delivery_total", 0.0),
                delivery_unitario=item.get("delivery_unitario", 0.0),
            ))

    print(f"  {len(productos)} productos extraidos.")
    print(f"  Inland China detectado: form_f={inland.form_f:.2f}  local={inland.local_charge:.2f}  long_veh={inland.long_vehicle:.2f}  steven_pi={inland.comision_steven:.2f}")
    if inland.no_reconocidos:
        print(f"  [!] {len(inland.no_reconocidos)} concepto(s) no reconocidos en PI.")

    return productos, inland, embarque, puerto

# ---------------------------------------------------------------------------
# LECTURA PL (CBM por producto)
# ---------------------------------------------------------------------------

def leer_pl(path: Path, productos: list[Producto]) -> float:
    """Lee el PL y rellena el campo cbm_total de cada producto. Devuelve total CBM."""
    print(f"\n[PL] Leyendo {path.name}...")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Header
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        for cell in row:
            if _norm(cell) in ("model",):
                header_row_idx = i
                break
        if header_row_idx:
            break
    if header_row_idx is None:
        header_row_idx = 1

    header_row = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    cols = {}
    for j, cell in enumerate(header_row):
        n = _norm(cell)
        if n == "model":                                 cols["model"] = j
        elif n in ("descripton", "description"):         cols["descripcion"] = j
        elif n in ("total cbm", "cbm total", "cbm"):     cols["cbm_total"] = j
        elif n in ("cbm/ctn",):                          cols["cbm_ctn"] = j
        elif n in ("total packages", "ctn"):             cols["ctn"] = j
        elif n in ("total q'ty", "total qty", "qty"):    cols["qty"] = j

    print(f"  Header fila {header_row_idx}. Columnas: {list(cols.keys())}")

    # Construir indice por model+descripcion para hacer match
    items_pl = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(c is not None for c in row):
            continue
        model = str(row[cols["model"]]).strip() if "model" in cols and row[cols["model"]] else ""
        if not model or _norm(model) in ("total", ""):
            continue
        cbm = _to_float(row[cols["cbm_total"]]) if "cbm_total" in cols else 0.0
        if cbm == 0 and "cbm_ctn" in cols and "ctn" in cols:
            cbm = _to_float(row[cols["cbm_ctn"]]) * _to_float(row[cols["ctn"]])
        items_pl.append({
            "model": model,
            "descripcion": _norm(row[cols.get("descripcion", -1)]) if "descripcion" in cols else "",
            "qty": _to_float(row[cols.get("qty", -1)]) if "qty" in cols else 0.0,
            "cbm": cbm,
        })

    # Match: buscar por model + descripcion mas cercana, fallback a model unicamente
    total_cbm = 0.0
    for prod in productos:
        match = None
        # 1) Match exacto por model + descripcion
        candidatos = [it for it in items_pl if it["model"].lower() == prod.model.lower()]
        if len(candidatos) == 1:
            match = candidatos[0]
        elif len(candidatos) > 1:
            # buscar por descripcion mas cercana
            d_norm = _norm(prod.descripcion)
            best = None
            best_score = 0
            for c in candidatos:
                # score por substring overlap
                if d_norm and c["descripcion"]:
                    score = sum(1 for w in d_norm.split() if w in c["descripcion"])
                    if score > best_score:
                        best, best_score = c, score
            match = best if best else candidatos[0]

        if match:
            prod.cbm_total = match["cbm"]
            total_cbm += match["cbm"]

    print(f"  Total CBM: {total_cbm:.4f}")
    return total_cbm

# ---------------------------------------------------------------------------
# LECTURA TARIFAS
# ---------------------------------------------------------------------------

def leer_tarifas(path: Path, puerto_default: str) -> Tarifas:
    print(f"\n[TARIFAS] Leyendo {path.name}...")
    wb = openpyxl.load_workbook(path, data_only=True)
    t = Tarifas()
    t.puerto = puerto_default
    t.puerto_nombre = BENCHMARKS.get(puerto_default, {}).get("nombre", puerto_default)

    # Recorrer todas las hojas y celdas buscando claves
    keymap = {
        "puerto": "puerto",
        "puerto origen": "puerto",
        "dolar": "dolar",
        "dolar aduana": "dolar",
        "fecha eta": "fecha_eta",
        "eta": "fecha_eta",
        "flete total": "flete_total_usd",
        "flete total 40hq": "flete_total_usd",
        "flete usd": "flete_total_usd",
        "capacidad 40hq": "capacidad_40hq_cbm",
        "capacidad cbm": "capacidad_40hq_cbm",
        "agente aduana": "pct_agente_aduana",
        "agente aduana (%)": "pct_agente_aduana",
    }
    gastos_clp_keys = [
        "gastos puerto", "puerto sti", "sti",
        "flete terrestre", "terrestre",
        "seimex",
        "desconsolidacion", "craft",
        "seguro", "seguro carga",
        "gastos despacho", "despacho",
        "gate in", "gate in maersk",
    ]

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for j, cell in enumerate(row):
                if cell is None:
                    continue
                key = _norm(cell)
                value = row[j + 1] if j + 1 < len(row) else None
                if key in keymap and value is not None:
                    field = keymap[key]
                    if field == "fecha_eta":
                        t.fecha_eta = str(value)
                    elif field == "puerto":
                        v = str(value).strip().upper()
                        if v in BENCHMARKS:
                            t.puerto = v
                            t.puerto_nombre = BENCHMARKS[v]["nombre"]
                    else:
                        setattr(t, field, _to_float(value))
                # Gastos CLP por palabra clave
                for gk in gastos_clp_keys:
                    if gk in key and value is not None and isinstance(value, (int, float)) and value > 0:
                        t.gastos_chile_clp[key] = _to_float(value)

    print(f"  Puerto={t.puerto} ({t.puerto_nombre})  Dolar={t.dolar}  Flete={t.flete_total_usd}  ETA={t.fecha_eta}")
    print(f"  Gastos CLP detectados: {len(t.gastos_chile_clp)}  total={sum(t.gastos_chile_clp.values()):,.0f} CLP")
    return t

# ---------------------------------------------------------------------------
# CALCULO PRE-COSTEO
# ---------------------------------------------------------------------------

def calcular_costeo(emb: Embarque):
    print(f"\n[CALC] Calculando costeo para {emb.numero}...")

    # Totales productos
    emb.total_pxq = sum(p.qty * p.price for p in emb.productos)
    emb.total_unidades = sum(p.qty for p in emb.productos)
    if emb.total_cbm == 0:
        emb.total_cbm = sum(p.cbm_total for p in emb.productos)

    # 1) CC1: EXW = sum(gift_box_real) + sum(delivery_total)
    emb.cc_exw = sum(p.gift_box_real for p in emb.productos) + sum(p.delivery_total for p in emb.productos)

    # 2) CC2: Inland China
    # Recalcular comision Steven 3% sobre (P×Q + delivery + local + long_vehicle)
    base_steven = emb.total_pxq + sum(p.delivery_total for p in emb.productos) + emb.inland_china.local_charge + emb.inland_china.long_vehicle
    emb.inland_china.comision_steven = base_steven * 0.03
    emb.cc_inland_china = (
        emb.inland_china.comision_steven
        + emb.inland_china.local_charge
        + emb.inland_china.long_vehicle
        + emb.inland_china.form_f
    )

    # 3) CC3: Flete (total)
    emb.cc_flete = emb.tarifas.flete_total_usd

    # 4) CIF y CC4: Inland Chile
    emb.total_cif = emb.total_pxq + emb.cc_exw + emb.cc_inland_china + emb.cc_flete
    agente_aduana_usd = emb.total_cif * emb.tarifas.pct_agente_aduana
    gastos_chile_clp_total = sum(emb.tarifas.gastos_chile_clp.values())
    emb.cc_inland_chile = agente_aduana_usd + (gastos_chile_clp_total / emb.tarifas.dolar if emb.tarifas.dolar else 0)

    # Sobrecosto
    emb.sobrecosto_usd = emb.cc_exw + emb.cc_inland_china + emb.cc_flete + emb.cc_inland_chile
    emb.sobrecosto_pct = (emb.sobrecosto_usd / emb.total_pxq * 100) if emb.total_pxq else 0
    emb.total_internado_clp = (emb.total_pxq + emb.sobrecosto_usd) * emb.tarifas.dolar

    # Costos por producto (prorrateado por CBM)
    total_exw = sum((p.qty * p.price) + p.gift_box_real + p.delivery_total for p in emb.productos)
    for p in emb.productos:
        p.pxq = p.qty * p.price
        p.exw_producto = p.pxq + p.gift_box_real + p.delivery_total
        p.pct_cbm = (p.cbm_total / emb.total_cbm) if emb.total_cbm else 0
        p.inland_china_producto = emb.cc_inland_china * (p.exw_producto / total_exw if total_exw else 0)
        p.flete_producto = emb.cc_flete * p.pct_cbm
        p.cif_producto = p.exw_producto + p.inland_china_producto + p.flete_producto
        agente_p = (emb.total_cif * emb.tarifas.pct_agente_aduana) * p.pct_cbm * emb.tarifas.dolar
        gastos_chile_p = gastos_chile_clp_total * p.pct_cbm
        p.inland_chile_producto = (agente_p / emb.tarifas.dolar if emb.tarifas.dolar else 0) + (gastos_chile_p / emb.tarifas.dolar if emb.tarifas.dolar else 0)
        p.costo_internado_total = (p.cif_producto * emb.tarifas.dolar) + agente_p + gastos_chile_p
        p.costo_internado_unit = p.costo_internado_total / p.qty if p.qty else 0

    print(f"  P×Q={emb.total_pxq:,.2f}  EXW={emb.cc_exw:,.2f}  InlChn={emb.cc_inland_china:,.2f}  Flete={emb.cc_flete:,.2f}  InlChl={emb.cc_inland_chile:,.2f}")
    print(f"  CIF={emb.total_cif:,.2f}  Sobrecosto={emb.sobrecosto_usd:,.2f} ({emb.sobrecosto_pct:.2f}%)  Internado CLP={emb.total_internado_clp:,.0f}")

# ---------------------------------------------------------------------------
# COMPARACION VS MAESTRA HISTORICA
# ---------------------------------------------------------------------------

def comparar_con_maestra(emb: Embarque, maestra_path: Optional[Path]):
    if not maestra_path or not maestra_path.exists():
        print("\n[COMPARE] Maestra no proporcionada o no existe. Todos los productos seran marcados sin historico.")
        for p in emb.productos:
            p.es_nuevo = True
        return

    print(f"\n[COMPARE] Comparando vs Maestra historica: {maestra_path.name}")
    try:
        df = pd.read_excel(maestra_path, sheet_name="Maestra", engine="openpyxl")
    except Exception as e:
        print(f"  [!] No se pudo leer pestania 'Maestra': {e}. Salteando comparacion.")
        for p in emb.productos:
            p.es_nuevo = True
        return

    sku_col = next((c for c in df.columns if _norm(c) in ("sku",)), None)
    costo_col = next((c for c in df.columns if "costo" in _norm(c) and ("neto" in _norm(c) or "internado" in _norm(c))), None)
    eta_col = next((c for c in df.columns if _norm(c) == "eta"), None)
    embarque_col = next((c for c in df.columns if "embarque" in _norm(c)), None)

    if not sku_col or not costo_col:
        print(f"  [!] No se encontraron columnas SKU/Costo en Maestra. Columnas: {list(df.columns)}")
        for p in emb.productos:
            p.es_nuevo = True
        return

    if eta_col:
        # ETA en la Maestra a veces es datetime y a veces texto (mezcla legacy).
        # Coerce a datetime para evitar TypeError al comparar; NaT al final.
        df['_eta_sort'] = pd.to_datetime(df[eta_col], errors='coerce')
        df = df.sort_values('_eta_sort', ascending=False, na_position='last').drop(columns='_eta_sort')

    for p in emb.productos:
        if not p.sku:
            p.es_nuevo = True
            continue
        hist = df[df[sku_col].astype(str).str.strip().str.lower() == p.sku.strip().lower()]
        if len(hist) == 0:
            p.es_nuevo = True
            continue
        ultimo = hist.iloc[0]
        p.ultimo_costo = _to_float(ultimo[costo_col])
        p.ultimo_embarque = str(ultimo[embarque_col]) if embarque_col else None
        if p.ultimo_costo > 0:
            p.variacion_pct = (p.costo_internado_unit - p.ultimo_costo) / p.ultimo_costo * 100

    n_var = sum(1 for p in emb.productos if p.variacion_pct is not None and abs(p.variacion_pct) > 10)
    n_new = sum(1 for p in emb.productos if p.es_nuevo)
    print(f"  Productos con variacion >10%: {n_var}.  Nuevos: {n_new}.")

# ---------------------------------------------------------------------------
# GENERAR PRE-COSTEO XLSX
# ---------------------------------------------------------------------------

def generar_precosteo_xlsx(emb: Embarque, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"Pre-costeo_x_CBM_{emb.numero}.xlsx"

    wb = openpyxl.Workbook()

    # Hoja resumen
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Embarque", emb.numero])
    ws.append(["Puerto", f"{emb.puerto} ({emb.puerto_nombre})"])
    ws.append(["ETA", emb.tarifas.fecha_eta])
    ws.append(["Dolar", emb.tarifas.dolar])
    ws.append([])
    ws.append(["Centro Costo", "USD", "% sobre P×Q"])
    base = emb.total_pxq if emb.total_pxq else 1
    ws.append(["Total Amount (P×Q)", emb.total_pxq, "Base"])
    ws.append(["EXW (Gift Box + Delivery)", emb.cc_exw, f"{emb.cc_exw/base*100:.2f}%"])
    ws.append(["Inland China", emb.cc_inland_china, f"{emb.cc_inland_china/base*100:.2f}%"])
    ws.append(["Flete Maritimo", emb.cc_flete, f"{emb.cc_flete/base*100:.2f}%"])
    ws.append(["Inland Chile", emb.cc_inland_chile, f"{emb.cc_inland_chile/base*100:.2f}%"])
    ws.append(["Sobrecosto Total", emb.sobrecosto_usd, f"{emb.sobrecosto_pct:.2f}%"])
    ws.append([])
    ws.append(["Total CIF", emb.total_cif])
    ws.append(["Total Internado (CLP)", emb.total_internado_clp])
    ws.append(["Total CBM", emb.total_cbm])
    ws.append(["Total Unidades", emb.total_unidades])

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 28

    # Hoja productos
    ws2 = wb.create_sheet("Productos")
    cols = [
        "Model", "SKU", "Descripcion", "Qty", "Price", "P×Q",
        "Gift Box PI", "Gift Box Real (×1.03)", "Delivery Total",
        "CBM Total", "% CBM",
        "EXW Producto", "Inland China Prod", "Flete Prod", "CIF Prod",
        "Inland Chile Prod", "Costo Internado Total (CLP)", "Costo Internado Unit (CLP)",
        "Ultimo Costo", "Ultimo Embarque", "Variacion %", "Estado"
    ]
    ws2.append(cols)
    for p in emb.productos:
        estado = "NUEVO" if p.es_nuevo else (
            f">10% AUMENTO" if p.variacion_pct and p.variacion_pct > 10 else
            f">10% MEJORA" if p.variacion_pct and p.variacion_pct < -10 else "OK"
        )
        ws2.append([
            p.model, p.sku, p.descripcion, p.qty, p.price, p.pxq,
            p.gift_box_pi, p.gift_box_real, p.delivery_total,
            p.cbm_total, p.pct_cbm,
            p.exw_producto, p.inland_china_producto, p.flete_producto, p.cif_producto,
            p.inland_chile_producto, p.costo_internado_total, p.costo_internado_unit,
            p.ultimo_costo, p.ultimo_embarque, p.variacion_pct, estado
        ])
    for j in range(1, len(cols) + 1):
        ws2.column_dimensions[get_column_letter(j)].width = 18
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A5276")

    # Hoja inland china detalle
    ws3 = wb.create_sheet("Inland China")
    ws3.append(["Concepto", "USD"])
    ws3.append(["Form F", emb.inland_china.form_f])
    ws3.append(["Local Charge", emb.inland_china.local_charge])
    ws3.append(["Long Vehicle", emb.inland_china.long_vehicle])
    ws3.append(["Comision Steven (3% recalc)", emb.inland_china.comision_steven])
    ws3.append(["TOTAL INLAND CHINA", emb.cc_inland_china])
    ws3.append([])
    ws3.append(["Conceptos detectados PI:"])
    for k, v in emb.inland_china.detectados.items():
        ws3.append([k, v])
    if emb.inland_china.no_reconocidos:
        ws3.append([])
        ws3.append(["NO RECONOCIDOS"])
        for nr in emb.inland_china.no_reconocidos:
            ws3.append([nr["descripcion"], nr["amount"]])

    # Hoja tarifas
    ws4 = wb.create_sheet("Tarifas")
    ws4.append(["Concepto", "Valor"])
    ws4.append(["Puerto", f"{emb.tarifas.puerto} ({emb.tarifas.puerto_nombre})"])
    ws4.append(["Dolar", emb.tarifas.dolar])
    ws4.append(["Fecha ETA", emb.tarifas.fecha_eta])
    ws4.append(["Flete Total USD", emb.tarifas.flete_total_usd])
    ws4.append(["Capacidad 40HQ CBM", emb.tarifas.capacidad_40hq_cbm])
    ws4.append(["Pct Agente Aduana", emb.tarifas.pct_agente_aduana])
    ws4.append([])
    ws4.append(["Gastos Chile (CLP)"])
    for k, v in emb.tarifas.gastos_chile_clp.items():
        ws4.append([k, v])
    ws4.append(["TOTAL CLP", sum(emb.tarifas.gastos_chile_clp.values())])

    wb.save(out_path)
    print(f"\n[OK] Pre-costeo guardado: {out_path}")
    return out_path

# ---------------------------------------------------------------------------
# ACTUALIZAR MAESTRA
# ---------------------------------------------------------------------------

def actualizar_maestra(emb: Embarque, maestra_path: Path):
    if not maestra_path.exists():
        print(f"\n[MAESTRA] No existe: {maestra_path}. Salteo actualizacion.")
        return

    print(f"\n[MAESTRA] Actualizando: {maestra_path.name}")
    backup_path = maestra_path.with_suffix(maestra_path.suffix + f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    import shutil
    shutil.copy2(maestra_path, backup_path)
    print(f"  Backup creado: {backup_path.name}")

    wb = openpyxl.load_workbook(maestra_path)

    # Maestra (filas por SKU)
    if "Maestra" in wb.sheetnames:
        ws = wb["Maestra"]
        # Encontrar columnas (asumiendo encabezados en fila 1)
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        next_row = ws.max_row + 1
        for p in emb.productos:
            ws.cell(next_row, headers.get("N° Embarque", 1), emb.numero)
            ws.cell(next_row, headers.get("SKU", 2), p.sku)
            ws.cell(next_row, headers.get("Model", 3), p.model)
            ws.cell(next_row, headers.get("Qty", 4), p.qty)
            ws.cell(next_row, headers.get("Price", 5), p.price)
            ws.cell(next_row, headers.get("Costo Neto Unitario", 6), p.costo_internado_unit)
            ws.cell(next_row, headers.get("ETA", 7), emb.tarifas.fecha_eta)
            ws.cell(next_row, headers.get("Puerto", 8), emb.puerto)
            # Resaltar fila nueva
            for c in range(1, ws.max_column + 1):
                ws.cell(next_row, c).fill = PatternFill("solid", fgColor="FFF3CD")
            next_row += 1
        print(f"  Pestaña 'Maestra': +{len(emb.productos)} filas")

    # 1. Apertura CC
    if "1. Apertura CC" in wb.sheetnames:
        ws = wb["1. Apertura CC"]
        next_row = ws.max_row + 1
        ws.cell(next_row, 1, emb.numero)
        ws.cell(next_row, 2, emb.puerto)
        ws.cell(next_row, 3, datetime.now().year)
        ws.cell(next_row, 4, emb.tarifas.fecha_eta)
        ws.cell(next_row, 5, len(emb.productos))
        ws.cell(next_row, 6, emb.total_pxq)
        ws.cell(next_row, 7, emb.cc_inland_china)
        ws.cell(next_row, 8, emb.total_pxq + emb.cc_exw + emb.cc_inland_china)  # Costo FOB
        ws.cell(next_row, 9, emb.cc_flete)
        ws.cell(next_row, 10, emb.total_cif)
        ws.cell(next_row, 11, emb.cc_inland_chile)
        ws.cell(next_row, 12, emb.total_internado_clp)
        ws.cell(next_row, 13, emb.sobrecosto_pct / 100)
        for c in range(1, 14):
            ws.cell(next_row, c).fill = PatternFill("solid", fgColor="FFF3CD")
        print(f"  Pestaña '1. Apertura CC': +1 fila")

    # 4. Matriz SKU
    if "4. Matriz SKU" in wb.sheetnames:
        ws = wb["4. Matriz SKU"]
        new_col = ws.max_column + 1
        ws.cell(1, new_col, emb.numero).font = Font(bold=True)
        # Mapear SKU -> fila
        sku_to_row = {}
        for r in range(2, ws.max_row + 1):
            sku_val = ws.cell(r, 1).value
            if sku_val:
                sku_to_row[str(sku_val).strip()] = r
        for p in emb.productos:
            if not p.sku:
                continue
            if p.sku in sku_to_row:
                ws.cell(sku_to_row[p.sku], new_col, round(p.costo_internado_unit, 0))
            else:
                # Agregar nuevo SKU
                next_row = ws.max_row + 1
                ws.cell(next_row, 1, p.sku)
                ws.cell(next_row, new_col, round(p.costo_internado_unit, 0))
        print(f"  Pestaña '4. Matriz SKU': nueva columna {emb.numero}")

    # 5. Resumen Variaciones
    if "5. Resumen Variaciones" in wb.sheetnames:
        ws = wb["5. Resumen Variaciones"]
        # Limpiar y recalcular top 20
        # (no implementamos limpieza completa; agregamos al final como nota informativa)
        next_row = ws.max_row + 2
        ws.cell(next_row, 1, f"Actualizado: {datetime.now().strftime('%Y-%m-%d')} - {emb.numero}").font = Font(italic=True, color="666666")
        next_row += 1
        ws.cell(next_row, 1, "SKU").font = Font(bold=True)
        ws.cell(next_row, 2, "Variacion %").font = Font(bold=True)
        ws.cell(next_row, 3, "Costo Actual").font = Font(bold=True)
        ws.cell(next_row, 4, "Ultimo Costo").font = Font(bold=True)
        next_row += 1
        var_prods = sorted([p for p in emb.productos if p.variacion_pct is not None],
                           key=lambda x: abs(x.variacion_pct or 0), reverse=True)[:20]
        for p in var_prods:
            ws.cell(next_row, 1, p.sku)
            ws.cell(next_row, 2, p.variacion_pct)
            ws.cell(next_row, 3, p.costo_internado_unit)
            ws.cell(next_row, 4, p.ultimo_costo)
            next_row += 1
        print(f"  Pestaña '5. Resumen Variaciones': +{len(var_prods)} filas (top 20)")

    wb.save(maestra_path)
    print(f"  [OK] Maestra guardada.")

# ---------------------------------------------------------------------------
# GENERAR EMAIL HTML
# ---------------------------------------------------------------------------

def _html_items_sin_sku(emb: 'Embarque') -> str:
    """Bloque HTML listando ítems del PI sin SKU asignado (samples / nuevos productos).
    Esos NO se cargan a la PO automática y requieren ingreso manual en Odoo."""
    sin_sku = []
    for p in emb.productos:
        sku = str(p.sku or '').strip()
        if not sku or sku.lower() in ('none', 'nan') or sku.lower().startswith(('samples', 'ice bag', 'cooler samples')):
            sin_sku.append(p)
    if not sin_sku:
        return ''
    filas = []
    for p in sin_sku:
        desc = (p.descripcion or p.model or '').replace('\n', ' ')[:120]
        filas.append(
            f'<tr><td style="padding:8px;border:1px solid #ddd">{p.sku or p.model or "—"}</td>'
            f'<td style="padding:8px;border:1px solid #ddd">{desc}</td>'
            f'<td style="padding:8px;border:1px solid #ddd;text-align:right">{p.qty:,.0f}</td>'
            f'<td style="padding:8px;border:1px solid #ddd;text-align:right">${p.price:,.2f}</td>'
            f'<td style="padding:8px;border:1px solid #ddd;text-align:right">${p.costo_internado_unit:,.0f}</td></tr>'
        )
    return f'''
<h3 style="color:#c0392b;margin-top:30px">⚠ 4. Ítems sin SKU — REQUIEREN INGRESO MANUAL</h3>
<p>El PI trae <b>{len(sin_sku)}</b> producto(s) sin SKU asignado (samples / nuevos). <b>NO se incluyeron en la PO automática a Odoo</b>. Es necesario crear el producto en Odoo y cargarlo manualmente para que ingresen al stock.</p>
<table style="width:100%;border-collapse:collapse;margin:15px 0;font-size:12px">
<thead><tr style="background:#c0392b;color:white">
<th style="padding:8px;border:1px solid #ddd">SKU / Tipo</th>
<th style="padding:8px;border:1px solid #ddd">Descripción</th>
<th style="padding:8px;border:1px solid #ddd">Qty</th>
<th style="padding:8px;border:1px solid #ddd">USD/u</th>
<th style="padding:8px;border:1px solid #ddd">CLP/u internado</th>
</tr></thead>
<tbody>
{"".join(filas)}
</tbody></table>'''


def generar_email_html(emb: Embarque, out_dir: Path) -> Path:
    benchmark = BENCHMARKS.get(emb.puerto, {}).get("benchmark", 16.0)
    diferencia_pp = emb.sobrecosto_pct - benchmark
    bajo_benchmark = diferencia_pp < 0

    bg_indicador = "#d5f5e3" if bajo_benchmark else "#fadbd8"
    color_indicador = "#1e8449" if bajo_benchmark else "#c0392b"
    simbolo = "&#10003;" if bajo_benchmark else "&#9888;&#65039;"
    direccion = "debajo" if bajo_benchmark else "encima"

    color_dif = "#27ae60" if bajo_benchmark else "#c0392b"
    signo_dif = "" if bajo_benchmark else "+"
    simbolo_dif = "&#10003;" if bajo_benchmark else "&#9888;&#65039;"

    # Productos con variacion
    productos_aumento = [p for p in emb.productos if p.variacion_pct is not None and p.variacion_pct > 10]
    productos_mejora = [p for p in emb.productos if p.variacion_pct is not None and p.variacion_pct < -10]
    productos_nuevos = [p for p in emb.productos if p.es_nuevo]

    base = emb.total_pxq if emb.total_pxq else 1
    total_pxq_fmt = f"{emb.total_pxq:,.2f}"
    cc_exw_fmt = f"{emb.cc_exw:,.2f}"
    cc_inland_china_fmt = f"{emb.cc_inland_china:,.2f}"
    cc_flete_fmt = f"{emb.cc_flete:,.2f}"
    cc_inland_chile_fmt = f"{emb.cc_inland_chile:,.2f}"
    sobrecosto_usd_fmt = f"{emb.sobrecosto_usd:,.2f}"

    def filas_aumento():
        rows = []
        for p in productos_aumento:
            rows.append(f"""
<tr style="background:#fadbd8">
<td style="padding:8px;border:1px solid #ddd">{p.descripcion}</td>
<td style="padding:8px;border:1px solid #ddd">{p.sku}</td>
<td style="padding:8px;text-align:right;border:1px solid #ddd">${p.costo_internado_unit:,.0f}</td>
<td style="padding:8px;text-align:right;border:1px solid #ddd">${p.ultimo_costo:,.0f}</td>
<td style="padding:8px;text-align:center;border:1px solid #ddd">{p.ultimo_embarque or "-"}</td>
<td style="padding:8px;text-align:center;border:1px solid #ddd;color:#c0392b"><b>+{p.variacion_pct:.1f}%</b></td>
</tr>""")
        return "".join(rows)

    def filas_mejora():
        rows = []
        for p in productos_mejora:
            rows.append(f"""
<tr style="background:#d5f5e3">
<td style="padding:8px;border:1px solid #ddd">{p.descripcion}</td>
<td style="padding:8px;border:1px solid #ddd">{p.sku}</td>
<td style="padding:8px;text-align:right;border:1px solid #ddd">${p.costo_internado_unit:,.0f}</td>
<td style="padding:8px;text-align:right;border:1px solid #ddd">${p.ultimo_costo:,.0f}</td>
<td style="padding:8px;text-align:center;border:1px solid #ddd">{p.ultimo_embarque or "-"}</td>
<td style="padding:8px;text-align:center;border:1px solid #ddd;color:#27ae60"><b>{p.variacion_pct:.1f}% &#10003;</b></td>
</tr>""")
        return "".join(rows)

    def filas_nuevos():
        rows = []
        for p in productos_nuevos[:20]:
            rows.append(f"""
<tr style="background:#f8f9fa">
<td style="padding:8px;border:1px solid #ddd">{p.descripcion}</td>
<td style="padding:8px;border:1px solid #ddd">{p.sku}</td>
<td style="padding:8px;text-align:right;border:1px solid #ddd">${p.costo_internado_unit:,.0f}</td>
<td style="padding:8px;text-align:right;border:1px solid #ddd">-</td>
<td style="padding:8px;text-align:center;border:1px solid #ddd">-</td>
<td style="padding:8px;text-align:center;border:1px solid #ddd"><span style="background:#e8f6f3;padding:2px 8px;border-radius:4px">NUEVO</span></td>
</tr>""")
        return "".join(rows)

    n_var = len(productos_aumento) + len(productos_mejora)
    n_nuevos = len(productos_nuevos)

    bg_concl = "#eafaf1" if bajo_benchmark else "#fadbd8"
    border_concl = "#27ae60" if bajo_benchmark else "#c0392b"
    color_concl = "#1e8449" if bajo_benchmark else "#c0392b"
    simbolo_concl = "&#10003;" if bajo_benchmark else "&#9888;&#65039;"
    texto_concl = (
        f"Embarque {emb.numero} con sobrecosto de {emb.sobrecosto_pct:.2f}% — "
        + ("por debajo del benchmark del puerto. Buen desempeño en costos." if bajo_benchmark else
           "por encima del benchmark del puerto. Revisar partidas con mayor desviación.")
    )

    html = f"""<div dir="ltr">
<div style="font-family:Arial,sans-serif;line-height:1.6;color:#333;max-width:800px;margin:0 auto">
<h2 style="color:#1a5276;border-bottom:2px solid #1a5276;padding-bottom:10px">&#128230; Análisis Embarque {emb.numero} | {emb.puerto} ({emb.puerto_nombre})</h2>

<table style="width:100%;background:{bg_indicador};border-radius:8px;margin:20px 0">
<tbody><tr><td style="padding:20px;text-align:center">
<span style="font-size:28px;font-weight:bold;color:{color_indicador}">{simbolo} Sobrecosto: {emb.sobrecosto_pct:.2f}%</span>
</td></tr></tbody></table>

<p style="font-size:16px">Embarque <b>{abs(diferencia_pp):.1f} pp por {direccion}</b> del benchmark de {emb.puerto} ({benchmark:.0f}% promedio últimos 6 meses).</p>

<h3 style="color:#1a5276;margin-top:30px">1. Desglose de Centros de Costo</h3>
<table style="width:100%;border-collapse:collapse;margin:15px 0">
<tbody>
<tr style="background:#1a5276;color:white">
<th style="padding:12px;text-align:left;border:1px solid #ddd">Centro de Costo</th>
<th style="padding:12px;text-align:right;border:1px solid #ddd">USD</th>
<th style="padding:12px;text-align:right;border:1px solid #ddd">% sobre P×Q</th>
</tr>
<tr style="background:#f8f9fa"><td style="padding:10px;border:1px solid #ddd">Total Amount (P×Q)</td><td style="padding:10px;text-align:right;border:1px solid #ddd">${total_pxq_fmt}</td><td style="padding:10px;text-align:right;border:1px solid #ddd">Base</td></tr>
<tr><td style="padding:10px;border:1px solid #ddd">EXW (Gift Box + Delivery)</td><td style="padding:10px;text-align:right;border:1px solid #ddd">${cc_exw_fmt}</td><td style="padding:10px;text-align:right;border:1px solid #ddd">{emb.cc_exw/base*100:.2f}%</td></tr>
<tr style="background:#f8f9fa"><td style="padding:10px;border:1px solid #ddd">Inland China</td><td style="padding:10px;text-align:right;border:1px solid #ddd">${cc_inland_china_fmt}</td><td style="padding:10px;text-align:right;border:1px solid #ddd">{emb.cc_inland_china/base*100:.2f}%</td></tr>
<tr><td style="padding:10px;border:1px solid #ddd">Flete Marítimo</td><td style="padding:10px;text-align:right;border:1px solid #ddd">${cc_flete_fmt}</td><td style="padding:10px;text-align:right;border:1px solid #ddd">{emb.cc_flete/base*100:.2f}%</td></tr>
<tr style="background:#f8f9fa"><td style="padding:10px;border:1px solid #ddd">Inland Chile</td><td style="padding:10px;text-align:right;border:1px solid #ddd">${cc_inland_chile_fmt}</td><td style="padding:10px;text-align:right;border:1px solid #ddd">{emb.cc_inland_chile/base*100:.2f}%</td></tr>
<tr style="background:#1a5276;color:white;font-weight:bold"><td style="padding:12px;border:1px solid #ddd">TOTAL SOBRECOSTO</td><td style="padding:12px;text-align:right;border:1px solid #ddd">${sobrecosto_usd_fmt}</td><td style="padding:12px;text-align:right;border:1px solid #ddd">{emb.sobrecosto_pct:.2f}%</td></tr>
</tbody></table>

<table style="width:100%;border-collapse:collapse;margin:15px 0">
<tbody>
<tr style="background:#2c3e50;color:white">
<th style="padding:10px;border:1px solid #ddd">Puerto</th>
<th style="padding:10px;border:1px solid #ddd">Benchmark 6m</th>
<th style="padding:10px;border:1px solid #ddd">Este Embarque</th>
<th style="padding:10px;border:1px solid #ddd">Diferencia</th>
</tr>
<tr>
<td style="padding:10px;text-align:center;border:1px solid #ddd"><b>{emb.puerto} ({emb.puerto_nombre})</b></td>
<td style="padding:10px;text-align:center;border:1px solid #ddd">{benchmark:.0f}%</td>
<td style="padding:10px;text-align:center;border:1px solid #ddd">{emb.sobrecosto_pct:.2f}%</td>
<td style="padding:10px;text-align:center;border:1px solid #ddd;color:{color_dif};font-weight:bold">{signo_dif}{diferencia_pp:.1f} pp {simbolo_dif}</td>
</tr>
</tbody></table>

<h3 style="color:#1a5276;margin-top:30px">2. Variación de Productos vs Último Costo</h3>
<p>Se detectan <b>{n_var} productos con variación significativa (&gt;10%)</b> y <b>{n_nuevos} productos nuevos</b>:</p>

<table style="width:100%;border-collapse:collapse;margin:15px 0;font-size:14px">
<tbody>
<tr style="background:#1a5276;color:white">
<th style="padding:10px;text-align:left;border:1px solid #ddd">Producto</th>
<th style="padding:10px;text-align:left;border:1px solid #ddd">SKU</th>
<th style="padding:10px;text-align:right;border:1px solid #ddd">Costo Actual</th>
<th style="padding:10px;text-align:right;border:1px solid #ddd">Último Costo</th>
<th style="padding:10px;text-align:center;border:1px solid #ddd">Embarque Ant.</th>
<th style="padding:10px;text-align:center;border:1px solid #ddd">Variación</th>
</tr>
{filas_aumento()}
{filas_mejora()}
{filas_nuevos()}
</tbody></table>

<div style="background:{bg_concl};border:1px solid {border_concl};border-radius:8px;padding:15px;margin:20px 0">
<strong style="color:{color_concl}">{simbolo_concl} Conclusión:</strong>
<p style="margin:10px 0">{texto_concl}</p>
</div>

<h3 style="color:#1a5276;margin-top:30px">3. Métricas del Embarque</h3>
<table style="width:100%;border-collapse:collapse;margin:15px 0">
<tbody>
<tr><td style="padding:10px;border:1px solid #ddd;width:25%"><b>Productos</b></td><td style="padding:10px;border:1px solid #ddd;width:25%">{len(emb.productos)}</td><td style="padding:10px;border:1px solid #ddd;width:25%"><b>CBM</b></td><td style="padding:10px;border:1px solid #ddd;width:25%">{emb.total_cbm:.2f}</td></tr>
<tr style="background:#f8f9fa"><td style="padding:10px;border:1px solid #ddd"><b>Unidades</b></td><td style="padding:10px;border:1px solid #ddd">{emb.total_unidades:,.0f}</td><td style="padding:10px;border:1px solid #ddd"><b>Dólar</b></td><td style="padding:10px;border:1px solid #ddd">${emb.tarifas.dolar:.0f}</td></tr>
<tr><td style="padding:10px;border:1px solid #ddd"><b>Amount (P×Q)</b></td><td style="padding:10px;border:1px solid #ddd">${total_pxq_fmt} USD</td><td style="padding:10px;border:1px solid #ddd"><b>Flete 40HQ</b></td><td style="padding:10px;border:1px solid #ddd">${cc_flete_fmt} USD</td></tr>
<tr style="background:#f8f9fa"><td style="padding:10px;border:1px solid #ddd"><b>Internado Total</b></td><td style="padding:10px;border:1px solid #ddd" colspan="3">${emb.total_internado_clp:,.0f} CLP</td></tr>
</tbody></table>

{_html_items_sin_sku(emb)}
</div>
<div style="font-family:Arial,sans-serif;line-height:1.6;color:#333;max-width:800px;margin:0 auto">Favor revisar precosteo y documentos para ingresar.</div>
<div style="font-family:Arial,sans-serif;line-height:1.6;color:#333;max-width:800px;margin:0 auto"><br></div>
<div style="font-family:Arial,sans-serif;line-height:1.6;color:#333;max-width:800px;margin:0 auto">Gracias.</div>
</div>"""

    out_path = out_dir / f"email_{emb.numero}.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"\n[OK] Email HTML guardado: {out_path}")

    # Subject
    subject = f"[{emb.numero}] Análisis de Costeo - {emb.puerto} - {emb.sobrecosto_pct:.2f}% ({n_var} productos con variación significativa)"
    (out_dir / f"email_{emb.numero}_subject.txt").write_text(subject, encoding="utf-8")
    return out_path

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Costear embarque COMEX (skill comex-workflow local)")
    parser.add_argument("--pi", required=True, help="Ruta al PI.xlsx")
    parser.add_argument("--pl", required=True, help="Ruta al PL.xlsx")
    parser.add_argument("--tarifas", required=True, help="Ruta al Tarifas_COMEX.xlsx")
    parser.add_argument("--maestra", default=None, help="Ruta a Maestra Importaciones.xlsx (opcional)")
    parser.add_argument("--out", default=None, help="Carpeta de salida (default: ./output_<embarque>)")
    args = parser.parse_args()

    pi_path = Path(args.pi)
    pl_path = Path(args.pl)
    tarifas_path = Path(args.tarifas)
    maestra_path = Path(args.maestra) if args.maestra else None

    for p in (pi_path, pl_path, tarifas_path):
        if not p.exists():
            print(f"[ERROR] No existe: {p}")
            sys.exit(1)

    # 1) Leer
    productos, inland, embarque_num, puerto = leer_pi(pi_path)
    if not productos:
        print("[ERROR] No se extrajeron productos del PI. Aborto.")
        sys.exit(1)

    total_cbm = leer_pl(pl_path, productos)
    tarifas = leer_tarifas(tarifas_path, puerto)

    emb = Embarque(
        numero=embarque_num,
        puerto=tarifas.puerto,
        puerto_nombre=tarifas.puerto_nombre,
        productos=productos,
        inland_china=inland,
        tarifas=tarifas,
        total_cbm=total_cbm,
    )

    # 2) Validacion conceptos
    print(f"\n{'='*70}\n VALIDACION DE CONCEPTOS\n{'='*70}")
    if inland.no_reconocidos or tarifas.no_reconocidos:
        print("\n[!] CONCEPTOS NO RECONOCIDOS:")
        for nr in inland.no_reconocidos:
            print(f"  PI fila {nr['row']}: {nr['descripcion']} - USD {nr['amount']}")
        for nr in tarifas.no_reconocidos:
            print(f"  Tarifas: {nr}")
        print("  Estos conceptos no fueron incluidos en el costeo. Revisar antes de validar.")
    else:
        print("  OK - Todos los conceptos reconocidos.")

    # 3) Calcular
    calcular_costeo(emb)

    # 4) Comparar vs maestra
    comparar_con_maestra(emb, maestra_path)

    # 5) Salidas
    out_dir = Path(args.out) if args.out else Path.cwd() / f"output_{emb.numero}"
    out_dir.mkdir(parents=True, exist_ok=True)

    precosteo_path = generar_precosteo_xlsx(emb, out_dir)
    email_path = generar_email_html(emb, out_dir)

    if maestra_path and maestra_path.exists():
        actualizar_maestra(emb, maestra_path)
    elif maestra_path:
        print(f"\n[!] Maestra no existe en: {maestra_path}. No se actualiza.")
    else:
        print("\n[i] No se proporciono --maestra. No se actualizan pestañas historicas.")

    # Resumen final
    benchmark = BENCHMARKS.get(emb.puerto, {}).get("benchmark", 16.0)
    print(f"\n{'='*70}\n RESUMEN EJECUTIVO\n{'='*70}")
    print(f"  Embarque:        {emb.numero}")
    print(f"  Puerto:          {emb.puerto} ({emb.puerto_nombre})")
    print(f"  ETA:             {emb.tarifas.fecha_eta}")
    print(f"  Productos:       {len(emb.productos)}")
    print(f"  CBM:             {emb.total_cbm:.2f}")
    print(f"  Unidades:        {emb.total_unidades:,.0f}")
    print(f"  Total P×Q:       USD {emb.total_pxq:,.2f}")
    print(f"  Sobrecosto:      USD {emb.sobrecosto_usd:,.2f} ({emb.sobrecosto_pct:.2f}%)")
    print(f"  Benchmark {emb.puerto}:    {benchmark:.0f}%")
    print(f"  Diferencia:      {(emb.sobrecosto_pct - benchmark):+.1f} pp")
    print(f"  Internado total: CLP {emb.total_internado_clp:,.0f}")
    print(f"\n  Pre-costeo:  {precosteo_path}")
    print(f"  Email HTML:  {email_path}")
    print(f"\n  Para crear el borrador en Gmail, abrir el HTML, copiar contenido y pegarlo en un nuevo email.")
    print("="*70)

if __name__ == "__main__":
    main()
